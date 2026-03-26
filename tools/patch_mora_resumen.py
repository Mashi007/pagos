"""Patch reportes_morosidad: first sheet RESUMEN with >121 days."""
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
p = ROOT / "backend/app/api/v1/endpoints/reportes/reportes_morosidad.py"
s = p.read_text(encoding="utf-8", errors="ignore")

if "from datetime import date, timedelta" not in s:
    s = s.replace("from datetime import date\n", "from datetime import date, timedelta\n", 1)

start = s.find("def exportar_morosidad_cedulas(")
if start == -1:
    raise SystemExit("exportar_morosidad_cedulas not found")

NEW = '''def exportar_morosidad_cedulas(
    db: Session = Depends(get_db),
    anos: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atras"),
):
    """Exporta Excel: pestana 1 RESUMEN (>121 dias); luego detalle por mes."""
    import openpyxl

    fc = date.today()
    cutoff_121 = fc - timedelta(days=121)

    periodos = _periodos_desde_filtros(anos, meses_list, meses)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_resumen = wb.create_sheet(title="RESUMEN", index=0)
    ws_resumen.append(
        [
            "Cedula",
            "Cantidad de cuotas (mayores a 121 dias)",
            "Deuda en dolares",
        ]
    )

    rows_resumen = db.execute(
        select(
            Cliente.cedula.label("cedula"),
            func.count(Cuota.id).label("total_cuotas"),
            func.coalesce(func.sum(Cuota.monto), 0).label("total_monto"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < cutoff_121,
        )
        .group_by(Cliente.cedula)
        .order_by(Cliente.cedula.asc())
    ).fetchall()

    total_cuotas_general = 0
    total_monto_general = 0.0
    for r in rows_resumen:
        cuotas = int(r.total_cuotas or 0)
        monto = round(_safe_float(r.total_monto), 2)
        total_cuotas_general += cuotas
        total_monto_general += monto
        ws_resumen.append([r.cedula or "", cuotas, monto])

    ws_resumen.append(["TOTAL", total_cuotas_general, round(total_monto_general, 2)])

    for (ano, mes) in periodos:
        ws = wb.create_sheet(title=f"MORA_{mes:02d}-{ano}")
        ws.append(
            ["Cedula", "Nombre", "Prestamo ID", "Cuota", "Fecha Vencimiento", "Monto USD"]
        )

        rows = db.execute(
            select(
                Cliente.cedula.label("cedula"),
                Cliente.nombres.label("nombre"),
                Prestamo.id.label("prestamo_id"),
                Cuota.numero_cuota.label("numero_cuota"),
                Cuota.fecha_vencimiento.label("fecha_vencimiento"),
                Cuota.monto.label("monto"),
            )
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                func.extract("year", Cuota.fecha_vencimiento) == ano,
                func.extract("month", Cuota.fecha_vencimiento) == mes,
            )
            .order_by(
                Cliente.cedula.asc(),
                Cuota.fecha_vencimiento.asc(),
                Cuota.numero_cuota.asc(),
            )
        ).fetchall()

        if rows:
            for r in rows:
                ws.append(
                    [
                        r.cedula or "",
                        r.nombre or "",
                        int(r.prestamo_id) if r.prestamo_id is not None else None,
                        int(r.numero_cuota) if r.numero_cuota is not None else None,
                        r.fecha_vencimiento.isoformat() if r.fecha_vencimiento else "",
                        round(_safe_float(r.monto), 2),
                    ]
                )
        else:
            ws.append(["", "Sin registros para este periodo", "", "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    hoy_str = date.today().isoformat()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_mora_cedulas_{hoy_str}.xlsx"
        },
    )

'''

p.write_text(s[:start] + NEW, encoding="utf-8")
print("patched", p)
