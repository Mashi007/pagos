# -*- coding: utf-8 -*-
"""Inserta endpoint Excel pagos sin aplicar en pagos.py"""
from pathlib import Path

P = Path(__file__).resolve().parent.parent / "backend" / "app" / "api" / "v1" / "endpoints" / "pagos.py"
text = P.read_text(encoding="utf-8")

old_imp = "from sqlalchemy import and_, case, func, or_, select\n"
new_imp = "from sqlalchemy import and_, case, delete, exists, func, or_, select\n"
if old_imp not in text:
    raise SystemExit("import line not found")
text = text.replace(old_imp, new_imp, 1)

marker = """    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/validar-filas-batch", response_model=dict)"""

insert = """    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/excel/pagos-sin-aplicar-cuotas")
def export_excel_pagos_sin_aplicar_cuotas(
    cohorte: str = Query(
        "fifo",
        description="fifo: cola tipo job (sin cuota_pagos, monto>0, prestamo_id, no ANULADO_IMPORT). "
        "sin_cupo: mismos filtros y ademas sin cupo aplicable en cuotas PENDIENTE/MORA/PARCIAL.",
    ),
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    \"\"\"
    Descarga Excel con pagos que aun no tienen filas en cuota_pagos.
    Requiere autenticacion. Maximo 50000 filas.
    \"\"\"
    c = (cohorte or "fifo").strip().lower()
    if c not in ("fifo", "sin_cupo"):
        raise HTTPException(status_code=400, detail="cohorte debe ser fifo o sin_cupo")

    import openpyxl

    tiene_cuota_pago = exists(select(CuotaPago.id).where(CuotaPago.pago_id == Pago.id))
    cond_base = and_(
        ~tiene_cuota_pago,
        func.coalesce(Pago.monto_pagado, 0) > 0,
        func.upper(func.coalesce(Pago.estado, "")) != "ANULADO_IMPORT",
        Pago.prestamo_id.isnot(None),
    )
    if c == "sin_cupo":
        aplicado_en_cuota = (
            select(func.coalesce(func.sum(CuotaPago.monto_aplicado), 0))
            .where(CuotaPago.cuota_id == Cuota.id)
            .correlate(Cuota)
            .scalar_subquery()
        )
        hay_cupo_aplicable = exists(
            select(1)
            .select_from(Cuota)
            .where(
                and_(
                    Cuota.prestamo_id == Pago.prestamo_id,
                    Cuota.estado.in_(["PENDIENTE", "MORA", "PARCIAL"]),
                    (func.coalesce(Cuota.monto, 0) - aplicado_en_cuota) > 0.01,
                )
            )
        )
        cond_final = and_(cond_base, ~hay_cupo_aplicable)
    else:
        cond_final = cond_base

    rows = (
        db.execute(
            select(Pago)
            .where(cond_final)
            .order_by(Pago.fecha_registro.asc(), Pago.id.asc())
            .limit(50000)
        )
        .scalars()
        .all()
    )
    rows = [r for r in rows if r is not None]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos sin aplicar"
    headers = [
        "pago_id",
        "fecha_registro",
        "fecha_pago",
        "prestamo_id",
        "cedula",
        "monto_pagado",
        "estado",
        "referencia_pago",
        "numero_documento",
        "conciliado",
        "usuario_registro",
        "cohorte_filtro",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, p in enumerate(rows, 2):
        fr = p.fecha_registro
        fp = p.fecha_pago
        ws.cell(row=row_idx, column=1, value=p.id)
        ws.cell(row=row_idx, column=2, value=fr.strftime("%Y-%m-%d %H:%M:%S") if fr else "")
        ws.cell(row=row_idx, column=3, value=fp.strftime("%Y-%m-%d %H:%M:%S") if fp else "")
        ws.cell(row=row_idx, column=4, value=p.prestamo_id)
        ws.cell(row=row_idx, column=5, value=(p.cedula_cliente or ""))
        ws.cell(row=row_idx, column=6, value=float(p.monto_pagado) if p.monto_pagado is not None else 0)
        ws.cell(row=row_idx, column=7, value=p.estado or "")
        ws.cell(row=row_idx, column=8, value=p.referencia_pago or "")
        ws.cell(row=row_idx, column=9, value=p.numero_documento or "")
        ws.cell(row=row_idx, column=10, value=bool(p.conciliado) if p.conciliado is not None else False)
        ws.cell(row=row_idx, column=11, value=p.usuario_registro or "")
        ws.cell(row=row_idx, column=12, value=c)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pagos_sin_aplicar_cuotas_{c}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/validar-filas-batch", response_model=dict)"""

if marker not in text:
    raise SystemExit("marker not found")
text = text.replace(marker, insert, 1)
P.write_text(text, encoding="utf-8")
print("OK", P)
