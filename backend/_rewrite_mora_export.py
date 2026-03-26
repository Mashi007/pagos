"""One-off patch: rewrite exportar_morosidad_cedulas."""
from pathlib import Path

p = Path(__file__).resolve().parent / "app/api/v1/endpoints/reportes/reportes_morosidad.py"
s = p.read_text(encoding="utf-8", errors="ignore")

if "from datetime import date, timedelta" not in s:
    s = s.replace("from datetime import date\n", "from datetime import date, timedelta\n", 1)

start = s.find("def exportar_morosidad_cedulas(")
if start == -1:
    raise SystemExit("function not found")
end = s.find("\n\n\n", s.find("return Response(", start))
# find end of function: after last return Response block of this function
end = s.find(
    '    return Response(\n        content=content,\n        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",\n        headers={\n            "Content-Disposition": f"attachment; filename=reporte_mora_cedulas_{hoy_str}.xlsx"\n        },\n    )\n',
    start,
)
if end == -1:
    raise SystemExit("end marker not found")
end = s.find("\n", end + 1)
# include closing newline after function
while end < len(s) and s[end] == "\n":
    end += 1
    if s[end:end + 1] not in ("\n", " ", "\t") and s[end:end + 4] != "def ":
        break
# simpler: find line after the closing paren of return Response
idx = s.find('filename=reporte_mora_cedulas_', start)
idx = s.find(")\n", idx) + 2
# skip blank lines until next @router or def at column 0
while idx < len(s) and s[idx] in " \n":
    idx += 1
end = idx

new_fn = '''def exportar_morosidad_cedulas(
    db: Session = Depends(get_db),
    anos: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atras"),
):
    """Exporta Excel: 1) pestana RESUMEN (cuotas >121 dias); 2) detalle por mes."""
    import openpyxl

    fc = date.today()
    cutoff_121 = fc - timedelta(days=121)

    periodos = _periodos_desde_filtros(anos, meses_list, meses)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Primera pestana: resumen por cedula (cuotas impagas con mas de 121 dias de atraso).
    ws_resumen = wb.create_sheet(title="RESUMEN", index=0)
    ws_resumen.append(
        [
            "Cedula",
            "Cantidad de cuotas (mayores a 121 dias)",
            "Deuda en dolares",
        ]
    )

    rows_resumen = db.execute(
        select(
            Cliente.cedula.label("cedula"),
            func.count(Cuota.id).label("total_cuotas"),
            func.coalesce(func.sum(Cuota.monto), 0).label("total_monto"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < cutoff_121,
        )
        .group_by(Cliente.cedula)
        .order_by(Cliente.cedula.asc())
    ).fetchall()

    total_cuotas_general = 0
    total_monto_general = 0.0
    for r in rows_resumen:
        cuotas = int(r.total_cuotas or 0)
        monto = round(_safe_float(r.total_monto), 2)
        total_cuotas_general += cuotas
        total_monto_general += monto
        ws_resumen.append([r.cedula or "", cuotas, monto])

    ws_resumen.append(["TOTAL", total_cuotas_general, round(total_monto_general, 2)])

    for (ano, mes) in periodos:
        ws = wb.create_sheet(title=f"MORA_{mes:02d}-{ano}")
        ws.append(
            ["Cedula", "Nombre", "Prestamo ID", "Cuota", "Fecha Vencimiento", "Monto USD"]
        )

        rows = db.execute(
            select(
                Cliente.cedula.label("cedula"),
                Cliente.nombres.label("nombre"),
                Prestamo.id.label("prestamo_id"),
                Cuota.numero_cuota.label("numero_cuota"),
                Cuota.fecha_vencimiento.label("fecha_vencimiento"),
                Cuota.monto.label("monto"),
            )
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                func.extract("year", Cuota.fecha_vencimiento) == ano,
                func.extract("month", Cuota.fecha_vencimiento) == mes,
            )
            .order_by(
                Cliente.cedula.asc(),
                Cuota.fecha_vencimiento.asc(),
                Cuota.numero_cuota.asc(),
            )
        ).fetchall()

        if rows:
            for r in rows:
                ws.append(
                    [
                        r.cedula or "",
                        r.nombre or "",
                        int(r.prestamo_id) if r.prestamo_id is not None else None,
                        int(r.numero_cuota) if r.numero_cuota is not None else None,
                        r.fecha_vencimiento.isoformat() if r.fecha_vencimiento else "",
                        round(_safe_float(r.monto), 2),
                    ]
                )
        else:
            ws.append(["", "Sin registros para este periodo", "", "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    hoy_str = date.today().isoformat()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_mora_cedulas_{hoy_str}.xlsx"
        },
    )

'''

# Find exact end: next line that starts with @router after function
rest = s[start:]
next_router = rest.find("\n@router", 1)
if next_router == -1:
    next_router = rest.find("\ndef exportar_morosidad", 1)
if next_router == -1:
    raise SystemExit("cannot find function end")
end_abs = start + next_router
s = s[:start] + new_fn + s[end_abs:]
p.write_text(s, encoding="utf-8")
print("OK", start, end_abs)
