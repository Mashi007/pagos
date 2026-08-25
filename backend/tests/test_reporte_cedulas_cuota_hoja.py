from datetime import date
from decimal import Decimal
from io import BytesIO

from app.services.reporte_cedulas_cuota_hoja import (
    FECHA_CORTE_JUNIO,
    conteo_cuotas_en_mora,
    cuota_unica_de_prestamos,
    estado_actual_de_prestamos,
    filas_cedula_cuota,
    generar_excel_cedulas_cuota,
    hay_pagos_parciales_a_cuotas_en_mora,
    metricas_corte_mora,
    parsear_cedulas_csv,
    pendiente_vencido,
    pagos_aplicados_a_vencido_o_mora,
    pagos_parciales_a_cuotas_en_mora,
    saldo_total_prestamo,
    saldo_vencido_solo_mora,
    _items_cuotas_para_informe,
)


def test_parsear_salta_encabezado_y_conserva_orden():
    raw = "Cédula\nE84491751\nV18739623\n".encode("utf-8")
    assert parsear_cedulas_csv(raw) == ["E84491751", "V18739623"]


def test_cuota_unica_un_aprobado():
    assert cuota_unica_de_prestamos([("APROBADO", Decimal("180.00"))]) == Decimal(
        "180.00"
    )


def test_cuota_unica_sin_prestamo_no_inventa():
    assert cuota_unica_de_prestamos([]) is None


def test_cuota_unica_montos_distintos_no_elige():
    assert (
        cuota_unica_de_prestamos(
            [("APROBADO", Decimal("180")), ("APROBADO", Decimal("200"))]
        )
        is None
    )


def test_cuota_unica_prioriza_aprobado():
    assert cuota_unica_de_prestamos(
        [("LIQUIDADO", Decimal("50")), ("APROBADO", Decimal("180"))]
    ) == Decimal("180.00")


def test_cuota_unica_usa_estado_en_revision():
    assert cuota_unica_de_prestamos(
        [("EN_REVISION", Decimal("160.00"))]
    ) == Decimal("160.00")


def test_cuota_unica_usa_desistimiento_si_no_hay_aprobado():
    assert cuota_unica_de_prestamos(
        [("DESISTIMIENTO", Decimal("175.00"))]
    ) == Decimal("175.00")


def test_estado_actual_prioriza_aprobado():
    assert estado_actual_de_prestamos(
        [("LIQUIDADO", Decimal("50")), ("APROBADO", Decimal("180"))]
    ) == "APROBADO"
    assert estado_actual_de_prestamos([("DESISTIMIENTO", 1)]) == "DESISTIMIENTO"
    assert estado_actual_de_prestamos([]) is None


def test_filas_cruza_prefijo_e_con_v_en_sistema():
    items = [
        (1, date(2025, 9, 1), Decimal("180"), Decimal("0"), None, 12),
        (2, date(2025, 10, 1), Decimal("180"), Decimal("0"), None, 12),
        (3, date(2025, 11, 1), Decimal("180"), Decimal("0"), None, 12),
        (4, date(2025, 12, 1), Decimal("180"), Decimal("0"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["E84491751"],
        {"V84491751": [("LIQUIDADO", Decimal("180.00"))]},
        {"V84491751": items},
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas[0]["cedula"] == "E84491751"
    assert filas[0]["estado"] == "LIQUIDADO"
    assert filas[0]["cuota"] == 180.0
    assert filas[0]["mora_hoy"] == 4
    assert filas[0]["saldo_total_prestamo"] == 720.0
    assert filas[0]["saldo_a_pagar"] == 720.0


def test_filas_celda_vacia_si_no_hay_cuota():
    filas = filas_cedula_cuota(
        ["E84491751", "V999"],
        {"E84491751": [("LIQUIDADO", Decimal("180.5"))]},
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas[0]["cuota"] == 180.5
    assert filas[0]["estado"] == "LIQUIDADO"
    assert filas[1]["cedula"] == "V999"
    assert filas[1]["cuota"] is None
    assert filas[1]["estado"] is None


def test_aprobado_muestra_conteo_mora_aunque_sea_menor_a_4():
    items = [
        (10, 1, date(2026, 1, 15), Decimal("500"), Decimal("0"), None, 12),
        (11, 2, date(2026, 2, 15), Decimal("500"), Decimal("0"), None, 12),
        (12, 3, date(2026, 3, 15), Decimal("500"), Decimal("0"), None, 12),
        (13, 4, date(2026, 6, 15), Decimal("500"), Decimal("100"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["E84491751"],
        {"E84491751": [("APROBADO", Decimal("500"))]},
        {"E84491751": items},
        fecha_hoy=date(2026, 8, 20),
    )
    assert conteo_cuotas_en_mora(items, date(2026, 8, 20)) == 3
    assert filas[0]["mora_hoy"] == 3
    # 3*500 MORA + 400 VENCIDO = 1900 total del préstamo
    assert filas[0]["saldo_total_prestamo"] == 1900.0


def test_aprobado_con_4_en_mora_muestra_junio_y_hoy():
    items = [
        (1, date(2025, 9, 1), Decimal("100"), Decimal("0"), None, 12),
        (2, date(2025, 10, 1), Decimal("100"), Decimal("0"), None, 12),
        (3, date(2025, 11, 1), Decimal("100"), Decimal("0"), None, 12),
        (4, date(2025, 12, 1), Decimal("100"), Decimal("0"), None, 12),
        (5, date(2026, 1, 1), Decimal("100"), Decimal("0"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["E84491751"],
        {"E84491751": [("APROBADO", Decimal("100"))]},
        {"E84491751": items},
        fecha_junio=FECHA_CORTE_JUNIO,
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas[0]["mora_junio"] == 5
    assert filas[0]["mora_hoy"] == 5
    assert filas[0]["saldo_total_prestamo"] == 500.0


def test_conteo_mora_junio_y_hoy_independiente_sin_ocultar():
    items = [
        (1, date(2025, 12, 1), Decimal("100"), Decimal("0"), None, 12),
        (2, date(2026, 1, 1), Decimal("100"), Decimal("0"), None, 12),
        (3, date(2026, 2, 1), Decimal("100"), Decimal("0"), None, 12),
        (4, date(2026, 3, 1), Decimal("100"), Decimal("0"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["E84491751"],
        {"E84491751": [("APROBADO", Decimal("100"))]},
        {"E84491751": items},
        fecha_junio=date(2026, 6, 1),
        fecha_hoy=date(2026, 8, 20),
    )
    assert conteo_cuotas_en_mora(items, date(2026, 6, 1)) == 2
    assert filas[0]["mora_junio"] == 2
    assert filas[0]["mora_hoy"] == 4
    assert filas[0]["saldo_total_prestamo"] == 400.0


def test_aprobado_informe_solo_cuotas_del_prestamo_aprobado():
    ref = date(2026, 8, 20)
    items_aprob = [
        (1, date(2026, 1, 15), Decimal("500"), Decimal("0"), None, 12),
        (2, date(2026, 2, 15), Decimal("500"), Decimal("0"), None, 12),
        (3, date(2026, 3, 15), Decimal("500"), Decimal("0"), None, 12),
    ]
    items_liq = [
        (1, date(2024, 1, 15), Decimal("200"), Decimal("0"), None, 12),
        (2, date(2024, 2, 15), Decimal("200"), Decimal("0"), None, 12),
        (3, date(2024, 3, 15), Decimal("200"), Decimal("0"), None, 12),
    ]
    items_all = items_aprob + items_liq
    items = _items_cuotas_para_informe(
        "25133615",
        {"25133615": items_all},
        {"25133615": items_aprob},
        {"25133615": {"APROBADO", "LIQUIDADO"}},
    )
    assert conteo_cuotas_en_mora(items, ref) == 3
    assert conteo_cuotas_en_mora(items_all, ref) == 6


def test_saldo_vencido_solo_cuotas_en_mora():
    ref = date(2026, 8, 20)
    items = [
        (1, date(2026, 1, 15), Decimal("100"), Decimal("0"), None, 12),  # MORA
        (2, date(2026, 2, 15), Decimal("100"), Decimal("0"), None, 12),  # MORA
        (3, date(2026, 3, 15), Decimal("100"), Decimal("0"), None, 12),  # MORA
        (4, date(2026, 6, 15), Decimal("100"), Decimal("0"), None, 12),  # VENCIDO
        (5, date(2026, 9, 15), Decimal("100"), Decimal("0"), None, 12),  # PENDIENTE
    ]
    assert conteo_cuotas_en_mora(items, ref) == 3
    assert saldo_vencido_solo_mora(items, ref) == Decimal("300.00")
    # Total del préstamo incluye vencido + pendiente
    assert saldo_total_prestamo(items, ref) == Decimal("500.00")


def test_pagos_junio_hasta_31_may_y_hoy_desde_1_jun():
    """Todos los pagos del préstamo: ≤31 may en corte junio; 1 jun–hoy en corte hoy."""
    items = [
        (101, 1, date(2026, 1, 15), Decimal("100"), Decimal("50"), None, 12),  # MORA
        (102, 2, date(2026, 6, 15), Decimal("100"), Decimal("30"), None, 12),  # VENCIDO
        (103, 3, date(2026, 9, 15), Decimal("100"), Decimal("0"), None, 12),  # PENDIENTE
    ]
    apps = [
        (date(2026, 5, 20), Decimal("40.00"), 101),
        (date(2026, 5, 31), Decimal("10.00"), 102),
        (date(2026, 6, 1), Decimal("20.00"), 101),
        (date(2026, 7, 10), Decimal("15.00"), 102),
        (date(2026, 7, 10), Decimal("99.00"), 103),  # pendiente: también cuenta
        (date(2026, 4, 1), Decimal("5.00"), 999),
    ]
    hoy = date(2026, 8, 20)
    n, sal, pag_j = metricas_corte_mora(
        items,
        date(2026, 6, 1),
        es_aprobado=False,
        aplicaciones=apps,
        pagos_desde=None,
        pagos_hasta=date(2026, 5, 31),
    )
    assert pag_j == Decimal("55.00")  # 40+10+5
    assert sal == Decimal("50.00")

    n2, sal2, pag_h = metricas_corte_mora(
        items,
        hoy,
        es_aprobado=False,
        aplicaciones=apps,
        pagos_desde=date(2026, 6, 1),
        pagos_hasta=hoy,
    )
    assert pag_h == Decimal("134.00")  # 20+15+99
    assert sal2 == Decimal("50.00")  # solo MORA cuota 101


def test_pagos_ventanas_exclusivas_por_fecha_pago():
    """Punto 1: ≤1 jun; punto 2: desde 2 jun. Sin solape."""
    from datetime import datetime

    from app.services.reporte_cedulas_cuota_hoja import sumar_pagos_ventanas_exclusivas

    pagos = [
        (date(2026, 5, 31), Decimal("100")),
        (datetime(2026, 5, 31, 23, 59, 0), Decimal("50")),
        (date(2026, 6, 1), Decimal("200")),  # punto 1
        (date(2026, 6, 2), Decimal("40")),  # punto 2
        (datetime(2026, 8, 20, 8, 0, 0), Decimal("25")),
        (date(2026, 8, 21), Decimal("999")),  # después de hoy
    ]
    p1, p2 = sumar_pagos_ventanas_exclusivas(
        pagos, corte_junio=date(2026, 6, 1), fecha_hoy=date(2026, 8, 20)
    )
    assert p1 == Decimal("350.00")  # 100+50+200
    assert p2 == Decimal("65.00")  # 40+25
    assert p1 + p2 == Decimal("415.00")


def test_saldo_a_pagar_es_mora_menos_todos_los_pagos():
    from app.services.reporte_cedulas_cuota_hoja import (
        saldo_a_pagar,
        saldo_neto_mora_menos_pagos,
    )

    assert saldo_a_pagar(
        Decimal("900.00"), Decimal("50.00"), Decimal("80.00")
    ) == Decimal("770.00")
    assert saldo_a_pagar(Decimal("100.00"), Decimal("0"), Decimal("0")) == Decimal(
        "100.00"
    )
    assert saldo_a_pagar(None, Decimal("10.00"), Decimal("5.00")) == Decimal("0.00")
    assert saldo_a_pagar(None, Decimal("0"), Decimal("0")) is None
    # Mora 720, pagos 900 → 0 (no negativo)
    assert saldo_neto_mora_menos_pagos(
        Decimal("720.00"), Decimal("900.00")
    ) == Decimal("0.00")


def test_pagos_aplicados_filtra_ventana():
    items = [(101, 1, date(2026, 1, 1), Decimal("100"), Decimal("0"), None, 12)]
    apps = [
        (date(2026, 5, 31), Decimal("10"), 101),
        (date(2026, 6, 1), Decimal("20"), 101),
    ]
    assert pagos_aplicados_a_vencido_o_mora(
        items,
        apps,
        as_of=date(2026, 8, 20),
        fecha_desde=None,
        fecha_hasta=date(2026, 5, 31),
    ) == Decimal("10.00")
    assert pagos_aplicados_a_vencido_o_mora(
        items,
        apps,
        as_of=date(2026, 8, 20),
        fecha_desde=date(2026, 6, 1),
        fecha_hasta=date(2026, 8, 20),
    ) == Decimal("20.00")


def test_metricas_corte_aprobado_muestra_conteo_menor_a_4():
    items = [
        (1, date(2026, 1, 15), Decimal("100"), Decimal("0"), None, 12),
        (2, date(2026, 2, 15), Decimal("100"), Decimal("0"), None, 12),
        (3, date(2026, 3, 15), Decimal("100"), Decimal("0"), None, 12),
        (4, date(2026, 6, 15), Decimal("100"), Decimal("0"), None, 12),
    ]
    n, s, p = metricas_corte_mora(items, date(2026, 8, 20), es_aprobado=True)
    assert n == 3
    assert s == Decimal("300.00")
    assert p is None


def test_pendiente_vencido_solo_si_ya_vencio_y_hay_saldo():
    ref = date(2026, 8, 18)
    assert pendiente_vencido(
        Decimal("180"), Decimal("0"), date(2026, 1, 15), None, ref
    ) == Decimal("180.00")
    assert (
        pendiente_vencido(
            Decimal("180"), Decimal("180"), date(2026, 1, 15), date(2026, 1, 20), ref
        )
        is None
    )


def test_fecha_pago_sin_monto_no_borra_mora_hoy():
    """fecha_pago sola con total_pagado=0 no cuenta como pagada (bug que ponía 0 hoy)."""
    items = [
        (1, date(2025, 9, 1), Decimal("100"), Decimal("0"), date(2026, 7, 1), 12),
        (2, date(2025, 10, 1), Decimal("100"), Decimal("0"), date(2026, 7, 1), 12),
        (3, date(2025, 11, 1), Decimal("100"), Decimal("0"), None, 12),
    ]
    hoy = date(2026, 8, 20)
    assert conteo_cuotas_en_mora(items, hoy) == 3
    filas = filas_cedula_cuota(
        ["V19208662"],
        {"V19208662": [("APROBADO", Decimal("100"))]},
        {"V19208662": items},
        fecha_hoy=hoy,
    )
    assert filas[0]["mora_hoy"] == 3
    assert filas[0]["mora_junio"] == 3


def test_conteo_mora_muestra_1_a_n_sin_ocultar():
    items = [
        (1, date(2025, 12, 1), Decimal("100"), Decimal("0"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["V1"],
        {"V1": [("APROBADO", Decimal("100"))]},
        {"V1": items},
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas[0]["mora_hoy"] == 1
    # Sin cuotas → 0 explícito
    filas0 = filas_cedula_cuota(
        ["V2"],
        {"V2": [("APROBADO", Decimal("100"))]},
        {},
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas0[0]["mora_hoy"] == 0


def test_saldo_total_incluye_mora_vencido_y_pendiente():
    items = [
        (1, date(2025, 9, 1), Decimal("180"), Decimal("0"), None, 12),
        (2, date(2025, 10, 1), Decimal("180"), Decimal("0"), None, 12),
        (3, date(2025, 11, 1), Decimal("180"), Decimal("0"), None, 12),
        (4, date(2025, 12, 1), Decimal("180"), Decimal("0"), None, 12),
        (5, date(2026, 9, 1), Decimal("180"), Decimal("50"), None, 12),  # pendiente parcial
    ]
    filas = filas_cedula_cuota(
        ["E84491751"],
        {"E84491751": [("APROBADO", Decimal("180"))]},
        {"E84491751": items},
        fecha_hoy=date(2026, 8, 20),
    )
    assert filas[0]["mora_hoy"] == 4
    assert filas[0]["saldo_total_prestamo"] == 850.0  # 4*180 + 130


def test_mora_con_abono_parcial_no_cuenta_en_informe():
    """Abono ≥ 0.10: no entra en mora_hoy (E); sí entra en mora_junio (D) si ya era MORA."""
    hoy = date(2026, 8, 20)
    items = [
        (1, date(2026, 1, 15), Decimal("100"), Decimal("0"), None, 12),  # MORA D y E
        (2, date(2026, 2, 15), Decimal("100"), Decimal("0.10"), None, 12),  # E: no (parcial)
        (3, date(2026, 3, 15), Decimal("100"), Decimal("40"), None, 12),  # E: no (parcial)
        (4, date(2026, 1, 1), Decimal("100"), Decimal("0.09"), None, 12),  # <0.10: D y E
        # Ya MORA al 1 jun (venc. 1 ene + 4m6d) con parcial: D sí, E no
        (5, date(2026, 1, 1), Decimal("100"), Decimal("50"), None, 12),
    ]
    assert conteo_cuotas_en_mora(items, hoy) == 5
    assert conteo_cuotas_en_mora(items, hoy, excluir_abono_parcial=True) == 2
    filas = filas_cedula_cuota(
        ["V1"],
        {"V1": [("APROBADO", Decimal("100"))]},
        {"V1": items},
        fecha_hoy=hoy,
    )
    assert filas[0]["mora_hoy"] == 2
    # Al 1 jun: MORA = 1, 4 y 5 (2 y 3 aún VENCIDO). D no quita la 5 por parcial.
    assert filas[0]["mora_junio"] == 3


def test_abono_parcial_posterior_al_corte_no_excluye_en_junio():
    """Pago con fecha_pago después del 1 jun no cuenta como abono al corte junio."""
    items = [
        (
            1,
            date(2026, 1, 15),
            Decimal("100"),
            Decimal("50"),
            date(2026, 7, 1),
            12,
        ),
    ]
    assert conteo_cuotas_en_mora(items, date(2026, 6, 1)) == 1
    assert (
        conteo_cuotas_en_mora(
            items, date(2026, 8, 20), excluir_abono_parcial=True
        )
        == 0
    )


def test_excel_cuotas_mora_y_saldo_total():
    import openpyxl

    content = generar_excel_cedulas_cuota(
        [
            {
                "cedula": "E1",
                "estado": "APROBADO",
                "cuota": 180.0,
                "fecha_punto_1": date(2026, 6, 1),
                "fecha_junio": date(2026, 6, 1),
                "fecha_hoy": date(2026, 8, 20),
                "mora_junio": 4,
                "mora_hoy": 5,
                "hay_pagos_parciales_mora": True,
                "pagos_parciales_mora": 40.0,
                "saldo_total_prestamo": 2160.0,
            },
            {"cedula": "V2", "estado": None, "cuota": None},
        ]
    )
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    assert ws["A1"].value == "Cédula"
    assert "Cuotas en mora al 1 jun 2026" in str(ws["D1"].value)
    assert "sin abono parcial" not in str(ws["D1"].value)
    assert "Cuotas en mora hoy" in str(ws["E1"].value)
    assert "sin abono parcial" in str(ws["E1"].value)
    assert "Pagos parciales a mora" in str(ws["F1"].value)
    assert ws["G1"].value == "Saldo total préstamo ($)"
    assert ws["F2"].value == "Sí ($40.00)"
    assert ws["F3"].value == "No"
    assert ws["G2"].value == 2160.0


def test_pagos_parciales_a_mora_en_ventana_junio_hoy():
    """Solo abonos parciales (es_pago_completo=False) a cuotas en MORA con saldo."""
    hoy = date(2026, 8, 20)
    items = [
        # MORA parcial
        (101, 1, date(2026, 1, 15), Decimal("100"), Decimal("40"), date(2026, 7, 1), 12),
        # MORA sin abono
        (102, 2, date(2026, 2, 15), Decimal("100"), Decimal("0"), None, 12),
        # VENCIDO parcial (no mora)
        (103, 3, date(2026, 6, 15), Decimal("100"), Decimal("30"), date(2026, 7, 5), 12),
    ]
    apps = [
        (date(2026, 5, 20), Decimal("10.00"), 101, False),  # antes de 1 jun
        (date(2026, 6, 1), Decimal("25.00"), 101, False),  # cuenta
        (date(2026, 7, 1), Decimal("15.00"), 101, False),  # cuenta
        (date(2026, 7, 5), Decimal("30.00"), 103, False),  # vencido: no
        (date(2026, 7, 10), Decimal("50.00"), 101, True),  # completo: no
        (date(2026, 8, 1), Decimal("5.00"), 999, False),  # otra cuota
    ]
    monto = pagos_parciales_a_cuotas_en_mora(
        items,
        apps,
        fecha_desde=date(2026, 6, 1),
        fecha_hasta=hoy,
        as_of=hoy,
    )
    assert monto == Decimal("40.00")  # 25+15
    assert hay_pagos_parciales_a_cuotas_en_mora(
        items,
        apps,
        fecha_desde=date(2026, 6, 1),
        fecha_hasta=hoy,
        as_of=hoy,
    )
    filas = filas_cedula_cuota(
        ["V1"],
        {"V1": [("APROBADO", Decimal("100"))]},
        {"V1": items},
        apps_cuota_por_norm={"V1": apps},
        fecha_hoy=hoy,
    )
    assert filas[0]["hay_pagos_parciales_mora"] is True
    assert filas[0]["pagos_parciales_mora"] == 40.0


def test_sin_pagos_parciales_a_mora_queda_no():
    hoy = date(2026, 8, 20)
    items = [
        (101, 1, date(2026, 1, 15), Decimal("100"), Decimal("0"), None, 12),
    ]
    filas = filas_cedula_cuota(
        ["V1"],
        {"V1": [("APROBADO", Decimal("100"))]},
        {"V1": items},
        fecha_hoy=hoy,
    )
    assert filas[0]["hay_pagos_parciales_mora"] is False
    assert filas[0]["pagos_parciales_mora"] is None
