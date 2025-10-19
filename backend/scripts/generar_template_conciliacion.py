#!/usr/bin/env python3
"""
Script para generar template Excel de conciliación bancaria
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

def crear_template_conciliacion():
    """Crear template Excel para conciliación bancaria"""
    
    # Crear workbook
    wb = Workbook()
    
    # HOJA 1: INSTRUCCIONES
    ws_instrucciones = wb.active
    ws_instrucciones.title = "Instrucciones"
    
    instrucciones = [
        ["INSTRUCCIONES PARA CONCILIACIÓN BANCARIA"],
        [""],
        ["1. FORMATO DE ARCHIVO:"],
        ["   - Archivo Excel (.xlsx)"],
        ["   - Primera fila: encabezados de columnas"],
        ["   - Datos desde la segunda fila"],
        ["   - Máximo 100 registros por archivo"],
        [""],
        ["2. COLUMNAS REQUERIDAS:"],
        ["   - fecha: Fecha real del pago (formato YYYY-MM-DD)"],
        ["   - numero_documento: Número de documento del pago"],
        [""],
        ["3. PROCESO DE CONCILIACIÓN:"],
        ["   - El sistema compara el número de documento con la base de datos"],
        ["   - Si hay coincidencia EXACTA: se marca como CONCILIADO"],
        ["   - Si NO hay coincidencia: se marca como PENDIENTE"],
        [""],
        ["4. VALIDACIONES:"],
        ["   - Fecha debe estar en formato YYYY-MM-DD"],
        ["   - Número de documento debe coincidir EXACTAMENTE"],
        ["   - No se permiten caracteres especiales adicionales"],
        [""],
        ["5. EJEMPLOS DE DATOS VÁLIDOS:"],
        ["   - fecha: '2024-01-15'"],
        ["   - numero_documento: 'DOC001234'"],
        [""],
        ["6. NOTAS IMPORTANTES:"],
        ["   - No eliminar las columnas"],
        ["   - No cambiar el orden de las columnas"],
        ["   - Usar solo caracteres ASCII"],
        ["   - Verificar que los números de documento sean exactos"],
        ["   - Un archivo por vez"],
        [""],
        ["7. RESULTADO:"],
        ["   - Los pagos conciliados aparecerán en el resumen"],
        ["   - Los pendientes requerirán revisión manual"],
        [""],
        ["8. DESCONCILIACIÓN:"],
        ["   - Se puede desconciliar un pago ya conciliado"],
        ["   - Requiere formulario con justificación"],
        ["   - Se registra en auditoría"],
        [""],
        ["FECHA DE GENERACIÓN: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["GENERADO POR: Sistema de Conciliación Bancaria"]
    ]
    
    # Agregar instrucciones a la hoja
    for i, instruccion in enumerate(instrucciones, 1):
        ws_instrucciones.cell(row=i, column=1, value=instruccion[0])
    
    # HOJA 2: TEMPLATE VACÍA
    ws_template = wb.create_sheet("Template_Conciliacion")
    
    # Encabezados
    encabezados = ["fecha", "numero_documento"]
    for i, encabezado in enumerate(encabezados, 1):
        ws_template.cell(row=1, column=i, value=encabezado)
    
    # Ejemplo de datos
    ejemplo = ["2024-01-15", "DOC001234"]
    for i, valor in enumerate(ejemplo, 1):
        ws_template.cell(row=2, column=i, value=valor)
    
    # Aplicar validaciones
    # Validación para fecha (formato YYYY-MM-DD)
    dv_fecha = DataValidation(type="date", formula1="2020-01-01", formula2="2030-12-31")
    dv_fecha.add("A2:A100")
    ws_template.add_data_validation(dv_fecha)
    
    # Guardar archivo
    filename = f"Template_Conciliacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    print(f"✅ Template generado: {filename}")
    print(f"📁 Ubicación: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    print("🚀 Generando template de conciliación bancaria...")
    filename = crear_template_conciliacion()
    print(f"🎯 Template listo para validar: {filename}")
