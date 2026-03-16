# Add GET /download-excel-temporal and GmailTemporal import
import os
base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path = os.path.join(base, "app", "api", "v1", "endpoints", "pagos_gmail.py")

with open(path, "r", encoding="utf-8", errors="replace") as f:
    c = f.read()

if "download-excel-temporal" in c or "download_excel_temporal" in c:
    print("Endpoint already present")
    exit(0)

# 1) Add GmailTemporal to import
c = c.replace(
    "from app.models.pagos_gmail_sync import PagosGmailSync, PagosGmailSyncItem",
    "from app.models.pagos_gmail_sync import PagosGmailSync, PagosGmailSyncItem, GmailTemporal",
)

# 2) Insert new endpoint before @router.get("/diagnostico")
endpoint = '''
@router.get("/download-excel-temporal")
def download_excel_temporal(db: Session = Depends(get_db)):
    """
    Genera Excel desde la tabla temporal gmail_temporal (cada procesamiento Gmail inserta a continuacion).
    Tras enviar el archivo se vacia gmail_temporal (TRUNCATE). Si no hay datos devuelve 404.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    items = db.execute(
        select(GmailTemporal).order_by(GmailTemporal.created_at)
    ).scalars().all()
    items = list(items)
    if not items:
        raise HTTPException(
            status_code=404,
            detail="Sin datos en tabla temporal. Procese correos Gmail primero; cada procesamiento se almacena a continuacion en gmail_temporal.",
        )
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append(["Correo Pagador", "Fecha Pago", "Cedula", "Monto", "Referencia", "Link", "Ver email"])
        link_font = Font(color="0563C1", underline="single")
        for row_idx, it in enumerate(items, start=2):
            link_url = (it.drive_link or "").strip()
            if link_url and not link_url.startswith("http"):
                link_url = "https://drive.google.com/file/d/" + link_url + "/view"
            email_url = (it.drive_email_link or "").strip()
            if email_url and not email_url.startswith("http"):
                email_url = "https://drive.google.com/file/d/" + email_url + "/view"
            email_text = "Ver email" if email_url else ""
            ws.append([
                it.correo_origen or "",
                it.fecha_pago or "",
                formatear_cedula(it.cedula or ""),
                it.monto or "",
                it.numero_referencia or "",
                "Ver imagen" if link_url else "",
                email_text,
            ])
            if link_url:
                c6 = ws.cell(row=row_idx, column=6)
                c6.hyperlink = link_url
                c6.value = "Ver imagen"
                c6.font = link_font
            if email_url:
                c7 = ws.cell(row=row_idx, column=7)
                c7.hyperlink = email_url
                c7.value = "Ver email"
                c7.font = link_font
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        logger.exception("[PAGOS_GMAIL] Error generando Excel temporal: %s", e)
        raise HTTPException(status_code=500, detail="Error al generar Excel.") from e

    # Vaciar tabla temporal despues de generar el Excel
    db.execute(delete(GmailTemporal))
    db.commit()

    from datetime import datetime as dt
    date_str = dt.utcnow().strftime("%Y-%m-%d")
    filename = f"Pagos_Gmail_temporal_{date_str}.xlsx"
    logger.info("[PAGOS_GMAIL] download-excel-temporal OK filas=%s tabla vaciada", len(items))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


'''

# Insert before @router.get("/diagnostico")
marker = '@router.get("/diagnostico")'
if marker not in c:
    marker = '@router.get(\'/diagnostico\')'
if marker not in c:
    print("diagnostico marker not found")
    exit(1)
c = c.replace(marker, endpoint.rstrip() + "\n\n" + marker)

with open(path, "w", encoding="utf-8") as f:
    f.write(c)
print("Added GET /download-excel-temporal and vaciar BD al descargar")
