#!/usr/bin/env python3
"""
Script de Actualización Automática de Plantilla Excel
=====================================================

Este script:
- Monitorea cambios en tablas de configuración (modelos, concesionarios, analistas)
- Regenera automáticamente la plantilla Excel cuando hay cambios
- Mantiene un cache de la última versión para evitar regeneraciones innecesarias
- Se puede ejecutar como servicio o cron job
- Notifica cuando hay cambios significativos

Uso:
    python auto_update_excel_template.py [--watch] [--force] [--notify]
"""

import os
import sys
import json
import hashlib
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional
import argparse

# Agregar el directorio raíz al path para imports
sys.path.append(str(Path(__file__).parent.parent))

from sqlalchemy.orm import Session
from app.db.session import get_db
from app.models.modelo_vehiculo import ModeloVehiculo
from app.models.concesionario import Concesionario
from app.models.analista import Analista
from app.models.cliente import Cliente
from openpyxl import Workbook
from openpyxl.styles import Font
import io

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_template_updates.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ExcelTemplateUpdater:
    """Actualizador automático de plantilla Excel"""
    
    def __init__(self, cache_file: str = "template_cache.json"):
        self.cache_file = Path(cache_file)
        self.cache_data = self._load_cache()
        
    def _load_cache(self) -> Dict:
        """Cargar cache de versiones anteriores"""
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.warning(f"Error cargando cache: {e}")
        return {
            "last_update": None,
            "modelos_hash": None,
            "concesionarios_hash": None,
            "analistas_hash": None,
            "total_clientes": 0,
            "version": 1
        }
    
    def _save_cache(self):
        """Guardar cache actualizado"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache_data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"Error guardando cache: {e}")
    
    def _get_data_hash(self, data: List[str]) -> str:
        """Generar hash de los datos para detectar cambios"""
        data_str = "|".join(sorted(data))
        return hashlib.md5(data_str.encode('utf-8')).hexdigest()
    
    def _get_current_data(self, db: Session) -> Dict:
        """Obtener datos actuales de la BD"""
        try:
            # Modelos de vehículos
            modelos_db = db.query(ModeloVehiculo).filter(ModeloVehiculo.activo == True).all()
            modelos_nombres = [m.modelo for m in modelos_db]
            
            # Concesionarios
            concesionarios_db = db.query(Concesionario).filter(Concesionario.activo == True).all()
            concesionarios_nombres = [c.nombre for c in concesionarios_db]
            
            # Analistas
            analistas_db = db.query(Analista).filter(Analista.activo == True).all()
            analistas_nombres = [a.nombre for a in analistas_db]
            
            # KPIs de clientes
            total_clientes = db.query(Cliente).filter(Cliente.activo == True).count()
            clientes_activos = db.query(Cliente).filter(Cliente.activo == True, Cliente.estado == 'ACTIVO').count()
            
            return {
                "modelos": modelos_nombres,
                "concesionarios": concesionarios_nombres,
                "analistas": analistas_nombres,
                "total_clientes": total_clientes,
                "clientes_activos": clientes_activos,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error obteniendo datos de BD: {e}")
            raise
    
    def _has_changes(self, current_data: Dict) -> bool:
        """Verificar si hay cambios desde la última actualización"""
        if not self.cache_data.get("last_update"):
            return True
        
        # Verificar hashes de cada tabla
        modelos_hash = self._get_data_hash(current_data["modelos"])
        concesionarios_hash = self._get_data_hash(current_data["concesionarios"])
        analistas_hash = self._get_data_hash(current_data["analistas"])
        
        # Comparar con cache
        if (modelos_hash != self.cache_data.get("modelos_hash") or
            concesionarios_hash != self.cache_data.get("concesionarios_hash") or
            analistas_hash != self.cache_data.get("analistas_hash") or
            current_data["total_clientes"] != self.cache_data.get("total_clientes")):
            
            logger.info("🔄 Cambios detectados en datos de configuración")
            logger.info(f"   Modelos: {len(current_data['modelos'])} (hash: {modelos_hash[:8]}...)")
            logger.info(f"   Concesionarios: {len(current_data['concesionarios'])} (hash: {concesionarios_hash[:8]}...)")
            logger.info(f"   Analistas: {len(current_data['analistas'])} (hash: {analistas_hash[:8]}...)")
            logger.info(f"   Clientes: {current_data['total_clientes']}")
            
            return True
        
        return False
    
    def _generate_excel_template(self, data: Dict) -> bytes:
        """Generar plantilla Excel con datos actuales"""
        try:
            # CREAR WORKBOOK
            wb = Workbook()
            
            # HOJA 1: INSTRUCCIONES
            ws_instrucciones = wb.active
            ws_instrucciones.title = "Instrucciones"
            
            instrucciones = [
                ["INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES"],
                [""],
                ["1. FORMATO DE ARCHIVO:"],
                ["   - Archivo Excel (.xlsx)"],
                ["   - Primera fila: encabezados de columnas"],
                ["   - Datos desde la segunda fila"],
                ["   - Máximo 100 registros por archivo"],
                [""],
                ["2. CAMPOS OBLIGATORIOS (marcados con *):"],
                ["   - cedula: Cédula del cliente (8-20 caracteres)"],
                ["   - nombres: Nombres del cliente (exactamente 2 palabras: nombre y apellido)"],
                ["   - apellidos: Apellidos del cliente (exactamente 2 palabras: paterno y materno)"],
                ["   - telefono: Teléfono del cliente (formato venezolano: +58 XXXXXXXXXX)"],
                ["   - email: Email del cliente (formato válido)"],
                ["   - direccion: Dirección completa del cliente"],
                ["   - fecha_nacimiento: Fecha de nacimiento (YYYY-MM-DD)"],
                ["   - ocupacion: Ocupación del cliente"],
                ["   - modelo_vehiculo: Modelo del vehículo (lista desplegable de configuración)"],
                ["   - concesionario: Concesionario (lista desplegable de configuración)"],
                ["   - analista: Analista asignado (lista desplegable de configuración)"],
                ["   - estado: Estado del cliente (ACTIVO/INACTIVO/FINALIZADO)"],
                [""],
                ["3. CAMPOS OPCIONALES:"],
                ["   - notas: Notas adicionales (si no llena, se pondrá 'NA')"],
                [""],
                ["4. VALIDACIONES:"],
                ["   - Cédula: Entre 8 y 20 caracteres"],
                ["   - Email: Formato válido usuario@dominio.com"],
                ["   - Teléfono: Formato venezolano +58 XXXXXXXXXX (10 dígitos)"],
                ["   - Fecha de nacimiento: Formato YYYY-MM-DD, no puede ser futura"],
                ["   - Nombres: Exactamente 2 palabras (nombre y apellido)"],
                ["   - Apellidos: Exactamente 2 palabras (paterno y materno)"],
                ["   - Modelo/Concesionario/Analista: COPIE EXACTAMENTE de la hoja 'Referencias'"],
                ["   - Estado: Solo ACTIVO, INACTIVO o FINALIZADO"],
                ["   - Activo: Solo TRUE o FALSE"],
                [""],
                ["5. ESTADÍSTICAS ACTUALES:"],
                [f"   - Total de clientes: {data['total_clientes']}"],
                [f"   - Clientes activos: {data['clientes_activos']}"],
                [f"   - Modelos disponibles: {len(data['modelos'])}"],
                [f"   - Concesionarios disponibles: {len(data['concesionarios'])}"],
                [f"   - Analistas disponibles: {len(data['analistas'])}"],
                [f"   - Última actualización: {data['timestamp']}"],
                [""],
                ["6. IMPORTANTE:"],
                ["   - No modifique los nombres de las columnas"],
                ["   - COPIE EXACTAMENTE los valores de la hoja 'Referencias'"],
                ["   - Todos los campos obligatorios deben estar completos"],
                ["   - La plantilla se actualiza automáticamente con los datos de configuración"],
            ]
            
            for row_data in instrucciones:
                ws_instrucciones.append(row_data)
            
            # HOJA 2: TEMPLATE VACÍO
            ws_template = wb.create_sheet("Template")
            
            # Encabezados con campos obligatorios marcados
            headers = [
                "cedula*", "nombres*", "apellidos*", "telefono*", "email*", 
                "direccion*", "fecha_nacimiento*", "ocupacion*", 
                "modelo_vehiculo*", "concesionario*", "analista*", 
                "estado*", "notas"
            ]
            ws_template.append(headers)
            
            # HOJA 3: REFERENCIAS - LISTAS PARA COPIAR Y PEGAR
            ws_referencias = wb.create_sheet("Referencias")
            
            # Título
            ws_referencias['A1'] = "REFERENCIAS DE CONFIGURACIÓN - COPIE Y PEGUE EXACTAMENTE"
            ws_referencias['A1'].font = Font(name='Calibri', size=14, bold=True)
            ws_referencias.merge_cells('A1:D1')
            
            # Modelos disponibles
            ws_referencias['A3'] = "MODELOS DE VEHÍCULOS (Columna I):"
            ws_referencias['A3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for modelo in data['modelos']:
                ws_referencias[f'A{row}'] = modelo
                row += 1
            
            # Concesionarios disponibles
            ws_referencias['B3'] = "CONCESIONARIOS (Columna J):"
            ws_referencias['B3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for concesionario in data['concesionarios']:
                ws_referencias[f'B{row}'] = concesionario
                row += 1
            
            # Analistas disponibles
            ws_referencias['C3'] = "ANALISTAS (Columna K):"
            ws_referencias['C3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for analista in data['analistas']:
                ws_referencias[f'C{row}'] = analista
                row += 1
            
            # Estados disponibles
            ws_referencias['D3'] = "ESTADOS (Columna L):"
            ws_referencias['D3'].font = Font(name='Calibri', size=12, bold=True)
            ws_referencias['D4'] = "ACTIVO"
            ws_referencias['D5'] = "INACTIVO"
            ws_referencias['D6'] = "FINALIZADO"
            
            # Ajustar ancho de columnas
            ws_referencias.column_dimensions['A'].width = 25
            ws_referencias.column_dimensions['B'].width = 25
            ws_referencias.column_dimensions['C'].width = 25
            ws_referencias.column_dimensions['D'].width = 15
            
            # GUARDAR EN BUFFER
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            raise
    
    def _save_template_to_file(self, excel_content: bytes, filename: Optional[str] = None):
        """Guardar plantilla en archivo local"""
        if not filename:
            filename = f"Plantilla_Clientes_Auto_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        
        output_dir = Path("generated_templates")
        output_dir.mkdir(exist_ok=True)
        
        file_path = output_dir / filename
        
        try:
            with open(file_path, 'wb') as f:
                f.write(excel_content)
            
            logger.info(f"✅ Plantilla guardada: {file_path}")
            logger.info(f"   Tamaño: {len(excel_content)} bytes")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error guardando archivo: {e}")
            raise
    
    def check_and_update(self, force: bool = False, save_file: bool = False) -> bool:
        """Verificar cambios y actualizar plantilla si es necesario"""
        try:
            logger.info("🔍 Verificando cambios en datos de configuración...")
            
            # Obtener datos actuales
            db = next(get_db())
            current_data = self._get_current_data(db)
            
            # Verificar si hay cambios
            if not force and not self._has_changes(current_data):
                logger.info("✅ No hay cambios en los datos de configuración")
                return False
            
            logger.info("🔄 Generando nueva plantilla Excel...")
            
            # Generar plantilla
            excel_content = self._generate_excel_template(current_data)
            
            # Guardar archivo si se solicita
            if save_file:
                self._save_template_to_file(excel_content)
            
            # Actualizar cache
            self.cache_data.update({
                "last_update": current_data["timestamp"],
                "modelos_hash": self._get_data_hash(current_data["modelos"]),
                "concesionarios_hash": self._get_data_hash(current_data["concesionarios"]),
                "analistas_hash": self._get_data_hash(current_data["analistas"]),
                "total_clientes": current_data["total_clientes"],
                "version": self.cache_data.get("version", 1) + 1
            })
            
            self._save_cache()
            
            logger.info("✅ Plantilla actualizada exitosamente")
            logger.info(f"   Versión: {self.cache_data['version']}")
            logger.info(f"   Modelos: {len(current_data['modelos'])}")
            logger.info(f"   Concesionarios: {len(current_data['concesionarios'])}")
            logger.info(f"   Analistas: {len(current_data['analistas'])}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error en actualización: {e}")
            raise
        finally:
            db.close()
    
    def watch_changes(self, interval_minutes: int = 5):
        """Monitorear cambios continuamente"""
        import time
        
        logger.info(f"👀 Iniciando monitoreo cada {interval_minutes} minutos...")
        logger.info("Presiona Ctrl+C para detener")
        
        try:
            while True:
                self.check_and_update(save_file=True)
                time.sleep(interval_minutes * 60)
                
        except KeyboardInterrupt:
            logger.info("🛑 Monitoreo detenido por el usuario")
        except Exception as e:
            logger.error(f"❌ Error en monitoreo: {e}")

def main():
    """Función principal"""
    parser = argparse.ArgumentParser(description="Actualizador automático de plantilla Excel")
    parser.add_argument("--watch", action="store_true", help="Monitorear cambios continuamente")
    parser.add_argument("--force", action="store_true", help="Forzar actualización sin verificar cambios")
    parser.add_argument("--save", action="store_true", help="Guardar plantilla en archivo local")
    parser.add_argument("--interval", type=int, default=5, help="Intervalo de monitoreo en minutos (default: 5)")
    
    args = parser.parse_args()
    
    updater = ExcelTemplateUpdater()
    
    try:
        if args.watch:
            updater.watch_changes(args.interval)
        else:
            updated = updater.check_and_update(force=args.force, save_file=args.save)
            if updated:
                print("✅ Plantilla actualizada")
            else:
                print("ℹ️ No hay cambios pendientes")
                
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
