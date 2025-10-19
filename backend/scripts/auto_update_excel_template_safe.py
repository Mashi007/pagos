#!/usr/bin/env python3
"""
Script de Actualización Automática de Plantilla Excel - VERSIÓN SEGURA
=======================================================================

Este script:
- Monitorea cambios en tablas de configuración (modelos, concesionarios, analistas)
- Regenera automáticamente la plantilla Excel cuando hay cambios
- Mantiene un cache de la última versión para evitar regeneraciones innecesarias
- NO INTERFIERE con otros procesos del sistema
- Usa conexiones de BD independientes
- Logging aislado
- Archivos en directorio separado

Uso:
    python auto_update_excel_template_safe.py [--watch] [--force] [--notify]
"""

import os
import sys
import json
import hashlib
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional
import argparse

# ✅ SEGURIDAD: Crear conexión de BD independiente
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font
import io

# ✅ SEGURIDAD: Configurar logging aislado
def setup_isolated_logging():
    """Configurar logging aislado que no interfiera con otros procesos"""
    # Crear directorio de logs específico
    log_dir = Path("excel_updater_logs")
    log_dir.mkdir(exist_ok=True)
    
    # Configurar logger específico
    logger = logging.getLogger("excel_template_updater")
    logger.setLevel(logging.INFO)
    
    # Limpiar handlers existentes para evitar conflictos
    logger.handlers.clear()
    
    # Handler para archivo
    file_handler = logging.FileHandler(
        log_dir / f"excel_updates_{datetime.now().strftime('%Y-%m-%d')}.log"
    )
    file_handler.setLevel(logging.INFO)
    
    # Handler para consola
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formato específico
    formatter = logging.Formatter(
        '%(asctime)s - [EXCEL-UPDATER] - %(levelname)s - %(message)s'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_isolated_logging()

class SafeExcelTemplateUpdater:
    """Actualizador seguro de plantilla Excel que no interfiere con otros procesos"""
    
    def __init__(self, cache_file: str = "excel_updater_cache/template_cache.json"):
        # ✅ SEGURIDAD: Directorio específico para cache
        self.cache_dir = Path("excel_updater_cache")
        self.cache_dir.mkdir(exist_ok=True)
        
        self.cache_file = self.cache_dir / cache_file
        self.cache_data = self._load_cache()
        
        # ✅ SEGURIDAD: Configuración de BD independiente
        self.db_config = self._get_db_config()
        
    def _get_db_config(self) -> Dict:
        """Obtener configuración de BD de variables de entorno"""
        return {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME', 'pagos_db'),
            'user': os.getenv('DB_USER', 'pagos_admin'),
            'password': os.getenv('DB_PASSWORD', '')
        }
    
    def _get_db_connection(self):
        """Crear conexión de BD independiente"""
        try:
            conn = psycopg2.connect(**self.db_config)
            return conn
        except Exception as e:
            logger.error(f"Error conectando a BD: {e}")
            raise
    
    def _load_cache(self) -> Dict:
        """Cargar cache de versiones anteriores"""
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.warning(f"Error cargando cache: {e}")
        return {
            "last_update": None,
            "modelos_hash": None,
            "concesionarios_hash": None,
            "analistas_hash": None,
            "total_clientes": 0,
            "version": 1
        }
    
    def _save_cache(self):
        """Guardar cache actualizado"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache_data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"Error guardando cache: {e}")
    
    def _get_data_hash(self, data: List[str]) -> str:
        """Generar hash de los datos para detectar cambios"""
        data_str = "|".join(sorted(data))
        return hashlib.md5(data_str.encode('utf-8')).hexdigest()
    
    def _get_current_data(self) -> Dict:
        """Obtener datos actuales de la BD usando conexión independiente"""
        conn = None
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # ✅ SEGURIDAD: Queries específicas y optimizadas
            # Modelos de vehículos
            cursor.execute("""
                SELECT modelo FROM modelo_vehiculos 
                WHERE activo = true 
                ORDER BY modelo
            """)
            modelos_nombres = [row['modelo'] for row in cursor.fetchall()]
            
            # Concesionarios
            cursor.execute("""
                SELECT nombre FROM concesionarios 
                WHERE activo = true 
                ORDER BY nombre
            """)
            concesionarios_nombres = [row['nombre'] for row in cursor.fetchall()]
            
            # Analistas
            cursor.execute("""
                SELECT nombre FROM analistas 
                WHERE activo = true 
                ORDER BY nombre
            """)
            analistas_nombres = [row['nombre'] for row in cursor.fetchall()]
            
            # KPIs de clientes
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_clientes,
                    COUNT(CASE WHEN estado = 'ACTIVO' THEN 1 END) as clientes_activos
                FROM clientes 
                WHERE activo = true
            """)
            kpis = cursor.fetchone()
            
            return {
                "modelos": modelos_nombres,
                "concesionarios": concesionarios_nombres,
                "analistas": analistas_nombres,
                "total_clientes": kpis['total_clientes'],
                "clientes_activos": kpis['clientes_activos'],
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"Error obteniendo datos de BD: {e}")
            raise
        finally:
            if conn:
                conn.close()
    
    def _has_changes(self, current_data: Dict) -> bool:
        """Verificar si hay cambios desde la última actualización"""
        if not self.cache_data.get("last_update"):
            return True
        
        # Verificar hashes de cada tabla
        modelos_hash = self._get_data_hash(current_data["modelos"])
        concesionarios_hash = self._get_data_hash(current_data["concesionarios"])
        analistas_hash = self._get_data_hash(current_data["analistas"])
        
        # Comparar con cache
        if (modelos_hash != self.cache_data.get("modelos_hash") or
            concesionarios_hash != self.cache_data.get("concesionarios_hash") or
            analistas_hash != self.cache_data.get("analistas_hash") or
            current_data["total_clientes"] != self.cache_data.get("total_clientes")):
            
            logger.info("🔄 Cambios detectados en datos de configuración")
            logger.info(f"   Modelos: {len(current_data['modelos'])} (hash: {modelos_hash[:8]}...)")
            logger.info(f"   Concesionarios: {len(current_data['concesionarios'])} (hash: {concesionarios_hash[:8]}...)")
            logger.info(f"   Analistas: {len(current_data['analistas'])} (hash: {analistas_hash[:8]}...)")
            logger.info(f"   Clientes: {current_data['total_clientes']}")
            
            return True
        
        return False
    
    def _generate_excel_template(self, data: Dict) -> bytes:
        """Generar plantilla Excel con datos actuales"""
        try:
            # CREAR WORKBOOK
            wb = Workbook()
            
            # HOJA 1: INSTRUCCIONES
            ws_instrucciones = wb.active
            ws_instrucciones.title = "Instrucciones"
            
            instrucciones = [
                ["INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES"],
                [""],
                ["1. FORMATO DE ARCHIVO:"],
                ["   - Archivo Excel (.xlsx)"],
                ["   - Primera fila: encabezados de columnas"],
                ["   - Datos desde la segunda fila"],
                ["   - Máximo 100 registros por archivo"],
                [""],
                ["2. CAMPOS OBLIGATORIOS (marcados con *):"],
                ["   - cedula: Cédula del cliente (8-20 caracteres)"],
                ["   - nombres: Nombres del cliente (exactamente 2 palabras: nombre y apellido)"],
                ["   - apellidos: Apellidos del cliente (exactamente 2 palabras: paterno y materno)"],
                ["   - telefono: Teléfono del cliente (formato venezolano: +58 XXXXXXXXXX)"],
                ["   - email: Email del cliente (formato válido)"],
                ["   - direccion: Dirección completa del cliente"],
                ["   - fecha_nacimiento: Fecha de nacimiento (YYYY-MM-DD)"],
                ["   - ocupacion: Ocupación del cliente"],
                ["   - modelo_vehiculo: Modelo del vehículo (lista desplegable de configuración)"],
                ["   - concesionario: Concesionario (lista desplegable de configuración)"],
                ["   - analista: Analista asignado (lista desplegable de configuración)"],
                ["   - estado: Estado del cliente (ACTIVO/INACTIVO/FINALIZADO)"],
                [""],
                ["3. CAMPOS OPCIONALES:"],
                ["   - notas: Notas adicionales (si no llena, se pondrá 'NA')"],
                [""],
                ["4. VALIDACIONES:"],
                ["   - Cédula: Entre 8 y 20 caracteres"],
                ["   - Email: Formato válido usuario@dominio.com"],
                ["   - Teléfono: Formato venezolano +58 XXXXXXXXXX (10 dígitos)"],
                ["   - Fecha de nacimiento: Formato YYYY-MM-DD, no puede ser futura"],
                ["   - Nombres: Exactamente 2 palabras (nombre y apellido)"],
                ["   - Apellidos: Exactamente 2 palabras (paterno y materno)"],
                ["   - Modelo/Concesionario/Analista: COPIE EXACTAMENTE de la hoja 'Referencias'"],
                ["   - Estado: Solo ACTIVO, INACTIVO o FINALIZADO"],
                ["   - Activo: Solo TRUE o FALSE"],
                [""],
                ["5. ESTADÍSTICAS ACTUALES:"],
                [f"   - Total de clientes: {data['total_clientes']}"],
                [f"   - Clientes activos: {data['clientes_activos']}"],
                [f"   - Modelos disponibles: {len(data['modelos'])}"],
                [f"   - Concesionarios disponibles: {len(data['concesionarios'])}"],
                [f"   - Analistas disponibles: {len(data['analistas'])}"],
                [f"   - Última actualización: {data['timestamp']}"],
                [""],
                ["6. IMPORTANTE:"],
                ["   - No modifique los nombres de las columnas"],
                ["   - COPIE EXACTAMENTE los valores de la hoja 'Referencias'"],
                ["   - Todos los campos obligatorios deben estar completos"],
                ["   - La plantilla se actualiza automáticamente con los datos de configuración"],
            ]
            
            for row_data in instrucciones:
                ws_instrucciones.append(row_data)
            
            # HOJA 2: TEMPLATE VACÍO
            ws_template = wb.create_sheet("Template")
            
            # Encabezados con campos obligatorios marcados
            headers = [
                "cedula*", "nombres*", "apellidos*", "telefono*", "email*", 
                "direccion*", "fecha_nacimiento*", "ocupacion*", 
                "modelo_vehiculo*", "concesionario*", "analista*", 
                "estado*", "notas"
            ]
            ws_template.append(headers)
            
            # HOJA 3: REFERENCIAS - LISTAS PARA COPIAR Y PEGAR
            ws_referencias = wb.create_sheet("Referencias")
            
            # Título
            ws_referencias['A1'] = "REFERENCIAS DE CONFIGURACIÓN - COPIE Y PEGUE EXACTAMENTE"
            ws_referencias['A1'].font = Font(name='Calibri', size=14, bold=True)
            ws_referencias.merge_cells('A1:D1')
            
            # Modelos disponibles
            ws_referencias['A3'] = "MODELOS DE VEHÍCULOS (Columna I):"
            ws_referencias['A3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for modelo in data['modelos']:
                ws_referencias[f'A{row}'] = modelo
                row += 1
            
            # Concesionarios disponibles
            ws_referencias['B3'] = "CONCESIONARIOS (Columna J):"
            ws_referencias['B3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for concesionario in data['concesionarios']:
                ws_referencias[f'B{row}'] = concesionario
                row += 1
            
            # Analistas disponibles
            ws_referencias['C3'] = "ANALISTAS (Columna K):"
            ws_referencias['C3'].font = Font(name='Calibri', size=12, bold=True)
            row = 4
            for analista in data['analistas']:
                ws_referencias[f'C{row}'] = analista
                row += 1
            
            # Estados disponibles
            ws_referencias['D3'] = "ESTADOS (Columna L):"
            ws_referencias['D3'].font = Font(name='Calibri', size=12, bold=True)
            ws_referencias['D4'] = "ACTIVO"
            ws_referencias['D5'] = "INACTIVO"
            ws_referencias['D6'] = "FINALIZADO"
            
            # Ajustar ancho de columnas
            ws_referencias.column_dimensions['A'].width = 25
            ws_referencias.column_dimensions['B'].width = 25
            ws_referencias.column_dimensions['C'].width = 25
            ws_referencias.column_dimensions['D'].width = 15
            
            # GUARDAR EN BUFFER
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            raise
    
    def _save_template_to_file(self, excel_content: bytes, filename: Optional[str] = None):
        """Guardar plantilla en archivo local"""
        if not filename:
            filename = f"Plantilla_Clientes_Auto_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        
        # ✅ SEGURIDAD: Directorio específico para archivos generados
        output_dir = Path("excel_updater_output")
        output_dir.mkdir(exist_ok=True)
        
        file_path = output_dir / filename
        
        try:
            with open(file_path, 'wb') as f:
                f.write(excel_content)
            
            logger.info(f"✅ Plantilla guardada: {file_path}")
            logger.info(f"   Tamaño: {len(excel_content)} bytes")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error guardando archivo: {e}")
            raise
    
    def check_and_update(self, force: bool = False, save_file: bool = False) -> bool:
        """Verificar cambios y actualizar plantilla si es necesario"""
        try:
            logger.info("🔍 Verificando cambios en datos de configuración...")
            
            # Obtener datos actuales
            current_data = self._get_current_data()
            
            # Verificar si hay cambios
            if not force and not self._has_changes(current_data):
                logger.info("✅ No hay cambios en los datos de configuración")
                return False
            
            logger.info("🔄 Generando nueva plantilla Excel...")
            
            # Generar plantilla
            excel_content = self._generate_excel_template(current_data)
            
            # Guardar archivo si se solicita
            if save_file:
                self._save_template_to_file(excel_content)
            
            # Actualizar cache
            self.cache_data.update({
                "last_update": current_data["timestamp"],
                "modelos_hash": self._get_data_hash(current_data["modelos"]),
                "concesionarios_hash": self._get_data_hash(current_data["concesionarios"]),
                "analistas_hash": self._get_data_hash(current_data["analistas"]),
                "total_clientes": current_data["total_clientes"],
                "version": self.cache_data.get("version", 1) + 1
            })
            
            self._save_cache()
            
            logger.info("✅ Plantilla actualizada exitosamente")
            logger.info(f"   Versión: {self.cache_data['version']}")
            logger.info(f"   Modelos: {len(current_data['modelos'])}")
            logger.info(f"   Concesionarios: {len(current_data['concesionarios'])}")
            logger.info(f"   Analistas: {len(current_data['analistas'])}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error en actualización: {e}")
            raise
    
    def watch_changes(self, interval_minutes: int = 5):
        """Monitorear cambios continuamente"""
        import time
        
        logger.info(f"👀 Iniciando monitoreo cada {interval_minutes} minutos...")
        logger.info("Presiona Ctrl+C para detener")
        
        try:
            while True:
                self.check_and_update(save_file=True)
                time.sleep(interval_minutes * 60)
                
        except KeyboardInterrupt:
            logger.info("🛑 Monitoreo detenido por el usuario")
        except Exception as e:
            logger.error(f"❌ Error en monitoreo: {e}")

def main():
    """Función principal"""
    parser = argparse.ArgumentParser(description="Actualizador seguro de plantilla Excel")
    parser.add_argument("--watch", action="store_true", help="Monitorear cambios continuamente")
    parser.add_argument("--force", action="store_true", help="Forzar actualización sin verificar cambios")
    parser.add_argument("--save", action="store_true", help="Guardar plantilla en archivo local")
    parser.add_argument("--interval", type=int, default=5, help="Intervalo de monitoreo en minutos (default: 5)")
    
    args = parser.parse_args()
    
    updater = SafeExcelTemplateUpdater()
    
    try:
        if args.watch:
            updater.watch_changes(args.interval)
        else:
            updated = updater.check_and_update(force=args.force, save_file=args.save)
            if updated:
                print("✅ Plantilla actualizada")
            else:
                print("ℹ️ No hay cambios pendientes")
                
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
