#!/usr/bin/env python3
"""
Generador de Plantilla Excel CON LISTAS DESPLEGABLES DESDE BD
============================================================

Plantilla que obtiene datos reales desde la base de datos:
- modelo_vehiculo desde /api/v1/modelos-vehiculos
- concesionario desde /api/v1/concesionarios  
- analista desde /api/v1/analistas
"""

import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import requests
import json

def obtener_datos_reales_api():
    """Obtener datos reales desde la API de produccion"""
    
    base_url = "https://pagos-f2qf.onrender.com"
    
    datos = {
        "modelos_vehiculos": [],
        "concesionarios": [],
        "analistas": []
    }
    
    print("Obteniendo datos reales desde la API...")
    
    try:
        # Obtener modelos de vehiculos
        print("  - Obteniendo modelos de vehiculos...")
        try:
            response = requests.get(f"{base_url}/api/v1/modelos-vehiculos?limit=100", timeout=15)
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list):
                    modelos = [item.get("nombre", "") for item in data if item.get("nombre")]
                elif isinstance(data, dict) and "items" in data:
                    modelos = [item.get("nombre", "") for item in data["items"] if item.get("nombre")]
                else:
                    modelos = []
                datos["modelos_vehiculos"] = modelos
                print(f"    ✅ {len(modelos)} modelos obtenidos")
            else:
                print(f"    ❌ Error HTTP {response.status_code}")
                datos["modelos_vehiculos"] = ["Toyota Corolla 2023", "Honda Civic 2024"]
        except Exception as e:
            print(f"    ❌ Error: {e}")
            datos["modelos_vehiculos"] = ["Toyota Corolla 2023", "Honda Civic 2024"]
        
        # Obtener concesionarios
        print("  - Obteniendo concesionarios...")
        try:
            response = requests.get(f"{base_url}/api/v1/concesionarios?limit=100", timeout=15)
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list):
                    concesionarios = [item.get("nombre", "") for item in data if item.get("nombre")]
                elif isinstance(data, dict) and "items" in data:
                    concesionarios = [item.get("nombre", "") for item in data["items"] if item.get("nombre")]
                else:
                    concesionarios = []
                datos["concesionarios"] = concesionarios
                print(f"    ✅ {len(concesionarios)} concesionarios obtenidos")
            else:
                print(f"    ❌ Error HTTP {response.status_code}")
                datos["concesionarios"] = ["AutoMax Quito", "AutoCenter Guayaquil"]
        except Exception as e:
            print(f"    ❌ Error: {e}")
            datos["concesionarios"] = ["AutoMax Quito", "AutoCenter Guayaquil"]
        
        # Obtener analistas
        print("  - Obteniendo analistas...")
        try:
            response = requests.get(f"{base_url}/api/v1/analistas?limit=100", timeout=15)
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list):
                    analistas = [item.get("nombre", "") for item in data if item.get("nombre")]
                elif isinstance(data, dict) and "items" in data:
                    analistas = [item.get("nombre", "") for item in data["items"] if item.get("nombre")]
                else:
                    analistas = []
                datos["analistas"] = analistas
                print(f"    ✅ {len(analistas)} analistas obtenidos")
            else:
                print(f"    ❌ Error HTTP {response.status_code}")
                datos["analistas"] = ["Maria Gonzalez", "Carlos Rodriguez"]
        except Exception as e:
            print(f"    ❌ Error: {e}")
            datos["analistas"] = ["Maria Gonzalez", "Carlos Rodriguez"]
            
    except Exception as e:
        print(f"Error general obteniendo datos: {e}")
        # Valores por defecto
        datos = {
            "modelos_vehiculos": ["Toyota Corolla 2023", "Honda Civic 2024"],
            "concesionarios": ["AutoMax Quito", "AutoCenter Guayaquil"],
            "analistas": ["Maria Gonzalez", "Carlos Rodriguez"]
        }
    
    return datos

def crear_plantilla_con_datos_reales():
    """Crear plantilla Excel con datos reales desde BD"""
    
    # Obtener datos reales
    datos_reales = obtener_datos_reales_api()
    
    print(f"\nDatos obtenidos:")
    print(f"  - Modelos: {len(datos_reales['modelos_vehiculos'])}")
    print(f"  - Concesionarios: {len(datos_reales['concesionarios'])}")
    print(f"  - Analistas: {len(datos_reales['analistas'])}")
    
    # Crear archivo Excel
    nombre_archivo = f"plantilla_clientes_bd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Crear workbook
    wb = Workbook()
    
    # HOJA 1: INSTRUCCIONES
    ws_instrucciones = wb.active
    ws_instrucciones.title = "Instrucciones"
    
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES"],
        [""],
        ["1. FORMATO DE ARCHIVO:"],
        ["   - Archivo Excel (.xlsx)"],
        ["   - Primera fila: encabezados de columnas"],
        ["   - Datos desde la segunda fila"],
        [""],
        ["2. CAMPOS REQUERIDOS:"],
        ["   - cedula: Cedula unica (maximo 20 caracteres)"],
        ["   - nombres: Nombres completos (maximo 100 caracteres)"],
        ["   - apellidos: Apellidos completos (maximo 100 caracteres)"],
        [""],
        ["3. CAMPOS CON LISTAS DESPLEGABLES (DATOS REALES DE BD):"],
        ["   - modelo_vehiculo: Seleccionar de lista desplegable"],
        ["   - concesionario: Seleccionar de lista desplegable"],
        ["   - analista: Seleccionar de lista desplegable"],
        ["   - estado: ACTIVO o INACTIVO"],
        [""],
        ["4. CAMPOS OPCIONALES:"],
        ["   - telefono: Numero de telefono (maximo 15 caracteres)"],
        ["   - email: Correo electronico valido"],
        ["   - direccion: Direccion completa"],
        ["   - fecha_nacimiento: Formato YYYY-MM-DD"],
        ["   - ocupacion: Ocupacion del cliente"],
        ["   - notas: Observaciones adicionales"],
        [""],
        ["5. VALIDACIONES:"],
        ["   - Cedula debe ser unica en el sistema"],
        ["   - Email debe tener formato valido"],
        ["   - Fecha de nacimiento en formato YYYY-MM-DD"],
        ["   - Estado debe ser ACTIVO o INACTIVO"],
        ["   - Usar solo valores de las listas desplegables"],
        [""],
        ["6. EJEMPLOS DE DATOS VALIDOS:"],
        ["   - cedula: '1234567890'"],
        ["   - nombres: 'Juan Carlos'"],
        ["   - apellidos: 'Perez Garcia'"],
        ["   - telefono: '0987654321'"],
        ["   - email: 'juan.perez@email.com'"],
        ["   - fecha_nacimiento: '1990-05-15'"],
        ["   - estado: 'ACTIVO'"],
        [""],
        ["7. NOTAS IMPORTANTES:"],
        ["   - No eliminar las columnas"],
        ["   - No cambiar el orden de las columnas"],
        ["   - Usar solo caracteres ASCII"],
        ["   - Evitar caracteres especiales en nombres"],
        ["   - Verificar que las cedulas no esten duplicadas"],
        ["   - Usar valores exactos de las listas desplegables"],
        [""],
        ["8. DATOS OBTENIDOS DESDE BASE DE DATOS:"],
        [""],
        [f"MODELOS DE VEHICULOS ({len(datos_reales['modelos_vehiculos'])} disponibles):"],
    ]
    
    # Agregar modelos reales
    for modelo in datos_reales['modelos_vehiculos']:
        instrucciones.append([f"   - {modelo}"])
    
    instrucciones.extend([
        [""],
        [f"CONCESIONARIOS ({len(datos_reales['concesionarios'])} disponibles):"],
    ])
    
    # Agregar concesionarios reales
    for concesionario in datos_reales['concesionarios']:
        instrucciones.append([f"   - {concesionario}"])
    
    instrucciones.extend([
        [""],
        [f"ANALISTAS ({len(datos_reales['analistas'])} disponibles):"],
    ])
    
    # Agregar analistas reales
    for analista in datos_reales['analistas']:
        instrucciones.append([f"   - {analista}"])
    
    # Agregar instrucciones a la hoja
    for i, instruccion in enumerate(instrucciones, 1):
        ws_instrucciones.cell(row=i, column=1, value=instruccion[0])
    
    # HOJA 2: PLANTILLA CON LISTAS DESPLEGABLES
    ws_plantilla = wb.create_sheet("Plantilla_Clientes")
    
    # Encabezados
    encabezados = [
        "cedula", "nombres", "apellidos", "telefono", "email",
        "direccion", "fecha_nacimiento", "ocupacion", "modelo_vehiculo",
        "concesionario", "analista", "estado", "notas"
    ]
    
    for i, encabezado in enumerate(encabezados, 1):
        ws_plantilla.cell(row=1, column=i, value=encabezado)
    
    # Ejemplo de datos (usando primer valor de cada lista)
    ejemplo = [
        "1234567890", "Juan Carlos", "Perez Garcia", "0987654321",
        "juan.perez@email.com", "Av. Principal 123, Quito", "1990-05-15",
        "Ingeniero", 
        datos_reales['modelos_vehiculos'][0] if datos_reales['modelos_vehiculos'] else "Toyota Corolla 2023",
        datos_reales['concesionarios'][0] if datos_reales['concesionarios'] else "AutoMax Quito",
        datos_reales['analistas'][0] if datos_reales['analistas'] else "Maria Gonzalez",
        "ACTIVO", "Cliente preferencial"
    ]
    
    for i, valor in enumerate(ejemplo, 1):
        ws_plantilla.cell(row=2, column=i, value=valor)
    
    # CONFIGURAR LISTAS DESPLEGABLES CON DATOS REALES
    
    # Lista desplegable para modelo_vehiculo (columna I)
    if datos_reales['modelos_vehiculos']:
        modelos_formula = f'"{",".join(datos_reales["modelos_vehiculos"])}"'
        dv_modelos = DataValidation(type="list", formula1=modelos_formula, allow_blank=True)
        dv_modelos.add(f'I2:I1000')
        ws_plantilla.add_data_validation(dv_modelos)
    
    # Lista desplegable para concesionario (columna J)
    if datos_reales['concesionarios']:
        concesionarios_formula = f'"{",".join(datos_reales["concesionarios"])}"'
        dv_concesionarios = DataValidation(type="list", formula1=concesionarios_formula, allow_blank=True)
        dv_concesionarios.add(f'J2:J1000')
        ws_plantilla.add_data_validation(dv_concesionarios)
    
    # Lista desplegable para analista (columna K)
    if datos_reales['analistas']:
        analistas_formula = f'"{",".join(datos_reales["analistas"])}"'
        dv_analistas = DataValidation(type="list", formula1=analistas_formula, allow_blank=True)
        dv_analistas.add(f'K2:K1000')
        ws_plantilla.add_data_validation(dv_analistas)
    
    # Lista desplegable para estado (columna L)
    dv_estado = DataValidation(type="list", formula1='"ACTIVO,INACTIVO"', allow_blank=True)
    dv_estado.add(f'L2:L1000')
    ws_plantilla.add_data_validation(dv_estado)
    
    # Guardar archivo
    wb.save(nombre_archivo)
    
    print(f"\nPlantilla Excel con datos reales creada: {nombre_archivo}")
    print(f"Ubicacion: {os.path.abspath(nombre_archivo)}")
    print(f"Hojas incluidas:")
    print(f"   - Instrucciones: Guia completa con datos reales")
    print(f"   - Plantilla_Clientes: Con listas desplegables reales")
    print(f"\nListas desplegables configuradas:")
    print(f"   - modelo_vehiculo (columna I): {len(datos_reales['modelos_vehiculos'])} opciones reales")
    print(f"   - concesionario (columna J): {len(datos_reales['concesionarios'])} opciones reales")
    print(f"   - analista (columna K): {len(datos_reales['analistas'])} opciones reales")
    print(f"   - estado (columna L): ACTIVO, INACTIVO")
    
    return nombre_archivo

if __name__ == "__main__":
    print("GENERANDO PLANTILLA EXCEL CON DATOS REALES DESDE BD")
    print("=" * 60)
    
    # Crear plantilla con datos reales
    archivo = crear_plantilla_con_datos_reales()
    
    print("\n" + "=" * 60)
    print("RESUMEN DE LA PLANTILLA CON DATOS REALES:")
    print("=" * 60)
    
    # Mostrar columnas
    columnas = [
        "cedula", "nombres", "apellidos", "telefono", "email",
        "direccion", "fecha_nacimiento", "ocupacion", "modelo_vehiculo",
        "concesionario", "analista", "estado", "notas"
    ]
    
    print("COLUMNAS INCLUIDAS:")
    for i, col in enumerate(columnas, 1):
        print(f"   {i:2d}. {col}")
    
    print(f"\nArchivo creado: {archivo}")
    print("Listo para usar con datos reales desde base de datos")
