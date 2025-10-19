#!/usr/bin/env python3
"""
Generador de Plantilla Excel CON LISTAS DESPLEGABLES
==================================================

Plantilla con listas desplegables reales para:
- modelo_vehiculo
- concesionario  
- analista
"""

import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows

def crear_plantilla_con_desplegables():
    """Crear plantilla Excel con listas desplegables reales"""
    
    # Datos de configuracion
    modelos_vehiculos = [
        "Toyota Corolla 2023",
        "Honda Civic 2024", 
        "Nissan Sentra 2023",
        "Hyundai Elantra 2024",
        "Mazda 3 2023",
        "Kia Forte 2024"
    ]
    
    concesionarios = [
        "AutoMax Quito",
        "AutoCenter Guayaquil", 
        "CarDealer Cuenca",
        "AutoShop Ambato",
        "MotorCity Manta",
        "AutoPlaza Loja"
    ]
    
    analistas = [
        "Maria Gonzalez",
        "Carlos Rodriguez",
        "Ana Fernandez", 
        "Luis Martinez",
        "Patricia Lopez",
        "Roberto Silva"
    ]
    
    # Crear archivo Excel
    nombre_archivo = f"plantilla_clientes_desplegables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Crear workbook
    wb = Workbook()
    
    # HOJA 1: INSTRUCCIONES
    ws_instrucciones = wb.active
    ws_instrucciones.title = "Instrucciones"
    
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES"],
        [""],
        ["1. FORMATO DE ARCHIVO:"],
        ["   - Archivo Excel (.xlsx)"],
        ["   - Primera fila: encabezados de columnas"],
        ["   - Datos desde la segunda fila"],
        [""],
        ["2. CAMPOS REQUERIDOS:"],
        ["   - cedula: Cedula unica (maximo 20 caracteres)"],
        ["   - nombres: Nombres completos (maximo 100 caracteres)"],
        ["   - apellidos: Apellidos completos (maximo 100 caracteres)"],
        [""],
        ["3. CAMPOS CON LISTAS DESPLEGABLES:"],
        ["   - modelo_vehiculo: Seleccionar de lista desplegable"],
        ["   - concesionario: Seleccionar de lista desplegable"],
        ["   - analista: Seleccionar de lista desplegable"],
        ["   - estado: ACTIVO o INACTIVO"],
        [""],
        ["4. CAMPOS OPCIONALES:"],
        ["   - telefono: Numero de telefono (maximo 15 caracteres)"],
        ["   - email: Correo electronico valido"],
        ["   - direccion: Direccion completa"],
        ["   - fecha_nacimiento: Formato YYYY-MM-DD"],
        ["   - ocupacion: Ocupacion del cliente"],
        ["   - notas: Observaciones adicionales"],
        [""],
        ["5. VALIDACIONES:"],
        ["   - Cedula debe ser unica en el sistema"],
        ["   - Email debe tener formato valido"],
        ["   - Fecha de nacimiento en formato YYYY-MM-DD"],
        ["   - Estado debe ser ACTIVO o INACTIVO"],
        ["   - Usar solo valores de las listas desplegables"],
        [""],
        ["6. EJEMPLOS DE DATOS VALIDOS:"],
        ["   - cedula: '1234567890'"],
        ["   - nombres: 'Juan Carlos'"],
        ["   - apellidos: 'Perez Garcia'"],
        ["   - telefono: '0987654321'"],
        ["   - email: 'juan.perez@email.com'"],
        ["   - fecha_nacimiento: '1990-05-15'"],
        ["   - estado: 'ACTIVO'"],
        [""],
        ["7. NOTAS IMPORTANTES:"],
        ["   - No eliminar las columnas"],
        ["   - No cambiar el orden de las columnas"],
        ["   - Usar solo caracteres ASCII"],
        ["   - Evitar caracteres especiales en nombres"],
        ["   - Verificar que las cedulas no esten duplicadas"],
        ["   - Usar valores exactos de las listas desplegables"]
    ]
    
    # Agregar instrucciones a la hoja
    for i, instruccion in enumerate(instrucciones, 1):
        ws_instrucciones.cell(row=i, column=1, value=instruccion[0])
    
    # HOJA 2: PLANTILLA CON LISTAS DESPLEGABLES
    ws_plantilla = wb.create_sheet("Plantilla_Clientes")
    
    # Encabezados
    encabezados = [
        "cedula", "nombres", "apellidos", "telefono", "email",
        "direccion", "fecha_nacimiento", "ocupacion", "modelo_vehiculo",
        "concesionario", "analista", "estado", "notas"
    ]
    
    for i, encabezado in enumerate(encabezados, 1):
        ws_plantilla.cell(row=1, column=i, value=encabezado)
    
    # Ejemplo de datos
    ejemplo = [
        "1234567890", "Juan Carlos", "Perez Garcia", "0987654321",
        "juan.perez@email.com", "Av. Principal 123, Quito", "1990-05-15",
        "Ingeniero", "Toyota Corolla 2023", "AutoMax Quito", 
        "Maria Gonzalez", "ACTIVO", "Cliente preferencial"
    ]
    
    for i, valor in enumerate(ejemplo, 1):
        ws_plantilla.cell(row=2, column=i, value=valor)
    
    # CONFIGURAR LISTAS DESPLEGABLES
    
    # Lista desplegable para modelo_vehiculo (columna I)
    modelos_formula = f'"{",".join(modelos_vehiculos)}"'
    dv_modelos = DataValidation(type="list", formula1=modelos_formula, allow_blank=True)
    dv_modelos.add(f'I2:I1000')  # Aplicar a todas las filas de datos
    ws_plantilla.add_data_validation(dv_modelos)
    
    # Lista desplegable para concesionario (columna J)
    concesionarios_formula = f'"{",".join(concesionarios)}"'
    dv_concesionarios = DataValidation(type="list", formula1=concesionarios_formula, allow_blank=True)
    dv_concesionarios.add(f'J2:J1000')  # Aplicar a todas las filas de datos
    ws_plantilla.add_data_validation(dv_concesionarios)
    
    # Lista desplegable para analista (columna K)
    analistas_formula = f'"{",".join(analistas)}"'
    dv_analistas = DataValidation(type="list", formula1=analistas_formula, allow_blank=True)
    dv_analistas.add(f'K2:K1000')  # Aplicar a todas las filas de datos
    ws_plantilla.add_data_validation(dv_analistas)
    
    # Lista desplegable para estado (columna L)
    dv_estado = DataValidation(type="list", formula1='"ACTIVO,INACTIVO"', allow_blank=True)
    dv_estado.add(f'L2:L1000')  # Aplicar a todas las filas de datos
    ws_plantilla.add_data_validation(dv_estado)
    
    # Guardar archivo
    wb.save(nombre_archivo)
    
    print(f"Plantilla Excel con listas desplegables creada: {nombre_archivo}")
    print(f"Ubicacion: {os.path.abspath(nombre_archivo)}")
    print(f"Hojas incluidas:")
    print(f"   - Instrucciones: Guia completa de uso")
    print(f"   - Plantilla_Clientes: Con listas desplegables reales")
    print(f"\nListas desplegables configuradas:")
    print(f"   - modelo_vehiculo (columna I): {len(modelos_vehiculos)} opciones")
    print(f"   - concesionario (columna J): {len(concesionarios)} opciones")
    print(f"   - analista (columna K): {len(analistas)} opciones")
    print(f"   - estado (columna L): ACTIVO, INACTIVO")
    
    return nombre_archivo

if __name__ == "__main__":
    print("GENERANDO PLANTILLA EXCEL CON LISTAS DESPLEGABLES")
    print("=" * 60)
    
    # Crear plantilla con desplegables
    archivo = crear_plantilla_con_desplegables()
    
    print("\n" + "=" * 60)
    print("RESUMEN DE LA PLANTILLA CON DESPLEGABLES:")
    print("=" * 60)
    
    # Mostrar columnas
    columnas = [
        "cedula", "nombres", "apellidos", "telefono", "email",
        "direccion", "fecha_nacimiento", "ocupacion", "modelo_vehiculo",
        "concesionario", "analista", "estado", "notas"
    ]
    
    print("COLUMNAS INCLUIDAS:")
    for i, col in enumerate(columnas, 1):
        print(f"   {i:2d}. {col}")
    
    print(f"\nArchivo creado: {archivo}")
    print("Listo para usar con listas desplegables reales")
