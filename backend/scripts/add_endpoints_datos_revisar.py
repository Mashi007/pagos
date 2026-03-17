# -*- coding: utf-8 -*-
"""Add GET datos-revisar and GET descargar-excel-errores for datos_importados_conerrores."""
path = "app/api/v1/endpoints/pagos.py"

with open(path, "r", encoding="utf-8") as f:
    c = f.read()

# Add imports
if "StreamingResponse" not in c:
    c = c.replace(
        "from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body",
        "from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body\nfrom fastapi.responses import StreamingResponse",
    )
if "from sqlalchemy import" in c and "delete" not in c:
    c = c.replace(
        "from sqlalchemy import and_, case, func, or_, select",
        "from sqlalchemy import and_, case, delete, func, or_, select",
    )

# Insert the two endpoints right after the closing of importar_reportados_aprobados_a_pagos return and before @router.post("/validar-filas-batch"
anchor = '''        "mensaje": f"Importados {registros_procesados} pagos desde Cobros; {len(ids_pagos_con_errores)} con error (revisar: descargar Excel).",
    }


@router.post("/validar-filas-batch"'''

new_endpoints = '''        "mensaje": f"Importados {registros_procesados} pagos desde Cobros; {len(ids_pagos_con_errores)} con error (revisar: descargar Excel).",
    }


@router.get("/importar-desde-cobros/datos-revisar", response_model=dict)
def get_datos_revisar_importados(
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Devuelve si hay datos que revisar (tabla datos_importados_conerrores) y el total. Para mostrar dialogo tras importar."""
    total = db.execute(select(func.count()).select_from(DatosImportadosConErrores)).scalar() or 0
    return {"tiene_datos": total > 0, "total": total}


@router.get("/importar-desde-cobros/descargar-excel-errores")
def descargar_excel_errores_importados(
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Genera Excel con registros de datos_importados_conerrores (mismas columnas que Revisar Pagos). Tras generar, vacia la tabla."""
    rows = db.execute(
        select(DatosImportadosConErrores).order_by(DatosImportadosConErrores.id)
    ).scalars().all()
    rows = [r for r in rows if r is not None]
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos con errores"
    headers = [
        "id", "cedula_cliente", "prestamo_id", "fecha_pago", "monto_pagado", "numero_documento",
        "estado", "referencia_pago", "errores_descripcion", "observaciones", "fila_origen", "referencia_interna", "created_at"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(rows, 2):
        fp = r.fecha_pago
        fecha_str = fp.strftime("%Y-%m-%d") if fp else ""
        err_str = "; ".join(r.errores_descripcion) if isinstance(r.errores_descripcion, list) else str(r.errores_descripcion or "")
        created_str = r.created_at.strftime("%Y-%m-%d %H:%M") if getattr(r, "created_at", None) else ""
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.cedula_cliente or "")
        ws.cell(row=row_idx, column=3, value=r.prestamo_id)
        ws.cell(row=row_idx, column=4, value=fecha_str)
        ws.cell(row=row_idx, column=5, value=float(r.monto_pagado) if r.monto_pagado is not None else 0)
        ws.cell(row=row_idx, column=6, value=r.numero_documento or "")
        ws.cell(row=row_idx, column=7, value=r.estado or "")
        ws.cell(row=row_idx, column=8, value=r.referencia_pago or "")
        ws.cell(row=row_idx, column=9, value=err_str)
        ws.cell(row=row_idx, column=10, value=r.observaciones or "")
        ws.cell(row=row_idx, column=11, value=r.fila_origen)
        ws.cell(row=row_idx, column=12, value=getattr(r, "referencia_interna", None) or "")
        ws.cell(row=row_idx, column=13, value=created_str)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    db.execute(delete(DatosImportadosConErrores))
    db.commit()
    from datetime import datetime as dt
    filename = f"datos_importados_con_errores_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/validar-filas-batch"'''

if anchor not in c:
    raise SystemExit("Anchor not found")
c = c.replace(anchor, new_endpoints, 1)

with open(path, "w", encoding="utf-8") as f:
    f.write(c)
print("OK: endpoints datos-revisar and descargar-excel-errores added.")
