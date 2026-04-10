"""
Endpoints para el pipeline Gmail -> Drive -> Gemini (modulo Pagos). Ejecucion solo manual (POST run-now desde la UI).
Solo correos con adjuntos (has:attachment); imagen/PDF desde cuerpo incrustado, adjuntos o .eml rfc822 (deduplicado).
Comprobantes plantilla 1 (A), 2 (B) o 3 (C Binance) con datos completos -> BD/Drive; por cada OK: etiqueta IMAGEN 1/2/3 + estrella.
Si ningun adjunto OK: sin estrella + no leido (solo con filtro unread).
- POST /pagos/gmail/run-now: ejecutar pipeline ahora
- GET /pagos/gmail/download-excel y download-excel-temporal: descargar Excel (solo lectura; no borran BD)
- GET /pagos/gmail/status: ultima ejecucion; escaneo automatico cada N h (solo pending_identification) si esta habilitado en settings
- POST /pagos/gmail/confirmar-dia: confirmacion si/no; si si, borrado de datos acumulados
"""
import io
import logging
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, delete, desc, select
from sqlalchemy.orm import Session

from app.core.database import SessionLocal, get_db
from app.core.deps import get_current_user
from app.models.pagos_gmail_sync import PagosGmailSync, PagosGmailSyncItem, GmailTemporal
from app.services.pagos_gmail.credentials import get_pagos_gmail_credentials, log_pagos_gmail_config_status
from app.services.pagos_gmail.helpers import format_monto_excel_pagos_gmail, formatear_cedula
from app.services.pagos_gmail.pipeline import run_pipeline

logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])


def _is_pipeline_running(db: Session) -> bool:
    """True si hay una sync en estado running. Solo se ignora si lleva >2h (huérfano por crash)."""
    # Ventana larga (2 h): evita que un pipeline legítimo deje de contar como "running"
    # y permita lanzar otro (doble ejecución). Huérfanos por crash se ignoran tras 2 h.
    cutoff = datetime.utcnow() - timedelta(hours=2)
    row = db.execute(
        select(PagosGmailSync).where(
            and_(
                PagosGmailSync.status == "running",
                PagosGmailSync.started_at >= cutoff,
            )
        ).limit(1)
    ).scalars().first()
    return row is not None


def _run_pipeline_background(sync_id: int, scan_filter: str = "all") -> None:
    """Ejecuta el pipeline en background con su propia sesion de BD. ón de BD (evita el timeout de 30s de Render/Axios)."""
    db = SessionLocal()
    logger.info("[PAGOS_GMAIL] [ETAPA] Inicio pipeline background sync_id=%s scan_filter=%s", sync_id, scan_filter)
    try:
        run_pipeline(db, existing_sync_id=sync_id, scan_filter=scan_filter)
    except Exception as e:
        logger.info("[PAGOS_GMAIL] [ETAPA] Pipeline finalizado sync_id=%s", sync_id)
        logger.exception("[PAGOS_GMAIL] [ETAPA] Error en background pipeline (sync_id=%s): %s", sync_id, e)
        try:
            sync = db.execute(select(PagosGmailSync).where(PagosGmailSync.id == sync_id)).scalars().first()
            if sync and sync.status == "running":
                sync.status = "error"
                sync.finished_at = datetime.utcnow()
                sync.error_message = str(e)[:2000]
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.get("/count-pending")
def count_pending(
    scan_filter: str = "all",
    db: Session = Depends(get_db),
):
    """
    Cuenta cuantos correos se procesarian sin iniciar el pipeline.
    El frontend puede mostrar "Se procesaran N correos. Iniciar? Si / No" y solo llamar
    POST /run-now si el usuario confirma (Si = inicia, No = no hace nada).
    scan_filter: "unread" | "read" | "all" | "pending_identification" (mismo que run-now; por defecto all).
    unread/read/all listan los mismos hilos (inbox + imagen/PDF, leidos y no leidos).
    Excluye correos con etiquetas de clasificacion del pipeline (MERCANTIL/BNC/BINANCE/BNV/MASTER/ERROR EMAIL/MANUAL; legado IMAGEN 5).
    """
    creds = get_pagos_gmail_credentials()
    if not creds:
        return {"count": 0, "scan_filter": scan_filter, "error": "no_credentials"}
    from app.services.pagos_gmail.gmail_service import build_gmail_service, count_messages_by_filter
    if scan_filter not in ("unread", "read", "all", "pending_identification"):
        scan_filter = "all"
    try:
        gmail_svc = build_gmail_service(creds)
        count = count_messages_by_filter(gmail_svc, scan_filter)
        return {"count": count, "scan_filter": scan_filter}
    except Exception as e:
        logger.warning("[PAGOS_GMAIL] count-pending error: %s", e)
        return {"count": 0, "scan_filter": scan_filter, "error": str(e)[:200]}


@router.post("/run-now")
def run_now(
    background_tasks: BackgroundTasks,
    force: bool = True,
    scan_filter: str = "all",
    db: Session = Depends(get_db),
):
    """
    Inicia el pipeline en segundo plano (Gmail -> Drive -> Gemini -> BD) y devuelve inmediatamente.
    Solo correos con adjuntos; candidatos imagen/PDF: incrustados, adjuntos y reenvios rfc822.
    scan_filter: "unread" | "read" | "all" | "pending_identification" (por defecto all).
    Listado: inbox con imagen/PDF, leidos y no leidos (unread/read/all equivalentes); sin ya etiquetados (MERCANTIL/BNC/BINANCE/BNV/MASTER/ERROR EMAIL/MANUAL; legado IMAGEN 5).
    Procesamiento en orden de fecha del correo: mas antiguo primero, mas reciente al final.
    pending_identification: ademas sin estrella (cola de pendientes).
    El frontend debe hacer polling a GET /status hasta que last_status sea 'success' o 'error'.
    El parametro force se mantiene por compatibilidad y no aplica ninguna restriccion.
    """
    _ = force
    if _is_pipeline_running(db):
        raise HTTPException(
            status_code=409,
            detail="Ya hay una sincronización en curso. Espere unos minutos.",
        )
    # Verificar credenciales de forma síncrona (respuesta inmediata si fallan)
    creds = get_pagos_gmail_credentials()
    if not creds:
        log_pagos_gmail_config_status()
        raise HTTPException(
            status_code=503,
            detail=(
                "Credenciales OAuth inválidas (Google devolvió 'invalid_client'). "
                "Vaya a Configuración > Informe de pagos: pegue el mismo Client ID y Client secret que en Google Cloud, guarde, "
                "y luego pulse «Conectar con Google» para obtener un nuevo token. Sin reconectar, el token antiguo no funciona con el secret nuevo."
            ),
        )
    # Crear registro de sync de inmediato (evita que un segundo click arranque otro pipeline)
    sync = PagosGmailSync(status="running", emails_processed=0, files_processed=0)
    db.add(sync)
    db.commit()
    db.refresh(sync)
    sync_id = sync.id
    # Validar scan_filter
    if scan_filter not in ("unread", "read", "all", "pending_identification"):
        scan_filter = "all"
    # Lanzar pipeline en segundo plano; el cliente hace polling a /status
    background_tasks.add_task(_run_pipeline_background, sync_id, scan_filter)
    return {
        "sync_id": sync_id,
        "status": "running",
        "emails_processed": 0,
        "files_processed": 0,
    }


@router.get("/status")
def status(db: Session = Depends(get_db)):
    """Ultima ejecucion manual; sin proxima ejecucion programada (pipeline solo desde la UI)."""
    last = db.execute(select(PagosGmailSync).order_by(desc(PagosGmailSync.started_at)).limit(1)).scalars().first()
    latest_data_date = _get_latest_date_with_data(db)
    marcados = 0
    if last is not None:
        marcados = int(getattr(last, "correos_marcados_revision", 0) or 0)
    return {
        "last_run": last.started_at.isoformat() if last and last.started_at else None,
        "last_status": last.status if last else None,
        "last_emails": last.emails_processed if last else 0,
        "last_files": last.files_processed if last else 0,
        "last_error": last.error_message if last and last.status == "error" else None,
        "next_run_approx": None,
        "latest_data_date": latest_data_date,
        "last_correos_marcados_revision": marcados,
    }


def _get_sheet_date_for_download() -> datetime:
    """Después de las 23:50 (America/Caracas) se considera el día siguiente."""
    from zoneinfo import ZoneInfo
    tz = ZoneInfo("America/Caracas")
    now = datetime.now(tz)
    base = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if now.hour == 23 and now.minute >= 50:
        base += timedelta(days=1)
    return base.replace(tzinfo=None)


def _sheet_date_from_fecha(fecha: Optional[str]) -> datetime:
    """Convierte fecha opcional 'YYYY-MM-DD' a datetime (00:00:00); si no, usa lógica del día actual."""
    if not fecha or not fecha.strip():
        return _get_sheet_date_for_download()
    try:
        return datetime.strptime(fecha.strip()[:10], "%Y-%m-%d")
    except ValueError:
        return _get_sheet_date_for_download()


class ConfirmarDiaBody(BaseModel):
    """Respuesta simple sí/no: si confirmado=true se borran los datos del día."""
    confirmado: bool
    fecha: Optional[str] = None  # YYYY-MM-DD; si no se envía, se usa el día actual (misma lógica que download)


@router.post("/confirmar-dia", response_model=dict)
def confirmar_dia(body: ConfirmarDiaBody = Body(...), db: Session = Depends(get_db)):
    """
    Confirmación desde la interfaz (sí/no).
    - Sin fecha: borra TODOS los ítems (limpiar el acumulado completo tras exportar).
    - Con fecha específica: borra solo los ítems de esa fecha (sheet_name exacto).
    Si confirmado=false no se hace nada.
    """
    if not body.confirmado:
        return {"confirmado": False, "mensaje": "Sin cambios. Los datos del día se mantienen."}
    if body.fecha and body.fecha.strip():
        from app.services.pagos_gmail.helpers import get_sheet_name_for_date
        sheet_date = _sheet_date_from_fecha(body.fecha)
        sheet_name = get_sheet_name_for_date(sheet_date)
        result = db.execute(delete(PagosGmailSyncItem).where(PagosGmailSyncItem.sheet_name == sheet_name))
        db.commit()
        deleted = result.rowcount if hasattr(result, "rowcount") else 0
        logger.info("Pagos Gmail confirmar-dia: borrados %d ítems de sheet_name=%s", deleted, sheet_name)
        return {
            "confirmado": True,
            "borrados": deleted,
            "mensaje": f"Datos de {sheet_date.strftime('%Y-%m-%d')} borrados ({deleted} filas)." if deleted else f"No había datos para {sheet_date.strftime('%Y-%m-%d')}.",
        }
    else:
        # Sin fecha: borrar todo el acumulado
        result = db.execute(delete(PagosGmailSyncItem))
        db.execute(delete(GmailTemporal))
        db.commit()
        deleted = result.rowcount if hasattr(result, "rowcount") else 0
        logger.info("Pagos Gmail confirmar-dia: borrados TODOS los ítems (%d)", deleted)
        return {
            "confirmado": True,
            "borrados": deleted,
            "mensaje": f"Acumulado completo borrado ({deleted} filas). Listo para el próximo ciclo.",
        }


def _find_most_recent_data(db: Session) -> tuple[Optional[str], Optional[datetime], list]:
    """
    Devuelve los ítems más recientes (acumulado), ordenados por fecha de creación ascendente.
    Por defecto incluye toda la bandeja (hasta 500k items). Opcional: PAGOS_GMAIL_DOWNLOAD_EXCEL_MAX_ITEMS.
    Devuelve (sheet_name_ref, email_date_ref, items) o (None, None, []).
    """
    from app.core.config import settings
    from app.services.pagos_gmail.helpers import parse_date_from_sheet_name
    max_items = getattr(settings, "PAGOS_GMAIL_DOWNLOAD_EXCEL_MAX_ITEMS", 0) or 0
    if max_items <= 0:
        max_items = 500000  # toda la bandeja en un solo archivo
    logger.info("[PAGOS_GMAIL] [ETAPA] _find_most_recent_data max_items=%s", max_items)
    rows = db.execute(
        select(PagosGmailSyncItem)
        .order_by(desc(PagosGmailSyncItem.created_at))
        .limit(max_items)
    ).scalars().all()
    items = list(reversed(rows))
    logger.info("[PAGOS_GMAIL] [ETAPA] _find_most_recent_data items_en_bd=%s devueltos=%s", len(rows), len(items))  # más recientes primero en el límite, pero orden ascendente en Excel
    if not items:
        return None, None, []
    last_sheet_name = items[-1].sheet_name or ""
    email_date = parse_date_from_sheet_name(last_sheet_name)
    return last_sheet_name, email_date, items


def _find_sheet_by_fecha(db: Session, fecha_date: datetime) -> tuple[Optional[str], list]:
    """Busca items para una fecha concreta (sheet_name exacto). Devuelve (sheet_name, items)."""
    from app.services.pagos_gmail.helpers import get_sheet_name_for_date
    sheet_name = get_sheet_name_for_date(fecha_date)
    items = db.execute(
        select(PagosGmailSyncItem)
        .where(PagosGmailSyncItem.sheet_name == sheet_name)
        .order_by(PagosGmailSyncItem.created_at)
    ).scalars().all()
    return (sheet_name, list(items)) if items else (sheet_name, [])


def _get_latest_date_with_data(db: Session) -> Optional[str]:
    """Devuelve la fecha del correo (YYYY-MM-DD) del ítem más reciente en BD, para guiar al usuario."""
    from app.services.pagos_gmail.helpers import parse_date_from_sheet_name
    row = db.execute(
        select(PagosGmailSyncItem.sheet_name)
        .order_by(desc(PagosGmailSyncItem.created_at))
        .limit(1)
    ).scalars().first()
    if not row:
        return None
    d = parse_date_from_sheet_name(row)
    return d.strftime("%Y-%m-%d") if d else None


@router.get("/download-excel")
def download_excel(fecha: Optional[str] = None, db: Session = Depends(get_db)):
    """
    Genera y devuelve un Excel con los ítems procesados (solo lectura en BD).
    No borra ni modifica filas: el acumulado sigue en el servidor para siguientes descargas y nuevos procesamientos.
    - Sin ?fecha: descarga el lote más reciente en BD, sin importar la fecha del correo
      (cubre backlog de cualquier antigüedad; los correos se procesan mientras estén no leídos).
    - Con ?fecha=YYYY-MM-DD: descarga exactamente esa fecha.
    Si no hay datos devuelve 404 (no se genera Excel vacío).
    Columnas A-E: Banco, Cedula, Fecha, Monto, Serial documento; luego Correo Pagador, Link, Ver email.
    Para vaciar tablas usar POST /pagos/gmail/confirmar-dia con confirmado=true.
    """
    from openpyxl import Workbook
    items: list = []
    sheet_date: Optional[datetime] = None
    logger.info("[PAGOS_GMAIL] [ETAPA] download-excel inicio fecha=%s", fecha or "(sin fecha = toda la bandeja)")

    if fecha and fecha.strip():
        # Fecha explícita: buscar ese día concreto
        fecha_date = _sheet_date_from_fecha(fecha)
        _, items = _find_sheet_by_fecha(db, fecha_date)
        logger.info("[PAGOS_GMAIL] [ETAPA] download-excel por fecha items=%s", len(items))
        sheet_date = fecha_date
    else:
        # Sin fecha: lote más reciente en BD (backlog de cualquier antigüedad)
        logger.info("[PAGOS_GMAIL] [ETAPA] download-excel sin fecha items=%s", len(items))
        _, sheet_date, items = _find_most_recent_data(db)

    if not items:
        raise HTTPException(
            status_code=404,
            detail=(
                "Sin datos disponibles. "
                "Pulse «Generar Excel desde Gmail» para procesar correos no leídos con adjuntos (imagen/PDF) "
                "y vuelva a descargar. Verifique que GEMINI_API_KEY esté configurado en el servidor."
                + (f" (buscado: {fecha})" if fecha else "")
            ),
        )

    from openpyxl.styles import Font

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append(
            [
                "Banco",
                "Cedula",
                "Fecha",
                "Monto",
                "Serial documento",
                "Correo Pagador",
                "Link",
                "Ver email",
            ]
        )
        link_font = Font(color="0563C1", underline="single")
        for row_idx, it in enumerate(items, start=2):  # fila 1 = cabecera
            link_url = (it.drive_link or "").strip()
            if link_url and not link_url.startswith("http"):
                link_url = "https://drive.google.com/file/d/" + link_url + "/view"
            link_text = link_url or ""
            email_url = (it.drive_email_link or "").strip()
            if email_url and not email_url.startswith("http"):
                email_url = "https://drive.google.com/file/d/" + email_url + "/view"
            email_text = email_url or ("—" if link_url else "")
            ws.append(
                [
                    it.banco or "",
                    formatear_cedula(it.cedula or ""),
                    it.fecha_pago or "",
                    format_monto_excel_pagos_gmail(it.monto) or (it.monto or ""),
                    it.numero_referencia or "",
                    it.correo_origen or "",
                    link_text,
                    email_text,
                ]
            )
            if link_url:
                c_link = ws.cell(row=row_idx, column=7)
                c_link.hyperlink = link_url
                c_link.value = link_url
                c_link.font = link_font
            if email_url:
                c_eml = ws.cell(row=row_idx, column=8)
                c_eml.hyperlink = email_url
                c_eml.value = email_url
                c_eml.font = link_font
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        logger.exception("[PAGOS_GMAIL] Error generando Excel: %s", e)
        raise HTTPException(
            status_code=500,
            detail="Error al generar el archivo Excel. Intente de nuevo o contacte soporte.",
        ) from e

    date_str = sheet_date.strftime("%Y-%m-%d") if sheet_date else "sin-fecha"
    filename = f"Pagos_Gmail_{date_str}.xlsx"
    logger.info("[PAGOS_GMAIL] [ETAPA] download-excel OK filas=%s filename=%s", len(items), filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



@router.get("/download-excel-temporal")
def download_excel_temporal(db: Session = Depends(get_db)):
    """
    Genera Excel desde la tabla temporal gmail_temporal (cada procesamiento Gmail inserta a continuacion).
    NO vacia la tabla: los datos solo se borran al usar el boton "Vaciar tabla (Generar Excel desde Gmail)". Si no hay datos devuelve 404.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    items = db.execute(
        select(GmailTemporal).order_by(GmailTemporal.created_at)
    ).scalars().all()
    items = list(items)
    if not items:
        raise HTTPException(
            status_code=404,
            detail="Sin datos en tabla temporal. Procese correos Gmail primero; cada procesamiento se almacena a continuacion en gmail_temporal.",
        )
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append(
            [
                "Banco",
                "Cedula",
                "Fecha",
                "Monto",
                "Serial documento",
                "Correo Pagador",
                "Link",
                "Ver email",
            ]
        )
        link_font = Font(color="0563C1", underline="single")
        for row_idx, it in enumerate(items, start=2):
            link_url = (it.drive_link or "").strip()
            if link_url and not link_url.startswith("http"):
                link_url = "https://drive.google.com/file/d/" + link_url + "/view"
            link_text = link_url or ""
            email_url = (it.drive_email_link or "").strip()
            if email_url and not email_url.startswith("http"):
                email_url = "https://drive.google.com/file/d/" + email_url + "/view"
            email_text = email_url or ("—" if link_url else "")
            ws.append(
                [
                    it.banco or "",
                    formatear_cedula(it.cedula or ""),
                    it.fecha_pago or "",
                    format_monto_excel_pagos_gmail(it.monto) or (it.monto or ""),
                    it.numero_referencia or "",
                    it.correo_origen or "",
                    link_text,
                    email_text,
                ]
            )
            if link_url:
                c_link = ws.cell(row=row_idx, column=7)
                c_link.hyperlink = link_url
                c_link.value = link_url
                c_link.font = link_font
            if email_url:
                c_eml = ws.cell(row=row_idx, column=8)
                c_eml.hyperlink = email_url
                c_eml.value = email_url
                c_eml.font = link_font
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        logger.exception("[PAGOS_GMAIL] Error generando Excel temporal: %s", e)
        raise HTTPException(status_code=500, detail="Error al generar Excel.") from e

    # No se vacia la BD aqui: solo el boton Vaciar tabla puede borrar.
    from datetime import datetime as dt
    date_str = dt.utcnow().strftime("%Y-%m-%d")
    filename = f"Pagos_Gmail_temporal_{date_str}.xlsx"
    logger.info("[PAGOS_GMAIL] download-excel-temporal OK filas=%s (tabla no se vacia)", len(items))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

@router.get("/diagnostico")
def diagnostico(db: Session = Depends(get_db)):
    """
    Diagnóstico integral del pipeline: verifica credenciales, Gmail, Drive, Sheets y Gemini
    con 1 correo real. Devuelve JSON detallado con el resultado de cada paso.
    """
    from app.services.pagos_gmail.credentials import get_pagos_gmail_credentials
    from app.services.pagos_gmail.gmail_service import (
        build_gmail_service,
        list_messages_by_filter,
        get_message_full_payload,
        get_pagos_gmail_image_pdf_files_for_pipeline,
    )
    from app.services.pagos_gmail.drive_service import build_drive_service
    from app.services.pagos_gmail.gemini_service import (
        classify_and_extract_pagos_gmail_attachment,
        check_gemini_available,
    )
    from app.core.config import settings
    import traceback

    result: dict = {
        "paso_1_credenciales": None,
        "paso_2_gmail_list": None,
        "paso_3_primer_correo": None,
        "paso_4_imagenes": None,
        "paso_5_gemini_health": None,
        "paso_6_gemini_extraccion": None,
        "config": {
            "GEMINI_API_KEY_set": bool(getattr(settings, "GEMINI_API_KEY", None)),
            "GEMINI_MODEL": getattr(settings, "GEMINI_MODEL", "gemini-2.5-flash"),
        }
    }

    # PASO 1: Credenciales
    try:
        creds = get_pagos_gmail_credentials()
        if creds:
            result["paso_1_credenciales"] = {"ok": True, "tipo": str(type(creds).__name__)}
        else:
            result["paso_1_credenciales"] = {"ok": False, "error": "get_pagos_gmail_credentials() devolvió None"}
            return result
    except Exception as e:
        result["paso_1_credenciales"] = {"ok": False, "error": str(e), "trace": traceback.format_exc()[-500:]}
        return result

    # PASO 2: Listar correos no leídos
    try:
        gmail_svc = build_gmail_service(creds)
        messages = list_messages_by_filter(gmail_svc, "unread")
        result["paso_2_gmail_list"] = {
            "ok": True,
            "total_no_leidos": len(messages),
            "primeros_3": [
                {"id": m["id"], "from": m["headers"].get("from", "")[:60],
                 "subject": m["headers"].get("subject", "")[:60]}
                for m in messages[:3]
            ]
        }
        if not messages:
            result["paso_2_gmail_list"]["advertencia"] = "No hay correos no leídos — nada que procesar"
            return result
    except Exception as e:
        result["paso_2_gmail_list"] = {"ok": False, "error": str(e), "trace": traceback.format_exc()[-500:]}
        return result

    # PASO 3: Primer correo — payload completo
    try:
        msg = messages[0]
        full_payload = get_message_full_payload(gmail_svc, msg["id"])
        mime_top = full_payload.get("mimeType", "desconocido") if full_payload else "payload vacío"
        parts_top = [p.get("mimeType", "?") for p in (full_payload or {}).get("parts", [])]
        result["paso_3_primer_correo"] = {
            "ok": True,
            "msg_id": msg["id"],
            "from": msg["headers"].get("from", "")[:80],
            "subject": msg["headers"].get("subject", "")[:80],
            "mimeType_top": mime_top,
            "parts_top_level": parts_top,
        }
    except Exception as e:
        result["paso_3_primer_correo"] = {"ok": False, "error": str(e), "trace": traceback.format_exc()[-500:]}
        return result

    # PASO 4: Extracción de imágenes del primer correo
    try:
        attachments = get_pagos_gmail_image_pdf_files_for_pipeline(
            gmail_svc, msg["id"], full_payload
        )
        result["paso_4_imagenes"] = {
            "ok": True,
            "nota": "Cuerpo incrustado + adjuntos imagen/PDF + message/rfc822 (.eml), deduplicado por contenido",
            "total_imagenes": len(attachments),
            "detalle": [
                {"nombre": f, "bytes": len(c), "mime": m}
                for f, c, m in attachments[:5]
            ]
        }
    except Exception as e:
        result["paso_4_imagenes"] = {"ok": False, "error": str(e), "trace": traceback.format_exc()[-500:]}
        return result

    # PASO 5: Health check de Gemini (prompt de texto simple)
    try:
        gemini_health = check_gemini_available()
        result["paso_5_gemini_health"] = gemini_health
    except Exception as e:
        result["paso_5_gemini_health"] = {"ok": False, "error": str(e)}

    # PASO 6: Extracción real con Gemini sobre la primera imagen (si existe)
    if attachments:
        fname, content, mime = attachments[0]
        try:
            from_hdr = (msg.get("headers") or {}).get("from") or ""
            fmt, data = classify_and_extract_pagos_gmail_attachment(
                content, fname, remitente_correo_header=from_hdr
            )
            result["paso_6_gemini_extraccion"] = {
                "ok": True,
                "archivo": fname,
                "bytes": len(content),
                "mime": mime,
                "formato": fmt,
                "resultado": data,
            }
        except Exception as e:
            result["paso_6_gemini_extraccion"] = {
                "ok": False,
                "archivo": fname,
                "bytes": len(content),
                "error": str(e),
                "trace": traceback.format_exc()[-500:],
            }
    else:
        result["paso_6_gemini_extraccion"] = {
            "ok": False,
            "motivo": "Sin imágenes >= 10KB en el primer correo — Gemini no fue llamado",
        }

    return result
