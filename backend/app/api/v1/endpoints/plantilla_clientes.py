"""
Endpoint para generar plantilla Excel dinÃ¡mica con datos reales de BD
====================================================================

Este endpoint:
- Obtiene datos reales desde las tablas de configuraciÃ³n
- Genera plantilla Excel con listas desplegables actualizadas
- Se actualiza automÃ¡ticamente cuando admin cambia las listas
- Campos obligatorios y validaciones completas
"""

from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from app.db.session import get_db
from app.models.user import User
from app.api.deps import get_current_user
from app.models.modelo_vehiculo import ModeloVehiculo
from app.models.concesionario import Concesionario
from app.models.analista import Analista
from app.models.cliente import Cliente
import logging
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import io

router = APIRouter()
logger = logging.getLogger(__name__)

@router.get("/plantilla-clientes")
async def generar_plantilla_clientes_dinamica(
    response: Response,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    ðŸ“Š Generar plantilla Excel dinÃ¡mica con datos reales de BD
    
    CaracterÃ­sticas:
    - Obtiene datos reales desde tablas de configuraciÃ³n
    - Listas desplegables con valores actuales
    - Se actualiza automÃ¡ticamente cuando admin cambia listas
    - Instrucciones completas incluidas
    - Campos obligatorios y validaciones
    """
    try:
        logger.info(f"Generando plantilla dinÃ¡mica - Usuario: {current_user.email}")
        
        # OBTENER DATOS REALES DESDE BD
        
        # Modelos de vehÃ­culos
        modelos_db = db.query(ModeloVehiculo).filter(ModeloVehiculo.activo == True).all()
        modelos_nombres = [m.nombre for m in modelos_db]
        logger.info(f"Modelos encontrados: {len(modelos_nombres)}")
        
        # Concesionarios
        concesionarios_db = db.query(Concesionario).filter(Concesionario.activo == True).all()
        concesionarios_nombres = [c.nombre for c in concesionarios_db]
        logger.info(f"Concesionarios encontrados: {len(concesionarios_nombres)}")
        
        # Analistas
        analistas_db = db.query(Analista).filter(Analista.activo == True).all()
        analistas_nombres = [a.nombre for a in analistas_db]
        logger.info(f"Analistas encontrados: {len(analistas_nombres)}")
        
        # KPIs de clientes
        total_clientes = db.query(Cliente).filter(Cliente.activo == True).count()
        clientes_activos = db.query(Cliente).filter(Cliente.activo == True, Cliente.estado == 'ACTIVO').count()
        
        # CREAR WORKBOOK
        wb = Workbook()
        
        # HOJA 1: INSTRUCCIONES
        ws_instrucciones = wb.active
        ws_instrucciones.title = "Instrucciones"
        
        instrucciones = [
            ["INSTRUCCIONES PARA CARGA MASIVA DE CLIENTES"],
            [""],
            ["1. FORMATO DE ARCHIVO:"],
            ["   - Archivo Excel (.xlsx)"],
            ["   - Primera fila: encabezados de columnas"],
            ["   - Datos desde la segunda fila"],
            ["   - MÃ¡ximo 100 registros por archivo"],
            [""],
            ["2. CAMPOS OBLIGATORIOS (marcados con *):"],
            ["   - cedula: CÃ©dula del cliente (8-20 caracteres)"],
            ["   - nombres: Nombres del cliente (exactamente 2 palabras: nombre y apellido)"],
            ["   - apellidos: Apellidos del cliente (exactamente 2 palabras: paterno y materno)"],
            ["   - telefono: TelÃ©fono del cliente (formato venezolano: +58 XXXXXXXXXX)"],
            ["   - email: Email del cliente (formato vÃ¡lido)"],
            ["   - direccion: DirecciÃ³n completa del cliente"],
            ["   - fecha_nacimiento: Fecha de nacimiento (YYYY-MM-DD)"],
            ["   - ocupacion: OcupaciÃ³n del cliente"],
            ["   - modelo_vehiculo: Modelo del vehÃ­culo (lista desplegable de configuraciÃ³n)"],
            ["   - concesionario: Concesionario (lista desplegable de configuraciÃ³n)"],
            ["   - analista: Analista asignado (lista desplegable de configuraciÃ³n)"],
            ["   - estado: Estado del cliente (ACTIVO/INACTIVO/FINALIZADO)"],
            [""],
            ["3. CAMPOS OPCIONALES:"],
            ["   - notas: Notas adicionales (si no llena, se pondrÃ¡ 'NA')"],
            [""],
            ["4. VALIDACIONES:"],
            ["   - CÃ©dula: Debe ser Ãºnica en el sistema (8-20 caracteres)"],
            ["   - Email: Debe tener formato vÃ¡lido usuario@dominio.com"],
            ["   - TelÃ©fono: Formato venezolano +58 XXXXXXXXXX (10 dÃ­gitos, primer dÃ­gito no puede ser 0)"],
            ["   - Fecha de nacimiento: No puede ser futura"],
            ["   - Nombres: Exactamente 2 palabras (nombre y apellido)"],
            ["   - Apellidos: Exactamente 2 palabras (paterno y materno)"],
            ["   - Listas desplegables: Solo valores de configuraciÃ³n actualizados"],
            [""],
            ["5. ESTADÃSTICAS ACTUALES:"],
            [f"   - Total de clientes: {total_clientes}"],
            [f"   - Clientes activos: {clientes_activos}"],
            [f"   - Modelos disponibles: {len(modelos_nombres)}"],
            [f"   - Concesionarios disponibles: {len(concesionarios_nombres)}"],
            [f"   - Analistas disponibles: {len(analistas_nombres)}"],
            [""],
            ["6. IMPORTANTE:"],
            ["   - No modifique los nombres de las columnas"],
            ["   - Use las listas desplegables para evitar errores"],
            ["   - Todos los campos obligatorios deben estar completos"],
            ["   - La plantilla se actualiza automÃ¡ticamente con los datos de configuraciÃ³n"],
        ]
        
        for row_data in instrucciones:
            ws_instrucciones.append(row_data)
        
        # HOJA 2: TEMPLATE VACÃO
        ws_template = wb.create_sheet("Template")
        
        # Encabezados con campos obligatorios marcados
        headers = [
            "cedula*", "nombres*", "apellidos*", "telefono*", "email*", 
            "direccion*", "fecha_nacimiento*", "ocupacion*", 
            "modelo_vehiculo*", "concesionario*", "analista*", 
            "estado*", "notas"
        ]
        ws_template.append(headers)
        
        # VALIDACIONES DE DATOS
        
        # ValidaciÃ³n para modelo_vehiculo (columna I)
        if modelos_nombres:
            dv_modelo = DataValidation(
                type="list",
                formula1=f'"{",".join(modelos_nombres)}"',
                showDropDown=True
            )
            dv_modelo.error = "Seleccione un modelo vÃ¡lido de la lista"
            dv_modelo.errorTitle = "Modelo invÃ¡lido"
            dv_modelo.add('I2:I101')  # Aplicar a 100 filas
            ws_template.add_data_validation(dv_modelo)
        
        # ValidaciÃ³n para concesionario (columna J)
        if concesionarios_nombres:
            dv_concesionario = DataValidation(
                type="list",
                formula1=f'"{",".join(concesionarios_nombres)}"',
                showDropDown=True
            )
            dv_concesionario.error = "Seleccione un concesionario vÃ¡lido de la lista"
            dv_concesionario.errorTitle = "Concesionario invÃ¡lido"
            dv_concesionario.add('J2:J101')  # Aplicar a 100 filas
            ws_template.add_data_validation(dv_concesionario)
        
        # ValidaciÃ³n para analista (columna K)
        if analistas_nombres:
            dv_analista = DataValidation(
                type="list",
                formula1=f'"{",".join(analistas_nombres)}"',
                showDropDown=True
            )
            dv_analista.error = "Seleccione un analista vÃ¡lido de la lista"
            dv_analista.errorTitle = "Analista invÃ¡lido"
            dv_analista.add('K2:K101')  # Aplicar a 100 filas
            ws_template.add_data_validation(dv_analista)
        
        # ValidaciÃ³n para estado (columna L)
        dv_estado = DataValidation(
            type="list",
            formula1='"ACTIVO,INACTIVO,FINALIZADO"',
            showDropDown=True
        )
        dv_estado.error = "Seleccione un estado vÃ¡lido: ACTIVO, INACTIVO o FINALIZADO"
        dv_estado.errorTitle = "Estado invÃ¡lido"
        dv_estado.add('L2:L101')  # Aplicar a 100 filas
        ws_template.add_data_validation(dv_estado)
        
        # ValidaciÃ³n para fecha de nacimiento (columna G)
        dv_fecha = DataValidation(
            type="date",
            operator="lessThanOrEqual",
            formula1=f'DATE({datetime.now().year},{datetime.now().month},{datetime.now().day})',
            showDropDown=True
        )
        dv_fecha.error = "La fecha de nacimiento no puede ser futura"
        dv_fecha.errorTitle = "Fecha invÃ¡lida"
        dv_fecha.add('G2:G101')  # Aplicar a 100 filas
        ws_template.add_data_validation(dv_fecha)
        
        # ValidaciÃ³n para email (columna E)
        dv_email = DataValidation(
            type="custom",
            formula1='AND(LEN(E2)>0,ISNUMBER(SEARCH("@",E2)),ISNUMBER(SEARCH(".",E2)))',
            showDropDown=True
        )
        dv_email.error = "Ingrese un email vÃ¡lido con formato usuario@dominio.com"
        dv_email.errorTitle = "Email invÃ¡lido"
        dv_email.add('E2:E101')  # Aplicar a 100 filas
        ws_template.add_data_validation(dv_email)
        
        # ValidaciÃ³n para cÃ©dula (columna A)
        dv_cedula = DataValidation(
            type="custom",
            formula1='AND(LEN(A2)>=8,LEN(A2)<=20)',
            showDropDown=True
        )
        dv_cedula.error = "La cÃ©dula debe tener entre 8 y 20 caracteres"
        dv_cedula.errorTitle = "CÃ©dula invÃ¡lida"
        dv_cedula.add('A2:A101')  # Aplicar a 100 filas
        ws_template.add_data_validation(dv_cedula)
        
        # ValidaciÃ³n para telÃ©fono (columna D)
        dv_telefono = DataValidation(
            type="custom",
            formula1='AND(LEN(D2)>=8,LEN(D2)<=15)',
            showDropDown=True
        )
        dv_telefono.error = "El telÃ©fono debe tener entre 8 y 15 caracteres"
        dv_telefono.errorTitle = "TelÃ©fono invÃ¡lido"
        dv_telefono.add('D2:D101')  # Aplicar a 100 filas
        ws_template.add_data_validation(dv_telefono)
        
        # HOJA 3: REFERENCIAS (OPCIONAL - para consulta rÃ¡pida)
        ws_referencias = wb.create_sheet("Referencias")
        ws_referencias.append(["REFERENCIAS DE CONFIGURACIÃ“N"])
        ws_referencias.append([""])
        
        # Modelos disponibles
        ws_referencias.append(["MODELOS DE VEHÃCULOS DISPONIBLES:"])
        for modelo in modelos_nombres:
            ws_referencias.append([f"- {modelo}"])
        ws_referencias.append([""])
        
        # Concesionarios disponibles
        ws_referencias.append(["CONCESIONARIOS DISPONIBLES:"])
        for concesionario in concesionarios_nombres:
            ws_referencias.append([f"- {concesionario}"])
        ws_referencias.append([""])
        
        # Analistas disponibles
        ws_referencias.append(["ANALISTAS DISPONIBLES:"])
        for analista in analistas_nombres:
            ws_referencias.append([f"- {analista}"])
        
        # GUARDAR EN BUFFER
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # CONFIGURAR RESPUESTA
        response.headers["Content-Disposition"] = "attachment; filename=plantilla_clientes_dinamica.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        logger.info(f"Plantilla generada exitosamente - Modelos: {len(modelos_nombres)}, Concesionarios: {len(concesionarios_nombres)}, Analistas: {len(analistas_nombres)}")
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Error generando plantilla dinÃ¡mica: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Error interno del servidor al generar plantilla"
        )