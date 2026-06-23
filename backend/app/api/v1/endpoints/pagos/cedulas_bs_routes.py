"""Pagos API: cedulas_bs."""
"""

Endpoints de pagos. Datos reales desde BD.

- Tabla pagos: GET/POST/PUT/DELETE /pagos/ (listado y CRUD para /pagos/pagos).

- GET /pagos/kpis, /stats, /ultimos; POST /upload, /conciliacion/upload, /{id}/aplicar-cuotas.



Nº documento / referencia de pago:

- Regla de cartera: **no puede repetirse** el valor almacenado en `pagos.numero_documento` (único en `pagos` y
  `pagos_con_errores`, salvo edición del mismo `pago_id`).

- **Única forma permitida** de reutilizar el mismo texto de comprobante del banco: desambiguar con **código**
  (`codigo_documento`). En BD se compone `base + §CD: + código` (`compose_numero_documento_almacenado`,
  máx. 100 caracteres). En revisión manual en préstamos, el token **A#### / P####** lo asigna **Visto** (sufijo
  operativo; mismo criterio que carga masiva).

- Formatos de comprobante: BNC/, BINANCE, VE/, etc. Carga masiva: columna opcional a la derecha del Nº = código.

- Además, huella funcional (prestamo + fecha + monto + ref_norm) evita duplicar el mismo pago operativo (409),
  vía `ux_pagos_fingerprint_activos`.

"""

import calendar

import io

import logging

import re

import time

import uuid

from datetime import date, datetime, time as dt_time

from decimal import Decimal

from typing import Any, Optional

from zoneinfo import ZoneInfo



from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body, Request

from fastapi.responses import StreamingResponse, Response

from pydantic import BaseModel, field_validator

from sqlalchemy import and_, case, delete, desc, exists, func, inspect, not_, or_, select, text

from sqlalchemy.orm import Session, aliased

from sqlalchemy.exc import IntegrityError, OperationalError



from app.core.database import get_db

from app.core.config import settings

from app.core.deps import (
    ComprobanteImagenReader,
    get_comprobante_imagen_reader,
    get_current_user,
    require_admin,
    require_operator_or_higher,
)
from app.core.rol_normalization import canonical_rol

from app.core.documento import (
    compose_numero_documento_almacenado,
    normalize_codigo_documento,
    normalize_documento,
    split_numero_documento_almacenado,
)
from app.utils.cedula_almacenamiento import (
    CedulaPagoFkError,
    asegurar_cedula_pago_para_fk,
    alinear_cedulas_clientes_existentes,
    normalizar_cedula_almacenamiento,
)
from app.services.pago_numero_documento import (
    numero_documento_ya_registrado,
    primer_pago_cartera_por_documento,
)

from app.core.serializers import to_float, format_date_iso

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.prestamo import Prestamo

from app.models.pago import Pago

from app.models.pago_comprobante_imagen import PagoComprobanteImagen

from app.models.pago_con_error import PagoConError

from app.models.revisar_pago import RevisarPago

from app.models.cuota_pago import CuotaPago

from app.models.pago_reportado import PagoReportado

from app.models.datos_importados_conerrores import DatosImportadosConErrores

from app.models.cedula_reportar_bs import CedulaReportarBs

from app.schemas.pago import (
    PagoCreate,
    PagoUpdate,
    PagoResponse,
    normalizar_link_comprobante,
)

from app.schemas.auth import UserResponse

from app.services.cobros.pago_reportado_documento import (
    claves_documento_pago_desde_campos,
    claves_documento_pago_para_reportado,
    claves_documento_para_lote_reportados,
    documento_numero_desde_pago_reportado,
    pago_reportado_colisiona_tabla_pagos,
)
from app.services.pagos.comprobante_link_desde_gmail import (
    enriquecer_items_link_comprobante_desde_gmail,
    enriquecer_items_link_comprobante_desde_pago_reportado,
)
from app.services.pagos.comprobante_imagen_finiquito_access import (
    comprobante_imagen_accesible_finiquito_portal,
)
from app.services.cuota_pago_integridad import (
    pago_tiene_aplicaciones_cuotas,
    validar_suma_aplicada_vs_monto_pago,
)
from app.services.pagos_cuotas_reaplicacion import (
    prestamo_requiere_correccion_cascada,
    reset_y_reaplicar_cascada_prestamo,
)
from app.services.pagos_cascada_aplicacion import (
    _aplicar_pago_a_cuotas_interno,
    _marcar_prestamo_liquidado_si_corresponde,
)
from app.services.pagos_aplicacion_prestamo import (
    aplicar_pagos_pendientes_prestamo,
    aplicar_pagos_pendientes_prestamo_con_diagnostico,
)
from app.services.pagos_cascada_mensajes import _mensaje_sin_aplicacion_cascada


from app.services.tasa_cambio_service import (
    convertir_bs_a_usd,
    normalizar_fuente_tasa,
    obtener_tasa_por_fecha,
    valor_tasa_para_fuente,
)

from app.services.pago_registro_moneda import (
    normalizar_moneda_registro,
    resolver_monto_registro_pago,
    preload_autorizados_bs,
)
from app.services.pago_huella_funcional import (
    conflicto_huella_para_creacion,
    contar_prestamos_con_huella_funcional_duplicada,
    HTTP_409_DETAIL_HUELLA_FUNCIONAL,
    mensaje_409_huella_funcional_con_id,
    primer_id_conflicto_huella_funcional,
    primer_par_huella_duplicada_prestamo,
    ref_norm_desde_campos,
)
from app.services.pago_huella_metricas import (
    registrar_rechazo_huella_funcional,
    snapshot_rechazos_huella_funcional,
)

from app.services.cobros.cedula_reportar_bs_service import (

    normalize_cedula_para_almacenar_lista_bs,

    normalize_cedula_lookup_key,

    load_autorizados_bs_claves,

    cedula_coincide_autorizados_bs,

    cedula_autorizada_para_bs,

    obtener_fuente_tasa_lista_bs,

    fuente_tasa_bs_efectiva_para_cedula,

)

from app.services.tasa_cambio_service import normalizar_fuente_tasa

from .payload_models import (
    AgregarCedulaReportarBsBody,
    ConsultarCedulasReportarBsBatchBody,
    GuardarFilaEditableBody,
    MoverRevisarPagosBody,
    PagoBatchBody,
    ValidarFilasBatchBody,
)
from app.services.cobros.cobros_publico_reporte_service import (
    mime_efectivo_comprobante_web,
    mime_efectivo_con_firma_archivo,
    preparar_adjunto_comprobante_para_vision,
    validate_file_magic,
)

from .comprobante_imagen_helpers import (
    _COMPROBANTE_IMG_CT,
    _MAX_COMPROBANTE_IMAGEN_BYTES,
    _normalizar_id_comprobante_imagen,
    _public_base_url_para_comprobante,
)
from app.services.pagos_gmail.comprobante_bd import url_comprobante_imagen_absoluta

from .constants import (
    TZ_NEGOCIO,
    _MAX_LEN_NUMERO_DOCUMENTO,
    _MAX_MONTO_PAGADO,
    _MIN_MONTO_PAGADO,
    _PRESTAMO_ID_MAX,
)
from .cascada_estado import _estado_pago_tras_aplicar_cascada
from .sql_where_pagos import (
    _where_pago_elegible_reaplicacion_cascada,
    _where_pago_excluido_operacion,
)
from .pago_integridad_db import _integridad_error_pgcode_y_constraint
from .pago_normalizacion import (
    _celda_a_string_documento,
    _normalizar_ref_fingerprint,
    _safe_float,
    _validar_monto,
)
from .pago_zona_horaria import _calcular_dias_mora, _hoy_local
from .pago_usuario_registro import _usuario_registro_desde_current_user
from .pago_conciliacion_estado import (
    _alinear_estado_si_toggle_conciliado_actualizar_pago,
    _estado_conciliacion_post_cascada,
)
from .pago_cascada_reglas import _debe_aplicar_cascada_pago
from .pago_serializacion_respuesta import (
    _enriquecer_items_duplicado_clave_misma_pagina,
    _enriquecer_items_tiene_aplicacion_cuotas,
    _enriquecer_pagos_pago_reportado_id,
    _pago_response_enriquecido,
    _pago_to_response,
)









logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])

@router.get("/cedulas-reportar-bs", response_model=dict)

def get_cedulas_reportar_bs(

    page: int = Query(1, ge=1, description="Página (1-based)"),

    page_size: int = Query(10, ge=1, le=100, description="Filas por página"),

    db: Session = Depends(get_db),

):

    """

    Total de cédulas en la lista y página ordenada del más reciente al más antiguo (creado_en DESC).

    Si la BD aún no tiene `creado_en` o `fuente_tasa_cambio`, se consulta solo lo existente

    (evita 500 hasta aplicar migraciones 052 / 068).

    """

    bind = db.get_bind()

    table_cols = {c["name"] for c in inspect(bind).get_columns("cedulas_reportar_bs")}

    has_creado_en = "creado_en" in table_cols

    has_fuente = "fuente_tasa_cambio" in table_cols

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    off = (page - 1) * page_size

    if has_creado_en and has_fuente:

        q = (

            db.query(CedulaReportarBs)

            .order_by(desc(CedulaReportarBs.creado_en), CedulaReportarBs.cedula.asc())

            .offset(off)

            .limit(page_size)

        )

        rows = q.all()

        items = [

            {

                "cedula": r.cedula,

                "creado_en": r.creado_en.isoformat() if r.creado_en else None,

                "fuente_tasa_cambio": normalizar_fuente_tasa(
                    getattr(r, "fuente_tasa_cambio", None)
                ),

            }

            for r in rows

        ]

    else:

        select_cols = ["cedula"]

        if has_creado_en:

            select_cols.append("creado_en")

        if has_fuente:

            select_cols.append("fuente_tasa_cambio")

        col_sql = ", ".join(select_cols)

        order_sql = (

            "creado_en DESC, cedula ASC" if has_creado_en else "cedula ASC"

        )

        stmt = text(

            f"SELECT {col_sql} FROM cedulas_reportar_bs ORDER BY {order_sql} "

            "LIMIT :lim OFFSET :off"

        )

        raw_rows = db.execute(

            stmt, {"lim": page_size, "off": off}

        ).fetchall()

        items = []

        for tup in raw_rows:

            m = dict(zip(select_cols, tup))

            creado = m.get("creado_en")

            items.append(

                {

                    "cedula": m.get("cedula"),

                    "creado_en": creado.isoformat() if creado else None,

                    "fuente_tasa_cambio": normalizar_fuente_tasa(

                        m.get("fuente_tasa_cambio")

                    ),

                }

            )

    return {

        "total": int(total),

        "page": page,

        "page_size": page_size,

        "items": items,

    }





@router.get("/cedulas-reportar-bs/consultar", response_model=dict)


def consultar_cedula_reportar_bs(

    cedula: str = Query(..., min_length=1, description="Cedula a verificar"),

    db: Session = Depends(get_db),

):

    """Indica si la cedula esta autorizada para reportar pagos en Bs (lista cedulas_reportar_bs)."""

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    raw = (cedula or "").strip()

    en_lista = cedula_autorizada_para_bs(db, raw)

    ced_norm = normalize_cedula_para_almacenar_lista_bs(raw)

    fuente_lista = (
        normalizar_fuente_tasa(obtener_fuente_tasa_lista_bs(db, raw))
        if en_lista
        else None
    )

    return {

        "cedula_ingresada": raw,

        "cedula_normalizada": ced_norm,

        "en_lista": en_lista,

        "fuente_tasa_cambio_lista_bs": fuente_lista,

        "total_en_lista": int(total),

    }





@router.post("/cedulas-reportar-bs/consultar-batch", response_model=dict)

def consultar_cedulas_reportar_bs_batch(

    payload: ConsultarCedulasReportarBsBatchBody,

    db: Session = Depends(get_db),

):

    """Varias cedulas en una sola peticion: carga la lista autorizada una vez (evita N+1)."""

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    claves = load_autorizados_bs_claves(db)

    por_cedula: dict = {}

    for raw in payload.cedulas or []:

        raw = (raw or "").strip()

        if not raw:

            continue

        norm = normalize_cedula_lookup_key(raw)

        en_lista = cedula_coincide_autorizados_bs(norm, claves)

        ced_norm = normalize_cedula_para_almacenar_lista_bs(raw)

        fuente_lista = (
            normalizar_fuente_tasa(obtener_fuente_tasa_lista_bs(db, raw))
            if en_lista
            else None
        )

        por_cedula[raw] = {

            "cedula_ingresada": raw,

            "cedula_normalizada": ced_norm,

            "en_lista": en_lista,

            "fuente_tasa_cambio_lista_bs": fuente_lista,

            "total_en_lista": int(total),

        }

    return {"total_en_lista": int(total), "por_cedula": por_cedula}





@router.post("/cedulas-reportar-bs/agregar", response_model=dict)

def agregar_cedula_reportar_bs(

    payload: AgregarCedulaReportarBsBody,

    db: Session = Depends(get_db),

):

    """

    Agrega una cédula a la lista de quienes pueden reportar en Bs (nuevo cliente que paga en bolívares).

    Si ya existe, no duplica. La cédula se normaliza (V/E/J/G + dígitos).

    """

    cedula_norm = normalize_cedula_para_almacenar_lista_bs(payload.cedula)

    if not cedula_norm:

        raise HTTPException(

            status_code=400,

            detail="Cédula inválida. Use letra V, E, J o G seguida de 6 a 11 dígitos (ej: V12345678).",

        )

    fuente = normalizar_fuente_tasa(payload.fuente_tasa_cambio)

    claves_actuales = load_autorizados_bs_claves(db)

    if cedula_coincide_autorizados_bs(cedula_norm, claves_actuales):

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        existente = db.get(CedulaReportarBs, cedula_norm)

        if existente is not None:

            prev = normalizar_fuente_tasa(getattr(existente, "fuente_tasa_cambio", None))

            if prev != fuente:

                existente.fuente_tasa_cambio = fuente

                db.commit()

                return {

                    "agregada": False,

                    "cedula": cedula_norm,

                    "fuente_tasa_cambio": fuente,

                    "total": total,

                    "mensaje": (

                        f"La cédula {cedula_norm} ya estaba en la lista. "

                        f"Se actualizó la fuente de tasa a «{fuente}»."

                    ),

                }

            return {

                "agregada": False,

                "cedula": cedula_norm,

                "fuente_tasa_cambio": prev,

                "total": total,

                "mensaje": f"La cédula {cedula_norm} ya estaba en la lista.",

            }

        return {

            "agregada": False,

            "cedula": cedula_norm,

            "fuente_tasa_cambio": fuente,

            "total": total,

            "mensaje": f"La cédula {cedula_norm} ya estaba en la lista.",

        }

    try:

        db.add(CedulaReportarBs(cedula=cedula_norm, fuente_tasa_cambio=fuente))

        db.commit()

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        return {

            "agregada": True,

            "cedula": cedula_norm,

            "fuente_tasa_cambio": fuente,

            "total": total,

            "mensaje": f"Cédula {cedula_norm} agregada. Ya puede reportar pagos en Bs en Cobros e Infopagos.",

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error agregando cedula_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al agregar la cédula.") from e





@router.post("/cedulas-reportar-bs/upload", response_model=dict)

def upload_cedulas_reportar_bs(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

):

    """

    Carga un Excel con columna 'cedula'. Reemplaza la lista de cédulas que pueden reportar en Bs.

    Afecta a rapicredit-cobros e infopagos: solo esas cédulas verán habilitada la opción Bs.

    """

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):

        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")

    try:

        import openpyxl

    except ImportError:

        raise HTTPException(status_code=500, detail="Excel library not available")

    content = file.file.read()

    if len(content) < 50:

        raise HTTPException(status_code=400, detail="El archivo está vacío o no es un Excel válido.")

    try:

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    except Exception as e:

        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {str(e)}") from e

    ws = wb.active

    if not ws:

        raise HTTPException(status_code=400, detail="El Excel no tiene hojas.")

    # Buscar columna "cedula" (primera fila)

    header = [ (c and str(c).strip().lower() if c is not None else "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) ]

    col_cedula = None

    for i, h in enumerate(header):

        if h == "cedula":

            col_cedula = i

            break

    if col_cedula is None:

        raise HTTPException(status_code=400, detail="El Excel debe tener una columna llamada 'cedula'.")

    col_fuente = None

    for i, h in enumerate(header):

        if h in ("fuente_tasa_cambio", "fuente"):

            col_fuente = i

            break

    cedulas_por_fuente: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row or len(row) <= col_cedula:

            continue

        val = row[col_cedula]

        cedula_norm = normalize_cedula_para_almacenar_lista_bs(str(val).strip() if val is not None else "")

        if not cedula_norm:

            continue

        raw_fuente = None

        if col_fuente is not None and len(row) > col_fuente:

            raw_fuente = row[col_fuente]

        fuente_fila = normalizar_fuente_tasa(

            str(raw_fuente).strip() if raw_fuente is not None and str(raw_fuente).strip() else None

        )

        cedulas_por_fuente[cedula_norm] = fuente_fila

    wb.close()

    # Reemplazar tabla

    try:

        db.query(CedulaReportarBs).delete()

        for ced, fte in cedulas_por_fuente.items():

            db.add(CedulaReportarBs(cedula=ced, fuente_tasa_cambio=fte))

        db.commit()

    except Exception as e:

        db.rollback()

        logger.exception("Error guardando cedulas_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al guardar la lista de cédulas.") from e

    n = len(cedulas_por_fuente)

    return {

        "total": n,

        "mensaje": f"Se cargaron {n} cédula(s). Solo ellas pueden reportar pagos en Bs en RapiCredit Cobros e Infopagos.",

    }


@router.delete("/cedulas-reportar-bs/eliminar", response_model=dict)

def eliminar_cedula_reportar_bs(

    payload: AgregarCedulaReportarBsBody,

    db: Session = Depends(get_db),

):

    """

    Elimina una cédula de la lista de quienes pueden reportar en Bs.

    La cédula se normaliza (V/E/J/G + dígitos) antes de eliminar.

    """

    cedula_norm = normalize_cedula_para_almacenar_lista_bs(payload.cedula)

    if not cedula_norm:

        raise HTTPException(

            status_code=400,

            detail="Cédula inválida. Use letra V, E, J o G seguida de 6 a 11 dígitos (ej: V12345678).",

        )

    try:

        # Buscar y eliminar la cédula

        cedula_encontrada = db.query(CedulaReportarBs).filter(

            CedulaReportarBs.cedula == cedula_norm

        ).first()

        if not cedula_encontrada:

            total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

            return {

                "eliminada": False,

                "cedula": cedula_norm,

                "total": total,

                "mensaje": f"La cédula {cedula_norm} no se encontró en la lista.",

            }

        db.delete(cedula_encontrada)

        db.commit()

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        return {

            "eliminada": True,

            "cedula": cedula_norm,

            "total": total,

            "mensaje": f"Cédula {cedula_norm} eliminada. Ya no podrá reportar pagos en Bs en Cobros e Infopagos.",

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error eliminando cedula_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al eliminar la cédula.") from e

# Compat: nombre historico
# Compat: nombre historico (importaciones antiguas).
_estado_pago_tras_aplicar_fifo = _estado_pago_tras_aplicar_cascada
