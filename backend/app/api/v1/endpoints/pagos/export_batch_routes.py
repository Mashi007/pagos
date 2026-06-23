"""Pagos API: export_batch."""
"""

Endpoints de pagos. Datos reales desde BD.

- Tabla pagos: GET/POST/PUT/DELETE /pagos/ (listado y CRUD para /pagos/pagos).

- GET /pagos/kpis, /stats, /ultimos; POST /upload, /conciliacion/upload, /{id}/aplicar-cuotas.



Nº documento / referencia de pago:

- Regla de cartera: **no puede repetirse** el valor almacenado en `pagos.numero_documento` (único en `pagos` y
  `pagos_con_errores`, salvo edición del mismo `pago_id`).

- **Única forma permitida** de reutilizar el mismo texto de comprobante del banco: desambiguar con **código**
  (`codigo_documento`). En BD se compone `base + §CD: + código` (`compose_numero_documento_almacenado`,
  máx. 100 caracteres). En revisión manual en préstamos, el token **A#### / P####** lo asigna **Visto** (sufijo
  operativo; mismo criterio que carga masiva).

- Formatos de comprobante: BNC/, BINANCE, VE/, etc. Carga masiva: columna opcional a la derecha del Nº = código.

- Además, huella funcional (prestamo + fecha + monto + ref_norm) evita duplicar el mismo pago operativo (409),
  vía `ux_pagos_fingerprint_activos`.

"""

import calendar

import io

import logging

import re

import time

import uuid

from datetime import date, datetime, time as dt_time

from decimal import Decimal

from typing import Any, Optional

from zoneinfo import ZoneInfo



from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body, Request

from fastapi.responses import StreamingResponse, Response

from pydantic import BaseModel, field_validator

from sqlalchemy import and_, case, delete, desc, exists, func, inspect, not_, or_, select, text

from sqlalchemy.orm import Session, aliased

from sqlalchemy.exc import IntegrityError, OperationalError



from app.core.database import get_db

from app.core.config import settings

from app.core.deps import (
    ComprobanteImagenReader,
    get_comprobante_imagen_reader,
    get_current_user,
    require_admin,
    require_operator_or_higher,
)
from app.core.rol_normalization import canonical_rol

from app.core.documento import (
    compose_numero_documento_almacenado,
    normalize_codigo_documento,
    normalize_documento,
    split_numero_documento_almacenado,
)
from app.utils.cedula_almacenamiento import (
    CedulaPagoFkError,
    asegurar_cedula_pago_para_fk,
    alinear_cedulas_clientes_existentes,
    normalizar_cedula_almacenamiento,
)
from app.services.pago_numero_documento import (
    numero_documento_ya_registrado,
    primer_pago_cartera_por_documento,
)

from app.core.serializers import to_float, format_date_iso

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.prestamo import Prestamo

from app.models.pago import Pago

from app.models.pago_comprobante_imagen import PagoComprobanteImagen

from app.models.pago_con_error import PagoConError

from app.models.revisar_pago import RevisarPago

from app.models.cuota_pago import CuotaPago

from app.models.pago_reportado import PagoReportado

from app.models.datos_importados_conerrores import DatosImportadosConErrores

from app.models.cedula_reportar_bs import CedulaReportarBs

from app.schemas.pago import (
    PagoCreate,
    PagoUpdate,
    PagoResponse,
    normalizar_link_comprobante,
)

from app.schemas.auth import UserResponse

from app.services.cobros.pago_reportado_documento import (
    claves_documento_pago_desde_campos,
    claves_documento_pago_para_reportado,
    claves_documento_para_lote_reportados,
    documento_numero_desde_pago_reportado,
    pago_reportado_colisiona_tabla_pagos,
)
from app.services.pagos.comprobante_link_desde_gmail import (
    enriquecer_items_link_comprobante_desde_gmail,
    enriquecer_items_link_comprobante_desde_pago_reportado,
)
from app.services.pagos.comprobante_imagen_finiquito_access import (
    comprobante_imagen_accesible_finiquito_portal,
)
from app.services.cuota_pago_integridad import (
    pago_tiene_aplicaciones_cuotas,
    validar_suma_aplicada_vs_monto_pago,
)
from app.services.pagos_cuotas_reaplicacion import (
    prestamo_requiere_correccion_cascada,
    reset_y_reaplicar_cascada_prestamo,
)
from app.services.pagos_cascada_aplicacion import (
    _aplicar_pago_a_cuotas_interno,
    _marcar_prestamo_liquidado_si_corresponde,
)
from app.services.pagos_aplicacion_prestamo import (
    aplicar_pagos_pendientes_prestamo,
    aplicar_pagos_pendientes_prestamo_con_diagnostico,
)
from app.services.pagos_cascada_mensajes import _mensaje_sin_aplicacion_cascada


from app.services.tasa_cambio_service import (
    convertir_bs_a_usd,
    normalizar_fuente_tasa,
    obtener_tasa_por_fecha,
    valor_tasa_para_fuente,
)

from app.services.pago_registro_moneda import (
    normalizar_moneda_registro,
    resolver_monto_registro_pago,
    preload_autorizados_bs,
)
from app.services.pago_huella_funcional import (
    conflicto_huella_para_creacion,
    contar_prestamos_con_huella_funcional_duplicada,
    HTTP_409_DETAIL_HUELLA_FUNCIONAL,
    mensaje_409_huella_funcional_con_id,
    primer_id_conflicto_huella_funcional,
    primer_par_huella_duplicada_prestamo,
    ref_norm_desde_campos,
)
from app.services.pago_huella_metricas import (
    registrar_rechazo_huella_funcional,
    snapshot_rechazos_huella_funcional,
)

from app.services.cobros.cedula_reportar_bs_service import (

    normalize_cedula_para_almacenar_lista_bs,

    normalize_cedula_lookup_key,

    load_autorizados_bs_claves,

    cedula_coincide_autorizados_bs,

    cedula_autorizada_para_bs,

    obtener_fuente_tasa_lista_bs,

    fuente_tasa_bs_efectiva_para_cedula,

)

from app.services.tasa_cambio_service import normalizar_fuente_tasa

from .payload_models import (
    AgregarCedulaReportarBsBody,
    ConsultarCedulasReportarBsBatchBody,
    GuardarFilaEditableBody,
    MoverRevisarPagosBody,
    PagoBatchBody,
    ValidarFilasBatchBody,
)
from app.services.cobros.cobros_publico_reporte_service import (
    mime_efectivo_comprobante_web,
    mime_efectivo_con_firma_archivo,
    preparar_adjunto_comprobante_para_vision,
    validate_file_magic,
)

from .comprobante_imagen_helpers import (
    _COMPROBANTE_IMG_CT,
    _MAX_COMPROBANTE_IMAGEN_BYTES,
    _normalizar_id_comprobante_imagen,
    _public_base_url_para_comprobante,
)
from .pagos_cedula_helpers import looks_like_cedula_vej
from app.services.pagos_gmail.comprobante_bd import url_comprobante_imagen_absoluta

from .constants import (
    TZ_NEGOCIO,
    _MAX_LEN_NUMERO_DOCUMENTO,
    _MAX_MONTO_PAGADO,
    _MIN_MONTO_PAGADO,
    _PRESTAMO_ID_MAX,
)
from .cascada_estado import _estado_pago_tras_aplicar_cascada
from .sql_where_pagos import (
    _where_pago_elegible_reaplicacion_cascada,
    _where_pago_excluido_operacion,
)
from .pago_integridad_db import _integridad_error_pgcode_y_constraint
from .pago_normalizacion import (
    _celda_a_string_documento,
    _normalizar_ref_fingerprint,
    _safe_float,
    _validar_monto,
)
from .pago_zona_horaria import _calcular_dias_mora, _hoy_local
from .pago_usuario_registro import _usuario_registro_desde_current_user
from .pago_conciliacion_estado import (
    _alinear_estado_si_toggle_conciliado_actualizar_pago,
    _estado_conciliacion_post_cascada,
)
from .pago_cascada_reglas import _debe_aplicar_cascada_pago
from .pago_serializacion_respuesta import (
    _enriquecer_items_duplicado_clave_misma_pagina,
    _enriquecer_items_tiene_aplicacion_cuotas,
    _enriquecer_pagos_pago_reportado_id,
    _pago_response_enriquecido,
    _pago_to_response,
)









logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])

@router.get("/export/excel/pagos-sin-aplicar-cuotas")

def export_excel_pagos_sin_aplicar_cuotas(

    cohorte: str = Query(

        "todos",

        description="todos (default): sin ninguna fila en cuota_pagos (no aplicados a cuotas). "

        "cascada: ademas monto>0, prestamo_id, no ANULADO_IMPORT (cola del job). "

        "sin_cupo: como cascada y sin cupo aplicable en cuotas PENDIENTE/MORA/PARCIAL. "

        "Nota: cohorte «cascada» es el nombre vigente; «fifo» se acepta como alias (compatibilidad).",

    ),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Descarga Excel con pagos que no tienen aplicacion a cuotas (sin filas en cuota_pagos).

    Por defecto cohorte=todos incluye todos esos pagos. Requiere autenticacion. Maximo 200000 filas.

    """

    c = (cohorte or "todos").strip().lower()

    if c == "fifo":

        c = "cascada"

    if c not in ("todos", "cascada", "sin_cupo"):

        raise HTTPException(
            status_code=400,
            detail="cohorte debe ser todos, cascada o sin_cupo (fifo aceptado como alias de cascada)",
        )



    import openpyxl



    tiene_cuota_pago = exists(select(CuotaPago.id).where(CuotaPago.pago_id == Pago.id))

    if c == "todos":

        cond_final = ~tiene_cuota_pago

    else:

        cond_base = and_(

            ~tiene_cuota_pago,

            func.coalesce(Pago.monto_pagado, 0) > 0,

            func.upper(func.coalesce(Pago.estado, "")) != "ANULADO_IMPORT",

            Pago.prestamo_id.isnot(None),

        )

        if c == "sin_cupo":

            aplicado_en_cuota = (

                select(func.coalesce(func.sum(CuotaPago.monto_aplicado), 0))

                .where(CuotaPago.cuota_id == Cuota.id)

                .correlate(Cuota)

                .scalar_subquery()

            )

            hay_cupo_aplicable = exists(

                select(1)

                .select_from(Cuota)

                .where(

                    and_(

                        Cuota.prestamo_id == Pago.prestamo_id,

                        Cuota.estado.in_(["PENDIENTE", "VENCIDO", "MORA", "PARCIAL"]),

                        (func.coalesce(Cuota.monto, 0) - aplicado_en_cuota) > 0.01,

                    )

                )

            )

            cond_final = and_(cond_base, ~hay_cupo_aplicable)

        else:

            cond_final = cond_base



    rows = (

        db.execute(

            select(Pago)

            .where(cond_final)

            .order_by(Pago.fecha_registro.asc(), Pago.id.asc())

            .limit(200000)

        )

        .scalars()

        .all()

    )

    rows = [r for r in rows if r is not None]



    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Pagos sin aplicar"

    headers = [

        "pago_id",

        "fecha_registro",

        "fecha_pago",

        "prestamo_id",

        "cedula",

        "monto_pagado",

        "estado",

        "referencia_pago",

        "numero_documento",

        "conciliado",

        "usuario_registro",

        "cohorte_filtro",

    ]

    for col, h in enumerate(headers, 1):

        ws.cell(row=1, column=col, value=h)

    for row_idx, p in enumerate(rows, 2):

        fr = p.fecha_registro

        fp = p.fecha_pago

        ws.cell(row=row_idx, column=1, value=p.id)

        ws.cell(row=row_idx, column=2, value=fr.strftime("%Y-%m-%d %H:%M:%S") if fr else "")

        ws.cell(row=row_idx, column=3, value=fp.strftime("%Y-%m-%d %H:%M:%S") if fp else "")

        ws.cell(row=row_idx, column=4, value=p.prestamo_id)

        ws.cell(row=row_idx, column=5, value=(p.cedula_cliente or ""))

        ws.cell(row=row_idx, column=6, value=float(p.monto_pagado) if p.monto_pagado is not None else 0)

        ws.cell(row=row_idx, column=7, value=p.estado or "")

        ws.cell(row=row_idx, column=8, value=p.referencia_pago or "")

        ws.cell(row=row_idx, column=9, value=p.numero_documento or "")

        ws.cell(row=row_idx, column=10, value=bool(p.conciliado) if p.conciliado is not None else False)

        ws.cell(row=row_idx, column=11, value=p.usuario_registro or "")

        ws.cell(row=row_idx, column=12, value=c)



    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    fname = f"pagos_sin_aplicar_cuotas_{c}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={fname}"},

    )





@router.post("/validar-filas-batch", response_model=dict)

def validar_filas_batch(

    body: ValidarFilasBatchBody = Body(...),

    db: Session = Depends(get_db),

):

    """

    Valida en lote (carga masiva / preview):

    - Cédulas: deben existir en tabla clientes (misma clave que FK fk_pagos_cedula al guardar pagos).

    - Nº documento: clave canónica única global (comprobante normalizado + código opcional compuesto

      como en POST /pagos); ya usada en `pagos` o en `pagos_con_errores` → duplicado.

      (Un solo registro por esa clave; las cuotas referencian ese pago vía pago_id.)

    """

    cedulas_norm = list(

        {

            c.strip().replace("-", "").upper()

            for c in (body.cedulas or [])

            if c and c.strip()

        }

    )

    cedulas_existentes: set[str] = set()

    if cedulas_norm:

        cc = func.upper(func.replace(Cliente.cedula, "-", ""))

        rows = db.execute(select(cc).where(cc.in_(cedulas_norm)).distinct()).all()

        cedulas_existentes = {r[0] for r in rows if r and r[0]}

    documentos_duplicados: list[dict] = []

    docs_norm_limpios: list[str] = []

    for d in body.documentos or []:

        if d is None:

            continue

        s = str(d).strip() if not isinstance(d, str) else (d or "").strip()

        if not s:

            continue

        n = normalize_documento(s)

        if n:

            docs_norm_limpios.append(n)

    docs_norm_limpios = list(dict.fromkeys(docs_norm_limpios))

    if docs_norm_limpios:

        doc_match_pagos = or_(

            Pago.numero_documento.in_(docs_norm_limpios),

            func.trim(Pago.numero_documento).in_(docs_norm_limpios),

        )

        rows_pagos = db.execute(

            select(

                Pago.numero_documento,

                Pago.id,

                Pago.cedula_cliente,

                Pago.fecha_pago,

                Pago.monto_pagado,

                Pago.prestamo_id,

            )

            .where(doc_match_pagos)

            .where(Pago.numero_documento.isnot(None))

            .distinct()

        ).all()

        docs_set = set(docs_norm_limpios)

        seen_pago_ids: set[int] = set()

        for row in rows_pagos:

            if normalize_documento(row[0]) not in docs_set:

                continue

            if row[1] in seen_pago_ids:

                continue

            seen_pago_ids.add(int(row[1]))

            documentos_duplicados.append(

                {

                    "numero_documento": row[0],

                    "pago_id": row[1],

                    "cedula": row[2],

                    "fecha_pago": row[3].isoformat() if row[3] else None,

                    "monto_pagado": float(row[4]) if row[4] else 0,

                    "prestamo_id": int(row[5]) if row[5] is not None else None,

                    "estado": "duplicado",

                    "origen": "pagos",

                }

            )

        doc_match_pce = or_(

            PagoConError.numero_documento.in_(docs_norm_limpios),

            func.trim(PagoConError.numero_documento).in_(docs_norm_limpios),

        )

        rows_pce = db.execute(

            select(

                PagoConError.numero_documento,

                PagoConError.id,

                PagoConError.cedula_cliente,

                PagoConError.fecha_pago,

                PagoConError.monto_pagado,

                PagoConError.prestamo_id,

            )

            .where(doc_match_pce)

            .where(PagoConError.numero_documento.isnot(None))

            .distinct()

        ).all()

        seen_pce_ids: set[int] = set()

        for row in rows_pce:

            if normalize_documento(row[0]) not in docs_set:

                continue

            if row[1] in seen_pce_ids:

                continue

            seen_pce_ids.add(int(row[1]))

            documentos_duplicados.append(

                {

                    "numero_documento": row[0],

                    "pago_con_error_id": row[1],

                    "cedula": row[2],

                    "fecha_pago": row[3].isoformat() if row[3] else None,

                    "monto_pagado": float(row[4]) if row[4] else 0,

                    "prestamo_id": int(row[5]) if row[5] is not None else None,

                    "estado": "duplicado",

                    "origen": "pagos_con_errores",

                }

            )

    return {

        "cedulas_existentes": list(cedulas_existentes),

        "documentos_duplicados": documentos_duplicados,

    }





@router.post("/guardar-fila-editable", response_model=dict)
@router.put("/guardar-fila-editable", response_model=dict)

def guardar_fila_editable(

    body: GuardarFilaEditableBody = Body(...),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Guarda una fila editable validada (desde Preview).

    Si cumple validadores, inserta en pagos, aplica cuotas y retorna éxito.

    Auto-marca como conciliado ('SI') para aplicar reglas de negocio inmediatamente.

    """

    try:

        cedula = (body.cedula or "").strip()

        monto = body.monto_pagado

        numero_doc = (body.numero_documento or "").strip() if body.numero_documento else None

        codigo_doc = (body.codigo_documento or "").strip() if getattr(body, "codigo_documento", None) else None

        prestamo_id = body.prestamo_id

        usuario_registro = _usuario_registro_desde_current_user(current_user)



        # Validaciones post-guardado

        if not cedula:

            raise HTTPException(status_code=400, detail="Cédula requerida")

        if not looks_like_cedula_vej(cedula):

            raise HTTPException(status_code=400, detail="Cédula inválida (debe ser V/E/J/Z + 6-11 dígitos)")


        cedula_norm_in = normalizar_cedula_almacenamiento(cedula)
        if not cedula_norm_in:
            raise HTTPException(status_code=400, detail="Cédula requerida")
        cn_in = cedula_norm_in.replace("-", "").upper()
        cc_cli = func.upper(func.replace(Cliente.cedula, "-", ""))
        cli_por_cedula = db.execute(
            select(Cliente.id).where(cc_cli == cn_in).limit(1)
        ).first()
        if not cli_por_cedula:
            raise HTTPException(
                status_code=400,
                detail="La cédula no está registrada en clientes. No se puede guardar el pago.",
            )

        if monto <= 0:

            raise HTTPException(status_code=400, detail="Monto debe ser > 0")

        if monto > _MAX_MONTO_PAGADO:

            raise HTTPException(status_code=400, detail=f"Monto excede límite máximo: {_MAX_MONTO_PAGADO}")



        # Parsear fecha

        try:

            fecha_pago = datetime.strptime(body.fecha_pago[:10], "%d-%m-%Y").date()

        except (ValueError, IndexError):

            raise HTTPException(status_code=400, detail="Fecha inválida (formato: DD-MM-YYYY)")



        # Comprobante + código opcional → valor único almacenado

        numero_doc_norm = compose_numero_documento_almacenado(numero_doc, codigo_doc)

        if numero_doc_norm and numero_documento_ya_registrado(db, numero_doc_norm):

            raise HTTPException(

                status_code=409,

                detail=(

                    "Conflicto numero_documento: ya existe un pago con la misma combinación "

                    "comprobante + código."

                ),

            )

        # Si prestamo_id es None, buscar por cédula en préstamos (normalizada)

        if prestamo_id is None:

            ced_norm = (

                cedula.strip().replace("-", "").upper()

                if cedula

                else ""

            )

            if ced_norm:

                pc = func.upper(func.replace(Prestamo.cedula, "-", ""))

                prest_row = db.execute(

                    select(Prestamo.id)

                    .where(pc == ced_norm)

                    .order_by(Prestamo.id.desc())

                    .limit(1)

                ).first()

                if prest_row:

                    prestamo_id = prest_row[0]



        if prestamo_id is None:

            raise HTTPException(

                status_code=400,

                detail="La cédula no tiene préstamo asociado; registre el crédito antes de cargar el pago.",

            )



        # Cedula en pagos debe existir en clientes (FK fk_pagos_cedula): usar la del cliente del prestamo.
        prest = db.get(Prestamo, prestamo_id)
        if not prest:
            raise HTTPException(status_code=404, detail="Prestamo no encontrado")
        cli = db.get(Cliente, prest.cliente_id)
        if not cli:
            raise HTTPException(
                status_code=400,
                detail="El prestamo no tiene cliente asociado en BD; no se puede registrar el pago.",
            )
        # Datos legacy: clientes.cedula en minusculas; FK pagos.cedula debe coincidir con clientes.cedula
        ced_norm_cli = normalizar_cedula_almacenamiento(cli.cedula)
        if ced_norm_cli and ced_norm_cli != (cli.cedula or ""):
            cli.cedula = ced_norm_cli
            db.flush()
        cedula_fk = normalizar_cedula_almacenamiento(cli.cedula) or normalizar_cedula_almacenamiento(
            prest.cedula
        )
        if not cedula_fk:
            raise HTTPException(status_code=400, detail="Cedula del cliente no disponible en BD.")
        cedula_input = normalizar_cedula_almacenamiento(cedula.strip())
        cn_body = (cedula_input or "").replace("-", "").upper()
        cn_fk = (cedula_fk or "").replace("-", "").upper()
        if cedula_input and cn_body != cn_fk:
            raise HTTPException(
                status_code=400,
                detail=f"La cedula no coincide con la del cliente del prestamo (en BD: {cedula_fk}).",
            )



        # Crear pago

        # [A2] Marcar conciliado=True y verificado_concordancia="SI" desde el momento de la creación,

        # ya que guardar-fila-editable implica que el pago fue revisado y validado manualmente.

        ref_pago = (numero_doc_norm or (numero_doc or "Carga"))[:_MAX_LEN_NUMERO_DOCUMENTO]

        ahora_conciliacion = datetime.now(ZoneInfo(TZ_NEGOCIO))

        moneda_r = (body.moneda_registro or "USD").strip().upper()

        tasa_man_dec = None

        if body.tasa_cambio_manual is not None:

            tasa_man_dec = Decimal(str(body.tasa_cambio_manual))

        monto_usd_g, moneda_fin_g, monto_bs_g, tasa_g, fecha_tasa_g = resolver_monto_registro_pago(

            db,

            cedula_normalizada=(cedula_fk or "").strip().upper(),

            fecha_pago=fecha_pago,

            monto_pagado=Decimal(str(round(monto, 2))),

            moneda_registro=moneda_r,

            tasa_cambio_manual=tasa_man_dec,

        )

        msg_h_fila = conflicto_huella_para_creacion(
            db,
            prestamo_id=prestamo_id,
            fecha_pago=fecha_pago,
            monto_pagado=monto_usd_g,
            numero_documento=numero_doc_norm,
            referencia_pago=ref_pago,
        )
        if msg_h_fila:
            registrar_rechazo_huella_funcional()
            raise HTTPException(status_code=409, detail=msg_h_fila)

        pago = Pago(

            cedula_cliente=cedula_fk,

            prestamo_id=prestamo_id,

            fecha_pago=datetime.combine(fecha_pago, dt_time.min),

            monto_pagado=monto_usd_g,

            numero_documento=numero_doc_norm,

            estado="PAGADO",

            referencia_pago=ref_pago,

            conciliado=True,  # [B2] Usar solo conciliado

            fecha_conciliacion=ahora_conciliacion,

            verificado_concordancia="SI",  # Legacy: sync con conciliado

            usuario_registro=usuario_registro,

            moneda_registro=moneda_fin_g,

            monto_bs_original=monto_bs_g,

            tasa_cambio_bs_usd=tasa_g,

            fecha_tasa_referencia=fecha_tasa_g,

        )

        db.add(pago)

        db.flush()



        # Aplicar a cuotas si prestamo_id existe

        cuotas_completadas = 0

        cuotas_parciales = 0

        if pago.prestamo_id and float(pago.monto_pagado or 0) > 0:

            cuotas_completadas, cuotas_parciales = _aplicar_pago_a_cuotas_interno(pago, db)

        pago.estado = _estado_conciliacion_post_cascada(pago, cuotas_completadas, cuotas_parciales)



        db.commit()



        return {

            "success": True,

            "pago_id": pago.id,

            "message": "Fila guardada exitosamente",

            "cuotas_completadas": cuotas_completadas,

            "cuotas_parciales": cuotas_parciales,

        }



    except HTTPException:

        raise

    except ValueError as e:

        db.rollback()

        logger.warning("guardar-fila-editable integridad: %s", e)

        raise HTTPException(status_code=409, detail=str(e)) from e

    except Exception as e:

        db.rollback()

        logger.exception("Error guardar-fila-editable: %s", e)

        raise HTTPException(status_code=500, detail=str(e)) from e


