"""

Endpoints de pagos. Datos reales desde BD.

- Tabla pagos: GET/POST/PUT/DELETE /pagos/ (listado y CRUD para /pagos/pagos).

- GET /pagos/kpis, /stats, /ultimos; POST /upload, /conciliacion/upload, /{id}/aplicar-cuotas.



Nº documento / referencia de pago:

- No puede repetirse el mismo valor almacenado en `pagos.numero_documento` (clave canónica compuesta).

- El mismo comprobante del banco puede repetirse si se envía un **codigo_documento** distinto: en BD se guarda
  comprobante + sufijo interno + código (hasta 100 caracteres en total).

- Se aceptan TODOS los formatos de comprobante (BNC/, BINANCE, VE/, etc.). Carga masiva: columna opcional
  inmediatamente a la derecha del Nº documento = código (formatos D, A, B, C, E según plantilla).

- Huella funcional (prestamo + fecha + monto + ref_norm): evita duplicar el mismo pago operativo (409),
  alineado con ux_pagos_fingerprint_activos.

"""

import calendar

import io

import logging

import re

import uuid

from datetime import date, datetime, time as dt_time

from decimal import Decimal

from typing import Any, Optional

from zoneinfo import ZoneInfo



from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body, Request

from fastapi.responses import StreamingResponse, Response

from sqlalchemy import and_, case, delete, desc, exists, func, not_, or_, select, text

from sqlalchemy.orm import Session

from sqlalchemy.exc import IntegrityError, OperationalError



from app.core.database import get_db

from app.core.config import settings

from app.core.deps import get_current_user

from app.core.documento import (
    compose_numero_documento_almacenado,
    normalize_codigo_documento,
    normalize_documento,
    split_numero_documento_almacenado,
)
from app.utils.cedula_almacenamiento import alinear_cedulas_clientes_existentes, normalizar_cedula_almacenamiento
from app.services.pago_numero_documento import numero_documento_ya_registrado

from app.core.serializers import to_float, format_date_iso

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.prestamo import Prestamo

from app.models.pago import Pago

from app.models.pago_comprobante_imagen import PagoComprobanteImagen

from app.models.pago_con_error import PagoConError

from app.models.revisar_pago import RevisarPago

from app.models.cuota_pago import CuotaPago

from app.models.pago_reportado import PagoReportado

from app.models.datos_importados_conerrores import DatosImportadosConErrores

from app.models.cedula_reportar_bs import CedulaReportarBs

from app.schemas.pago import (
    PagoCreate,
    PagoUpdate,
    PagoResponse,
    normalizar_link_comprobante,
)

from app.schemas.auth import UserResponse

from app.services.cobros.pago_reportado_documento import (
    claves_documento_pago_desde_campos,
    claves_documento_pago_para_reportado,
    claves_documento_para_lote_reportados,
    documento_numero_desde_pago_reportado,
)
from app.services.pagos.comprobante_link_desde_gmail import (
    enriquecer_items_link_comprobante_desde_gmail,
)
from app.services.cuota_pago_integridad import (
    pago_tiene_aplicaciones_cuotas,
    validar_suma_aplicada_vs_monto_pago,
)
from app.services.pagos_cuotas_reaplicacion import (
    prestamo_requiere_correccion_cascada,
    reset_y_reaplicar_cascada_prestamo,
)
from app.services.pagos_cascada_aplicacion import (
    _aplicar_pago_a_cuotas_interno,
    _marcar_prestamo_liquidado_si_corresponde,
)


from app.services.tasa_cambio_service import (

    convertir_bs_a_usd,

    obtener_tasa_por_fecha,

)

from app.services.pago_registro_moneda import (
    normalizar_moneda_registro,
    resolver_monto_registro_pago,
    preload_autorizados_bs,
)
from app.services.pago_huella_funcional import (
    conflicto_huella_para_creacion,
    contar_prestamos_con_huella_funcional_duplicada,
    HTTP_409_DETAIL_HUELLA_FUNCIONAL,
    mensaje_409_huella_funcional_con_id,
    primer_id_conflicto_huella_funcional,
    ref_norm_desde_campos,
)
from app.services.pago_huella_metricas import (
    registrar_rechazo_huella_funcional,
    snapshot_rechazos_huella_funcional,
)

from app.services.cobros.cedula_reportar_bs_service import (

    normalize_cedula_para_almacenar_lista_bs,

    normalize_cedula_lookup_key,

    load_autorizados_bs_claves,

    cedula_coincide_autorizados_bs,

    cedula_autorizada_para_bs,

)

from .payload_models import (
    AgregarCedulaReportarBsBody,
    ConsultarCedulasReportarBsBatchBody,
    GuardarFilaEditableBody,
    MoverRevisarPagosBody,
    PagoBatchBody,
    ValidarFilasBatchBody,
)
from .comprobante_imagen_helpers import (
    _COMPROBANTE_IMG_CT,
    _MAX_COMPROBANTE_IMAGEN_BYTES,
    _normalizar_id_comprobante_imagen,
    _public_base_url_para_comprobante,
)
from app.services.pagos_gmail.comprobante_bd import url_comprobante_imagen_absoluta

from .constants import (
    TZ_NEGOCIO,
    _MAX_LEN_NUMERO_DOCUMENTO,
    _MAX_MONTO_PAGADO,
    _MIN_MONTO_PAGADO,
    _PRESTAMO_ID_MAX,
)
from .cascada_estado import _estado_pago_tras_aplicar_cascada
from .sql_where_pagos import (
    _where_pago_elegible_reaplicacion_cascada,
    _where_pago_excluido_operacion,
)
from .pago_integridad_db import _integridad_error_pgcode_y_constraint
from .pago_normalizacion import (
    _celda_a_string_documento,
    _normalizar_ref_fingerprint,
    _safe_float,
    _validar_monto,
)
from .pago_zona_horaria import _calcular_dias_mora, _hoy_local
from .pago_usuario_registro import _usuario_registro_desde_current_user
from .pago_conciliacion_estado import (
    _alinear_estado_si_toggle_conciliado_actualizar_pago,
    _estado_conciliacion_post_cascada,
)
from .pago_cascada_reglas import _debe_aplicar_cascada_pago
from .pago_serializacion_respuesta import (
    _enriquecer_items_tiene_aplicacion_cuotas,
    _enriquecer_pagos_pago_reportado_id,
    _pago_response_enriquecido,
    _pago_to_response,
)








logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])


@router.post("/comprobante-imagen", response_model=dict)
async def upload_pago_comprobante_imagen(
    request: Request,
    file: UploadFile = File(..., alias="file"),
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Sube una imagen de comprobante. Devuelve URL absoluta para persistir en pagos.link_comprobante.
    """
    ct_raw = (file.content_type or "").split(";")[0].strip().lower()
    if ct_raw not in _COMPROBANTE_IMG_CT:
        raise HTTPException(
            status_code=400,
            detail="Solo se permiten imagenes JPEG, PNG, WebP o GIF.",
        )
    data = await file.read()
    if len(data) > _MAX_COMPROBANTE_IMAGEN_BYTES:
        raise HTTPException(
            status_code=400,
            detail="La imagen no puede superar 8 MB.",
        )
    uid = uuid.uuid4().hex
    row = PagoComprobanteImagen(
        id=uid,
        content_type=ct_raw,
        imagen_data=data,
    )
    db.add(row)
    db.commit()
    base = _public_base_url_para_comprobante(request)
    rel_path = f"{settings.API_V1_STR}/pagos/comprobante-imagen/{uid}"
    url = f"{base}{rel_path}"
    return {"url": url, "id": uid}


@router.get("/comprobante-imagen/{comprobante_id}")
def get_pago_comprobante_imagen(
    comprobante_id: str,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Sirve la imagen subida (requiere sesion; el enlace en link_comprobante es para personal autenticado)."""
    cid = _normalizar_id_comprobante_imagen(comprobante_id)
    if not cid:
        raise HTTPException(
            status_code=400,
            detail="Identificador de comprobante no valido.",
        )
    row = db.get(PagoComprobanteImagen, cid)
    if row is None:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado.")
    return Response(
        content=row.imagen_data,
        media_type=(row.content_type or "application/octet-stream"),
        headers={
            "Cache-Control": "private, no-store",
            "Pragma": "no-cache",
        },
    )



@router.get("", response_model=dict)

def listar_pagos(

    page: int = Query(1, ge=1),

    per_page: int = Query(20, ge=1, le=100),

    cedula: Optional[str] = Query(None),

    estado: Optional[str] = Query(None),

    fecha_desde: Optional[str] = Query(None),

    fecha_hasta: Optional[str] = Query(None),

    analista: Optional[str] = Query(None),

    conciliado: Optional[str] = Query(None, description="si=conciliados, no=no conciliados, vacío=todos"),

    sin_prestamo: Optional[str] = Query(None, description="si=solo pagos sin crédito asignado (prestamo_id NULL)"),

    prestamo_cartera: str = Query(
        "activa",
        description="activa=solo pagos sin crédito o con préstamo APROBADO (oculta LIQUIDADO y otros). todos=sin filtrar por estado del préstamo.",
    ),

    resumen_prestamo_id: Optional[int] = Query(
        None,
        ge=1,
        description=(
            "Si se indica, la respuesta incluye resumen_prestamo: agregados de la tabla pagos "
            "para ese prestamo_id (cantidad, suma de montos, pendiente vs PAGADO). Opcionalmente "
            "cruzado con filtro cédula si viene en la misma petición. No altera el listado paginado."
        ),
    ),

    db: Session = Depends(get_db),

):

    """Listado paginado desde la tabla pagos. Filtros: cedula, estado, fecha_desde, fecha_hasta, analista, conciliado, sin_prestamo, prestamo_cartera."""

    try:

        # Invocación directa (p. ej. finiquito): defaults sin resolver siguen siendo Query().

        def _solo_str_lp(v: Any) -> Optional[str]:

            return v if isinstance(v, str) else None

        page = page if isinstance(page, int) else 1

        per_page = per_page if isinstance(per_page, int) else 20

        cedula = _solo_str_lp(cedula)

        estado = _solo_str_lp(estado)

        fecha_desde = _solo_str_lp(fecha_desde)

        fecha_hasta = _solo_str_lp(fecha_hasta)

        analista = _solo_str_lp(analista)

        conciliado = _solo_str_lp(conciliado)

        sin_prestamo = _solo_str_lp(sin_prestamo)

        prestamo_cartera = _solo_str_lp(prestamo_cartera) or "activa"

        q = select(Pago)

        count_q = select(func.count()).select_from(Pago)

        sum_q = select(func.coalesce(func.sum(Pago.monto_pagado), 0)).select_from(Pago)

        _pc = (prestamo_cartera or "activa").strip().lower()

        _solo_prestamos_aprobados = _pc not in ("todos", "all", "t")

        def _estado_prestamo_aprobado_sql():
            return func.upper(
                func.trim(func.coalesce(Prestamo.estado, ""))
            ) == "APROBADO"

        if sin_prestamo and sin_prestamo.strip().lower() == "si":

            q = q.where(Pago.prestamo_id.is_(None))

            count_q = count_q.where(Pago.prestamo_id.is_(None))

            sum_q = sum_q.where(Pago.prestamo_id.is_(None))

            # Excluir pagos ya movidos a revisar_pagos (tabla temporal de validación)

            revisar_subq = select(RevisarPago.pago_id)

            q = q.where(~Pago.id.in_(revisar_subq))

            count_q = count_q.where(~Pago.id.in_(revisar_subq))

            sum_q = sum_q.where(~Pago.id.in_(revisar_subq))

        if conciliado and conciliado.strip().lower() == "si":

            q = q.where(Pago.conciliado == True)

            count_q = count_q.where(Pago.conciliado == True)

            sum_q = sum_q.where(Pago.conciliado == True)

        elif conciliado and conciliado.strip().lower() == "no":

            q = q.where(or_(Pago.conciliado == False, Pago.conciliado.is_(None)))

            count_q = count_q.where(or_(Pago.conciliado == False, Pago.conciliado.is_(None)))

            sum_q = sum_q.where(or_(Pago.conciliado == False, Pago.conciliado.is_(None)))

        if cedula and cedula.strip():

            q = q.where(Pago.cedula_cliente.ilike(f"%{cedula.strip()}%"))

            count_q = count_q.where(Pago.cedula_cliente.ilike(f"%{cedula.strip()}%"))

            sum_q = sum_q.where(Pago.cedula_cliente.ilike(f"%{cedula.strip()}%"))

        if estado and estado.strip():

            q = q.where(Pago.estado == estado.strip().upper())

            count_q = count_q.where(Pago.estado == estado.strip().upper())

            sum_q = sum_q.where(Pago.estado == estado.strip().upper())

        if fecha_desde:

            try:

                fd = date.fromisoformat(fecha_desde)

                q = q.where(Pago.fecha_pago >= datetime.combine(fd, dt_time.min))

                count_q = count_q.where(Pago.fecha_pago >= datetime.combine(fd, dt_time.min))

                sum_q = sum_q.where(Pago.fecha_pago >= datetime.combine(fd, dt_time.min))

            except ValueError:

                pass

        if fecha_hasta:

            try:

                fh = date.fromisoformat(fecha_hasta)

                q = q.where(Pago.fecha_pago <= datetime.combine(fh, dt_time.max))

                count_q = count_q.where(Pago.fecha_pago <= datetime.combine(fh, dt_time.max))

                sum_q = sum_q.where(Pago.fecha_pago <= datetime.combine(fh, dt_time.max))

            except ValueError:

                pass

        if analista and analista.strip():

            q = q.join(Prestamo, Pago.prestamo_id == Prestamo.id).where(Prestamo.analista == analista.strip())

            count_q = count_q.join(Prestamo, Pago.prestamo_id == Prestamo.id).where(Prestamo.analista == analista.strip())

            sum_q = sum_q.join(Prestamo, Pago.prestamo_id == Prestamo.id).where(Prestamo.analista == analista.strip())

            if _solo_prestamos_aprobados:

                q = q.where(_estado_prestamo_aprobado_sql())

                count_q = count_q.where(_estado_prestamo_aprobado_sql())

                sum_q = sum_q.where(_estado_prestamo_aprobado_sql())

        elif _solo_prestamos_aprobados:

            q = q.outerjoin(Prestamo, Pago.prestamo_id == Prestamo.id).where(

                or_(Pago.prestamo_id.is_(None), _estado_prestamo_aprobado_sql()),

            )

            count_q = count_q.outerjoin(Prestamo, Pago.prestamo_id == Prestamo.id).where(

                or_(Pago.prestamo_id.is_(None), _estado_prestamo_aprobado_sql()),

            )

            sum_q = sum_q.outerjoin(Prestamo, Pago.prestamo_id == Prestamo.id).where(

                or_(Pago.prestamo_id.is_(None), _estado_prestamo_aprobado_sql()),

            )

        total = db.scalar(count_q) or 0

        # Orden: más reciente primero (fecha_pago desc, luego id desc)

        q = q.order_by(Pago.fecha_pago.desc().nullslast(), Pago.id.desc()).offset((page - 1) * per_page).limit(per_page)

        rows = db.execute(q).scalars().all()

        items = [_pago_to_response(r) for r in rows]

        _enriquecer_items_tiene_aplicacion_cuotas(db, items)

        _enriquecer_pagos_pago_reportado_id(db, items)

        enriquecer_items_link_comprobante_desde_gmail(db, items)

        total_pages = (total + per_page - 1) // per_page if total else 0

        out: dict = {

            "pagos": items,

            "total": total,

            "page": page,

            "per_page": per_page,

            "total_pages": total_pages,

        }

        # Solo con filtro cédula: total de monto_pagado de todos los pagos que coinciden (no solo la página).
        if cedula and cedula.strip():

            raw_sum = db.scalar(sum_q)

            out["sum_monto_pagado_cedula"] = float(raw_sum or 0)

        # Resumen por crédito (sin filtrar el listado principal): auditoría revisión manual / coherencia con cuotas.
        if resumen_prestamo_id is not None and isinstance(resumen_prestamo_id, int) and resumen_prestamo_id > 0:

            pid = int(resumen_prestamo_id)

            rp_conds = [Pago.prestamo_id == pid]

            if cedula and cedula.strip():

                rp_conds.append(Pago.cedula_cliente.ilike(f"%{cedula.strip()}%"))

            rp_where = and_(*rp_conds)

            n_all = db.scalar(select(func.count()).select_from(Pago).where(rp_where)) or 0

            s_all = db.scalar(

                select(func.coalesce(func.sum(Pago.monto_pagado), 0)).select_from(Pago).where(rp_where)

            ) or 0

            pend_where = and_(rp_where, Pago.estado == "PENDIENTE")

            n_pend = db.scalar(select(func.count()).select_from(Pago).where(pend_where)) or 0

            s_pend = db.scalar(

                select(func.coalesce(func.sum(Pago.monto_pagado), 0)).select_from(Pago).where(pend_where)

            ) or 0

            pag_where = and_(rp_where, Pago.estado == "PAGADO")

            n_pag = db.scalar(select(func.count()).select_from(Pago).where(pag_where)) or 0

            s_pag = db.scalar(

                select(func.coalesce(func.sum(Pago.monto_pagado), 0)).select_from(Pago).where(pag_where)

            ) or 0

            out["resumen_prestamo"] = {

                "prestamo_id": pid,

                "cantidad": int(n_all),

                "suma_monto_pagado": float(s_all or 0),

                "cantidad_pendiente": int(n_pend),

                "suma_monto_pendiente": float(s_pend or 0),

                "cantidad_pagado": int(n_pag),

                "suma_monto_estado_pagado": float(s_pag or 0),

            }

        return out

    except Exception as e:

        logger.exception("Error en GET /pagos: %s", e)

        db.rollback()

        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/sin-prestamo/sugerir", response_model=dict)
def sugerir_prestamos_sin_asignar(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """
    Retorna pagos sin prestamo_id y sugiere asignación automática si el cliente tiene exactamente 1 crédito activo.
    
    Respuesta incluye:
    - pago_id: ID del pago
    - cedula_cliente: Cédula del cliente
    - prestamo_sugerido: ID del crédito (si hay exactamente 1) o None
    - num_creditos_activos: Cantidad de créditos activos del cliente
    - acciones_necesarias: "auto" si puede auto-asignarse, "manual" si requiere intervención
    """
    try:
        # Contar pagos sin prestamo_id
        count_q = select(func.count()).select_from(Pago).where(Pago.prestamo_id.is_(None))
        total = db.scalar(count_q) or 0
        
        # Obtener pagos sin prestamo_id
        q = select(Pago).where(Pago.prestamo_id.is_(None)).order_by(Pago.id.desc()).offset((page - 1) * per_page).limit(per_page)
        pagos_sin_prestamo = db.execute(q).scalars().all()
        
        sugerencias = []
        
        for pago in pagos_sin_prestamo:
            cedula_norm = pago.cedula_cliente.strip().upper() if pago.cedula_cliente else ""
            
            prestamos_activos = []
            prestamo_sugerido = None
            num_creditos = 0
            acciones = "manual"
            
            if cedula_norm:
                # Buscar créditos activos para esta cédula
                prestamos_activos = db.execute(
                    select(Prestamo.id)
                    .select_from(Prestamo)
                    .join(Cliente, Prestamo.cliente_id == Cliente.id)
                    .where(
                        Cliente.cedula == cedula_norm,
                        Prestamo.estado == "APROBADO",
                    )
                    .order_by(Prestamo.id)
                ).scalars().all()
                
                num_creditos = len(prestamos_activos)
                
                # Si exactamente 1 crédito activo → sugerencia automática
                if num_creditos == 1:
                    prestamo_sugerido = prestamos_activos[0]
                    acciones = "auto"
            
            sugerencias.append({
                "pago_id": pago.id,
                "cedula_cliente": pago.cedula_cliente,
                "fecha_pago": pago.fecha_pago.isoformat() if pago.fecha_pago else None,
                "monto_pagado": float(pago.monto_pagado) if pago.monto_pagado else 0,
                "numero_documento": pago.numero_documento,
                "prestamo_sugerido": prestamo_sugerido,
                "num_creditos_activos": num_creditos,
                "acciones_necesarias": acciones,
            })
        
        total_pages = (total + per_page - 1) // per_page if total else 0
        
        return {
            "sugerencias": sugerencias,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
            "resumen": {
                "total_pagos_sin_prestamo": total,
                "can_auto_asignar": sum(1 for s in sugerencias if s["acciones_necesarias"] == "auto"),
                "requieren_manual": sum(1 for s in sugerencias if s["acciones_necesarias"] == "manual"),
            }
        }
    
    except Exception as e:
        logger.exception("Error en GET /pagos/sin-prestamo/sugerir: %s", e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/sin-prestamo/asignar-automatico", response_model=dict)
def asignar_automaticamente_prestamos(
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Auto-asigna prestamo_id a pagos que:
    1. Tienen prestamo_id = NULL
    2. El cliente tiene EXACTAMENTE 1 crédito activo (APROBADO)
    
    Retorna cantidad de pagos actualizados y detalles.
    """
    try:
        # Obtener todos los pagos sin prestamo_id
        pagos_sin_prestamo = db.execute(
            select(Pago).where(Pago.prestamo_id.is_(None))
        ).scalars().all()
        
        asignados = []
        no_asignables = []
        
        for pago in pagos_sin_prestamo:
            cedula_norm = pago.cedula_cliente.strip().upper() if pago.cedula_cliente else ""
            
            if not cedula_norm:
                no_asignables.append({
                    "pago_id": pago.id,
                    "razon": "Cédula vacía"
                })
                continue
            
            # Buscar créditos activos
            prestamos_activos = db.execute(
                select(Prestamo.id)
                .select_from(Prestamo)
                .join(Cliente, Prestamo.cliente_id == Cliente.id)
                .where(
                    Cliente.cedula == cedula_norm,
                    Prestamo.estado == "APROBADO",
                )
                .order_by(Prestamo.id)
            ).scalars().all()
            
            if len(prestamos_activos) == 1:
                # Auto-asignar
                pago.prestamo_id = prestamos_activos[0]
                db.add(pago)
                asignados.append({
                    "pago_id": pago.id,
                    "cedula_cliente": pago.cedula_cliente,
                    "prestamo_id_asignado": prestamos_activos[0],
                })
            else:
                no_asignables.append({
                    "pago_id": pago.id,
                    "cedula_cliente": pago.cedula_cliente,
                    "razon": f"Cliente tiene {len(prestamos_activos)} créditos activos (requiere intervención manual)" if len(prestamos_activos) > 1 else "Cliente sin créditos activos"
                })
        
        # Commit
        db.commit()
        
        logger.info(
            "Auto-asignación de prestamos: %d asignados, %d no asignables",
            len(asignados),
            len(no_asignables)
        )
        
        return {
            "asignados": len(asignados),
            "no_asignables": len(no_asignables),
            "detalles_asignados": asignados[:50],  # Limitar a 50 para la respuesta
            "detalles_no_asignables": no_asignables[:50],
            "mensaje": f"Se asignaron {len(asignados)} pagos automáticamente. {len(no_asignables)} requieren intervención manual.",
        }
    
    except Exception as e:
        logger.exception("Error en POST /pagos/sin-prestamo/asignar-automatico: %s", e)
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e)) from e



@router.get("/ultimos", response_model=dict)

def get_ultimos_pagos(

    page: int = Query(1, ge=1),

    per_page: int = Query(20, ge=1, le=100),

    cedula: Optional[str] = Query(None),

    estado: Optional[str] = Query(None),

    db: Session = Depends(get_db),

):

    """

    Resumen de últimos pagos por cédula (para PagosListResumen).

    Items: cedula, pago_id, prestamo_id, estado_pago, monto_ultimo_pago, fecha_ultimo_pago,

    cuotas_atrasadas, saldo_vencido, total_prestamos.

    """

    hoy = _hoy_local()

    # Subconsulta: cédulas distintas ordenadas por pago más reciente (más actual a más antiguo)

    subq = (

        select(

            Pago.cedula_cliente,

            func.max(Pago.fecha_pago).label("max_fecha"),

            func.max(Pago.id).label("max_id"),

        )

        .where(

            Pago.cedula_cliente.isnot(None),

            Pago.cedula_cliente != "",

        )

        .group_by(Pago.cedula_cliente)

    )

    if cedula and cedula.strip():

        subq = subq.where(Pago.cedula_cliente.ilike(f"%{cedula.strip()}%"))

    subq = subq.subquery()

    total_cedulas = db.scalar(select(func.count()).select_from(subq)) or 0

    total_pages = (total_cedulas + per_page - 1) // per_page if total_cedulas else 0

    q_cedulas = (

        select(subq.c.cedula_cliente)

        .order_by(subq.c.max_fecha.desc(), subq.c.max_id.desc())

        .offset((page - 1) * per_page)

        .limit(per_page)

    )

    cedulas_page = db.execute(q_cedulas).scalars().all()

    cedulas_list = [c[0] for c in cedulas_page if c[0]]

    items = []

    for ced in cedulas_list:

        row_ultimo = db.execute(

            select(Pago)

            .where(Pago.cedula_cliente == ced)

            .order_by(Pago.id.desc())

            .limit(1)

        ).first()

        ultimo = row_ultimo[0] if row_ultimo else None

        if not ultimo:

            continue

        if estado and estado.strip() and (ultimo.estado or "").upper() != estado.strip().upper():

            continue

        prestamo_id = ultimo.prestamo_id

        # Cuotas atrasadas y saldo vencido para este cliente (por prestamos con esa cedula)

        prestamos_cliente = db.execute(

            select(Prestamo.id).select_from(Prestamo).join(Cliente, Prestamo.cliente_id == Cliente.id).where(Cliente.cedula == ced)

        ).scalars().all()

        prestamo_ids = [p[0] for p in prestamos_cliente]

        cuotas_atrasadas = 0

        saldo_vencido = 0.0

        if prestamo_ids:

            q_mora = (

                select(func.count(), func.coalesce(func.sum(Cuota.monto), 0))

                .select_from(Cuota)

                .where(

                    Cuota.prestamo_id.in_(prestamo_ids),

                    Cuota.fecha_pago.is_(None),

                    Cuota.fecha_vencimiento < hoy,

                )

            )

            row_mora = db.execute(q_mora).first()

            if row_mora:

                cuotas_atrasadas = int(row_mora[0] or 0)

                saldo_vencido = _safe_float(row_mora[1])

        total_prestamos = len(prestamo_ids)

        items.append({

            "cedula": ced,

            "pago_id": ultimo.id,

            "prestamo_id": prestamo_id,

            "estado_pago": ultimo.estado or "PENDIENTE",

            "monto_ultimo_pago": _safe_float(ultimo.monto_pagado),

            "fecha_ultimo_pago": ultimo.fecha_pago.date().isoformat() if hasattr(ultimo.fecha_pago, "date") and ultimo.fecha_pago else (ultimo.fecha_pago.isoformat()[:10] if ultimo.fecha_pago else None),

            "cuotas_atrasadas": cuotas_atrasadas,

            "saldo_vencido": saldo_vencido,

            "total_prestamos": total_prestamos,

        })

    return {

        "items": items,

        "total": total_cedulas,

        "page": page,

        "per_page": per_page,

        "total_pages": total_pages,

    }





@router.post("/upload", response_model=dict)

async def upload_excel_pagos(

    file: UploadFile = File(..., alias="file"),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Carga masiva de pagos desde Excel (subir y procesar todo en el servidor).

    Formatos de columnas soportados (primera fila = cabecera; datos desde fila 2):

    - Formato D (principal): Cédula | Monto | Fecha | Nº documento [| Código opcional]

    - Formato E: Banco | Cédula | Fecha | Monto | Nº documento [| Código opcional]

    - Formato A: Documento | Cédula | Fecha | Monto [| Código opcional]

    - Formato B: Fecha | Cédula | Monto | Documento [| Código opcional]

    - Formato C: Cédula | ID Préstamo | Fecha | Monto | Nº documento [| Código opcional]

    Recomendado: hasta 2.500 filas para evitar timeouts; máximo 10.000.

    Timeout: en producción, configurar timeout del servidor (uvicorn/gunicorn o proxy)

    suficientemente alto para archivos grandes (p. ej. 120 s) para POST /pagos/upload.

    """

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):

        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB (alineado con frontend)

    MAX_ROWS = 10000  # Límite máximo de filas (rechazo si se supera)

    MAX_ROWS_RECOMENDADO = 2500  # Sin sobrecarga ni timeouts en servidor típico

    try:

        import openpyxl

        content = await file.read()

        if len(content) > MAX_FILE_SIZE:

            raise HTTPException(

                status_code=400,

                detail=f"El archivo es demasiado grande. Tamano maximo: {MAX_FILE_SIZE // (1024 * 1024)} MB.",

            )

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        ws = wb.active

        if not ws:

            return {"message": "Archivo sin hojas", "registros_procesados": 0, "errores": []}

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if len(rows) > MAX_ROWS:

            raise HTTPException(

                status_code=400,

                detail=f"El archivo tiene demasiadas filas ({len(rows)}). Máximo permitido: {MAX_ROWS}. Para evitar sobrecarga, se recomienda hasta {MAX_ROWS_RECOMENDADO} filas.",

            )

        if len(rows) > MAX_ROWS_RECOMENDADO:

            logger.warning(

                "Carga masiva con %s filas (recomendado hasta %s). Puede haber timeouts o lentitud.",

                len(rows),

                MAX_ROWS_RECOMENDADO,

            )



        def _looks_like_cedula(v: Any) -> bool:

            """Cédula válida: solo V, E o J + 6-11 dígitos (no se admite Z)."""

            if v is None:

                return False

            s = str(v).strip()

            return bool(re.match(r"^[VEJ]\d{6,11}$", s, re.IGNORECASE))



        def _looks_like_documento(v: Any) -> bool:

            """True si el valor puede ser Nº documento. REGLA: aceptar TODOS los formatos; única restricción = no duplicados."""

            if v is None or (isinstance(v, str) and not v.strip()):

                return False

            s = _celda_a_string_documento(v)

            if not s:

                return False

            if _looks_like_cedula(v):

                return False  # No confundir cédula con documento

            # Cualquier otro valor no vacío (1+ caracteres, hasta límite BD) se acepta como documento

            return len(s) <= _MAX_LEN_NUMERO_DOCUMENTO



        def _looks_like_date(v: Any) -> bool:

            if v is None:

                return False

            if isinstance(v, (datetime, date)):

                return True

            s = str(v).strip()

            return bool(re.search(r"\d{1,4}[-\/]\d{1,2}[-\/]\d{1,4}", s))



        def _celda_parece_banco_excel(v: Any) -> bool:

            """Primera columna tipo nombre de banco (p. ej. BINANCE, BNC), no cédula ni fecha ni referencia larga."""

            if v is None:

                return False

            s = str(v).strip()

            if not s or len(s) > 80:

                return False

            if _looks_like_date(v):

                return False

            if _looks_like_cedula(v):

                return False

            if re.search(r"\d{10,}", s):

                return False

            if len(s) > 28 and re.search(r"\d{4,}", s):

                return False

            return True



        def _extraer_documento_de_fila(row: tuple, col_documento: Optional[int]) -> str:

            """Obtiene el valor de documento: primero columna indicada; si vacío, busca en todas las celdas (fallback)."""

            if col_documento is not None and col_documento < len(row) and row[col_documento] is not None:

                s = _celda_a_string_documento(row[col_documento])

                if (s or "").strip():

                    return s

            for idx, cell in enumerate(row):

                if cell is None:

                    continue

                if _looks_like_documento(cell):

                    s = _celda_a_string_documento(cell)

                    if (s or "").strip():

                        return s

            return ""





        # Initialize error tracking

        errores: list[str] = []

        errores_detalle: list[dict] = []

        def _parse_fecha(v: Any) -> date:

            if isinstance(v, datetime):

                return v.date()

            if isinstance(v, date):

                return v

            if v is None:

                return date.today()

            s = str(v).strip()

            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):

                try:

                    return datetime.strptime(s[:10], fmt).date()

                except ValueError:

                    continue

            return date.today()



        # --- FASE 1: Parsear todas las filas (ingresar todos los datos para validar después) ---

        FilasParseadas: list[dict] = []

        filas_omitidas = 0

        pagos_con_error_list: list[dict] = []  # Filas con error para guardar en pagos_con_errores

        errores: list[str] = []

        errores_detalle: list[dict] = []

        for i, row in enumerate(rows):

            if not row or all(cell is None for cell in row):

                continue

            try:

                cedula = ""

                prestamo_id: Optional[int] = None

                fecha_val: Any = None

                monto = 0.0

                numero_doc = ""

                col_doc: Optional[int] = None

                codigo_doc_raw = ""

                institucion_bancaria: Optional[str] = None

                # Formato D (PRINCIPAL): Cédula, Monto, Fecha, Nº documento

                if len(row) >= 4 and _looks_like_cedula(row[0]) and row[1] is not None and _looks_like_date(row[2]):

                    cedula = str(row[0]).strip()

                    es_valido, monto, err_msg = _validar_monto(row[1])

                    if not es_valido and monto != 0.0:

                        errores.append(f'Fila {i + 2} (Formato D - Principal): {err_msg}')

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": cedula,

                            "prestamo_id": prestamo_id,

                            "fecha_val": fecha_val,

                            "monto": monto,

                            "numero_doc": numero_doc,

                            "errores": [err_msg]

                        })

                        continue

                    fecha_val = row[2]

                    numero_doc = _celda_a_string_documento(row[3]) if len(row) > 3 else ""

                    col_doc = 3

                    prestamo_id = None

                    if len(row) > 4 and row[4] is not None:

                        codigo_doc_raw = str(row[4]).strip()

                # Formato E: Banco, Cédula, Fecha, Monto [, Nº documento]

                elif (
                    len(row) >= 4
                    and _celda_parece_banco_excel(row[0])
                    and _looks_like_cedula(row[1])
                    and _looks_like_date(row[2])
                ):

                    _banco_raw = str(row[0]).strip()[:255]

                    institucion_bancaria = _banco_raw or None

                    cedula = str(row[1]).strip()

                    fecha_val = row[2]

                    es_valido, monto, err_msg = _validar_monto(row[3])



                    if not es_valido and monto != 0.0:

                        errores.append(f'Fila {i + 2} (Formato E - Banco): {err_msg}')

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": cedula,

                            "prestamo_id": None,

                            "fecha_val": fecha_val,

                            "monto": monto,

                            "numero_doc": "",

                            "errores": [err_msg],

                            "institucion_bancaria": institucion_bancaria,

                        })

                        continue

                    numero_doc = _celda_a_string_documento(row[4]) if len(row) > 4 else ""

                    col_doc = 4 if len(row) > 4 else None

                    prestamo_id = None

                    if len(row) > 5 and row[5] is not None:

                        codigo_doc_raw = str(row[5]).strip()

                # Formato A: Documento, Cédula, Fecha, Monto

                elif len(row) >= 4 and _looks_like_documento(row[0]) and _looks_like_cedula(row[1]):

                    numero_doc = _celda_a_string_documento(row[0])

                    col_doc = 0

                    if len(row) > 4 and row[4] is not None:

                        codigo_doc_raw = str(row[4]).strip()

                    cedula = str(row[1]).strip()

                    fecha_val = row[2] if len(row) > 2 else None

                    es_valido, monto, err_msg = _validar_monto(row[3]) if len(row) > 3 else (True, 0.0, '')



                    if not es_valido and monto != 0.0:

                        errores.append(f'Fila {i + 2} (Formato A): {err_msg}')

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": cedula,

                            "prestamo_id": None,

                            "fecha_val": row[2] if len(row) > 2 else None,

                            "monto": monto,

                            "numero_doc": numero_doc,

                            "errores": [err_msg],

                        })

                        continue

                # Formato B: Fecha, Cédula, Monto, Documento

                elif len(row) >= 4 and _looks_like_date(row[0]) and _looks_like_cedula(row[1]):

                    cedula = str(row[1]).strip()

                    es_valido, monto, err_msg = _validar_monto(row[2])



                    if not es_valido and monto != 0.0:

                        errores.append(f'Fila {i + 2} (Formato B): {err_msg}')

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": cedula,

                            "prestamo_id": None,

                            "fecha_val": row[0],

                            "monto": monto,

                            "numero_doc": _celda_a_string_documento(row[3]) if len(row) > 3 else "",

                            "errores": [err_msg],

                        })

                        continue

                    fecha_val = row[0]

                    numero_doc = _celda_a_string_documento(row[3])

                    col_doc = 3

                    if len(row) > 4 and row[4] is not None:

                        codigo_doc_raw = str(row[4]).strip()

                else:

                    # Fila con datos pero no matchea D, A ni B. Si primera columna no es cédula ? formato no reconocido.

                    row_has_content = any(cell is not None and str(cell).strip() for cell in row)

                    first_cell = str(row[0]).strip() if row[0] is not None else ""

                    if row_has_content and first_cell and not _looks_like_cedula(row[0]):

                        err_formato = "Formato de fila no reconocido. Use Cédula | Monto | Fecha | Documento (o los formatos soportados)."

                        errores.append(f"Fila {i + 2}: {err_formato}")

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": first_cell[:100] or "",

                            "prestamo_id": None,

                            "fecha_val": row[2] if len(row) > 2 else None,

                            "monto": 0.0,

                            "numero_doc": _celda_a_string_documento(row[4]) if len(row) > 4 else "",

                            "errores": [err_formato],

                        })

                        continue

                    # Formato C (Alternativo): Cédula, ID Préstamo, Fecha, Monto, Nº documento

                    cedula = first_cell or (str(row[0]).strip() if row[0] is not None else "")

                    _val_prestamo = row[1] if len(row) > 1 else None

                    if _val_prestamo is None:

                        prestamo_id = None

                    else:

                        _s = str(_val_prestamo).strip()

                        try:

                            _pid = int(_s) if (_s and _s.isdigit()) else None

                        except ValueError:

                            _pid = None

                        if _pid is not None and (_pid < 1 or _pid > _PRESTAMO_ID_MAX):

                            prestamo_id = None

                        else:

                            prestamo_id = _pid

                    fecha_val = row[2] if len(row) > 2 else None

                    _monto_raw = row[3] if len(row) > 3 else None

                    es_valido, monto, err_msg = _validar_monto(_monto_raw)

                    if not es_valido and monto != 0.0:

                        errores.append(f'Fila {i + 2} (Formato C): {err_msg}')

                        pagos_con_error_list.append({

                            "fila_idx": i + 2,

                            "cedula": cedula,

                            "prestamo_id": prestamo_id,

                            "fecha_val": fecha_val,

                            "monto": monto,

                            "numero_doc": _celda_a_string_documento(row[4]) if len(row) > 4 else "",

                            "errores": [err_msg],

                        })

                        continue

                    numero_doc = _celda_a_string_documento(row[4]) if len(row) > 4 else ""

                    col_doc = 4 if len(row) > 4 else None

                    if len(row) > 5 and row[5] is not None:

                        codigo_doc_raw = str(row[5]).strip()

                # Fallback: si documento vacío, buscar en cualquier celda de la fila

                if not (numero_doc or "").strip():

                    numero_doc = _extraer_documento_de_fila(row, col_doc)



                if not cedula or monto <= 0:

                    filas_omitidas += 1

                    # Guardar en lista de errores para despues guardar en BD

                    pagos_con_error_list.append({

                        "fila_idx": i + 2,

                        "cedula": cedula or "",

                        "prestamo_id": prestamo_id,

                        "fecha_val": fecha_val,

                        "monto": monto,

                        "numero_doc": numero_doc,

                        "errores": ["Cedula vacia o monto <= 0"]

                    })

                    continue



                FilasParseadas.append({

                    "fila_idx": i + 2,

                    "cedula": cedula,

                    "prestamo_id": prestamo_id,

                    "fecha_val": fecha_val,

                    "monto": monto,

                    "numero_doc_raw": (numero_doc or "").strip(),

                    "codigo_doc_raw": (codigo_doc_raw or "").strip(),

                    "institucion_bancaria": institucion_bancaria,

                })

            except Exception as e:

                errores.append(f"Fila {i + 2}: {e}")

                errores_detalle.append({"fila": i + 2, "cedula": "", "error": str(e), "datos": {}})



        # --- FASE 2: Insertar (documento+código único en BD; huella funcional por fila y lote) ---

        compuestos_archivo: set[str] = set()

        for _it in FilasParseadas:

            _c = compose_numero_documento_almacenado(

                _it.get("numero_doc_raw"),

                _it.get("codigo_doc_raw"),

            )

            if _c:

                compuestos_archivo.add(_c)

        documentos_ya_en_bd_excel: set[str] = set()

        if compuestos_archivo:

            _lista_c = list(compuestos_archivo)

            _chunk = 1000

            for _i0 in range(0, len(_lista_c), _chunk):

                _ch = _lista_c[_i0 : _i0 + _chunk]

                _ex = db.execute(select(Pago.numero_documento).where(Pago.numero_documento.in_(_ch))).scalars().all()

                documentos_ya_en_bd_excel.update(str(d) for d in _ex if d)

        numeros_doc_en_lote_excel: set[str] = set()

        registros = 0

        pagos_con_prestamo: list[Pago] = []

        huellas_lote_excel: set[tuple[int, str, str, str]] = set()

        for item in FilasParseadas:

            i = item["fila_idx"]

            cedula = item["cedula"]

            prestamo_id = item["prestamo_id"]

            fecha_val = item["fecha_val"]

            monto = item["monto"]

            numero_doc = item["numero_doc_raw"]

            ib_carga = item.get("institucion_bancaria")

            if isinstance(ib_carga, str):

                ib_carga = ib_carga.strip()[:255] or None

            else:

                ib_carga = None



            numero_doc_norm = compose_numero_documento_almacenado(

                numero_doc,

                item.get("codigo_doc_raw"),

            )

            if numero_doc_norm and numero_doc_norm in numeros_doc_en_lote_excel:

                err_msg = (

                    "Misma combinación comprobante + código repetida en este archivo. "

                    "Use códigos distintos o una sola fila por clave."

                )

                errores.append(f"Fila {i}: {err_msg}")

                errores_detalle.append(

                    {

                        "fila": i,

                        "cedula": cedula,

                        "error": err_msg,

                        "datos": {

                            "cedula": cedula,

                            "prestamo_id": prestamo_id,

                            "fecha_pago": fecha_val,

                            "monto_pagado": monto,

                            "numero_documento": numero_doc or "",

                        },

                    }

                )

                pagos_con_error_list.append(

                    {

                        "fila_idx": i,

                        "cedula": cedula or "",

                        "prestamo_id": prestamo_id,

                        "fecha_val": fecha_val,

                        "monto": monto,

                        "numero_doc": numero_doc or "",

                        "errores": [err_msg],

                    }

                )

                continue

            if numero_doc_norm and numero_doc_norm in documentos_ya_en_bd_excel:

                err_msg = "Ya existe un pago con la misma combinación comprobante + código."

                errores.append(f"Fila {i}: {err_msg}")

                errores_detalle.append(

                    {

                        "fila": i,

                        "cedula": cedula,

                        "error": err_msg,

                        "datos": {

                            "cedula": cedula,

                            "prestamo_id": prestamo_id,

                            "fecha_pago": fecha_val,

                            "monto_pagado": monto,

                            "numero_documento": numero_doc or "",

                        },

                    }

                )

                pagos_con_error_list.append(

                    {

                        "fila_idx": i,

                        "cedula": cedula or "",

                        "prestamo_id": prestamo_id,

                        "fecha_val": fecha_val,

                        "monto": monto,

                        "numero_doc": numero_doc or "",

                        "errores": [err_msg],

                    }

                )

                continue

            if numero_doc_norm:

                numeros_doc_en_lote_excel.add(numero_doc_norm)

            # Identificación automática de préstamo: si la cédula tiene exactamente 1 crédito activo, asignarlo

            if prestamo_id is None and cedula.strip():

                cedula_norm = cedula.strip().upper()

                prestamos_activos = (

                    db.execute(

                        select(Prestamo.id)

                        .select_from(Prestamo)

                        .join(Cliente, Prestamo.cliente_id == Cliente.id)

                        .where(

                            Cliente.cedula == cedula_norm,

                            Prestamo.estado == "APROBADO",

                        )

                        .order_by(Prestamo.id)

                    )

                ).scalars().all()

                count_prestamos = len(prestamos_activos)

                if count_prestamos > 1:

                    err_msg = f"Esta persona tiene {count_prestamos} préstamos. Debe indicar el ID del préstamo."

                    errores.append(f"Fila {i}: La cédula {cedula} tiene {count_prestamos} préstamos. Debe indicar el ID del préstamo.")

                    errores_detalle.append({"fila": i, "cedula": cedula, "error": err_msg, "datos": {"cedula": cedula, "prestamo_id": prestamo_id, "fecha_pago": fecha_val, "monto_pagado": monto, "numero_documento": numero_doc or ""}})

                    pagos_con_error_list.append({

                        "fila_idx": i,

                        "cedula": cedula or "",

                        "prestamo_id": prestamo_id,

                        "fecha_val": fecha_val,

                        "monto": monto,

                        "numero_doc": numero_doc or "",

                        "errores": [err_msg],

                    })

                    continue

                if count_prestamos == 1:

                    prestamo_id = prestamos_activos[0][0]



            try:

                fecha_pago = _parse_fecha(fecha_val)

                ref_pago = ((numero_doc_norm or (numero_doc or "").strip()) or "Carga")[:_MAX_LEN_NUMERO_DOCUMENTO]

                usuario_registro = _usuario_registro_desde_current_user(current_user)

                if prestamo_id:
                    msg_h_excel = conflicto_huella_para_creacion(
                        db,
                        prestamo_id=prestamo_id,
                        fecha_pago=fecha_pago,
                        monto_pagado=monto,
                        numero_documento=numero_doc_norm,
                        referencia_pago=ref_pago,
                        huellas_en_mismo_lote=huellas_lote_excel,
                    )
                    if msg_h_excel:
                        registrar_rechazo_huella_funcional()
                        err_msg = msg_h_excel
                        errores.append(f"Fila {i}: {err_msg}")
                        errores_detalle.append(
                            {
                                "fila": i,
                                "cedula": cedula,
                                "error": err_msg,
                                "datos": {
                                    "cedula": cedula,
                                    "prestamo_id": prestamo_id,
                                    "fecha_pago": fecha_val,
                                    "monto_pagado": monto,
                                    "numero_documento": numero_doc or "",
                                },
                            }
                        )
                        pagos_con_error_list.append(
                            {
                                "fila_idx": i,
                                "cedula": cedula or "",
                                "prestamo_id": prestamo_id,
                                "fecha_val": fecha_val,
                                "monto": monto,
                                "numero_doc": numero_doc or "",
                                "errores": [err_msg],
                            }
                        )
                        continue

                # Autoconciliar: pagos creados por carga Excel se marcan conciliados (aplicados a cuotas después)

                ahora_up = datetime.now(ZoneInfo(TZ_NEGOCIO))

                p = Pago(

                    cedula_cliente=cedula.strip().upper() if cedula else "",

                    prestamo_id=prestamo_id,

                    fecha_pago=datetime.combine(fecha_pago, dt_time.min),

                    monto_pagado=monto,

                    numero_documento=numero_doc_norm,

                    institucion_bancaria=ib_carga,

                    estado="PAGADO",

                    referencia_pago=ref_pago,

                    usuario_registro=usuario_registro,  # [MEJORADO] Usuario real desde JWT

                    conciliado=True,

                    fecha_conciliacion=ahora_up,

                    verificado_concordancia="SI",

                )

                db.add(p)

                registros += 1

                if prestamo_id and monto > 0:

                    pagos_con_prestamo.append(p)

            except Exception as e:

                err_msg = str(e)

                errores.append(f"Fila {i}: {e}")

                errores_detalle.append({"fila": i, "cedula": cedula, "error": err_msg, "datos": {"cedula": cedula, "prestamo_id": prestamo_id, "fecha_pago": fecha_val, "monto_pagado": monto, "numero_documento": numero_doc or ""}})

                pagos_con_error_list.append({

                    "fila_idx": i,

                    "cedula": cedula or "",

                    "prestamo_id": prestamo_id,

                    "fecha_val": fecha_val,

                    "monto": monto,

                    "numero_doc": numero_doc or "",

                    "errores": [err_msg],

                })

        db.flush()  # Asigna IDs a los pagos insertados

        # --- GUARDAR PAGOS CON ERRORES EN BD ---

        for pce_data in pagos_con_error_list:

            try:

                pce = PagoConError(

                    cedula_cliente=pce_data["cedula"].strip().upper() if pce_data["cedula"] else "",  # Normalize

                    prestamo_id=pce_data["prestamo_id"],

                    fecha_pago=datetime.combine(_parse_fecha(pce_data["fecha_val"]), dt_time.min) if pce_data["fecha_val"] else datetime.now(),

                    monto_pagado=pce_data["monto"],

                    numero_documento=normalize_documento(pce_data.get("numero_doc")),

                    institucion_bancaria=(

                        None

                        if pce_data.get("institucion_bancaria") in (None, "")

                        else str(pce_data.get("institucion_bancaria")).strip()[:255] or None

                    ),

                    estado="PENDIENTE",

                    errores_descripcion=pce_data["errores"],

                    observaciones="validacion",

                    fila_origen=pce_data["fila_idx"]

                )

                db.add(pce)

            except Exception as e:

                logger.warning(f"No se pudo guardar error de fila {pce_data['fila_idx']}: {e}")

        # Reglas de negocio: aplicar pagos con prestamo_id a cuotas

        cuotas_aplicadas = 0

        pagos_articulados = 0  # Track how many payments were successfully articulated

        for p in pagos_con_prestamo:

            try:

                cc, cp = _aplicar_pago_a_cuotas_interno(p, db)

                if cc > 0 or cp > 0:

                    p.estado = "PAGADO"

                    cuotas_aplicadas += cc + cp

                    pagos_articulados += 1

                    logger.info(f"Pago {p.id}: articulado a {cc + cp} cuota(s)")

                else:

                    logger.warning(f"Pago {p.id} (monto={p.monto_pagado}) no se pudo aplicar a cuotas")

            except Exception as e:

                logger.warning("Carga masiva: no se pudo aplicar pago id=%s a cuotas: %s", getattr(p, "id", "?"), e)

        db.commit()

        errores_limit = 50

        errores_detalle_limit = 100

        total_errores = len(errores)

        total_errores_detalle = len(errores_detalle)

        return {

            "message": "Carga finalizada",

            "registros_procesados": registros,

            "registros_con_error": len(pagos_con_error_list),

            "cuotas_aplicadas": cuotas_aplicadas,

            "pagos_articulados": pagos_articulados,  # [NUEVA] Número de pagos que se articularon a cuotas

            "filas_omitidas": filas_omitidas,

            "pagos_con_errores": [

                {

                    "id": pce.id,

                    "fila_origen": pce.fila_origen,

                    "cedula": pce.cedula_cliente,

                    "monto": float(pce.monto_pagado) if pce.monto_pagado else 0,

                    "errores": pce.errores_descripcion or [],

                    "accion": "revisar"

                }

                for pce in (db.query(PagoConError).filter(PagoConError.fila_origen.in_([p['fila_idx'] for p in pagos_con_error_list])).all() if pagos_con_error_list else [])

            ],

            "errores": errores[:errores_limit],

            "errores_detalle": errores_detalle[:errores_detalle_limit],

            "errores_total": total_errores,

            "errores_detalle_total": total_errores_detalle,

            "errores_truncados": total_errores > errores_limit or total_errores_detalle > errores_detalle_limit,

            "max_filas_recomendado": MAX_ROWS_RECOMENDADO,

            "max_filas_permitido": MAX_ROWS,

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error upload Excel pagos: %s", e)

        raise HTTPException(status_code=500, detail=str(e)) from e





def importar_un_pago_reportado_a_pagos(

    db: Session,

    pr: PagoReportado,

    *,

    usuario_email: str,

    documentos_ya_en_bd: set[str],

    docs_en_lote: set[str],

    registrar_error_en_tabla: bool = True,

    huellas_funcional_lote: Optional[set[tuple[int, str, str, str]]] = None,

) -> dict:

    """

    Crea un Pago desde un PagoReportado con las mismas validaciones que importar-desde-cobros.

    El Nº documento del pago toma `numero_operacion` del reporte cuando viene informado;
    si no, `referencia_interna` (RPC-…), para alinear duplicados con el comprobante real.

    No hace commit ni aplica a cuotas (el lote o el caller aplican después).

    Retorna {"ok": True, "pago": Pago} o {"ok": False, "error": str, "referencia": ..., "pce_id": int|None}.

    """

    ref_display = (pr.referencia_interna or "").strip()[:100] or None

    conciliado_reportado = True



    def _err(msg: str) -> dict:

        return {"ok": False, "error": msg, "referencia": pr.referencia_interna, "pce_id": None}



    def _err_con_pce(msg: str, **pce_kw) -> dict:

        if not registrar_error_en_tabla:

            return _err(msg)

        ref = (pr.referencia_interna or "").strip()[:100] or "N/A"

        pce = DatosImportadosConErrores(

            cedula_cliente=pce_kw.get("cedula_cliente", ""),

            prestamo_id=pce_kw.get("prestamo_id"),

            fecha_pago=datetime.combine(pr.fecha_pago, dt_time.min) if pr.fecha_pago else datetime.now(),

            monto_pagado=float(pr.monto or 0),

            numero_documento=(documento_numero_desde_pago_reportado(pr)[0] or pr.referencia_interna or "")[:100],

            estado="PAGADO" if conciliado_reportado else "PENDIENTE",

            referencia_pago=ref,

            errores_descripcion=[msg],

            observaciones=ORIGEN_COBROS_REPORTADOS,

            referencia_interna=(pr.referencia_interna or "")[:100] or None,

            fila_origen=pr.id,

        )

        db.add(pce)

        db.flush()

        return {"ok": False, "error": msg, "referencia": pr.referencia_interna, "pce_id": pce.id}



    cedula_raw = f"{pr.tipo_cedula or ''}{pr.numero_cedula or ''}".replace("-", "").strip().upper()

    if not cedula_raw:

        return _err_con_pce("Cédula vacía", cedula_cliente="")



    if not _looks_like_cedula_inline(cedula_raw):

        return _err_con_pce(

            "Cédula inválida. Formato: V, E o J + 6-11 dígitos (igual que Pagos desde Excel).",

            cedula_cliente=cedula_raw,

        )



    numero_doc_raw, numero_doc_norm = documento_numero_desde_pago_reportado(pr)

    claves_pr = claves_documento_pago_para_reportado(pr)

    if claves_pr:

        if any(k in documentos_ya_en_bd for k in claves_pr):

            return _err_con_pce(

                "Ya existe un pago enlazado a este comprobante o referencia (duplicado en BD)",

                cedula_cliente=cedula_raw,

            )

        if any(k in docs_en_lote for k in claves_pr):

            return _err_con_pce(

                "Duplicado en este lote: mismo comprobante o referencia que otra fila",

                cedula_cliente=cedula_raw,

            )



    cliente = db.execute(

        select(Cliente).where(func.replace(Cliente.cedula, "-", "") == cedula_raw)

    ).scalars().first()

    if not cliente:

        return _err_con_pce("Cédula no encontrada en clientes", cedula_cliente=cedula_raw)



    prestamos = db.execute(

        select(Prestamo)

        .where(Prestamo.cliente_id == cliente.id, Prestamo.estado == "APROBADO")

        .order_by(Prestamo.id)

    ).scalars().all()

    prestamos = [p for p in prestamos if p is not None]

    if len(prestamos) == 0:

        return _err_con_pce("Sin crédito activo (APROBADO)", cedula_cliente=cedula_raw)

    if len(prestamos) > 1:

        return _err_con_pce(

            f"Cédula con {len(prestamos)} préstamos; indique ID del crédito",

            cedula_cliente=cedula_raw,

        )



    prestamo_id = prestamos[0].id

    moneda_pr = (getattr(pr, "moneda", None) or "USD").strip().upper()

    if moneda_pr == "USDT":

        moneda_pr = "USD"

    monto = float(pr.monto or 0)

    monto_bs_original = None

    tasa_aplicada = None

    fecha_tasa_ref = None

    if moneda_pr == "BS":

        if not cedula_autorizada_para_bs(db, cedula_raw):

            return _err_con_pce(

                "Cedula no autorizada para pagos en bolivares",

                cedula_cliente=cedula_raw,

                prestamo_id=prestamo_id,

            )

        if not pr.fecha_pago:

            return _err_con_pce(

                "Fecha de pago requerida para convertir bolivares a USD",

                cedula_cliente=cedula_raw,

                prestamo_id=prestamo_id,

            )

        tasa_obj = obtener_tasa_por_fecha(db, pr.fecha_pago)

        if tasa_obj is None:

            return _err_con_pce(

                "No hay tasa de cambio registrada para la fecha de pago "

                f"{pr.fecha_pago.isoformat()}; no se puede importar en bolivares",

                cedula_cliente=cedula_raw,

                prestamo_id=prestamo_id,

            )

        tasa_aplicada = float(tasa_obj.tasa_oficial)

        monto_bs_original = monto

        try:

            monto = convertir_bs_a_usd(monto_bs_original, tasa_aplicada)

        except ValueError as e:

            return _err_con_pce(str(e), cedula_cliente=cedula_raw, prestamo_id=prestamo_id)

        fecha_tasa_ref = pr.fecha_pago

    elif moneda_pr != "USD":

        return _err_con_pce(

            f"Moneda no soportada en importacion: {moneda_pr}",

            cedula_cliente=cedula_raw,

            prestamo_id=prestamo_id,

        )

    if monto < _MIN_MONTO_PAGADO:

        return _err_con_pce(

            f"Monto debe ser mayor a {_MIN_MONTO_PAGADO}",

            cedula_cliente=cedula_raw,

            prestamo_id=prestamo_id,

        )



    ref_pago = (numero_doc_norm or numero_doc_raw or "Cobros")[:_MAX_LEN_NUMERO_DOCUMENTO]

    ahora_imp = datetime.now(ZoneInfo(TZ_NEGOCIO))

    fecha_huella_cobros = (
        pr.fecha_pago
        if pr.fecha_pago is not None
        else datetime.now(ZoneInfo(TZ_NEGOCIO)).date()
    )

    monto_dec_cobros = Decimal(str(round(monto, 2)))

    msg_h_cobros = conflicto_huella_para_creacion(
        db,
        prestamo_id=prestamo_id,
        fecha_pago=fecha_huella_cobros,
        monto_pagado=monto_dec_cobros,
        numero_documento=numero_doc_norm,
        referencia_pago=ref_pago,
        huellas_en_mismo_lote=huellas_funcional_lote,
    )

    if msg_h_cobros:
        registrar_rechazo_huella_funcional()
        return _err_con_pce(msg_h_cobros, cedula_cliente=cedula_raw, prestamo_id=prestamo_id)

    rpc_tr = (pr.referencia_interna or "").strip()[:100]
    notas_pago = None
    if rpc_tr and numero_doc_raw and rpc_tr != numero_doc_raw:
        notas_pago = f"Ref. interna reporte: {rpc_tr}"

    img_id = (getattr(pr, "comprobante_imagen_id", None) or "").strip()
    link_comp = url_comprobante_imagen_absoluta(img_id) if img_id else None
    doc_nom = ((pr.comprobante_nombre or "").strip()[:255] or None) if img_id else None

    p = Pago(

        cedula_cliente=cedula_raw,

        prestamo_id=prestamo_id,

        fecha_pago=datetime.combine(pr.fecha_pago, dt_time.min) if pr.fecha_pago else datetime.now(),

        monto_pagado=monto_dec_cobros,

        numero_documento=numero_doc_norm,

        # Evita violar chk_pagos_conciliado_pendiente_inconsistente
        estado="PAGADO",

        referencia_pago=ref_pago,

        notas=notas_pago,

        usuario_registro=usuario_email,

        conciliado=True,

        fecha_conciliacion=ahora_imp,

        verificado_concordancia="SI",

        moneda_registro=moneda_pr,

        monto_bs_original=Decimal(str(round(monto_bs_original, 2))) if monto_bs_original is not None else None,

        tasa_cambio_bs_usd=Decimal(str(tasa_aplicada)) if tasa_aplicada is not None else None,

        fecha_tasa_referencia=fecha_tasa_ref,

        link_comprobante=link_comp,

        documento_nombre=doc_nom,

    )

    db.add(p)

    db.flush()

    for _k in claves_pr:

        if _k:

            docs_en_lote.add(_k)

            documentos_ya_en_bd.add(_k)

    return {"ok": True, "pago": p, "referencia": ref_display, "pce_id": None}







ORIGEN_COBROS_REPORTADOS = "Cobros (reportados aprobados)"





@router.post("/importar-desde-cobros", response_model=dict)

def importar_reportados_aprobados_a_pagos(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Pasa los pagos reportados (módulo Cobros) con estado aprobado a la tabla pagos.

    Mismas reglas que carga masiva: documento único, cédula con 1 crédito activo (APROBADO).

    Los que no cumplen se guardan en datos_importados_conerrores; descargar Excel desde el frontend.

    Los pagos válidos se aplican a cuotas del préstamo (asignación en cascada) en el mismo request.

    """

    rows = db.execute(

        select(PagoReportado).where(PagoReportado.estado == "aprobado").order_by(PagoReportado.id)

    ).scalars().all()

    reportados = [r for r in rows if r is not None]

    if not reportados:

        return {

            "registros_procesados": 0,

            "registros_con_error": 0,

            "errores_detalle": [],

            "ids_pagos_con_errores": [],

            "total_datos_revisar": 0,

            "mensaje": "No hay pagos reportados aprobados para importar.",

            "cuotas_aplicadas": 0,

            "operaciones_cuota_total": 0,

            "pagos_con_aplicacion_a_cuotas": 0,

            "pagos_sin_aplicacion_cuotas": [],

            "pagos_sin_aplicacion_cuotas_total": 0,

            "pagos_sin_aplicacion_cuotas_truncados": False,

        }



    usuario_registro = _usuario_registro_desde_current_user(current_user)

    documentos_ya_en_bd: set[str] = set()

    todas_claves_lote = claves_documento_para_lote_reportados(reportados)

    if todas_claves_lote:

        cl_list = list(todas_claves_lote)

        _chunk_claves = 800

        for i in range(0, len(cl_list), _chunk_claves):

            chunk = cl_list[i : i + _chunk_claves]

            existentes = db.execute(

                select(Pago.numero_documento).where(Pago.numero_documento.in_(chunk))

            ).scalars().all()

            documentos_ya_en_bd.update(str(d) for d in existentes if d)



    registros_procesados = 0

    ids_pagos_con_errores: list[int] = []

    errores_detalle: list[dict] = []

    docs_en_lote: set[str] = set()

    huellas_funcional_lote_cobros: set[tuple[int, str, str, str]] = set()

    pagos_creados: list[Pago] = []



    for pr in reportados:

        res = importar_un_pago_reportado_a_pagos(

            db,

            pr,

            usuario_email=usuario_registro,

            documentos_ya_en_bd=documentos_ya_en_bd,

            docs_en_lote=docs_en_lote,

            registrar_error_en_tabla=True,

            huellas_funcional_lote=huellas_funcional_lote_cobros,

        )

        if res.get("ok"):

            registros_procesados += 1

            pagos_creados.append(res["pago"])

            continue

        pce_id = res.get("pce_id")

        if pce_id is not None:

            ids_pagos_con_errores.append(pce_id)

        errores_detalle.append(

            {"referencia": res.get("referencia"), "error": res.get("error", "Error")}

        )





    operaciones_cuota_total = 0

    pagos_con_aplicacion_a_cuotas = 0

    pagos_sin_aplicacion_cuotas: list[dict] = []

    for p in pagos_creados:

        pid = getattr(p, "id", None)

        ced = (getattr(p, "cedula_cliente", None) or "") or ""

        pr_id = getattr(p, "prestamo_id", None)

        try:

            cc, cp = _aplicar_pago_a_cuotas_interno(p, db)

            p.estado = _estado_pago_tras_aplicar_cascada(cc, cp)

            if cc > 0 or cp > 0:

                operaciones_cuota_total += cc + cp

                pagos_con_aplicacion_a_cuotas += 1

            elif pr_id and float(p.monto_pagado or 0) > 0:

                pagos_sin_aplicacion_cuotas.append(

                    {

                        "pago_id": pid,

                        "cedula_cliente": ced,

                        "prestamo_id": pr_id,

                        "motivo": "sin_cuotas_afectadas",

                        "detalle": "Ninguna cuota recibió monto; revise cuotas pendientes o use Aplicar a cuotas.",

                    }

                )

        except Exception as e:

            logger.warning(

                "Importar Cobros: no se pudo aplicar pago id=%s a cuotas: %s",

                getattr(p, "id", "?"),

                e,

            )

            pagos_sin_aplicacion_cuotas.append(

                {

                    "pago_id": pid,

                    "cedula_cliente": ced,

                    "prestamo_id": pr_id,

                    "motivo": "error",

                    "detalle": str(e),

                }

            )

    db.commit()

    total_datos_revisar = db.execute(select(func.count()).select_from(DatosImportadosConErrores)).scalar() or 0

    n_sin_aplicacion = len(pagos_sin_aplicacion_cuotas)

    _limite_sin_aplicacion = 50

    mensaje_base = (

        f"Importados {registros_procesados} pagos desde Cobros; "

        f"{len(ids_pagos_con_errores)} con error (revisar: descargar Excel)."

    )

    if n_sin_aplicacion > 0:

        mensaje_base += (

            f" Atención: {n_sin_aplicacion} pago(s) creado(s) pero sin aplicar a cuotas "

            "(error o sin cuotas pendientes según reglas); use Aplicar a cuotas o revise el préstamo."

        )

    return {

        "registros_procesados": registros_procesados,

        "registros_con_error": len(ids_pagos_con_errores),

        "errores_detalle": errores_detalle[:100],

        "ids_pagos_con_errores": ids_pagos_con_errores,

        "cuotas_aplicadas": operaciones_cuota_total,

        "operaciones_cuota_total": operaciones_cuota_total,

        "pagos_con_aplicacion_a_cuotas": pagos_con_aplicacion_a_cuotas,

        "pagos_sin_aplicacion_cuotas": pagos_sin_aplicacion_cuotas[:_limite_sin_aplicacion],

        "pagos_sin_aplicacion_cuotas_total": n_sin_aplicacion,

        "pagos_sin_aplicacion_cuotas_truncados": n_sin_aplicacion > _limite_sin_aplicacion,

        "total_datos_revisar": total_datos_revisar,

        "mensaje": mensaje_base,

    }





@router.get("/importar-desde-cobros/datos-revisar", response_model=dict)

def get_datos_revisar_importados(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """Devuelve si hay datos que revisar (tabla datos_importados_conerrores) y el total. Para mostrar dialogo tras importar."""

    total = db.execute(select(func.count()).select_from(DatosImportadosConErrores)).scalar() or 0

    return {"tiene_datos": total > 0, "total": total}





@router.get("/importar-desde-cobros/descargar-excel-errores")

def descargar_excel_errores_importados(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """Genera Excel con registros de datos_importados_conerrores (mismas columnas que Revisar Pagos). Tras generar, vacia la tabla."""

    rows = db.execute(

        select(DatosImportadosConErrores).order_by(DatosImportadosConErrores.id)

    ).scalars().all()

    rows = [r for r in rows if r is not None]

    import openpyxl

    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Datos con errores"

    headers = [

        "id", "cedula_cliente", "prestamo_id", "fecha_pago", "monto_pagado", "numero_documento",

        "estado", "referencia_pago", "errores_descripcion", "observaciones", "fila_origen", "referencia_interna", "created_at"

    ]

    for col, h in enumerate(headers, 1):

        ws.cell(row=1, column=col, value=h)

    for row_idx, r in enumerate(rows, 2):

        fp = r.fecha_pago

        fecha_str = fp.strftime("%Y-%m-%d") if fp else ""

        err_str = "; ".join(r.errores_descripcion) if isinstance(r.errores_descripcion, list) else str(r.errores_descripcion or "")

        created_str = r.created_at.strftime("%Y-%m-%d %H:%M") if getattr(r, "created_at", None) else ""

        ws.cell(row=row_idx, column=1, value=r.id)

        ws.cell(row=row_idx, column=2, value=r.cedula_cliente or "")

        ws.cell(row=row_idx, column=3, value=r.prestamo_id)

        ws.cell(row=row_idx, column=4, value=fecha_str)

        ws.cell(row=row_idx, column=5, value=float(r.monto_pagado) if r.monto_pagado is not None else 0)

        ws.cell(row=row_idx, column=6, value=r.numero_documento or "")

        ws.cell(row=row_idx, column=7, value=r.estado or "")

        ws.cell(row=row_idx, column=8, value=r.referencia_pago or "")

        ws.cell(row=row_idx, column=9, value=err_str)

        ws.cell(row=row_idx, column=10, value=r.observaciones or "")

        ws.cell(row=row_idx, column=11, value=r.fila_origen)

        ws.cell(row=row_idx, column=12, value=getattr(r, "referencia_interna", None) or "")

        ws.cell(row=row_idx, column=13, value=created_str)

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    db.execute(delete(DatosImportadosConErrores))

    db.commit()

    filename = f"datos_importados_con_errores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={filename}"},

    )





@router.get("/export/excel/pagos-sin-aplicar-cuotas")

def export_excel_pagos_sin_aplicar_cuotas(

    cohorte: str = Query(

        "todos",

        description="todos (default): sin ninguna fila en cuota_pagos (no aplicados a cuotas). "

        "cascada: ademas monto>0, prestamo_id, no ANULADO_IMPORT (cola del job). "

        "sin_cupo: como cascada y sin cupo aplicable en cuotas PENDIENTE/MORA/PARCIAL. "

        "Nota: cohorte «cascada» es el nombre vigente; «fifo» se acepta como alias (compatibilidad).",

    ),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Descarga Excel con pagos que no tienen aplicacion a cuotas (sin filas en cuota_pagos).

    Por defecto cohorte=todos incluye todos esos pagos. Requiere autenticacion. Maximo 200000 filas.

    """

    c = (cohorte or "todos").strip().lower()

    if c == "fifo":

        c = "cascada"

    if c not in ("todos", "cascada", "sin_cupo"):

        raise HTTPException(
            status_code=400,
            detail="cohorte debe ser todos, cascada o sin_cupo (fifo aceptado como alias de cascada)",
        )



    import openpyxl



    tiene_cuota_pago = exists(select(CuotaPago.id).where(CuotaPago.pago_id == Pago.id))

    if c == "todos":

        cond_final = ~tiene_cuota_pago

    else:

        cond_base = and_(

            ~tiene_cuota_pago,

            func.coalesce(Pago.monto_pagado, 0) > 0,

            func.upper(func.coalesce(Pago.estado, "")) != "ANULADO_IMPORT",

            Pago.prestamo_id.isnot(None),

        )

        if c == "sin_cupo":

            aplicado_en_cuota = (

                select(func.coalesce(func.sum(CuotaPago.monto_aplicado), 0))

                .where(CuotaPago.cuota_id == Cuota.id)

                .correlate(Cuota)

                .scalar_subquery()

            )

            hay_cupo_aplicable = exists(

                select(1)

                .select_from(Cuota)

                .where(

                    and_(

                        Cuota.prestamo_id == Pago.prestamo_id,

                        Cuota.estado.in_(["PENDIENTE", "VENCIDO", "MORA", "PARCIAL"]),

                        (func.coalesce(Cuota.monto, 0) - aplicado_en_cuota) > 0.01,

                    )

                )

            )

            cond_final = and_(cond_base, ~hay_cupo_aplicable)

        else:

            cond_final = cond_base



    rows = (

        db.execute(

            select(Pago)

            .where(cond_final)

            .order_by(Pago.fecha_registro.asc(), Pago.id.asc())

            .limit(200000)

        )

        .scalars()

        .all()

    )

    rows = [r for r in rows if r is not None]



    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Pagos sin aplicar"

    headers = [

        "pago_id",

        "fecha_registro",

        "fecha_pago",

        "prestamo_id",

        "cedula",

        "monto_pagado",

        "estado",

        "referencia_pago",

        "numero_documento",

        "conciliado",

        "usuario_registro",

        "cohorte_filtro",

    ]

    for col, h in enumerate(headers, 1):

        ws.cell(row=1, column=col, value=h)

    for row_idx, p in enumerate(rows, 2):

        fr = p.fecha_registro

        fp = p.fecha_pago

        ws.cell(row=row_idx, column=1, value=p.id)

        ws.cell(row=row_idx, column=2, value=fr.strftime("%Y-%m-%d %H:%M:%S") if fr else "")

        ws.cell(row=row_idx, column=3, value=fp.strftime("%Y-%m-%d %H:%M:%S") if fp else "")

        ws.cell(row=row_idx, column=4, value=p.prestamo_id)

        ws.cell(row=row_idx, column=5, value=(p.cedula_cliente or ""))

        ws.cell(row=row_idx, column=6, value=float(p.monto_pagado) if p.monto_pagado is not None else 0)

        ws.cell(row=row_idx, column=7, value=p.estado or "")

        ws.cell(row=row_idx, column=8, value=p.referencia_pago or "")

        ws.cell(row=row_idx, column=9, value=p.numero_documento or "")

        ws.cell(row=row_idx, column=10, value=bool(p.conciliado) if p.conciliado is not None else False)

        ws.cell(row=row_idx, column=11, value=p.usuario_registro or "")

        ws.cell(row=row_idx, column=12, value=c)



    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    fname = f"pagos_sin_aplicar_cuotas_{c}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={fname}"},

    )





@router.post("/validar-filas-batch", response_model=dict)

def validar_filas_batch(

    body: ValidarFilasBatchBody = Body(...),

    db: Session = Depends(get_db),

):

    """

    Valida en lote (carga masiva / preview):

    - Cédulas: deben existir en tabla clientes (misma clave que FK fk_pagos_cedula al guardar pagos).

    - Nº documento: clave canónica única global (comprobante normalizado + código opcional compuesto

      como en POST /pagos); ya usada en `pagos` o en `pagos_con_errores` → duplicado.

      (Un solo registro por esa clave; las cuotas referencian ese pago vía pago_id.)

    """

    cedulas_norm = list(

        {

            c.strip().replace("-", "").upper()

            for c in (body.cedulas or [])

            if c and c.strip()

        }

    )

    cedulas_existentes: set[str] = set()

    if cedulas_norm:

        cc = func.upper(func.replace(Cliente.cedula, "-", ""))

        rows = db.execute(select(cc).where(cc.in_(cedulas_norm)).distinct()).all()

        cedulas_existentes = {r[0] for r in rows if r and r[0]}

    documentos_duplicados: list[dict] = []

    docs_norm_limpios: list[str] = []

    for d in body.documentos or []:

        if d is None:

            continue

        s = str(d).strip() if not isinstance(d, str) else (d or "").strip()

        if not s:

            continue

        n = normalize_documento(s)

        if n:

            docs_norm_limpios.append(n)

    docs_norm_limpios = list(dict.fromkeys(docs_norm_limpios))

    if docs_norm_limpios:

        doc_match_pagos = or_(

            Pago.numero_documento.in_(docs_norm_limpios),

            func.trim(Pago.numero_documento).in_(docs_norm_limpios),

        )

        rows_pagos = db.execute(

            select(

                Pago.numero_documento,

                Pago.id,

                Pago.cedula_cliente,

                Pago.fecha_pago,

                Pago.monto_pagado,

                Pago.prestamo_id,

            )

            .where(doc_match_pagos)

            .where(Pago.numero_documento.isnot(None))

            .distinct()

        ).all()

        docs_set = set(docs_norm_limpios)

        seen_pago_ids: set[int] = set()

        for row in rows_pagos:

            if normalize_documento(row[0]) not in docs_set:

                continue

            if row[1] in seen_pago_ids:

                continue

            seen_pago_ids.add(int(row[1]))

            documentos_duplicados.append(

                {

                    "numero_documento": row[0],

                    "pago_id": row[1],

                    "cedula": row[2],

                    "fecha_pago": row[3].isoformat() if row[3] else None,

                    "monto_pagado": float(row[4]) if row[4] else 0,

                    "prestamo_id": int(row[5]) if row[5] is not None else None,

                    "estado": "duplicado",

                    "origen": "pagos",

                }

            )

        doc_match_pce = or_(

            PagoConError.numero_documento.in_(docs_norm_limpios),

            func.trim(PagoConError.numero_documento).in_(docs_norm_limpios),

        )

        rows_pce = db.execute(

            select(

                PagoConError.numero_documento,

                PagoConError.id,

                PagoConError.cedula_cliente,

                PagoConError.fecha_pago,

                PagoConError.monto_pagado,

                PagoConError.prestamo_id,

            )

            .where(doc_match_pce)

            .where(PagoConError.numero_documento.isnot(None))

            .distinct()

        ).all()

        seen_pce_ids: set[int] = set()

        for row in rows_pce:

            if normalize_documento(row[0]) not in docs_set:

                continue

            if row[1] in seen_pce_ids:

                continue

            seen_pce_ids.add(int(row[1]))

            documentos_duplicados.append(

                {

                    "numero_documento": row[0],

                    "pago_con_error_id": row[1],

                    "cedula": row[2],

                    "fecha_pago": row[3].isoformat() if row[3] else None,

                    "monto_pagado": float(row[4]) if row[4] else 0,

                    "prestamo_id": int(row[5]) if row[5] is not None else None,

                    "estado": "duplicado",

                    "origen": "pagos_con_errores",

                }

            )

    return {

        "cedulas_existentes": list(cedulas_existentes),

        "documentos_duplicados": documentos_duplicados,

    }





@router.post("/guardar-fila-editable", response_model=dict)
@router.put("/guardar-fila-editable", response_model=dict)

def guardar_fila_editable(

    body: GuardarFilaEditableBody = Body(...),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Guarda una fila editable validada (desde Preview).

    Si cumple validadores, inserta en pagos, aplica cuotas y retorna éxito.

    Auto-marca como conciliado ('SI') para aplicar reglas de negocio inmediatamente.

    """

    try:

        cedula = (body.cedula or "").strip()

        monto = body.monto_pagado

        numero_doc = (body.numero_documento or "").strip() if body.numero_documento else None

        codigo_doc = (body.codigo_documento or "").strip() if getattr(body, "codigo_documento", None) else None

        prestamo_id = body.prestamo_id

        usuario_registro = _usuario_registro_desde_current_user(current_user)



        # Validaciones post-guardado

        if not cedula:

            raise HTTPException(status_code=400, detail="Cédula requerida")

        if not _looks_like_cedula_inline(cedula):

            raise HTTPException(status_code=400, detail="Cédula inválida (debe ser V/E/J/Z + 6-11 dígitos)")


        cedula_norm_in = normalizar_cedula_almacenamiento(cedula)
        if not cedula_norm_in:
            raise HTTPException(status_code=400, detail="Cédula requerida")
        cn_in = cedula_norm_in.replace("-", "").upper()
        cc_cli = func.upper(func.replace(Cliente.cedula, "-", ""))
        cli_por_cedula = db.execute(
            select(Cliente.id).where(cc_cli == cn_in).limit(1)
        ).first()
        if not cli_por_cedula:
            raise HTTPException(
                status_code=400,
                detail="La cédula no está registrada en clientes. No se puede guardar el pago.",
            )

        if monto <= 0:

            raise HTTPException(status_code=400, detail="Monto debe ser > 0")

        if monto > _MAX_MONTO_PAGADO:

            raise HTTPException(status_code=400, detail=f"Monto excede límite máximo: {_MAX_MONTO_PAGADO}")



        # Parsear fecha

        try:

            fecha_pago = datetime.strptime(body.fecha_pago[:10], "%d-%m-%Y").date()

        except (ValueError, IndexError):

            raise HTTPException(status_code=400, detail="Fecha inválida (formato: DD-MM-YYYY)")



        # Comprobante + código opcional → valor único almacenado

        numero_doc_norm = compose_numero_documento_almacenado(numero_doc, codigo_doc)

        if numero_doc_norm and numero_documento_ya_registrado(db, numero_doc_norm):

            raise HTTPException(

                status_code=409,

                detail=(

                    "Conflicto numero_documento: ya existe un pago con la misma combinación "

                    "comprobante + código."

                ),

            )

        # Si prestamo_id es None, buscar por cédula en préstamos (normalizada)

        if prestamo_id is None:

            ced_norm = (

                cedula.strip().replace("-", "").upper()

                if cedula

                else ""

            )

            if ced_norm:

                pc = func.upper(func.replace(Prestamo.cedula, "-", ""))

                prest_row = db.execute(

                    select(Prestamo.id)

                    .where(pc == ced_norm)

                    .order_by(Prestamo.id.desc())

                    .limit(1)

                ).first()

                if prest_row:

                    prestamo_id = prest_row[0]



        if prestamo_id is None:

            raise HTTPException(

                status_code=400,

                detail="La cédula no tiene préstamo asociado; registre el crédito antes de cargar el pago.",

            )



        # Cedula en pagos debe existir en clientes (FK fk_pagos_cedula): usar la del cliente del prestamo.
        prest = db.get(Prestamo, prestamo_id)
        if not prest:
            raise HTTPException(status_code=404, detail="Prestamo no encontrado")
        cli = db.get(Cliente, prest.cliente_id)
        if not cli:
            raise HTTPException(
                status_code=400,
                detail="El prestamo no tiene cliente asociado en BD; no se puede registrar el pago.",
            )
        # Datos legacy: clientes.cedula en minusculas; FK pagos.cedula debe coincidir con clientes.cedula
        ced_norm_cli = normalizar_cedula_almacenamiento(cli.cedula)
        if ced_norm_cli and ced_norm_cli != (cli.cedula or ""):
            cli.cedula = ced_norm_cli
            db.flush()
        cedula_fk = normalizar_cedula_almacenamiento(cli.cedula) or normalizar_cedula_almacenamiento(
            prest.cedula
        )
        if not cedula_fk:
            raise HTTPException(status_code=400, detail="Cedula del cliente no disponible en BD.")
        cedula_input = normalizar_cedula_almacenamiento(cedula.strip())
        cn_body = (cedula_input or "").replace("-", "").upper()
        cn_fk = (cedula_fk or "").replace("-", "").upper()
        if cedula_input and cn_body != cn_fk:
            raise HTTPException(
                status_code=400,
                detail=f"La cedula no coincide con la del cliente del prestamo (en BD: {cedula_fk}).",
            )



        # Crear pago

        # [A2] Marcar conciliado=True y verificado_concordancia="SI" desde el momento de la creación,

        # ya que guardar-fila-editable implica que el pago fue revisado y validado manualmente.

        ref_pago = (numero_doc_norm or (numero_doc or "Carga"))[:_MAX_LEN_NUMERO_DOCUMENTO]

        ahora_conciliacion = datetime.now(ZoneInfo(TZ_NEGOCIO))

        moneda_r = (body.moneda_registro or "USD").strip().upper()

        tasa_man_dec = None

        if body.tasa_cambio_manual is not None:

            tasa_man_dec = Decimal(str(body.tasa_cambio_manual))

        monto_usd_g, moneda_fin_g, monto_bs_g, tasa_g, fecha_tasa_g = resolver_monto_registro_pago(

            db,

            cedula_normalizada=(cedula_fk or "").strip().upper(),

            fecha_pago=fecha_pago,

            monto_pagado=Decimal(str(round(monto, 2))),

            moneda_registro=moneda_r,

            tasa_cambio_manual=tasa_man_dec,

        )

        msg_h_fila = conflicto_huella_para_creacion(
            db,
            prestamo_id=prestamo_id,
            fecha_pago=fecha_pago,
            monto_pagado=monto_usd_g,
            numero_documento=numero_doc_norm,
            referencia_pago=ref_pago,
        )
        if msg_h_fila:
            registrar_rechazo_huella_funcional()
            raise HTTPException(status_code=409, detail=msg_h_fila)

        pago = Pago(

            cedula_cliente=cedula_fk,

            prestamo_id=prestamo_id,

            fecha_pago=datetime.combine(fecha_pago, dt_time.min),

            monto_pagado=monto_usd_g,

            numero_documento=numero_doc_norm,

            estado="PAGADO",

            referencia_pago=ref_pago,

            conciliado=True,  # [B2] Usar solo conciliado

            fecha_conciliacion=ahora_conciliacion,

            verificado_concordancia="SI",  # Legacy: sync con conciliado

            usuario_registro=usuario_registro,

            moneda_registro=moneda_fin_g,

            monto_bs_original=monto_bs_g,

            tasa_cambio_bs_usd=tasa_g,

            fecha_tasa_referencia=fecha_tasa_g,

        )

        db.add(pago)

        db.flush()



        # Aplicar a cuotas si prestamo_id existe

        cuotas_completadas = 0

        cuotas_parciales = 0

        if pago.prestamo_id and float(pago.monto_pagado or 0) > 0:

            cuotas_completadas, cuotas_parciales = _aplicar_pago_a_cuotas_interno(pago, db)

        pago.estado = _estado_conciliacion_post_cascada(pago, cuotas_completadas, cuotas_parciales)



        db.commit()



        return {

            "success": True,

            "pago_id": pago.id,

            "message": "Fila guardada exitosamente",

            "cuotas_completadas": cuotas_completadas,

            "cuotas_parciales": cuotas_parciales,

        }



    except HTTPException:

        raise

    except ValueError as e:

        db.rollback()

        logger.warning("guardar-fila-editable integridad: %s", e)

        raise HTTPException(status_code=409, detail=str(e)) from e

    except Exception as e:

        db.rollback()

        logger.exception("Error guardar-fila-editable: %s", e)

        raise HTTPException(status_code=500, detail=str(e)) from e





def _looks_like_cedula_inline(cedula: str) -> bool:

    """Validar cédula inline: solo V, E o J + 6-11 dígitos (no se admite Z)."""

    return bool(re.match(r"^[VEJ]\d{6,11}$", cedula.strip(), re.IGNORECASE))





@router.post("/conciliacion/upload", response_model=dict)

async def upload_conciliacion(

    file: UploadFile = File(..., alias="file"),

    db: Session = Depends(get_db),

):

    """

    Carga archivo de conciliación (Excel: Fecha de Depósito, Número de Documento).

    Marca pagos encontrados por numero_documento como conciliados.

    """

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):

        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")

    try:

        import openpyxl

        content = await file.read()

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        ws = wb.active

        if not ws:

            return {

                "pagos_conciliados": 0,

                "pagos_no_encontrados": 0,

                "documentos_no_encontrados": [],

                "errores": 0,

                "errores_detalle": ["Archivo sin datos"],

            }

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        ahora = datetime.now(ZoneInfo(TZ_NEGOCIO))

        pagos_conciliados = 0

        documentos_no_encontrados = []

        errores_detalle = []

        pagos_para_aplicar: list[Pago] = []

        for i, row in enumerate(rows):

            if not row or (row[0] is None and row[1] is None):

                continue

            try:

                numero_doc_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

                if not numero_doc_raw:

                    continue

                # Misma clave canónica que carga/crear: cualquier formato reconocido, búsqueda por valor normalizado

                numero_doc = normalize_documento(numero_doc_raw)

                pago_row = db.execute(

                    select(Pago).where(Pago.numero_documento == numero_doc).limit(1)

                ).first()

                if not pago_row:

                    documentos_no_encontrados.append(numero_doc_raw)

                    continue

                pago = pago_row[0]

                pago.conciliado = True

                pago.fecha_conciliacion = ahora

                pagos_conciliados += 1

                if pago.prestamo_id and float(pago.monto_pagado or 0) > 0:

                    pagos_para_aplicar.append(pago)

            except Exception as e:

                errores_detalle.append(f"Fila {i + 2}: {e}")

        cuotas_aplicadas = 0

        for pago in pagos_para_aplicar:

            try:

                cc, cp = _aplicar_pago_a_cuotas_interno(pago, db)

                if cc > 0 or cp > 0:

                    pago.estado = "PAGADO"

                    cuotas_aplicadas += cc + cp

            except Exception as e:

                logger.warning("Conciliacion: no se pudo aplicar pago id=%s a cuotas: %s", pago.id, e)

        db.commit()

        return {

            "pagos_conciliados": pagos_conciliados,

            "cuotas_aplicadas": cuotas_aplicadas,

            "pagos_no_encontrados": len(documentos_no_encontrados),

            "documentos_no_encontrados": documentos_no_encontrados[:100],

            "errores": len(errores_detalle),

            "errores_detalle": errores_detalle[:50],

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error upload conciliación: %s", e)

        raise HTTPException(status_code=500, detail=str(e)) from e





@router.get("/kpis")

def get_pagos_kpis(

    mes: Optional[int] = Query(None, ge=1, le=12),

    anio: Optional[int] = Query(None, ge=2000, le=2100),

    fecha_inicio: Optional[str] = Query(None),

    fecha_fin: Optional[str] = Query(None),

    db: Session = Depends(get_db),

):

    """

    KPIs de pagos para un mes:

    1. montoACobrarMes: cuánto dinero debería cobrarse en el mes (cuotas con vencimiento en el mes).

    2. montoCobradoMes: cuánto dinero se ha cobrado = pagado en el mes.

    3. morosidadMensualPorcentaje: pago vencido mensual en % (cuotas vencidas no cobradas / cartera * 100).

    Parámetros: mes (1-12) y anio (2000-2100). Si no se envían, se usa el mes actual.

    """

    try:

        hoy = _hoy_local()

        if mes is not None and anio is not None:

            inicio_mes = hoy.replace(year=anio, month=mes, day=1)

            _, ultimo_dia = calendar.monthrange(anio, mes)

            fin_mes = inicio_mes.replace(day=ultimo_dia)

        elif fecha_inicio and fecha_fin:

            try:

                inicio_mes = datetime.strptime(fecha_inicio[:10], "%Y-%m-%d").date()

                fin_mes = datetime.strptime(fecha_fin[:10], "%Y-%m-%d").date()

            except (ValueError, TypeError):

                inicio_mes = hoy.replace(day=1)

                _, ultimo_dia = calendar.monthrange(hoy.year, hoy.month)

                fin_mes = inicio_mes.replace(day=ultimo_dia)

        else:

            inicio_mes = hoy.replace(day=1)

            _, ultimo_dia = calendar.monthrange(hoy.year, hoy.month)

            fin_mes = inicio_mes.replace(day=ultimo_dia)

        # Fecha de referencia: para meses pasados = fin_mes; para mes actual = hoy

        fecha_referencia = min(fin_mes, hoy)



        # Condiciones base: solo clientes ACTIVOS

        conds_activo = [

            Cuota.prestamo_id == Prestamo.id,

            Prestamo.cliente_id == Cliente.id,

            Cliente.estado == "ACTIVO",

            Prestamo.estado == "APROBADO",

        ]

        base_where = and_(*conds_activo)



        # Una sola consulta con agregación condicional para los 5 montos (menos round-trips a BD)

        aggr = select(

            func.coalesce(

                func.sum(

                    case(

                        (

                            and_(

                                Cuota.fecha_vencimiento >= inicio_mes,

                                Cuota.fecha_vencimiento <= fin_mes,

                            ),

                            Cuota.monto,

                        ),

                        else_=0,

                    )

                ),

                0,

            ).label("monto_a_cobrar_mes"),

            func.coalesce(

                func.sum(

                    case(

                        (

                            and_(

                                Cuota.fecha_pago.isnot(None),

                                Cuota.fecha_pago >= inicio_mes,

                                Cuota.fecha_pago <= fecha_referencia,

                            ),

                            Cuota.monto,

                        ),

                        else_=0,

                    )

                ),

                0,

            ).label("monto_cobrado_mes"),

            func.coalesce(

                func.sum(

                    case(

                        (

                            and_(

                                Cuota.fecha_vencimiento >= inicio_mes,

                                Cuota.fecha_vencimiento <= fin_mes,

                                Cuota.fecha_vencimiento < fecha_referencia,

                            ),

                            Cuota.monto,

                        ),

                        else_=0,

                    )

                ),

                0,

            ).label("total_vencido_mes"),

            func.coalesce(

                func.sum(

                    case(

                        (

                            and_(

                                Cuota.fecha_vencimiento >= inicio_mes,

                                Cuota.fecha_vencimiento <= fin_mes,

                                Cuota.fecha_vencimiento < fecha_referencia,

                                Cuota.fecha_pago.is_(None),

                            ),

                            Cuota.monto,

                        ),

                        else_=0,

                    )

                ),

                0,

            ).label("no_cobrado_mes"),

            func.coalesce(

                func.sum(case((Cuota.fecha_pago.is_(None), Cuota.monto), else_=0)),

                0,

            ).label("cartera_pendiente"),

        ).select_from(Cuota).join(Prestamo, Cuota.prestamo_id == Prestamo.id).join(

            Cliente, Prestamo.cliente_id == Cliente.id

        ).where(base_where)



        row = db.execute(aggr).one()

        monto_a_cobrar_mes = row.monto_a_cobrar_mes or 0

        monto_cobrado_mes = row.monto_cobrado_mes or 0

        total_vencido_mes = row.total_vencido_mes or 0

        no_cobrado_mes = row.no_cobrado_mes or 0

        cartera_pendiente = row.cartera_pendiente or 0



        morosidad_porcentaje = (

            (_safe_float(no_cobrado_mes) / _safe_float(total_vencido_mes) * 100.0)

            if total_vencido_mes and _safe_float(total_vencido_mes) > 0

            else 0.0

        )



        # Conteos: clientes en mora y clientes con préstamo (2 consultas ligeras)

        subq = (

            select(Prestamo.cliente_id)

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(

                and_(*conds_activo),

                Cuota.fecha_pago.is_(None),

                Cuota.fecha_vencimiento < fecha_referencia,

            )

            .distinct()

        )

        clientes_en_mora = db.scalar(select(func.count()).select_from(subq.subquery())) or 0

        clientes_con_prestamo = db.scalar(

            select(func.count(func.distinct(Prestamo.cliente_id)))

            .select_from(Prestamo)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")

        ) or 0

        clientes_al_dia = max(0, clientes_con_prestamo - clientes_en_mora)



        mes_resp = mes if mes is not None else inicio_mes.month

        anio_resp = anio if anio is not None else inicio_mes.year

        n_prestamos_huella_dup = contar_prestamos_con_huella_funcional_duplicada(db)

        calidad_huella = {

            **snapshot_rechazos_huella_funcional(),

            "prestamos_con_alerta_control_huella_funcional_duplicada": n_prestamos_huella_dup,

            "codigo_control_auditoria": "pagos_huella_funcional_duplicada",

            "nota": (

                "rechazos_409_* cuenta rechazos API desde el ultimo arranque del proceso; "

                "prestamos_con_alerta_* es conteo en tiempo real en BD (misma regla que auditoria cartera)."

            ),

        }

        return {

            "montoACobrarMes": _safe_float(monto_a_cobrar_mes),

            "montoCobradoMes": _safe_float(monto_cobrado_mes),

            "morosidadMensualPorcentaje": round(morosidad_porcentaje, 2),

            "mes": mes_resp,

            "anio": anio_resp,

            "saldoPorCobrar": _safe_float(cartera_pendiente),

            "clientesEnMora": clientes_en_mora,

            "clientesAlDia": clientes_al_dia,

            "calidad_carga_pagos_huella": calidad_huella,

        }

    except Exception as e:

        logger.exception("Error en GET /pagos/kpis: %s", e)

        try:

            db.rollback()

        except Exception:

            pass

        hoy = _hoy_local()

        return {

            "montoACobrarMes": 0.0,

            "montoCobradoMes": 0.0,

            "morosidadMensualPorcentaje": 0.0,

            "mes": mes if mes is not None else hoy.month,

            "anio": anio if anio is not None else hoy.year,

            "saldoPorCobrar": 0.0,

            "clientesEnMora": 0,

            "clientesAlDia": 0,

            "calidad_carga_pagos_huella": {

                **snapshot_rechazos_huella_funcional(),

                "prestamos_con_alerta_control_huella_funcional_duplicada": None,

                "codigo_control_auditoria": "pagos_huella_funcional_duplicada",

                "nota": "KPI principal no disponible por error; conteo BD omitido en este fallback.",

            },

        }





def _stats_conds_cuota(analista: Optional[str], concesionario: Optional[str], modelo: Optional[str]):

    """Condiciones base para filtrar cuotas por préstamo (solo clientes ACTIVOS + analista/concesionario/modelo)."""

    conds = [

        Cuota.prestamo_id == Prestamo.id,

        Prestamo.cliente_id == Cliente.id,

        Cliente.estado == "ACTIVO",

        Prestamo.estado == "APROBADO",

    ]

    if analista:

        conds.append(Prestamo.analista == analista)

    if concesionario:

        conds.append(Prestamo.concesionario == concesionario)

    if modelo:

        conds.append(Prestamo.modelo_vehiculo == modelo)

    return conds





@router.get("/stats")

def get_pagos_stats(

    fecha_inicio: Optional[str] = Query(None),

    fecha_fin: Optional[str] = Query(None),

    analista: Optional[str] = Query(None),

    concesionario: Optional[str] = Query(None),

    modelo: Optional[str] = Query(None),

    db: Session = Depends(get_db),

):

    """

    Estadísticas de pagos desde BD (solo clientes ACTIVOS): total_pagos, total_pagado, pagos_por_estado,

    cuotas_pagadas, cuotas_pendientes, cuotas_atrasadas, pagos_hoy.

    """

    hoy = _hoy_local()

    use_filters = bool(analista or concesionario or modelo)



    def _q_cuotas():

        return (

            select(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(and_(*_stats_conds_cuota(analista, concesionario, modelo)))

        )



    def _count(q):

        subq = q.subquery()

        return int(db.scalar(select(func.count()).select_from(subq)) or 0)



    try:

        # Cuotas pagadas / pendientes / atrasadas (solo clientes ACTIVOS)

        cuotas_pagadas = _count(_q_cuotas().where(Cuota.fecha_pago.isnot(None)))

        cuotas_pendientes = _count(_q_cuotas().where(Cuota.fecha_pago.is_(None)))

        cuotas_atrasadas = _count(

            _q_cuotas().where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < hoy)

        )

        # Total pagado (solo clientes ACTIVOS)

        q_sum = (

            select(func.coalesce(func.sum(Cuota.monto), 0))

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(and_(*_stats_conds_cuota(analista, concesionario, modelo), Cuota.fecha_pago.isnot(None))

            )

        )

        total_pagado = db.scalar(q_sum) or 0

        # Pagos hoy

        pagos_hoy = _count(

            _q_cuotas().where(

                Cuota.fecha_pago.isnot(None),

                func.date(Cuota.fecha_pago) == hoy,

            )

        )

        # Pagos por estado

        q_estado = (

            select(Cuota.estado, func.count())

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(and_(*_stats_conds_cuota(analista, concesionario, modelo)))

            .group_by(Cuota.estado)

        )

        rows_estado = db.execute(q_estado).all()

        pagos_por_estado = [{"estado": str(r[0]) if r[0] is not None else "N/A", "count": int(r[1])} for r in rows_estado]

        total_pagos = cuotas_pagadas + cuotas_pendientes

        return {

            "total_pagos": total_pagos,

            "total_pagado": _safe_float(total_pagado),

            "pagos_por_estado": pagos_por_estado,

            "cuotas_pagadas": cuotas_pagadas,

            "cuotas_pendientes": cuotas_pendientes,

            "cuotas_atrasadas": cuotas_atrasadas,

            "pagos_hoy": pagos_hoy,

        }

    except Exception as e:

        logger.exception("Error en GET /pagos/stats: %s", e)

        try:

            db.rollback()

        except Exception:

            pass

        return {

            "total_pagos": 0,

            "total_pagado": 0.0,

            "pagos_por_estado": [],

            "cuotas_pagadas": 0,

            "cuotas_pendientes": 0,

            "cuotas_atrasadas": 0,

            "pagos_hoy": 0,

        }





@router.post("/revisar-pagos/mover", response_model=dict)

def mover_a_revisar_pagos(payload: MoverRevisarPagosBody = Body(...), db: Session = Depends(get_db)):

    """

    Mueve los pagos exportados a la tabla revisar_pagos (temporal de validación).

    Tras descargar Excel y guardar en PC, estos pagos dejan de mostrarse en Revisar Pagos.

    No interfiere con procesos ni reglas de negocio.

    """

    if not payload.pago_ids:

        return {"movidos": 0, "mensaje": "No hay pagos para mover"}

    insertados = 0

    for pid in payload.pago_ids:

        if not isinstance(pid, int) or pid <= 0:

            continue

        existing = db.scalar(select(RevisarPago.id).where(RevisarPago.pago_id == pid))

        if existing:

            continue

        db.add(RevisarPago(pago_id=pid))

        insertados += 1

    db.commit()

    return {"movidos": insertados, "mensaje": f"{insertados} pagos movidos a revisar_pagos"}





@router.post("/batch", response_model=dict)

def crear_pagos_batch(

    body: PagoBatchBody,

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Crea varios pagos en una sola petición (máx. 500).

    Devuelve éxitos y errores por índice para reducir rondas y timeouts en "Guardar todos".

    Optimizado: precarga de préstamos y clientes en lugar de N consultas por fila.

    Mismo comprobante puede repetirse con codigo_documento distinto; duplicado = mismo valor almacenado.

    """

    try:

        usuario_registro = _usuario_registro_desde_current_user(current_user)

        pagos_list = body.pagos

        docs_compuestos = [

            compose_numero_documento_almacenado(p.numero_documento, getattr(p, "codigo_documento", None))

            for p in pagos_list

        ]

        docs_no_vacios = [d for d in docs_compuestos if d]

        existing_docs: set[str] = set()

        if docs_no_vacios:

            rows_bd = db.execute(select(Pago.numero_documento).where(Pago.numero_documento.in_(docs_no_vacios))).scalars().all()

            existing_docs = {r for r in rows_bd if r}

            rows_pe = db.execute(

                select(PagoConError.numero_documento).where(PagoConError.numero_documento.in_(docs_no_vacios))

            ).scalars().all()

            existing_docs.update({r for r in rows_pe if r})

        cedulas_payload = list(

            {

                (p.cedula_cliente or "").strip().replace("-", "").upper()

                for p in pagos_list

                if (p.cedula_cliente or "").strip()

            }

        )

        # Preload: ids de préstamos válidos (una sola consulta)

        prestamo_ids = [p.prestamo_id for p in pagos_list if p.prestamo_id]

        pc_prest = func.upper(func.replace(Prestamo.cedula, "-", ""))

        prestamos_activos_por_cedula: dict[str, list[int]] = {}

        if cedulas_payload:

            rows_act = db.execute(

                select(Prestamo.id, pc_prest)

                .where(pc_prest.in_(cedulas_payload))

                .where(Prestamo.estado.in_(("APROBADO", "DESEMBOLSADO")))

            ).all()

            for pid, ccell in rows_act:

                if pid is None:

                    continue

                ck = (str(ccell) if ccell is not None else "").strip().replace("-", "").upper()

                if ck:

                    prestamos_activos_por_cedula.setdefault(ck, []).append(int(pid))

            for _ck in prestamos_activos_por_cedula:

                prestamos_activos_por_cedula[_ck] = sorted(set(prestamos_activos_por_cedula[_ck]))

        all_pids_batch: set[int] = set(int(x) for x in prestamo_ids if x is not None)

        for _ids in prestamos_activos_por_cedula.values():

            all_pids_batch.update(_ids)

        valid_prestamo_ids: set[int] = set()

        if all_pids_batch:

            ids_rows = db.execute(select(Prestamo.id).where(Prestamo.id.in_(all_pids_batch))).scalars().all()

            valid_prestamo_ids = {int(r) for r in ids_rows if r is not None}

        prestamo_estado_por_id: dict[int, str] = {}

        if all_pids_batch:

            er_rows = db.execute(select(Prestamo.id, Prestamo.estado).where(Prestamo.id.in_(all_pids_batch))).all()

            prestamo_estado_por_id = {int(r[0]): (r[1] or "") for r in er_rows if r[0] is not None}


        # Preload: cédulas que tienen al menos un préstamo (normalizadas)

        valid_cedulas_prestamo: set[str] = set()

        if cedulas_payload:

            pc = func.upper(func.replace(Prestamo.cedula, "-", ""))

            ced_rows = db.execute(select(pc).where(pc.in_(cedulas_payload)).distinct()).scalars().all()

            valid_cedulas_prestamo = {(r or "").strip().replace("-", "").upper() for r in ced_rows if r}

        todas_cedulas_upper = list(

            {

                (p.cedula_cliente or "").strip().upper()

                for p in pagos_list

                if (p.cedula_cliente or "").strip()

            }

        )

        valid_cedulas: set[str] = set()

        if todas_cedulas_upper:

            ced_rows_c = db.execute(

                select(Cliente.cedula).where(func.upper(Cliente.cedula).in_(todas_cedulas_upper))

            ).scalars().all()

            valid_cedulas = {(r or "").strip().upper() for r in ced_rows_c if r}

        # Fase 1: validacion (sin insertar). Errores por indice; las filas validas se insertan en fase 2.

        errors_by_index: dict[int, dict] = {}

        resolved_prestamo_id_by_index: dict[int, int] = {}

        docs_added_in_batch: set[str] = set()

        for idx, payload in enumerate(pagos_list):

            num_doc = (

                docs_compuestos[idx]

                if idx < len(docs_compuestos)

                else compose_numero_documento_almacenado(

                    payload.numero_documento,

                    getattr(payload, "codigo_documento", None),

                )

            )

            if num_doc and (num_doc in existing_docs or num_doc in docs_added_in_batch):

                errors_by_index[idx] = {

                    "error": "Ya existe un pago con la misma combinación comprobante + código.",

                    "status_code": 409,

                }

                continue

            cedula_normalizada = (payload.cedula_cliente or "").strip().upper()

            ced_norm_prest = (payload.cedula_cliente or "").strip().replace("-", "").upper()

            if ced_norm_prest and ced_norm_prest not in valid_cedulas_prestamo:

                errors_by_index[idx] = {

                        "error": f"La cédula no tiene préstamo registrado: {cedula_normalizada}",

                        "status_code": 400,

                    }

                continue

            activos_ced = prestamos_activos_por_cedula.get(ced_norm_prest, [])

            effective_prestamo_id: Optional[int] = None

            raw_pid = getattr(payload, "prestamo_id", None)

            if raw_pid is not None and int(raw_pid) > 0:

                effective_prestamo_id = int(raw_pid)

            elif len(activos_ced) == 1:

                effective_prestamo_id = activos_ced[0]

            elif len(activos_ced) == 0:

                errors_by_index[idx] = {

                    "error": (

                        f"Sin credito activo (APROBADO/DESEMBOLSADO) para {cedula_normalizada}; "

                        "indique prestamo_id o asocie un credito."

                    ),

                    "status_code": 400,

                }

                continue

            else:

                errors_by_index[idx] = {

                    "error": (

                        f"La cedula {cedula_normalizada} tiene varios creditos activos; "

                        "elija prestamo_id en la columna Credito (obligatorio en lote)."

                    ),

                    "status_code": 400,

                }

                continue

            if effective_prestamo_id not in valid_prestamo_ids:

                errors_by_index[idx] = {"error": f"Credito #{effective_prestamo_id} no existe.", "status_code": 400}

                continue

            if (prestamo_estado_por_id.get(effective_prestamo_id) or "").strip().upper() == "DESISTIMIENTO":

                errors_by_index[idx] = {

                    "error": "El prestamo esta en desistimiento; no se registran pagos.",

                    "status_code": 400,

                }

                continue

            if cedula_normalizada and cedula_normalizada not in valid_cedulas:

                errors_by_index[idx] = {"error": f"No existe cliente con cedula {cedula_normalizada}", "status_code": 404}

                continue

            resolved_prestamo_id_by_index[idx] = effective_prestamo_id

            if num_doc:

                docs_added_in_batch.add(num_doc)

        if len(errors_by_index) == len(pagos_list):

            results = [
                {
                    "index": i,
                    "success": False,
                    "error": errors_by_index[i]["error"],
                    "status_code": errors_by_index[i]["status_code"],
                }
                for i in sorted(errors_by_index)
            ]

            return {"results": results, "ok_count": 0, "fail_count": len(results)}

        cedulas_lote = {
            (pagos_list[i].cedula_cliente or "").strip().upper()
            for i in range(len(pagos_list))
            if i not in errors_by_index and (pagos_list[i].cedula_cliente or "").strip()
        }

        alinear_cedulas_clientes_existentes(db, cedulas_lote)



        # Fase 2: transaccion unica. Crear pagos validos (flush), aplicar a cuotas, un commit al final.

        results: list[dict] = []

        try:

            autorizados_bs = preload_autorizados_bs(db)

            huellas_lote_batch: set[tuple[int, str, str, str]] = set()

            for idx, payload in enumerate(pagos_list):

                num_doc = (

                    docs_compuestos[idx]

                    if idx < len(docs_compuestos)

                    else compose_numero_documento_almacenado(

                        payload.numero_documento,

                        getattr(payload, "codigo_documento", None),

                    )

                )

                if idx in errors_by_index:

                    err = errors_by_index[idx]

                    results.append(

                        {

                            "index": idx,

                            "success": False,

                            "error": err["error"],

                            "status_code": err["status_code"],

                        }

                    )

                    continue

                ref = (num_doc or "N/A")[:_MAX_LEN_NUMERO_DOCUMENTO]

                fecha_pago_ts = datetime.combine(payload.fecha_pago, dt_time.min)

                _pid_res = resolved_prestamo_id_by_index.get(idx)

                if _pid_res is None:

                    raise HTTPException(status_code=500, detail=f"Lote fila {idx + 1}: falta prestamo_id resuelto (error interno).")

                conciliado = payload.conciliado if payload.conciliado is not None else True

                cedula_normalizada = (payload.cedula_cliente or "").strip().upper()

                es_valido_b, monto_val_b, err_msg_b = _validar_monto(payload.monto_pagado)

                if not es_valido_b:

                    raise HTTPException(status_code=400, detail=f"Lote fila {idx + 1}: {err_msg_b}")

                try:

                    monto_usd_b, moneda_fin_b, monto_bs_b, tasa_b, fecha_tasa_b = resolver_monto_registro_pago(

                        db,

                        cedula_normalizada=cedula_normalizada,

                        fecha_pago=payload.fecha_pago,

                        monto_pagado=Decimal(str(round(monto_val_b, 2))),

                        moneda_registro=payload.moneda_registro or "USD",

                        tasa_cambio_manual=payload.tasa_cambio_manual,

                        autorizados_bs=autorizados_bs,

                    )

                except HTTPException as e:

                    d = e.detail if isinstance(e.detail, str) else str(e.detail)

                    raise HTTPException(status_code=e.status_code, detail=f"Lote fila {idx + 1}: {d}") from e

                msg_h_batch = conflicto_huella_para_creacion(
                    db,
                    prestamo_id=_pid_res,
                    fecha_pago=payload.fecha_pago,
                    monto_pagado=monto_usd_b,
                    numero_documento=num_doc,
                    referencia_pago=ref,
                    huellas_en_mismo_lote=huellas_lote_batch,
                )
                if msg_h_batch:
                    registrar_rechazo_huella_funcional()
                    raise HTTPException(
                        status_code=409,
                        detail=f"Lote fila {idx + 1}: {msg_h_batch}",
                    )

                row = Pago(

                    cedula_cliente=cedula_normalizada,

                    prestamo_id=_pid_res,

                    fecha_pago=fecha_pago_ts,

                    monto_pagado=monto_usd_b,

                    numero_documento=num_doc,

                    institucion_bancaria=payload.institucion_bancaria.strip() if payload.institucion_bancaria else None,

                    estado="PAGADO" if conciliado else "PENDIENTE",

                    notas=payload.notas.strip() if payload.notas else None,

                    referencia_pago=ref,

                    conciliado=conciliado,

                    fecha_conciliacion=datetime.now(ZoneInfo(TZ_NEGOCIO)) if conciliado else None,

                    verificado_concordancia="SI" if conciliado else "",

                    usuario_registro=usuario_registro,

                    moneda_registro=moneda_fin_b,

                    monto_bs_original=monto_bs_b,

                    tasa_cambio_bs_usd=tasa_b,

                    fecha_tasa_referencia=fecha_tasa_b,

                    link_comprobante=(
                        payload.link_comprobante.strip()
                        if payload.link_comprobante
                        else None
                    ),

                )

                db.add(row)

                db.flush()

                db.refresh(row)

                if _debe_aplicar_cascada_pago(row):

                    cc_b, cp_b = _aplicar_pago_a_cuotas_interno(row, db)

                    row.estado = _estado_conciliacion_post_cascada(row, cc_b, cp_b)

                results.append({"index": idx, "success": True, "pago": _pago_to_response(row)})

            db.commit()

            # Persistir link_comprobante desde pagos_gmail_sync_item si el Excel no trajo URL
            try:
                ok_ids = [
                    int(r["pago"]["id"])
                    for r in results
                    if r.get("success") and r.get("pago", {}).get("id") is not None
                ]
                if ok_ids:
                    rows_lc = db.execute(select(Pago).where(Pago.id.in_(ok_ids))).scalars().all()
                    to_persist = False
                    for row_lc in rows_lc:
                        if (getattr(row_lc, "link_comprobante", None) or "").strip():
                            continue
                        out_lc = _pago_to_response(row_lc)
                        enriquecer_items_link_comprobante_desde_gmail(db, [out_lc])
                        url_lc = (out_lc.get("link_comprobante") or "").strip()
                        if url_lc:
                            row_lc.link_comprobante = url_lc
                            to_persist = True
                    if to_persist:
                        db.commit()
                        for r in results:
                            if not r.get("success"):
                                continue
                            pid = r.get("pago", {}).get("id")
                            if pid is None:
                                continue
                            row_u = db.get(Pago, int(pid))
                            if row_u:
                                r["pago"] = _pago_to_response(row_u)
            except Exception as ex_lc:
                logger.warning(
                    "Batch: no se pudo persistir link_comprobante desde Gmail sync: %s",
                    ex_lc,
                )
                db.rollback()

            results.sort(key=lambda r: r["index"])

            ok_count = sum(1 for r in results if r.get("success"))

            fail_count = len(results) - ok_count

            return {"results": results, "ok_count": ok_count, "fail_count": fail_count}

        except HTTPException:

            db.rollback()

            raise

        except Exception as e:

            db.rollback()

            logger.exception("Batch: error en transaccion unica: %s", e)

            raise HTTPException(status_code=500, detail=f"Error al guardar el lote. Ningun pago fue creado. Detalle: {str(e)}")

    except HTTPException:

        raise

    except OperationalError:

        raise HTTPException(

            status_code=503,

            detail="Servicio no disponible. Reintente en unos segundos.",

        )

    except Exception as e:

        logger.exception("Batch: error inesperado: %s", e)

        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}") from e







@router.get("/{pago_id:int}", response_model=dict)

def obtener_pago(pago_id: int, db: Session = Depends(get_db)):

    """Obtiene un pago por ID desde la tabla pagos."""

    row = db.get(Pago, pago_id)

    if not row:

        raise HTTPException(status_code=404, detail="Pago no encontrado")

    return _pago_response_enriquecido(db, row)





@router.post("", response_model=dict, status_code=201)

def crear_pago(payload: PagoCreate, db: Session = Depends(get_db), current_user: UserResponse = Depends(get_current_user)):

    """Crea un pago. Clave única = comprobante + código opcional; 409 si ya existe o por huella funcional."""

    if payload.prestamo_id is None:

        raise HTTPException(status_code=400, detail="prestamo_id es obligatorio para crear pagos.")

    num_stored = compose_numero_documento_almacenado(
        payload.numero_documento,
        payload.codigo_documento,
    )

    if not num_stored:

        raise HTTPException(status_code=400, detail="numero_documento es obligatorio para crear pagos.")

    if numero_documento_ya_registrado(db, num_stored):

        raise HTTPException(

            status_code=409,

            detail=(

                "Conflicto numero_documento: ya existe un pago con la misma combinación "

                "comprobante + código (incluye misma referencia con distinto uso de mayúsculas). "

                "Use un código distinto en el campo «Código» o verifique duplicados."

            ),

        )

    ref = (num_stored or "N/A")[:_MAX_LEN_NUMERO_DOCUMENTO]

    ref_norm = _normalizar_ref_fingerprint(num_stored or ref)

    fecha_pago_ts = datetime.combine(payload.fecha_pago, dt_time.min)

    # Autoconciliar cuando se asigna a un préstamo y no se indica explícitamente conciliado

    conciliado = payload.conciliado if payload.conciliado is not None else (True if payload.prestamo_id else False)

    usuario_registro = _usuario_registro_desde_current_user(current_user)

    

    # Normalizar cédula: uppercase para evitar FK mismatch

    cedula_normalizada = payload.cedula_cliente.strip().upper() if payload.cedula_cliente else ""

    

    # Validar que el crédito (prestamo_id) existe en prestamos (evita IntegrityError FK)

    if payload.prestamo_id:

        prestamo = db.execute(select(Prestamo).where(Prestamo.id == payload.prestamo_id)).scalars().first()

        if not prestamo:

            raise HTTPException(

                status_code=400,

                detail=f"El crédito #{payload.prestamo_id} no existe. Elija un crédito de la lista (no use el número de documento como crédito).",

            )
        if prestamo and (prestamo.estado or "").strip().upper() == "DESISTIMIENTO":

            raise HTTPException(

                status_code=400,

                detail="El prestamo esta en desistimiento; no se registran pagos.",

            )


    # Validar que cedula existe en clientes si se proporciona y hay prestamo_id

    if cedula_normalizada and payload.prestamo_id:

        cliente = db.execute(

            select(Cliente).where(func.upper(Cliente.cedula) == cedula_normalizada)

        ).scalars().first()

        if not cliente:

            raise HTTPException(

                status_code=404,

                detail=f"No existe cliente con cedula {cedula_normalizada}"

            )



    # Validar monto (mismo criterio que carga Excel y guardar-fila-editable)

    monto_raw = payload.monto_pagado

    es_valido, monto_val, err_msg = _validar_monto(monto_raw)

    if not es_valido:

        raise HTTPException(status_code=400, detail=err_msg)



    if cedula_normalizada:

        alinear_cedulas_clientes_existentes(db, [cedula_normalizada])



    try:

        monto_usd, moneda_fin, monto_bs_o, tasa_o, fecha_tasa_o = resolver_monto_registro_pago(

            db,

            cedula_normalizada=cedula_normalizada,

            fecha_pago=payload.fecha_pago,

            monto_pagado=Decimal(str(round(monto_val, 2))),

            moneda_registro=payload.moneda_registro or "USD",

            tasa_cambio_manual=payload.tasa_cambio_manual,

        )

        msg_huella = conflicto_huella_para_creacion(
            db,
            prestamo_id=payload.prestamo_id,
            fecha_pago=payload.fecha_pago,
            monto_pagado=monto_usd,
            numero_documento=num_stored,
            referencia_pago=ref,
        )
        if msg_huella:
            registrar_rechazo_huella_funcional()
            raise HTTPException(status_code=409, detail=msg_huella)

        row = Pago(

            cedula_cliente=cedula_normalizada,

            prestamo_id=payload.prestamo_id,

            fecha_pago=fecha_pago_ts,

            monto_pagado=monto_usd,

            numero_documento=num_stored,

            institucion_bancaria=payload.institucion_bancaria.strip() if payload.institucion_bancaria else None,

            estado="PAGADO" if conciliado else "PENDIENTE",

            notas=payload.notas.strip() if payload.notas else None,

            referencia_pago=ref,

            ref_norm=ref_norm,

            conciliado=conciliado,  # [B2] Usar solo conciliado, no verificado_concordancia

            fecha_conciliacion=datetime.now(ZoneInfo(TZ_NEGOCIO)) if conciliado else None,

            verificado_concordancia="SI" if conciliado else "",  # Legacy: sync con conciliado

            usuario_registro=usuario_registro,  # [MEJORADO] Usuario real desde JWT

            moneda_registro=moneda_fin,

            monto_bs_original=monto_bs_o,

            tasa_cambio_bs_usd=tasa_o,

            fecha_tasa_referencia=fecha_tasa_o,

            link_comprobante=(payload.link_comprobante.strip() if payload.link_comprobante else None),

        )

        db.add(row)

        db.flush()

        db.refresh(row)

        # [C3] Aplicar cascada a cuotas en la misma transacción para que préstamos y estado de cuenta se actualicen

        if _debe_aplicar_cascada_pago(row):

            cc_n, cp_n = _aplicar_pago_a_cuotas_interno(row, db)

            row.estado = _estado_conciliacion_post_cascada(row, cc_n, cp_n)

        db.commit()

        db.refresh(row)

        return _pago_response_enriquecido(db, row)

    except HTTPException:

        raise

    except IntegrityError as e:

        db.rollback()

        # Unique violation (p. ej. numero_documento duplicado) ? 409

        if getattr(getattr(e, "orig", None), "pgcode", None) == "23505":

            constraint = getattr(getattr(getattr(e, "orig", None), "diag", None), "constraint_name", "") or ""

            if "ux_pagos_fingerprint_activos" in constraint:

                registrar_rechazo_huella_funcional()

                raise HTTPException(

                    status_code=409,

                    detail=HTTP_409_DETAIL_HUELLA_FUNCIONAL,

                )

            raise HTTPException(

                status_code=409,

                detail=(

                    "Violación de unicidad en base de datos (restricción distinta a la huella funcional). "

                    "Si el mensaje menciona numero_documento, ejecute en BD: "

                    "ALTER TABLE public.pagos DROP CONSTRAINT IF EXISTS uq_pagos_numero_documento;"

                ),

            )

        raise HTTPException(status_code=500, detail="Error de integridad en la base de datos.")

    except Exception as e:

        logger.exception("Error en POST /pagos (crear_pago): %s", e)

        db.rollback()

        raise HTTPException(

            status_code=500,

            detail="Error al crear el pago. Revise los datos (cédula, crédito, monto, fecha, Nº documento) o contacte soporte.",

        )





@router.put("/{pago_id:int}", response_model=dict)

def actualizar_pago(pago_id: int, payload: PagoUpdate, db: Session = Depends(get_db)):

    """Actualiza un pago en la tabla pagos. Mismo texto de documento puede existir en otros pagos.

    Si el pago ya estaba articulado a cuotas (cuota_pagos) y cambian monto, fecha de pago o préstamo,
    se ejecuta la misma reconstrucción en cascada que POST .../prestamos/{id}/reaplicar-cascada-aplicacion
    sobre el o los préstamos afectados para que la amortización y servicios alineados vuelvan a cuadrar.
    """

    row = db.get(Pago, pago_id)

    if not row:

        raise HTTPException(status_code=404, detail="Pago no encontrado")

    had_cuota_pagos_antes = pago_tiene_aplicaciones_cuotas(db, pago_id)

    old_prestamo_id = row.prestamo_id

    old_monto_pagado = row.monto_pagado

    old_fecha_pago = row.fecha_pago

    data = payload.model_dump(exclude_unset=True)

    _doc_touch = False

    if "numero_documento" in data:

        _doc_touch = True

    if "codigo_documento" in data:

        _doc_touch = True

    if _doc_touch:

        b0, c0 = split_numero_documento_almacenado(row.numero_documento)

        nb = (

            normalize_documento(data["numero_documento"])

            if "numero_documento" in data and data["numero_documento"] is not None

            else b0

        )

        if not nb:

            raise HTTPException(status_code=400, detail="numero_documento no puede estar vacio.")

        if "codigo_documento" in data:

            nc = normalize_codigo_documento(data["codigo_documento"])

        else:

            nc = normalize_codigo_documento(c0) if c0 else None

        new_stored = compose_numero_documento_almacenado(nb, nc)

        old_stored = (row.numero_documento or "").strip()

        if (new_stored or "") != old_stored and (

            bool(row.conciliado) or str(row.estado or "").upper() in ("PAGADO", "PAGO_ADELANTADO")

        ):

            raise HTTPException(

                status_code=409,

                detail="No se permite cambiar comprobante/código en pagos conciliados o pagados.",

            )

        if new_stored and numero_documento_ya_registrado(db, new_stored, exclude_pago_id=pago_id):

            raise HTTPException(

                status_code=409,

                detail="Ya existe otro pago con la misma combinación comprobante + código.",

            )

        data.pop("numero_documento", None)

        data.pop("codigo_documento", None)

        row.numero_documento = new_stored

    aplicar_conciliado = False

    _put_mon_keys = ("monto_pagado", "moneda_registro", "tasa_cambio_manual", "fecha_pago")

    monetario_up = {k: data.pop(k) for k in _put_mon_keys if k in data}

    for k, v in data.items():

        if k == "notas" and v is not None:

            setattr(row, k, v.strip() or None)

        elif k == "institucion_bancaria" and v is not None:

            setattr(row, k, v.strip() or None)

        elif k == "cedula_cliente" and v is not None:

            setattr(row, k, v.strip())

        elif k == "conciliado" and v is not None:

            row.conciliado = bool(v)

            row.fecha_conciliacion = datetime.now(ZoneInfo(TZ_NEGOCIO)) if v else None

            row.verificado_concordancia = "SI" if v else "NO"

            aplicar_conciliado = bool(v)

            _alinear_estado_si_toggle_conciliado_actualizar_pago(row, bool(v))

        elif k == "verificado_concordancia" and v is not None:

            val = (v or "").strip().upper()

            row.verificado_concordancia = val if val in ("SI", "NO") else ("" if v == "" else str(v)[:10])

        else:

            setattr(row, k, v)

    if monetario_up:

        def _fp_row_a_date(fp):

            if fp is None:

                return None

            return fp.date() if hasattr(fp, "date") and callable(getattr(fp, "date", None)) else fp

        try:

            moneda_prev = normalizar_moneda_registro(row.moneda_registro or "USD")

        except HTTPException:

            moneda_prev = "USD"

        if "moneda_registro" in monetario_up:

            moneda_target = normalizar_moneda_registro(monetario_up["moneda_registro"])

        else:

            moneda_target = moneda_prev

        if moneda_target == "BS":

            fecha_eff = monetario_up.get("fecha_pago")

            if fecha_eff is None:

                fecha_eff = _fp_row_a_date(row.fecha_pago)

            if fecha_eff is None:

                raise HTTPException(

                    status_code=400,

                    detail="fecha_pago es requerida para pagos en bolivares.",

                )

            if "monto_pagado" in monetario_up:

                es_valido, monto_val, err_msg = _validar_monto(monetario_up["monto_pagado"])

                if not es_valido:

                    raise HTTPException(status_code=400, detail=err_msg)

                monto_bs_input = Decimal(str(round(monto_val, 2)))

            else:

                if row.monto_bs_original is not None:

                    monto_bs_input = Decimal(str(round(float(row.monto_bs_original), 2)))

                else:

                    raise HTTPException(

                        status_code=400,

                        detail="Indique el monto en bolivares; no hay monto Bs previo en este pago.",

                    )

            cedula_norm = (row.cedula_cliente or "").strip().upper()

            if cedula_norm:

                alinear_cedulas_clientes_existentes(db, [cedula_norm])

            tasa_manual_put = monetario_up.get("tasa_cambio_manual")

            monto_usd, moneda_fin, monto_bs_o, tasa_o, fecha_tasa_o = resolver_monto_registro_pago(

                db,

                cedula_normalizada=cedula_norm,

                fecha_pago=fecha_eff,

                monto_pagado=monto_bs_input,

                moneda_registro="BS",

                tasa_cambio_manual=tasa_manual_put,

            )

            row.monto_pagado = monto_usd

            row.moneda_registro = moneda_fin

            row.monto_bs_original = monto_bs_o

            row.tasa_cambio_bs_usd = tasa_o

            row.fecha_tasa_referencia = fecha_tasa_o

            if "fecha_pago" in monetario_up:

                fpv = monetario_up["fecha_pago"]

                row.fecha_pago = (

                    datetime.combine(fpv, dt_time.min)

                    if isinstance(fpv, date) and not isinstance(fpv, datetime)

                    else fpv

                )

        else:

            if "fecha_pago" in monetario_up:

                fpv = monetario_up["fecha_pago"]

                row.fecha_pago = (

                    datetime.combine(fpv, dt_time.min)

                    if isinstance(fpv, date) and not isinstance(fpv, datetime)

                    else fpv

                )

            if "monto_pagado" in monetario_up:

                es_valido, monto_val, err_msg = _validar_monto(monetario_up["monto_pagado"])

                if not es_valido:

                    raise HTTPException(status_code=400, detail=err_msg)

                row.monto_pagado = Decimal(str(round(monto_val, 2)))

            row.moneda_registro = "USD"

            row.monto_bs_original = None

            row.tasa_cambio_bs_usd = None

            row.fecha_tasa_referencia = None

    fp_row = row.fecha_pago

    if isinstance(fp_row, datetime):

        fecha_huella_up = fp_row.date()

    elif isinstance(fp_row, date):

        fecha_huella_up = fp_row

    else:

        fecha_huella_up = None

    rn_up = ref_norm_desde_campos(row.numero_documento, row.referencia_pago).strip()

    if row.prestamo_id and fecha_huella_up is not None and rn_up:

        cid_up = primer_id_conflicto_huella_funcional(
            db,
            prestamo_id=int(row.prestamo_id),
            fecha_pago=fecha_huella_up,
            monto_pagado=row.monto_pagado,
            ref_norm=rn_up,
            exclude_pago_id=pago_id,
        )
        if cid_up is not None:
            db.rollback()

            registrar_rechazo_huella_funcional()

            raise HTTPException(status_code=409, detail=mensaje_409_huella_funcional_con_id(cid_up))

    try:

        db.commit()

        db.refresh(row)

    except IntegrityError as e:

        db.rollback()

        pgcode, cname = _integridad_error_pgcode_y_constraint(e)

        if pgcode == "23505":

            if "ux_pagos_fingerprint_activos" in cname:

                registrar_rechazo_huella_funcional()

                raise HTTPException(status_code=409, detail=HTTP_409_DETAIL_HUELLA_FUNCIONAL)

            raise HTTPException(

                status_code=409,

                detail=(

                    "Violación de unicidad en base de datos. Si aplica a numero_documento, "

                    "ejecute: ALTER TABLE public.pagos DROP CONSTRAINT IF EXISTS uq_pagos_numero_documento;"

                ),

            )

        if pgcode == "23514":

            logger.warning(
                "actualizar_pago pago_id=%s: violacion CHECK (23514) constraint=%s detail=%s",
                pago_id,
                cname,
                str(e)[:400],
            )

            raise HTTPException(
                status_code=400,
                detail=(
                    "La base de datos rechazó la combinación estado/conciliación del pago "
                    f"(restricción: {cname or 'CHECK'}). "
                    "Pruebe de nuevo tras recargar; si quitó validación cartera, el estado debe pasar a Pendiente."
                ),
            )

        logger.warning(
            "actualizar_pago pago_id=%s IntegrityError pgcode=%s constraint=%s msg=%s",
            pago_id,
            pgcode,
            cname,
            str(e)[:500],
        )

        raise HTTPException(
            status_code=400,
            detail=(
                "La base de datos rechazó la actualización del pago "
                f"(código SQL {pgcode or 'desconocido'}, restricción {cname or 'N/A'}). "
                "Revise estado, conciliación y documento; copie este mensaje si contacta soporte."
            ),
        )

    def _fecha_pago_a_date(fp):

        if fp is None:

            return None

        return fp.date() if hasattr(fp, "date") and callable(getattr(fp, "date", None)) else fp

    new_fd = _fecha_pago_a_date(row.fecha_pago)

    old_fd = _fecha_pago_a_date(old_fecha_pago)

    monto_changed = had_cuota_pagos_antes and abs(float(old_monto_pagado or 0) - float(row.monto_pagado or 0)) >= 0.01

    fecha_changed = had_cuota_pagos_antes and (old_fd != new_fd)

    prestamo_changed = had_cuota_pagos_antes and (old_prestamo_id != row.prestamo_id)

    articulacion_afectada = monto_changed or fecha_changed or prestamo_changed

    if articulacion_afectada:

        from app.services.pagos_cuotas_reaplicacion import reset_y_reaplicar_cascada_prestamo

        try:

            for pid in sorted({p for p in (old_prestamo_id, row.prestamo_id) if p}):

                r = reset_y_reaplicar_cascada_prestamo(db, pid)

                if not r.get("ok"):

                    raise HTTPException(

                        status_code=400,

                        detail=(

                            f"No se pudo sincronizar la tabla de amortización del préstamo {pid}: "

                            f"{r.get('error') or 'error desconocido'}."

                        ),

                    )

            db.commit()

            db.refresh(row)

        except HTTPException:

            db.rollback()

            raise

        except Exception as e:

            logger.exception(

                "actualizar_pago pago_id=%s: fallo reaplicar cascada tras cambio de articulacion",

                pago_id,

            )

            db.rollback()

            raise HTTPException(

                status_code=500,

                detail=f"Error al recalcular cuotas tras editar el pago: {str(e)[:280]}",

            ) from e

        return _pago_response_enriquecido(db, row)

    # Regla: si el pago cumple validadores (prestamo_id + monto), aplicar automáticamente a cuotas en cualquier canal

    if row.prestamo_id and float(row.monto_pagado or 0) > 0:

        try:

            cuotas_completadas, cuotas_parciales = _aplicar_pago_a_cuotas_interno(row, db)

            # Misma regla que crear_pago: alinear estado y conciliado si no hubo abono en cuotas.
            row.estado = _estado_conciliacion_post_cascada(row, cuotas_completadas, cuotas_parciales)

            db.commit()

            db.refresh(row)

        except IntegrityError as e:

            try:
                db.rollback()
            except Exception:
                logger.exception("Rollback tras IntegrityError cascada (actualizar_pago pago_id=%s)", pago_id)

            pg2, cn2 = _integridad_error_pgcode_y_constraint(e)

            logger.warning(
                "actualizar_pago pago_id=%s commit post-cascada IntegrityError pgcode=%s cname=%s %s",
                pago_id,
                pg2,
                cn2,
                str(e)[:400],
            )

            raise HTTPException(
                status_code=400,
                detail=(
                    "Integridad al guardar tras aplicar cuotas al pago "
                    f"(código {pg2 or '?'}, restricción {cn2 or 'N/A'}). "
                    "Revise montos y cuotas; use «Aplicar pagos a cuotas» por préstamo si corresponde."
                ),
            ) from e

        except Exception as e:

            logger.warning(
                "Al actualizar pago, no se pudo aplicar a cuotas: %s",
                e,
                exc_info=True,
            )
            try:
                db.rollback()
            except Exception:
                logger.exception("Rollback tras fallo aplicacion cuotas (actualizar_pago pago_id=%s)", pago_id)
            else:
                try:
                    db.refresh(row)
                except Exception:
                    logger.warning(
                        "No se pudo refrescar pago id=%s tras rollback de cascada; se responde con estado en memoria.",
                        pago_id,
                    )

    return _pago_response_enriquecido(db, row)


def _mensaje_sin_aplicacion_cascada(diagnostico: dict[str, Any]) -> str:
    """Texto explicativo cuando pagos_con_aplicacion == 0 sin reaplicación completa."""
    n_no = int(diagnostico.get("pagos_no_elegibles_sin_cuota_pagos") or 0)
    n_eleg = int(diagnostico.get("pagos_elegibles_cascada_sin_cuota_pagos") or 0)
    n_oper = int(diagnostico.get("pagos_operativos_sin_cuota_pagos") or 0)
    sin_abono = diagnostico.get("pagos_con_intento_sin_abono_ids") or []
    errs = diagnostico.get("errores_por_pago") or []
    partes: list[str] = []
    if n_no > 0 and n_eleg == 0:
        partes.append(
            f"Hay {n_no} pago(s) sin articulación en cuota_pagos que no cumplen criterio de elegibilidad "
            "(conciliado, verificado SÍ o estado PAGADO; excluye anulados/rechazados). "
            "Concilie o verifique esos pagos en el módulo Pagos."
        )
    elif n_eleg > 0 and isinstance(sin_abono, list) and len(sin_abono) > 0:
        muestra = ", ".join(str(x) for x in sin_abono[:20])
        suf = "…" if len(sin_abono) > 20 else ""
        partes.append(
            f"Se intentó con {n_eleg} pago(s) elegible(s) ordenados por fecha, "
            f"pero ninguno generó abono en cuotas (sin saldo pendiente en cuotas en BD o bloqueo). "
            f"IDs: {muestra}{suf}."
        )
    elif n_oper == 0:
        partes.append(
            "No hay pagos operativos con monto > 0 pendientes de articulación para este crédito "
            "(todos tienen al menos una fila en cuota_pagos a nivel global, o no hay registros aplicables). "
            "Eso no implica que el dinero esté bien repartido en las cuotas de este préstamo: "
            "si el sistema detecta desajuste entre pagos elegibles y monto articulado a sus cuotas, "
            "reconstruye la cascada automáticamente al pulsar de nuevo."
        )
    if isinstance(errs, list) and len(errs) > 0:
        partes.append(f"Fallos al aplicar en {len(errs)} pago(s); revise logs o integridad.")
    if not partes:
        partes.append(
            "No quedaban pagos elegibles sin filas en cuota_pagos (o monto 0). "
            "La reaplicación completa corre si hay inconsistencia de integridad en cuotas "
            "o si el monto elegible en pagos supera lo articulado a las cuotas de este préstamo."
        )
    return "Ningún pago nuevo se articuló en cuotas. " + " ".join(partes)


@router.post("/por-prestamo/{prestamo_id:int}/aplicar-pagos-cuotas", response_model=dict)

def aplicar_pagos_pendientes_cuotas_por_prestamo(

    prestamo_id: int,

    db: Session = Depends(get_db),

):

    """

    Aplica en cascada (por fecha_pago, luego id) los pagos del préstamo que aún no tienen

    filas en cuota_pagos y cumplen criterios de elegibilidad (conciliado / verificado / PAGADO).

    Por cada pago, el reparto a cuotas sigue el orden numero_cuota (cascada / waterfall).

    Persiste en BD. Útil tras editar/crear pagos en revisión manual o regenerar cuotas.

    """

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    diagnostico: dict[str, Any] = {}

    try:

        res_primera = aplicar_pagos_pendientes_prestamo_con_diagnostico(prestamo_id, db)

        n = int(res_primera.get("pagos_con_aplicacion") or 0)

        diagnostico = dict(res_primera.get("diagnostico") or {})

        reaplicacion_completa = False

        detalle_reaplicacion: Optional[dict] = None

        if n == 0 and prestamo_requiere_correccion_cascada(db, prestamo_id):

            detalle_reaplicacion = reset_y_reaplicar_cascada_prestamo(db, prestamo_id)

            reaplicacion_completa = True

            if not detalle_reaplicacion.get("ok"):

                raise HTTPException(

                    status_code=400,

                    detail=str(

                        detalle_reaplicacion.get("error")

                        or "No se pudo reconstruir la cascada de cuotas."

                    ),

                )

            n = int(detalle_reaplicacion.get("pagos_reaplicados") or 0)

        db.commit()

    except HTTPException:

        db.rollback()

        raise

    except Exception as e:

        db.rollback()

        logger.exception(

            "aplicar-pagos-cuotas por prestamo_id=%s: %s",

            prestamo_id,

            e,

        )

        raise HTTPException(

            status_code=500,

            detail=f"Error al aplicar pagos a cuotas: {str(e)}",

        ) from e

    if n > 0:

        if reaplicacion_completa:

            mensaje = (

                f"Amortización recalculada: se reinició la aplicación a cuotas y "

                f"{n} pago(s) quedaron distribuidos (cascada)."

            )

        else:

            mensaje = f"Cascada aplicada: {n} pago(s) con abono efectivo en cuotas."

    elif reaplicacion_completa:

        mensaje = (

            "Tabla de amortización reiniciada; no había pagos elegibles para volver a aplicar "

            "(conciliado / verificado / PAGADO, monto > 0). Revise la conciliación de los pagos."

        )

    else:

        mensaje = _mensaje_sin_aplicacion_cascada(diagnostico)

    return {

        "prestamo_id": prestamo_id,

        "pagos_con_aplicacion": n,

        "reaplicacion_completa": reaplicacion_completa,

        "detalle_reaplicacion": detalle_reaplicacion,

        "mensaje": mensaje,

        "diagnostico": diagnostico,

    }





@router.delete("/por-prestamo/{prestamo_id:int}/todos", response_model=dict)

def eliminar_todos_pagos_por_prestamo(prestamo_id: int, db: Session = Depends(get_db)):

    """

    Elimina todos los pagos asociados al préstamo, limpia dependencias y reinicia totales en cuotas.

    Solo préstamos en estado APROBADO (flujo «reemplazar pagos» antes de carga masiva desde Excel).

    """

    from app.services.pagos_cuotas_reaplicacion import eliminar_todos_pagos_prestamo

    try:

        r = eliminar_todos_pagos_prestamo(db, prestamo_id)

        if not r.get("ok"):

            msg = str(r.get("error") or "No se pudo eliminar")

            if "no encontrado" in msg.lower():

                raise HTTPException(status_code=404, detail=msg)

            raise HTTPException(status_code=400, detail=msg)

        db.commit()

    except HTTPException:

        db.rollback()

        raise

    except Exception as e:

        db.rollback()

        logger.exception(

            "eliminar_todos_pagos_por_prestamo prestamo_id=%s: %s",

            prestamo_id,

            e,

        )

        raise HTTPException(status_code=500, detail=str(e)[:300]) from e

    return r





@router.delete("/{pago_id:int}", status_code=204)

def eliminar_pago(pago_id: int, db: Session = Depends(get_db)):

    """Elimina un pago, limpia dependencias y, si tiene crédito, alinea cuotas vía cascada.

    Con `prestamo_id`: tras borrar el pago se ejecuta `reset_y_reaplicar_cascada_prestamo`
    (limpia `reporte_contable_cache`, recalcula `total_pagado` / estados y reaplica pagos restantes).
    Sin crédito: basta el borrado del pago (CASCADE en `cuota_pagos` si hubiera filas huérfanas).
    """

    row = db.get(Pago, pago_id)

    if not row:

        raise HTTPException(status_code=404, detail="Pago no encontrado")

    prestamo_id_previo = row.prestamo_id

    try:
        db.execute(text("DELETE FROM auditoria_conciliacion_manual WHERE pago_id = :pid"), {"pid": pago_id})
        db.execute(text("DELETE FROM cuota_pagos WHERE pago_id = :pid"), {"pid": pago_id})
        db.execute(text("UPDATE cuotas SET pago_id = NULL WHERE pago_id = :pid"), {"pid": pago_id})
        db.execute(text("DELETE FROM revisar_pagos WHERE pago_id = :pid"), {"pid": pago_id})

        db.delete(row)
        db.flush()

        if prestamo_id_previo:
            from app.services.pagos_cuotas_reaplicacion import reset_y_reaplicar_cascada_prestamo

            r = reset_y_reaplicar_cascada_prestamo(db, prestamo_id_previo)
            if not r.get("ok"):
                raise HTTPException(
                    status_code=500,
                    detail=(r.get("error") or "No se pudo alinear cuotas tras eliminar el pago")[:300],
                )

        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        import logging
        logging.getLogger(__name__).error("Error eliminando pago %s: %s", pago_id, e)
        raise HTTPException(
            status_code=500,
            detail=f"Error al eliminar pago {pago_id}: {str(e)[:200]}"
        ) from e

    return None


@router.get("/diagnostico/pago/{pago_id}", response_model=dict)
def diagnostico_pago(pago_id: int, db: Session = Depends(get_db)):
    """Verifica si un pago existe en BD y muestra sus dependencias."""
    info: dict = {"pago_id": pago_id, "existe_en_pagos": False, "dependencias": {}}

    row = db.execute(
        text("SELECT id, prestamo_id, cedula_cliente, monto_pagado, fecha_pago, numero_documento, estado, ref_norm FROM pagos WHERE id = :pid"),
        {"pid": pago_id},
    ).first()

    if row:
        info["existe_en_pagos"] = True
        info["pago"] = {
            "id": row[0], "prestamo_id": row[1], "cedula": row[2],
            "monto": float(row[3]) if row[3] else None,
            "fecha": str(row[4]) if row[4] else None,
            "documento": row[5], "estado": row[6], "ref_norm": row[7],
        }

    cp = db.execute(text("SELECT COUNT(*) FROM cuota_pagos WHERE pago_id = :pid"), {"pid": pago_id}).scalar()
    info["dependencias"]["cuota_pagos"] = cp or 0

    acm = db.execute(text("SELECT COUNT(*) FROM auditoria_conciliacion_manual WHERE pago_id = :pid"), {"pid": pago_id}).scalar()
    info["dependencias"]["auditoria_conciliacion_manual"] = acm or 0

    cuotas_ref = db.execute(text("SELECT COUNT(*) FROM cuotas WHERE pago_id = :pid"), {"pid": pago_id}).scalar()
    info["dependencias"]["cuotas_con_pago_id"] = cuotas_ref or 0

    rp = db.execute(text("SELECT COUNT(*) FROM revisar_pagos WHERE pago_id = :pid"), {"pid": pago_id}).scalar()
    info["dependencias"]["revisar_pagos"] = rp or 0

    return info


@router.delete("/forzar-eliminar/{pago_id}", response_model=dict)
@router.post("/forzar-eliminar/{pago_id}", response_model=dict)
@router.get("/forzar-eliminar/{pago_id}", response_model=dict)
def forzar_eliminar_pago(pago_id: int, db: Session = Depends(get_db)):
    """Eliminacion forzada: limpia TODAS las dependencias y borra el pago con SQL directo.

    Si el pago tenia `prestamo_id`, tras borrarlo se reaplica la cascada del credito
    (misma logica que DELETE normal: cuotas alineadas y cache contable invalidada).
    """
    import logging
    log = logging.getLogger(__name__)

    row0 = db.execute(
        text("SELECT id, prestamo_id FROM pagos WHERE id = :pid"),
        {"pid": pago_id},
    ).first()
    if not row0:
        return {"ok": False, "detail": f"Pago {pago_id} no existe en tabla pagos"}
    prestamo_id_previo = row0[1]

    try:
        r1 = db.execute(text("DELETE FROM cuota_pagos WHERE pago_id = :pid"), {"pid": pago_id})
        r2 = db.execute(text("DELETE FROM auditoria_conciliacion_manual WHERE pago_id = :pid"), {"pid": pago_id})
        r3 = db.execute(text("UPDATE cuotas SET pago_id = NULL WHERE pago_id = :pid"), {"pid": pago_id})
        r4 = db.execute(text("DELETE FROM revisar_pagos WHERE pago_id = :pid"), {"pid": pago_id})

        # Buscar cualquier otra FK que referencie pagos.id
        fk_check = db.execute(text("""
            SELECT tc.table_name, kcu.column_name
            FROM information_schema.table_constraints tc
            JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
            JOIN information_schema.constraint_column_usage ccu ON tc.constraint_name = ccu.constraint_name
            WHERE ccu.table_name = 'pagos' AND ccu.column_name = 'id'
              AND tc.constraint_type = 'FOREIGN KEY'
              AND tc.table_name NOT IN ('cuota_pagos', 'auditoria_conciliacion_manual', 'cuotas', 'revisar_pagos')
        """)).fetchall()

        extra_cleaned = []
        for fk_row in fk_check:
            tbl, col = fk_row[0], fk_row[1]
            db.execute(text(f'DELETE FROM "{tbl}" WHERE "{col}" = :pid'), {"pid": pago_id})
            extra_cleaned.append(f"{tbl}.{col}")

        r5 = db.execute(text("DELETE FROM pagos WHERE id = :pid"), {"pid": pago_id})

        reaplicado: Optional[dict] = None
        if prestamo_id_previo:
            from app.services.pagos_cuotas_reaplicacion import reset_y_reaplicar_cascada_prestamo

            r = reset_y_reaplicar_cascada_prestamo(db, prestamo_id_previo)
            if not r.get("ok"):
                db.rollback()
                raise HTTPException(
                    status_code=500,
                    detail=(r.get("error") or "No se pudo alinear cuotas tras forzar eliminación")[:300],
                )
            reaplicado = {
                "prestamo_id": prestamo_id_previo,
                "pagos_reaplicados": r.get("pagos_reaplicados"),
                "cache_contable_eliminadas": r.get("cache_contable_eliminadas"),
            }

        db.commit()

        log.info("Pago %s eliminado forzadamente. cuota_pagos=%s acm=%s cuotas=%s revisar=%s pagos=%s extra=%s",
                 pago_id, r1.rowcount, r2.rowcount, r3.rowcount, r4.rowcount, r5.rowcount, extra_cleaned)

        out: dict = {
            "ok": True,
            "eliminado": {
                "cuota_pagos": r1.rowcount, "auditoria_conciliacion_manual": r2.rowcount,
                "cuotas_nulled": r3.rowcount, "revisar_pagos": r4.rowcount,
                "pagos": r5.rowcount, "extra_fks": extra_cleaned,
            },
        }
        if reaplicado is not None:
            out["cascada_tras_eliminar"] = reaplicado
        return out
    except Exception as e:
        db.rollback()
        log.error("Error forzar-eliminar pago %s: %s", pago_id, e)
        raise HTTPException(status_code=500, detail=f"Error: {str(e)[:300]}") from e




def aplicar_pagos_pendientes_prestamo_con_diagnostico(prestamo_id: int, db: Session) -> dict[str, Any]:
    """
    Igual que aplicar_pagos_pendientes_prestamo pero devuelve diagnóstico para UI y soporte.

    diagnostico incluye conteos antes de aplicar y listas de pagos sin abono o con error.
    """
    vacio: dict[str, Any] = {
        "pagos_operativos_sin_cuota_pagos": 0,
        "pagos_elegibles_cascada_sin_cuota_pagos": 0,
        "pagos_no_elegibles_sin_cuota_pagos": 0,
        "pagos_con_intento_sin_abono_ids": [],
        "errores_por_pago": [],
    }
    prestamo_chk = db.get(Prestamo, prestamo_id)
    if prestamo_chk and (prestamo_chk.estado or "").strip().upper() == "DESISTIMIENTO":
        return {"pagos_con_aplicacion": 0, "diagnostico": vacio}

    subq = select(CuotaPago.pago_id).where(CuotaPago.pago_id.isnot(None)).distinct()
    base_operativo = and_(
        Pago.prestamo_id == prestamo_id,
        Pago.monto_pagado > 0,
        ~Pago.id.in_(subq),
        not_(_where_pago_excluido_operacion()),
    )
    n_oper = int(db.scalar(select(func.count()).select_from(Pago).where(base_operativo)) or 0)

    rows = db.execute(
        select(Pago)
        .where(
            Pago.prestamo_id == prestamo_id,
            _where_pago_elegible_reaplicacion_cascada(),
            Pago.monto_pagado > 0,
            ~Pago.id.in_(subq),
        )
        .order_by(Pago.fecha_pago.asc().nulls_last(), Pago.id.asc())
    ).scalars().all()

    n_eleg = len(rows)
    n_no_eleg = max(0, n_oper - n_eleg)

    n = 0
    sin_abono: list[int] = []
    errores: list[dict[str, Any]] = []

    for pago in rows:
        try:
            cc, cp = _aplicar_pago_a_cuotas_interno(pago, db)
            if cc > 0 or cp > 0:
                pago.estado = "PAGADO"
                n += 1
            else:
                sin_abono.append(int(pago.id))
        except Exception as e:
            logger.warning(
                "aplicar_pagos_pendientes_prestamo prestamo_id=%s pago id=%s: %s",
                prestamo_id,
                pago.id,
                e,
            )
            errores.append({"pago_id": int(pago.id), "error": str(e)})

    return {
        "pagos_con_aplicacion": n,
        "diagnostico": {
            "pagos_operativos_sin_cuota_pagos": n_oper,
            "pagos_elegibles_cascada_sin_cuota_pagos": n_eleg,
            "pagos_no_elegibles_sin_cuota_pagos": n_no_eleg,
            "pagos_con_intento_sin_abono_ids": sin_abono,
            "errores_por_pago": errores,
        },
    }


def aplicar_pagos_pendientes_prestamo(prestamo_id: int, db: Session) -> int:
    """
    Aplica a cuotas los pagos del préstamo que aún no tienen enlaces en cuota_pagos.

    Criterio de elegibilidad: conciliado, verificado_concordancia SI, o estado PAGADO;
    excluye anulados/reversados/duplicado declarado (alineado con auditoria cartera).

    No hace commit. Retorna el número de pagos a los que se les aplicó algo (cc o cp > 0).
    """
    return int(aplicar_pagos_pendientes_prestamo_con_diagnostico(prestamo_id, db)["pagos_con_aplicacion"])





@router.post("/{pago_id:int}/aplicar-cuotas", response_model=dict)

def aplicar_pago_a_cuotas(pago_id: int, db: Session = Depends(get_db)):

    """

    Aplica el monto del pago a cuotas del prestamo (orden numero_cuota).

    pago.estado = PAGADO solo si hubo articulacion (cc o cp > 0); si no hubo cupo, PENDIENTE.

    """

    pago = db.get(Pago, pago_id)

    if not pago:

        raise HTTPException(status_code=404, detail="Pago no encontrado")

    if not pago.prestamo_id:

        return {

            "success": False,

            "cuotas_completadas": 0,

            "cuotas_parciales": 0,

            "message": "El pago no tiene préstamo asociado.",

        }

    monto_restante = float(pago.monto_pagado) if pago.monto_pagado else 0

    if monto_restante <= 0:

        return {"success": True, "cuotas_completadas": 0, "cuotas_parciales": 0, "message": "Monto del pago es cero."}

    if pago_tiene_aplicaciones_cuotas(db, pago.id):
        return {
            "success": True,
            "ya_aplicado": True,
            "cuotas_completadas": 0,
            "cuotas_parciales": 0,
            "message": "El pago ya tiene aplicaciones en cuota_pagos. No se duplico la cascada.",
        }

    try:

        cuotas_completadas, cuotas_parciales = _aplicar_pago_a_cuotas_interno(pago, db)

        pago.estado = _estado_conciliacion_post_cascada(pago, cuotas_completadas, cuotas_parciales)

        db.commit()

        return {

            "success": True,

            "ya_aplicado": False,

            "cuotas_completadas": cuotas_completadas,

            "cuotas_parciales": cuotas_parciales,

            "message": f"Se aplicó el pago: {cuotas_completadas} cuota(s) completadas, {cuotas_parciales} parcial(es).",

        }

    except HTTPException:

        raise

    except ValueError as e:

        db.rollback()

        logger.warning("aplicar-cuotas integridad pago_id=%s: %s", pago_id, e)

        raise HTTPException(status_code=409, detail=str(e)) from e

    except Exception as e:

        db.rollback()

        logger.exception("Error aplicar-cuotas pago_id=%s: %s", pago_id, e)

        raise HTTPException(

            status_code=500,

            detail=f"Error al aplicar el pago a cuotas: {str(e)}",

        ) from e





# --- Cédulas permitidas para reportar en Bs (rapicredit-cobros / infopagos) ---
# Normalización canónica: app.services.cobros.cedula_reportar_bs_service


@router.get("/cedulas-reportar-bs", response_model=dict)

def get_cedulas_reportar_bs(

    page: int = Query(1, ge=1, description="Página (1-based)"),

    page_size: int = Query(10, ge=1, le=100, description="Filas por página"),

    db: Session = Depends(get_db),

):

    """

    Total de cédulas en la lista y página ordenada del más reciente al más antiguo (creado_en DESC).

    """

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    q = (

        db.query(CedulaReportarBs)

        .order_by(desc(CedulaReportarBs.creado_en), CedulaReportarBs.cedula.asc())

        .offset((page - 1) * page_size)

        .limit(page_size)

    )

    rows = q.all()

    items = [

        {

            "cedula": r.cedula,

            "creado_en": r.creado_en.isoformat() if r.creado_en else None,

        }

        for r in rows

    ]

    return {

        "total": int(total),

        "page": page,

        "page_size": page_size,

        "items": items,

    }





@router.get("/cedulas-reportar-bs/consultar", response_model=dict)


def consultar_cedula_reportar_bs(

    cedula: str = Query(..., min_length=1, description="Cedula a verificar"),

    db: Session = Depends(get_db),

):

    """Indica si la cedula esta autorizada para reportar pagos en Bs (lista cedulas_reportar_bs)."""

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    raw = (cedula or "").strip()

    en_lista = cedula_autorizada_para_bs(db, raw)

    ced_norm = normalize_cedula_para_almacenar_lista_bs(raw)

    return {

        "cedula_ingresada": raw,

        "cedula_normalizada": ced_norm,

        "en_lista": en_lista,

        "total_en_lista": int(total),

    }





@router.post("/cedulas-reportar-bs/consultar-batch", response_model=dict)

def consultar_cedulas_reportar_bs_batch(

    payload: ConsultarCedulasReportarBsBatchBody,

    db: Session = Depends(get_db),

):

    """Varias cedulas en una sola peticion: carga la lista autorizada una vez (evita N+1)."""

    total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

    claves = load_autorizados_bs_claves(db)

    por_cedula: dict = {}

    for raw in payload.cedulas or []:

        raw = (raw or "").strip()

        if not raw:

            continue

        norm = normalize_cedula_lookup_key(raw)

        en_lista = cedula_coincide_autorizados_bs(norm, claves)

        ced_norm = normalize_cedula_para_almacenar_lista_bs(raw)

        por_cedula[raw] = {

            "cedula_ingresada": raw,

            "cedula_normalizada": ced_norm,

            "en_lista": en_lista,

            "total_en_lista": int(total),

        }

    return {"total_en_lista": int(total), "por_cedula": por_cedula}





@router.post("/cedulas-reportar-bs/agregar", response_model=dict)

def agregar_cedula_reportar_bs(

    payload: AgregarCedulaReportarBsBody,

    db: Session = Depends(get_db),

):

    """

    Agrega una cédula a la lista de quienes pueden reportar en Bs (nuevo cliente que paga en bolívares).

    Si ya existe, no duplica. La cédula se normaliza (V/E/J/G + dígitos).

    """

    cedula_norm = normalize_cedula_para_almacenar_lista_bs(payload.cedula)

    if not cedula_norm:

        raise HTTPException(

            status_code=400,

            detail="Cédula inválida. Use letra V, E, J o G seguida de 6 a 11 dígitos (ej: V12345678).",

        )

    claves_actuales = load_autorizados_bs_claves(db)

    if cedula_coincide_autorizados_bs(cedula_norm, claves_actuales):

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        return {

            "agregada": False,

            "cedula": cedula_norm,

            "total": total,

            "mensaje": f"La cédula {cedula_norm} ya estaba en la lista.",

        }

    try:

        db.add(CedulaReportarBs(cedula=cedula_norm))

        db.commit()

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        return {

            "agregada": True,

            "cedula": cedula_norm,

            "total": total,

            "mensaje": f"Cédula {cedula_norm} agregada. Ya puede reportar pagos en Bs en Cobros e Infopagos.",

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error agregando cedula_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al agregar la cédula.") from e





@router.post("/cedulas-reportar-bs/upload", response_model=dict)

def upload_cedulas_reportar_bs(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

):

    """

    Carga un Excel con columna 'cedula'. Reemplaza la lista de cédulas que pueden reportar en Bs.

    Afecta a rapicredit-cobros e infopagos: solo esas cédulas verán habilitada la opción Bs.

    """

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):

        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")

    try:

        import openpyxl

    except ImportError:

        raise HTTPException(status_code=500, detail="Excel library not available")

    content = file.file.read()

    if len(content) < 50:

        raise HTTPException(status_code=400, detail="El archivo está vacío o no es un Excel válido.")

    try:

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    except Exception as e:

        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {str(e)}") from e

    ws = wb.active

    if not ws:

        raise HTTPException(status_code=400, detail="El Excel no tiene hojas.")

    # Buscar columna "cedula" (primera fila)

    header = [ (c and str(c).strip().lower() if c is not None else "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) ]

    col_cedula = None

    for i, h in enumerate(header):

        if h == "cedula":

            col_cedula = i

            break

    if col_cedula is None:

        raise HTTPException(status_code=400, detail="El Excel debe tener una columna llamada 'cedula'.")

    cedulas_unicas: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row or len(row) <= col_cedula:

            continue

        val = row[col_cedula]

        cedula_norm = normalize_cedula_para_almacenar_lista_bs(str(val).strip() if val is not None else "")

        if cedula_norm:

            cedulas_unicas.add(cedula_norm)

    wb.close()

    # Reemplazar tabla

    try:

        db.query(CedulaReportarBs).delete()

        for c in cedulas_unicas:

            db.add(CedulaReportarBs(cedula=c))

        db.commit()

    except Exception as e:

        db.rollback()

        logger.exception("Error guardando cedulas_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al guardar la lista de cédulas.") from e

    return {

        "total": len(cedulas_unicas),

        "mensaje": f"Se cargaron {len(cedulas_unicas)} cédula(s). Solo ellas pueden reportar pagos en Bs en RapiCredit Cobros e Infopagos.",

    }


@router.delete("/cedulas-reportar-bs/eliminar", response_model=dict)

def eliminar_cedula_reportar_bs(

    payload: AgregarCedulaReportarBsBody,

    db: Session = Depends(get_db),

):

    """

    Elimina una cédula de la lista de quienes pueden reportar en Bs.

    La cédula se normaliza (V/E/J/G + dígitos) antes de eliminar.

    """

    cedula_norm = normalize_cedula_para_almacenar_lista_bs(payload.cedula)

    if not cedula_norm:

        raise HTTPException(

            status_code=400,

            detail="Cédula inválida. Use letra V, E, J o G seguida de 6 a 11 dígitos (ej: V12345678).",

        )

    try:

        # Buscar y eliminar la cédula

        cedula_encontrada = db.query(CedulaReportarBs).filter(

            CedulaReportarBs.cedula == cedula_norm

        ).first()

        if not cedula_encontrada:

            total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

            return {

                "eliminada": False,

                "cedula": cedula_norm,

                "total": total,

                "mensaje": f"La cédula {cedula_norm} no se encontró en la lista.",

            }

        db.delete(cedula_encontrada)

        db.commit()

        total = db.query(func.count(CedulaReportarBs.cedula)).scalar() or 0

        return {

            "eliminada": True,

            "cedula": cedula_norm,

            "total": total,

            "mensaje": f"Cédula {cedula_norm} eliminada. Ya no podrá reportar pagos en Bs en Cobros e Infopagos.",

        }

    except Exception as e:

        db.rollback()

        logger.exception("Error eliminando cedula_reportar_bs: %s", e)

        raise HTTPException(status_code=500, detail="Error al eliminar la cédula.") from e

# Compat: nombre historico
# Compat: nombre historico (importaciones antiguas).
_estado_pago_tras_aplicar_fifo = _estado_pago_tras_aplicar_cascada
