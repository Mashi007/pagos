"""Pagos API: importar_cobros."""
"""

Endpoints de pagos. Datos reales desde BD.

- Tabla pagos: GET/POST/PUT/DELETE /pagos/ (listado y CRUD para /pagos/pagos).

- GET /pagos/kpis, /stats, /ultimos; POST /upload, /conciliacion/upload, /{id}/aplicar-cuotas.



Nº documento / referencia de pago:

- Regla de cartera: **no puede repetirse** el valor almacenado en `pagos.numero_documento` (único en `pagos` y
  `pagos_con_errores`, salvo edición del mismo `pago_id`).

- **Única forma permitida** de reutilizar el mismo texto de comprobante del banco: desambiguar con **código**
  (`codigo_documento`). En BD se compone `base + §CD: + código` (`compose_numero_documento_almacenado`,
  máx. 100 caracteres). En revisión manual en préstamos, el token **A#### / P####** lo asigna **Visto** (sufijo
  operativo; mismo criterio que carga masiva).

- Formatos de comprobante: BNC/, BINANCE, VE/, etc. Carga masiva: columna opcional a la derecha del Nº = código.

- Además, huella funcional (prestamo + fecha + monto + ref_norm) evita duplicar el mismo pago operativo (409),
  vía `ux_pagos_fingerprint_activos`.

"""

import calendar

import io

import logging

import re

import time

import uuid

from datetime import date, datetime, time as dt_time

from decimal import Decimal

from typing import Any, Optional

from zoneinfo import ZoneInfo



from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body, Request

from fastapi.responses import StreamingResponse, Response

from pydantic import BaseModel, field_validator

from sqlalchemy import and_, case, delete, desc, exists, func, inspect, not_, or_, select, text

from sqlalchemy.orm import Session, aliased

from sqlalchemy.exc import IntegrityError, OperationalError



from app.core.database import get_db

from app.core.config import settings

from app.core.deps import (
    ComprobanteImagenReader,
    get_comprobante_imagen_reader,
    get_current_user,
    require_admin,
    require_operator_or_higher,
)
from app.core.rol_normalization import canonical_rol

from app.core.documento import (
    compose_numero_documento_almacenado,
    normalize_codigo_documento,
    normalize_documento,
    split_numero_documento_almacenado,
)
from app.utils.cedula_almacenamiento import (
    CedulaPagoFkError,
    asegurar_cedula_pago_para_fk,
    alinear_cedulas_clientes_existentes,
    normalizar_cedula_almacenamiento,
)
from app.services.pago_numero_documento import (
    numero_documento_ya_registrado,
    primer_pago_cartera_por_documento,
)

from app.core.serializers import to_float, format_date_iso

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.prestamo import Prestamo

from app.models.pago import Pago

from app.models.pago_comprobante_imagen import PagoComprobanteImagen

from app.models.pago_con_error import PagoConError

from app.models.revisar_pago import RevisarPago

from app.models.cuota_pago import CuotaPago

from app.models.pago_reportado import PagoReportado

from app.models.datos_importados_conerrores import DatosImportadosConErrores

from app.models.cedula_reportar_bs import CedulaReportarBs

from app.schemas.pago import (
    PagoCreate,
    PagoUpdate,
    PagoResponse,
    normalizar_link_comprobante,
)

from app.schemas.auth import UserResponse

from app.services.cobros.pago_reportado_documento import (
    claves_documento_pago_desde_campos,
    claves_documento_pago_para_reportado,
    claves_documento_para_lote_reportados,
    documento_numero_desde_pago_reportado,
    pago_reportado_colisiona_tabla_pagos,
)
from app.services.pagos.comprobante_link_desde_gmail import (
    enriquecer_items_link_comprobante_desde_gmail,
    enriquecer_items_link_comprobante_desde_pago_reportado,
)
from app.services.pagos.comprobante_imagen_finiquito_access import (
    comprobante_imagen_accesible_finiquito_portal,
)
from app.services.cuota_pago_integridad import (
    pago_tiene_aplicaciones_cuotas,
    validar_suma_aplicada_vs_monto_pago,
)
from app.services.pagos_cuotas_reaplicacion import (
    prestamo_requiere_correccion_cascada,
    reset_y_reaplicar_cascada_prestamo,
)
from app.services.pagos_cascada_aplicacion import (
    _aplicar_pago_a_cuotas_interno,
    _marcar_prestamo_liquidado_si_corresponde,
)
from app.services.pagos_aplicacion_prestamo import (
    aplicar_pagos_pendientes_prestamo,
    aplicar_pagos_pendientes_prestamo_con_diagnostico,
)
from app.services.pagos_cascada_mensajes import _mensaje_sin_aplicacion_cascada


from app.services.tasa_cambio_service import (
    convertir_bs_a_usd,
    normalizar_fuente_tasa,
    obtener_tasa_por_fecha,
    valor_tasa_para_fuente,
)

from app.services.pago_registro_moneda import (
    normalizar_moneda_registro,
    resolver_monto_registro_pago,
    preload_autorizados_bs,
)
from app.services.pago_huella_funcional import (
    conflicto_huella_para_creacion,
    contar_prestamos_con_huella_funcional_duplicada,
    HTTP_409_DETAIL_HUELLA_FUNCIONAL,
    mensaje_409_huella_funcional_con_id,
    primer_id_conflicto_huella_funcional,
    primer_par_huella_duplicada_prestamo,
    ref_norm_desde_campos,
)
from app.services.pago_huella_metricas import (
    registrar_rechazo_huella_funcional,
    snapshot_rechazos_huella_funcional,
)

from app.services.cobros.cedula_reportar_bs_service import (

    normalize_cedula_para_almacenar_lista_bs,

    normalize_cedula_lookup_key,

    load_autorizados_bs_claves,

    cedula_coincide_autorizados_bs,

    cedula_autorizada_para_bs,

    obtener_fuente_tasa_lista_bs,

    fuente_tasa_bs_efectiva_para_cedula,

)

from app.services.tasa_cambio_service import normalizar_fuente_tasa

from .payload_models import (
    AgregarCedulaReportarBsBody,
    ConsultarCedulasReportarBsBatchBody,
    GuardarFilaEditableBody,
    MoverRevisarPagosBody,
    PagoBatchBody,
    ValidarFilasBatchBody,
)
from app.services.cobros.cobros_publico_reporte_service import (
    mime_efectivo_comprobante_web,
    mime_efectivo_con_firma_archivo,
    preparar_adjunto_comprobante_para_vision,
    validate_file_magic,
)

from .comprobante_imagen_helpers import (
    _COMPROBANTE_IMG_CT,
    _MAX_COMPROBANTE_IMAGEN_BYTES,
    _normalizar_id_comprobante_imagen,
    _public_base_url_para_comprobante,
)
from app.services.pagos_gmail.comprobante_bd import url_comprobante_imagen_absoluta

from .constants import (
    TZ_NEGOCIO,
    _MAX_LEN_NUMERO_DOCUMENTO,
    _MAX_MONTO_PAGADO,
    _MIN_MONTO_PAGADO,
    _PRESTAMO_ID_MAX,
)
from .cascada_estado import _estado_pago_tras_aplicar_cascada
from .sql_where_pagos import (
    _where_pago_elegible_reaplicacion_cascada,
    _where_pago_excluido_operacion,
)
from .pago_integridad_db import _integridad_error_pgcode_y_constraint
from .pago_normalizacion import (
    _celda_a_string_documento,
    _normalizar_ref_fingerprint,
    _safe_float,
    _validar_monto,
)
from .pago_zona_horaria import _calcular_dias_mora, _hoy_local
from .pago_usuario_registro import _usuario_registro_desde_current_user
from .pago_conciliacion_estado import (
    _alinear_estado_si_toggle_conciliado_actualizar_pago,
    _estado_conciliacion_post_cascada,
)
from .pago_cascada_reglas import _debe_aplicar_cascada_pago
from .pago_serializacion_respuesta import (
    _enriquecer_items_duplicado_clave_misma_pagina,
    _enriquecer_items_tiene_aplicacion_cuotas,
    _enriquecer_pagos_pago_reportado_id,
    _pago_response_enriquecido,
    _pago_to_response,
)









logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])

@router.post("/importar-desde-cobros", response_model=dict)

def importar_reportados_aprobados_a_pagos(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Pasa los pagos reportados (módulo Cobros) con estado aprobado a la tabla pagos.

    Mismas reglas que carga masiva: documento único, cédula con 1 crédito activo (APROBADO).

    Los que no cumplen se guardan en datos_importados_conerrores; descargar Excel desde el frontend.

    Los pagos válidos se aplican a cuotas del préstamo (asignación en cascada) en el mismo request.

    """

    rows = db.execute(

        select(PagoReportado).where(PagoReportado.estado == "aprobado").order_by(PagoReportado.id)

    ).scalars().all()

    reportados = [r for r in rows if r is not None]

    if not reportados:

        return {

            "registros_procesados": 0,

            "registros_con_error": 0,

            "errores_detalle": [],

            "ids_pagos_con_errores": [],

            "total_datos_revisar": 0,

            "mensaje": "No hay pagos reportados aprobados para importar.",

            "cuotas_aplicadas": 0,

            "operaciones_cuota_total": 0,

            "pagos_con_aplicacion_a_cuotas": 0,

            "pagos_sin_aplicacion_cuotas": [],

            "pagos_sin_aplicacion_cuotas_total": 0,

            "pagos_sin_aplicacion_cuotas_truncados": False,

        }



    usuario_registro = _usuario_registro_desde_current_user(current_user)

    documentos_ya_en_bd: set[str] = set()

    todas_claves_lote = claves_documento_para_lote_reportados(reportados)

    if todas_claves_lote:

        cl_list = list(todas_claves_lote)

        _chunk_claves = 800

        for i in range(0, len(cl_list), _chunk_claves):

            chunk = cl_list[i : i + _chunk_claves]

            existentes = db.execute(

                select(Pago.numero_documento).where(Pago.numero_documento.in_(chunk))

            ).scalars().all()

            documentos_ya_en_bd.update(str(d) for d in existentes if d)



    registros_procesados = 0

    ids_pagos_con_errores: list[int] = []

    errores_detalle: list[dict] = []

    docs_en_lote: set[str] = set()

    huellas_funcional_lote_cobros: set[tuple[int, str, str, str]] = set()

    pagos_creados: list[Pago] = []



    for pr in reportados:

        res = importar_un_pago_reportado_a_pagos(

            db,

            pr,

            usuario_email=usuario_registro,

            documentos_ya_en_bd=documentos_ya_en_bd,

            docs_en_lote=docs_en_lote,

            registrar_error_en_tabla=True,

            huellas_funcional_lote=huellas_funcional_lote_cobros,

        )

        if res.get("ok"):

            registros_procesados += 1

            pagos_creados.append(res["pago"])

            continue

        pce_id = res.get("pce_id")

        if pce_id is not None:

            ids_pagos_con_errores.append(pce_id)

        errores_detalle.append(

            {"referencia": res.get("referencia"), "error": res.get("error", "Error")}

        )





    operaciones_cuota_total = 0

    pagos_con_aplicacion_a_cuotas = 0

    pagos_sin_aplicacion_cuotas: list[dict] = []

    for p in pagos_creados:

        pid = getattr(p, "id", None)

        ced = (getattr(p, "cedula_cliente", None) or "") or ""

        pr_id = getattr(p, "prestamo_id", None)

        try:

            cc, cp = _aplicar_pago_a_cuotas_interno(p, db)

            p.estado = _estado_pago_tras_aplicar_cascada(cc, cp)

            if cc > 0 or cp > 0:

                operaciones_cuota_total += cc + cp

                pagos_con_aplicacion_a_cuotas += 1

            elif pr_id and float(p.monto_pagado or 0) > 0:

                pagos_sin_aplicacion_cuotas.append(

                    {

                        "pago_id": pid,

                        "cedula_cliente": ced,

                        "prestamo_id": pr_id,

                        "motivo": "sin_cuotas_afectadas",

                        "detalle": "Ninguna cuota recibió monto; revise cuotas pendientes o use Aplicar a cuotas.",

                    }

                )

        except Exception as e:

            logger.warning(

                "Importar Cobros: no se pudo aplicar pago id=%s a cuotas: %s",

                getattr(p, "id", "?"),

                e,

            )

            pagos_sin_aplicacion_cuotas.append(

                {

                    "pago_id": pid,

                    "cedula_cliente": ced,

                    "prestamo_id": pr_id,

                    "motivo": "error",

                    "detalle": str(e),

                }

            )

    db.commit()

    total_datos_revisar = db.execute(select(func.count()).select_from(DatosImportadosConErrores)).scalar() or 0

    n_sin_aplicacion = len(pagos_sin_aplicacion_cuotas)

    _limite_sin_aplicacion = 50

    mensaje_base = (

        f"Importados {registros_procesados} pagos desde Cobros; "

        f"{len(ids_pagos_con_errores)} con error (revisar: descargar Excel)."

    )

    if n_sin_aplicacion > 0:

        mensaje_base += (

            f" Atención: {n_sin_aplicacion} pago(s) creado(s) pero sin aplicar a cuotas "

            "(error o sin cuotas pendientes según reglas); use Aplicar a cuotas o revise el préstamo."

        )

    return {

        "registros_procesados": registros_procesados,

        "registros_con_error": len(ids_pagos_con_errores),

        "errores_detalle": errores_detalle[:100],

        "ids_pagos_con_errores": ids_pagos_con_errores,

        "cuotas_aplicadas": operaciones_cuota_total,

        "operaciones_cuota_total": operaciones_cuota_total,

        "pagos_con_aplicacion_a_cuotas": pagos_con_aplicacion_a_cuotas,

        "pagos_sin_aplicacion_cuotas": pagos_sin_aplicacion_cuotas[:_limite_sin_aplicacion],

        "pagos_sin_aplicacion_cuotas_total": n_sin_aplicacion,

        "pagos_sin_aplicacion_cuotas_truncados": n_sin_aplicacion > _limite_sin_aplicacion,

        "total_datos_revisar": total_datos_revisar,

        "mensaje": mensaje_base,

    }





@router.get("/importar-desde-cobros/datos-revisar", response_model=dict)

def get_datos_revisar_importados(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """Devuelve si hay datos que revisar (tabla datos_importados_conerrores) y el total. Para mostrar dialogo tras importar."""

    total = db.execute(select(func.count()).select_from(DatosImportadosConErrores)).scalar() or 0

    return {"tiene_datos": total > 0, "total": total}





@router.get("/importar-desde-cobros/descargar-excel-errores")

def descargar_excel_errores_importados(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """Genera Excel con registros de datos_importados_conerrores (mismas columnas que Revisar Pagos). Tras generar, vacia la tabla."""

    rows = db.execute(

        select(DatosImportadosConErrores).order_by(DatosImportadosConErrores.id)

    ).scalars().all()

    rows = [r for r in rows if r is not None]

    import openpyxl

    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Datos con errores"

    headers = [

        "id", "cedula_cliente", "prestamo_id", "fecha_pago", "monto_pagado", "numero_documento",

        "estado", "referencia_pago", "errores_descripcion", "observaciones", "fila_origen", "referencia_interna", "created_at"

    ]

    for col, h in enumerate(headers, 1):

        ws.cell(row=1, column=col, value=h)

    for row_idx, r in enumerate(rows, 2):

        fp = r.fecha_pago

        fecha_str = fp.strftime("%Y-%m-%d") if fp else ""

        err_str = "; ".join(r.errores_descripcion) if isinstance(r.errores_descripcion, list) else str(r.errores_descripcion or "")

        created_str = r.created_at.strftime("%Y-%m-%d %H:%M") if getattr(r, "created_at", None) else ""

        ws.cell(row=row_idx, column=1, value=r.id)

        ws.cell(row=row_idx, column=2, value=r.cedula_cliente or "")

        ws.cell(row=row_idx, column=3, value=r.prestamo_id)

        ws.cell(row=row_idx, column=4, value=fecha_str)

        ws.cell(row=row_idx, column=5, value=float(r.monto_pagado) if r.monto_pagado is not None else 0)

        ws.cell(row=row_idx, column=6, value=r.numero_documento or "")

        ws.cell(row=row_idx, column=7, value=r.estado or "")

        ws.cell(row=row_idx, column=8, value=r.referencia_pago or "")

        ws.cell(row=row_idx, column=9, value=err_str)

        ws.cell(row=row_idx, column=10, value=r.observaciones or "")

        ws.cell(row=row_idx, column=11, value=r.fila_origen)

        ws.cell(row=row_idx, column=12, value=getattr(r, "referencia_interna", None) or "")

        ws.cell(row=row_idx, column=13, value=created_str)

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    db.execute(delete(DatosImportadosConErrores))

    db.commit()

    filename = f"datos_importados_con_errores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={filename}"},

    )





