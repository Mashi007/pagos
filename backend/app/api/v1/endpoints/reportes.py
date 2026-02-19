"""
Endpoints de reportes. Datos reales desde BD.
Dashboard resumen, cartera, pagos, morosidad, financiero, asesores, productos.
Exportación Excel y PDF para cada tipo según corresponda.
"""
import calendar
import io
from datetime import date, datetime, timezone, timedelta
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy import func, select, and_, or_, delete
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.cliente import Cliente
from app.models.cuota import Cuota
from app.models.pago import Pago
from app.models.prestamo import Prestamo
from app.models.reporte_contable_cache import ReporteContableCache

router = APIRouter(dependencies=[Depends(get_current_user)])


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _parse_fecha(s: Optional[str]) -> date:
    if not s:
        return date.today()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return date.today()


def _periodos_desde_filtros(
    años_str: Optional[str],
    meses_str: Optional[str],
    meses_default: int = 12,
) -> List[tuple]:
    """
    Retorna lista de (año, mes) ordenada descendente.
    Si años_str y meses_str están presentes, usa esos. Si no, usa últimos meses_default meses.
    """
    if años_str and meses_str:
        try:
            años = sorted([int(x.strip()) for x in años_str.split(",") if x.strip()], reverse=True)
            meses = sorted([int(x.strip()) for x in meses_str.split(",") if x.strip() and 1 <= int(x.strip()) <= 12])
            if años and meses:
                periodos = [(a, m) for a in años for m in meses]
                periodos.sort(key=lambda p: (-p[0], -p[1]))
                return periodos
        except (ValueError, TypeError):
            pass
    hoy = date.today()
    result = []
    for i in range(meses_default):
        año = hoy.year
        mes = hoy.month - i
        while mes <= 0:
            mes += 12
            año -= 1
        result.append((año, mes))
    return result


# ---------- Dashboard resumen ----------
@router.get("/dashboard/resumen")
def get_resumen_dashboard(db: Session = Depends(get_db)):
    """
    Resumen para el dashboard de reportes: total_clientes, total_prestamos, total_pagos,
    cartera_activa, prestamos_mora, pagos_mes, fecha_actualizacion. Datos reales desde BD.
    """
    hoy = date.today()
    now_utc = datetime.now(timezone.utc)
    inicio_mes = hoy.replace(day=1)

    # KPIs solo incluyen clientes ACTIVOS
    total_clientes = db.scalar(
        select(func.count()).select_from(Cliente).where(Cliente.estado == "ACTIVO")
    ) or 0
    total_prestamos = db.scalar(
        select(func.count())
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
    ) or 0
    total_pagos = db.scalar(
        select(func.count())
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.isnot(None),
        )
    ) or 0
    cartera_activa = db.scalar(
        select(func.coalesce(func.sum(Cuota.monto), 0))
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
        )
    ) or 0
    subq_mora = (
        select(Cuota.prestamo_id)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < (hoy - timedelta(days=60)),  # MORA: vencido hace 61+ días
        )
        .distinct()
    )
    prestamos_mora = db.scalar(select(func.count()).select_from(subq_mora.subquery())) or 0
    # Monto total cobrado este mes (cuotas con fecha_pago en el mes actual)
    pagos_mes = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.isnot(None),
                func.date(Cuota.fecha_pago) >= inicio_mes,
                func.date(Cuota.fecha_pago) <= hoy,
            )
        )
        or 0
    )

    return {
        "total_clientes": total_clientes,
        "total_prestamos": total_prestamos,
        "total_pagos": total_pagos,
        "cartera_activa": _safe_float(cartera_activa),
        "pagos_vencidos": prestamos_mora,
        "pagos_mes": pagos_mes,
        "fecha_actualizacion": now_utc.isoformat(),
    }


# ---------- Reporte pendientes por cliente (PDF) ----------
def _generar_pdf_pendientes_cliente(cedula: str, nombre: str, cuotas: List[dict], total_pendiente: float) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Cuotas pendientes por cliente", styles["Title"]))
    story.append(Paragraph(f"Cédula: {cedula}", styles["Normal"]))
    story.append(Paragraph(f"Cliente: {nombre or '—'}", styles["Normal"]))
    story.append(Paragraph(f"Total pendiente: {total_pendiente:.2f}", styles["Normal"]))
    story.append(Spacer(1, 12))
    if not cuotas:
        story.append(Paragraph("No hay cuotas pendientes.", styles["Normal"]))
    else:
        rows = [["Préstamo", "Nº Cuota", "Vencimiento", "Monto", "Estado"]]
        for c in cuotas:
            rows.append([
                str(c.get("prestamo_id", "")),
                str(c.get("numero_cuota", "")),
                c.get("fecha_vencimiento", ""),
                str(c.get("monto", 0)),
                c.get("estado", ""),
            ])
        t = Table(rows)
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t)
    doc.build(story)
    return buf.getvalue()


@router.get("/cliente/{cedula}/pendientes.pdf")
def get_pendientes_cliente_pdf(cedula: str, db: Session = Depends(get_db)):
    """Genera PDF con cuotas pendientes del cliente por cédula."""
    from fastapi import HTTPException
    row = db.execute(select(Cliente).where(Cliente.cedula == cedula.strip())).first()
    if not row:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente = row[0]
    nombre = getattr(cliente, "nombres", None) or ""
    prestamos = db.execute(
        select(Prestamo).where(Prestamo.cliente_id == cliente.id)
    ).scalars().all()
    prestamo_ids = [p.id for p in prestamos]
    cuotas_list = []
    total_pendiente = 0.0
    if prestamo_ids:
        cuotas_rows = db.execute(
            select(Cuota)
            .where(Cuota.prestamo_id.in_(prestamo_ids), Cuota.fecha_pago.is_(None))
            .order_by(Cuota.prestamo_id, Cuota.numero_cuota)
        ).scalars().all()
        for c in cuotas_rows:
            m = float(c.monto) if c.monto is not None else 0
            total_pendiente += m
            cuotas_list.append({
                "prestamo_id": c.prestamo_id,
                "numero_cuota": c.numero_cuota,
                "fecha_vencimiento": c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else "",
                "monto": m,
                "estado": c.estado or "PENDIENTE",
            })
    content = _generar_pdf_pendientes_cliente(cedula, nombre, cuotas_list, total_pendiente)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=pendientes_{cedula}.pdf"},
    )


# ---------- PDF Tabla de amortización completa por cédula ----------
def _generar_pdf_amortizacion_cliente(cedula: str, nombre: str, cuotas: List[dict]) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Tabla de amortización", styles["Title"]))
    story.append(Paragraph(f"Cédula: {cedula}", styles["Normal"]))
    story.append(Paragraph(f"Cliente: {nombre or '—'}", styles["Normal"]))
    story.append(Spacer(1, 12))
    if not cuotas:
        story.append(Paragraph("No hay cuotas para este cliente.", styles["Normal"]))
    else:
        rows = [["Préstamo", "Nº Cuota", "Vencimiento", "Fecha pago", "Monto", "Estado"]]
        for c in cuotas:
            rows.append([
                str(c.get("prestamo_id", "")),
                str(c.get("numero_cuota", "")),
                c.get("fecha_vencimiento", ""),
                c.get("fecha_pago") or "—",
                str(c.get("monto", 0)),
                c.get("estado", ""),
            ])
        t = Table(rows)
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t)
    doc.build(story)
    return buf.getvalue()


@router.get("/cliente/{cedula}/amortizacion.pdf")
def get_amortizacion_cliente_pdf(cedula: str, db: Session = Depends(get_db)):
    """Genera PDF con tabla de amortización completa (todas las cuotas) del cliente por cédula."""
    from fastapi import HTTPException
    row = db.execute(select(Cliente).where(Cliente.cedula == cedula.strip())).first()
    if not row:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente = row[0]
    nombre = getattr(cliente, "nombres", None) or ""
    prestamos = db.execute(
        select(Prestamo).where(Prestamo.cliente_id == cliente.id)
    ).scalars().all()
    prestamo_ids = [p.id for p in prestamos]
    cuotas_list = []
    if prestamo_ids:
        cuotas_rows = db.execute(
            select(Cuota)
            .where(Cuota.prestamo_id.in_(prestamo_ids))
            .order_by(Cuota.prestamo_id, Cuota.numero_cuota)
        ).scalars().all()
        for c in cuotas_rows:
            m = float(c.monto) if c.monto is not None else 0
            fp = getattr(c, "fecha_pago", None)
            fecha_pago_str = fp.isoformat() if fp else "—"
            cuotas_list.append({
                "prestamo_id": c.prestamo_id,
                "numero_cuota": c.numero_cuota,
                "fecha_vencimiento": c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else "",
                "fecha_pago": fecha_pago_str,
                "monto": m,
                "estado": c.estado or "PENDIENTE",
            })
    content = _generar_pdf_amortizacion_cliente(cedula, nombre, cuotas_list)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=amortizacion_{cedula}.pdf"},
    )


# ---------- Reporte Cartera ----------
def _datos_cartera(db: Session, fecha_corte: date) -> dict:
    """Obtiene datos para reporte de cartera a una fecha de corte (solo clientes ACTIVOS)."""
    cuotas_pendientes = (
        select(func.coalesce(func.sum(Cuota.monto), 0))
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
        )
    )
    cartera_total = _safe_float(db.scalar(cuotas_pendientes) or 0)

    prestamos_activos = db.scalar(
        select(func.count())
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
    ) or 0

    subq_mora = (
        select(Cuota.prestamo_id)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < fecha_corte,
        )
        .distinct()
    )
    prestamos_mora = db.scalar(select(func.count()).select_from(subq_mora.subquery())) or 0

    mora_total = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento < fecha_corte,
            )
        )
        or 0
    )

    # Distribución por monto (rangos simples)
    rangos = [(0, 5000), (5001, 15000), (15001, 50000), (50001, 999999999)]
    distribucion_por_monto: List[dict] = []
    for low, high in rangos:
        subq = (
            select(Cuota.prestamo_id)
            .where(Cuota.fecha_pago.is_(None))
            .group_by(Cuota.prestamo_id)
            .having(and_(func.sum(Cuota.monto) >= low, func.sum(Cuota.monto) <= high))
        )
        # Contar préstamos con saldo en ese rango (aproximado por suma de cuotas)
        q = (
            select(func.count(func.distinct(Cuota.prestamo_id)), func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .where(Cuota.fecha_pago.is_(None))
        )
        # Simplificado: un solo rango "Total" para no complicar la query
        pass
    distribucion_por_monto = [
        {"rango": "0 - 5.000", "cantidad": 0, "monto": 0},
        {"rango": "5.001 - 15.000", "cantidad": 0, "monto": 0},
        {"rango": "15.001 - 50.000", "cantidad": 0, "monto": 0},
        {"rango": "> 50.000", "cantidad": 0, "monto": 0},
    ]
    saldos = db.execute(
        select(Cuota.prestamo_id, func.sum(Cuota.monto).label("saldo"))
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
        )
        .group_by(Cuota.prestamo_id)
    ).all()
    for pid, saldo in saldos:
        s = _safe_float(saldo)
        if s <= 5000:
            distribucion_por_monto[0]["cantidad"] += 1
            distribucion_por_monto[0]["monto"] += s
        elif s <= 15000:
            distribucion_por_monto[1]["cantidad"] += 1
            distribucion_por_monto[1]["monto"] += s
        elif s <= 50000:
            distribucion_por_monto[2]["cantidad"] += 1
            distribucion_por_monto[2]["monto"] += s
        else:
            distribucion_por_monto[3]["cantidad"] += 1
            distribucion_por_monto[3]["monto"] += s

    # Distribución por mora (días)
    distribucion_por_mora: List[dict] = []
    for label, dias_min, dias_max in [
        ("0-30 días", 0, 30),
        ("31-60 días", 31, 60),
        ("61-90 días", 61, 90),
        ("> 90 días", 91, 9999),
    ]:
        delta_min = fecha_corte - timedelta(days=dias_max)
        delta_max = fecha_corte - timedelta(days=dias_min)
        q = (
            select(func.count(Cuota.id), func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento <= delta_max,
                Cuota.fecha_vencimiento >= delta_min,
            )
        )
        row = db.execute(q).one_or_none()
        cnt = (row[0] or 0) if row else 0
        monto = _safe_float(row[1] or 0) if row else 0
        distribucion_por_mora.append({"rango": label, "cantidad": cnt, "monto_total": monto})

    return {
        "fecha_corte": fecha_corte.isoformat(),
        "cartera_total": cartera_total,
        "capital_pendiente": cartera_total,
        "intereses_pendientes": 0,
        "mora_total": mora_total,
        "cantidad_prestamos_activos": prestamos_activos,
        "cantidad_prestamos_mora": prestamos_mora,
        "distribucion_por_monto": distribucion_por_monto,
        "distribucion_por_mora": distribucion_por_mora,
    }


def _cartera_por_periodos(db: Session, periodos: List[tuple]) -> dict:
    """Genera datos cartera para lista de (año, mes)."""
    resultado: dict = {"meses": []}
    for (año, mes) in periodos:
        inicio = date(año, mes, 1)
        _, ultimo = calendar.monthrange(año, mes)
        fin = date(año, mes, ultimo)

        # Agrupar por día del mes (1-31): cuotas con fecha_vencimiento en ese día
        rows = db.execute(
            select(
                func.extract("day", Cuota.fecha_vencimiento).label("dia"),
                func.coalesce(func.sum(Cuota.monto), 0).label("monto_cobrar"),
                func.count(Cuota.id).label("cantidad_cuotas"),
            )
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento >= inicio,
                Cuota.fecha_vencimiento <= fin,
            )
            .group_by(func.extract("day", Cuota.fecha_vencimiento))
            .order_by(func.extract("day", Cuota.fecha_vencimiento))
        ).fetchall()

        # Mapa día -> {cantidad_cuotas, monto_cobrar}
        por_dia: dict = {}
        for r in rows:
            d = int(r.dia) if r.dia is not None else 0
            por_dia[d] = {
                "cantidad_cuotas": r.cantidad_cuotas or 0,
                "monto_cobrar": round(_safe_float(r.monto_cobrar), 2),
            }

        # Una fila por cada día del mes (1 a ultimo)
        items: List[dict] = []
        for d in range(1, ultimo + 1):
            data = por_dia.get(d, {"cantidad_cuotas": 0, "monto_cobrar": 0})
            items.append({
                "dia": d,
                "cantidad_cuotas": data["cantidad_cuotas"],
                "monto_cobrar": data["monto_cobrar"],
            })

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "items": items,
        })

    return resultado


@router.get("/cartera/por-mes")
def get_cartera_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    años: Optional[str] = Query(None, description="Años separados por coma, ej: 2023,2024"),
    meses_list: Optional[str] = Query(None, description="Meses 1-12 separados por coma, ej: 1,2,3"),
):
    """Cuentas por cobrar: una pestaña por mes (MM/YYYY). Por día del mes: cuotas por cobrar ese día."""
    periodos = _periodos_desde_filtros(años, meses_list, meses)
    return _cartera_por_periodos(db, periodos)


@router.get("/cartera")
def get_reporte_cartera(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None, description="Fecha de corte YYYY-MM-DD"),
):
    """Reporte de cartera en JSON. Datos reales desde BD."""
    fc = _parse_fecha(fecha_corte)
    return _datos_cartera(db, fc)


def _generar_excel_cartera_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (MM/YYYY). Columnas: Día, Cuotas por cobrar, Monto ($)."""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])

    # Diccionario de meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        año = mes_data.get("año", datetime.now().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{año}")
        sheet_name = f"{mes_nombre} {año}"[:31]
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Reporte de Cartera", label])
        ws.append(["Cuotas por cobrar por día del mes (dato actualizado al aprobar/eliminar préstamos)"])
        ws.append([])
        ws.append(["Día", "Cuotas por cobrar", "Monto ($)"])
        
        for item in mes_data.get("items", []):
            row_num = ws.max_row + 1
            ws.append([
                item.get("dia", 0),
                item.get("cantidad_cuotas", 0),
                item.get("monto_cobrar", 0)
            ])
            # Aplicar formato dólares con 2 decimales a la columna C (Monto)
            cell = ws.cell(row=row_num, column=3)
            cell.number_format = '$#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_cartera(data: dict) -> bytes:
    """Excel clásico de cartera (resumen + distribuciones)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera"
    ws.append(["Reporte de Cartera", data.get("fecha_corte", "")])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    ws.append(["Cartera total", data.get("cartera_total", 0)])
    ws.append(["Capital pendiente", data.get("capital_pendiente", 0)])
    ws.append(["Mora total", data.get("mora_total", 0)])
    ws.append(["Préstamos activos", data.get("cantidad_prestamos_activos", 0)])
    ws.append(["Préstamos en mora", data.get("cantidad_prestamos_mora", 0)])
    ws.append([])
    ws.append(["Distribución por monto"])
    ws.append(["Rango", "Cantidad", "Monto"])
    for r in data.get("distribucion_por_monto", []):
        ws.append([r.get("rango", ""), r.get("cantidad", 0), r.get("monto", 0)])
    ws2 = wb.create_sheet("Mora")
    ws2.append(["Rango", "Cantidad", "Monto total"])
    for r in data.get("distribucion_por_mora", []):
        ws2.append([r.get("rango", ""), r.get("cantidad", 0), r.get("monto_total", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_cartera(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Cartera", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Resumen", styles["Heading2"]))
    resumen = [
        ["Cartera total", str(data.get("cartera_total", 0))],
        ["Capital pendiente", str(data.get("capital_pendiente", 0))],
        ["Mora total", str(data.get("mora_total", 0))],
        ["Préstamos activos", str(data.get("cantidad_prestamos_activos", 0))],
        ["Préstamos en mora", str(data.get("cantidad_prestamos_mora", 0))],
    ]
    t = Table(resumen)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    story.append(Spacer(1, 12))
    story.append(Paragraph("Distribución por monto", styles["Heading2"]))
    rows = [["Rango", "Cantidad", "Monto"]]
    for r in data.get("distribucion_por_monto", []):
        rows.append([r.get("rango", ""), str(r.get("cantidad", 0)), str(r.get("monto", 0))])
    t2 = Table(rows)
    t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t2)
    doc.build(story)
    return buf.getvalue()


@router.get("/exportar/cartera")
def exportar_cartera(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses (una pestaña por mes)"),
    años: Optional[str] = Query(None, description="Años separados por coma, ej: 2023,2024"),
    meses_list: Optional[str] = Query(None, description="Meses 1-12 separados por coma, ej: 1,2,3"),
):
    """Exporta reporte de cartera. Excel: una pestaña por mes (MM/YYYY), columnas Día | Cuotas por cobrar | Monto.
    PDF: resumen clásico. Datos desde cuotas (actualizados al aprobar/eliminar préstamos)."""
    fc = _parse_fecha(fecha_corte)
    if formato == "excel":
        data_por_mes = get_cartera_por_mes(db=db, meses=meses, años=años, meses_list=meses_list)
        content = _generar_excel_cartera_por_mes(data_por_mes)
        hoy_str = date.today().isoformat()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_cartera_{hoy_str}.xlsx"},
        )
    data = _datos_cartera(db, fc)
    content = _generar_pdf_cartera(data)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_cartera_{fc.isoformat()}.pdf"},
    )


# ---------- Reporte Pagos ----------
@router.get("/pagos")
def get_reporte_pagos(
    db: Session = Depends(get_db),
    fecha_inicio: str = Query(..., description="YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="YYYY-MM-DD"),
):
    """Reporte de pagos en un rango de fechas. Datos reales desde BD."""
    fi = _parse_fecha(fecha_inicio)
    ff = _parse_fecha(fecha_fin)
    if fi > ff:
        fi, ff = ff, fi
    # Solo clientes ACTIVOS
    total_pagos = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.isnot(None),
                func.date(Cuota.fecha_pago) >= fi,
                func.date(Cuota.fecha_pago) <= ff,
            )
        )
        or 0
    )
    cantidad_pagos = db.scalar(
        select(func.count())
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.isnot(None),
            func.date(Cuota.fecha_pago) >= fi,
            func.date(Cuota.fecha_pago) <= ff,
        )
    ) or 0
    pagos_por_dia = (
        db.execute(
            select(func.date(Cuota.fecha_pago).label("fecha"), func.count().label("cantidad"), func.sum(Cuota.monto).label("monto"))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.isnot(None),
                func.date(Cuota.fecha_pago) >= fi,
                func.date(Cuota.fecha_pago) <= ff,
            )
            .group_by(func.date(Cuota.fecha_pago))
            .order_by(func.date(Cuota.fecha_pago))
        )
    ).fetchall()
    return {
        "fecha_inicio": fi.isoformat(),
        "fecha_fin": ff.isoformat(),
        "total_pagos": total_pagos,
        "cantidad_pagos": cantidad_pagos,
        "pagos_por_metodo": [],
        "pagos_por_dia": [
            {"fecha": str(r.fecha), "cantidad": r.cantidad, "monto": _safe_float(r.monto)}
            for r in pagos_por_dia
        ],
    }


def _pagos_por_dia_periodos(db: Session, periodos: List[tuple]) -> dict:
    """Genera datos pagos por día para lista de (año, mes)."""
    resultado: dict = {"meses": []}
    for (año, mes) in periodos:
        inicio = date(año, mes, 1)
        _, ultimo = calendar.monthrange(año, mes)
        fin = date(año, mes, ultimo)

        rows = db.execute(
            select(
                func.extract("day", Pago.fecha_pago).label("dia"),
                func.count(Pago.id).label("cantidad_pagos"),
                func.count(func.distinct(Pago.cedula_cliente)).label("cantidad_cedulas"),
                func.coalesce(func.sum(Pago.monto_pagado), 0).label("monto_total"),
            )
            .select_from(Pago)
            .where(
                func.date(Pago.fecha_pago) >= inicio,
                func.date(Pago.fecha_pago) <= fin,
            )
            .group_by(func.extract("day", Pago.fecha_pago))
            .order_by(func.extract("day", Pago.fecha_pago))
        ).fetchall()

        por_dia: dict = {}
        for r in rows:
            d = int(r.dia) if r.dia is not None else 0
            por_dia[d] = {
                "cantidad_pagos": r.cantidad_pagos or 0,
                "cantidad_cedulas": r.cantidad_cedulas or 0,
                "monto_total": round(_safe_float(r.monto_total), 2),
            }

        items: List[dict] = []
        for d in range(1, ultimo + 1):
            data = por_dia.get(d, {"cantidad_pagos": 0, "cantidad_cedulas": 0, "monto_total": 0})
            items.append({
                "dia": d,
                "cantidad_pagos": data["cantidad_pagos"],
                "cantidad_cedulas": data["cantidad_cedulas"],
                "monto_total": data["monto_total"],
            })

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "items": items,
        })

    return resultado


@router.get("/pagos/por-dia-mes")
def get_pagos_por_dia_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    años: Optional[str] = Query(None, description="Años separados por coma"),
    meses_list: Optional[str] = Query(None, description="Meses 1-12 separados por coma"),
):
    """Pagos por día del mes: una pestaña por mes (MM/YYYY). Por día: cantidad pagos, cantidad cédulas, monto ($)."""
    periodos = _periodos_desde_filtros(años, meses_list, meses)
    return _pagos_por_dia_periodos(db, periodos)


@router.get("/pagos/por-mes")
def get_pagos_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
):
    """Pagos agrupados por mes/año. Cada mes tiene lista de pagos: Fecha, id préstamo, cédula, nombre, monto, documento. Orden descendente por fecha."""
    hoy = date.today()
    resultado: dict = {"meses": []}

    for i in range(meses):
        año = hoy.year
        mes = hoy.month - i
        while mes <= 0:
            mes += 12
            año -= 1
        inicio = date(año, mes, 1)
        _, ultimo = calendar.monthrange(año, mes)
        fin = date(año, mes, ultimo)

        q = (
            select(
                Pago.id,
                Pago.fecha_pago,
                Pago.prestamo_id,
                Pago.cedula_cliente,
                Pago.monto_pagado,
                Pago.numero_documento,
                func.coalesce(Prestamo.nombres, Cliente.nombres).label("nombres"),
            )
            .select_from(Pago)
            .outerjoin(Prestamo, Pago.prestamo_id == Prestamo.id)
            .outerjoin(Cliente, Pago.cedula_cliente == Cliente.cedula)
            .where(
                func.date(Pago.fecha_pago) >= inicio,
                func.date(Pago.fecha_pago) <= fin,
            )
            .order_by(Pago.fecha_pago.desc())
        )
        rows = db.execute(q).fetchall()

        items = []
        for r in rows:
            fp = r.fecha_pago
            fecha_str = fp.date().isoformat() if hasattr(fp, "date") else (fp.isoformat()[:10] if fp else None)
            items.append({
                "pago_id": r.id,
                "fecha": fecha_str,
                "prestamo_id": r.prestamo_id,
                "cedula": r.cedula_cliente or "",
                "nombre": (r.nombres or "").strip(),
                "monto_pago": _safe_float(r.monto_pagado),
                "documento": r.numero_documento or "",
            })

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "items": items,
        })

    return resultado


def _generar_excel_pagos_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (MM/YYYY). Columnas: Día | Cantidad pagos | Cantidad cédulas | Monto ($)."""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])

    # Diccionario de meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        año = mes_data.get("año", datetime.now().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{año}")
        sheet_name = f"{mes_nombre} {año}"[:31]
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Reporte de Pagos", label])
        ws.append([])
        
        # Calcular totales
        items = mes_data.get("items", [])
        total_pagos = sum(item.get("cantidad_pagos", 0) for item in items)
        total_cedulas = sum(item.get("cantidad_cedulas", 0) for item in items)
        total_monto = sum(item.get("monto_total", 0) for item in items)
        
        # Agregar fila de totales
        ws.append([
            f"Total pagos: {total_pagos} | Total cédulas: {total_cedulas} | Total monto: ${total_monto:.2f}"
        ])
        ws.append([])
        
        ws.append(["Día", "Cantidad de pagos", "Cantidad de cédulas", "Monto ($)"])
        for item in items:
            ws.append([
                item.get("dia", 0),
                item.get("cantidad_pagos", 0),
                item.get("cantidad_cedulas", 0),
                item.get("monto_total", 0),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_pagos(data: dict) -> bytes:
    """Excel clásico de pagos por rango de fechas."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.append(["Reporte de Pagos", data.get("fecha_inicio", ""), "a", data.get("fecha_fin", "")])
    ws.append([])
    ws.append(["Total pagos", data.get("total_pagos", 0)])
    ws.append(["Cantidad de pagos", data.get("cantidad_pagos", 0)])
    ws.append([])
    ws.append(["Fecha", "Cantidad", "Monto"])
    for r in data.get("pagos_por_dia", []):
        ws.append([r.get("fecha", ""), r.get("cantidad", 0), r.get("monto", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_pagos(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Pagos", styles["Title"]))
    story.append(Paragraph(f"Período: {data.get('fecha_inicio', '')} a {data.get('fecha_fin', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    resumen = [
        ["Total pagos (monto)", str(data.get("total_pagos", 0))],
        ["Cantidad de pagos", str(data.get("cantidad_pagos", 0))],
    ]
    t = Table(resumen)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    pagos_por_dia = data.get("pagos_por_dia", [])
    if pagos_por_dia:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Pagos por día", styles["Heading2"]))
        rows = [["Fecha", "Cantidad", "Monto"]]
        for r in pagos_por_dia:
            rows.append([r.get("fecha", ""), str(r.get("cantidad", 0)), str(r.get("monto", 0))])
        t2 = Table(rows)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


@router.get("/exportar/pagos")
def exportar_pagos(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses (una pestaña por mes)"),
    años: Optional[str] = Query(None, description="Años separados por coma"),
    meses_list: Optional[str] = Query(None, description="Meses 1-12 separados por coma"),
):
    """Exporta reporte de pagos. Excel: una pestaña por mes (MM/YYYY), columnas Día | Cantidad pagos | Cantidad cédulas | Monto ($).
    PDF: requiere fecha_inicio y fecha_fin."""
    if formato == "excel":
        data_por_mes = get_pagos_por_dia_mes(db=db, meses=meses, años=años, meses_list=meses_list)
        content = _generar_excel_pagos_por_mes(data_por_mes)
        hoy_str = date.today().isoformat()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_pagos_{hoy_str}.xlsx"},
        )
    fi = _parse_fecha(fecha_inicio or date.today().isoformat())
    ff = _parse_fecha(fecha_fin or date.today().isoformat())
    if fi > ff:
        fi, ff = ff, fi
    data = get_reporte_pagos(db=db, fecha_inicio=fi.isoformat(), fecha_fin=ff.isoformat())
    content = _generar_pdf_pagos(data)
    return Response(content=content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_pagos_{fi.isoformat()}_{ff.isoformat()}.pdf"})


# ---------- Reporte Morosidad ----------
@router.get("/morosidad")
def get_reporte_morosidad(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Reporte de pago vencido. Datos reales desde BD (solo clientes ACTIVOS).
    Concepto: Pago vencido = fecha_vencimiento < fc y fecha_pago IS NULL. Moroso = 61+ días de atraso."""
    fc = _parse_fecha(fecha_corte)
    subq_mora = (
        select(Cuota.prestamo_id)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < fc,
        )
        .distinct()
    )
    total_prestamos_mora = db.scalar(select(func.count()).select_from(subq_mora.subquery())) or 0
    prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]
    total_clientes_mora = 0
    if prestamos_ids:
        total_clientes_mora = db.scalar(
            select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(prestamos_ids))
        ) or 0
    monto_total_mora = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento < fc,
            )
        )
        or 0
    )
    cuotas_mora = db.execute(
        select(Cuota.prestamo_id, Cuota.fecha_vencimiento)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < fc,
        )
    ).fetchall()
    dias_list = [(fc - r.fecha_vencimiento).days for r in cuotas_mora]
    promedio_dias_mora = sum(dias_list) / len(dias_list) if dias_list else 0
    distribucion_por_rango: List[dict] = []
    morosidad_por_analista: List[dict] = []
    analistas = db.execute(select(Prestamo.analista).where(Prestamo.id.in_(prestamos_ids)).distinct()).fetchall()
    for (analista,) in analistas:
        if not analista:
            continue
        ids_ana = db.execute(select(Prestamo.id).where(Prestamo.analista == analista, Prestamo.id.in_(prestamos_ids))).fetchall()
        ids_ana = [x[0] for x in ids_ana]
        monto_ana = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc, Cuota.prestamo_id.in_(ids_ana)))) or 0
        cuotas_ana = db.execute(
            select(Cuota.fecha_vencimiento).select_from(Cuota).where(
                Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc, Cuota.prestamo_id.in_(ids_ana)
            )
        ).fetchall()
        dias_ana = [(fc - r.fecha_vencimiento).days for r in cuotas_ana]
        prom_dias_ana = sum(dias_ana) / len(dias_ana) if dias_ana else 0
        morosidad_por_analista.append({
            "analista": analista,
            "cantidad_prestamos": len(ids_ana),
            "cantidad_clientes": db.scalar(select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(ids_ana))) or 0,
            "monto_total_mora": monto_ana,
            "promedio_dias_mora": prom_dias_ana,
        })
    detalle: List[dict] = []
    for pid in prestamos_ids[:200]:
        p = db.get(Prestamo, pid)
        if not p:
            continue
        monto_mora = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)) or 0)
        cuotas_en_mora = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)) or 0
        primera = db.execute(select(func.min(Cuota.fecha_vencimiento)).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)).scalar()
        max_dias = (fc - primera).days if primera else 0
        detalle.append({
            "prestamo_id": pid,
            "cedula": p.cedula or "",
            "nombres": p.nombres or "",
            "total_financiamiento": _safe_float(p.total_financiamiento),
            "analista": p.analista or "",
            "concesionario": p.concesionario or "",
            "cuotas_en_mora": cuotas_en_mora,
            "monto_total_mora": monto_mora,
            "max_dias_mora": max_dias,
            "primera_cuota_vencida": primera.isoformat() if primera else None,
        })
    return {
        "fecha_corte": fc.isoformat(),
        "total_prestamos_mora": total_prestamos_mora,
        "total_clientes_mora": total_clientes_mora,
        "monto_total_mora": monto_total_mora,
        "promedio_dias_mora": round(promedio_dias_mora, 2),
        "distribucion_por_rango": distribucion_por_rango,
        "morosidad_por_analista": morosidad_por_analista,
        "detalle_prestamos": detalle,
    }


@router.get("/morosidad/por-mes")
def get_morosidad_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    años: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Morosidad por mes: una pestaña por mes (mes/año). Cuotas vencidas sin pagar a fin de cada mes, agrupadas por cédula."""
    resultado: dict = {"meses": []}
    periodos = _periodos_desde_filtros(años, meses_list, meses)

    for (año, mes) in periodos:
        _, ultimo = calendar.monthrange(año, mes)
        fc = date(año, mes, ultimo)

        subq_mora = (
            select(Cuota.prestamo_id)
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento < fc,
            )
            .distinct()
        )
        prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]
        total_prestamos_mora = len(prestamos_ids)
        total_clientes_mora = 0
        monto_total_mora = 0.0
        if prestamos_ids:
            total_clientes_mora = db.scalar(
                select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(prestamos_ids))
            ) or 0
            monto_total_mora = _safe_float(
                db.scalar(
                    select(func.coalesce(func.sum(Cuota.monto), 0))
                    .select_from(Cuota)
                    .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
                    .join(Cliente, Prestamo.cliente_id == Cliente.id)
                    .where(
                        Cliente.estado == "ACTIVO",
                        Prestamo.estado == "APROBADO",
                        Cuota.fecha_pago.is_(None),
                        Cuota.fecha_vencimiento < fc,
                    )
                )
            ) or 0

        cuotas_mora = db.execute(
            select(Cuota.prestamo_id, Cuota.fecha_vencimiento)
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento < fc,
            )
        ).fetchall()
        dias_list = [(fc - r.fecha_vencimiento).days for r in cuotas_mora]
        promedio_dias_mora = sum(dias_list) / len(dias_list) if dias_list else 0

        morosidad_por_cedula: List[dict] = []
        # Una sola consulta agrupada por cédula (evita N+1 con miles de cédulas)
        if prestamos_ids:
            from sqlalchemy import cast, Date
            from sqlalchemy.sql import literal_column
            # PostgreSQL: date - date da días como entero
            fc_date = literal_column(f"'{fc}'::date")
            dias_expr = fc_date - cast(Cuota.fecha_vencimiento, Date)
            rows_cedula = db.execute(
                select(
                    Prestamo.cedula,
                    Prestamo.nombres,
                    func.count(func.distinct(Cuota.prestamo_id)).label("cantidad_prestamos"),
                    func.coalesce(func.sum(Cuota.monto), 0).label("monto_total_mora"),
                    func.avg(dias_expr).label("promedio_dias_mora"),
                )
                .select_from(Cuota)
                .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
                .where(
                    Prestamo.id.in_(prestamos_ids),
                    Prestamo.cedula.isnot(None),
                    Prestamo.cedula != "",
                    Cuota.fecha_pago.is_(None),
                    Cuota.fecha_vencimiento < fc,
                )
                .group_by(Prestamo.cedula, Prestamo.nombres)
                .order_by(func.sum(Cuota.monto).desc())
            ).fetchall()
            for r in rows_cedula:
                morosidad_por_cedula.append({
                    "cedula": r.cedula or "",
                    "nombres": r.nombres or "",
                    "cantidad_prestamos": r.cantidad_prestamos or 0,
                    "cantidad_clientes": 1,
                    "monto_total_mora": round(_safe_float(r.monto_total_mora), 2),
                    "promedio_dias_mora": round(_safe_float(r.promedio_dias_mora), 1),
                })

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "fecha_corte": fc.isoformat(),
            "total_prestamos_mora": total_prestamos_mora,
            "total_clientes_mora": total_clientes_mora,
            "monto_total_mora": monto_total_mora,
            "promedio_dias_mora": round(promedio_dias_mora, 1),
            "morosidad_por_cedula": morosidad_por_cedula,
        })

    return resultado


# Rangos de días atrasados para informe pago vencido en pestañas
RANGOS_ATRASO = [
    {"key": "1_dia", "label": "1 día atrasado", "min_dias": 1, "max_dias": 14},
    {"key": "15_dias", "label": "15 días atrasado", "min_dias": 15, "max_dias": 29},
    {"key": "30_dias", "label": "30 días atrasado", "min_dias": 30, "max_dias": 59},
    {"key": "2_meses", "label": "2 meses atrasado", "min_dias": 60, "max_dias": 60},
    {"key": "61_dias", "label": "Más de 61 días", "min_dias": 61, "max_dias": 99999},
]


@router.get("/morosidad/por-rangos")
def get_morosidad_por_rangos(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Informe pago vencido por rangos de días. Datos para viñetas: cédula, nombres, total financiamiento,
    pagos totales, saldo, último pago (fecha), próximo pago (fecha). Una pestaña por rango."""
    fc = _parse_fecha(fecha_corte)
    subq_mora = (
        select(Cuota.prestamo_id)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < fc,
        )
        .distinct()
    )
    prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]

    resultado: dict = {"fecha_corte": fc.isoformat(), "rangos": {}}


    for rango in RANGOS_ATRASO:
        min_d, max_d = rango["min_dias"], rango["max_dias"]
        items: List[dict] = []

        for pid in prestamos_ids:
            p = db.get(Prestamo, pid)
            if not p:
                continue

            primera = db.execute(
                select(func.min(Cuota.fecha_vencimiento))
                .select_from(Cuota)
                .where(
                    Cuota.prestamo_id == pid,
                    Cuota.fecha_pago.is_(None),
                    Cuota.fecha_vencimiento < fc,
                )
            ).scalar()
            if not primera:
                continue
            dias_atraso = (fc - primera).days
            if dias_atraso < min_d or dias_atraso > max_d:
                continue

            saldo = _safe_float(
                db.scalar(
                    select(func.coalesce(func.sum(Cuota.monto), 0))
                    .select_from(Cuota)
                    .where(
                        Cuota.prestamo_id == pid,
                        Cuota.fecha_pago.is_(None),
                    )
                )
            ) or 0

            pagos_totales = _safe_float(
                db.scalar(
                    select(func.coalesce(func.sum(Cuota.total_pagado), 0))
                    .select_from(Cuota)
                    .where(Cuota.prestamo_id == pid)
                )
            ) or 0

            ultimo_pago = db.execute(
                select(func.max(Cuota.fecha_pago))
                .select_from(Cuota)
                .where(Cuota.prestamo_id == pid, Cuota.fecha_pago.isnot(None))
            ).scalar()

            proximo_pago = db.execute(
                select(func.min(Cuota.fecha_vencimiento))
                .select_from(Cuota)
                .where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None))
            ).scalar()

            def _fecha_iso(d):
                if d is None:
                    return None
                if hasattr(d, "date"):
                    d = d.date() if callable(getattr(d, "date", None)) else d
                return d.isoformat()[:10] if hasattr(d, "isoformat") else str(d)[:10]

            items.append({
                "prestamo_id": pid,
                "cedula": p.cedula or "",
                "nombres": p.nombres or "",
                "total_financiamiento": _safe_float(p.total_financiamiento),
                "pagos_totales": pagos_totales,
                "saldo": saldo,
                "ultimo_pago_fecha": _fecha_iso(ultimo_pago),
                "proximo_pago_fecha": _fecha_iso(proximo_pago),
                "dias_atraso": dias_atraso,
            })

        resultado["rangos"][rango["key"]] = {
            "label": rango["label"],
            "items": items,
        }

    return resultado


def _generar_excel_morosidad_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (mes/año). Agrupado por cédula."""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])

    # Diccionario de meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        año = mes_data.get("año", datetime.now().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{año}")
        sheet_name = f"{mes_nombre} {año}"[:31]
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Informe de Vencimiento de Pagos", label])
        ws.append([])
        ws.append(["Total préstamos con pago vencido", mes_data.get("total_prestamos_mora", 0)])
        ws.append(["Total clientes con pago vencido", mes_data.get("total_clientes_mora", 0)])
        ws.append(["Monto total en dólares", mes_data.get("monto_total_mora", 0)])
        ws.append(["Promedio días de atraso", round(_safe_float(mes_data.get("promedio_dias_mora", 0)), 1)])
        ws.append([])
        ws.append(["Pago vencido por cédula"])
        ws.append(["Cédula", "Nombre", "Cant. préstamos", "Cant. clientes", "Monto en dólares", "Prom. días"])
        
        for r in mes_data.get("morosidad_por_cedula", []):
            prom_dias = round(_safe_float(r.get("promedio_dias_mora", 0)), 1)
            ws.append([
                r.get("cedula", ""),
                r.get("nombres", ""),
                r.get("cantidad_prestamos", 0),
                r.get("cantidad_clientes", 0),
                r.get("monto_total_mora", 0),
                prom_dias
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_morosidad(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Informe de Vencimiento de Pagos", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    resumen = [
        ["Total préstamos con pago vencido", str(data.get("total_prestamos_mora", 0))],
        ["Total clientes con pago vencido", str(data.get("total_clientes_mora", 0))],
        ["Monto total en dólares", str(data.get("monto_total_mora", 0))],
        ["Promedio días de atraso", str(round(_safe_float(data.get("promedio_dias_mora", 0)), 1))],
    ]
    t = Table(resumen)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    mora_analista = data.get("morosidad_por_analista", [])
    if mora_analista:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Pago vencido por analista", styles["Heading2"]))
        rows = [["Analista", "Préstamos", "Clientes", "Monto en dólares", "Prom. días"]]
        for r in mora_analista:
            prom_dias = round(_safe_float(r.get("promedio_dias_mora", 0)), 1)
            rows.append([r.get("analista", ""), str(r.get("cantidad_prestamos", 0)), str(r.get("cantidad_clientes", 0)), str(r.get("monto_total_mora", 0)), str(prom_dias)])
        t2 = Table(rows)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


@router.get("/exportar/morosidad")
def exportar_morosidad(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses (una pestaña por mes)"),
    años: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Exporta reporte de morosidad. Excel: una pestaña por mes (mes/año). PDF: fecha de corte única."""
    if formato == "pdf":
        data = get_reporte_morosidad(db=db, fecha_corte=fecha_corte)
        content = _generar_pdf_morosidad(data)
        return Response(content=content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=informe_vencimiento_pagos_{data['fecha_corte']}.pdf"})
    data_por_mes = get_morosidad_por_mes(db=db, meses=meses, años=años, meses_list=meses_list)
    content = _generar_excel_morosidad_por_mes(data_por_mes)
    hoy_str = date.today().isoformat()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=informe_vencimiento_pagos_{hoy_str}.xlsx"},
    )


# ---------- Reporte Financiero ----------
@router.get("/financiero")
def get_reporte_financiero(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Reporte financiero. Datos reales desde BD (solo clientes ACTIVOS)."""
    fc = _parse_fecha(fecha_corte)
    conds_activo = [
        Cuota.prestamo_id == Prestamo.id,
        Prestamo.cliente_id == Cliente.id,
        Cliente.estado == "ACTIVO",
        Prestamo.estado == "APROBADO",
    ]
    total_ingresos = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(and_(*conds_activo, Cuota.fecha_pago.isnot(None)))
        )
        or 0
    )
    cantidad_pagos = db.scalar(
        select(func.count())
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(and_(*conds_activo, Cuota.fecha_pago.isnot(None)))
    ) or 0
    cartera_total = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(and_(*conds_activo, Cuota.fecha_pago.is_(None)))
        )
        or 0
    )
    morosidad_total = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(and_(*conds_activo, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc))
        )
        or 0
    )
    saldo_pendiente = cartera_total
    total_desembolsado = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Prestamo.total_financiamiento), 0))
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        )
        or 0
    )
    porcentaje_cobrado = (total_ingresos / total_desembolsado * 100) if total_desembolsado else 0

    # Ingresos por mes: cuotas pagadas agrupadas por mes de fecha_pago
    q_ingresos = (
        select(
            func.to_char(func.date_trunc("month", Cuota.fecha_pago), "YYYY-MM").label("mes"),
            func.count().label("cantidad_pagos"),
            func.coalesce(func.sum(Cuota.monto), 0).label("monto_total"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(and_(*conds_activo, Cuota.fecha_pago.isnot(None)))
        .group_by(func.date_trunc("month", Cuota.fecha_pago))
        .order_by(func.date_trunc("month", Cuota.fecha_pago))
    )
    rows_ing = db.execute(q_ingresos).all()
    ingresos_por_mes = [
        {"mes": r.mes, "cantidad_pagos": r.cantidad_pagos, "monto_total": round(_safe_float(r.monto_total), 2)}
        for r in rows_ing
    ]

    # Egresos programados: cuotas pendientes de pago agrupadas por mes de vencimiento
    q_egresos = (
        select(
            func.to_char(func.date_trunc("month", Cuota.fecha_vencimiento), "YYYY-MM").label("mes"),
            func.count().label("cantidad_cuotas"),
            func.coalesce(func.sum(Cuota.monto), 0).label("monto_programado"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(and_(*conds_activo, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento.isnot(None)))
        .group_by(func.date_trunc("month", Cuota.fecha_vencimiento))
        .order_by(func.date_trunc("month", Cuota.fecha_vencimiento))
    )
    rows_egr = db.execute(q_egresos).all()
    egresos_programados = [
        {"mes": r.mes, "cantidad_cuotas": r.cantidad_cuotas, "monto_programado": round(_safe_float(r.monto_programado), 2)}
        for r in rows_egr
    ]

    # Flujo de caja: ingresos - egresos por mes (meses que aparecen en ingresos o egresos)
    meses_set = {x["mes"] for x in ingresos_por_mes} | {x["mes"] for x in egresos_programados}
    ing_map = {x["mes"]: x["monto_total"] for x in ingresos_por_mes}
    egr_map = {x["mes"]: x["monto_programado"] for x in egresos_programados}
    flujo_caja = []
    for m in sorted(meses_set):
        ing = ing_map.get(m, 0)
        egr = egr_map.get(m, 0)
        flujo_caja.append({
            "mes": m,
            "ingresos": ing,
            "egresos_programados": egr,
            "flujo_neto": round(ing - egr, 2),
        })

    return {
        "fecha_corte": fc.isoformat(),
        "total_ingresos": total_ingresos,
        "cantidad_pagos": cantidad_pagos,
        "cartera_total": cartera_total,
        "cartera_pendiente": cartera_total,
        "morosidad_total": morosidad_total,
        "saldo_pendiente": saldo_pendiente,
        "porcentaje_cobrado": round(porcentaje_cobrado, 2),
        "ingresos_por_mes": ingresos_por_mes,
        "egresos_programados": egresos_programados,
        "flujo_caja": flujo_caja,
    }


@router.get("/exportar/financiero")
def exportar_financiero(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
):
    """Exporta reporte financiero en Excel o PDF."""
    data = get_reporte_financiero(db=db, fecha_corte=fecha_corte)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financiero"
    ws.append(["Reporte Financiero", data.get("fecha_corte", "")])
    ws.append(["Total ingresos", data.get("total_ingresos", 0)])
    ws.append(["Cantidad pagos", data.get("cantidad_pagos", 0)])
    ws.append(["Cartera total", data.get("cartera_total", 0)])
    ws.append(["Morosidad total", data.get("morosidad_total", 0)])
    ws.append(["Porcentaje cobrado", data.get("porcentaje_cobrado", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    if formato == "pdf":
        pdf_content = _generar_pdf_financiero(data)
        return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_financiero_{data['fecha_corte']}.pdf"})
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_financiero_{data['fecha_corte']}.xlsx"})


def _generar_pdf_financiero(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte Financiero", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    resumen = [
        ["Total ingresos", str(data.get("total_ingresos", 0))],
        ["Cantidad pagos", str(data.get("cantidad_pagos", 0))],
        ["Cartera total", str(data.get("cartera_total", 0))],
        ["Morosidad total", str(data.get("morosidad_total", 0))],
        ["Porcentaje cobrado (%)", str(data.get("porcentaje_cobrado", 0))],
    ]
    t = Table(resumen)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    flujo = data.get("flujo_caja", [])
    if flujo:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Flujo de caja", styles["Heading2"]))
        rows = [["Mes", "Ingresos", "Egresos programados", "Flujo neto"]]
        for r in flujo:
            rows.append([r.get("mes", ""), str(r.get("ingresos", 0)), str(r.get("egresos_programados", 0)), str(r.get("flujo_neto", 0))])
        t2 = Table(rows)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


# ---------- Reporte Asesores ----------
@router.get("/asesores/por-mes")
def get_asesores_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    años: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Asesores por mes: una pestaña por mes. Solo lo que sucede en ese mes: cuotas con fecha_vencimiento en el mes y sin pagar."""
    resultado: dict = {"meses": []}
    periodos = _periodos_desde_filtros(años, meses_list, meses)

    for (año, mes) in periodos:
        inicio = date(año, mes, 1)
        _, ultimo = calendar.monthrange(año, mes)
        fin = date(año, mes, ultimo)

        # Solo cuotas que vencieron en este mes (fecha_vencimiento en el mes) y no están pagadas
        rows = db.execute(
            select(
                Prestamo.analista,
                func.coalesce(func.sum(Cuota.monto), 0).label("vencimiento_total"),
                func.count(func.distinct(Cuota.prestamo_id)).label("total_prestamos"),
            )
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento >= inicio,
                Cuota.fecha_vencimiento <= fin,
            )
            .group_by(Prestamo.analista)
            .order_by(func.sum(Cuota.monto).desc())
        ).fetchall()

        items = [
            {
                "analista": r.analista or "Sin asignar",
                "vencimiento_total": round(_safe_float(r.vencimiento_total), 2),
                "total_prestamos": r.total_prestamos or 0,
            }
            for r in rows
        ]

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "items": items,
        })

    return resultado


@router.get("/asesores")
def get_reporte_asesores(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Reporte por asesor/analista. Datos reales desde BD (solo clientes ACTIVOS)."""
    fc = _parse_fecha(fecha_corte)
    analistas = db.execute(
        select(Prestamo.analista)
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        .distinct()
    ).fetchall()
    resumen_por_analista: List[dict] = []
    for (analista,) in analistas:
        if not analista:
            continue
        prestamos = db.execute(
            select(Prestamo.id)
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.analista == analista,
                Prestamo.estado == "APROBADO",
            )
        ).fetchall()
        ids = [x[0] for x in prestamos]
        total_prestamos = len(ids)
        total_clientes = db.scalar(select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(ids))) or 0
        cartera_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.prestamo_id.in_(ids)))) or 0
        morosidad_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc, Cuota.prestamo_id.in_(ids)))) or 0
        total_cobrado = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.isnot(None), Cuota.prestamo_id.in_(ids)))) or 0
        porcentaje_cobrado = (total_cobrado / (total_cobrado + cartera_total) * 100) if (total_cobrado + cartera_total) else 0
        porcentaje_morosidad = (morosidad_total / cartera_total * 100) if cartera_total else 0
        resumen_por_analista.append({
            "analista": analista,
            "total_prestamos": total_prestamos,
            "total_clientes": total_clientes,
            "cartera_total": cartera_total,
            "morosidad_total": morosidad_total,
            "total_cobrado": total_cobrado,
            "porcentaje_cobrado": round(porcentaje_cobrado, 2),
            "porcentaje_morosidad": round(porcentaje_morosidad, 2),
        })
    return {
        "fecha_corte": fc.isoformat(),
        "resumen_por_analista": resumen_por_analista,
        "desempeno_mensual": [],
        "clientes_por_analista": [],
    }


def _generar_excel_asesores_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (MM/YYYY). Solo datos del mes reportado. Columnas: Analista, Total préstamos, Total vencimiento."""
    import openpyxl
    from openpyxl.styles import numbers
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])

    # Diccionario de meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        año = mes_data.get("año", datetime.now().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{año}")
        sheet_name = f"{mes_nombre} {año}"[:31]
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Reporte de Analistas", label])
        ws.append(["Solo cuotas que vencieron en este mes y no están pagadas"])
        ws.append([])
        ws.append(["Analista", "Total préstamos", "Total vencimiento ($)"])
        
        for r in mes_data.get("items", []):
            row_num = ws.max_row + 1
            ws.append([
                r.get("analista", ""), 
                r.get("total_prestamos", 0), 
                r.get("vencimiento_total", 0)
            ])
            # Aplicar formato contable con 2 decimales a la columna C (Total vencimiento)
            cell = ws.cell(row=row_num, column=3)
            cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# DEPRECATED: Este endpoint fue reemplazado por /exportar/morosidad
# El reporte de asesores ahora se genera desde el endpoint de morosidad (agrupado por cédula)
# Se mantiene comentado para referencia histórica
# @router.get("/exportar/asesores")
# def exportar_asesores(
#     db: Session = Depends(get_db),
#     formato: str = Query("excel", pattern="^(excel|pdf)$"),
#     fecha_corte: Optional[str] = Query(None),
#     meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses (una pestaña por mes)"),
#     años: Optional[str] = Query(None),
#     meses_list: Optional[str] = Query(None),
# ):
#     """DEPRECATED: Exporta reporte asesores. Excel: una pestaña por mes (MM/YYYY), solo datos del mes. PDF: fecha de corte única."""
#     if formato == "excel":
#         data_por_mes = get_asesores_por_mes(db=db, meses=meses, años=años, meses_list=meses_list)
#         content = _generar_excel_asesores_por_mes(data_por_mes)
#         hoy_str = date.today().isoformat()
#         return Response(
#             content=content,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename=reporte_asesores_{hoy_str}.xlsx"},
#         )
#     data = get_reporte_asesores(db=db, fecha_corte=fecha_corte)
#     import openpyxl
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Asesores"
#     ws.append(["Analista", "Préstamos", "Clientes", "Cartera", "Vencimiento", "Cobrado", "% Cobrado", "% Vencimiento"])
#     for r in data.get("resumen_por_analista", []):
#         ws.append([r.get("analista", ""), r.get("total_prestamos", 0), r.get("total_clientes", 0), r.get("cartera_total", 0), r.get("morosidad_total", 0), r.get("total_cobrado", 0), r.get("porcentaje_cobrado", 0), r.get("porcentaje_morosidad", 0)])
#     buf = io.BytesIO()
#     wb.save(buf)
#     content = buf.getvalue()
#     if formato == "pdf":
#         pdf_content = _generar_pdf_asesores(data)
#         return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_asesores_{data['fecha_corte']}.pdf"})
#     return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_asesores_{data['fecha_corte']}.xlsx"})


def _generar_pdf_asesores(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Asesores", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    rows = [["Analista", "Préstamos", "Clientes", "Cartera", "Vencimiento", "Cobrado", "% Cobrado", "% Vencimiento"]]
    for r in data.get("resumen_por_analista", []):
        rows.append([r.get("analista", ""), str(r.get("total_prestamos", 0)), str(r.get("total_clientes", 0)), str(r.get("cartera_total", 0)), str(r.get("morosidad_total", 0)), str(r.get("total_cobrado", 0)), str(r.get("porcentaje_cobrado", 0)), str(r.get("porcentaje_morosidad", 0))])
    t = Table(rows)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ---------- Reporte Productos ----------
@router.get("/productos/por-mes")
def get_productos_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    años: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Productos por mes: una pestaña por mes. Columnas: Modelo vehículo, Total financiamiento, Valor activo (70%)."""
    resultado: dict = {"meses": []}
    periodos = _periodos_desde_filtros(años, meses_list, meses)
    fecha_ref = func.coalesce(func.date(Prestamo.fecha_aprobacion), func.date(Prestamo.fecha_registro))

    for (año, mes) in periodos:
        inicio = date(año, mes, 1)
        _, ultimo = calendar.monthrange(año, mes)
        fin = date(año, mes, ultimo)

        rows = db.execute(
            select(
                Prestamo.modelo_vehiculo,
                func.coalesce(func.sum(Prestamo.total_financiamiento), 0).label("total_financiamiento"),
            )
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                fecha_ref >= inicio,
                fecha_ref <= fin,
            )
            .group_by(Prestamo.modelo_vehiculo)
            .order_by(func.sum(Prestamo.total_financiamiento).desc())
        ).fetchall()

        items = []
        for r in rows:
            total_fin = round(_safe_float(r.total_financiamiento), 2)
            valor_activo = round(total_fin * 0.70, 2)
            items.append({
                "modelo": r.modelo_vehiculo or "Sin modelo",
                "total_financiamiento": total_fin,
                "valor_activo": valor_activo,
            })

        resultado["meses"].append({
            "mes": mes,
            "año": año,
            "label": f"{mes:02d}/{año}",
            "items": items,
        })

    return resultado


@router.get("/productos")
def get_reporte_productos(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Reporte por producto. Datos reales desde BD."""
    fc = _parse_fecha(fecha_corte)
    productos = db.execute(
        select(Prestamo.producto)
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        .distinct()
    ).fetchall()
    resumen_por_producto: List[dict] = []
    for (producto,) in productos:
        if not producto:
            continue
        prestamos = db.execute(
            select(Prestamo.id)
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.producto == producto,
                Prestamo.estado == "APROBADO",
            )
        ).fetchall()
        ids = [x[0] for x in prestamos]
        total_prestamos = len(ids)
        total_clientes = db.scalar(select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(ids))) or 0
        cartera_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.prestamo_id.in_(ids)))) or 0
        promedio_prestamo = _safe_float(db.scalar(select(func.avg(Prestamo.total_financiamiento)).select_from(Prestamo).where(Prestamo.id.in_(ids)))) or 0
        total_cobrado = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.isnot(None), Cuota.prestamo_id.in_(ids)))) or 0
        morosidad_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc, Cuota.prestamo_id.in_(ids)))) or 0
        porcentaje_cobrado = (total_cobrado / (total_cobrado + cartera_total) * 100) if (total_cobrado + cartera_total) else 0
        resumen_por_producto.append({
            "producto": producto,
            "total_prestamos": total_prestamos,
            "total_clientes": total_clientes,
            "cartera_total": cartera_total,
            "promedio_prestamo": round(promedio_prestamo, 2),
            "total_cobrado": total_cobrado,
            "morosidad_total": morosidad_total,
            "porcentaje_cobrado": round(porcentaje_cobrado, 2),
        })
    productos_por_concesionario: List[dict] = []
    concesionarios = db.execute(
        select(Prestamo.concesionario)
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        .distinct()
    ).fetchall()
    for (conc,) in concesionarios:
        if not conc:
            continue
        for (producto,) in productos:
            if not producto:
                continue
            cnt = db.scalar(
                select(func.count())
                .select_from(Prestamo)
                .join(Cliente, Prestamo.cliente_id == Cliente.id)
                .where(
                    Cliente.estado == "ACTIVO",
                    Prestamo.concesionario == conc,
                    Prestamo.producto == producto,
                    Prestamo.estado == "APROBADO",
                )
            ) or 0
            if cnt:
                monto = _safe_float(
                    db.scalar(
                        select(func.sum(Prestamo.total_financiamiento))
                        .select_from(Prestamo)
                        .join(Cliente, Prestamo.cliente_id == Cliente.id)
                        .where(
                            Cliente.estado == "ACTIVO",
                            Prestamo.concesionario == conc,
                            Prestamo.producto == producto,
                            Prestamo.estado == "APROBADO",
                        )
                    )
                    or 0
                )
                productos_por_concesionario.append({"concesionario": conc, "producto": producto, "cantidad_prestamos": cnt, "monto_total": monto})
    return {
        "fecha_corte": fc.isoformat(),
        "resumen_por_producto": resumen_por_producto,
        "productos_por_concesionario": productos_por_concesionario,
        "tendencia_mensual": [],
    }


def _generar_excel_productos_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (MM/YYYY). Columnas: Modelo vehículo | Total financiamiento | Valor activo (70%)."""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])

    # Diccionario de meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        año = mes_data.get("año", datetime.now().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{año}")
        sheet_name = f"{mes_nombre} {año}"[:31]  # Título en español con límite de 31 caracteres
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Encabezado
        ws.append(["Reporte de Productos", label])
        ws.append([])
        ws.append(["Modelo de vehículo", "Total financiamiento ($)", "Valor del activo ($)"])
        
        # Datos
        items = mes_data.get("items", [])
        if items:
            for r in items:
                ws.append([
                    r.get("modelo", ""),
                    r.get("total_financiamiento", 0),
                    r.get("valor_activo", 0),
                ])
        else:
            # Si no hay datos, agregar fila informativa
            ws.append(["Sin datos", 0, 0])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/exportar/productos")
def exportar_productos(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses (una pestaña por mes)"),
    años: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Exporta reporte productos. Excel: una pestaña por mes (MM/YYYY), columnas Modelo | Total financiamiento | Valor activo (70%). PDF: clásico."""
    if formato == "excel":
        data_por_mes = get_productos_por_mes(db=db, meses=meses, años=años, meses_list=meses_list)
        content = _generar_excel_productos_por_mes(data_por_mes)
        hoy_str = date.today().isoformat()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_productos_{hoy_str}.xlsx"},
        )
    data = get_reporte_productos(db=db, fecha_corte=fecha_corte)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["Producto", "Préstamos", "Clientes", "Cartera", "Promedio", "Cobrado", "Mora", "% Cobrado"])
    for r in data.get("resumen_por_producto", []):
        ws.append([r.get("producto", ""), r.get("total_prestamos", 0), r.get("total_clientes", 0), r.get("cartera_total", 0), r.get("promedio_prestamo", 0), r.get("total_cobrado", 0), r.get("morosidad_total", 0), r.get("porcentaje_cobrado", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    if formato == "pdf":
        pdf_content = _generar_pdf_productos(data)
        return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_productos_{data['fecha_corte']}.pdf"})
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_productos_{data['fecha_corte']}.xlsx"})


@router.get("/por-cedula")
def get_reportes_por_cedula(db: Session = Depends(get_db)):
    """
    Reporte por cédula: id préstamo, cédula, nombre, total financiamiento, total abono,
    cuotas totales, cuotas pagadas, cuotas atrasadas (estado != PAGADO), monto cuotas atrasadas.
    """
    prestamos = (
        db.execute(
            select(Prestamo)
            .where(Prestamo.estado == "APROBADO")
            .order_by(Prestamo.id)
        )
    ).scalars().all()

    ids = [p.id for p in prestamos]
    abono_map: dict = {}
    pagadas_map: dict = {}
    atrasadas_map: dict = {}
    monto_atrasadas_map: dict = {}

    if ids:
        rows_abono = db.execute(
            select(Pago.prestamo_id, func.coalesce(func.sum(Pago.monto_pagado), 0))
            .where(Pago.prestamo_id.in_(ids))
            .group_by(Pago.prestamo_id)
        ).fetchall()
        abono_map = {r[0]: _safe_float(r[1]) for r in rows_abono}

        rows_pagadas = db.execute(
            select(Cuota.prestamo_id, func.count())
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado == "PAGADO")
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        pagadas_map = {r[0]: int(r[1]) for r in rows_pagadas}

        rows_atrasadas = db.execute(
            select(Cuota.prestamo_id, func.count())
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado != "PAGADO")
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        atrasadas_map = {r[0]: int(r[1]) for r in rows_atrasadas}

        # Monto total de cuotas atrasadas (pendientes, no pagadas)
        rows_monto_atrasadas = db.execute(
            select(Cuota.prestamo_id, func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado != "PAGADO")
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        monto_atrasadas_map = {r[0]: _safe_float(r[1]) for r in rows_monto_atrasadas}

    items: List[dict] = []
    for p in prestamos:
        cedula = p.cedula or ""
        nombre = p.nombres or ""
        if not cedula or not nombre:
            row_cli = db.execute(select(Cliente.cedula, Cliente.nombres).where(Cliente.id == p.cliente_id)).first()
            if row_cli:
                cedula = cedula or (row_cli.cedula or "")
                nombre = nombre or (row_cli.nombres or "")
        items.append({
            "id_prestamo": p.id,
            "cedula": cedula,
            "nombre": nombre,
            "total_financiamiento": round(_safe_float(p.total_financiamiento), 2),
            "total_abono": round(abono_map.get(p.id, 0), 2),
            "cuotas_totales": p.numero_cuotas or 0,
            "cuotas_pagadas": pagadas_map.get(p.id, 0),
            "cuotas_atrasadas": atrasadas_map.get(p.id, 0),
            "monto_cuotas_atrasadas": round(monto_atrasadas_map.get(p.id, 0), 2),
        })
    return {"items": items}


def _generar_excel_por_cedula(items: List[dict]) -> bytes:
    """Genera Excel reporte por cédula. Columnas A-I con totales en cabecera."""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Obtener mes y año actual
    hoy = datetime.now()
    mes_nombre = hoy.strftime("%B %Y")  # Ej: "February 2026"
    # Para español: usar meses en español
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    mes_es = meses_es.get(hoy.month, "")
    año = hoy.year
    titulo_pestana = f"Por cédula - {mes_es} {año}"
    
    ws.title = titulo_pestana
    ws.append(["Reporte por cédula"])
    ws.append([])
    
    # Calcular totales
    total_financiamiento = sum(r.get("total_financiamiento", 0) for r in items)
    total_abono = sum(r.get("total_abono", 0) for r in items)
    total_atrasos = sum(r.get("monto_cuotas_atrasadas", 0) for r in items)
    
    # Agregar fila de totales
    ws.append([
        f"Total financiamiento: ${total_financiamiento:.2f} | Total abono: ${total_abono:.2f} | Total atrasos: ${total_atrasos:.2f}"
    ])
    ws.append([])
    
    ws.append([
        "ID Préstamo",
        "Cédula",
        "Nombre",
        "Total financiamiento ($)",
        "Total abono ($)",
        "Cuotas totales",
        "Cuotas pagadas",
        "Cuotas atrasadas",
        "Cuotas atrasadas ($)",
    ])
    for r in items:
        ws.append([
            str(r.get("id_prestamo", "")),
            r.get("cedula", ""),
            r.get("nombre", ""),
            r.get("total_financiamiento", 0),
            r.get("total_abono", 0),
            r.get("cuotas_totales", 0),
            r.get("cuotas_pagadas", 0),
            r.get("cuotas_atrasadas", 0),
            r.get("monto_cuotas_atrasadas", 0),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/exportar/cedula")
def exportar_cedula(db: Session = Depends(get_db)):
    """Exporta reporte por cédula en Excel. Columnas: ID préstamo | Cédula | Nombre | Total financiamiento | Total abono | Cuotas totales | Cuotas pagadas | Cuotas atrasadas | Cuotas atrasadas ($)."""
    data = get_reportes_por_cedula(db=db)
    items = data.get("items", [])
    content = _generar_excel_por_cedula(items)
    hoy_str = date.today().isoformat()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_por_cedula_{hoy_str}.xlsx"},
    )


def _generar_pdf_productos(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Productos", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    rows = [["Producto", "Préstamos", "Clientes", "Cartera", "Promedio", "Cobrado", "Mora", "% Cobrado"]]
    for r in data.get("resumen_por_producto", []):
        rows.append([r.get("producto", ""), str(r.get("total_prestamos", 0)), str(r.get("total_clientes", 0)), str(r.get("cartera_total", 0)), str(r.get("promedio_prestamo", 0)), str(r.get("total_cobrado", 0)), str(r.get("morosidad_total", 0)), str(r.get("porcentaje_cobrado", 0))])
    t = Table(rows)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    prod_conc = data.get("productos_por_concesionario", [])
    if prod_conc:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Productos por concesionario", styles["Heading2"]))
        rows2 = [["Concesionario", "Producto", "Cant. préstamos", "Monto total"]]
        for r in prod_conc:
            rows2.append([r.get("concesionario", ""), r.get("producto", ""), str(r.get("cantidad_prestamos", 0)), str(r.get("monto_total", 0))])
        t2 = Table(rows2)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


# ---------- Reporte Contable (con cache) ----------
CONTABLE_CACHE_DIAS_ACTUALIZABLES = 7


def _obtener_tasa_usd_bs(fecha: date) -> float:
    """Obtiene tasa USD a Bolívares (Venezuela) para la fecha. Usa API externa o TASA_USD_BS_DEFAULT."""
    from app.core.config import settings
    import urllib.request
    import json

    if settings.TASA_USD_BS_DEFAULT is not None and fecha < date.today():
        return float(settings.TASA_USD_BS_DEFAULT)

    try:
        req = urllib.request.Request(
            settings.EXCHANGERATE_API_URL,
            headers={"User-Agent": "RapiCredit/1.0"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        rates = data.get("rates", {})
        tasa = rates.get("VES")
        if tasa is not None:
            return float(tasa)
    except Exception:
        pass

    if settings.TASA_USD_BS_DEFAULT is not None:
        return float(settings.TASA_USD_BS_DEFAULT)
    return 36.0


def _cuotas_a_filas_contable(rows, tasas_cache: dict) -> List[dict]:
    """Convierte filas de cuotas a formato contable con tasa guardada.
    Si total_pagado es NULL o 0 (datos legacy), usa monto_cuota como importe."""
    items: List[dict] = []
    for r in rows:
        fp = r.fecha_pago or r.fecha_pago_real
        fp_date = fp.date() if hasattr(fp, "date") else (date.fromisoformat(str(fp)[:10]) if fp else date.today())
        if fp_date not in tasas_cache:
            tasas_cache[fp_date] = _obtener_tasa_usd_bs(fp_date)
        tasa = tasas_cache[fp_date]

        monto_cuota = _safe_float(r.monto)
        total_pagado_raw = _safe_float(r.total_pagado)
        total_pagado = total_pagado_raw if total_pagado_raw > 0 else monto_cuota
        pago_completo = total_pagado >= monto_cuota - 0.01
        tipo_doc = f"Cuota {r.numero_cuota}" if pago_completo else "Abono"
        importe_ml = round(total_pagado * tasa, 2)

        items.append({
            "cuota_id": r.id,
            "cedula": r.cedula or "",
            "nombre": (r.nombres or "").strip(),
            "tipo_documento": tipo_doc,
            "fecha_vencimiento": r.fecha_vencimiento,
            "fecha_pago": fp_date,
            "importe_md": round(total_pagado, 2),
            "moneda_documento": "USD",
            "tasa": tasa,
            "importe_ml": importe_ml,
            "moneda_local": "Bs.",
        })
    return items


def _query_cuotas_contable(db: Session, fecha_inicio: date, fecha_fin: date):
    """Consulta cuotas con pago en el rango dado."""
    q = (
        select(
            Cuota.id,
            Cuota.numero_cuota,
            Cuota.fecha_vencimiento,
            Cuota.fecha_pago,
            Cuota.monto,
            Cuota.total_pagado,
            Prestamo.cedula,
            Prestamo.nombres,
            Pago.fecha_pago.label("fecha_pago_real"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .outerjoin(Pago, Cuota.pago_id == Pago.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.isnot(None),
            Cuota.fecha_pago >= fecha_inicio,
            Cuota.fecha_pago <= fecha_fin,
        )
    )
    return db.execute(q).fetchall()


def sync_reporte_contable_completo(db: Session) -> int:
    """Sincroniza todo el histórico al cache. Retorna cantidad de filas insertadas."""
    fi = date(2000, 1, 1)
    ff = date.today()
    rows = _query_cuotas_contable(db, fi, ff)
    tasas: dict = {}
    filas = _cuotas_a_filas_contable(rows, tasas)

    existentes = set(
        r[0] for r in db.execute(select(ReporteContableCache.cuota_id)).fetchall()
    )
    insertadas = 0
    for f in filas:
        if f["cuota_id"] not in existentes:
            db.add(ReporteContableCache(
                cuota_id=f["cuota_id"],
                cedula=f["cedula"],
                nombre=f["nombre"],
                tipo_documento=f["tipo_documento"],
                fecha_vencimiento=f["fecha_vencimiento"],
                fecha_pago=f["fecha_pago"],
                importe_md=f["importe_md"],
                moneda_documento=f["moneda_documento"],
                tasa=f["tasa"],
                importe_ml=f["importe_ml"],
                moneda_local=f["moneda_local"],
            ))
            insertadas += 1
    if insertadas:
        db.commit()
    return insertadas


def refresh_cache_ultimos_7_dias(db: Session) -> int:
    """Elimina y reinserta en cache solo los últimos 7 días. Retorna filas actualizadas."""
    hoy = date.today()
    limite = hoy - timedelta(days=CONTABLE_CACHE_DIAS_ACTUALIZABLES)

    # Eliminar filas con fecha_pago en ventana actualizable
    db.execute(delete(ReporteContableCache).where(ReporteContableCache.fecha_pago >= limite))
    db.commit()

    rows = _query_cuotas_contable(db, limite, hoy)
    tasas: dict = {}
    filas = _cuotas_a_filas_contable(rows, tasas)

    for f in filas:
        db.add(ReporteContableCache(
            cuota_id=f["cuota_id"],
            cedula=f["cedula"],
            nombre=f["nombre"],
            tipo_documento=f["tipo_documento"],
            fecha_vencimiento=f["fecha_vencimiento"],
            fecha_pago=f["fecha_pago"],
            importe_md=f["importe_md"],
            moneda_documento=f["moneda_documento"],
            tasa=f["tasa"],
            importe_ml=f["importe_ml"],
            moneda_local=f["moneda_local"],
        ))
    db.commit()
    return len(filas)


def get_reporte_contable_desde_cache(
    db: Session,
    años: List[int],
    meses: List[int],
    cedulas: Optional[List[str]] = None,
) -> dict:
    """Lee del cache filtrando por años, meses y opcionalmente cédulas."""
    if not años or not meses:
        return {"items": [], "fecha_inicio": "", "fecha_fin": ""}

    periodos = [(a, m) for a in años for m in meses]
    if not periodos:
        return {"items": [], "fecha_inicio": "", "fecha_fin": ""}

    condiciones = []
    for año, mes in periodos:
        _, ultimo = calendar.monthrange(año, mes)
        inicio = date(año, mes, 1)
        fin = date(año, mes, ultimo)
        condiciones.append(and_(ReporteContableCache.fecha_pago >= inicio, ReporteContableCache.fecha_pago <= fin))

    q = (
        select(ReporteContableCache)
        .where(or_(*condiciones) if len(condiciones) > 1 else condiciones[0])
        .order_by(ReporteContableCache.fecha_pago.asc(), ReporteContableCache.cedula.asc())
    )
    if cedulas:
        q = q.where(ReporteContableCache.cedula.in_(cedulas))

    rows = db.execute(q).scalars().all()
    items = [
        {
            "cedula": r.cedula,
            "nombre": r.nombre,
            "tipo_documento": r.tipo_documento,
            "fecha_vencimiento": r.fecha_vencimiento.isoformat() if r.fecha_vencimiento else "",
            "fecha_pago": r.fecha_pago.isoformat() if r.fecha_pago else "",
            "importe_md": _safe_float(r.importe_md),
            "moneda_documento": r.moneda_documento or "USD",
            "tasa": _safe_float(r.tasa),
            "importe_ml": _safe_float(r.importe_ml),
            "moneda_local": r.moneda_local or "Bs.",
        }
        for r in rows
    ]

    min_periodo = min(periodos, key=lambda p: (p[0], p[1]))
    max_periodo = max(periodos, key=lambda p: (p[0], p[1]))
    fi = date(min_periodo[0], min_periodo[1], 1)
    _, ult = calendar.monthrange(max_periodo[0], max_periodo[1])
    ff = date(max_periodo[0], max_periodo[1], ult)

    return {"items": items, "fecha_inicio": fi.isoformat(), "fecha_fin": ff.isoformat()}


def _generar_excel_contable(data: dict) -> bytes:
    """Genera Excel reporte contable. Columnas A-J según especificación."""
    import openpyxl
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Contable"

    items = data.get("items", [])
    ws.append(["Reporte Contable"])
    ws.append([f"Período: {data.get('fecha_inicio', '')} a {data.get('fecha_fin', '')}"])
    ws.append([])

    headers = [
        "Cédula",
        "Nombre",
        "Tipo de documento",
        "Fecha de vencimiento",
        "Fecha de pago",
        "Importe MD",
        "Moneda documento",
        "Tasa",
        "Importe en ML",
        "Moneda Local",
    ]
    ws.append(headers)

    for r in items:
        ws.append([
            r.get("cedula", ""),
            r.get("nombre", ""),
            r.get("tipo_documento", ""),
            r.get("fecha_vencimiento", ""),
            r.get("fecha_pago", ""),
            r.get("importe_md", 0),
            r.get("moneda_documento", "USD"),
            r.get("tasa", 0),
            r.get("importe_ml", 0),
            r.get("moneda_local", "Bs."),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/contable/cedulas")
def buscar_cedulas_contable(
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None, description="Búsqueda por cédula o nombre"),
    limit: int = Query(50, ge=1, le=200),
):
    """Busca cédulas para filtrar reporte contable. Usa cache si existe; si no, consulta fuente."""
    try:
        ReporteContableCache.__table__.create(db.get_bind(), checkfirst=True)
    except Exception:
        pass

    count_cache = db.scalar(select(func.count()).select_from(ReporteContableCache)) or 0
    if count_cache > 0:
        subq = select(ReporteContableCache.cedula, ReporteContableCache.nombre).distinct()
        if q and q.strip():
            term = f"%{q.strip()}%"
            subq = subq.where(
                or_(
                    ReporteContableCache.cedula.ilike(term),
                    ReporteContableCache.nombre.ilike(term),
                )
            )
        rows = db.execute(subq.limit(limit)).fetchall()
    else:
        subq = (
            select(Prestamo.cedula, Prestamo.nombres)
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.isnot(None),
                Cuota.total_pagado.isnot(None),
                Cuota.total_pagado > 0,
            )
            .distinct()
        )
        if q and q.strip():
            term = f"%{q.strip()}%"
            subq = subq.where(or_(Prestamo.cedula.ilike(term), Prestamo.nombres.ilike(term)))
        rows = db.execute(subq.limit(limit)).fetchall()
    return {"cedulas": [{"cedula": r[0] or "", "nombre": (r[1] or "").strip()} for r in rows]}


@router.get("/exportar/contable")
def exportar_contable(
    db: Session = Depends(get_db),
    años: Optional[str] = Query(None, description="Años separados por coma, ej. 2024,2025"),
    meses: Optional[str] = Query(None, description="Meses 1-12 separados por coma, ej. 1,6,12"),
    cedulas: Optional[str] = Query(None, description="Cédulas separadas por coma (vacío = todas)"),
):
    """
    Exporta reporte contable en Excel desde cache.
    Política: cache con todo el histórico; solo últimos 7 días se actualizan en cada descarga.
    Filtros: años, meses, cedulas (opcional).
    """
    # Crear tabla si no existe (compatible sin Alembic)
    try:
        ReporteContableCache.__table__.create(db.get_bind(), checkfirst=True)
    except Exception:
        pass

    # Si cache vacío, sincronizar histórico completo
    count_cache = db.scalar(select(func.count()).select_from(ReporteContableCache)) or 0
    if count_cache == 0:
        sync_reporte_contable_completo(db)
    else:
        refresh_cache_ultimos_7_dias(db)

    # Parsear filtros
    años_list = []
    if años:
        try:
            años_list = sorted([int(x.strip()) for x in años.split(",") if x.strip()], reverse=True)
        except (ValueError, TypeError):
            años_list = [date.today().year]
    if not años_list:
        años_list = [date.today().year]

    meses_list = []
    if meses:
        try:
            meses_list = sorted([int(x.strip()) for x in meses.split(",") if x.strip() and 1 <= int(x.strip()) <= 12])
        except (ValueError, TypeError):
            meses_list = list(range(1, 13))
    if not meses_list:
        meses_list = list(range(1, 13))

    cedulas_list = None
    if cedulas and cedulas.strip():
        cedulas_list = [c.strip() for c in cedulas.split(",") if c.strip()]

    data = get_reporte_contable_desde_cache(db, años=años_list, meses=meses_list, cedulas=cedulas_list)
    content = _generar_excel_contable(data)
    hoy_str = date.today().isoformat()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_contable_{hoy_str}.xlsx"},
    )
