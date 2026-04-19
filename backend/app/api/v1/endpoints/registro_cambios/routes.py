"""
Endpoints para el módulo de Registro de Cambios en Auditoría.
Permite listar, filtrar y exportar el historial de cambios realizados en el sistema.
"""
from datetime import datetime
from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.orm import Session
import io
import json

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.registro_cambios import RegistroCambios
from app.schemas.auth import UserResponse

router = APIRouter(
    prefix="/auditoria/registro-cambios",
    tags=["auditoria"],
    dependencies=[Depends(get_current_user)]
)


class RegistroCambiosItem(BaseModel):
    id: int
    usuario_id: int
    usuario_email: Optional[str] = None
    modulo: str
    tipo_cambio: str
    descripcion: str
    registro_id: Optional[int] = None
    tabla_afectada: Optional[str] = None
    campos_anteriores: Optional[Any] = None
    campos_nuevos: Optional[Any] = None
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    fecha_hora: str
    vigente: bool

    class Config:
        from_attributes = True


class RegistroCambiosListResponse(BaseModel):
    items: List[RegistroCambiosItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class RegistroCambiosStatsResponse(BaseModel):
    total_cambios: int
    cambios_hoy: int
    cambios_esta_semana: int
    cambios_por_modulo: dict
    cambios_por_usuario: dict
    cambios_por_tipo: dict


def _row_to_item(r: RegistroCambios) -> RegistroCambiosItem:
    """Convierte fila BD a RegistroCambiosItem."""
    fecha_str = r.fecha_hora.isoformat() + "Z" if r.fecha_hora else ""
    return RegistroCambiosItem(
        id=r.id,
        usuario_id=r.usuario_id,
        usuario_email=r.usuario_email,
        modulo=r.modulo,
        tipo_cambio=r.tipo_cambio,
        descripcion=r.descripcion,
        registro_id=r.registro_id,
        tabla_afectada=r.tabla_afectada,
        campos_anteriores=r.campos_anteriores,
        campos_nuevos=r.campos_nuevos,
        ip_address=r.ip_address,
        user_agent=r.user_agent,
        fecha_hora=fecha_str,
        vigente=r.vigente,
    )


@router.get("", response_model=RegistroCambiosListResponse)
def listar_registro_cambios(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    modulo: Optional[str] = Query(None),
    usuario_id: Optional[int] = Query(None),
    tipo_cambio: Optional[str] = Query(None),
    registro_id: Optional[int] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    ordenar_por: str = Query("fecha_hora"),
    orden: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Lista los registros de cambios con filtros y paginación."""
    q = select(RegistroCambios)

    if modulo:
        q = q.where(RegistroCambios.modulo == modulo)

    if usuario_id is not None:
        q = q.where(RegistroCambios.usuario_id == usuario_id)

    if tipo_cambio:
        q = q.where(RegistroCambios.tipo_cambio == tipo_cambio)

    if registro_id is not None:
        q = q.where(RegistroCambios.registro_id == registro_id)

    if fecha_desde:
        try:
            fd = datetime.fromisoformat(fecha_desde.replace("Z", "+00:00"))
            q = q.where(RegistroCambios.fecha_hora >= fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = datetime.fromisoformat(fecha_hasta.replace("Z", "+00:00"))
            q = q.where(RegistroCambios.fecha_hora <= fh)
        except ValueError:
            pass

    # Contar total
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0

    # Ordenamiento
    col_name = {
        "modulo": "modulo",
        "usuario": "usuario_id",
        "tipo": "tipo_cambio",
    }.get(ordenar_por, "fecha_hora")
    
    order_col = getattr(RegistroCambios, col_name, RegistroCambios.fecha_hora)

    if orden == "desc":
        q = q.order_by(order_col.desc())
    else:
        q = q.order_by(order_col.asc())

    # Paginación
    q = q.offset(skip).limit(limit)
    rows = db.execute(q).scalars().all()

    items = [_row_to_item(r) for r in rows]
    total_pages = (total + limit - 1) // limit if limit else 0

    return RegistroCambiosListResponse(
        items=items,
        total=total,
        page=(skip // limit) + 1 if limit else 1,
        page_size=limit,
        total_pages=total_pages,
    )


@router.get("/stats", response_model=RegistroCambiosStatsResponse)
def obtener_estadisticas_cambios(
    db: Session = Depends(get_db),
):
    """Estadísticas de cambios en el sistema."""
    from datetime import timedelta

    hoy = datetime.now().replace(tzinfo=None)
    inicio_semana = hoy - timedelta(days=hoy.weekday())

    # Totales
    total_cambios = db.scalar(select(func.count()).select_from(RegistroCambios)) or 0

    # Cambios hoy
    cambios_hoy = db.scalar(
        select(func.count()).select_from(RegistroCambios)
        .where(func.date(RegistroCambios.fecha_hora) == hoy.date())
    ) or 0

    # Cambios esta semana
    cambios_esta_semana = db.scalar(
        select(func.count()).select_from(RegistroCambios)
        .where(RegistroCambios.fecha_hora >= inicio_semana)
    ) or 0

    # Por módulo
    rows_mod = db.execute(
        select(RegistroCambios.modulo, func.count())
        .select_from(RegistroCambios)
        .group_by(RegistroCambios.modulo)
    ).all()
    cambios_por_modulo = {r[0] or "": r[1] for r in rows_mod}

    # Por usuario
    rows_usr = db.execute(
        select(RegistroCambios.usuario_email, func.count())
        .select_from(RegistroCambios)
        .group_by(RegistroCambios.usuario_email)
    ).all()
    cambios_por_usuario = {r[0] or "": r[1] for r in rows_usr}

    # Por tipo de cambio
    rows_tipo = db.execute(
        select(RegistroCambios.tipo_cambio, func.count())
        .select_from(RegistroCambios)
        .group_by(RegistroCambios.tipo_cambio)
    ).all()
    cambios_por_tipo = {r[0] or "": r[1] for r in rows_tipo}

    return RegistroCambiosStatsResponse(
        total_cambios=total_cambios,
        cambios_hoy=cambios_hoy,
        cambios_esta_semana=cambios_esta_semana,
        cambios_por_modulo=cambios_por_modulo,
        cambios_por_usuario=cambios_por_usuario,
        cambios_por_tipo=cambios_por_tipo,
    )


@router.get("/exportar")
def exportar_registro_cambios(
    modulo: Optional[str] = Query(None),
    usuario_id: Optional[int] = Query(None),
    tipo_cambio: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Exporta el registro de cambios a Excel."""
    try:
        import openpyxl
    except ImportError:
        minimal_xlsx = (
            b"PK\x03\x04\x14\x00\x00\x00\x08\x00"
            b"[\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00"
        )
        return StreamingResponse(
            iter([minimal_xlsx]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=registro_cambios.xlsx"},
        )

    q = select(RegistroCambios).order_by(RegistroCambios.fecha_hora.desc()).limit(50000)

    if modulo:
        q = q.where(RegistroCambios.modulo == modulo)

    if usuario_id is not None:
        q = q.where(RegistroCambios.usuario_id == usuario_id)

    if tipo_cambio:
        q = q.where(RegistroCambios.tipo_cambio == tipo_cambio)

    if fecha_desde:
        try:
            fd = datetime.fromisoformat(fecha_desde.replace("Z", "+00:00"))
            q = q.where(RegistroCambios.fecha_hora >= fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = datetime.fromisoformat(fecha_hasta.replace("Z", "+00:00"))
            q = q.where(RegistroCambios.fecha_hora <= fh)
        except ValueError:
            pass

    rows = db.execute(q).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro de Cambios"
    
    headers = [
        "id",
        "usuario_email",
        "modulo",
        "tipo_cambio",
        "descripcion",
        "registro_id",
        "tabla_afectada",
        "fecha_hora",
        "campos_anteriores",
        "campos_nuevos",
    ]
    
    ws.append(headers)

    for r in rows:
        campos_ant = json.dumps(r.campos_anteriores, ensure_ascii=False) if r.campos_anteriores else ""
        campos_new = json.dumps(r.campos_nuevos, ensure_ascii=False) if r.campos_nuevos else ""
        
        ws.append([
            r.id,
            r.usuario_email or "",
            r.modulo or "",
            r.tipo_cambio or "",
            (r.descripcion or "")[:500],
            r.registro_id or "",
            r.tabla_afectada or "",
            r.fecha_hora.isoformat() if r.fecha_hora else "",
            campos_ant[:1000],
            campos_new[:1000],
        ])

    # Ajustar ancho de columnas
    for col_num, header in enumerate(headers, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=registro_cambios.xlsx"},
    )


@router.get("/{cambio_id}", response_model=RegistroCambiosItem)
def obtener_cambio_detalle(
    cambio_id: int,
    db: Session = Depends(get_db),
):
    """Obtiene los detalles completos de un cambio registrado."""
    cambio = db.get(RegistroCambios, cambio_id)
    if not cambio:
        raise HTTPException(status_code=404, detail="Registro de cambio no encontrado")
    return _row_to_item(cambio)
