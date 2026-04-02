"""

Endpoints de préstamos. Datos reales desde BD (tabla prestamos).

Todos los endpoints usan Depends(get_db). No hay stubs ni datos demo.

"""

import calendar

import io

import logging

from datetime import date, datetime, timedelta, time

from decimal import Decimal

from typing import List, Optional



from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request

from fastapi.responses import Response

from pydantic import BaseModel, field_validator

from sqlalchemy import cast, delete, exists, func, or_, select, text, update

from sqlalchemy.exc import IntegrityError, OperationalError, ProgrammingError

from sqlalchemy.orm import Session

from sqlalchemy.types import Date



from app.core.database import get_db

from app.core.deps import get_current_user

from app.schemas.auth import UserResponse

from app.models.auditoria import Auditoria

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.cuota_pago import CuotaPago

from app.models.pago import Pago

from app.models.pago_con_error import PagoConError

from app.models.prestamo import Prestamo

from app.models.analista import Analista

from app.models.user import User

from app.models.revision_manual_prestamo import RevisionManualPrestamo

from app.models.auditoria_cambios_estado_prestamo import AuditoriaCambiosEstadoPrestamo
from app.models.auditoria_cartera_revision import AuditoriaCarteraRevision

from app.models.envio_notificacion import EnvioNotificacion

from app.models.datos_importados_conerrores import DatosImportadosConErrores

from app.models.auditoria_conciliacion_manual import AuditoriaConciliacionManual

from app.models.reporte_contable_cache import ReporteContableCache

from app.schemas.prestamo import PrestamoCreate, PrestamoResponse, PrestamoUpdate, PrestamoListResponse

from app.constants.prestamo_estados import prestamo_estado_exige_fecha_aprobacion

from app.api.v1.endpoints.pagos import aplicar_pagos_pendientes_prestamo

from app.services.pagos_cuotas_sincronizacion import sincronizar_pagos_pendientes_a_prestamos
from app.services.estado_cuenta_datos import obtener_pago_para_recibo_cuota, texto_institucion_recibo_cuota
from app.services.pagos_cuotas_reaplicacion import (
    integridad_cuotas_prestamo,
    reset_y_reaplicar_cascada_prestamo,
)

from app.services.cobros.recibo_cuota_amortizacion import generar_recibo_cuota_amortizacion
from app.services.cobros.recibo_cuota_moneda import contexto_moneda_montos_recibo_cuota

from app.services.analistas_catalogo_sync import (
    sincronizar_analistas_desde_prestamos_si_catalogo_vacio,
)

from app.services.cuota_estado import (
    estado_cuota_para_mostrar,
    etiqueta_estado_cuota,
    hoy_negocio,
    sincronizar_columna_estado_cuotas,
)

from app.services.prestamos.cupo_cedula_aprobados import validar_cupo_nuevo_prestamo_aprobado
from app.services.prestamos.prestamo_cedula_cliente_coherencia import (
    PrestamoCedulaClienteError,
    asegurar_prestamo_alineado_con_cliente,
    exigir_cliente_cedula_para_prestamo_aprobado,
)
from app.services.prestamos.prestamo_huella import ensure_no_duplicate_aprobado_huella
from app.services.prestamos.fechas_prestamo_coherencia import (
    alinear_fecha_aprobacion_y_base_calculo,
)
from app.services.prestamos.prestamo_fecha_referencia_query import (
    prestamo_fecha_referencia_negocio,
)

from app.services.prestamo_estado_coherencia import (
    condicion_filtro_estado_prestamo,
    prestamo_bloquea_insertar_filas_cuota_si_liquidado_bd,
    prestamo_bloquea_nuevas_cuotas_o_cambio_plazo,
    prestamo_ids_aprobados_todas_cuotas_cubiertas,
)
from app.services.prestamo_db_compat import (
    fetch_prestamos_fecha_desistimiento_map,
    prestamos_tiene_columna_fecha_desistimiento,
)



logger = logging.getLogger(__name__)

router = APIRouter(dependencies=[Depends(get_current_user)])


def _escape_ilike_pattern(fragment: str) -> str:
    """Escape LIKE metacharacters for ILIKE with escape backslash (PostgreSQL)."""
    return fragment.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _resolver_analista_para_prestamo(
    db: Session,
    analista: Optional[str],
    analista_id: Optional[int],
) -> tuple[str, Optional[int]]:
    """
    Catálogo analistas: prioriza analista_id; si solo viene texto y existe en catálogo, enlaza id.
    Texto sin coincidencia se guarda en prestamos.analista sin id (cargas legacy / Excel).
    """
    sincronizar_analistas_desde_prestamos_si_catalogo_vacio(db)
    if analista_id is not None:
        row_a = db.get(Analista, analista_id)
        if not row_a:
            raise HTTPException(status_code=400, detail="analista_id no existe en el catálogo")
        if not row_a.activo:
            raise HTTPException(
                status_code=400,
                detail="El analista está inactivo; reactívelo en Analistas o elija otro.",
            )
        return row_a.nombre, row_a.id
    if analista is not None and str(analista).strip():
        n = str(analista).strip()
        found = db.execute(select(Analista).where(Analista.nombre == n)).scalar_one_or_none()
        if found:
            if not found.activo:
                raise HTTPException(
                    status_code=400,
                    detail="El analista está inactivo; reactívelo en Analistas o elija otro.",
                )
            return found.nombre, found.id
        return n, None
    return "", None


class PrestamoIdsCuotasBody(BaseModel):
    """IDs de prestamos para devolver todas sus cuotas en una sola respuesta (plana)."""

    prestamo_ids: List[int]

    @field_validator("prestamo_ids")
    @classmethod
    def cap_ids(cls, v: List[int]) -> List[int]:
        if len(v) > 200:
            raise ValueError("Maximo 200 prestamo_ids por solicitud")
        return v


class CuotaUpdateAPI(BaseModel):
    fecha_vencimiento: Optional[date] = None
    fecha_pago: Optional[date] = None
    monto_cuota: Optional[float] = None
    total_pagado: Optional[float] = None
    observaciones: Optional[str] = None


def _listado_cuotas_prestamo_dicts(db: Session, prestamo_id: int) -> Optional[List[dict]]:
    row = db.get(Prestamo, prestamo_id)
    if not row:
        return None
    n_aplicados = sincronizar_pagos_pendientes_a_prestamos(db, [prestamo_id])
    if n_aplicados > 0:
        logger.info(
            "cuotas prestamo %s: aplicados %s pago(s) pendientes antes de listar",
            prestamo_id,
            n_aplicados,
        )
    cuotas = db.execute(
        select(Cuota).where(Cuota.prestamo_id == prestamo_id).order_by(Cuota.numero_cuota)
    ).scalars().all()
    sincronizar_columna_estado_cuotas(db, cuotas, commit=True)
    resultado: List[dict] = []
    for c in cuotas:
        try:
            pago_conciliado_flag = False
            pago_monto_conciliado = 0.0
            pago_verificado_concordancia = ""
            pago_monto_conciliado = float(c.total_pagado or 0)
            cp_links = db.execute(
                select(Pago).join(CuotaPago, CuotaPago.pago_id == Pago.id).where(
                    CuotaPago.cuota_id == c.id,
                    or_(
                        Pago.conciliado.is_(True),
                        func.coalesce(func.upper(func.trim(Pago.verificado_concordancia)), "") == "SI",
                    ),
                )
            ).scalars().all()
            if cp_links:
                pago_conciliado_flag = True
            if not pago_conciliado_flag and c.pago_id:
                pago = db.get(Pago, c.pago_id)
                if pago:
                    pago_conciliado_flag = bool(pago.conciliado)
                    pago_verificado_concordancia = str(pago.verificado_concordancia or "").strip().upper()
                    if pago_verificado_concordancia == "SI":
                        pago_conciliado_flag = True
            if not pago_conciliado_flag:
                if c.fecha_vencimiento:
                    fecha_inicio = c.fecha_vencimiento - timedelta(days=15)
                    fecha_fin = c.fecha_vencimiento + timedelta(days=15)
                    pagos_en_rango = db.execute(
                        select(Pago)
                        .where(
                            Pago.prestamo_id == prestamo_id,
                            func.date(Pago.fecha_pago) >= fecha_inicio,
                            func.date(Pago.fecha_pago) <= fecha_fin,
                        )
                        .order_by(Pago.fecha_pago.desc())
                    ).scalars().all()
                    for pago in pagos_en_rango:
                        if pago.conciliado or (str(pago.verificado_concordancia or "").strip().upper() == "SI"):
                            pago_conciliado_flag = True
                            break
            monto_cuota_f = float(c.monto or 0)
            total_pagado_f = float(c.total_pagado or 0)
            fv = c.fecha_vencimiento
            fv_date = fv.date() if fv and hasattr(fv, "date") else fv
            estado_mostrar = estado_cuota_para_mostrar(total_pagado_f, monto_cuota_f, fv_date, hoy_negocio())
            resultado.append(
                {
                    "id": c.id,
                    "prestamo_id": c.prestamo_id,
                    "pago_id": c.pago_id,
                    "numero_cuota": c.numero_cuota,
                    "fecha_vencimiento": c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else None,
                    "monto": float(c.monto) if c.monto is not None else 0,
                    "monto_cuota": float(c.monto) if c.monto is not None else 0,
                    "monto_capital": float(c.monto_capital) if c.monto_capital is not None else 0,
                    "monto_interes": float(c.monto_interes) if c.monto_interes is not None else 0,
                    "saldo_capital_inicial": float(c.saldo_capital_inicial) if c.saldo_capital_inicial is not None else 0,
                    "saldo_capital_final": float(c.saldo_capital_final) if c.saldo_capital_final is not None else 0,
                    "capital_pagado": None,
                    "interes_pagado": None,
                    "total_pagado": float(c.total_pagado) if c.total_pagado is not None else 0,
                    "fecha_pago": c.fecha_pago.isoformat() if c.fecha_pago else None,
                    "estado": estado_mostrar,
                    "estado_etiqueta": etiqueta_estado_cuota(estado_mostrar),
                    "dias_mora": c.dias_mora if c.dias_mora is not None else 0,
                    "dias_morosidad": c.dias_morosidad if c.dias_morosidad is not None else 0,
                    "pago_conciliado": pago_conciliado_flag,
                    "pago_monto_conciliado": pago_monto_conciliado,
                }
            )
        except Exception as e:
            logger.exception(
                "Error procesando cuota %s del prestamo %s: %s",
                c.id if hasattr(c, 'id') else 'unknown',
                prestamo_id,
                str(e)
            )
            raise HTTPException(
                status_code=500,
                detail=f"Error procesando cuota del préstamo {prestamo_id}: {str(e)}"
            )
    return resultado







def _audit_user_id(db: Session, current_user: UserResponse) -> int:

    """

    Devuelve un usuario_id que exista en la tabla usuarios (la que referencia la FK de auditoria).

    Usa consultas explícitas a public.usuarios para no depender del mapeo del modelo User.

    Verifica que el id exista antes de devolverlo (evita FK violation si hay desincronización).

    """

    def _verify_exists(uid: int) -> bool:

        """Comprueba que el id exista en usuarios (misma tabla que la FK de auditoria)."""

        r = db.execute(text("SELECT 1 FROM public.usuarios WHERE id = :id"), {"id": uid}).first()

        return r is not None



    # 1) Probar current_user.id si viene de BD (evita usar id de tabla users u otra fuente)

    uid = getattr(current_user, "id", None)

    if uid is not None and _verify_exists(uid):

        return uid



    # 2) Buscar por email (case-insensitive)

    email = (getattr(current_user, "email", None) or "").strip().lower()

    if email:

        r = db.execute(

            text("SELECT id FROM public.usuarios WHERE LOWER(email) = :e AND is_active = true LIMIT 1"),

            {"e": email},

        ).first()

        if r and _verify_exists(r[0]):

            return r[0]



    # 3) Cualquier administrador activo

    r = db.execute(

        text("SELECT id FROM public.usuarios WHERE rol = 'admin' AND is_active = true LIMIT 1"),

    ).first()

    if r and _verify_exists(r[0]):

        return r[0]



    # 4) Cualquier usuario activo

    r = db.execute(

        text("SELECT id FROM public.usuarios WHERE is_active = true ORDER BY id LIMIT 1"),

    ).first()

    if r and _verify_exists(r[0]):

        return r[0]



    # 5) Cualquier usuario

    r = db.execute(text("SELECT id FROM public.usuarios ORDER BY id LIMIT 1")).first()

    if r and _verify_exists(r[0]):

        return r[0]



    raise HTTPException(

        status_code=503,

        detail=(

            "No hay usuarios en la tabla usuarios; no se puede registrar la aprobación. "

            "Cree al menos un usuario activo en la aplicación."

        ),

    )





def _registrar_en_revision_manual(db: Session, prestamo_id: int) -> None:

    """

    Registra el préstamo en Revisión Manual (estado pendiente) si no existe.

    Se llama al aprobar un crédito para que aparezca automáticamente en la lista.

    """

    existente = db.execute(

        select(RevisionManualPrestamo).where(RevisionManualPrestamo.prestamo_id == prestamo_id)

    ).scalars().first()

    if not existente:

        rev = RevisionManualPrestamo(prestamo_id=prestamo_id, estado_revision="pendiente")

        db.add(rev)





# --- Schemas para body de endpoints adicionales ---

class AplicarCondicionesBody(BaseModel):

    tasa_interes: Optional[float] = 0.0  # Siempre 0% por defecto

    plazo_maximo: Optional[int] = None

    # Solo fecha_aprobacion: la base de calculo se copia en servidor (no enviar solo fecha_base).
    fecha_aprobacion: Optional[date] = None

    observaciones: Optional[str] = None





class EvaluarRiesgoBody(BaseModel):

    ml_impago_nivel_riesgo_manual: Optional[str] = None

    ml_impago_probabilidad_manual: Optional[float] = None

    requiere_revision: Optional[bool] = None

    estado: Optional[str] = None





class AsignarFechaAprobacionBody(BaseModel):

    fecha_aprobacion: date





class AprobarManualBody(BaseModel):

    """Body para aprobación manual de riesgo: una fecha, confirmaciones y datos editables."""

    fecha_aprobacion: date

    acepta_declaracion: bool  # Declaración políticas RapiCredit y riesgo

    documentos_analizados: bool  # Confirmación de que se analizaron documentos

    total_financiamiento: Optional[float] = None

    numero_cuotas: Optional[int] = None

    modalidad_pago: Optional[str] = None



    cuota_periodo: Optional[float] = None

    tasa_interes: Optional[float] = 0.0  # Siempre 0% por defecto

    observaciones: Optional[str] = None



    @field_validator("numero_cuotas")

    @classmethod

    def numero_cuotas_rango(cls, v: Optional[int]) -> Optional[int]:

        if v is not None and (v < 1 or v > 50):

            raise ValueError("numero_cuotas debe ser un entero entre 1 y 50")

        return v





@router.get("", response_model=dict)

def listar_prestamos(

    page: int = Query(1, ge=1),

    per_page: int = Query(20, ge=1, le=100),

    cliente_id: Optional[int] = Query(None),

    estado: Optional[str] = Query(None),

    analista: Optional[str] = Query(None),

    concesionario: Optional[str] = Query(None),

    cedula: Optional[str] = Query(None),

    fecha_inicio: Optional[str] = Query(None),

    fecha_fin: Optional[str] = Query(None),

    requiere_revision: Optional[bool] = Query(None),

    modelo: Optional[str] = Query(None),

    search: Optional[str] = Query(None),

    db: Session = Depends(get_db),

):

    """Listado paginado de préstamos desde BD con nombres y cédula del cliente (join)."""

    q = select(Prestamo, Cliente.nombres, Cliente.cedula).select_from(Prestamo).join(

        Cliente, Prestamo.cliente_id == Cliente.id

    )

    count_q = select(func.count()).select_from(Prestamo).join(Cliente, Prestamo.cliente_id == Cliente.id)



    if cliente_id is not None:

        q = q.where(Prestamo.cliente_id == cliente_id)

        count_q = count_q.where(Prestamo.cliente_id == cliente_id)

    if estado and estado.strip():

        est = estado.strip().upper()

        cond_est = condicion_filtro_estado_prestamo(est)

        if cond_est is not None:

            q = q.where(cond_est)

            count_q = count_q.where(cond_est)

        else:

            q = q.where(Prestamo.estado == est)

            count_q = count_q.where(Prestamo.estado == est)

    if analista and analista.strip():

        q = q.where(Prestamo.analista == analista.strip())

        count_q = count_q.where(Prestamo.analista == analista.strip())

    if concesionario and concesionario.strip():

        q = q.where(Prestamo.concesionario == concesionario.strip())

        count_q = count_q.where(Prestamo.concesionario == concesionario.strip())

    if cedula and cedula.strip():

        ced_clean = cedula.strip()

        q = q.where(Cliente.cedula == ced_clean)

        count_q = count_q.where(Cliente.cedula == ced_clean)

    if search and search.strip():

        search_clean = search.strip()

        # Si el search es numérico, permitir buscar directamente por ID de préstamo

        prestamo_id_search = None

        try:

            if search_clean.isdigit():

                prestamo_id_search = int(search_clean)

        except ValueError:

            prestamo_id_search = None

        pat = f"%{_escape_ilike_pattern(search_clean)}%"

        condicion_search = Cliente.cedula.ilike(pat, escape="\\") | Cliente.nombres.ilike(pat, escape="\\")

        if prestamo_id_search is not None:

            condicion_search = condicion_search | (Prestamo.id == prestamo_id_search)

        q = q.where(condicion_search)

        count_q = count_q.where(condicion_search)

    if modelo and modelo.strip():

        q = q.where(Prestamo.modelo_vehiculo == modelo.strip())

        count_q = count_q.where(Prestamo.modelo_vehiculo == modelo.strip())

    if requiere_revision is not None:

        q = q.where(Prestamo.requiere_revision == requiere_revision)

        count_q = count_q.where(Prestamo.requiere_revision == requiere_revision)

    if fecha_inicio:

        try:

            fd = datetime.fromisoformat(fecha_inicio.replace("Z", "+00:00")).date()

            q = q.where(prestamo_fecha_referencia_negocio() >= fd)

            count_q = count_q.where(prestamo_fecha_referencia_negocio() >= fd)

        except ValueError:

            pass

    if fecha_fin:

        try:

            fh = datetime.fromisoformat(fecha_fin.replace("Z", "+00:00")).date()

            q = q.where(prestamo_fecha_referencia_negocio() <= fh)

            count_q = count_q.where(prestamo_fecha_referencia_negocio() <= fh)

        except ValueError:

            pass



    total = db.scalar(count_q) or 0

    # Orden alineado con la columna "Fecha" (fecha_aprobacion) en la UI: más reciente primero.
    # Secundario: fecha_requerimiento e id (no fecha_registro).

    q = (

        q.order_by(

            Prestamo.fecha_aprobacion.desc().nullslast(),

            Prestamo.fecha_requerimiento.desc(),

            Prestamo.id.desc(),

        )

        .offset((page - 1) * per_page)

        .limit(per_page)

    )

    rows = db.execute(q).all()

    prestamo_ids = [row[0].id for row in rows]

    fd_desist_map = fetch_prestamos_fecha_desistimiento_map(db, prestamo_ids)

    # Conteo de cuotas por préstamo (para mostrar en columna Cuotas)

    cuotas_por_prestamo = {}

    if prestamo_ids:

        cuenta = select(Cuota.prestamo_id, func.count()).select_from(Cuota).where(

            Cuota.prestamo_id.in_(prestamo_ids)

        ).group_by(Cuota.prestamo_id)

        for pid, cnt in db.execute(cuenta).all():

            cuotas_por_prestamo[pid] = cnt

    

    # Estados de revisión manual

    revision_manual_estados = {}

    if prestamo_ids:

        try:

            rev_q = select(RevisionManualPrestamo.prestamo_id, RevisionManualPrestamo.estado_revision).where(

                RevisionManualPrestamo.prestamo_id.in_(prestamo_ids)

            )

            for pid, estado in db.execute(rev_q).all():

                revision_manual_estados[pid] = estado

        except (ProgrammingError, OperationalError) as e:

            logger.warning(

                "revision_manual_prestamos no disponible o error de BD al listar prestamos: %s",

                e,

            )

    

    try:

        liquidacion_efectiva_ids = prestamo_ids_aprobados_todas_cuotas_cubiertas(
            db, prestamo_ids
        )

    except (ProgrammingError, OperationalError) as e:

        logger.warning(
            "liquidacion efectiva (cuotas/cuota_pagos) no disponible al listar prestamos: %s",
            e,
        )

        liquidacion_efectiva_ids = set()

    items = []

    for row in rows:

        p, nombres_cliente, cedula_cliente = row[0], row[1], row[2]

        # Cuotas: preferir conteo desde tabla cuotas; si no hay, usar columna numero_cuotas

        numero_cuotas = cuotas_por_prestamo.get(p.id) if cuotas_por_prestamo.get(p.id) is not None else p.numero_cuotas

        estado_resp = p.estado or "DRAFT"

        if p.id in liquidacion_efectiva_ids:

            estado_resp = "LIQUIDADO"

        item = PrestamoListResponse(

            id=p.id,

            cliente_id=p.cliente_id,

            total_financiamiento=(p.total_financiamiento if p.total_financiamiento is not None else Decimal("0")),

            estado=estado_resp,

            concesionario=p.concesionario,

            modelo=p.modelo,

            analista=p.analista or "",

            fecha_creacion=p.fecha_creacion,

            fecha_actualizacion=p.fecha_actualizacion,

            fecha_registro=p.fecha_registro,

            fecha_aprobacion=p.fecha_aprobacion,

            nombres=nombres_cliente or p.nombres,

            cedula=cedula_cliente or p.cedula,

            numero_cuotas=numero_cuotas,

            modalidad_pago=p.modalidad_pago,

            revision_manual_estado=revision_manual_estados.get(p.id),  # None si no existe

            fecha_desistimiento=fd_desist_map.get(p.id),

        )

        items.append(item)

    total_pages = (total + per_page - 1) // per_page if total else 0

    return {

        "prestamos": items,

        "total": total,

        "page": page,

        "per_page": per_page,

        "total_pages": total_pages,

    }





@router.get("/stats", response_model=dict)

def get_prestamos_stats(

    analista: Optional[str] = Query(None),

    concesionario: Optional[str] = Query(None),

    modelo: Optional[str] = Query(None),

    mes: Optional[int] = Query(None),

    anio: Optional[int] = Query(None),

    db: Session = Depends(get_db),

):

    """Estadísticas de préstamos mensuales desde BD (solo clientes ACTIVOS).

    a) total_financiamiento: suma de total_financiamiento de préstamos aprobados en el mes.

    b) total: cantidad de préstamos aprobados en el mes.

    c) cartera_vigente: suma de monto de cuotas con vencimiento en el mes no cobradas.

    d) Cuenta por fecha_aprobacion por mes (solo préstamos con fecha_aprobacion en el mes). Estado APROBADO."""

    # Usar mes/anio de la BD cuando no se pasan, para coincidir con fechas de aprobacion/registro

    if mes is not None and 1 <= mes <= 12 and anio is not None and anio >= 2000:

        mes_u = mes

        anio_u = anio

        _, ultimo_dia = calendar.monthrange(anio_u, mes_u)

        inicio_mes = date(anio_u, mes_u, 1)

        fin_mes = date(anio_u, mes_u, ultimo_dia)

    else:

        # Zona Venezuela para coincidir con negocio

        primer_dia_mes = db.scalar(text(

            "SELECT date_trunc('month', (CURRENT_TIMESTAMP AT TIME ZONE 'America/Caracas'))::date"

        ))

        ultimo_dia_mes = db.scalar(text(

            "SELECT (date_trunc('month', (CURRENT_TIMESTAMP AT TIME ZONE 'America/Caracas') + INTERVAL '1 month') - INTERVAL '1 day')::date"

        ))

        if primer_dia_mes is None or ultimo_dia_mes is None:

            hoy = date.today()

            inicio_mes = hoy.replace(day=1)

            fin_mes = hoy

            mes_u = inicio_mes.month

            anio_u = inicio_mes.year

        else:

            inicio_mes = primer_dia_mes

            fin_mes = ultimo_dia_mes

            mes_u = inicio_mes.month

            anio_u = inicio_mes.year



    # Estados que cuentan como "aprobados en el mes"

    _estados_aprobados_kpi = ("APROBADO",)

    # Fecha de referencia: solo fecha_aprobacion (cuenta por fecha de aprobación por mes)

    # Solo clientes ACTIVOS (consistente con dashboard, pagos, reportes)

    fecha_ref = cast(Prestamo.fecha_aprobacion, Date)

    q_base = (

        select(Prestamo)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado.in_(_estados_aprobados_kpi),

            Prestamo.fecha_aprobacion.isnot(None),

            fecha_ref >= inicio_mes,

            fecha_ref <= fin_mes,

        )

    )

    if analista and analista.strip():

        q_base = q_base.where(Prestamo.analista == analista.strip())

    if concesionario and concesionario.strip():

        q_base = q_base.where(Prestamo.concesionario == concesionario.strip())

    if modelo and modelo.strip():

        q_base = q_base.where(Prestamo.modelo_vehiculo == modelo.strip())

    subq = q_base.subquery()

    total = db.scalar(select(func.count()).select_from(subq)) or 0

    fecha_ref2 = cast(Prestamo.fecha_aprobacion, Date)

    q_estado = (

        select(Prestamo.estado, func.count())

        .select_from(Prestamo)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado.in_(_estados_aprobados_kpi),

            Prestamo.fecha_aprobacion.isnot(None),

            fecha_ref2 >= inicio_mes,

            fecha_ref2 <= fin_mes,

        )

    )

    if analista and analista.strip():

        q_estado = q_estado.where(Prestamo.analista == analista.strip())

    if concesionario and concesionario.strip():

        q_estado = q_estado.where(Prestamo.concesionario == concesionario.strip())

    if modelo and modelo.strip():

        q_estado = q_estado.where(Prestamo.modelo_vehiculo == modelo.strip())

    q_estado = q_estado.group_by(Prestamo.estado)

    rows = db.execute(q_estado).all()

    por_estado = {r[0]: r[1] for r in rows}

    total_fin = db.scalar(select(func.coalesce(func.sum(subq.c.total_financiamiento), 0)).select_from(subq)) or 0

    total_fin = float(total_fin)

    promedio_monto = (total_fin / total) if total else 0

    # Por cobrar (mensual): suma de cuotas no cobradas de préstamos aprobados en el mes (por fecha_aprobacion)

    fecha_aprob_date = cast(Prestamo.fecha_aprobacion, Date)

    conds_cartera = [

        Cliente.estado == "ACTIVO",

        Prestamo.estado.in_(_estados_aprobados_kpi),

        Prestamo.fecha_aprobacion.isnot(None),

        fecha_aprob_date >= inicio_mes,

        fecha_aprob_date <= fin_mes,

        Cuota.fecha_pago.is_(None),

    ]

    if analista and analista.strip():

        conds_cartera.append(Prestamo.analista == analista.strip())

    if concesionario and concesionario.strip():

        conds_cartera.append(Prestamo.concesionario == concesionario.strip())

    if modelo and modelo.strip():

        conds_cartera.append(Prestamo.modelo_vehiculo == modelo.strip())

    cartera_vigente = float(

        db.scalar(

            select(func.coalesce(func.sum(Cuota.monto), 0))

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(*conds_cartera)

        ) or 0

    )

    return {

        "total": total,

        "por_estado": por_estado,

        "total_financiamiento": total_fin,

        "promedio_monto": promedio_monto,

        "cartera_vigente": cartera_vigente,

        "mes": mes_u,

        "anio": anio_u,

    }





def _normalizar_cedula_para_busqueda(cedula: str) -> str:

    """Normaliza cédula para búsqueda: sin guiones, sin espacios, mayúsculas."""

    if not cedula:

        return ""

    return (cedula or "").strip().upper().replace("-", "").replace(" ", "")





class PrestamosPorCedulasBatchBody(BaseModel):

    """Body para consulta batch de préstamos por múltiples cédulas."""



    cedulas: list[str] = []





@router.post("/cedula/batch", response_model=dict)

def listar_prestamos_por_cedulas_batch(

    body: PrestamosPorCedulasBatchBody,

    db: Session = Depends(get_db),

):

    """Consulta batch OPTIMIZADA: búsqueda rápida con normalización.

    

    Estrategia:

    1. Búsqueda exacta primero (si cédula en request = cédula en BD) - muy rápido

    2. Si no encuentra, busca todos los préstamos (sin WHERE) y filtra por normalización en Python

       (más rápido que SQL con func.upper/replace para strings)

    """

    cedulas_clean = [(c or "").strip() for c in (body.cedulas or []) if (c or "").strip()]

    cedulas_clean = list(dict.fromkeys(cedulas_clean))

    if not cedulas_clean:

        return {"prestamos": {}}



    resultado: dict[str, list] = {c: [] for c in cedulas_clean}

    cedulas_encontradas = set()

    

    # PASO 1: Búsqueda exacta (muy rápida, usa índices)

    q_exacto = (

        select(Prestamo.id, Prestamo.cliente_id, Prestamo.estado, Prestamo.cedula, Cliente.cedula)

        .select_from(Prestamo)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(or_(

            Cliente.cedula.in_(cedulas_clean),

            Prestamo.cedula.in_(cedulas_clean),

        ))

        .order_by(Prestamo.id.desc())

        .limit(50000)

    )

    

    for p_id, cli_id, p_estado, p_cedula, cli_cedula in db.execute(q_exacto):

        cedula_cli = (cli_cedula or p_cedula or "").strip()

        cedula_cli_norm = cedula_cli.replace("-", "").replace(" ", "").upper()

        # Buscar coincidencia normalizada en cedulas_clean
        cedula_encontrada = None
        for ced_clean in cedulas_clean:
            ced_clean_norm = ced_clean.replace("-", "").replace(" ", "").upper()
            if cedula_cli_norm == ced_clean_norm:
                cedula_encontrada = ced_clean
                break

        if cedula_encontrada:

            cedulas_encontradas.add(cedula_encontrada)

            resultado[cedula_encontrada].append({

                "id": p_id,

                "cliente_id": cli_id,

                "estado": p_estado,

                "cedula": cedula_cli,

            })

    

    # PASO 2: Para cédulas NO encontradas, buscar con normalización (sin guiones)

    cedulas_faltantes = [c for c in cedulas_clean if c not in cedulas_encontradas]

    if cedulas_faltantes:

        # Mapeo: cédula normalizada (sin guiones) → cédula original solicitada

        cedulas_norm_map = {}

        for ced in cedulas_faltantes:

            ced_norm = ced.replace("-", "").replace(" ", "").upper()

            cedulas_norm_map[ced_norm] = ced

        

        # Búsqueda más eficiente: usar OR en BD en lugar de traer todo el dataset

        # Construir condiciones OR para cada cédula normalizada

        or_conditions = []

        for ced_norm in cedulas_norm_map.keys():

            # Buscar: Cliente.cedula normalizada = ced_norm O Prestamo.cedula normalizada = ced_norm

            or_conditions.append(

                func.upper(func.replace(func.replace(Cliente.cedula, "-", ""), " ", "")) == ced_norm

            )

            or_conditions.append(

                func.upper(func.replace(func.replace(Prestamo.cedula, "-", ""), " ", "")) == ced_norm

            )

        

        # Ejecutar búsqueda con todas las condiciones

        q_faltantes = (

            select(Prestamo.id, Prestamo.cliente_id, Prestamo.estado, Prestamo.cedula, Cliente.cedula)

            .select_from(Prestamo)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(or_(*or_conditions) if or_conditions else False)

            .order_by(Prestamo.id.desc())

        )

        

        for p_id, cli_id, p_estado, p_cedula, cli_cedula in db.execute(q_faltantes):

            cedula_cli = (cli_cedula or p_cedula or "").strip()

            cedula_cli_norm = cedula_cli.replace("-", "").replace(" ", "").upper()

            

            # Match normalizado

            if cedula_cli_norm in cedulas_norm_map:

                ced_original = cedulas_norm_map[cedula_cli_norm]

                cedulas_encontradas.add(ced_original)

                resultado[ced_original].append({

                    "id": p_id,

                    "cliente_id": cli_id,

                    "estado": p_estado,

                    "cedula": cedula_cli,

                })

                # Eliminar del mapa para no procesar nuevamente

                del cedulas_norm_map[cedula_cli_norm]

                if not cedulas_norm_map:

                    break  # Ya encontró todas las cédulas faltantes



    return {"prestamos": resultado}





@router.get("/cedula/{cedula}", response_model=dict)

def listar_prestamos_por_cedula(cedula: str, db: Session = Depends(get_db)):

    """Listado de préstamos por cédula del cliente (integrado con frontend).

    Acepta coincidencia exacta o normalizada (sin guiones, mayúsculas) para Cliente.cedula y Prestamo.cedula.

    """

    try:

        cedula_clean = (cedula or "").strip()

        if not cedula_clean:

            return {"prestamos": [], "total": 0}

        cedula_norm = _normalizar_cedula_para_busqueda(cedula_clean)

        # Coincidencia: exacta o normalizada (V17709701 = V-17709701 = v17709701)

        cond_cliente = or_(

            Cliente.cedula == cedula_clean,

            func.upper(func.replace(func.replace(Cliente.cedula, "-", ""), " ", "")) == cedula_norm,

        )

        cond_prestamo = or_(

            Prestamo.cedula == cedula_clean,

            func.upper(func.replace(func.replace(Prestamo.cedula, "-", ""), " ", "")) == cedula_norm,

        )

        q = (

            select(Prestamo, Cliente.nombres, Cliente.cedula)

            .select_from(Prestamo)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(or_(cond_cliente, cond_prestamo))

            .order_by(Prestamo.id.desc())

        )

        rows = db.execute(q).all()

        prestamo_ids = [row[0].id for row in rows]

        fd_desist_map_ced = fetch_prestamos_fecha_desistimiento_map(db, prestamo_ids)

        cuotas_por_prestamo = {}

        if prestamo_ids:

            cuenta = (

                select(Cuota.prestamo_id, func.count())

                .select_from(Cuota)

                .where(Cuota.prestamo_id.in_(prestamo_ids))

                .group_by(Cuota.prestamo_id)

            )

            for pid, cnt in db.execute(cuenta).all():

                cuotas_por_prestamo[pid] = cnt

        liquidacion_efectiva_ids = prestamo_ids_aprobados_todas_cuotas_cubiertas(
            db, prestamo_ids
        )

        items = []

        for row in rows:

            p, nombres_cliente, cedula_cliente = row[0], row[1], row[2]

            numero_cuotas = cuotas_por_prestamo.get(p.id) if cuotas_por_prestamo.get(p.id) is not None else p.numero_cuotas

            estado_resp = p.estado or "DRAFT"

            if p.id in liquidacion_efectiva_ids:

                estado_resp = "LIQUIDADO"

            items.append(

                PrestamoListResponse(

                    id=p.id,

                    cliente_id=p.cliente_id,

                    total_financiamiento=p.total_financiamiento,

                    estado=estado_resp,

                    concesionario=p.concesionario,

                    modelo=p.modelo,

                    analista=p.analista,

                    fecha_creacion=p.fecha_creacion,

                    fecha_actualizacion=p.fecha_actualizacion,

                    fecha_registro=p.fecha_registro,

                    fecha_aprobacion=p.fecha_aprobacion,

                    nombres=nombres_cliente or p.nombres,

                    cedula=cedula_cliente or p.cedula,

                    numero_cuotas=numero_cuotas,

                    modalidad_pago=p.modalidad_pago,

                    fecha_desistimiento=fd_desist_map_ced.get(p.id),

                )

            )

        return {"prestamos": [i.model_dump() for i in items], "total": len(items)}







    except Exception as e:

        logger.exception("listar_prestamos_por_cedula error: %s", e)

        return {"prestamos": [], "total": 0}

@router.get("/cedula/{cedula}/resumen", response_model=dict)

def resumen_prestamos_por_cedula(cedula: str, db: Session = Depends(get_db)):

    """Resumen de préstamos por cédula: total, saldo pendiente, cuotas en mora (integrado con frontend)."""

    cedula_clean = (cedula or "").strip()

    if not cedula_clean:

        return {

            "tiene_prestamos": False,

            "total_prestamos": 0,

            "total_saldo_pendiente": 0,

            "total_cuotas_mora": 0,

            "prestamos": [],

        }

    q = (

        select(Prestamo)

        .select_from(Prestamo)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(Cliente.cedula == cedula_clean)

        .order_by(Prestamo.id.desc())

    )

    prestamos = db.execute(q).scalars().all()

    if not prestamos:

        return {

            "tiene_prestamos": False,

            "total_prestamos": 0,

            "total_saldo_pendiente": 0,

            "total_cuotas_mora": 0,

            "prestamos": [],

        }

    from datetime import date

    hoy = date.today()

    prestamo_ids = [p.id for p in prestamos]

    # Saldo pendiente y cuotas en mora por préstamo (solo cuotas sin fecha_pago)

    saldo_q = (

        select(Cuota.prestamo_id, func.coalesce(func.sum(Cuota.monto), 0))

        .select_from(Cuota)

        .where(Cuota.prestamo_id.in_(prestamo_ids), Cuota.fecha_pago.is_(None))

        .group_by(Cuota.prestamo_id)

    )

    saldos = {r[0]: float(r[1]) for r in db.execute(saldo_q).all()}

    mora_q = (

        select(Cuota.prestamo_id, func.count())

        .select_from(Cuota)

        .where(

            Cuota.prestamo_id.in_(prestamo_ids),

            Cuota.fecha_pago.is_(None),

            Cuota.fecha_vencimiento < hoy,

        )

        .group_by(Cuota.prestamo_id)

    )

    moras = {r[0]: r[1] for r in db.execute(mora_q).all()}

    total_saldo = sum(saldos.get(pid, 0) for pid in prestamo_ids)

    total_mora = sum(moras.get(pid, 0) for pid in prestamo_ids)

    list_prestamos = []

    for p in prestamos:

        list_prestamos.append({

            "id": p.id,

            "modelo_vehiculo": p.modelo or "",

            "total_financiamiento": float(p.total_financiamiento) if p.total_financiamiento is not None else 0,

            "saldo_pendiente": saldos.get(p.id, 0),

            "cuotas_en_mora": moras.get(p.id, 0),

            "estado": p.estado or "",

            "fecha_registro": p.fecha_registro.isoformat() if p.fecha_registro else None,

        })

    return {

        "tiene_prestamos": True,

        "total_prestamos": len(prestamos),

        "total_saldo_pendiente": total_saldo,

        "total_cuotas_mora": total_mora,

        "prestamos": list_prestamos,

    }





@router.get("/{prestamo_id}", response_model=PrestamoResponse)

def get_prestamo(prestamo_id: int, db: Session = Depends(get_db)):

    """Obtiene un préstamo por ID desde BD. Incluye cedula/nombres del cliente (join si faltan en prestamo)."""

    row = db.execute(

        select(Prestamo, Cliente.nombres, Cliente.cedula)

        .select_from(Prestamo)

        .outerjoin(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(Prestamo.id == prestamo_id)

    ).first()

    if not row:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    p, nombres_cliente, cedula_cliente = row[0], row[1], row[2]

    resp = PrestamoResponse.model_validate(p)

    fd_one = fetch_prestamos_fecha_desistimiento_map(db, [prestamo_id])

    if prestamo_id in fd_one:

        resp.fecha_desistimiento = fd_one[prestamo_id]

    # Preferir cedula/nombres del cliente (join) si faltan o vacíos en prestamo

    resp.nombres = nombres_cliente or p.nombres or ""

    resp.cedula = cedula_cliente or p.cedula or ""

    if prestamo_id in prestamo_ids_aprobados_todas_cuotas_cubiertas(db, [prestamo_id]):

        resp.estado = "LIQUIDADO"

    return resp





def _calcular_monto_cuota_frances(total: float, tasa_periodo: float, n: int) -> float:

    """

    Calcula la cuota fija de un préstamo con amortización francesa (sistema francés / annuity).

    Fórmula: C = P * [i * (1+i)^n] / [(1+i)^n - 1]

    donde P=capital, i=tasa por período, n=número de cuotas.

    Si tasa_periodo == 0 devuelve la cuota plana (P/n).

    """

    if tasa_periodo <= 0 or n <= 0:

        return total / n if n else 0

    factor = (1 + tasa_periodo) ** n

    return total * (tasa_periodo * factor) / (factor - 1)





def _fecha_base_mas_meses(fecha_base: date, meses: int) -> date:

    """

    Suma `meses` meses a `fecha_base` manteniendo el mismo dia del mes.

    Si el dia no existe en el mes resultante (ej. 31 en febrero), usa el ultimo dia del mes.

    Ejemplo: 17 julio 2025 + 1 mes = 17 agosto 2025; + 2 meses = 17 septiembre 2025.

    """

    if meses <= 0:

        return fecha_base

    year = fecha_base.year + (fecha_base.month + meses - 1) // 12

    month = (fecha_base.month + meses - 1) % 12 + 1

    _, ultimo_dia = calendar.monthrange(year, month)

    day = min(fecha_base.day, ultimo_dia)

    return date(year, month, day)





def _fecha_requerimiento_date(p: "Prestamo") -> Optional[date]:

    """Devuelve fecha_requerimiento del préstamo como date, o None."""

    fr = getattr(p, "fecha_requerimiento", None)

    if not fr:

        return None

    if hasattr(fr, "date") and callable(getattr(fr, "date", None)):

        return fr.date()

    if isinstance(fr, date):

        return fr

    return None





def _ajustar_req_si_mayor_que_aprobacion(
    db: Session,
    row: Prestamo,
    *,
    origen: str,
    usuario_id: int,
) -> None:
    """
    Si fecha_requerimiento > dia de fecha_aprobacion, iguala requerimiento a ese dia.
    Deja rastro en observaciones y en tabla auditoria.
    """
    if not row.fecha_aprobacion or not row.fecha_requerimiento:
        return
    ap_date = (
        row.fecha_aprobacion.date()
        if hasattr(row.fecha_aprobacion, "date")
        else row.fecha_aprobacion
    )
    req_date = row.fecha_requerimiento
    if req_date <= ap_date:
        return
    req_old = req_date
    row.fecha_requerimiento = ap_date
    logger.info(
        "[%s] Auto-ajuste fecha_requerimiento %s -> %s (fecha_aprobacion %s)",
        origen,
        req_old,
        ap_date,
        ap_date,
    )
    note = (
        f"\n[AUTO {origen}] fecha_requerimiento ajustada de {req_old} a {ap_date} "
        f"(alineada con fecha_aprobacion)."
    )
    prev_obs = (row.observaciones or "").strip()
    row.observaciones = (prev_obs + note).strip() if prev_obs else note.strip()
    db.add(
        Auditoria(
            usuario_id=usuario_id,
            accion="AJUSTE_FECHA_REQUERIMIENTO",
            entidad="prestamos",
            entidad_id=getattr(row, "id", None),
            detalles=(
                f"{origen}: fecha_requerimiento {req_old} -> {ap_date} "
                f"(coherencia con fecha_aprobacion)."
            ),
            exito=True,
        )
    )





def _fecha_para_amortizacion(p: "Prestamo") -> Optional[date]:

    """

    Regla unica: la amortizacion usa fecha_base_calculo (copia de la fecha de aprobacion en formularios).

    Compatibilidad: si fecha_base_calculo es NULL (datos antiguos), se usa la parte fecha de fecha_aprobacion.

    """

    fb = getattr(p, "fecha_base_calculo", None)

    if fb is not None:

        if isinstance(fb, datetime):

            return fb.date()

        if isinstance(fb, date):

            return fb

    fa = getattr(p, "fecha_aprobacion", None)

    if not fa:

        return None

    if hasattr(fa, "date") and callable(getattr(fa, "date", None)):

        return fa.date()

    if isinstance(fa, date):

        return fa

    return None





def _resolver_monto_cuota(prestamo: "Prestamo", total: float, numero_cuotas: int) -> float:

    """

    [C1] Determina el monto de cuota considerando la tasa de interés del préstamo.

    - tasa_interes == 0 (o NULL): cuota plana = total / numero_cuotas.

    - tasa_interes > 0: amortización francesa.

      La tasa se almacena como porcentaje anual (ej. 12 = 12% anual).

      Se convierte al período: MENSUAL → tasa/12, QUINCENAL → tasa/24, SEMANAL → tasa/52.

    """

    tasa_anual = float(prestamo.tasa_interes or 0)

    if tasa_anual <= 0 or numero_cuotas <= 0:

        return total / numero_cuotas if numero_cuotas else 0

    modalidad = (prestamo.modalidad_pago or "MENSUAL").upper()

    periodos_por_anio = 12 if modalidad == "MENSUAL" else (24 if modalidad == "QUINCENAL" else 52)

    tasa_periodo = (tasa_anual / 100) / periodos_por_anio

    return _calcular_monto_cuota_frances(total, tasa_periodo, numero_cuotas)





def _generar_cuotas_amortizacion(db: Session, p: Prestamo, fecha_base: date, numero_cuotas: int, monto_cuota: float) -> int:

    """

    Genera filas en tabla cuotas para el préstamo dado.

    Regla unica: fecha_base debe ser fecha_base_calculo del prestamo (misma fecha calendario que fecha_aprobacion).



    Fecha de vencimiento:

      - MENSUAL: mismo día del mes; cuota n = fecha_base + n meses (ej. 2 ene → 2 feb, 2 mar).

      - QUINCENAL: fin de cada quincena; cuota n = fecha_base + (15*n - 1) días (ej. 1 ene → 15 ene, 30 ene, 14 feb).

      - SEMANAL: fin de cada semana; cuota n = fecha_base + (7*n - 1) días.



    [C1] Cálculo del monto de cuota:

    - Si tasa_interes == 0 (o NULL): cuota plana = total_financiamiento / numero_cuotas.

    - Si tasa_interes > 0: amortización francesa. La tasa se interpreta como tasa ANUAL

      nominal; se convierte al período según modalidad_pago antes de calcular.

      El `monto_cuota` que recibe este helper ya viene calculado por el caller; si el

      caller quiere usar interés debe llamar a _calcular_monto_cuota_frances primero.



    Los campos monto_capital e monto_interes no se almacenan en cuotas (ver auditoría B3);

    se derivan en el frontend desde saldo_capital_inicial / saldo_capital_final.

    """

    bloqueo_ins = prestamo_bloquea_insertar_filas_cuota_si_liquidado_bd(p)

    if bloqueo_ins:

        raise HTTPException(status_code=400, detail=bloqueo_ins)

    modalidad = (p.modalidad_pago or "MENSUAL").upper()

    delta_dias = 15 if modalidad == "QUINCENAL" else (7 if modalidad == "SEMANAL" else None)

    cliente_id = p.cliente_id

    total = monto_cuota * numero_cuotas

    monto_cuota_dec = Decimal(str(round(monto_cuota, 2)))

    creadas = 0

    for n in range(1, numero_cuotas + 1):

        if modalidad == "MENSUAL":

            next_date = _fecha_base_mas_meses(fecha_base, n)

        else:

            # QUINCENAL: 15*n-1 días (1 ene → 15 ene, 30 ene, 14 feb); SEMANAL: 7*n-1

            next_date = fecha_base + timedelta(days=delta_dias * n - 1)

        saldo_inicial = Decimal(str(round(total - (n - 1) * monto_cuota, 2)))

        saldo_final = Decimal(str(round(total - n * monto_cuota, 2)))

        if saldo_final < 0:

            saldo_final = Decimal("0")

        c = Cuota(

            prestamo_id=p.id,

            cliente_id=cliente_id,

            numero_cuota=n,

            fecha_vencimiento=next_date,

            monto=monto_cuota_dec,

            saldo_capital_inicial=saldo_inicial,

            saldo_capital_final=saldo_final,

            # [B3] Calcular desglose: capital = diferencia de saldos; interés = residuo

            monto_capital=saldo_inicial - saldo_final,

            monto_interes=monto_cuota_dec - (saldo_inicial - saldo_final),

            estado="PENDIENTE",

            dias_mora=0,

        )

        db.add(c)

        creadas += 1

    return creadas


def _recalcular_fechas_vencimiento_cuotas(db: Session, p: Prestamo, fecha_base: date) -> dict:
    """
    Recalcula SOLO las fechas de vencimiento de las cuotas existentes.
    Mantiene: montos, pagos, saldos de capital.
    Recalcula: fecha_vencimiento y estado de cuota (VENCIDO, PENDIENTE, MOROSO, etc.).
    
    Regla: fecha_base debe ser la nueva fecha_base_calculo del prestamo (alineada con fecha_aprobacion).
    
    Returns: dict con estadísticas de actualización.
    """
    from app.services.cuota_estado import clasificar_estado_cuota, hoy_negocio
    
    cuotas = db.query(Cuota).filter(Cuota.prestamo_id == p.id).order_by(Cuota.numero_cuota).all()
    
    if not cuotas:
        return {"message": "Sin cuotas para recalcular", "actualizadas": 0}
    
    modalidad = (p.modalidad_pago or "MENSUAL").upper()
    delta_dias = 15 if modalidad == "QUINCENAL" else (7 if modalidad == "SEMANAL" else None)
    hoy = hoy_negocio()
    
    actualizadas = 0
    
    for cuota in cuotas:
        numero_cuota = cuota.numero_cuota
        
        # Calcular nueva fecha de vencimiento usando mismo algoritmo que generación
        if modalidad == "MENSUAL":
            nueva_fecha = _fecha_base_mas_meses(fecha_base, numero_cuota)
        else:
            # QUINCENAL: 15*n-1 días; SEMANAL: 7*n-1
            nueva_fecha = fecha_base + timedelta(days=delta_dias * numero_cuota - 1)
        
        # Actualizar fecha de vencimiento
        cuota.fecha_vencimiento = nueva_fecha
        
        # Recalcular estado de cuota según nueva fecha de vencimiento y pagos existentes
        total_pagado = float(getattr(cuota, "total_pagado", None) or 0)
        monto_cuota = float(getattr(cuota, "monto", None) or 0)
        
        nuevo_estado = clasificar_estado_cuota(
            total_pagado,
            monto_cuota,
            nueva_fecha,
            hoy
        )
        
        cuota.estado = nuevo_estado
        
        db.add(cuota)
        actualizadas += 1
    
    db.commit()
    
    return {
        "message": "Fechas de vencimiento recalculadas exitosamente",
        "actualizadas": actualizadas
    }



@router.post("/cuotas/by-prestamo-ids", response_model=list)
def post_cuotas_by_prestamo_ids(
    body: PrestamoIdsCuotasBody,
    db: Session = Depends(get_db),
):
    """
    Devuelve en una lista plana todas las cuotas de los prestamos indicados (mismo shape que GET .../cuotas).
    Omite IDs inexistentes. Maximo 200 prestamos por solicitud.
    """
    out: List[dict] = []
    seen: set[int] = set()
    for pid in body.prestamo_ids:
        if pid in seen:
            continue
        seen.add(pid)
        part = _listado_cuotas_prestamo_dicts(db, pid)
        if part:
            out.extend(part)
    return out




@router.put("/{prestamo_id}/cuotas/{cuota_id}", response_model=dict)
def put_cuota_prestamo(
    prestamo_id: int,
    cuota_id: int,
    body: CuotaUpdateAPI,
    db: Session = Depends(get_db),
):
    c = db.get(Cuota, cuota_id)
    if not c or c.prestamo_id != prestamo_id:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    if body.fecha_vencimiento is not None:
        c.fecha_vencimiento = body.fecha_vencimiento
    if body.fecha_pago is not None:
        c.fecha_pago = body.fecha_pago
    if body.monto_cuota is not None:
        c.monto = Decimal(str(round(float(body.monto_cuota), 2)))
    if body.total_pagado is not None:
        c.total_pagado = Decimal(str(round(float(body.total_pagado), 2)))
    if body.observaciones is not None:
        c.observaciones = body.observaciones
    db.flush()
    todas = db.execute(
        select(Cuota).where(Cuota.prestamo_id == prestamo_id).order_by(Cuota.numero_cuota)
    ).scalars().all()
    sincronizar_columna_estado_cuotas(db, todas, commit=False)
    db.commit()
    refreshed = _listado_cuotas_prestamo_dicts(db, prestamo_id) or []
    for row in refreshed:
        if row.get("id") == cuota_id:
            return row
    raise HTTPException(status_code=500, detail="No se pudo reconstruir la cuota actualizada")


@router.delete("/{prestamo_id}/cuotas/{cuota_id}", status_code=204)
def delete_cuota_prestamo(
    prestamo_id: int,
    cuota_id: int,
    db: Session = Depends(get_db),
):
    c = db.get(Cuota, cuota_id)
    if not c or c.prestamo_id != prestamo_id:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    db.delete(c)
    db.commit()
    return None


@router.get("/{prestamo_id}/cuotas", response_model=list)

def get_cuotas_prestamo(prestamo_id: int, db: Session = Depends(get_db)):
    """
    Lista las cuotas (tabla de amortizacion) de un prestamo, con info de pago conciliado.
    Aplica pagos pendientes, alinea columna `cuotas.estado` con la regla unificada y devuelve estado + etiqueta.
    """
    out = _listado_cuotas_prestamo_dicts(db, prestamo_id)
    if out is None:
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    return out









@router.get("/{prestamo_id}/cuotas/{cuota_id}/recibo.pdf")

def get_recibo_cuota_pdf(prestamo_id: int, cuota_id: int, db: Session = Depends(get_db)):

    """Genera el recibo PDF de una cuota (mismo formato que cobros)."""

    prestamo = db.get(Prestamo, prestamo_id)

    if not prestamo:

        raise HTTPException(status_code=404, detail="Prestamo no encontrado")

    cuota = db.get(Cuota, cuota_id)

    if not cuota or cuota.prestamo_id != prestamo_id:

        raise HTTPException(status_code=404, detail="Cuota no encontrada")

    monto_cuota = float(cuota.monto or 0)

    total_pagado = float(cuota.total_pagado or 0)

    if total_pagado <= 0 and monto_cuota <= 0:

        raise HTTPException(status_code=400, detail="La cuota no tiene monto pagado")

    referencia = f"Cuota-{cuota.numero_cuota}-Prestamo-{prestamo_id}"

    institucion = "N/A"

    numero_operacion = referencia

    fecha_recep = None

    fecha_pago_date = None

    pago = obtener_pago_para_recibo_cuota(db, cuota)

    if pago:

        institucion = texto_institucion_recibo_cuota(db, pago)

        nd = (pago.numero_documento or "").strip()
        rp = (pago.referencia_pago or "").strip()
        if nd.upper().startswith("COB-") and rp and not rp.upper().startswith("COB-"):
            numero_operacion = rp[:100]
        else:
            numero_operacion = (nd or rp or referencia)[:100]

        if pago.fecha_pago:

            fecha_recep = pago.fecha_pago

            fecha_pago_date = pago.fecha_pago.date() if hasattr(pago.fecha_pago, "date") else pago.fecha_pago

    if not fecha_recep and cuota.fecha_pago:

        fp_c = cuota.fecha_pago

        if isinstance(fp_c, datetime):

            fecha_recep = fp_c

            fecha_pago_date = fp_c.date()

        else:

            fecha_pago_date = fp_c

            fecha_recep = datetime.combine(fp_c, datetime.min.time())

    fv_c = cuota.fecha_vencimiento

    fv_date_c = fv_c.date() if fv_c and hasattr(fv_c, "date") else fv_c

    estado_codigo = estado_cuota_para_mostrar(total_pagado, monto_cuota, fv_date_c, hoy_negocio())

    estado_cuota_lbl = etiqueta_estado_cuota(estado_codigo)

    fpd = "-"

    if fecha_pago_date:

        fpd = fecha_pago_date.strftime("%d/%m/%Y")

    fecha_reporte_aprobacion_display = None
    if pago:
        f_rep = getattr(pago, "fecha_conciliacion", None) or getattr(pago, "fecha_registro", None)
        if f_rep and hasattr(f_rep, "strftime"):
            fecha_reporte_aprobacion_display = f_rep.strftime("%d/%m/%Y %H:%M")

    ctx = contexto_moneda_montos_recibo_cuota(db, prestamo, cuota, pago)

    pdf_bytes = generar_recibo_cuota_amortizacion(

        referencia_interna=referencia,

        nombres_completos=(prestamo.nombres or "").strip(),

        cedula=(prestamo.cedula or "").strip(),

        institucion_financiera=institucion,

        monto=ctx.monto_str,

        numero_operacion=numero_operacion,

        fecha_recepcion=fecha_recep,

        fecha_pago=fecha_pago_date,

        fecha_reporte_aprobacion_display=fecha_reporte_aprobacion_display,

        moneda=ctx.moneda,

        tasa_cambio=ctx.tasa_cambio,

        aplicado_a_cuotas=f"Cuota {cuota.numero_cuota}",

        saldo_inicial=ctx.saldo_inicial,

        saldo_final=ctx.saldo_final,

        numero_cuota=cuota.numero_cuota,

        fecha_pago_display=fpd,

        estado_cuota=estado_cuota_lbl,

    )

    return Response(

        content=pdf_bytes,

        media_type="application/pdf",

        headers={"Content-Disposition": f'inline; filename="recibo_{referencia}.pdf"'},

    )



def _obtener_cuotas_para_export(db: Session, prestamo_id: int, prestamo: Prestamo) -> list:

    """Obtiene cuotas del préstamo con datos formateados para exportación Excel/PDF."""

    cuotas = db.execute(

        select(Cuota).where(Cuota.prestamo_id == prestamo_id).order_by(Cuota.numero_cuota)

    ).scalars().all()

    sincronizar_columna_estado_cuotas(db, cuotas, commit=True)



    resultado = []

    hoy = hoy_negocio()

    for c in cuotas:

        saldo_inicial = float(c.saldo_capital_inicial) if c.saldo_capital_inicial is not None else 0

        saldo_final = float(c.saldo_capital_final) if c.saldo_capital_final is not None else 0

        monto_cuota = float(c.monto) if c.monto is not None else 0

        total_pagado = float(c.total_pagado or 0)

        monto_capital = max(0, saldo_inicial - saldo_final)

        monto_interes = max(0, monto_cuota - monto_capital)

        fv = c.fecha_vencimiento

        fv_date = fv.date() if fv and hasattr(fv, "date") else fv

        estado_mostrar = estado_cuota_para_mostrar(total_pagado, monto_cuota, fv_date, hoy)



        resultado.append({

            "numero_cuota": c.numero_cuota,

            "fecha_vencimiento": c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else "",

            "monto_capital": monto_capital,

            "monto_interes": monto_interes,

            "monto_cuota": monto_cuota,

            "saldo_capital_final": saldo_final,

            "estado": estado_mostrar,

            "estado_etiqueta": etiqueta_estado_cuota(estado_mostrar),

        })

    return resultado





def _generar_excel_amortizacion(cuotas: list, prestamo: Prestamo) -> bytes:

    """Genera archivo Excel con tabla de amortización del préstamo."""

    import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment

    

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Tabla de Amortización"

    

    # Encabezados

    headers = ["Cuota", "Fecha Vencimiento", "Capital", "Interés", "Total", "Saldo Pendiente", "Estado"]

    ws.append(headers)

    header_row = ws[1]

    for cell in header_row:

        cell.font = Font(bold=True)

        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    

    total_capital = 0

    total_interes = 0

    total_general = 0

    

    for c in cuotas:

        ws.append([

            c["numero_cuota"],

            c["fecha_vencimiento"],

            c["monto_capital"],

            c["monto_interes"],

            c["monto_cuota"],

            c["saldo_capital_final"],

            c.get("estado_etiqueta") or c.get("estado") or "-",

        ])

        total_capital += c["monto_capital"]

        total_interes += c["monto_interes"]

        total_general += c["monto_cuota"]

    

    # Fila de resumen

    ws.append([])

    ws.append(["RESUMEN", "", total_capital, total_interes, total_general, "", ""])

    

    # Formato moneda para columnas numéricas

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=6):

        for cell in row:

            cell.number_format = '$#,##0.00'

    

    # Ajustar anchos de columna

    ws.column_dimensions["A"].width = 8

    ws.column_dimensions["B"].width = 18

    ws.column_dimensions["C"].width = 12

    ws.column_dimensions["D"].width = 12

    ws.column_dimensions["E"].width = 12

    ws.column_dimensions["F"].width = 15

    ws.column_dimensions["G"].width = 15

    

    buf = io.BytesIO()

    wb.save(buf)

    return buf.getvalue()





def _generar_pdf_amortizacion(cuotas: list, prestamo: Prestamo) -> bytes:

    """Genera archivo PDF con tabla de amortización del préstamo."""

    from reportlab.lib.pagesizes import letter

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    from reportlab.lib.styles import getSampleStyleSheet

    

    buf = io.BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=letter)

    styles = getSampleStyleSheet()

    story = []

    

    story.append(Paragraph("Tabla de Amortización", styles["Title"]))

    story.append(Paragraph(f"Cliente: {prestamo.nombres}", styles["Normal"]))

    story.append(Paragraph(f"Cédula: {prestamo.cedula}", styles["Normal"]))

    story.append(Paragraph(f"Préstamo #{prestamo.id}", styles["Normal"]))

    story.append(Paragraph(f"Total: ${float(prestamo.total_financiamiento or 0):,.2f}", styles["Normal"]))

    story.append(Spacer(1, 12))

    

    if not cuotas:

        story.append(Paragraph("No hay cuotas para este préstamo.", styles["Normal"]))

    else:

        rows = [["Cuota", "Fecha Venc.", "Capital", "Interés", "Total", "Saldo Pend.", "Estado"]]

        for c in cuotas:

            rows.append([

                str(c["numero_cuota"]),

                c["fecha_vencimiento"],

                f"${c['monto_capital']:,.2f}",

                f"${c['monto_interes']:,.2f}",

                f"${c['monto_cuota']:,.2f}",

                f"${c['saldo_capital_final']:,.2f}",

                c.get("estado_etiqueta") or c.get("estado") or "-",

            ])

        t = Table(rows)

        t.setStyle(TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), "#3B82F6"),

            ("TEXTCOLOR", (0, 0), (-1, 0), "white"),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("FONTSIZE", (0, 0), (-1, 0), 10),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            ("ALIGN", (2, 0), (5, -1), "RIGHT"),

            ("GRID", (0, 0), (-1, -1), 0.5, "#ccc"),

            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [0xffffff, 0xf5f7fa]),

        ]))

        story.append(t)

    

    doc.build(story)

    return buf.getvalue()





@router.get("/{prestamo_id}/amortizacion/excel")

def exportar_amortizacion_excel(prestamo_id: int, db: Session = Depends(get_db)):

    """Exporta la tabla de amortización del préstamo en formato Excel."""

    prestamo = db.get(Prestamo, prestamo_id)

    if not prestamo:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    

    cuotas = _obtener_cuotas_para_export(db, prestamo_id, prestamo)

    content = _generar_excel_amortizacion(cuotas, prestamo)

    filename = f"Tabla_Amortizacion_{prestamo.cedula}_{prestamo.id}.xlsx"

    

    return Response(

        content=content,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename={filename}"},

    )





@router.get("/{prestamo_id}/amortizacion/pdf")

def exportar_amortizacion_pdf(prestamo_id: int, db: Session = Depends(get_db)):

    """Exporta la tabla de amortización del préstamo en formato PDF."""

    prestamo = db.get(Prestamo, prestamo_id)

    if not prestamo:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    

    cuotas = _obtener_cuotas_para_export(db, prestamo_id, prestamo)

    content = _generar_pdf_amortizacion(cuotas, prestamo)

    filename = f"Tabla_Amortizacion_{prestamo.cedula}_{prestamo.id}.pdf"

    

    return Response(

        content=content,

        media_type="application/pdf",

        headers={"Content-Disposition": f"inline; filename={filename}"},

    )





@router.post("/{prestamo_id}/generar-amortizacion", response_model=dict)

def generar_amortizacion(prestamo_id: int, db: Session = Depends(get_db)):

    """Genera la tabla de amortización (cuotas) para el préstamo. No crea si ya existen cuotas."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    existentes = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == prestamo_id)) or 0

    if existentes > 0:

        return {"message": "Ya existe tabla de amortización.", "cuotas": existentes, "creadas": 0}

    numero_cuotas = p.numero_cuotas or 12

    total = float(p.total_financiamiento or 0)

    if numero_cuotas <= 0 or total <= 0:

        raise HTTPException(status_code=400, detail="Préstamo sin número de cuotas o monto válido.")

    monto_cuota = _resolver_monto_cuota(p, total, numero_cuotas)  # [C1] usa amortización francesa si tasa > 0

    # Regla unica: amortizacion con fecha_base_calculo (alineada con fecha de aprobacion en formularios).

    fecha_base = _fecha_para_amortizacion(p)

    if not fecha_base:

        raise HTTPException(

            status_code=400,

            detail="El préstamo debe tener fecha base de cálculo (o fecha de aprobación en datos antiguos) para generar la tabla de amortización.",

        )

    creadas = _generar_cuotas_amortizacion(db, p, fecha_base, numero_cuotas, monto_cuota)

    aplicar_pagos_pendientes_prestamo(p.id, db)

    db.commit()

    return {"message": "Tabla de amortización generada.", "cuotas": creadas, "creadas": creadas}





@router.post("/{prestamo_id}/recalcular-fechas-amortizacion", response_model=dict)
def recalcular_fechas_amortizacion(
    prestamo_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Recalcula SOLO las fechas de vencimiento de las cuotas existentes cuando cambia fecha_base_calculo
    (misma fecha calendario que fecha de aprobacion en formularios).
    
    Mantiene:
    - Montos de cuota iguales
    - Pagos asociados a cada cuota
    - Saldos de capital iniciales y finales
    
    Recalcula:
    - Fechas de vencimiento desde la nueva fecha base
    - Estados de cuota (VENCIDO, PENDIENTE, MOROSO, etc.) según nueva fecha de vencimiento
    
    Debe llamarse después de alinear fecha_aprobacion y fecha_base_calculo en el prestamo.
    """
    p = db.get(Prestamo, prestamo_id)
    
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    
    fecha_base = _fecha_para_amortizacion(p)
    
    if not fecha_base:
        raise HTTPException(
            status_code=400,
            detail="El préstamo debe tener fecha base de cálculo (o fecha de aprobación en datos antiguos) para recalcular fechas de vencimiento."
        )
    
    # Verificar que existen cuotas
    existentes = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == prestamo_id)) or 0
    
    if existentes == 0:
        raise HTTPException(
            status_code=400,
            detail="El préstamo no tiene cuotas para recalcular."
        )
    
    # Recalcular fechas y estados
    resultado = _recalcular_fechas_vencimiento_cuotas(db, p, fecha_base)

    _fallback_uid = db.execute(
        text("SELECT id FROM public.usuarios ORDER BY id LIMIT 1")
    ).scalar() or 1
    audit_recalc = Auditoria(
        usuario_id=_fallback_uid,
        accion="RECALCULO_FECHAS_AMORTIZACION",
        entidad="prestamos",
        entidad_id=prestamo_id,
        detalles=(
            f"Recalculo manual de fechas de cuotas (endpoint). "
            f"Fecha base: {fecha_base}. "
            f"Cuotas actualizadas: {resultado.get('actualizadas', 0)}"
        ),
        exito=True,
    )
    db.add(audit_recalc)
    db.commit()

    return resultado


@router.post("/{prestamo_id}/aplicar-condiciones-aprobacion", response_model=PrestamoResponse)

def aplicar_condiciones_aprobacion(prestamo_id: int, payload: AplicarCondicionesBody, db: Session = Depends(get_db)):

    """Aplica condiciones de aprobación: actualiza préstamo y opcionalmente genera cuotas."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    # [A3] Solo se pueden aplicar condiciones en estados permitidos

    _ESTADOS_APROBABLES = ("DRAFT", "EN_REVISION", "APROBADO", "EVALUADO")

    if p.estado not in _ESTADOS_APROBABLES:

        raise HTTPException(

            status_code=400,

            detail=f"No se pueden aplicar condiciones a un préstamo en estado '{p.estado}'. "

                   f"Estados permitidos: {', '.join(_ESTADOS_APROBABLES)}.",

        )

    if payload.tasa_interes is not None:

        p.tasa_interes = Decimal(str(payload.tasa_interes))

    if payload.plazo_maximo is not None:

        p.numero_cuotas = payload.plazo_maximo

    if payload.observaciones is not None:

        p.observaciones = payload.observaciones

    p.estado = "APROBADO"

    # fecha_aprobacion: solo payload explicito o la ya guardada; nunca base, requerimiento, registro ni hoy.
    if payload.fecha_aprobacion is not None:

        fecha_calendario = payload.fecha_aprobacion

        fecha_req = _fecha_requerimiento_date(p)

        if fecha_req and fecha_calendario < fecha_req:

            raise HTTPException(

                status_code=400,

                detail=f"La fecha de aprobación ({fecha_calendario}) debe ser igual o posterior a la fecha de requerimiento ({fecha_req}).",

            )

        p.fecha_base_calculo = fecha_calendario

        p.fecha_aprobacion = datetime.combine(fecha_calendario, datetime.min.time())

    elif p.fecha_aprobacion is None:

        raise HTTPException(

            status_code=400,

            detail=(

                "Debe indicar la fecha de aprobación / desembolso en la solicitud. "

                "No se infiere de la fecha de cálculo, ni de la fecha de registro, ni de la fecha actual."

            ),

        )

    validar_cupo_nuevo_prestamo_aprobado(db, p.cedula or "", exclude_prestamo_id=p.id)

    ensure_no_duplicate_aprobado_huella(db, p, exclude_prestamo_id=p.id)

    _registrar_en_revision_manual(db, prestamo_id)

    if prestamo_estado_exige_fecha_aprobacion(p.estado) and p.fecha_aprobacion is None:

        raise HTTPException(

            status_code=400,

            detail="Falta la fecha de aprobación. Los préstamos aprobados, desembolsados o liquidados deben tener fecha de aprobación.",

        )

    alinear_fecha_aprobacion_y_base_calculo(p)

    db.commit()

    # Generar cuotas si no existen (fecha_base_calculo alineada con fecha de aprobacion)

    existentes = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == prestamo_id)) or 0

    if existentes == 0:

        fecha_base = _fecha_para_amortizacion(p)

        if fecha_base:

            numero_cuotas = p.numero_cuotas or 12

            total = float(p.total_financiamiento or 0)

            monto_cuota = _resolver_monto_cuota(p, total, numero_cuotas)  # [C1] usa amortización francesa si tasa > 0

            _generar_cuotas_amortizacion(db, p, fecha_base, numero_cuotas, monto_cuota)

            aplicar_pagos_pendientes_prestamo(p.id, db)

            db.commit()

    db.refresh(p)

    return PrestamoResponse.model_validate(p)





@router.post("/{prestamo_id}/evaluar-riesgo", response_model=PrestamoResponse)

def evaluar_riesgo(prestamo_id: int, payload: EvaluarRiesgoBody, db: Session = Depends(get_db)):

    """Registra evaluación de riesgo manual (ML) y opcionalmente estado/requiere_revision."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    if payload.ml_impago_nivel_riesgo_manual is not None:

        p.ml_impago_nivel_riesgo_manual = payload.ml_impago_nivel_riesgo_manual

    if payload.ml_impago_probabilidad_manual is not None:

        p.ml_impago_probabilidad_manual = Decimal(str(payload.ml_impago_probabilidad_manual))

    if payload.requiere_revision is not None:

        p.requiere_revision = payload.requiere_revision

    if payload.estado is not None:

        p.estado = payload.estado.strip().upper()

    db.commit()

    db.refresh(p)

    return PrestamoResponse.model_validate(p)





@router.patch("/{prestamo_id}/marcar-revision", response_model=PrestamoResponse)

def marcar_revision(

    prestamo_id: int,

    requiere_revision: bool = Query(..., alias="requiere_revision"),

    db: Session = Depends(get_db),

):

    """Marca o desmarca el préstamo como requiere_revision."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    p.requiere_revision = requiere_revision

    db.commit()

    db.refresh(p)

    return PrestamoResponse.model_validate(p)





@router.post("/{prestamo_id}/asignar-fecha-aprobacion", response_model=dict)

def asignar_fecha_aprobacion(prestamo_id: int, payload: AsignarFechaAprobacionBody, db: Session = Depends(get_db)):

    """Asigna fecha de aprobación/desembolso (misma fecha), mantiene estado APROBADO y genera cuotas si no existen.

    Útil si el préstamo quedó en APROBADO sin fecha; en flujo normal, aprobar-manual también deja estado APROBADO."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    # [A3] Solo se puede asignar fecha a un préstamo en estados previos a rechazo

    _ESTADOS_PERMITIDOS = ("DRAFT", "EN_REVISION", "APROBADO", "EVALUADO")

    if p.estado not in _ESTADOS_PERMITIDOS:

        raise HTTPException(

            status_code=400,

            detail=f"No se puede asignar fecha de aprobación a un préstamo en estado '{p.estado}'. "

                   f"Estados permitidos: {', '.join(_ESTADOS_PERMITIDOS)}.",

        )

    # Coherencia: la fecha de aprobación debe ser >= fecha de requerimiento

    fecha_ap_date = payload.fecha_aprobacion if isinstance(payload.fecha_aprobacion, date) else (payload.fecha_aprobacion.date() if hasattr(payload.fecha_aprobacion, "date") and callable(getattr(payload.fecha_aprobacion, "date", None)) else None)

    fecha_req = _fecha_requerimiento_date(p)

    if fecha_req and fecha_ap_date and fecha_ap_date < fecha_req:

        raise HTTPException(

            status_code=400,

            detail=f"La fecha de aprobación ({fecha_ap_date}) debe ser igual o posterior a la fecha de requerimiento ({fecha_req}).",

        )

    p.fecha_aprobacion = datetime.combine(payload.fecha_aprobacion, datetime.min.time()) if isinstance(payload.fecha_aprobacion, date) else payload.fecha_aprobacion

    # fecha_base_calculo siempre igual a fecha_aprobacion
    p.fecha_base_calculo = fecha_ap_date

    p.estado = "APROBADO"

    validar_cupo_nuevo_prestamo_aprobado(db, p.cedula or "", exclude_prestamo_id=p.id)

    ensure_no_duplicate_aprobado_huella(db, p, exclude_prestamo_id=p.id)

    _registrar_en_revision_manual(db, prestamo_id)

    existentes = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == prestamo_id)) or 0

    cuotas_recalculadas = 0

    if existentes == 0:

        numero_cuotas = p.numero_cuotas or 12

        total = float(p.total_financiamiento or 0)

        monto_cuota = _resolver_monto_cuota(p, total, numero_cuotas)  # [C1] usa amortización francesa si tasa > 0

        fecha_base = fecha_ap_date

        cuotas_recalculadas = _generar_cuotas_amortizacion(db, p, fecha_base, numero_cuotas, monto_cuota)

        aplicar_pagos_pendientes_prestamo(p.id, db)

    db.commit()

    db.refresh(p)

    return {"prestamo": PrestamoResponse.model_validate(p), "cuotas_recalculadas": cuotas_recalculadas}





@router.post("/{prestamo_id}/aprobar-manual", response_model=dict)

def aprobar_manual(

    prestamo_id: int,

    payload: AprobarManualBody,

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Aprobación manual de riesgo: una fecha (aprobación = desembolso, misma fecha). Confirmación de documentos

    y declaración de políticas. Actualiza datos editables del préstamo, genera tabla de amortización

    y registra en auditoría. Solo préstamos en DRAFT o EN_REVISION. Al aprobar queda estado APROBADO

    con fecha_aprobación = fecha de aprobación/desembolso.

    """

    if (getattr(current_user, "rol", None) or "").lower() != "admin":

        raise HTTPException(

            status_code=403,

            detail="Solo administración puede aprobar préstamos (aprobación manual de riesgo).",

        )

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    if p.estado not in ("DRAFT", "EN_REVISION"):

        raise HTTPException(

            status_code=400,

            detail=f"Solo se puede aprobar manualmente un préstamo en DRAFT o EN_REVISION. Estado actual: {p.estado}",

        )

    if not payload.acepta_declaracion:

        raise HTTPException(status_code=400, detail="Debe aceptar la declaración de políticas y riesgo.")

    if not payload.documentos_analizados:

        raise HTTPException(status_code=400, detail="Debe confirmar que se analizaron los documentos del cliente.")



    fecha_ap = payload.fecha_aprobacion

    if hasattr(fecha_ap, "date"):

        fecha_ap = fecha_ap.date() if callable(getattr(fecha_ap, "date", None)) else fecha_ap



    # Coherencia: la fecha de aprobación debe ser >= fecha de requerimiento

    fecha_req = _fecha_requerimiento_date(p)

    if fecha_req and fecha_ap and fecha_ap < fecha_req:

        raise HTTPException(

            status_code=400,

            detail=f"La fecha de aprobación ({fecha_ap}) debe ser igual o posterior a la fecha de requerimiento ({fecha_req}).",

        )



    try:

        if payload.total_financiamiento is not None:

            p.total_financiamiento = Decimal(str(payload.total_financiamiento))

        if payload.numero_cuotas is not None:

            p.numero_cuotas = payload.numero_cuotas

        if payload.modalidad_pago is not None:

            p.modalidad_pago = payload.modalidad_pago.strip().upper() or p.modalidad_pago

        if payload.cuota_periodo is not None:

            p.cuota_periodo = Decimal(str(payload.cuota_periodo))

        if payload.tasa_interes is not None:

            p.tasa_interes = Decimal(str(payload.tasa_interes))

        if payload.observaciones is not None:

            p.observaciones = payload.observaciones



        p.fecha_aprobacion = datetime.combine(fecha_ap, datetime.min.time())

        p.fecha_base_calculo = fecha_ap

        p.usuario_aprobador = current_user.email

        p.estado = "APROBADO"

        validar_cupo_nuevo_prestamo_aprobado(db, p.cedula or "", exclude_prestamo_id=p.id)

        ensure_no_duplicate_aprobado_huella(db, p, exclude_prestamo_id=p.id)



        db.execute(delete(Cuota).where(Cuota.prestamo_id == prestamo_id))

        numero_cuotas = p.numero_cuotas or 12

        total = float(p.total_financiamiento or 0)

        if numero_cuotas <= 0 or total <= 0:

            raise HTTPException(status_code=400, detail="Número de cuotas o monto de financiamiento inválido.")

        monto_cuota = _resolver_monto_cuota(p, total, numero_cuotas)  # [C1] usa amortización francesa si tasa > 0

        creadas = _generar_cuotas_amortizacion(db, p, fecha_ap, numero_cuotas, monto_cuota)

        aplicar_pagos_pendientes_prestamo(prestamo_id, db)



        audit = Auditoria(

            usuario_id=_audit_user_id(db, current_user),

            accion="APROBACION_MANUAL",

            entidad="prestamos",

            entidad_id=prestamo_id,

            detalles=f"Aprobación manual de riesgo. Préstamo {prestamo_id}. Fecha aprobación: {fecha_ap}. Usuario: {current_user.email}.",

            exito=True,

        )

        db.add(audit)

        _registrar_en_revision_manual(db, prestamo_id)

        db.commit()

        db.refresh(p)

        return {"prestamo": PrestamoResponse.model_validate(p), "cuotas_generadas": creadas}

    except HTTPException:

        db.rollback()

        raise

    except IntegrityError as e:

        db.rollback()

        logger.exception("aprobar_manual prestamo_id=%s IntegrityError: %s", prestamo_id, e)

        err_msg = str(e.orig) if getattr(e, "orig", None) else str(e)

        if "auditoria" in err_msg.lower() and ("usuario_id" in err_msg.lower() or "users" in err_msg.lower()):

            raise HTTPException(

                status_code=500,

                detail=(

                    "Error de integridad en auditoría (usuario_id). "

                    "Ejecuta la migración backend/sql/migracion_auditoria_fk_usuarios.sql en la BD."

                ),

            ) from e

        raise HTTPException(

            status_code=500,

            detail=f"Error de integridad en BD al aprobar: {err_msg[:200]}",

        ) from e

    except Exception as e:

        db.rollback()

        logger.exception("aprobar_manual prestamo_id=%s: %s", prestamo_id, e)

        raise HTTPException(

            status_code=500,

            detail="Error al aprobar el préstamo. Revisa los logs del servidor.",

        ) from e





@router.post("/{prestamo_id}/rechazar", response_model=PrestamoResponse)

def rechazar_prestamo(

    prestamo_id: int,

    payload: dict,  # {"motivo_rechazo": "string"}

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Rechaza un préstamo en estados tempranos (DRAFT, EN_REVISION, EVALUADO).

    Solo usuario con rol 'administrador' puede rechazar.

    Registra motivo en auditoría.

    """

    if (getattr(current_user, "rol", None) or "").lower() != "admin":

        raise HTTPException(

            status_code=403,

            detail="Solo administración puede rechazar préstamos.",

        )

    

    prestamo = db.get(Prestamo, prestamo_id)

    if not prestamo:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    

    # Solo puedo rechazar en estados tempranos

    if prestamo.estado not in ("DRAFT", "EN_REVISION", "EVALUADO"):

        raise HTTPException(

            status_code=400,

            detail=f"No se puede rechazar préstamo en estado {prestamo.estado}. Solo DRAFT, EN_REVISION, EVALUADO.",

        )

    

    motivo = (payload.get("motivo_rechazo") or "Sin motivo especificado").strip()[:500]

    

    try:

        prestamo.estado = "RECHAZADO"

        prestamo.usuario_aprobador = current_user.email

        # Rechazo: no es aprobacion; no se usa fecha del sistema como fecha de aprobacion ni base.
        prestamo.fecha_aprobacion = None

        prestamo.fecha_base_calculo = None

        prestamo.observaciones = f"[RECHAZADO] {motivo}" if prestamo.observaciones else f"Rechazado: {motivo}"

        

        # Auditar rechazo

        audit = Auditoria(

            usuario_id=_audit_user_id(db, current_user),

            accion="RECHAZO_PRESTAMO",

            entidad="prestamos",

            entidad_id=prestamo_id,

            detalles=f"Rechazo de préstamo. Motivo: {motivo}. Usuario: {current_user.email}.",

            exito=True,

        )

        db.add(audit)

        db.commit()

        db.refresh(prestamo)

        return PrestamoResponse.model_validate(prestamo)

    except Exception as e:

        db.rollback()

        logger.exception("Error rechazar prestamo_id=%s: %s", prestamo_id, e)

        raise HTTPException(

            status_code=500,

            detail="Error al rechazar el préstamo.",

        ) from e





@router.get("/{prestamo_id}/evaluacion-riesgo", response_model=dict)

def get_evaluacion_riesgo(prestamo_id: int, db: Session = Depends(get_db)):

    """Devuelve datos de evaluación de riesgo (ML manual/calculado) del préstamo."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    return {

        "ml_impago_nivel_riesgo_manual": getattr(p, "ml_impago_nivel_riesgo_manual", None),

        "ml_impago_probabilidad_manual": float(p.ml_impago_probabilidad_manual) if getattr(p, "ml_impago_probabilidad_manual", None) is not None else None,

        "ml_impago_nivel_riesgo_calculado": getattr(p, "ml_impago_nivel_riesgo_calculado", None),

        "ml_impago_probabilidad_calculada": float(p.ml_impago_probabilidad_calculada) if getattr(p, "ml_impago_probabilidad_calculada", None) is not None else None,

        "requiere_revision": getattr(p, "requiere_revision", False),

    }





@router.get("/{prestamo_id}/auditoria", response_model=list)

def get_auditoria_prestamo(prestamo_id: int, db: Session = Depends(get_db)):

    """Lista registros de auditoría asociados al préstamo (entidad=prestamos, entidad_id=prestamo_id)."""

    p = db.get(Prestamo, prestamo_id)

    if not p:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    q = select(Auditoria).where(Auditoria.entidad == "prestamos", Auditoria.entidad_id == prestamo_id).order_by(Auditoria.fecha.desc())

    registros = db.execute(q).scalars().all()

    return [

        {

            "id": r.id,

            "usuario_id": r.usuario_id,

            "accion": r.accion,

            "modulo": r.entidad,

            "descripcion": r.detalles,

            "fecha": r.fecha.isoformat() + "Z" if r.fecha else "",

        }

        for r in registros

    ]





@router.post("/generar-cuotas-aprobados-sin-cuotas", response_model=dict)

def generar_cuotas_aprobados_sin_cuotas(

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Genera cuotas para todos los préstamos en estado APROBADO que no tienen cuotas.

    Regla unica: se usa fecha_base_calculo para el calculo; se omiten prestamos sin fecha base ni fecha de aprobacion (legacy).

    Solo administrador. Útil para regularizar datos legacy o préstamos que quedaron sin tabla de amortización.

    """

    if (getattr(current_user, "rol", None) or "").lower() != "admin":

        raise HTTPException(

            status_code=403,

            detail="Solo administración puede ejecutar generación masiva de cuotas.",

        )

    # Préstamos APROBADO sin ninguna cuota

    q = select(Prestamo).where(

        Prestamo.estado == "APROBADO",

        ~exists(select(1).select_from(Cuota).where(Cuota.prestamo_id == Prestamo.id)),

    )

    rows = db.execute(q).scalars().all()

    prestamos_sin_cuotas = list(rows) if rows else []

    if not prestamos_sin_cuotas:

        return {

            "procesados": 0,

            "cuotas_creadas": 0,

            "errores": [],

            "mensaje": "No hay préstamos APROBADO sin cuotas.",

        }

    total_cuotas = 0

    errores = []

    for p in prestamos_sin_cuotas:

        try:

            fecha_base = _fecha_para_amortizacion(p)

            if not fecha_base:

                errores.append({"prestamo_id": p.id, "error": "Sin fecha base de cálculo ni fecha de aprobación; obligatorias para generar amortización."})

                continue

            numero_cuotas = p.numero_cuotas or 12

            total = float(p.total_financiamiento or 0)

            if numero_cuotas <= 0 or total <= 0:

                errores.append({"prestamo_id": p.id, "error": "numero_cuotas o total_financiamiento inválido"})

                continue

            monto_cuota = _resolver_monto_cuota(p, total, numero_cuotas)

            creadas = _generar_cuotas_amortizacion(db, p, fecha_base, numero_cuotas, monto_cuota)

            aplicar_pagos_pendientes_prestamo(p.id, db)

            db.commit()

            total_cuotas += creadas

            logger.info("Préstamo %s: %s cuotas generadas (fecha_base=%s)", p.id, creadas, fecha_base)

        except Exception as e:

            db.rollback()

            logger.exception("Error generando cuotas para préstamo %s: %s", p.id, e)

            errores.append({"prestamo_id": p.id, "error": str(e)})

    return {

        "procesados": len(prestamos_sin_cuotas) - len(errores),

        "cuotas_creadas": total_cuotas,

        "errores": errores,

        "mensaje": f"Procesados {len(prestamos_sin_cuotas) - len(errores)} préstamos, {total_cuotas} cuotas creadas."

        + (f" {len(errores)} error(es)." if errores else ""),

    }







class ReaplicarCascadaMasivaBody(BaseModel):
    """Lista de prestamo_id a reaplicar en cascada (reset cuota_pagos + aplicar de nuevo). Maximo 500."""

    prestamo_ids: List[int]


# Compat: nombre historico del body (OpenAPI / clientes antiguos).
ReaplicarFifoMasivaBody = ReaplicarCascadaMasivaBody




@router.get("/{prestamo_id}/integridad-cuotas", response_model=dict)
def get_integridad_cuotas_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Diagnostico: total_pagado vs SUM(cuota_pagos) por cuota (y diff global).
    Solo administrador. No modifica datos.
    """
    if (getattr(current_user, "rol", None) or "").lower() != "admin":
        raise HTTPException(
            status_code=403,
            detail="Solo administracion puede consultar integridad de cuotas.",
        )
    r = integridad_cuotas_prestamo(db, prestamo_id)
    if not r.get("ok"):
        raise HTTPException(status_code=404, detail=r.get("error") or "Prestamo no encontrado")
    return r


@router.post("/{prestamo_id}/reaplicar-cascada-aplicacion", response_model=dict)
@router.post("/{prestamo_id}/reaplicar-fifo-aplicacion", response_model=dict)
def reaplicar_cascada_aplicacion_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Reaplicacion integral en cascada para un prestamo: borra cuota_pagos, resetea totales en cuotas
    y vuelve a aplicar todos los pagos conciliados en orden (fecha_pago, id).

    Ruta principal: POST `.../reaplicar-cascada-aplicacion`.
    Compatibilidad: POST `.../reaplicar-fifo-aplicacion` (mismo comportamiento).

    Usar cuando la tabla de amortizacion no refleja los pagos pese a filas en `pagos`
    (articulacion vieja, regeneracion de cuotas, o desalineacion total_pagado vs cuota_pagos).

    Solo administrador.
    """
    if (getattr(current_user, "rol", None) or "").lower() != "admin":
        raise HTTPException(
            status_code=403,
            detail="Solo administracion puede reaplicar la cascada sobre cuotas.",
        )
    try:
        integridad_antes = integridad_cuotas_prestamo(db, prestamo_id)
        r = reset_y_reaplicar_cascada_prestamo(db, prestamo_id)
        if not r.get("ok"):
            raise HTTPException(status_code=404, detail=r.get("error") or "No se pudo reaplicar")
        db.commit()
        db.expire_all()
        integridad_despues = integridad_cuotas_prestamo(db, prestamo_id)
        return {
            **r,
            "integridad_antes": integridad_antes,
            "integridad_despues": integridad_despues,
            "mensaje": "Cascada reaplicada: cuota_pagos reiniciado y pagos conciliados aplicados de nuevo.",
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("reaplicar-cascada-aplicacion prestamo_id=%s: %s", prestamo_id, e)
        raise HTTPException(status_code=500, detail=str(e))


# Compat: nombre de handler historico (importaciones / tests).
reaplicar_fifo_aplicacion_prestamo = reaplicar_cascada_aplicacion_prestamo


@router.post("/reaplicar-cascada-aplicacion-masiva", response_model=dict)
@router.post("/reaplicar-fifo-aplicacion-masiva", response_model=dict)
def reaplicar_cascada_aplicacion_masiva(
    body: ReaplicarCascadaMasivaBody,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """
    Igual que /{prestamo_id}/reaplicar-cascada-aplicacion pero para varios prestamos.
    Ruta principal: POST `.../reaplicar-cascada-aplicacion-masiva`.
    Compatibilidad: POST `.../reaplicar-fifo-aplicacion-masiva` (mismo comportamiento).
    Solo administrador. Maximo 500 IDs por solicitud.
    """
    if (getattr(current_user, "rol", None) or "").lower() != "admin":
        raise HTTPException(
            status_code=403,
            detail="Solo administracion puede reaplicar cascada masivo.",
        )
    ids = [int(x) for x in (body.prestamo_ids or []) if x is not None]
    if not ids:
        raise HTTPException(status_code=400, detail="prestamo_ids requerido")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="Maximo 500 prestamo_ids por solicitud")
    ok: list[dict] = []
    errores: list[dict] = []
    for pid in ids:
        try:
            r = reset_y_reaplicar_cascada_prestamo(db, pid)
            if r.get("ok"):
                ok.append(r)
                db.commit()
            else:
                db.rollback()
                errores.append({"prestamo_id": pid, "error": r.get("error") or "fallo"})
        except HTTPException as he:
            db.rollback()
            errores.append({"prestamo_id": pid, "error": he.detail})
        except Exception as e:
            db.rollback()
            logger.exception("reaplicar-cascada-masiva prestamo_id=%s: %s", pid, e)
            errores.append({"prestamo_id": pid, "error": str(e)})
    return {
        "procesados": len(ids),
        "exitosos": len(ok),
        "resultados": ok,
        "errores": errores,
        "mensaje": f"Cascada masiva: {len(ok)} ok, {len(errores)} error(es).",
    }


# Compat: nombre de handler historico.
reaplicar_fifo_aplicacion_masiva = reaplicar_cascada_aplicacion_masiva


class ConciliarAmortizacionMasivaBody(BaseModel):

    """Opcional: lista de prestamo_id a conciliar. Si vacío o ausente, se procesan todos los que tengan pagos conciliados sin aplicar."""

    prestamo_ids: Optional[List[int]] = None





@router.post("/conciliar-amortizacion-masiva", response_model=dict)

def conciliar_amortizacion_masiva(

    body: ConciliarAmortizacionMasivaBody = Body(default=ConciliarAmortizacionMasivaBody()),

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """

    Aplica a cuotas los pagos conciliados que aún no tienen enlace en cuota_pagos

    (p. ej. tras regenerar la tabla de amortización). Por préstamo llama a aplicar_pagos_pendientes_prestamo.

    Si body.prestamo_ids viene vacío o null, se detectan automáticamente los préstamos con pagos pendientes de aplicar.

    """

    if body.prestamo_ids:

        ids = list(body.prestamo_ids)

    else:

        # Préstamos que tienen al menos un pago conciliado con monto y sin fila en cuota_pagos

        subq = select(CuotaPago.pago_id).where(CuotaPago.pago_id.isnot(None)).distinct()

        r = db.execute(

            select(Pago.prestamo_id)

            .where(

                Pago.prestamo_id.isnot(None),

                Pago.conciliado == True,

                Pago.monto_pagado > 0,

                ~Pago.id.in_(subq),

            )

            .distinct()

        ).scalars().all()

        ids = [row[0] for row in r if row[0] is not None]

    if not ids:

        return {

            "procesados": 0,

            "pagos_aplicados_total": 0,

            "errores": [],

            "mensaje": "No hay préstamos con pagos pendientes de aplicar a cuotas.",

        }

    pagos_aplicados_total = 0

    errores = []

    for prestamo_id in ids:

        try:

            n = aplicar_pagos_pendientes_prestamo(prestamo_id, db)

            if n > 0:

                db.commit()

                pagos_aplicados_total += n

        except Exception as e:

            db.rollback()

            logger.exception("conciliar-amortizacion-masiva prestamo_id=%s: %s", prestamo_id, e)

            errores.append({"prestamo_id": prestamo_id, "error": str(e)})

    return {

        "procesados": len(ids),

        "pagos_aplicados_total": pagos_aplicados_total,

        "errores": errores,

        "mensaje": f"Conciliación masiva: {len(ids)} préstamo(s), {pagos_aplicados_total} pago(s) aplicados a cuotas."

        + (f" {len(errores)} error(es)." if errores else ""),

    }





@router.post("", response_model=PrestamoResponse, status_code=201)

def create_prestamo(payload: PrestamoCreate, db: Session = Depends(get_db), current_user: UserResponse = Depends(get_current_user)):

    """Crea un préstamo en BD. Valida que cliente_id exista. cedula/nombres se toman del Cliente.

    Automáticamente genera las cuotas de amortización."""

    cliente = db.get(Cliente, payload.cliente_id)

    if not cliente:

        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Usar email del usuario actual (no hardcoded)

    usuario_proponente_email = current_user.email if current_user else "itmaster@rapicreditca.com"

    

    # Siempre crear como APROBADO (individual y carga masiva); cuotas se generan con fecha_aprobacion

    estado_inicial = "APROBADO"

    # fecha_registro: la pone la BD al insertar; no sustituye fecha_aprobacion ni fecha_base_calculo.
    fa_d = payload.fecha_aprobacion
    req_d = payload.fecha_requerimiento
    if req_d and fa_d < req_d:
        raise HTTPException(
            status_code=400,
            detail=f"La fecha de aprobación ({fa_d}) debe ser igual o posterior a la fecha de requerimiento ({req_d}).",
        )
    fecha_aprob = datetime.combine(fa_d, time.min)

    analista_nombre, analista_row_id = _resolver_analista_para_prestamo(
        db, payload.analista, payload.analista_id
    )

    row = Prestamo(

        cliente_id=payload.cliente_id,

        cedula=cliente.cedula or "",

        nombres=cliente.nombres or "",

        total_financiamiento=payload.total_financiamiento,

        fecha_requerimiento=payload.fecha_requerimiento,

        modalidad_pago=payload.modalidad_pago or "MENSUAL",

        numero_cuotas=payload.numero_cuotas or 12,

        cuota_periodo=payload.cuota_periodo or 0,

        producto=payload.producto or "Financiamiento",

        estado=estado_inicial,

        fecha_aprobacion=fecha_aprob,

        concesionario=payload.concesionario,

        modelo_vehiculo=payload.modelo,

        analista=analista_nombre or (payload.analista or ""),

        analista_id=analista_row_id,

        usuario_proponente=usuario_proponente_email,

    )

    try:

        asegurar_prestamo_alineado_con_cliente(
            db, row, cliente=cliente, estado_para_validar=estado_inicial
        )

    except PrestamoCedulaClienteError as e:

        raise HTTPException(status_code=400, detail=str(e)) from e

    if not payload.omitir_validacion_huella_duplicada:

        ensure_no_duplicate_aprobado_huella(db, row, exclude_prestamo_id=None)

    validar_cupo_nuevo_prestamo_aprobado(db, row.cedula or "", exclude_prestamo_id=None)

    alinear_fecha_aprobacion_y_base_calculo(row)

    db.add(row)

    db.flush()

    _ajustar_req_si_mayor_que_aprobacion(
        db,
        row,
        origen="POST_prestamos",
        usuario_id=_audit_user_id(db, current_user),
    )

    db.commit()

    db.refresh(row)

    # Generar cuotas con fecha_base_calculo (alineada con fecha de aprobacion). Creacion individual y carga masiva ya dejan APROBADO y fechas.

    numero_cuotas = payload.numero_cuotas or 12

    total_financiamiento = float(payload.total_financiamiento)

    monto_cuota = _resolver_monto_cuota(row, total_financiamiento, numero_cuotas)

    prestamo_id = row.id  # guardar antes del try para no acceder a row tras rollback



    try:

        fecha_base_cuotas = _fecha_para_amortizacion(row)

        if fecha_base_cuotas:

            cuotas_generadas = _generar_cuotas_amortizacion(db, row, fecha_base_cuotas, numero_cuotas, monto_cuota)

            aplicar_pagos_pendientes_prestamo(row.id, db)

            db.commit()

            logger.info(f"Préstamo {prestamo_id}: {cuotas_generadas} cuotas generadas automáticamente (fecha_base={fecha_base_cuotas})")

        # Si no hay fecha base ni aprobacion (ej. DRAFT): no generar cuotas; se generaran al aprobar con fecha.

    except Exception as e:

        db.rollback()

        logger.error("Error generando cuotas para préstamo %s: %s", prestamo_id, str(e))

        raise HTTPException(

            status_code=500,

            detail=f"Error al generar cuotas: {str(e)}"

        )



    _registrar_en_revision_manual(db, row.id)

    db.commit()

    return PrestamoResponse.model_validate(row)





@router.put("/{prestamo_id}", response_model=PrestamoResponse)

def update_prestamo(
    prestamo_id: int,
    payload: PrestamoUpdate,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):

    """Actualiza un préstamo en BD."""

    row = db.get(Prestamo, prestamo_id)

    if not row:

        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    # Debug logging
    logger.info(f"[update_prestamo] Inicio: prestamo_id={prestamo_id}")
    logger.info(f"[update_prestamo] Payload raw recibido: fecha_requerimiento={payload.fecha_requerimiento} (type={type(payload.fecha_requerimiento).__name__}), fecha_aprobacion={payload.fecha_aprobacion} (type={type(payload.fecha_aprobacion).__name__})")
    logger.info(f"[update_prestamo] BD antes: fecha_requerimiento={row.fecha_requerimiento} (type={type(row.fecha_requerimiento).__name__}), fecha_aprobacion={row.fecha_aprobacion} (type={type(row.fecha_aprobacion).__name__})")

    est_antes = (row.estado or "").strip().upper()
    fecha_base_amort_antes = _fecha_para_amortizacion(row)

    if payload.fecha_base_calculo is not None and payload.fecha_aprobacion is None:

        raise HTTPException(

            status_code=400,

            detail=(

                "No se puede actualizar solo fecha_base_calculo: indique fecha_aprobacion. "

                "La fecha de calculo se copia siempre de la fecha de aprobacion (no se inventa)."

            ),

        )

    if est_antes == "DESISTIMIENTO":

        raise HTTPException(

            status_code=400,

            detail="El prestamo esta en desistimiento; no admite modificaciones.",

        )

    if payload.cliente_id is not None:

        cliente = db.get(Cliente, payload.cliente_id)

        if not cliente:

            raise HTTPException(status_code=404, detail="Cliente no encontrado")

        row.cliente_id = payload.cliente_id

    if payload.total_financiamiento is not None:

        row.total_financiamiento = payload.total_financiamiento

    if payload.estado is not None:

        row.estado = payload.estado

    if payload.concesionario is not None:

        row.concesionario = payload.concesionario

    if payload.modelo is not None:

        row.modelo_vehiculo = payload.modelo

    if payload.analista_id is not None or payload.analista is not None:

        an_txt, an_id = _resolver_analista_para_prestamo(
            db,
            payload.analista if payload.analista is not None else row.analista,
            payload.analista_id,
        )

        row.analista = an_txt or row.analista

        row.analista_id = an_id

    if payload.modalidad_pago is not None:

        row.modalidad_pago = payload.modalidad_pago

    if payload.numero_cuotas is not None:

        if payload.numero_cuotas != row.numero_cuotas:

            bloqueo_plazo = prestamo_bloquea_nuevas_cuotas_o_cambio_plazo(db, row)

            if bloqueo_plazo:

                raise HTTPException(status_code=400, detail=bloqueo_plazo)

        row.numero_cuotas = payload.numero_cuotas

    if payload.fecha_requerimiento is not None:

        row.fecha_requerimiento = payload.fecha_requerimiento

    if payload.fecha_aprobacion is not None:

        row.fecha_aprobacion = payload.fecha_aprobacion

        fa_date = (
            payload.fecha_aprobacion.date()
            if hasattr(payload.fecha_aprobacion, "date")
            and callable(getattr(payload.fecha_aprobacion, "date", None))
            else payload.fecha_aprobacion
        )

        row.fecha_base_calculo = fa_date

    if payload.cuota_periodo is not None:

        row.cuota_periodo = payload.cuota_periodo

    if payload.producto is not None:

        row.producto = payload.producto

    if payload.valor_activo is not None:

        row.valor_activo = payload.valor_activo

    if payload.observaciones is not None:

        row.observaciones = payload.observaciones

    if payload.tasa_interes is not None:

        row.tasa_interes = payload.tasa_interes

    # Reparar desalineaciones legacy: la fecha de aprobacion manda; base = misma fecha calendario.
    alinear_fecha_aprobacion_y_base_calculo(row)

    logger.info(f"[update_prestamo] BD después de aplicar cambios: fecha_requerimiento={row.fecha_requerimiento} (type={type(row.fecha_requerimiento).__name__}), fecha_aprobacion={row.fecha_aprobacion} (type={type(row.fecha_aprobacion).__name__})")

    _ajustar_req_si_mayor_que_aprobacion(
        db,
        row,
        origen="PUT_prestamos",
        usuario_id=_audit_user_id(db, current_user),
    )

    est_despues = (row.estado or "").strip().upper()

    if est_despues == "DESISTIMIENTO" and est_antes != "DESISTIMIENTO":

        if prestamos_tiene_columna_fecha_desistimiento(db):

            db.execute(

                text(

                    "UPDATE prestamos SET fecha_desistimiento = :fd "

                    "WHERE id = :pid AND fecha_desistimiento IS NULL"

                ),

                {"fd": hoy_negocio(), "pid": prestamo_id},

            )

        cli_des = db.get(Cliente, row.cliente_id)

        if cli_des is not None:

            cli_des.estado = "FINALIZADO"

    if prestamo_estado_exige_fecha_aprobacion(row.estado) and row.fecha_aprobacion is None:

        raise HTTPException(

            status_code=400,

            detail="Falta la fecha de aprobación. Los préstamos aprobados, desembolsados o liquidados deben tener fecha de aprobación.",

        )

    try:

        try:

            asegurar_prestamo_alineado_con_cliente(db, row)

        except PrestamoCedulaClienteError as e:

            raise HTTPException(status_code=400, detail=str(e)) from e

        if (row.estado or "").upper() == "APROBADO":

            validar_cupo_nuevo_prestamo_aprobado(db, row.cedula or "", exclude_prestamo_id=row.id)

            ensure_no_duplicate_aprobado_huella(db, row, exclude_prestamo_id=row.id)

        db.commit()

        db.refresh(row)

    except Exception:

        db.rollback()

        raise

    # APROBADO: sin cuotas, generar tabla; APROBADO o LIQUIDADO con cuotas: recalcular fechas si cambió la fecha base.

    _est_amort = (row.estado or "").strip().upper()

    if _est_amort in ("APROBADO", "LIQUIDADO"):

        existentes = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == prestamo_id)) or 0
        fecha_base = _fecha_para_amortizacion(row)

        if existentes == 0 and _est_amort == "APROBADO":

            if fecha_base:

                numero_cuotas = row.numero_cuotas or 12

                total_fin = float(row.total_financiamiento or 0)

                if numero_cuotas > 0 and total_fin > 0:

                    monto_cuota = _resolver_monto_cuota(row, total_fin, numero_cuotas)

                    _generar_cuotas_amortizacion(db, row, fecha_base, numero_cuotas, monto_cuota)

                    aplicar_pagos_pendientes_prestamo(prestamo_id, db)

                    try:

                        db.commit()

                        db.refresh(row)

                    except Exception:

                        db.rollback()

                        raise

        elif existentes > 0 and fecha_base and fecha_base != fecha_base_amort_antes:
            logger.info(
                f"[update_prestamo] fecha base amortizacion cambio de {fecha_base_amort_antes} a {fecha_base} — "
                f"recalculando fechas de vencimiento de {existentes} cuota(s)"
            )
            resultado_recalc = _recalcular_fechas_vencimiento_cuotas(db, row, fecha_base)
            _fallback_uid = db.execute(
                text("SELECT id FROM public.usuarios ORDER BY id LIMIT 1")
            ).scalar() or 1
            audit_recalc = Auditoria(
                usuario_id=_fallback_uid,
                accion="RECALCULO_FECHAS_AMORTIZACION",
                entidad="prestamos",
                entidad_id=prestamo_id,
                detalles=(
                    f"Recalculo automatico de fechas de cuotas. "
                    f"Fecha base anterior: {fecha_base_amort_antes}, "
                    f"nueva: {fecha_base}. "
                    f"Cuotas actualizadas: {resultado_recalc.get('actualizadas', 0)}"
                ),
                exito=True,
            )
            db.add(audit_recalc)
            db.commit()
            db.refresh(row)

    return PrestamoResponse.model_validate(row)





@router.delete("/{prestamo_id}", status_code=204)

def delete_prestamo(prestamo_id: int, db: Session = Depends(get_db)):

    """Elimina un prestamo y limpia filas hijas (cuotas, cuota_pagos, cache contable, etc.)."""

    row = db.get(Prestamo, prestamo_id)

    if not row:

        raise HTTPException(status_code=404, detail="Prestamo no encontrado")

    try:

        db.execute(
            delete(RevisionManualPrestamo).where(
                RevisionManualPrestamo.prestamo_id == prestamo_id
            )
        )

        db.execute(
            delete(AuditoriaCambiosEstadoPrestamo).where(
                AuditoriaCambiosEstadoPrestamo.prestamo_id == prestamo_id
            )
        )

        db.execute(
            delete(AuditoriaCarteraRevision).where(
                AuditoriaCarteraRevision.prestamo_id == prestamo_id
            )
        )

        db.execute(
            update(EnvioNotificacion)
            .where(EnvioNotificacion.prestamo_id == prestamo_id)
            .values(prestamo_id=None)
        )

        db.execute(
            update(DatosImportadosConErrores)
            .where(DatosImportadosConErrores.prestamo_id == prestamo_id)
            .values(prestamo_id=None)
        )

        # Al desvincular pagos, prestamo_id pasa a NULL; chk_pagos_prestamo_id_not_null exige
        # estado = ANULADO_IMPORT cuando no hay prestamo (mismo criterio que FK ON DELETE SET NULL).
        # Dos pasos en SQL crudo: evita PAGADO + prestamo_id NULL si el FK SET NULL actua sin estado.
        db.execute(
            text("UPDATE pagos SET estado = :estado WHERE prestamo_id = :pid"),
            {"estado": "ANULADO_IMPORT", "pid": prestamo_id},
        )
        db.execute(
            text("UPDATE pagos SET prestamo_id = NULL WHERE prestamo_id = :pid"),
            {"pid": prestamo_id},
        )

        db.execute(
            update(PagoConError)
            .where(PagoConError.prestamo_id == prestamo_id)
            .values(prestamo_id=None)
        )

        cuota_ids = list(
            db.scalars(
                select(Cuota.id).where(Cuota.prestamo_id == prestamo_id)
            ).all()
        )

        if cuota_ids:

            db.execute(delete(CuotaPago).where(CuotaPago.cuota_id.in_(cuota_ids)))

            db.execute(
                delete(AuditoriaConciliacionManual).where(
                    AuditoriaConciliacionManual.cuota_id.in_(cuota_ids)
                )
            )

            db.execute(
                delete(ReporteContableCache).where(
                    ReporteContableCache.cuota_id.in_(cuota_ids)
                )
            )

        db.execute(delete(Cuota).where(Cuota.prestamo_id == prestamo_id))

        db.flush()

        db.delete(row)

        db.commit()

    except IntegrityError as e:

        db.rollback()

        orig = getattr(e, "orig", None)

        detalle = str(orig) if orig is not None else str(e)

        logger.warning("delete_prestamo IntegrityError prestamo_id=%s: %s", prestamo_id, detalle)

        raise HTTPException(

            status_code=409,

            detail=f"No se puede eliminar el prestamo: restriccion en base de datos. {detalle}",

        ) from e

    except Exception as e:

        db.rollback()

        logger.exception("delete_prestamo prestamo_id=%s", prestamo_id)

        raise HTTPException(

            status_code=500,

            detail=f"Error al eliminar el prestamo: {e}",

        ) from e

    return None





# ============================================================================

# CARGA MASIVA DE PRÉSTAMOS DESDE EXCEL

# ============================================================================



from fastapi import UploadFile, File

from app.models.prestamo_con_error import PrestamoConError



try:

    import openpyxl

except ImportError:

    openpyxl = None





def _validate_modalidad_pago(modalidad: str) -> bool:

    """Valida que modalidad sea MENSUAL, QUINCENAL o SEMANAL."""

    if not modalidad:

        return False

    mod = str(modalidad).strip().upper()

    return mod in ("MENSUAL", "QUINCENAL", "SEMANAL")





def _parse_decimal(value: any) -> Decimal:

    """Parsea valor a Decimal."""

    if value is None:

        return Decimal("0.00")

    try:

        return Decimal(str(value))

    except:

        return None





@router.post("/upload-excel", response_model=dict)

async def upload_prestamos_excel(

    file: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user = Depends(get_current_user),

):

    """

    Carga masiva de préstamos desde Excel.

    Formato esperado: Cédula Cliente | Monto Financiamiento | Modalidad Pago | Nº Cuotas | Producto | Analista | Concesionario

    

    Validaciones:

    - Cédula: válida en BD (cliente existe)

    - Monto: > 0

    - Modalidad: MENSUAL|QUINCENAL|SEMANAL

    - Nº Cuotas: 1-12

    - Analista: requerido

    - Concesionario: opcional

    

    Respuesta: {registros_creados, registros_con_error}

    """

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):

        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")

    

    if not openpyxl:

        raise HTTPException(status_code=500, detail="Excel library not available")

    

    try:

        content = await file.read()

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        ws = wb.active

        

        if not ws:

            return {

                "registros_creados": 0,

                "registros_con_error": 0,

                "clientes_con_errores": []

            }

        

        registros_creados = 0

        registros_con_error = 0

        prestamos_con_errores = []

        

        # Pre-cargar clientes para validar

        clientes_cedulas = {}

        for cliente in db.execute(select(Cliente.id, Cliente.cedula)).all():

            clientes_cedulas[cliente.cedula.upper() if cliente.cedula else ""] = cliente.id

        

        usuario_email = current_user.email if hasattr(current_user, 'email') else "sistema@rapicredit.com"

        from openpyxl.utils.datetime import from_excel as _openpyxl_from_excel

        def _excel_cell_a_date_bulk(v) -> Optional[date]:

            if v is None:

                return None

            if isinstance(v, datetime):

                return v.date()

            if isinstance(v, date):

                return v

            if isinstance(v, (int, float)):

                try:

                    dt = _openpyxl_from_excel(v)

                    return dt.date() if hasattr(dt, "date") else None

                except Exception:

                    return None

            return None

        

        # Procesar filas (saltar header)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            if not row or all(cell is None for cell in row):

                continue

            

            try:

                cedula_raw = row[0] if len(row) > 0 else None

                monto_raw = row[1] if len(row) > 1 else None

                modalidad_raw = row[2] if len(row) > 2 else None

                cuotas_raw = row[3] if len(row) > 3 else None

                producto_raw = row[4] if len(row) > 4 else None

                analista_raw = row[5] if len(row) > 5 else None

                concesionario_raw = row[6] if len(row) > 6 else None

                fecha_req_raw = row[7] if len(row) > 7 else None

                fecha_aprob_raw = row[8] if len(row) > 8 else None

                

                errores = []

                

                # Validar cédula

                cedula = str(cedula_raw or "").strip().upper()

                if not cedula:

                    errores.append("Cédula es requerida")

                elif cedula not in clientes_cedulas:

                    errores.append(f"Cliente con cédula {cedula} no existe en BD")

                

                # Validar monto

                monto = _parse_decimal(monto_raw)

                if monto is None:

                    errores.append("Monto Financiamiento inválido")

                elif monto <= 0:

                    errores.append("Monto debe ser mayor a 0")

                

                # Validar modalidad

                modalidad = str(modalidad_raw or "").strip().upper()

                if not modalidad:

                    errores.append("Modalidad Pago es requerida (MENSUAL|QUINCENAL|SEMANAL)")

                elif not _validate_modalidad_pago(modalidad):

                    errores.append("Modalidad debe ser MENSUAL, QUINCENAL o SEMANAL")

                

                # Validar cuotas

                try:

                    cuotas = int(cuotas_raw) if cuotas_raw else None

                    if cuotas is None:

                        errores.append("Nº Cuotas es requerido")

                    elif cuotas < 1 or cuotas > 12:

                        errores.append("Nº Cuotas debe estar entre 1 y 12")

                except (ValueError, TypeError):

                    errores.append("Nº Cuotas debe ser número entero")

                

                # Validar producto

                producto = str(producto_raw or "").strip()

                if not producto:

                    errores.append("Producto es requerido")

                

                # Validar analista

                analista = str(analista_raw or "").strip()

                if not analista:

                    errores.append("Analista es requerido")

                

                # Concesionario es opcional

                concesionario = str(concesionario_raw or "").strip() if concesionario_raw else None

                fecha_requerimiento_x = _excel_cell_a_date_bulk(fecha_req_raw)

                fecha_aprobacion_x = _excel_cell_a_date_bulk(fecha_aprob_raw)

                if fecha_requerimiento_x is None:

                    errores.append("Fecha requerimiento es requerida (columna H, formato fecha)")

                if fecha_aprobacion_x is None:

                    errores.append("Fecha aprobacion/desembolso es requerida (columna I, formato fecha)")

                if (

                    fecha_requerimiento_x is not None

                    and fecha_aprobacion_x is not None

                    and fecha_aprobacion_x < fecha_requerimiento_x

                ):

                    errores.append("Fecha aprobacion debe ser >= fecha requerimiento")

                

                # Si hay errores, agregar a lista de revisión

                if errores:

                    prestamo_error = PrestamoConError(

                        cedula_cliente=cedula or None,

                        total_financiamiento=monto,

                        modalidad_pago=modalidad or None,

                        numero_cuotas=cuotas if 'cuotas' in locals() and isinstance(cuotas, int) else None,

                        producto=producto or None,

                        analista=analista or None,

                        concesionario=concesionario,

                        estado="PENDIENTE",

                        errores_descripcion="; ".join(errores),

                        fila_origen=idx,

                        usuario_registro=usuario_email

                    )

                    db.add(prestamo_error)

                    registros_con_error += 1

                    prestamos_con_errores.append({

                        "cedula": cedula,

                        "fila": idx,

                        "errores": "; ".join(errores)

                    })

                    continue

                

                # Crear préstamo

                cliente_id = clientes_cedulas[cedula]

                cliente = db.get(Cliente, cliente_id)

                try:

                    exigir_cliente_cedula_para_prestamo_aprobado(cliente, "APROBADO")

                    validar_cupo_nuevo_prestamo_aprobado(
                        db, cliente.cedula or "", exclude_prestamo_id=None
                    )

                except PrestamoCedulaClienteError as ced_exc:

                    ced_msg = str(ced_exc)

                    prestamo_error = PrestamoConError(

                        cedula_cliente=cedula,

                        total_financiamiento=monto,

                        modalidad_pago=modalidad,

                        numero_cuotas=cuotas,

                        producto=producto,

                        analista=analista,

                        concesionario=concesionario,

                        estado="PENDIENTE",

                        errores_descripcion=ced_msg,

                        fila_origen=idx,

                        usuario_registro=usuario_email,

                    )

                    db.add(prestamo_error)

                    registros_con_error += 1

                    prestamos_con_errores.append(

                        {"cedula": cedula, "fila": idx, "errores": ced_msg}

                    )

                    continue

                except HTTPException as cupo_exc:

                    cupo_msg = str(cupo_exc.detail)

                    prestamo_error = PrestamoConError(

                        cedula_cliente=cedula,

                        total_financiamiento=monto,

                        modalidad_pago=modalidad,

                        numero_cuotas=cuotas,

                        producto=producto,

                        analista=analista,

                        concesionario=concesionario,

                        estado="PENDIENTE",

                        errores_descripcion=cupo_msg,

                        fila_origen=idx,

                        usuario_registro=usuario_email,

                    )

                    db.add(prestamo_error)

                    registros_con_error += 1

                    prestamos_con_errores.append(

                        {"cedula": cedula, "fila": idx, "errores": cupo_msg}

                    )

                    continue

                

                # Carga masiva: fechas explicitas en Excel (columnas H e I); no se usa fecha_registro ni hoy.

                prestamo = Prestamo(

                    cliente_id=cliente_id,

                    cedula=cliente.cedula or "",

                    nombres=cliente.nombres or "",

                    total_financiamiento=monto,

                    fecha_requerimiento=fecha_requerimiento_x,

                    fecha_aprobacion=datetime.combine(fecha_aprobacion_x, time.min),

                    fecha_base_calculo=fecha_aprobacion_x,

                    modalidad_pago=modalidad,

                    numero_cuotas=cuotas,

                    cuota_periodo=Decimal("0.00"),

                    producto=producto,

                    estado="APROBADO",

                    concesionario=concesionario,

                    analista=analista,

                    usuario_proponente=usuario_email,

                )

                db.add(prestamo)

                db.flush()  # Obtener ID del préstamo

                monto_cuota = _resolver_monto_cuota(prestamo, float(monto), cuotas)

                _generar_cuotas_amortizacion(db, prestamo, fecha_aprobacion_x, cuotas, monto_cuota)

                aplicar_pagos_pendientes_prestamo(prestamo.id, db)

                # Registrar en revisión manual

                _registrar_en_revision_manual(db, prestamo.id)

                

                registros_creados += 1

                

            except Exception as e:

                logger.error(f"Error procesando fila {idx}: {e}")

                prestamo_error = PrestamoConError(

                    cedula_cliente=str(row[0]) if len(row) > 0 else None,

                    total_financiamiento=_parse_decimal(row[1]) if len(row) > 1 else None,

                    errores_descripcion=f"Error general: {str(e)[:200]}",

                    fila_origen=idx,

                    usuario_registro=usuario_email

                )

                db.add(prestamo_error)

                registros_con_error += 1

                prestamos_con_errores.append({

                    "cedula": str(row[0]) if len(row) > 0 else "?",

                    "fila": idx,

                    "errores": str(e)[:100]

                })

        

        db.commit()

        

        return {

            "registros_creados": registros_creados,

            "registros_con_error": registros_con_error,

            "mensaje": f"Se crearon {registros_creados} préstamo(s) y {registros_con_error} con error(es)",

            "prestamos_con_errores": prestamos_con_errores[:10]  # Mostrar primeros 10

        }

        

    except Exception as e:

        db.rollback()

        logger.exception("Error cargando Excel de préstamos: %s", e)

        raise HTTPException(

            status_code=500,

            detail=f"Error procesando archivo: {str(e)[:200]}"

        )





class RevisarPrestamoAgregarBody(BaseModel):

    """Datos de una fila para enviar a revisión manual (prestamos_con_errores)."""

    cedula_cliente: Optional[str] = None

    total_financiamiento: Optional[float] = None

    modalidad_pago: Optional[str] = None

    numero_cuotas: Optional[int] = None

    producto: Optional[str] = None

    analista: Optional[str] = None

    concesionario: Optional[str] = None

    errores_descripcion: Optional[str] = None

    fila_origen: Optional[int] = None





@router.post("/revisar/agregar", response_model=dict)

def agregar_prestamo_a_revisar(

    body: RevisarPrestamoAgregarBody,

    db: Session = Depends(get_db),

    current_user=Depends(get_current_user),

):

    """

    Envía una fila a la tabla prestamos_con_errores (revisar préstamos).

    Usado desde la carga masiva cuando el usuario pulsa "Enviar a Revisar Préstamos".

    """

    usuario_email = getattr(current_user, "email", None) or "sistema@rapicredit.com"

    row = PrestamoConError(

        cedula_cliente=body.cedula_cliente,

        total_financiamiento=Decimal(str(body.total_financiamiento)) if body.total_financiamiento is not None else None,

        modalidad_pago=body.modalidad_pago,

        numero_cuotas=body.numero_cuotas,

        producto=body.producto,

        analista=body.analista,

        concesionario=body.concesionario,

        estado="PENDIENTE",

        errores_descripcion=body.errores_descripcion or "Enviado a revisión desde carga masiva",

        fila_origen=body.fila_origen,

        usuario_registro=usuario_email,

    )

    db.add(row)

    db.commit()

    db.refresh(row)

    return {"id": row.id, "mensaje": "Préstamo enviado a Revisar Préstamos"}





@router.get("/revisar/lista", response_model=dict)

def get_prestamos_con_errores(

    page: int = Query(1, ge=1),

    per_page: int = Query(20, ge=1, le=100),

    db: Session = Depends(get_db),

):

    """

    Listado de préstamos con errores de validación (pendientes de revisión).

    Paginado para facilitar corrección manual.

    """

    total = db.scalar(select(func.count()).select_from(PrestamoConError)) or 0

    rows = db.execute(

        select(PrestamoConError)

        .offset((page - 1) * per_page)

        .limit(per_page)

    ).scalars().all()

    

    return {

        "total": total,

        "page": page,

        "per_page": per_page,

        "items": [

            {

                "id": r.id,

                "cedula_cliente": r.cedula_cliente,

                "total_financiamiento": str(r.total_financiamiento) if r.total_financiamiento else None,

                "modalidad_pago": r.modalidad_pago,

                "numero_cuotas": r.numero_cuotas,

                "producto": r.producto,

                "analista": r.analista,

                "concesionario": r.concesionario,

                "errores": r.errores_descripcion,

                "fila_origen": r.fila_origen,

                "estado": r.estado,

                "fecha_registro": r.fecha_registro.isoformat() if r.fecha_registro else None,

            }

            for r in rows

        ]

    }





class EliminarPorDescargaBody(BaseModel):

    """IDs de registros exportados a Excel para eliminar de prestamos_con_errores."""

    ids: list[int]





@router.post("/revisar/eliminar-por-descarga", response_model=dict)

def eliminar_prestamos_por_descarga(

    payload: EliminarPorDescargaBody = Body(...),

    db: Session = Depends(get_db),

):

    """

    Elimina de prestamos_con_errores los registros descargados (borrado en lote).

    Misma regla que Pagos: al descargar Excel se vacía la lista; se rellena al enviar desde Carga Masiva.

    """

    valid_ids = [i for i in payload.ids if isinstance(i, int) and i > 0]

    if not valid_ids:

        return {"eliminados": 0, "mensaje": "No hay IDs"}

    result = db.execute(delete(PrestamoConError).where(PrestamoConError.id.in_(valid_ids)))

    eliminados = result.rowcount

    db.commit()

    return {"eliminados": eliminados, "mensaje": f"{eliminados} eliminados de prestamos_con_errores"}





@router.delete("/revisar/{error_id}", status_code=204)

def resolver_prestamo_error(error_id: int, db: Session = Depends(get_db)):

    """Marcar préstamo con error como resuelto (eliminar de la lista)."""

    row = db.get(PrestamoConError, error_id)

    if not row:

        raise HTTPException(status_code=404, detail="Registro no encontrado")

    

    db.delete(row)

    db.commit()

    return None








@router.get("/{prestamo_id}/estado-cuenta")
def get_estado_cuenta_prestamo_json(
    prestamo_id: int,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Mismo payload que el PDF de estado de cuenta, en JSON serializable. Fuente: estado_cuenta_datos."""
    from app.services.estado_cuenta_datos import (
        obtener_datos_estado_cuenta_prestamo,
        serializar_estado_cuenta_payload_json,
    )

    datos = obtener_datos_estado_cuenta_prestamo(db, prestamo_id)
    if not datos:
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    return serializar_estado_cuenta_payload_json(datos)


@router.get("/{prestamo_id}/estado-cuenta/pdf")

def get_estado_cuenta_prestamo_pdf(

    prestamo_id: int,

    request: Request,

    db: Session = Depends(get_db),

    current_user: UserResponse = Depends(get_current_user),

):

    """Genera PDF de estado de cuenta para un prestamo especifico."""

    from app.services.estado_cuenta_datos import obtener_datos_estado_cuenta_prestamo
    from app.services.estado_cuenta_pdf import generar_pdf_estado_cuenta

    

    try:

        datos = obtener_datos_estado_cuenta_prestamo(db, prestamo_id)

    except Exception:

        logger.exception("Error cargando datos para PDF estado de cuenta prestamo_id=%s", prestamo_id)

        raise HTTPException(

            status_code=500,

            detail="Error al preparar los datos del estado de cuenta. Verifique migraciones de BD o contacte soporte.",

        )

    if not datos:

        raise HTTPException(status_code=404, detail="Prestamo no encontrado")

    

    try:

        cedula_display = datos.get("cedula_display") or ""

        nombre = datos.get("nombre") or ""

        prestamos_list = datos.get("prestamos_list") or []


        fecha_corte = datos.get("fecha_corte") or date.today()

        amortizaciones_por_prestamo = datos.get("amortizaciones_por_prestamo") or []

        

        from app.api.v1.endpoints.validadores import validate_cedula

        from app.core.security import create_recibo_token

        cedula_lookup = ""

        if cedula_display:

            vr = validate_cedula(cedula_display.strip())

            if vr.get("valido"):

                vf = (vr.get("valor_formateado") or "").replace("-", "")

                cedula_lookup = vf

        recibo_token = (

            create_recibo_token(cedula_lookup, expire_hours=2) if cedula_lookup else None

        )

        base_url = str(request.base_url).rstrip("/")

        pdf_bytes = generar_pdf_estado_cuenta(

            cedula=cedula_display,

            nombre=nombre,

            prestamos=prestamos_list,

            fecha_corte=fecha_corte,

            amortizaciones_por_prestamo=amortizaciones_por_prestamo,

            pagos_realizados=datos.get("pagos_realizados") or [],

            recibos=None,

            recibo_token=recibo_token,

            base_url=base_url,

        )

        

        if not pdf_bytes or len(pdf_bytes) < 8 or not pdf_bytes.startswith(b"%PDF"):

            logger.error(

                "PDF invalido generado para estado de cuenta prestamo_id=%s len=%s",

                prestamo_id,

                len(pdf_bytes) if pdf_bytes else 0,

            )

            raise HTTPException(

                status_code=500,

                detail="Error generando PDF del estado de cuenta (salida invalida).",

            )

        filename = f"Estado_Cuenta_Prestamo_{prestamo_id}.pdf"

        return Response(

            content=pdf_bytes,

            media_type="application/pdf",

            headers={

                "Content-Disposition": f'inline; filename="{filename}"',

                "X-Content-Type-Options": "nosniff",

            },

        )

    except Exception as e:

        logger.exception("Error generando PDF: prestamo_id=%s error=%s", prestamo_id, e)

        raise HTTPException(status_code=500, detail="Error generando PDF del estado de cuenta")

