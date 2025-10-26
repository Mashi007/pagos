import io
import logging
from datetime import datetime

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import Response
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_db
from app.models.pago import Pago
from app.models.user import User
from app.schemas.conciliacion import ConciliacionResponse

# Sistema de Conciliación Bancaria

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/template-conciliacion")
async def generar_template_conciliacion(
    current_user: User = Depends(get_current_user),
):
    # Generar template Excel para conciliación bancaria
    try:
        logger.info(
            f"Generando template de conciliación - Usuario: {current_user.email}"
        )

        # Crear workbook
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()

        # HOJA 1: INSTRUCCIONES
        ws_instrucciones = wb.active
        ws_instrucciones.title = "Instrucciones"

        instrucciones = [
            ["INSTRUCCIONES PARA CONCILIACIÓN BANCARIA"],
            [""],
            ["1. Esta plantilla debe completarse con los datos del banco"],
            ["2. La segunda hoja contiene los pagos del sistema"],
            ["3. Complete la columna 'CONCILIAR' con 'SI' o 'NO'"],
            ["4. Suba el archivo completado para procesar la conciliación"],
            [""],
            ["COLUMNAS REQUERIDAS:"],
            ["- FECHA: Fecha del movimiento bancario"],
            ["- MONTO: Monto del movimiento"],
            ["- CONCILIAR: SI/NO para indicar si conciliar"],
            ["- OBSERVACIONES: Comentarios adicionales"],
        ]

        for row, instruction in enumerate(instrucciones, 1):
            ws_instrucciones.cell(row=row, column=1, value=instruction[0])

        # HOJA 2: DATOS PARA CONCILIAR
        ws_datos = wb.create_sheet("Datos Conciliación")

        # Encabezados
        headers = ["FECHA", "MONTO", "CONCILIAR", "OBSERVACIONES"]
        for col, header in enumerate(headers, 1):
            ws_datos.cell(row=1, column=col, value=header)

        # Validación para columna CONCILIAR
        validation = DataValidation(type="list", formula1='"SI,NO"')
        validation.add("D2:D1000")
        ws_datos.add_data_validation(validation)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=template_conciliacion.xlsx"
            },
        )

    except Exception as e:
        logger.error(f"Error generando template de conciliación: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor: {str(e)}",
        )


def _validar_archivo_conciliacion(file: UploadFile) -> None:
    """Validar que el archivo sea Excel"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Archivo debe ser Excel (.xlsx o .xls)",
        )


def _validar_columnas_conciliacion(df: pd.DataFrame) -> None:
    """Validar que el DataFrame tenga las columnas requeridas"""
    required_columns = ["FECHA", "MONTO", "CONCILIAR"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas requeridas: {', '.join(missing_columns)}",
        )


def _procesar_fila_conciliacion(
    row: pd.Series, index: int, db: Session, current_user: User
) -> tuple[bool, str]:
    """Procesar una fila individual de conciliación"""
    try:
        if row["CONCILIAR"].upper() != "SI":
            return False, ""

        # Buscar pago por fecha y monto
        fecha = pd.to_datetime(row["FECHA"]).date()
        monto = float(row["MONTO"])

        pago = (
            db.query(Pago)
            .filter(
                and_(
                    Pago.fecha_pago == fecha,
                    Pago.monto == monto,
                    ~Pago.conciliado,
                )
            )
            .first()
        )

        if pago:
            pago.conciliado = True
            pago.fecha_conciliacion = datetime.now()
            pago.usuario_conciliacion = int(current_user.id)
            return True, ""
        else:
            return (
                False,
                f"Fila {index + 2}: No se encontró pago para fecha {fecha} y monto {monto}",
            )

    except Exception as e:
        return False, f"Fila {index + 2}: Error procesando - {str(e)}"


@router.post("/procesar-conciliacion")
async def procesar_conciliacion(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Procesar archivo Excel de conciliación bancaria
    try:
        logger.info(f"Procesando conciliación - Usuario: {current_user.email}")

        # Validar archivo
        _validar_archivo_conciliacion(file)

        # Leer archivo Excel
        file_content = await file.read()
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=1)  # Segunda hoja

        # Validar columnas
        _validar_columnas_conciliacion(df)

        # Procesar conciliaciones
        conciliaciones_procesadas = 0
        errores = []

        for index, row in df.iterrows():
            procesado, error = _procesar_fila_conciliacion(row, index, db, current_user)
            if procesado:
                conciliaciones_procesadas += 1
            elif error:
                errores.append(error)

        db.commit()

        return {
            "mensaje": "Conciliación procesada exitosamente",
            "conciliaciones_procesadas": conciliaciones_procesadas,
            "errores": errores,
            "total_errores": len(errores),
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error procesando conciliación: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor: {str(e)}",
        )


@router.post(
    "/desconciliar-pago",
    response_model=ConciliacionResponse,
    status_code=status.HTTP_201_CREATED,
)
async def desconciliar_pago(
    pago_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Desconciliar un pago ya conciliado
    try:
        logger.info(f"Desconciliando pago {pago_id} - Usuario: {current_user.email}")

        # Buscar el pago a desconciliar
        pago = db.query(Pago).filter(Pago.id == pago_id).first()

        if not pago:
            raise HTTPException(status_code=404, detail="Pago no encontrado")

        if not pago.conciliado:
            raise HTTPException(status_code=400, detail="El pago no está conciliado")

        # Desconciliar
        pago.conciliado = False
        pago.fecha_conciliacion = None
        pago.usuario_conciliacion = None

        db.commit()

        return ConciliacionResponse(
            pago_id=pago.id,
            conciliado=False,
            fecha_conciliacion=None,
            usuario_conciliacion=None,
            mensaje="Pago desconciliado exitosamente",
        )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error desconciliando pago: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor: {str(e)}",
        )


@router.get("/estado-conciliacion")
async def obtener_estado_conciliacion(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Obtener estado general de conciliación
    try:
        # Estadísticas generales
        total_pagos = db.query(Pago).filter(Pago.activo).count()
        pagos_conciliados = (
            db.query(Pago).filter(and_(Pago.activo, Pago.conciliado)).count()
        )

        # Porcentaje de conciliación
        porcentaje_conciliacion = (
            (pagos_conciliados / total_pagos * 100) if total_pagos > 0 else 0
        )

        return {
            "total_pagos": total_pagos,
            "pagos_conciliados": pagos_conciliados,
            "porcentaje_conciliacion": round(porcentaje_conciliacion, 2),
        }

    except Exception as e:
        logger.error(f"Error obteniendo estado de conciliación: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno del servidor: {str(e)}",
        )
