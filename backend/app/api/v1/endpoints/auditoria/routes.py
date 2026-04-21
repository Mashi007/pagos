"""
Endpoints de auditoria.

- Tabla `auditoria`: GET listado, GET stats, GET exportar, GET /{id}, POST /registrar.
- Cartera (prestamos/cartera/*): chequeos paginados, resumen sin items, meta persistida, ejecutar/corregir,
  POST sincronizar-estados-cuotas (solo alinea cuotas.estado, sin meta ni controles).
  Ver `docs/auditoria-api-cartera.md` para contrato y parametros (`solo_alertas` es historico y no filtra).
"""
import io
import json
from datetime import datetime, timedelta, timezone
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user, require_admin, require_auditoria_cartera_access
from app.core.rol_normalization import canonical_rol
from app.models.auditoria import Auditoria
from app.models.auditoria_cartera_revision import AuditoriaCarteraRevision
from app.models.prestamo import Prestamo
from app.models.user import User
from app.schemas.auth import UserResponse
from app.services.auditoria_cartera_revision_service import (
    CONTROLES_CARTERA_VALIDOS,
    NOTA_MIN_REVOCAR_OK,
    TIPOS_REVISION_VALIDOS,
    iso_utc,
    listar_codigos_control_con_excepcion_historica,
    listar_ocultos_marcar_ok,
    ultimo_tipo_revision_par,
)
from app.services.auditoria_cartera_revision_snapshot import (
    construir_payload_snapshot_marcar_ok,
    payload_minimo_revocar_ok,
)
from app.services.pago_control5_visto_service import (
    aplicar_visto_control5_duplicado_fecha_monto,
    listar_pagos_duplicados_fecha_monto_por_prestamo,
)
from app.services.auditoria_liquidados_intensiva import (
    filtrar_filas_cierre,
    hallazgos_cierre_prestamos_liquidados,
    paginar_filas,
    resumen_cierre_desde_filas,
)
from app.services.prestamo_cartera_auditoria import (
    ejecutar_auditoria_cartera,
    leer_meta_ejecucion,
    listar_pagos_sin_aplicacion_cuotas_por_prestamo,
    persistir_meta_ejecucion,
)

router = APIRouter(dependencies=[Depends(get_current_user)])


def _codigo_control_cartera_opcional(raw: Optional[str]) -> Optional[str]:
    if raw is None or not str(raw).strip():
        return None
    c = str(raw).strip()
    if c not in CONTROLES_CARTERA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="codigo_control no es un control de cartera conocido",
        )
    return c


# --- Schemas (compatibles con frontend) ---

class AuditoriaItem(BaseModel):
    id: int
    usuario_id: Optional[int] = None
    usuario_email: Optional[str] = None
    accion: str
    modulo: str
    tabla: str
    registro_id: Optional[int] = None
    descripcion: Optional[str] = None
    campo: Optional[str] = None
    datos_anteriores: Optional[Any] = None
    datos_nuevos: Optional[Any] = None
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    resultado: str
    mensaje_error: Optional[str] = None
    fecha: str

    class Config:
        from_attributes = True


class AuditoriaListResponse(BaseModel):
    items: List[AuditoriaItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class AuditoriaStats(BaseModel):
    total_acciones: int
    acciones_por_modulo: dict
    acciones_por_usuario: dict
    acciones_hoy: int
    acciones_esta_semana: int
    acciones_este_mes: int


class RegistrarAuditoriaBody(BaseModel):
    modulo: str
    accion: str
    descripcion: str
    registro_id: Optional[int] = None


def _row_to_item(r: Auditoria) -> AuditoriaItem:
    """Convierte fila BD a AuditoriaItem (entidad->modulo, entidad_id->registro_id, detalles->descripcion, exito->resultado)."""
    fecha_str = r.fecha.isoformat() + "Z" if r.fecha else ""
    return AuditoriaItem(
        id=r.id,
        usuario_id=r.usuario_id,
        usuario_email=None,
        accion=r.accion,
        modulo=r.entidad or "",
        tabla=r.entidad or "",
        registro_id=r.entidad_id,
        descripcion=r.detalles,
        campo=None,
        datos_anteriores=None,
        datos_nuevos=None,
        ip_address=r.ip_address,
        user_agent=r.user_agent,
        resultado="EXITOSO" if r.exito else "ERROR",
        mensaje_error=r.mensaje_error,
        fecha=fecha_str,
    )


@router.get("", response_model=AuditoriaListResponse)
def listar_auditoria(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    usuario_email: Optional[str] = Query(None),
    modulo: Optional[str] = Query(None),
    registro_id: Optional[int] = Query(None),
    accion: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    ordenar_por: str = Query("fecha"),
    orden: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Lista registros de auditoría con filtros y paginación. Datos desde BD."""
    q = select(Auditoria)
    if modulo:
        q = q.where(Auditoria.entidad == modulo)
    if registro_id is not None:
        q = q.where(Auditoria.entidad_id == registro_id)
    if accion:
        q = q.where(Auditoria.accion == accion)
    if fecha_desde:
        try:
            fd = datetime.fromisoformat(fecha_desde.replace("Z", "+00:00")).date()
            q = q.where(func.date(Auditoria.fecha) >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.fromisoformat(fecha_hasta.replace("Z", "+00:00")).date()
            q = q.where(func.date(Auditoria.fecha) <= fh)
        except ValueError:
            pass
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0
    col_name = {"modulo": "entidad", "registro_id": "entidad_id"}.get(ordenar_por, ordenar_por)
    order_col = getattr(Auditoria, col_name, Auditoria.fecha)
    if orden == "desc":
        q = q.order_by(order_col.desc())
    else:
        q = q.order_by(order_col.asc())
    q = q.offset(skip).limit(limit)
    rows = db.execute(q).scalars().all()
    items = [_row_to_item(r) for r in rows]
    total_pages = (total + limit - 1) // limit if limit else 0
    return AuditoriaListResponse(
        items=items,
        total=total,
        page=(skip // limit) + 1 if limit else 1,
        page_size=limit,
        total_pages=total_pages,
    )


@router.get("/stats", response_model=AuditoriaStats)
def obtener_estadisticas(db: Session = Depends(get_db)):
    """Estadísticas de auditoría desde BD (totales, por módulo, por usuario, hoy/semana/mes)."""
    hoy = datetime.now(timezone.utc).replace(tzinfo=None)
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total_acciones = db.scalar(select(func.count()).select_from(Auditoria)) or 0
    acciones_hoy = db.scalar(
        select(func.count()).select_from(Auditoria).where(func.date(Auditoria.fecha) == hoy.date())
    ) or 0
    acciones_esta_semana = db.scalar(
        select(func.count()).select_from(Auditoria).where(Auditoria.fecha >= inicio_semana)) or 0
    acciones_este_mes = db.scalar(
        select(func.count()).select_from(Auditoria).where(Auditoria.fecha >= inicio_mes)) or 0
    rows_mod = db.execute(
        select(Auditoria.entidad, func.count()).select_from(Auditoria).group_by(Auditoria.entidad)
    ).all()
    acciones_por_modulo = {r[0] or "": r[1] for r in rows_mod}
    rows_usr = db.execute(
        select(Auditoria.usuario_id, func.count()).select_from(Auditoria).group_by(Auditoria.usuario_id)
    ).all()
    acciones_por_usuario = {str(r[0]) if r[0] is not None else "": r[1] for r in rows_usr}
    return AuditoriaStats(
        total_acciones=total_acciones,
        acciones_por_modulo=acciones_por_modulo,
        acciones_por_usuario=acciones_por_usuario,
        acciones_hoy=acciones_hoy,
        acciones_esta_semana=acciones_esta_semana,
        acciones_este_mes=acciones_este_mes,
    )


@router.get("/exportar")
def exportar_auditoria(
    usuario_email: Optional[str] = Query(None),
    modulo: Optional[str] = Query(None),
    accion: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Exporta auditoría a Excel. Datos desde BD."""
    try:
        import openpyxl
    except ImportError:
        minimal_xlsx = (
            b"PK\x03\x04\x14\x00\x00\x00\x08\x00"
            b"[\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00"
        )
        return StreamingResponse(
            iter([minimal_xlsx]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=auditoria.xlsx"},
        )
    q = select(Auditoria).order_by(Auditoria.fecha.desc()).limit(10000)
    if modulo:
        q = q.where(Auditoria.entidad == modulo)
    if accion:
        q = q.where(Auditoria.accion == accion)
    if fecha_desde:
        try:
            fd = datetime.fromisoformat(fecha_desde.replace("Z", "+00:00")).date()
            q = q.where(func.date(Auditoria.fecha) >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.fromisoformat(fecha_hasta.replace("Z", "+00:00")).date()
            q = q.where(func.date(Auditoria.fecha) <= fh)
        except ValueError:
            pass
    rows = db.execute(q).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    ws.append(["id", "fecha", "usuario_id", "entidad", "accion", "entidad_id", "detalles", "exito"])
    for r in rows:
        ws.append([
            r.id,
            r.fecha.isoformat() if r.fecha else "",
            r.usuario_id or "",
            r.entidad or "",
            r.accion or "",
            r.entidad_id or "",
            (r.detalles or "")[:500],
            "SI" if r.exito else "NO",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=auditoria.xlsx"},
    )


class ControlCarteraItem(BaseModel):
    codigo: str
    titulo: str
    alerta: str
    detalle: str


class PrestamoCarteraChequeoItem(BaseModel):
    prestamo_id: int
    cliente_id: int
    cedula: str
    nombres: str
    estado_prestamo: str
    cliente_email: str
    tiene_alerta: bool
    controles: List[ControlCarteraItem]


class PrestamoCarteraChequeoResponse(BaseModel):
    items: List[PrestamoCarteraChequeoItem]
    resumen: dict = Field(
        ...,
        description=(
            "Totales de la corrida: prestamos_evaluados, prestamos_con_alerta, prestamos_listados_total, "
            "prestamos_listados (pagina), conteos_por_control, reglas_version, fecha_referencia, pagina_skip, pagina_limit."
        ),
    )
    meta_ultima_corrida: dict
    sincronizar_estado_cuotas: Optional[dict[str, Any]] = Field(
        default=None,
        description="Solo en POST ejecutar/corregir: alineacion masiva de cuotas.estado antes/después del flujo.",
    )


class LiquidadosCierreChequeoResponse(BaseModel):
    items: List[PrestamoCarteraChequeoItem]
    resumen: dict = Field(
        ...,
        description=(
            "Hallazgos de cierre solo para prestamos LIQUIDADO (fecha_liquidado, finiquito_casos, documentos). "
            "Totales reflejan el filtro (cedula/prestamo_id) antes de paginar; prestamos_listados es la pagina."
        ),
    )


class LiquidadosIntensivaResponse(BaseModel):
    cartera: PrestamoCarteraChequeoResponse
    cierre: LiquidadosCierreChequeoResponse


class CarteraCorregirBody(BaseModel):
    """Correcciones de datos orientadas por la auditoria de cartera (solo administrador)."""

    sincronizar_estados: bool = True
    reaplicar_cascada_desajuste_pagos: bool = Field(
        False,
        description="Prestamos con control total_pagado_vs_aplicado_cuotas en SI (típ. LIQUIDADO descuadrado).",
    )
    reaplicar_cascada_pagos_sin_aplicacion_cuotas: bool = Field(
        False,
        description="Prestamos con control pagos_sin_aplicacion_a_cuotas en SI (sin cuota_pagos o saldo > 0.02 USD).",
    )
    max_reaplicaciones: int = Field(50, ge=1, le=500)


class CarteraCorreccionResponse(BaseModel):
    items: List[PrestamoCarteraChequeoItem]
    resumen: dict = Field(
        ...,
        description="Igual estructura que en GET `/prestamos/cartera/chequeos` (respuesta completa, sin paginar).",
    )
    meta_ultima_corrida: dict
    sincronizar_estado_cuotas: Optional[dict[str, Any]] = None
    reaplicar_cascada: List[dict[str, Any]] = Field(default_factory=list)


class CarteraRevisionCrearBody(BaseModel):
    prestamo_id: int = Field(..., ge=1)
    codigo_control: str = Field(..., max_length=80)
    tipo: str = Field("MARCAR_OK", max_length=30)
    nota: Optional[str] = Field(None, max_length=2000)


class CarteraRevisionOcultoPar(BaseModel):
    prestamo_id: int
    codigo_control: str


class CarteraRevisionOcultosResponse(BaseModel):
    ocultos: List[CarteraRevisionOcultoPar]


class CarteraRevisionOcultosBody(BaseModel):
    prestamo_ids: List[int] = Field(
        default_factory=list,
        description="IDs de prestamos de la pagina actual (maximo 500 recomendado).",
    )


class CarteraRevisionItemResponse(BaseModel):
    id: int
    prestamo_id: int
    codigo_control: str
    tipo: str
    usuario_id: int
    usuario_email: Optional[str] = None
    nota: Optional[str] = None
    creado_en: str
    payload_snapshot: Optional[Any] = None


class AuditoriaCarteraResumenResponse(BaseModel):
    """Respuesta sin items: solo KPIs y conteos por control (GET /prestamos/cartera/resumen)."""

    resumen: dict = Field(
        ...,
        description="Misma forma que `resumen` en GET chequeos, con prestamos_listados=0 y sin filas.",
    )
    meta_ultima_corrida: dict


class SincronizarEstadosCuotasResponse(BaseModel):
    cuotas_escaneadas: int = Field(..., description="Cuotas en prestamos APROBADO/LIQUIDADO recorridas.")
    estados_actualizados: int = Field(..., description="Filas cuyo estado en BD se actualizo al valor calculado.")


def _persistir_meta_desde_resumen_cartera(db: Session, resumen: dict[str, Any]) -> None:
    if resumen.get("excluye_marcar_ok") is True:
        raise HTTPException(
            status_code=500,
            detail=(
                "No se persiste meta de auditoria de cartera con exclusion MARCAR_OK: "
                "los KPIs guardados deben reflejar solo el motor objetivo sobre la BD."
            ),
        )
    raw_conteos = resumen.get("conteos_por_control")
    persistir_meta_ejecucion(
        db,
        total_evaluados=int(resumen.get("prestamos_evaluados") or 0),
        con_alerta=int(resumen.get("prestamos_con_alerta") or 0),
        conteos_por_control=raw_conteos if isinstance(raw_conteos, dict) else None,
        reglas_version=str(resumen.get("reglas_version") or ""),
        commit=True,
    )


def _prestamos_cartera_dicts_a_items(rows: List[dict[str, Any]]) -> List[PrestamoCarteraChequeoItem]:
    return [
        PrestamoCarteraChequeoItem(
            prestamo_id=r["prestamo_id"],
            cliente_id=r["cliente_id"],
            cedula=r["cedula"],
            nombres=r["nombres"],
            estado_prestamo=r["estado_prestamo"],
            cliente_email=r["cliente_email"],
            tiene_alerta=r["tiene_alerta"],
            controles=[ControlCarteraItem(**c) for c in r["controles"]],
        )
        for r in rows
    ]


@router.get("/prestamos/cartera/meta")
def meta_auditoria_cartera(
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """Ultima corrida automatica (03:00) y totales guardados en `configuracion` (JSON en clave auditoria_cartera_ultima_resumen)."""
    return leer_meta_ejecucion(db)


@router.post(
    "/prestamos/cartera/sincronizar-estados-cuotas",
    response_model=SincronizarEstadosCuotasResponse,
)
def sincronizar_estados_cuotas_cartera_endpoint(
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Alinea `cuotas.estado` con la regla de negocio (vencimiento y total_pagado vs monto), misma que el control
    estado_cuota_vs_calculo y que el paso inicial de POST ejecutar. No recalcula controles ni persiste meta de cartera.
    """
    from app.services.cuota_estado import sincronizar_estado_cuotas_cartera

    stats = sincronizar_estado_cuotas_cartera(db, commit=True)
    return SincronizarEstadosCuotasResponse(
        cuotas_escaneadas=int(stats.get("cuotas_escaneadas") or 0),
        estados_actualizados=int(stats.get("estados_actualizados") or 0),
    )


@router.get("/prestamos/cartera/resumen", response_model=AuditoriaCarteraResumenResponse)
def resumen_auditoria_cartera(
    prestamo_id: Optional[int] = Query(
        None,
        ge=1,
        description="Opcional. Acota la evaluacion a un prestamo APROBADO/LIQUIDADO.",
    ),
    cedula: Optional[str] = Query(
        None,
        description="Opcional. Fragmento de cedula del prestamo (misma regla que GET chequeos).",
    ),
    excluir_marcar_ok: bool = Query(
        False,
        description=(
            "Vista operativa: si true, la cola excluye solo pares con ultimo MARCAR_OK (aceptado); "
            "fuera de eso un caso sale de la lista solo si el motor ya no marca SI (causa raiz). "
            "Meta persistida (job/manual ejecutar) usa siempre motor sin esta capa."
        ),
    ),
    codigo_control: Optional[str] = Query(
        None,
        description="Opcional. Acota prestamos_con_alerta al subconjunto con este control en SI (conteos_por_control siguen globales).",
    ),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Misma logica que GET `/prestamos/cartera/chequeos`, pero **sin** devolver la lista de prestamos.
    Util para KPIs o dashboards con menos payload; el coste de CPU en BD es el mismo que una corrida completa.
    """
    cod_cc = _codigo_control_cartera_opcional(codigo_control)
    _rows, resumen = ejecutar_auditoria_cartera(
        db,
        solo_con_alerta=True,
        prestamo_id=prestamo_id,
        cedula_contiene=cedula,
        skip=0,
        limit=None,
        incluir_filas=False,
        excluir_marcar_ok=excluir_marcar_ok,
        codigo_control=cod_cc,
    )
    meta = leer_meta_ejecucion(db)
    return AuditoriaCarteraResumenResponse(resumen=resumen, meta_ultima_corrida=meta)


@router.get("/prestamos/cartera/chequeos", response_model=PrestamoCarteraChequeoResponse)
def listar_chequeos_cartera(
    solo_alertas: bool = Query(
        True,
        description=(
            "Parametro historico sin efecto en el resultado: siempre solo prestamos con alerta SI "
            "y solo controles en SI. Mantener en clientes legacy."
        ),
    ),
    skip: int = Query(0, ge=0, description="Offset sobre la lista de prestamos con alerta (orden por id)."),
    limit: int = Query(100, ge=1, le=5000, description="Tamano de pagina; hasta 5000 (export masivo)."),
    prestamo_id: Optional[int] = Query(None, ge=1, description="Solo evaluar este prestamo si esta APROBADO/LIQUIDADO."),
    cedula: Optional[str] = Query(
        None,
        description="Fragmento de cedula del prestamo (coincidencia parcial, sin normalizar al cliente).",
    ),
    excluir_marcar_ok: bool = Query(
        False,
        description=(
            "Cola operativa: si true, omite solo (prestamo, control) con ultimo MARCAR_OK; "
            "el resto solo desaparece si el motor deja SI. No altera SI/NO del motor. "
            "Job 03:00 y meta persistida usan false."
        ),
    ),
    codigo_control: Optional[str] = Query(
        None,
        description="Opcional. Lista y paginacion solo sobre prestamos con este control en SI; conteos_por_control siguen globales.",
    ),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Revision de cartera en tiempo real desde tablas prestamos, clientes, cuotas, pagos, cuota_pagos.
    Cada control devuelve alerta SI/NO (SI = revisar por un humano).
    Con `excluir_marcar_ok=true`, un caso deja de listarse solo por aceptacion MARCAR_OK o porque el motor ya no marca SI.
    """
    cod_cc = _codigo_control_cartera_opcional(codigo_control)
    rows, resumen = ejecutar_auditoria_cartera(
        db,
        solo_con_alerta=solo_alertas,
        prestamo_id=prestamo_id,
        cedula_contiene=cedula,
        skip=skip,
        limit=limit,
        excluir_marcar_ok=excluir_marcar_ok,
        codigo_control=cod_cc,
    )
    meta = leer_meta_ejecucion(db)
    return PrestamoCarteraChequeoResponse(
        items=_prestamos_cartera_dicts_a_items(rows),
        resumen=resumen,
        meta_ultima_corrida=meta,
    )


@router.get(
    "/prestamos/liquidados/auditoria-intensiva",
    response_model=LiquidadosIntensivaResponse,
)
def auditoria_intensiva_prestamos_liquidados(
    skip: int = Query(0, ge=0, description="Offset compartido para cartera y cierre (listas independientes)."),
    limit: int = Query(50, ge=1, le=5000, description="Tamano de pagina compartido para cartera y cierre."),
    prestamo_id: Optional[int] = Query(None, ge=1, description="Filtra ambas listas al mismo prestamo."),
    cedula: Optional[str] = Query(
        None,
        description="Fragmento de cedula del prestamo (misma regla que cartera; aplica a ambas listas).",
    ),
    excluir_marcar_ok: bool = Query(
        False,
        description="Solo aplica a la seccion cartera (misma semantica que GET /prestamos/cartera/chequeos).",
    ),
    codigo_control: Optional[str] = Query(
        None,
        description="Solo aplica a la seccion cartera (misma semantica que GET /prestamos/cartera/chequeos).",
    ),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Auditoria intensiva para prestamos **LIQUIDADO**:
    - **cartera**: misma logica que `/prestamos/cartera/chequeos`, pero el universo evaluado es solo LIQUIDADO.
    - **cierre**: hallazgos adicionales (fecha_liquidado, finiquito_casos, documentos duplicados, riesgo doc en otro prestamo).

    La seccion `cierre` no usa bitacora MARCAR_OK (motor objetivo sobre tablas reales).
    """
    cod_cc = _codigo_control_cartera_opcional(codigo_control)
    rows, resumen = ejecutar_auditoria_cartera(
        db,
        solo_con_alerta=True,
        prestamo_id=prestamo_id,
        cedula_contiene=cedula,
        skip=skip,
        limit=limit,
        excluir_marcar_ok=excluir_marcar_ok,
        codigo_control=cod_cc,
        estados_filas_prestamo=("LIQUIDADO",),
    )
    meta = leer_meta_ejecucion(db)
    cartera = PrestamoCarteraChequeoResponse(
        items=_prestamos_cartera_dicts_a_items(rows),
        resumen=resumen,
        meta_ultima_corrida=meta,
    )

    hall_full, _ = hallazgos_cierre_prestamos_liquidados(db)
    hall_filtrado = filtrar_filas_cierre(
        hall_full,
        prestamo_id=prestamo_id,
        cedula_contiene=cedula,
    )
    res_global = resumen_cierre_desde_filas(hall_filtrado)
    hall_page = paginar_filas(hall_filtrado, skip=skip, limit=limit)
    cierre_resumen = {
        **res_global,
        "pagina_skip": int(skip),
        "pagina_limit": int(limit),
        "prestamos_listados": len(hall_page),
    }
    cierre = LiquidadosCierreChequeoResponse(
        items=_prestamos_cartera_dicts_a_items(hall_page),
        resumen=cierre_resumen,
    )
    return LiquidadosIntensivaResponse(cartera=cartera, cierre=cierre)


@router.post("/prestamos/cartera/ejecutar", response_model=PrestamoCarteraChequeoResponse)
def ejecutar_y_persistir_auditoria_cartera(
    solo_alertas: bool = Query(
        True,
        description=(
            "Parametro historico sin efecto: respuesta completa de alertas (sin paginar en este POST). "
            "Igual que GET en cuanto a criterios SI/NO."
        ),
    ),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Alinea `cuotas.estado` con la regla de negocio (misma que la auditoria), recalcula chequeos,
    persiste metadatos de ejecucion y devuelve alertas restantes.

    La evaluacion y la meta persistida son siempre el **motor objetivo** (`excluir_marcar_ok=false`):
    la bitacora MARCAR_OK no altera conteos guardados en `configuracion`.
    """
    _ = solo_alertas  # compat. API; ejecutar_auditoria_cartera ya solo devuelve filas con alerta SI
    from app.services.cuota_estado import sincronizar_estado_cuotas_cartera

    sync_stats = sincronizar_estado_cuotas_cartera(db, commit=True)
    rows, resumen = ejecutar_auditoria_cartera(
        db,
        solo_con_alerta=False,
        skip=0,
        limit=None,
        excluir_marcar_ok=False,
        codigo_control=None,
    )
    _persistir_meta_desde_resumen_cartera(db, resumen)
    meta = leer_meta_ejecucion(db)
    return PrestamoCarteraChequeoResponse(
        items=_prestamos_cartera_dicts_a_items(rows),
        resumen=resumen,
        meta_ultima_corrida=meta,
        sincronizar_estado_cuotas=sync_stats,
    )


@router.post("/prestamos/cartera/corregir", response_model=CarteraCorreccionResponse)
def corregir_auditoria_cartera(
    body: CarteraCorregirBody,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Opciones de correccion acotadas: sincronizar estados de cuota en cartera y/o reaplicar cascada
    en prestamos que la auditoria marca con desajuste suma pagos vs aplicado a cuotas y/o con
    pagos operativos sin aplicacion a cuotas (control 15). Si ambos flags estan activos, se unen
    los prestamos afectados y se respeta max_reaplicaciones en total.
    Solo administrador. Tras las acciones vuelve a ejecutar auditoria y persiste meta.
    """
    if canonical_rol(current_user.rol) != "admin":
        raise HTTPException(
            status_code=403,
            detail="Solo administracion puede ejecutar correcciones de cartera.",
        )
    from app.services.cuota_estado import sincronizar_estado_cuotas_cartera
    from app.services.pagos_cuotas_reaplicacion import reset_y_reaplicar_cascada_prestamo

    reaplicar_resultados: list[dict[str, Any]] = []
    need_reap = body.reaplicar_cascada_desajuste_pagos or body.reaplicar_cascada_pagos_sin_aplicacion_cuotas
    if need_reap:
        _, meta_pre = ejecutar_auditoria_cartera(
            db,
            solo_con_alerta=False,
            skip=0,
            limit=None,
            incluir_filas=False,
            incluir_mapa_ids_por_control=True,
            excluir_marcar_ok=False,
            codigo_control=None,
        )
        mapa = meta_pre.get("prestamo_ids_alerta_por_control") or {}
        pids_set: set[int] = set()
        if body.reaplicar_cascada_desajuste_pagos:
            pids_set.update(int(x) for x in mapa.get("total_pagado_vs_aplicado_cuotas", []))
        if body.reaplicar_cascada_pagos_sin_aplicacion_cuotas:
            pids_set.update(int(x) for x in mapa.get("pagos_sin_aplicacion_a_cuotas", []))
        pids = sorted(pids_set)[: int(body.max_reaplicaciones)]
        for pid in pids:
            try:
                r = reset_y_reaplicar_cascada_prestamo(db, pid)
                if r.get("ok"):
                    db.commit()
                    reaplicar_resultados.append({"prestamo_id": pid, "ok": True, **r})
                else:
                    db.rollback()
                    reaplicar_resultados.append({"prestamo_id": pid, "ok": False, **r})
            except Exception as e:
                db.rollback()
                reaplicar_resultados.append({"prestamo_id": pid, "ok": False, "error": str(e)})

    sync_stats: Optional[dict[str, Any]] = None
    if body.sincronizar_estados:
        sync_stats = sincronizar_estado_cuotas_cartera(db, commit=True)

    rows, resumen = ejecutar_auditoria_cartera(
        db,
        solo_con_alerta=False,
        skip=0,
        limit=None,
        excluir_marcar_ok=False,
        codigo_control=None,
    )
    _persistir_meta_desde_resumen_cartera(db, resumen)
    meta = leer_meta_ejecucion(db)
    return CarteraCorreccionResponse(
        items=_prestamos_cartera_dicts_a_items(rows),
        resumen=resumen,
        meta_ultima_corrida=meta,
        sincronizar_estado_cuotas=sync_stats,
        reaplicar_cascada=reaplicar_resultados,
    )


@router.post("/prestamos/cartera/revisiones", response_model=CarteraRevisionItemResponse)
def crear_revision_cartera(
    body: CarteraRevisionCrearBody,
    db: Session = Depends(get_db),
    current_user: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Registra revision humana (append-only). `MARCAR_OK` documenta excepcion de negocio aceptada:
    requiere nota (min. 15 caracteres). Es una de las dos vias legitimas para que el caso deje la cola
    operativa (la otra es que el motor deje de marcar SI tras corregir datos). No borra la condicion del motor.
    `REVOCAR_OK` anula la vigencia operativa de la ultima aceptacion (ultimo evento debe ser MARCAR_OK).
    """
    cod = body.codigo_control.strip()
    if cod not in CONTROLES_CARTERA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="codigo_control no es un control de cartera conocido",
        )
    tipo_u = (body.tipo or "MARCAR_OK").strip().upper()
    if tipo_u not in TIPOS_REVISION_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail=f"tipo no permitido en esta version: {tipo_u}",
        )
    if not db.get(Prestamo, body.prestamo_id):
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    nota_val = (body.nota or "").strip() or None
    payload_snap: Optional[Any] = None
    if tipo_u == "MARCAR_OK":
        if not nota_val or len(nota_val) < 15:
            raise HTTPException(
                status_code=400,
                detail="MARCAR_OK requiere nota (minimo 15 caracteres) documentando la excepcion de negocio.",
            )
        payload_snap = construir_payload_snapshot_marcar_ok(
            db, prestamo_id=body.prestamo_id, codigo_control=cod
        )
    elif tipo_u == "REVOCAR_OK":
        if not nota_val or len(nota_val) < NOTA_MIN_REVOCAR_OK:
            raise HTTPException(
                status_code=400,
                detail=f"REVOCAR_OK requiere nota (minimo {NOTA_MIN_REVOCAR_OK} caracteres).",
            )
        ult = ultimo_tipo_revision_par(db, body.prestamo_id, cod)
        if ult != "MARCAR_OK":
            raise HTTPException(
                status_code=400,
                detail="No hay excepcion vigente (ultimo evento del par no es MARCAR_OK).",
            )
        payload_snap = payload_minimo_revocar_ok(
            prestamo_id=body.prestamo_id, codigo_control=cod
        )
    row = AuditoriaCarteraRevision(
        prestamo_id=body.prestamo_id,
        codigo_control=cod,
        tipo=tipo_u,
        usuario_id=current_user.id,
        nota=nota_val,
        payload_snapshot=payload_snap,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    u = db.get(User, row.usuario_id)
    email = u.email if u else None
    return CarteraRevisionItemResponse(
        id=row.id,
        prestamo_id=row.prestamo_id,
        codigo_control=row.codigo_control,
        tipo=row.tipo,
        usuario_id=row.usuario_id,
        usuario_email=email,
        nota=row.nota,
        creado_en=iso_utc(row.creado_en),
        payload_snapshot=row.payload_snapshot,
    )


@router.post("/prestamos/cartera/revisiones/ocultos", response_model=CarteraRevisionOcultosResponse)
def ocultos_revision_cartera(
    body: CarteraRevisionOcultosBody,
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Pares (prestamo_id, codigo_control) cuyo **ultimo** evento en bitacora es `MARCAR_OK`.
    POST (no GET) para enviar muchos `prestamo_ids` sin limites de URL.
    """
    ids = body.prestamo_ids or []
    if len(ids) > 5000:
        raise HTTPException(status_code=400, detail="Maximo 5000 prestamo_ids por solicitud")
    ocultos = listar_ocultos_marcar_ok(db, ids)
    return CarteraRevisionOcultosResponse(
        ocultos=[CarteraRevisionOcultoPar(**x) for x in ocultos]
    )


@router.get(
    "/prestamos/cartera/revisiones/controles-con-excepciones",
    response_model=List[str],
)
def listar_controles_con_excepciones_historicas(
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """Codigos de control con al menos un MARCAR_OK en bitacora (historico)."""
    return listar_codigos_control_con_excepcion_historica(db)


@router.get("/prestamos/cartera/revisiones/export-excel")
def exportar_revisiones_cartera_excel(
    codigo_control: str = Query(..., min_length=1, max_length=80),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """
    Descarga Excel con **todas** las filas de bitacora del control indicado (todos los tipos),
    incluyendo snapshot JSON cuando exista.
    """
    cod = codigo_control.strip()
    if cod not in CONTROLES_CARTERA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="codigo_control no es un control de cartera conocido",
        )
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = (
        db.query(AuditoriaCarteraRevision, User.email)
        .outerjoin(User, User.id == AuditoriaCarteraRevision.usuario_id)
        .filter(AuditoriaCarteraRevision.codigo_control == cod)
        .order_by(AuditoriaCarteraRevision.creado_en.asc())
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "revisiones"
    headers = [
        "id",
        "prestamo_id",
        "codigo_control",
        "tipo",
        "usuario_email",
        "nota",
        "creado_en",
        "payload_snapshot_json",
    ]
    ws.append(headers)
    for r, email in rows:
        snap = r.payload_snapshot
        snap_txt = json.dumps(snap, ensure_ascii=False) if snap is not None else ""
        ws.append(
            [
                r.id,
                r.prestamo_id,
                r.codigo_control,
                r.tipo,
                email or "",
                r.nota or "",
                iso_utc(r.creado_en),
                snap_txt,
            ]
        )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(len(headers))].width = 80
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_cod = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in cod)[:60]
    fname = f"auditoria_cartera_revisiones_{safe_cod}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/prestamos/cartera/revisiones/historial", response_model=List[CarteraRevisionItemResponse])
def historial_revision_cartera(
    prestamo_id: int = Query(..., ge=1),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    _aud: UserResponse = Depends(require_auditoria_cartera_access),
):
    """Historial cronologico descendente de revisiones para un prestamo."""
    rows = (
        db.query(AuditoriaCarteraRevision, User.email)
        .outerjoin(User, User.id == AuditoriaCarteraRevision.usuario_id)
        .filter(AuditoriaCarteraRevision.prestamo_id == prestamo_id)
        .order_by(AuditoriaCarteraRevision.creado_en.desc())
        .limit(limit)
        .all()
    )
    out: List[CarteraRevisionItemResponse] = []
    for r, email in rows:
        out.append(
            CarteraRevisionItemResponse(
                id=r.id,
                prestamo_id=r.prestamo_id,
                codigo_control=r.codigo_control,
                tipo=r.tipo,
                usuario_id=r.usuario_id,
                usuario_email=email,
                nota=r.nota,
                creado_en=iso_utc(r.creado_en),
                payload_snapshot=r.payload_snapshot,
            )
        )
    return out


class Control5PagoDuplicadoItem(BaseModel):
    pago_id: int
    prestamo_id: Optional[int] = None
    fecha_pago: Optional[str] = None
    monto_pagado: Optional[float] = None
    conciliado: bool = False
    estado_pago: str = ""
    numero_documento: str = ""
    referencia_pago: str = ""
    institucion_bancaria: str = ""


class Control5DuplicadosListaResponse(BaseModel):
    items: List[Control5PagoDuplicadoItem]


class Control5VistoAplicarResponse(BaseModel):
    pago_id: int
    prestamo_id: Optional[int] = None
    numero_documento_anterior: Optional[str] = None
    numero_documento_nuevo: str
    sufijo_cuatro_digitos: str
    auditoria_id: int


class Control15PagoSinAplicacionItem(BaseModel):
    pago_id: int
    prestamo_id: int
    fecha_pago: Optional[str] = None
    monto_pagado: float
    sum_monto_aplicado_cuotas: float
    saldo_sin_aplicar_usd: float
    motivo: str


class Control15PagosSinAplicacionListaResponse(BaseModel):
    items: List[Control15PagoSinAplicacionItem]


@router.get(
    "/prestamos/cartera/control-5-pagos-duplicados-fecha-monto/{prestamo_id}",
    response_model=Control5DuplicadosListaResponse,
)
def listar_control5_pagos_duplicados_por_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _admin: UserResponse = Depends(require_admin),
):
    """
    Control 5: pagos operativos que comparten (prestamo, fecha calendario, monto) con otro pago.
    Solo administrador. Usado para aplicar Visto (sufijo aleatorio en documento + bitacora).
    """
    if not db.get(Prestamo, prestamo_id):
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    raw = listar_pagos_duplicados_fecha_monto_por_prestamo(db, prestamo_id)
    return Control5DuplicadosListaResponse(
        items=[Control5PagoDuplicadoItem(**x) for x in raw]
    )


@router.post(
    "/prestamos/cartera/control-5-pagos-duplicados-fecha-monto/{pago_id}/visto",
    response_model=Control5VistoAplicarResponse,
)
def aplicar_control5_visto_duplicado_fecha_monto(
    pago_id: int,
    db: Session = Depends(get_db),
    admin: UserResponse = Depends(require_admin),
):
    """
    Solo administrador. Anexa `_A####` o `_P####` (4 digitos aleatorios) a `numero_documento`:
    A = mismo prestamo / contexto varias cuotas; P = documento duplicado respecto a otro prestamo.
    Marca exclusion del motor control 5 y registra fila en `auditoria_pago_control5_visto`.
    """
    try:
        out = aplicar_visto_control5_duplicado_fecha_monto(db, pago_id, admin.id)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Conflicto de unicidad en numero_documento. Reintente; se generara otro sufijo.",
        )
    except Exception:
        db.rollback()
        raise
    return Control5VistoAplicarResponse(**out)


@router.get(
    "/prestamos/cartera/control-15-pagos-sin-aplicacion-cuotas/{prestamo_id}",
    response_model=Control15PagosSinAplicacionListaResponse,
)
def listar_control15_pagos_sin_aplicacion_cuotas_por_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    _admin: UserResponse = Depends(require_admin),
):
    """
    Control 15: pagos operativos del prestamo sin articulacion completa en cuota_pagos.
    Solo administrador. Misma regla que el motor `pagos_sin_aplicacion_a_cuotas`.
    """
    if not db.get(Prestamo, prestamo_id):
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")
    raw = listar_pagos_sin_aplicacion_cuotas_por_prestamo(db, prestamo_id)
    return Control15PagosSinAplicacionListaResponse(
        items=[Control15PagoSinAplicacionItem(**x) for x in raw]
    )


@router.get("/{auditoria_id}", response_model=AuditoriaItem)
def obtener_auditoria(auditoria_id: int, db: Session = Depends(get_db)):
    """Obtiene un registro de auditoría por ID desde BD."""
    row = db.get(Auditoria, auditoria_id)
    if not row:
        raise HTTPException(status_code=404, detail="Registro de auditoría no encontrado")
    return _row_to_item(row)


@router.post("/registrar", response_model=AuditoriaItem)
def registrar_evento(
    body: RegistrarAuditoriaBody,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Registra un evento de auditoría en BD. usuario_id se toma del usuario autenticado."""
    now = datetime.now(timezone.utc)
    row = Auditoria(
        usuario_id=current_user.id,
        accion=body.accion,
        entidad=body.modulo,
        entidad_id=body.registro_id,
        detalles=body.descripcion,
        exito=True,
        fecha=now,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _row_to_item(row)
