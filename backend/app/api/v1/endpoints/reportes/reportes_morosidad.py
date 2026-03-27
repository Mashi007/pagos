"""
Reportes de morosidad.

AUDITORIA — conexion API base: /api/v1/reportes

Solo estado calculado MORA (CASE con c.total_pagado; mismo criterio que el Excel del reporte):
  - GET /morosidad/clientes
  - GET /exportar/morosidad-clientes
  - GET /exportar/morosidad-cedulas  (una fila por cedula: reporte principal)
  - GET /morosidad/auditoria/mora-por-cliente?cedula= | cliente_id=  (auditoria alineada al reporte por cedula)
  - GET /morosidad/auditoria/mora-por-prestamo?prestamo_id=  (opcional: desglose por un solo prestamo)

Criterio distinto "pago vencido" (4 meses + impago; incluye logica no igual a solo MORA):
  - GET /morosidad, /morosidad/por-mes, /morosidad/por-rangos, GET /exportar/morosidad

Frontend Reportes: tipo MOROSIDAD usa exportarReporteMorosidadCedulas -> /exportar/morosidad-cedulas.
"""

import calendar

import logging

import io

from datetime import date

from typing import List, Optional



from fastapi import APIRouter, Depends, HTTPException, Query

from fastapi.responses import Response

from sqlalchemy import func, literal_column, select, text

from sqlalchemy.orm import Session, aliased



from app.core.database import get_db

from app.core.deps import get_current_user

from app.models.cliente import Cliente

from app.models.cuota import Cuota

from app.models.cuota_pago import CuotaPago

from app.models.prestamo import Prestamo

from app.services.cuota_estado import (
    SQL_PG_ESTADO_CUOTA_CASE_CORRELATED,
    SQL_PG_ESTADO_CUOTA_CASE_CORRELATED_TOTAL_PAGADO,
    hoy_negocio,
)



from app.api.v1.endpoints.reportes_utils import _safe_float, _parse_fecha, _periodos_desde_filtros



router = APIRouter(dependencies=[Depends(get_current_user)])

logger = logging.getLogger(__name__)

# Desviacion maxima aceptada (USD) entre cuotas.total_pagado y SUM(cuota_pagos) antes de advertir en auditoria.
TOL_ALINEACION_PAGADO_CUOTA = 0.02


# Export /exportar/morosidad-cedulas: mismo CASE que clasificar_estado_cuota / tabla de amortizacion.
# Lista cerrada de codigos posibles (6): PAGADO, PAGO_ADELANTADO, PENDIENTE, PARCIAL, VENCIDO, MORA.
# No negociable: solo MORA entra al informe; VENCIDO y todos los demas quedan fuera.
REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS = ("MORA",)
REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS = (
    "PAGADO",
    "PAGO_ADELANTADO",
    "PENDIENTE",
    "PARCIAL",
    "VENCIDO",
)
assert not (
    set(REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS) & set(REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS)
), "reporte morosidad: inclusion y exclusion no deben solaparse"
assert (
    set(REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS) | set(REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS)
    == {"PAGADO", "PAGO_ADELANTADO", "PENDIENTE", "PARCIAL", "VENCIDO", "MORA"}
), "reporte morosidad: debe cubrir exactamente los 6 estados de cuota_estado"


def _sql_expr_trim_estado_cuota_alias_c() -> str:
    """TRIM(CASE...) con alias `c`; usa total_pagado de columna como la lista de amortizacion."""


    return f"TRIM(BOTH FROM ({SQL_PG_ESTADO_CUOTA_CASE_CORRELATED_TOTAL_PAGADO}))"


def _filtro_cuota_solo_estado_mora_sql() -> text:
    """Estado calculado == MORA unicamente (UI: Mora 4+ meses). VENCIDO y resto quedan fuera."""


    return text(f"({_sql_expr_trim_estado_cuota_alias_c()}) = 'MORA'")


def _filtro_cuota_mora_desde_case_sql(case_sql: str) -> text:
    """Mismo patron que solo MORA pero con otro CASE (p. ej. SUM cuota_pagos vs total_pagado)."""


    return text(f"(TRIM(BOTH FROM ({case_sql}))) = 'MORA'")


RANGOS_ATRASO = [

    {"key": "1_dia", "label": "1 dÃ­a atrasado", "min_dias": 1, "max_dias": 14},

    {"key": "15_dias", "label": "15 dÃ­as atrasado", "min_dias": 15, "max_dias": 29},

    {"key": "30_dias", "label": "30 dÃ­as atrasado", "min_dias": 30, "max_dias": 59},

    {"key": "2_meses", "label": "2 meses atrasado", "min_dias": 60, "max_dias": 89},

    {"key": "4_meses", "label": "4+ meses (moroso)", "min_dias": 121, "max_dias": 99999},

]


@router.get("/morosidad/auditoria/mora-por-cliente")
def get_auditoria_mora_por_cliente(
    db: Session = Depends(get_db),
    cedula: Optional[str] = Query(
        None,
        description="Cedula del cliente: mismo alcance que GET /exportar/morosidad-cedulas (una fila por cedula)",
    ),
    cliente_id: Optional[int] = Query(
        None,
        ge=1,
        description="ID del cliente (alternativa a cedula)",
    ),
):
    """
    Auditoria del **reporte de morosidad** (Excel por cedula): cuenta cuotas MORA en todos los prestamos
    APROBADO del cliente ACTIVO. Es el criterio del informe, no un modulo de prestamos.
    """
    has_ced = cedula is not None and str(cedula).strip() != ""
    has_id = cliente_id is not None
    if has_ced == has_id:
        raise HTTPException(
            status_code=400,
            detail="Indique exactamente uno: cedula o cliente_id",
        )

    if has_id:
        row_cl = db.get(Cliente, cliente_id)
    else:
        ced_norm = str(cedula).strip()
        row_cl = db.execute(
            select(Cliente).where(Cliente.cedula == ced_norm)
        ).scalars().first()

    if not row_cl:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    if row_cl.estado != "ACTIVO":
        raise HTTPException(
            status_code=400,
            detail="Cliente no ACTIVO: el reporte de morosidad solo incluye clientes ACTIVOS",
        )

    c = aliased(Cuota, name="c")
    filtro_mora = _filtro_cuota_solo_estado_mora_sql()
    filtro_mora_sum_pagos = _filtro_cuota_mora_desde_case_sql(SQL_PG_ESTADO_CUOTA_CASE_CORRELATED)

    base_filters = [
        Cliente.id == row_cl.id,
        Cliente.estado == "ACTIVO",
        Prestamo.estado == "APROBADO",
    ]

    agg = db.execute(
        select(func.count(c.id), func.coalesce(func.sum(c.monto), 0))
        .select_from(c)
        .join(Prestamo, c.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(*base_filters, filtro_mora)
    ).one()

    n_mora_si_sum_cuota_pagos = int(
        db.execute(
            select(func.count(c.id))
            .select_from(c)
            .join(Prestamo, c.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(*base_filters, filtro_mora_sum_pagos)
        ).scalar_one()
        or 0
    )

    n_mora_estado_columna_bd = int(
        db.execute(
            select(func.count(c.id))
            .select_from(c)
            .join(Prestamo, c.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(*base_filters, c.estado == "MORA")
        ).scalar_one()
        or 0
    )

    detalle = db.execute(
        select(c.id, c.numero_cuota, c.monto, Prestamo.id.label("prestamo_id"))
        .select_from(c)
        .join(Prestamo, c.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(*base_filters, filtro_mora)
        .order_by(Prestamo.id, c.numero_cuota.asc())
    ).fetchall()

    sum_aplicado_sq = (
        select(func.coalesce(func.sum(CuotaPago.monto_aplicado), 0))
        .where(CuotaPago.cuota_id == Cuota.id)
        .correlate(Cuota)
        .scalar_subquery()
    )
    alin_rows = db.execute(
        select(
            Cuota.id,
            Cuota.prestamo_id,
            Cuota.numero_cuota,
            Cuota.total_pagado,
            sum_aplicado_sq.label("sum_aplicado"),
        )
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .where(Prestamo.cliente_id == row_cl.id, Prestamo.estado == "APROBADO")
        .order_by(Cuota.prestamo_id, Cuota.numero_cuota.asc())
    ).fetchall()
    desv: List[dict] = []
    for ar in alin_rows:
        col = _safe_float(ar.total_pagado)
        sap = _safe_float(ar.sum_aplicado)
        if abs(col - sap) > TOL_ALINEACION_PAGADO_CUOTA:
            desv.append(
                {
                    "cuota_id": ar.id,
                    "prestamo_id": ar.prestamo_id,
                    "numero_cuota": ar.numero_cuota,
                    "total_pagado_columna": round(col, 2),
                    "sum_cuota_pagos": round(sap, 2),
                    "diff_usd": round(col - sap, 2),
                }
            )

    n_mora_reporte = int(agg[0] or 0)

    return {
        "alcance": "reporte_morosidad_cedulas",
        "cliente_id": row_cl.id,
        "cedula": row_cl.cedula or "",
        "nombres": row_cl.nombres or "",
        "cantidad_cuotas_mora": n_mora_reporte,
        "total_monto_usd": round(_safe_float(agg[1]), 2),
        "criterio": "Informe morosidad: todas las cuotas MORA del cliente en prestamos APROBADO (ACTIVO).",
        "cantidad_cuotas_mora_si_sum_cuota_pagos": n_mora_si_sum_cuota_pagos,
        "cantidad_cuotas_estado_columna_bd_mora": n_mora_estado_columna_bd,
        "conteos_coinciden": n_mora_reporte
        == n_mora_si_sum_cuota_pagos
        == n_mora_estado_columna_bd,
        "tolerancia_alineacion_usd": TOL_ALINEACION_PAGADO_CUOTA,
        "cuotas_desalineadas_pagos": desv,
        "precision": "Valide el Excel de morosidad por cedula con estos numeros. Si difiere cantidad_cuotas_mora_si_sum_cuota_pagos, total_pagado no coincide con SUM(cuota_pagos). Diagnostico por un solo prestamo: GET .../mora-por-prestamo.",
        "cuotas": [
            {
                "cuota_id": r.id,
                "prestamo_id": r.prestamo_id,
                "numero_cuota": r.numero_cuota,
                "monto_usd": round(_safe_float(r.monto), 2),
            }
            for r in detalle
        ],
    }


@router.get("/morosidad/auditoria/mora-por-prestamo")
def get_auditoria_mora_por_prestamo(
    prestamo_id: int = Query(
        ...,
        ge=1,
        description="Solo diagnostico por un prestamo; el reporte de morosidad es por cliente/cedula (ver mora-por-cliente)",
    ),
    db: Session = Depends(get_db),
):
    """Opcional: desglose por un prestamo (amortizacion). El informe oficial es por cedula: usar mora-por-cliente."""

    row_p = db.get(Prestamo, prestamo_id)
    if not row_p:
        raise HTTPException(status_code=404, detail="Prestamo no encontrado")

    c = aliased(Cuota, name="c")
    filtro_mora = _filtro_cuota_solo_estado_mora_sql()

    agg = db.execute(
        select(func.count(c.id), func.coalesce(func.sum(c.monto), 0))
        .select_from(c)
        .where(c.prestamo_id == prestamo_id, filtro_mora)
    ).one()

    filtro_mora_sum_pagos = _filtro_cuota_mora_desde_case_sql(SQL_PG_ESTADO_CUOTA_CASE_CORRELATED)
    n_mora_si_sum_cuota_pagos = int(
        db.execute(
            select(func.count(c.id))
            .select_from(c)
            .where(c.prestamo_id == prestamo_id, filtro_mora_sum_pagos)
        ).scalar_one()
        or 0
    )

    n_mora_estado_columna_bd = int(
        db.execute(
            select(func.count(Cuota.id)).where(
                Cuota.prestamo_id == prestamo_id,
                Cuota.estado == "MORA",
            )
        ).scalar_one()
        or 0
    )

    detalle = db.execute(
        select(c.id, c.numero_cuota, c.monto)
        .select_from(c)
        .where(c.prestamo_id == prestamo_id, filtro_mora)
        .order_by(c.numero_cuota.asc())
    ).fetchall()

    sum_aplicado_sq = (
        select(func.coalesce(func.sum(CuotaPago.monto_aplicado), 0))
        .where(CuotaPago.cuota_id == Cuota.id)
        .correlate(Cuota)
        .scalar_subquery()
    )
    alin_rows = db.execute(
        select(
            Cuota.id,
            Cuota.numero_cuota,
            Cuota.total_pagado,
            sum_aplicado_sq.label("sum_aplicado"),
        )
        .where(Cuota.prestamo_id == prestamo_id)
        .order_by(Cuota.numero_cuota.asc())
    ).fetchall()
    desv: List[dict] = []
    for ar in alin_rows:
        col = _safe_float(ar.total_pagado)
        sap = _safe_float(ar.sum_aplicado)
        if abs(col - sap) > TOL_ALINEACION_PAGADO_CUOTA:
            desv.append(
                {
                    "cuota_id": ar.id,
                    "numero_cuota": ar.numero_cuota,
                    "total_pagado_columna": round(col, 2),
                    "sum_cuota_pagos": round(sap, 2),
                    "diff_usd": round(col - sap, 2),
                }
            )

    n_mora_reporte = int(agg[0] or 0)

    return {
        "alcance": "diagnostico_por_prestamo",
        "reporte_morosidad_usar": "GET /reportes/morosidad/auditoria/mora-por-cliente (Excel va por cedula, no por prestamo)",
        "prestamo_id": prestamo_id,
        "cedula_prestamo": row_p.cedula or "",
        "cantidad_cuotas_mora": n_mora_reporte,
        "total_monto_usd": round(_safe_float(agg[1]), 2),
        "criterio": "Un solo prestamo: estado MORA con total_pagado (como lista de cuotas).",
        "cantidad_cuotas_mora_si_sum_cuota_pagos": n_mora_si_sum_cuota_pagos,
        "cantidad_cuotas_estado_columna_bd_mora": n_mora_estado_columna_bd,
        "conteos_coinciden": n_mora_reporte == n_mora_si_sum_cuota_pagos == n_mora_estado_columna_bd,
        "tolerancia_alineacion_usd": TOL_ALINEACION_PAGADO_CUOTA,
        "cuotas_desalineadas_pagos": desv,
        "precision": "Auxiliar para contrastar una tabla de amortizacion. El reporte de morosidad agrega por cliente: mora-por-cliente.",
        "cuotas": [
            {
                "cuota_id": r.id,
                "numero_cuota": r.numero_cuota,
                "monto_usd": round(_safe_float(r.monto), 2),
            }
            for r in detalle
        ],
    }


@router.get("/morosidad/clientes")

def get_morosidad_clientes(

    db: Session = Depends(get_db),

    fecha_corte: Optional[str] = Query(None),

):

    """Lista por cedula: agrega todas las cuotas MORA del cliente (todos sus prestamos APROBADO).

    Mismo filtro que exportar/morosidad-cedulas (solo estado calculado MORA).
    """

    fc = _parse_fecha(fecha_corte)

    c = aliased(Cuota, name="c")

    filtro_mora = _filtro_cuota_solo_estado_mora_sql()

    q = (

        select(

            Cliente.id.label("cliente_id"),

            Cliente.nombres.label("nombre"),

            Cliente.cedula.label("cedula"),

            func.count(c.id).label("cantidad_cuotas"),

            func.coalesce(func.sum(c.monto), 0).label("total_usd"),

        )

        .select_from(c)

        .join(Prestamo, c.prestamo_id == Prestamo.id)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado == "APROBADO",

            filtro_mora,

        )

        .group_by(Cliente.id, Cliente.nombres, Cliente.cedula)

        .order_by(func.coalesce(func.sum(c.monto), 0).desc())

    )

    rows = db.execute(q).fetchall()

    return {

        "fecha_corte": fc.isoformat(),

        "solo_estados_cuota": list(REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS),

        "excluye_estados_cuota": list(REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS),

        "criterio": "Solo cuotas con estado calculado MORA (4+ meses; excluye VENCIDO y demas).",

        "items": [

            {

                "nombre": r.nombre or "",

                "cedula": r.cedula or "",

                "cantidad_cuotas_morosidad": int(r.cantidad_cuotas or 0),

                "total_usd_morosidad": round(_safe_float(r.total_usd), 2),

            }

            for r in rows

        ],

    }





@router.get("/exportar/morosidad-clientes")

def exportar_morosidad_clientes(

    db: Session = Depends(get_db),

    fecha_corte: Optional[str] = Query(None),

):

    """Exporta Excel: Nombre, Cedula, cantidad/total solo cuotas estado MORA (no VENCIDO)."""

    import time

    import openpyxl

    t0 = time.perf_counter()

    data = get_morosidad_clientes(db=db, fecha_corte=fecha_corte)

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Solo estado MORA"

    ws.append(
        [
            "Nombre",
            "C\u00e9dula",
            "Cantidad cuotas (solo MORA)",
            "Total USD (solo MORA)",
        ]
    )

    total_cuotas = 0

    total_monto = 0.0

    for r in data["items"]:

        n = int(r.get("cantidad_cuotas_morosidad", 0))

        m = round(_safe_float(r.get("total_usd_morosidad", 0)), 2)

        total_cuotas += n

        total_monto += m

        ws.append(

            [

                r.get("nombre", ""),

                r.get("cedula", ""),

                n,

                m,

            ]

        )

    ws.append(["TOTAL", "", total_cuotas, round(total_monto, 2)])

    buf = io.BytesIO()

    wb.save(buf)

    content = buf.getvalue()

    hoy_str = data["fecha_corte"]

    n_filas = len(data["items"])

    ms = (time.perf_counter() - t0) * 1000

    logger.info(

        "exportar_morosidad_clientes incluye=%s excluye=%s filas_clientes=%s total_cuotas=%s ms=%.1f fecha_corte=%s",

        ",".join(REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS),

        ",".join(REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS),

        n_filas,

        total_cuotas,

        ms,

        hoy_str,

    )

    return Response(

        content=content,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={

            "Content-Disposition": f"attachment; filename=reporte_mora_clientes_{hoy_str}.xlsx",

            "X-Reporte-Morosidad-Clientes": str(n_filas),

            "X-Reporte-Morosidad-Fecha-Corte": hoy_str,

            "X-Reporte-Solo-Estado-Cuota": "MORA",

        },

    )





@router.get("/morosidad")

def get_reporte_morosidad(

    db: Session = Depends(get_db),

    fecha_corte: Optional[str] = Query(None),

):

    """Reporte de pago vencido. Moroso = 4 meses calendario + 1 dia de atraso."""

    fc = _parse_fecha(fecha_corte)

    subq_mora = (

        select(Cuota.prestamo_id)

        .select_from(Cuota)

        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado == "APROBADO",

            Cuota.fecha_pago.is_(None),

            Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

        )

        .distinct()

    )

    total_prestamos_mora = db.scalar(select(func.count()).select_from(subq_mora.subquery())) or 0

    prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]

    total_clientes_mora = 0

    if prestamos_ids:

        total_clientes_mora = db.scalar(

            select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(prestamos_ids))

        ) or 0

    monto_total_mora = _safe_float(

        db.scalar(

            select(func.coalesce(func.sum(Cuota.monto), 0))

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(

                Cliente.estado == "ACTIVO",

                Prestamo.estado == "APROBADO",

                Cuota.fecha_pago.is_(None),

                Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

            )

        )

        or 0

    )

    cuotas_mora = db.execute(

        select(Cuota.prestamo_id, Cuota.fecha_vencimiento)

        .select_from(Cuota)

        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado == "APROBADO",

            Cuota.fecha_pago.is_(None),

            Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

        )

    ).fetchall()

    dias_list = [(fc - r.fecha_vencimiento).days for r in cuotas_mora]

    promedio_dias_mora = sum(dias_list) / len(dias_list) if dias_list else 0

    distribucion_por_rango: List[dict] = []

    morosidad_por_analista: List[dict] = []

    analistas = db.execute(select(Prestamo.analista).where(Prestamo.id.in_(prestamos_ids)).distinct()).fetchall()

    for (analista,) in analistas:

        if not analista:

            continue

        ids_ana = db.execute(select(Prestamo.id).where(Prestamo.analista == analista, Prestamo.id.in_(prestamos_ids))).fetchall()

        ids_ana = [x[0] for x in ids_ana]

        monto_ana = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc, Cuota.prestamo_id.in_(ids_ana)))) or 0

        cuotas_ana = db.execute(

            select(Cuota.fecha_vencimiento).select_from(Cuota).where(

                Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc, Cuota.prestamo_id.in_(ids_ana)

            )

        ).fetchall()

        dias_ana = [(fc - r.fecha_vencimiento).days for r in cuotas_ana]

        prom_dias_ana = sum(dias_ana) / len(dias_ana) if dias_ana else 0

        morosidad_por_analista.append({

            "analista": analista,

            "cantidad_prestamos": len(ids_ana),

            "cantidad_clientes": db.scalar(select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(ids_ana))) or 0,

            "monto_total_mora": monto_ana,

            "promedio_dias_mora": prom_dias_ana,

        })

    detalle: List[dict] = []

    for pid in prestamos_ids[:200]:

        p = db.get(Prestamo, pid)

        if not p:

            continue

        monto_mora = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)) or 0)

        cuotas_en_mora = db.scalar(select(func.count()).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)) or 0

        primera = db.execute(select(func.min(Cuota.fecha_vencimiento)).select_from(Cuota).where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc)).scalar()

        max_dias = (fc - primera).days if primera else 0

        detalle.append({

            "prestamo_id": pid,

            "cedula": p.cedula or "",

            "nombres": p.nombres or "",

            "total_financiamiento": _safe_float(p.total_financiamiento),

            "analista": p.analista or "",

            "concesionario": p.concesionario or "",

            "cuotas_en_mora": cuotas_en_mora,

            "monto_total_mora": monto_mora,

            "max_dias_mora": max_dias,

            "primera_cuota_vencida": primera.isoformat() if primera else None,

        })

    return {

        "fecha_corte": fc.isoformat(),

        "total_prestamos_mora": total_prestamos_mora,

        "total_clientes_mora": total_clientes_mora,

        "monto_total_mora": monto_total_mora,

        "promedio_dias_mora": round(promedio_dias_mora, 2),

        "distribucion_por_rango": distribucion_por_rango,

        "morosidad_por_analista": morosidad_por_analista,

        "detalle_prestamos": detalle,

    }





@router.get("/morosidad/por-mes")

def get_morosidad_por_mes(

    db: Session = Depends(get_db),

    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrÃ¡s"),

    anos: Optional[str] = Query(None),

    meses_list: Optional[str] = Query(None),

):

    """Morosidad por mes: una pestaÃ±a por mes. Cuotas vencidas sin pagar a fin de cada mes, agrupadas por cÃ©dula."""

    resultado: dict = {"meses": []}

    periodos = _periodos_desde_filtros(anos, meses_list, meses)



    for (ano, mes) in periodos:

        _, ultimo = calendar.monthrange(ano, mes)

        fc = date(ano, mes, ultimo)



        subq_mora = (

            select(Cuota.prestamo_id)

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(

                Cliente.estado == "ACTIVO",

                Prestamo.estado == "APROBADO",

                Cuota.fecha_pago.is_(None),

                Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

            )

            .distinct()

        )

        prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]

        total_prestamos_mora = len(prestamos_ids)

        total_clientes_mora = 0

        monto_total_mora = 0.0

        if prestamos_ids:

            total_clientes_mora = db.scalar(

                select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(prestamos_ids))

            ) or 0

            monto_total_mora = _safe_float(

                db.scalar(

                    select(func.coalesce(func.sum(Cuota.monto), 0))

                    .select_from(Cuota)

                    .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

                    .join(Cliente, Prestamo.cliente_id == Cliente.id)

                    .where(

                        Cliente.estado == "ACTIVO",

                        Prestamo.estado == "APROBADO",

                        Cuota.fecha_pago.is_(None),

                        Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

                    )

                )

            ) or 0



        cuotas_mora = db.execute(

            select(Cuota.prestamo_id, Cuota.fecha_vencimiento)

            .select_from(Cuota)

            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

            .join(Cliente, Prestamo.cliente_id == Cliente.id)

            .where(

                Cliente.estado == "ACTIVO",

                Prestamo.estado == "APROBADO",

                Cuota.fecha_pago.is_(None),

                Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

            )

        ).fetchall()

        dias_list = [(fc - r.fecha_vencimiento).days for r in cuotas_mora]

        promedio_dias_mora = sum(dias_list) / len(dias_list) if dias_list else 0



        morosidad_por_cedula: List[dict] = []

        if prestamos_ids:

            monto_por_prestamo_rows = db.execute(

                select(Cuota.prestamo_id, func.coalesce(func.sum(Cuota.monto), 0).label("monto"))

                .select_from(Cuota)

                .where(

                    Cuota.prestamo_id.in_(prestamos_ids),

                    Cuota.fecha_pago.is_(None),

                    Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

                )

                .group_by(Cuota.prestamo_id)

            ).fetchall()

            monto_por_prestamo = {r.prestamo_id: _safe_float(r.monto) for r in monto_por_prestamo_rows}

            prestamos_data = db.execute(

                select(Prestamo.id, Prestamo.cedula, Prestamo.nombres)

                .where(Prestamo.id.in_(prestamos_ids), Prestamo.cedula.isnot(None), Prestamo.cedula != "")

            ).fetchall()

            pid_to_cedula: dict = {}

            cedula_agg: dict = {}

            for pid, cedula, nombres in prestamos_data:

                cedula = cedula or ""

                pid_to_cedula[pid] = cedula

                if cedula not in cedula_agg:

                    cedula_agg[cedula] = {"nombres": nombres or "", "prestamos": set(), "dias": []}

                cedula_agg[cedula]["prestamos"].add(pid)

            for r in cuotas_mora:

                cedula = pid_to_cedula.get(r.prestamo_id)

                if cedula and cedula in cedula_agg:

                    cedula_agg[cedula]["dias"].append((fc - r.fecha_vencimiento).days)

            for cedula, data in cedula_agg.items():

                monto_total = sum(monto_por_prestamo.get(pid, 0) for pid in data["prestamos"])

                prom_dias = sum(data["dias"]) / len(data["dias"]) if data["dias"] else 0

                morosidad_por_cedula.append({

                    "cedula": cedula,

                    "nombres": data["nombres"],

                    "cantidad_prestamos": len(data["prestamos"]),

                    "cantidad_clientes": 1,

                    "monto_total_mora": round(monto_total, 2),

                    "promedio_dias_mora": round(prom_dias, 1),

                })

            morosidad_por_cedula.sort(key=lambda x: -x["monto_total_mora"])



        resultado["meses"].append({

            "mes": mes,

            "ano": ano,

            "label": f"{mes:02d}/{ano}",

            "fecha_corte": fc.isoformat(),

            "total_prestamos_mora": total_prestamos_mora,

            "total_clientes_mora": total_clientes_mora,

            "monto_total_mora": monto_total_mora,

            "promedio_dias_mora": round(promedio_dias_mora, 1),

            "morosidad_por_cedula": morosidad_por_cedula,

        })



    return resultado





@router.get("/morosidad/por-rangos")

def get_morosidad_por_rangos(

    db: Session = Depends(get_db),

    fecha_corte: Optional[str] = Query(None),

):

    """Informe pago vencido por rangos de dÃ­as."""

    fc = _parse_fecha(fecha_corte)

    subq_mora = (

        select(Cuota.prestamo_id)

        .select_from(Cuota)

        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)

        .join(Cliente, Prestamo.cliente_id == Cliente.id)

        .where(

            Cliente.estado == "ACTIVO",

            Prestamo.estado == "APROBADO",

            Cuota.fecha_pago.is_(None),

            Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

        )

        .distinct()

    )

    prestamos_ids = [r[0] for r in db.execute(subq_mora).fetchall()]



    resultado: dict = {"fecha_corte": fc.isoformat(), "rangos": {}}



    def _fecha_iso(d):

        if d is None:

            return None

        if hasattr(d, "date"):

            d = d.date() if callable(getattr(d, "date", None)) else d

        return d.isoformat()[:10] if hasattr(d, "isoformat") else str(d)[:10]



    for rango in RANGOS_ATRASO:

        min_d, max_d = rango["min_dias"], rango["max_dias"]

        items: List[dict] = []



        for pid in prestamos_ids:

            p = db.get(Prestamo, pid)

            if not p:

                continue



            primera = db.execute(

                select(func.min(Cuota.fecha_vencimiento))

                .select_from(Cuota)

                .where(

                    Cuota.prestamo_id == pid,

                    Cuota.fecha_pago.is_(None),

                    Cuota.fecha_vencimiento + text("INTERVAL '4 months 1 day'") <= fc,

                )

            ).scalar()

            if not primera:

                continue

            dias_atraso = (fc - primera).days

            if dias_atraso < min_d or dias_atraso > max_d:

                continue



            saldo = _safe_float(

                db.scalar(

                    select(func.coalesce(func.sum(Cuota.monto), 0))

                    .select_from(Cuota)

                    .where(

                        Cuota.prestamo_id == pid,

                        Cuota.fecha_pago.is_(None),

                    )

                )

            ) or 0



            pagos_totales = _safe_float(

                db.scalar(

                    select(func.coalesce(func.sum(Cuota.total_pagado), 0))

                    .select_from(Cuota)

                    .where(Cuota.prestamo_id == pid)

                )

            ) or 0



            ultimo_pago = db.execute(

                select(func.max(Cuota.fecha_pago))

                .select_from(Cuota)

                .where(Cuota.prestamo_id == pid, Cuota.fecha_pago.isnot(None))

            ).scalar()



            proximo_pago = db.execute(

                select(func.min(Cuota.fecha_vencimiento))

                .select_from(Cuota)

                .where(Cuota.prestamo_id == pid, Cuota.fecha_pago.is_(None))

            ).scalar()



            items.append({

                "prestamo_id": pid,

                "cedula": p.cedula or "",

                "nombres": p.nombres or "",

                "total_financiamiento": _safe_float(p.total_financiamiento),

                "pagos_totales": pagos_totales,

                "saldo": saldo,

                "ultimo_pago_fecha": _fecha_iso(ultimo_pago),

                "proximo_pago_fecha": _fecha_iso(proximo_pago),

                "dias_atraso": dias_atraso,

            })



        resultado["rangos"][rango["key"]] = {

            "label": rango["label"],

            "items": items,

        }



    return resultado





def _generar_excel_morosidad_por_mes(data_por_mes: dict) -> bytes:

    """Genera Excel con una pestaÃ±a por mes."""

    import openpyxl



    wb = openpyxl.Workbook()

    meses_data = data_por_mes.get("meses", [])

    meses_es = {

        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",

        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"

    }



    for idx, mes_data in enumerate(meses_data):

        mes = mes_data.get("mes", 1)

        ano = mes_data.get("ano", date.today().year)

        mes_nombre = meses_es.get(mes, "")

        label = mes_data.get("label", f"{mes:02d}/{ano}")

        sheet_name = f"{mes_nombre} {ano}"[:31]



        if idx == 0:

            ws = wb.active

            ws.title = sheet_name

        else:

            ws = wb.create_sheet(title=sheet_name)



        ws.append(["Informe de Vencimiento de Pagos", label])

        ws.append([])

        ws.append(["Total prÃ©stamos con pago vencido", mes_data.get("total_prestamos_mora", 0)])

        ws.append(["Total clientes con pago vencido", mes_data.get("total_clientes_mora", 0)])

        ws.append(["Monto total en dÃ³lares", mes_data.get("monto_total_mora", 0)])

        ws.append(["Promedio dÃ­as de atraso", round(_safe_float(mes_data.get("promedio_dias_mora", 0)), 1)])

        ws.append([])

        ws.append(["Pago vencido por cÃ©dula"])

        ws.append(["CÃ©dula", "Nombre", "Cant. prÃ©stamos", "Cant. clientes", "Monto en dÃ³lares", "Prom. dÃ­as"])

        for r in mes_data.get("morosidad_por_cedula", []):

            prom_dias = round(_safe_float(r.get("promedio_dias_mora", 0)), 1)

            ws.append([

                r.get("cedula", ""),

                r.get("nombres", ""),

                r.get("cantidad_prestamos", 0),

                r.get("cantidad_clientes", 0),

                r.get("monto_total_mora", 0),

                prom_dias

            ])



    buf = io.BytesIO()

    wb.save(buf)

    return buf.getvalue()





def _generar_pdf_morosidad(data: dict) -> bytes:

    from reportlab.lib.pagesizes import letter

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=letter)

    styles = getSampleStyleSheet()

    story = []

    story.append(Paragraph("Informe de Vencimiento de Pagos", styles["Title"]))

    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))

    story.append(Spacer(1, 12))

    resumen = [

        ["Total prÃ©stamos con pago vencido", str(data.get("total_prestamos_mora", 0))],

        ["Total clientes con pago vencido", str(data.get("total_clientes_mora", 0))],

        ["Monto total en dÃ³lares", str(data.get("monto_total_mora", 0))],

        ["Promedio dÃ­as de atraso", str(round(_safe_float(data.get("promedio_dias_mora", 0)), 1))],

    ]

    t = Table(resumen)

    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))

    story.append(t)

    mora_analista = data.get("morosidad_por_analista", [])

    if mora_analista:

        story.append(Spacer(1, 12))

        story.append(Paragraph("Pago vencido por analista", styles["Heading2"]))

        rows = [["Analista", "PrÃ©stamos", "Clientes", "Monto en dÃ³lares", "Prom. dÃ­as"]]

        for r in mora_analista:

            prom_dias = round(_safe_float(r.get("promedio_dias_mora", 0)), 1)

            rows.append([r.get("analista", ""), str(r.get("cantidad_prestamos", 0)), str(r.get("cantidad_clientes", 0)), str(r.get("monto_total_mora", 0)), str(prom_dias)])

        t2 = Table(rows)

        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))

        story.append(t2)

    doc.build(story)

    return buf.getvalue()





@router.get("/exportar/morosidad")

def exportar_morosidad(

    db: Session = Depends(get_db),

    formato: str = Query("excel", pattern="^(excel|pdf)$"),

    fecha_corte: Optional[str] = Query(None),

    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses"),

    anos: Optional[str] = Query(None),

    meses_list: Optional[str] = Query(None),

):

    """Exporta reporte de morosidad. Excel: una pestaÃ±a por mes. PDF: fecha de corte Ãºnica."""

    if formato == "pdf":

        data = get_reporte_morosidad(db=db, fecha_corte=fecha_corte)

        content = _generar_pdf_morosidad(data)

        return Response(

            content=content,

            media_type="application/pdf",

            headers={"Content-Disposition": f"attachment; filename=informe_vencimiento_pagos_{data['fecha_corte']}.pdf"},

        )

    data_por_mes = get_morosidad_por_mes(db=db, meses=meses, anos=anos, meses_list=meses_list)

    content = _generar_excel_morosidad_por_mes(data_por_mes)

    hoy_str = date.today().isoformat()

    return Response(

        content=content,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f"attachment; filename=informe_vencimiento_pagos_{hoy_str}.xlsx"},

    )


@router.get("/exportar/morosidad-cedulas")
def exportar_morosidad_cedulas(db: Session = Depends(get_db)):
    """Una fila por cedula (cliente): cuenta y suma TODAS las cuotas con estado MORA.

    Incluye cuotas MORA de todos los prestamos APROBADO del cliente (no solo un prestamo).
    Cliente ACTIVO. Etiqueta UI: 'Mora (4 meses+)'. VENCIDO excluido.
    Corte: hoy America/Caracas.
    """
    import time

    import openpyxl

    t0 = time.perf_counter()

    # Alias `c` debe coincidir con el SQL embebido (referencias c.id, c.monto_cuota, ...).
    c = aliased(Cuota, name="c")
    estado_cuota_en_morosidad = _filtro_cuota_solo_estado_mora_sql()
    modalidad_por_cliente = literal_column(
        "string_agg(DISTINCT prestamos.modalidad_pago, ', ' "
        "ORDER BY prestamos.modalidad_pago)"
    ).label("modalidad")

    rows_resumen = db.execute(
        select(
            Cliente.id.label("cliente_id"),
            Cliente.cedula.label("cedula"),
            modalidad_por_cliente,
            func.count(c.id).label("total_cuotas"),
            func.coalesce(func.sum(c.monto), 0).label("total_monto"),
        )
        .select_from(c)
        .join(Prestamo, c.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            estado_cuota_en_morosidad,
        )
        .group_by(Cliente.id, Cliente.cedula)
        .order_by(Cliente.cedula.asc())
    ).fetchall()

    fc = hoy_negocio()
    hoy_str = fc.isoformat()

    wb = openpyxl.Workbook()
    ws = wb.active
    # Estructura fija del informe por cedula: una fila por persona; no renombrar columnas ni pestana.
    ws.title = "Morosidad"
    ws.append(
        [
            "C\u00e9dula",
            "Modalidad",
            "Cantidad de cuotas",
            "Total USD",
        ]
    )

    total_cuotas_general = 0
    total_monto_general = 0.0
    for r in rows_resumen:
        cuotas = int(r.total_cuotas or 0)
        monto = round(_safe_float(r.total_monto), 2)
        total_cuotas_general += cuotas
        total_monto_general += monto
        ws.append(
            [
                r.cedula or "",
                r.modalidad or "",
                cuotas,
                monto,
            ]
        )

    ws.append(
        [
            "TOTAL",
            "",
            total_cuotas_general,
            round(total_monto_general, 2),
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    ms = (time.perf_counter() - t0) * 1000
    logger.info(
        "exportar_morosidad_cedulas incluye=%s excluye=%s filas_clientes=%s total_cuotas=%s ms=%.1f fecha_corte=%s",
        ",".join(REPORTE_MOROSIDAD_ESTADOS_INCLUIDOS),
        ",".join(REPORTE_MOROSIDAD_ESTADOS_EXCLUIDOS),
        len(rows_resumen),
        total_cuotas_general,
        ms,
        hoy_str,
    )
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_morosidad_{hoy_str}.xlsx",
            "X-Reporte-Morosidad-Clientes": str(len(rows_resumen)),
            "X-Reporte-Morosidad-Fecha-Corte": hoy_str,
        },
    )

