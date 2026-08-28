"""
Reportes de cartera.
"""
import calendar
import io
from datetime import date, timedelta
from typing import List, Optional, Set

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy import and_, case, func, or_, select, text
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.cliente import Cliente
from app.models.cuota import Cuota
from app.models.prestamo import Prestamo
from app.models.pago import Pago
from app.models.cuota_pago import CuotaPago

from app.api.v1.endpoints.reportes_utils import _safe_float, _parse_fecha, _periodos_desde_filtros
from app.utils.cedula_almacenamiento import expr_cedula_normalizada_para_comparar
from app.services.cuota_estado import SQL_PG_INTERVAL_INICIO_MORA

router = APIRouter(dependencies=[Depends(get_current_user)])


def _datos_cartera(db: Session, fecha_corte: date) -> dict:
    """Obtiene datos para reporte de cartera a una fecha de corte (solo clientes ACTIVOS)."""
    cuotas_pendientes = (
        select(func.coalesce(func.sum(Cuota.monto), 0))
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
        )
    )
    cartera_total = _safe_float(db.scalar(cuotas_pendientes) or 0)

    prestamos_activos = db.scalar(
        select(func.count())
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
    ) or 0

    subq_mora = (
        select(Cuota.prestamo_id)
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento + text(SQL_PG_INTERVAL_INICIO_MORA) <= fecha_corte,
        )
        .distinct()
    )
    prestamos_mora = db.scalar(select(func.count()).select_from(subq_mora.subquery())) or 0

    mora_total = _safe_float(
        db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento + text(SQL_PG_INTERVAL_INICIO_MORA) <= fecha_corte,
            )
        )
        or 0
    )

    distribucion_por_monto = [
        {"rango": "0 - 5.000", "cantidad": 0, "monto": 0},
        {"rango": "5.001 - 15.000", "cantidad": 0, "monto": 0},
        {"rango": "15.001 - 50.000", "cantidad": 0, "monto": 0},
        {"rango": "> 50.000", "cantidad": 0, "monto": 0},
    ]
    saldos = db.execute(
        select(Cuota.prestamo_id, func.sum(Cuota.monto).label("saldo"))
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cliente.estado == "ACTIVO",
            Prestamo.estado == "APROBADO",
            Cuota.fecha_pago.is_(None),
        )
        .group_by(Cuota.prestamo_id)
    ).all()
    for pid, saldo in saldos:
        s = _safe_float(saldo)
        if s <= 5000:
            distribucion_por_monto[0]["cantidad"] += 1
            distribucion_por_monto[0]["monto"] += s
        elif s <= 15000:
            distribucion_por_monto[1]["cantidad"] += 1
            distribucion_por_monto[1]["monto"] += s
        elif s <= 50000:
            distribucion_por_monto[2]["cantidad"] += 1
            distribucion_por_monto[2]["monto"] += s
        else:
            distribucion_por_monto[3]["cantidad"] += 1
            distribucion_por_monto[3]["monto"] += s

    distribucion_por_mora: List[dict] = []
    for label, dias_min, dias_max in [
        ("1-30 días", 1, 30),
        ("31-60 días", 31, 60),
        ("61-89 días", 61, 89),
        ("4+ meses (moroso)", 121, 9999),
    ]:
        delta_min = fecha_corte - timedelta(days=dias_max)
        delta_max = fecha_corte - timedelta(days=dias_min)
        q = (
            select(func.count(Cuota.id), func.coalesce(func.sum(Cuota.monto), 0))
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento <= delta_max,
                Cuota.fecha_vencimiento >= delta_min,
            )
        )
        row = db.execute(q).one_or_none()
        cnt = (row[0] or 0) if row else 0
        monto = _safe_float(row[1] or 0) if row else 0
        distribucion_por_mora.append({"rango": label, "cantidad": cnt, "monto_total": monto})

    return {
        "fecha_corte": fecha_corte.isoformat(),
        "cartera_total": cartera_total,
        "capital_pendiente": cartera_total,
        "intereses_pendientes": 0,
        "mora_total": mora_total,
        "cantidad_prestamos_activos": prestamos_activos,
        "cantidad_prestamos_mora": prestamos_mora,
        "distribucion_por_monto": distribucion_por_monto,
        "distribucion_por_mora": distribucion_por_mora,
    }


def _cartera_por_periodos(db: Session, periodos: List[tuple]) -> dict:
    """Genera datos cartera para lista de (año, mes)."""
    resultado: dict = {"meses": []}
    for (ano, mes) in periodos:
        inicio = date(ano, mes, 1)
        _, ultimo = calendar.monthrange(ano, mes)
        fin = date(ano, mes, ultimo)

        rows = db.execute(
            select(
                func.extract("day", Cuota.fecha_vencimiento).label("dia"),
                func.coalesce(func.sum(Cuota.monto), 0).label("monto_cobrar"),
                func.count(Cuota.id).label("cantidad_cuotas"),
            )
            .select_from(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                Cuota.fecha_pago.is_(None),
                Cuota.fecha_vencimiento >= inicio,
                Cuota.fecha_vencimiento <= fin,
            )
            .group_by(func.extract("day", Cuota.fecha_vencimiento))
            .order_by(func.extract("day", Cuota.fecha_vencimiento))
        ).fetchall()

        por_dia: dict = {}
        for r in rows:
            d = int(r.dia) if r.dia is not None else 0
            por_dia[d] = {
                "cantidad_cuotas": r.cantidad_cuotas or 0,
                "monto_cobrar": round(_safe_float(r.monto_cobrar), 2),
            }

        items: List[dict] = []
        for d in range(1, ultimo + 1):
            data = por_dia.get(d, {"cantidad_cuotas": 0, "monto_cobrar": 0})
            items.append({
                "dia": d,
                "cantidad_cuotas": data["cantidad_cuotas"],
                "monto_cobrar": data["monto_cobrar"],
            })

        resultado["meses"].append({
            "mes": mes,
            "ano": ano,
            "label": f"{mes:02d}/{ano}",
            "items": items,
        })

    return resultado


@router.get("/cartera/por-mes")
def get_cartera_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    anos: Optional[str] = Query(None, description="Años separados por coma, ej: 2023,2024"),
    meses_list: Optional[str] = Query(None, description="Meses 1-12 separados por coma, ej: 1,2,3"),
):
    """Cuentas por cobrar: una pestaña por mes (MM/YYYY). Por día del mes: cuotas por cobrar ese día."""
    periodos = _periodos_desde_filtros(anos, meses_list, meses)
    return _cartera_por_periodos(db, periodos)


@router.get("/cartera")
def get_reporte_cartera(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None, description="Fecha de corte YYYY-MM-DD"),
):
    """Reporte de cartera en JSON. Datos reales desde BD."""
    fc = _parse_fecha(fecha_corte)
    return _datos_cartera(db, fc)


def _generar_excel_cartera_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes (MM/YYYY)."""
    import openpyxl

    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        ano = mes_data.get("ano", date.today().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{ano}")
        sheet_name = f"{mes_nombre} {ano}"[:31]

        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Reporte de Cartera", label])
        ws.append(["Cuotas por cobrar por día del mes"])
        ws.append([])
        ws.append(["Día", "Cuotas por cobrar", "Monto ($)"])

        for item in mes_data.get("items", []):
            row_num = ws.max_row + 1
            ws.append([
                item.get("dia", 0),
                item.get("cantidad_cuotas", 0),
                item.get("monto_cobrar", 0)
            ])
            cell = ws.cell(row=row_num, column=3)
            cell.number_format = '$#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_cartera(data: dict) -> bytes:
    """Excel clásico de cartera (resumen + distribuciones)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera"
    ws.append(["Reporte de Cartera", data.get("fecha_corte", "")])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    ws.append(["Cartera total", data.get("cartera_total", 0)])
    ws.append(["Capital pendiente", data.get("capital_pendiente", 0)])
    ws.append(["Mora total", data.get("mora_total", 0)])
    ws.append(["Préstamos activos", data.get("cantidad_prestamos_activos", 0)])
    ws.append(["Préstamos en mora", data.get("cantidad_prestamos_mora", 0)])
    ws.append([])
    ws.append(["Distribución por monto"])
    ws.append(["Rango", "Cantidad", "Monto"])
    for r in data.get("distribucion_por_monto", []):
        ws.append([r.get("rango", ""), r.get("cantidad", 0), r.get("monto", 0)])
    ws2 = wb.create_sheet("Mora")
    ws2.append(["Rango", "Cantidad", "Monto total"])
    for r in data.get("distribucion_por_mora", []):
        ws2.append([r.get("rango", ""), r.get("cantidad", 0), r.get("monto_total", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_cartera(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Cartera", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Resumen", styles["Heading2"]))
    resumen = [
        ["Cartera total", str(data.get("cartera_total", 0))],
        ["Capital pendiente", str(data.get("capital_pendiente", 0))],
        ["Mora total", str(data.get("mora_total", 0))],
        ["Préstamos activos", str(data.get("cantidad_prestamos_activos", 0))],
        ["Préstamos en mora", str(data.get("cantidad_prestamos_mora", 0))],
    ]
    t = Table(resumen)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    story.append(Spacer(1, 12))
    story.append(Paragraph("Distribución por monto", styles["Heading2"]))
    rows = [["Rango", "Cantidad", "Monto"]]
    for r in data.get("distribucion_por_monto", []):
        rows.append([r.get("rango", ""), str(r.get("cantidad", 0)), str(r.get("monto", 0))])
    t2 = Table(rows)
    t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t2)
    doc.build(story)
    return buf.getvalue()




def _agg_impagas_en_fecha(
    db: Session,
    fecha: date,
    cedulas_norm: Optional[Set[str]] = None,
) -> dict:
    """
    Snapshot a la fecha: cuotas en MORA oficial (mismo umbral que el sistema).

    MORA = vencimiento + 4 meses + 6 días <= fecha, no cubierta al 100% (tol 0.01);
    excluye CANCELADA. No incluye solo VENCIDO.
    Solo prestamos APROBADO (excluye LIQUIDADO, DESISTIMIENTO y demas estados).
    Si cedulas_norm: solo esas cedulas (universo Aseguradora u otro).
    """
    if cedulas_norm is not None and len(cedulas_norm) == 0:
        return {}
    total_pagado_n = func.coalesce(Cuota.total_pagado, 0)
    impaga = and_(
        total_pagado_n < (Cuota.monto - 0.01),
        Cuota.estado.is_distinct_from("CANCELADA"),
    )
    saldo_cuota = func.greatest(Cuota.monto - total_pagado_n, 0)
    prestamo_aprobado = func.upper(func.trim(Prestamo.estado)) == "APROBADO"
    where_parts = [
        Cliente.estado == "ACTIVO",
        prestamo_aprobado,
        impaga,
        Cuota.fecha_vencimiento + text(SQL_PG_INTERVAL_INICIO_MORA) <= fecha,
    ]
    if cedulas_norm is not None:
        where_parts.append(
            expr_cedula_normalizada_para_comparar(Prestamo.cedula).in_(list(cedulas_norm))
        )
    rows = db.execute(
        select(
            Prestamo.id.label("prestamo_id"),
            Prestamo.cedula,
            Prestamo.nombres,
            func.count(Cuota.id).label("cuotas"),
            func.coalesce(func.sum(saldo_cuota), 0).label("monto"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(*where_parts)
        .group_by(Prestamo.id, Prestamo.cedula, Prestamo.nombres)
    ).fetchall()
    out: dict = {}
    for r in rows:
        out[int(r.prestamo_id)] = {
            "prestamo_id": int(r.prestamo_id),
            "cedula": (r.cedula or "").strip(),
            "nombres": (r.nombres or "").strip(),
            "cuotas": int(r.cuotas or 0),
            "monto": round(_safe_float(r.monto), 2),
        }
    return out



def _agg_impagas_en_fecha_historico(
    db: Session,
    fecha: date,
    cedulas_norm: Optional[Set[str]] = None,
) -> dict:
    """
    Corte historico real a `fecha` (Aseguradora).

    Pagado a la fecha = suma de cuota_pagos cuyo pago.fecha_pago <= fecha
    (excluye ANULADO*/DUPLICADO). Si no hay articulos pero cuota.fecha_pago <= fecha,
    se considera pagada al 100%.

    Impaga: vencimiento <= fecha, no CANCELADA, pagado_asof < monto - 0.01.
    Saldo = max(monto - pagado_asof, 0).

    Incluye APROBADO y LIQUIDADO: si el credito se liquido al terminar de pagar,
    debe seguir apareciendo en el corte (c2=0) para no perder esos casos.
    """
    if cedulas_norm is not None and len(cedulas_norm) == 0:
        return {}

    # Limite exclusivo: todos los timestamps del dia `fecha`
    limite = fecha + timedelta(days=1)
    estado_pago = func.upper(func.trim(func.coalesce(Pago.estado, "")))
    pago_operativo = and_(
        ~estado_pago.like("ANULADO%"),
        estado_pago.is_distinct_from("DUPLICADO"),
    )
    pagado_subq = (
        select(
            CuotaPago.cuota_id.label("cuota_id"),
            func.coalesce(func.sum(CuotaPago.monto_aplicado), 0).label("pagado_asof"),
        )
        .select_from(CuotaPago)
        .join(Pago, Pago.id == CuotaPago.pago_id)
        .where(
            Pago.fecha_pago < limite,
            pago_operativo,
        )
        .group_by(CuotaPago.cuota_id)
        .subquery()
    )

    pagado_join = func.coalesce(pagado_subq.c.pagado_asof, 0)
    # Legacy: cuota marcada pagada en/antes de la fecha sin filas en cuota_pagos
    pagado_asof = case(
        (
            and_(
                pagado_join <= 0.009,
                Cuota.fecha_pago.is_not(None),
                Cuota.fecha_pago <= fecha,
            ),
            Cuota.monto,
        ),
        else_=pagado_join,
    )
    impaga = and_(
        pagado_asof < (Cuota.monto - 0.01),
        Cuota.estado.is_distinct_from("CANCELADA"),
    )
    saldo_cuota = func.greatest(Cuota.monto - pagado_asof, 0)
    estado_prestamo = func.upper(func.trim(Prestamo.estado))
    # APROBADO + LIQUIDADO: liquidados terminaron de pagar y deben verse en c2=0.
    prestamo_en_cartera_hist = estado_prestamo.in_(("APROBADO", "LIQUIDADO"))
    where_parts = [
        Cliente.estado == "ACTIVO",
        prestamo_en_cartera_hist,
        impaga,
        Cuota.fecha_vencimiento <= fecha,
    ]
    if cedulas_norm is not None:
        where_parts.append(
            expr_cedula_normalizada_para_comparar(Prestamo.cedula).in_(list(cedulas_norm))
        )

    rows = db.execute(
        select(
            Prestamo.id.label("prestamo_id"),
            Prestamo.cedula,
            Prestamo.nombres,
            func.count(Cuota.id).label("cuotas"),
            func.coalesce(func.sum(saldo_cuota), 0).label("monto"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .outerjoin(pagado_subq, pagado_subq.c.cuota_id == Cuota.id)
        .where(*where_parts)
        .group_by(Prestamo.id, Prestamo.cedula, Prestamo.nombres)
    ).fetchall()

    out: dict = {}
    for r in rows:
        out[int(r.prestamo_id)] = {
            "prestamo_id": int(r.prestamo_id),
            "cedula": (r.cedula or "").strip(),
            "nombres": (r.nombres or "").strip(),
            "cuotas": int(r.cuotas or 0),
            "monto": round(_safe_float(r.monto), 2),
        }
    return out


def _ultimo_dia_mes(anio: int, mes: int) -> date:
    return date(anio, mes, calendar.monthrange(anio, mes)[1])


def _pct_variacion(actual: float, base: float) -> Optional[float]:
    """Variacion intermensual %. None = no comparable (base ~0 y actual > 0)."""
    if abs(base) < 0.005:
        if abs(actual) < 0.005:
            return 0.0
        return None
    return round(((actual - base) / abs(base)) * 100.0, 2)


def _serie_mensual_impagas(
    db: Session,
    n_meses: int = 6,
    ref: Optional[date] = None,
    cuotas_impagas_min: int = 1,
    cuotas_impagas_max: int = 15,
    cedulas_norm: Optional[Set[str]] = None,
) -> List[dict]:
    """
    Totales de cartera impaga (APROBADO) a cierre de cada mes.
    Ultimos n_meses hasta el mes de `ref` (hoy por defecto).
    Mes en curso: corte = min(ultimo dia del mes, ref).
    Mismo filtro min/max de cuotas impagas que el detalle (por conteo en ese mes).
    """
    hoy = ref or date.today()
    n = max(1, min(12, int(n_meses)))
    min_n = max(1, min(99, int(cuotas_impagas_min)))
    max_n = max(1, min(99, int(cuotas_impagas_max)))
    if min_n > max_n:
        min_n, max_n = max_n, min_n
    y, m = hoy.year, hoy.month
    periodos: List[tuple] = []
    for _ in range(n):
        periodos.append((y, m))
        m -= 1
        if m < 1:
            m = 12
            y -= 1
    periodos.reverse()

    out: List[dict] = []
    prev_monto: Optional[float] = None
    for anio, mes in periodos:
        corte = _ultimo_dia_mes(anio, mes)
        if corte > hoy:
            corte = hoy
        snap = _agg_impagas_en_fecha(db, corte, cedulas_norm=cedulas_norm)
        filtrados = [
            v
            for v in snap.values()
            if min_n <= int(v.get("cuotas") or 0) <= max_n
        ]
        n_prest = len(filtrados)
        n_cuotas = sum(int(v.get("cuotas") or 0) for v in filtrados)
        monto = round(sum(float(v.get("monto") or 0) for v in filtrados), 2)
        var_pct = None if prev_monto is None else _pct_variacion(monto, prev_monto)
        out.append(
            {
                "anio": anio,
                "mes": mes,
                "periodo": f"{anio}-{mes:02d}",
                "fecha_corte": corte.isoformat(),
                "prestamos": n_prest,
                "cuotas": n_cuotas,
                "monto": monto,
                "var_pct_vs_mes_anterior": var_pct,
                "cuotas_impagas_min": min_n,
                "cuotas_impagas_max": max_n,
            }
        )
        prev_monto = monto
    return out



def _datos_cuentas_por_cobrar(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    cuotas_impagas_min: int,
    cuotas_impagas_max: int,
    cedulas_norm: Optional[Set[str]] = None,
    titulo_informe: str = "Cuentas por cobrar",
    incluir_serie_mensual: bool = True,
    corte_historico: bool = False,
    corte_unico: bool = False,
) -> dict:
    """
    Compara dos cortes en orden cronologico (fecha menor -> fecha mayor).

    En cada fecha (Cuentas por cobrar): cuotas en MORA oficial a esa fecha
    (vencimiento + 4 meses + 6 dias <= corte, saldo pendiente).
    Con corte_historico (Aseguradora/Impagas): cuotas impagas con vencimiento <= fecha.
    Filtro min-max: se aplica al conteo de la fecha mayor (hasta).
    Con corte_historico: tambien entran quienes redujeron impagas
    o terminaron de pagar (c2=0), aunque queden fuera del rango.
    """
    # Tope amplio: incluir prestamos con cualquier cantidad de cuotas en mora.
    min_n = max(1, min(99, int(cuotas_impagas_min)))
    max_n = max(1, min(99, int(cuotas_impagas_max)))
    if min_n > max_n:
        min_n, max_n = max_n, min_n

    agg_fn = (
        _agg_impagas_en_fecha_historico if corte_historico else _agg_impagas_en_fecha
    )

    if corte_unico:
        snap2 = agg_fn(db, fecha_hasta, cedulas_norm=cedulas_norm)
        items_u: List[dict] = []
        tot_c2_u = 0
        tot_m2_u = 0.0
        for pid, b in snap2.items():
            c2 = int(b["cuotas"])
            m2 = float(b["monto"])
            if not (min_n <= c2 <= max_n):
                continue
            tot_c2_u += c2
            tot_m2_u += m2
            items_u.append(
                {
                    "prestamo_id": pid,
                    "cedula": b.get("cedula", ""),
                    "nombres": b.get("nombres", ""),
                    "cuotas_f1": 0,
                    "monto_f1": 0.0,
                    "cuotas_f2": c2,
                    "monto_f2": round(m2, 2),
                    "redujo_impagas": False,
                    "abono_parcial": False,
                }
            )
        items_u.sort(key=lambda x: (x.get("cedula") or "", x.get("prestamo_id") or 0))
        serie_u: List[dict] = []
        if incluir_serie_mensual:
            serie_u = _serie_mensual_impagas(
                db,
                n_meses=6,
                ref=fecha_hasta,
                cuotas_impagas_min=min_n,
                cuotas_impagas_max=max_n,
                cedulas_norm=cedulas_norm,
            )
        return {
            "titulo_informe": titulo_informe,
            "corte_unico": True,
            "fecha_desde": fecha_hasta.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat(),
            "fecha_1": fecha_hasta.isoformat(),
            "fecha_2": fecha_hasta.isoformat(),
            "cuotas_impagas_min": min_n,
            "cuotas_impagas_max": max_n,
            "cantidad_prestamos": len(items_u),
            "total_cuotas_f1": 0,
            "total_monto_f1": 0.0,
            "total_cuotas_f2": tot_c2_u,
            "total_monto_f2": round(tot_m2_u, 2),
            "serie_mensual": serie_u,
            "universo_cedulas": len(cedulas_norm) if cedulas_norm is not None else None,
            "corte_historico": bool(corte_historico),
            "items": items_u,
        }

    snap1 = agg_fn(db, fecha_desde, cedulas_norm=cedulas_norm)
    snap2 = agg_fn(db, fecha_hasta, cedulas_norm=cedulas_norm)
    ids = set(snap1.keys()) | set(snap2.keys())

    items: List[dict] = []
    tot_c1 = tot_c2 = 0
    tot_m1 = tot_m2 = 0.0
    for pid in ids:
        a = snap1.get(pid)
        b = snap2.get(pid)
        base = a or b or {}
        c1 = int(a["cuotas"]) if a else 0
        m1 = float(a["monto"]) if a else 0.0
        c2 = int(b["cuotas"]) if b else 0
        m2 = float(b["monto"]) if b else 0.0
        en_filtro = min_n <= c2 <= max_n
        redujo_impagas = c2 < c1  # incluye termino (c2=0)
        abono_parcial = (not redujo_impagas) and (m2 + 0.009 < m1)
        # Filtro 1-15 en F2; con historico: tambien quien bajo/termino o abono parcial.
        if corte_historico:
            if not (en_filtro or redujo_impagas or abono_parcial):
                continue
        elif not en_filtro:
            continue
        tot_c1 += c1
        tot_c2 += c2
        tot_m1 += m1
        tot_m2 += m2
        items.append(
            {
                "prestamo_id": pid,
                "cedula": base.get("cedula", ""),
                "nombres": base.get("nombres", ""),
                "cuotas_f1": c1,
                "monto_f1": round(m1, 2),
                "cuotas_f2": c2,
                "monto_f2": round(m2, 2),
                "redujo_impagas": redujo_impagas,
                "abono_parcial": abono_parcial,
            }
        )

    def _orden_item(x: dict):
        # Primero bajaron cuotas (mayor baja primero); luego bajaron solo monto; resto.
        c1 = int(x.get("cuotas_f1") or 0)
        c2 = int(x.get("cuotas_f2") or 0)
        m1 = float(x.get("monto_f1") or 0)
        m2 = float(x.get("monto_f2") or 0)
        if c2 < c1:
            grupo = 0
            score = float(c1 - c2)
        elif m2 + 0.009 < m1:
            grupo = 1
            score = m1 - m2
        else:
            grupo = 2
            score = 0.0
        return (grupo, -score, x.get("cedula") or "", x.get("prestamo_id") or 0)

    items.sort(key=_orden_item)
    # Panorama 6 meses (opcional; Aseguradora lo omite).
    serie_mensual: List[dict] = []
    if incluir_serie_mensual:
        serie_mensual = _serie_mensual_impagas(
            db,
            n_meses=6,
            ref=fecha_hasta,
            cuotas_impagas_min=min_n,
            cuotas_impagas_max=max_n,
            cedulas_norm=cedulas_norm,
        )
    return {
        "titulo_informe": titulo_informe,
        "corte_unico": False,
        "fecha_desde": fecha_desde.isoformat(),
        "fecha_hasta": fecha_hasta.isoformat(),
        "fecha_1": fecha_desde.isoformat(),
        "fecha_2": fecha_hasta.isoformat(),
        "cuotas_impagas_min": min_n,
        "cuotas_impagas_max": max_n,
        "cantidad_prestamos": len(items),
        "total_cuotas_f1": tot_c1,
        "total_monto_f1": round(tot_m1, 2),
        "total_cuotas_f2": tot_c2,
        "total_monto_f2": round(tot_m2, 2),
        "serie_mensual": serie_mensual,
        "universo_cedulas": len(cedulas_norm) if cedulas_norm is not None else None,
        "corte_historico": bool(corte_historico),
        "items": items,
    }


def _generar_excel_cuentas_por_cobrar(data: dict) -> bytes:
    """Excel en filas (corrido hacia abajo)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    titulo = (data.get("titulo_informe") or "Cuentas por cobrar").strip()
    ws.title = titulo[:31]
    f1 = data.get("fecha_1") or data.get("fecha_desde", "")
    f2 = data.get("fecha_2") or data.get("fecha_hasta", "")

    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    meta_font = Font(name="Calibri", size=10, color="44546A")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill_l = PatternFill("solid", fgColor="1F4E79")
    header_fill_r = PatternFill("solid", fgColor="2E75B6")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    ws.append([titulo])
    ws["A1"].font = title_font
    univ = data.get("universo_cedulas")
    univ_part = f"   |   Universo hoja: {univ} cedulas" if univ is not None else ""
    hist_part = "   |   Corte historico (APROBADO+LIQUIDADO, pagos por fecha_pago)" if data.get("corte_historico") else ""
    corte_unico = bool(data.get("corte_unico"))
    filtro_label = (
        "Filtro cuotas impagas"
        if data.get("corte_historico")
        else "Filtro cuotas en mora"
    )
    if corte_unico:
        ws.append(
            [
                f"Corte hasta: {f2}   |   "
                f"{filtro_label}: {data.get('cuotas_impagas_min')}-{data.get('cuotas_impagas_max')}"
                f"{univ_part}{hist_part}"
            ]
        )
    else:
        ws.append(
            [
                f"Desde (corte): {f1}   |   Hasta (corte): {f2}   |   "
                f"{filtro_label}: {data.get('cuotas_impagas_min')}-{data.get('cuotas_impagas_max')}"
                f"{univ_part}{hist_part}"
            ]
        )
    ws["A2"].font = meta_font
    if not corte_unico:
        ws.append(
            [
                f"Prestamos: {data.get('cantidad_prestamos', 0)}   |   "
                f"F1: {data.get('total_cuotas_f1', 0)} cuotas / ${data.get('total_monto_f1', 0):,.2f}   |   "
                f"F2: {data.get('total_cuotas_f2', 0)} cuotas / ${data.get('total_monto_f2', 0):,.2f}"
            ]
        )
        ws["A3"].font = meta_font
        ws.append(
            [
                "Colores: verde = redujo impagas o termino de pagar; "
                "naranja = abono parcial (bajo monto, mismas cuotas)."
            ]
        )
        ws["A4"].font = meta_font
        ws.append([])
    if corte_unico:
        # Quitar filas A3 (totales dual) y A4 (colores) del flujo dual.
        while ws.max_row > 2:
            ws.delete_rows(3)
        ws.append(
            [
                f"Prestamos: {data.get('cantidad_prestamos', 0)}   |   "
                f"Cuotas en mora: {data.get('total_cuotas_f2', 0)}   |   "
                f"Monto mora: ${data.get('total_monto_f2', 0):,.2f}"
            ]
        )
        ws["A3"].font = meta_font
        ws.append([])
        headers = [
            "Cedula",
            "Cliente",
            f"Cuotas en mora ({f2})",
            f"Monto mora ({f2})",
        ]
        header_row = 5
    else:
        cuotas_hdr = "Cuotas impagas" if data.get("corte_historico") else "Cuotas en mora"
        monto_hdr = "Monto impagas" if data.get("corte_historico") else "Monto mora"
        headers = [
            "Cedula",
            "Cliente",
            f"{cuotas_hdr} desde ({f1})",
            f"{monto_hdr} desde ({f1})",
            f"{cuotas_hdr} hasta ({f2})",
            f"{monto_hdr} hasta ({f2})",
        ]
        header_row = 6
    ws.append(headers)
    for col, cell in enumerate(ws[header_row], start=1):
        cell.font = header_font
        if corte_unico:
            cell.fill = header_fill_l if col <= 2 else header_fill_r
        else:
            cell.fill = header_fill_l if col <= 4 else header_fill_r
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    fill_baja = PatternFill("solid", fgColor="E8F5E9")  # verde: redujo cuotas / termino
    fill_abono = PatternFill("solid", fgColor="FFF3E0")  # naranja: abono parcial
    for item in data.get("items", []):
        row_num = ws.max_row + 1
        if corte_unico:
            ws.append(
                [
                    item.get("cedula", ""),
                    item.get("nombres", ""),
                    item.get("cuotas_f2", 0),
                    item.get("monto_f2", 0),
                ]
            )
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).border = thin
            ws.cell(row=row_num, column=4).number_format = '"$"#,##0.00'
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")
        else:
            ws.append(
                [
                    item.get("cedula", ""),
                    item.get("nombres", ""),
                    item.get("cuotas_f1", 0),
                    item.get("monto_f1", 0),
                    item.get("cuotas_f2", 0),
                    item.get("monto_f2", 0),
                ]
            )
            c1 = int(item.get("cuotas_f1") or 0)
            c2 = int(item.get("cuotas_f2") or 0)
            m1 = float(item.get("monto_f1") or 0)
            m2 = float(item.get("monto_f2") or 0)
            redujo = bool(item.get("redujo_impagas")) if "redujo_impagas" in item else c2 < c1
            abono = (
                bool(item.get("abono_parcial"))
                if "abono_parcial" in item
                else ((not redujo) and (m2 + 0.009 < m1))
            )
            fill = fill_baja if redujo else (fill_abono if abono else None)
            for col in range(1, 7):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin
                if fill is not None:
                    cell.fill = fill
            ws.cell(row=row_num, column=4).number_format = '"$"#,##0.00'
            ws.cell(row=row_num, column=6).number_format = '"$"#,##0.00'
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    if not corte_unico:
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16
    ws.freeze_panes = "A6" if corte_unico else "A7"

    # Hoja: evolucion mensual (omitida si no hay serie, p. ej. Aseguradora)
    if data.get("serie_mensual"):
        ws2 = wb.create_sheet("Evolucion 6 meses")
        ws2.append(["Evolucion mensual - cuentas por cobrar (cuotas en mora a corte)"])
        ws2["A1"].font = title_font
        ws2.append(
            [
                "Ultimos 6 meses hasta el mes de la fecha hasta. "
                f"Filtro cuotas en mora: {data.get('cuotas_impagas_min')}-{data.get('cuotas_impagas_max')} "
                "(mismo que el detalle). Var % = cambio vs mes anterior."
            ]
        )
        ws2["A2"].font = meta_font
        ws2.append([])
        h2 = [
            "Periodo",
            "Fecha corte",
            "Prestamos",
            "Cuotas en mora",
            "Pendiente USD",
            "Var % vs mes ant.",
        ]
        ws2.append(h2)
        for col, cell in enumerate(ws2[4], start=1):
            cell.font = header_font
            cell.fill = header_fill_l
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        for row in data.get("serie_mensual") or []:
            var = row.get("var_pct_vs_mes_anterior")
            if var is None:
                var_txt = "n/d"
            else:
                var_txt = f"{var:+.2f}%"
            rnum = ws2.max_row + 1
            ws2.append(
                [
                    row.get("periodo", ""),
                    row.get("fecha_corte", ""),
                    row.get("prestamos", 0),
                    row.get("cuotas", 0),
                    row.get("monto", 0),
                    var_txt,
                ]
            )
            for col in range(1, 7):
                ws2.cell(row=rnum, column=col).border = thin
            ws2.cell(row=rnum, column=5).number_format = '"$"#,##0.00'
            ws2.cell(row=rnum, column=3).alignment = Alignment(horizontal="center")
            ws2.cell(row=rnum, column=4).alignment = Alignment(horizontal="center")
            ws2.cell(row=rnum, column=6).alignment = Alignment(horizontal="center")
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 16
        ws2.column_dimensions["F"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_cuentas_por_cobrar(data: dict) -> bytes:
    """PDF landscape profesional: encabezado, resumen de cortes y pagina X de Y."""
    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        KeepTogether,
    )

    buf = io.BytesIO()
    f1 = data.get("fecha_1") or data.get("fecha_desde", "")
    f2 = data.get("fecha_2") or data.get("fecha_hasta", "")
    items = list(data.get("items") or [])
    n_prestamos = int(data.get("cantidad_prestamos") or len(items))
    tot_c1 = int(data.get("total_cuotas_f1") or 0)
    tot_m1 = float(data.get("total_monto_f1") or 0)
    tot_c2 = int(data.get("total_cuotas_f2") or 0)
    tot_m2 = float(data.get("total_monto_f2") or 0)
    prestamos_f1 = sum(
        1 for it in items if int(it.get("cuotas_f1") or 0) > 0 or float(it.get("monto_f1") or 0) > 0
    )
    prestamos_f2 = sum(
        1 for it in items if int(it.get("cuotas_f2") or 0) > 0 or float(it.get("monto_f2") or 0) > 0
    )
    delta_m = round(tot_m2 - tot_m1, 2)
    filtro_min = data.get("cuotas_impagas_min")
    filtro_max = data.get("cuotas_impagas_max")
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")

    page = landscape(letter)
    page_w, page_h = page
    top_m = 0.85 * inch
    bottom_m = 0.55 * inch
    side_m = 0.45 * inch

    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=side_m,
        rightMargin=side_m,
        topMargin=top_m,
        bottomMargin=bottom_m,
    )
    styles = getSampleStyleSheet()
    subtitle_style = ParagraphStyle(
        "cpc_sub",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#44546A"),
        leading=11,
        spaceAfter=6,
    )
    summary_label = ParagraphStyle(
        "cpc_sum_lbl",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        textColor=colors.HexColor("#44546A"),
        leading=9,
        alignment=1,
    )
    summary_value = ParagraphStyle(
        "cpc_sum_val",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=colors.HexColor("#1F4E79"),
        leading=12,
        alignment=1,
    )
    summary_hint = ParagraphStyle(
        "cpc_sum_hint",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        textColor=colors.HexColor("#667085"),
        leading=9,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        "cpc_cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#1A1A1A"),
    )
    # Paragraph ignora TEXTCOLOR de TableStyle: header necesita color propio.
    header_cell_style = ParagraphStyle(
        "cpc_header_cell",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=1,
    )
    empty_style = ParagraphStyle(
        "cpc_empty",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        textColor=colors.HexColor("#C00000"),
        alignment=1,
        spaceBefore=20,
    )

    class _NumberedCanvas(pdf_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            pdf_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_chrome(page_count)
                pdf_canvas.Canvas.showPage(self)
            pdf_canvas.Canvas.save(self)

        def _draw_chrome(self, page_count: int) -> None:
            self.saveState()
            # Encabezado
            self.setFillColor(colors.HexColor("#1F4E79"))
            self.rect(0, page_h - 0.42 * inch, page_w, 0.42 * inch, fill=1, stroke=0)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 11)
            self.drawString(
                side_m,
                page_h - 0.27 * inch,
                (data.get("titulo_informe") or "CUENTAS POR COBRAR").upper()[:40],
            )
            self.setFont("Helvetica", 8)
            self.drawRightString(
                page_w - side_m,
                page_h - 0.27 * inch,
                (
                    f"Corte hasta {f2}"
                    if data.get("corte_unico")
                    else f"Cortes {f1}  |  {f2}"
                ),
            )
            self.setStrokeColor(colors.HexColor("#2E75B6"))
            self.setLineWidth(2)
            self.line(0, page_h - 0.42 * inch, page_w, page_h - 0.42 * inch)

            # Pie
            self.setStrokeColor(colors.HexColor("#D0D5DD"))
            self.setLineWidth(0.6)
            self.line(side_m, 0.38 * inch, page_w - side_m, 0.38 * inch)
            self.setFillColor(colors.HexColor("#667085"))
            self.setFont("Helvetica", 7.5)
            self.drawString(side_m, 0.22 * inch, f"Generado: {generado}")
            self.drawCentredString(
                page_w / 2.0,
                0.22 * inch,
                f"Pagina {self._pageNumber} de {page_count}",
            )
            self.drawRightString(
                page_w - side_m,
                0.22 * inch,
                "Confidencial - uso interno",
            )
            self.restoreState()

    story = []
    # Titulo solo en la barra superior del canvas (evitar duplicar "Aseguradora"/titulo).
    univ = data.get("universo_cedulas")
    univ_txt = (
        f" &nbsp;&nbsp;|&nbsp;&nbsp; <b>Universo hoja:</b> {univ} cedulas"
        if univ is not None
        else ""
    )
    hist_txt = (
        " &nbsp;&nbsp;|&nbsp;&nbsp; <b>Corte historico</b> (APROBADO+LIQUIDADO; pagos por fecha_pago)"
        if data.get("corte_historico")
        else ""
    )
    if data.get("corte_unico"):
        story.append(
            Paragraph(
                f"Corte hasta: <b>{f2}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Filtro cuotas impagas: <b>{filtro_min}-{filtro_max}</b>"
                f"{univ_txt}{hist_txt}",
                subtitle_style,
            )
        )
    else:
        story.append(
            Paragraph(
                f"Corte menor (antes): <b>{f1}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Corte mayor (hoy / hasta): <b>{f2}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Filtro cuotas impagas en fecha mayor: <b>{filtro_min}-{filtro_max}</b>"
                f"{univ_txt}{hist_txt}",
                subtitle_style,
            )
        )
        story.append(
            Paragraph(
                "<b>Verde:</b> redujo impagas o termino de pagar "
                "(cuotas en 0 e incluye LIQUIDADO). "
                "<b>Naranja:</b> abono parcial (bajo el pendiente, mismas cuotas).",
                subtitle_style,
            )
        )

    # Resumen profesional (4 tarjetas)
    usable_w = page_w - 2 * side_m
    card_w = usable_w / 4.0
    summary_data = [
        [
            [
                Paragraph("PRESTAMOS EN INFORME", summary_label),
                Paragraph(f"{n_prestamos:,}", summary_value),
                Paragraph("Filtrados por impagas en fecha mayor", summary_hint),
            ],
            [
                Paragraph(f"ANTES · {f1}", summary_label),
                Paragraph(f"${tot_m1:,.2f}", summary_value),
                Paragraph(
                    f"{prestamos_f1:,} prestamos · {tot_c1:,} cuotas pend.",
                    summary_hint,
                ),
            ],
            [
                Paragraph(f"HOY · {f2}", summary_label),
                Paragraph(f"${tot_m2:,.2f}", summary_value),
                Paragraph(
                    f"{prestamos_f2:,} prestamos · {tot_c2:,} cuotas pend.",
                    summary_hint,
                ),
            ],
            [
                Paragraph("VARIACION MONTO", summary_label),
                Paragraph(
                    f"{'+' if delta_m > 0 else ''}{delta_m:,.2f}",
                    summary_value,
                ),
                Paragraph("Pendiente hoy menos pendiente antes", summary_hint),
            ],
        ]
    ]
    # Flatten for Table: each cell is a nested mini-table or flowables list
    summary_row = []
    for cell_flowables in summary_data[0]:
        inner = Table(
            [[flow] for flow in cell_flowables],
            colWidths=[card_w - 8],
        )
        inner.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        summary_row.append(inner)

    sum_tbl = Table([summary_row], colWidths=[card_w] * 4)
    sum_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#EEF2F7")),
                ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#E8F1FB")),
                ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#E8F8F0")),
                ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#FFF4E5")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#1F4E79")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D5DD")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(KeepTogether([sum_tbl]))
    story.append(Spacer(1, 10))

    # Indicadores mes a mes (ultimos 6 meses)
    serie = list(data.get("serie_mensual") or [])
    if serie:
        mes_title = ParagraphStyle(
            "cpc_mes_title",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.HexColor("#1F4E79"),
            spaceAfter=4,
        )
        story.append(
            Paragraph(
                f"Evolucion mensual (ultimos 6 meses) — filtro impagas {filtro_min}-{filtro_max}",
                mes_title,
            )
        )
        story.append(
            Paragraph(
                "Cada fila: saldo impago con vencimiento &lt;= cierre de mes "
                "(mes actual: hasta hoy). Var % = intermensual vs mes anterior.",
                summary_hint,
            )
        )
        story.append(Spacer(1, 4))
        mes_header = [
            Paragraph("<b>Periodo</b>", cell_style),
            Paragraph("<b>Corte</b>", cell_style),
            Paragraph("<b>Prestamos</b>", cell_style),
            Paragraph("<b>Cuotas</b>", cell_style),
            Paragraph("<b>Pendiente USD</b>", cell_style),
            Paragraph("<b>Var % mes</b>", cell_style),
        ]
        mes_rows = [mes_header]
        for row in serie:
            var = row.get("var_pct_vs_mes_anterior")
            if var is None:
                var_txt = "n/d"
            elif var > 0:
                var_txt = f"<font color='#B42318'><b>+{var:.2f}%</b></font>"
            elif var < 0:
                var_txt = f"<font color='#027A48'><b>{var:.2f}%</b></font>"
            else:
                var_txt = "0.00%"
            mes_rows.append(
                [
                    str(row.get("periodo", "")),
                    str(row.get("fecha_corte", "")),
                    f"{int(row.get('prestamos') or 0):,}",
                    f"{int(row.get('cuotas') or 0):,}",
                    f"${float(row.get('monto') or 0):,.2f}",
                    Paragraph(var_txt, cell_style),
                ]
            )
        mes_w = [
            usable_w * 0.12,
            usable_w * 0.16,
            usable_w * 0.14,
            usable_w * 0.14,
            usable_w * 0.24,
            usable_w * 0.20,
        ]
        mes_tbl = Table(mes_rows, colWidths=mes_w, repeatRows=1)
        mes_tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("ALIGN", (2, 1), (3, -1), "CENTER"),
                    ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                    ("ALIGN", (5, 1), (5, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BFBFBF")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F2F2F2")],
                    ),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    (
                        "BACKGROUND",
                        (0, -1),
                        (-1, -1),
                        colors.HexColor("#EEF2F7"),
                    ),
                ]
            )
        )
        story.append(KeepTogether([mes_tbl]))
        story.append(Spacer(1, 10))

    if not items:
        story.append(
            Paragraph(
                "Sin resultados para las fechas y el filtro de cuotas impagas seleccionados.",
                empty_style,
            )
        )
        doc.build(story, canvasmaker=_NumberedCanvas)
        return buf.getvalue()

    col_widths = [
        usable_w * 0.13,
        usable_w * 0.27,
        usable_w * 0.12,
        usable_w * 0.14,
        usable_w * 0.12,
        usable_w * 0.14,
    ]
    header = [
        Paragraph("Cedula", header_cell_style),
        Paragraph("Cliente", header_cell_style),
        Paragraph(f"Cuotas<br/>antes {f1}", header_cell_style),
        Paragraph(f"Pendiente<br/>antes {f1}", header_cell_style),
        Paragraph(f"Cuotas<br/>hoy {f2}", header_cell_style),
        Paragraph(f"Pendiente<br/>hoy {f2}", header_cell_style),
    ]
    rows = [header]
    for it in items:
        nombre = (it.get("nombres") or "")[:42]
        rows.append(
            [
                Paragraph(str(it.get("cedula", "")), cell_style),
                Paragraph(nombre, cell_style),
                str(it.get("cuotas_f1", 0)),
                f"${it.get('monto_f1', 0):,.2f}",
                str(it.get("cuotas_f2", 0)),
                f"${it.get('monto_f2', 0):,.2f}",
            ]
        )

    # Fila de totales al final de la tabla
    rows.append(
        [
            Paragraph("<b>TOTAL</b>", cell_style),
            Paragraph(f"<b>{n_prestamos:,} prestamos</b>", cell_style),
            Paragraph(f"<b>{tot_c1:,}</b>", cell_style),
            Paragraph(f"<b>${tot_m1:,.2f}</b>", cell_style),
            Paragraph(f"<b>{tot_c2:,}</b>", cell_style),
            Paragraph(f"<b>${tot_m2:,.2f}</b>", cell_style),
        ]
    )

    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    last = len(rows) - 1
    style_cmds = [
        ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#1F4E79")),
        ("BACKGROUND", (4, 0), (5, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, last - 1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#EEF2F7")),
        ("LINEABOVE", (0, last), (-1, last), 1.0, colors.HexColor("#1F4E79")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEBEFORE", (4, 0), (4, -1), 1.2, colors.HexColor("#1F4E79")),
    ]
    # Filas de detalle: header=0, items empiezan en 1, total=last
    for i, it in enumerate(items, start=1):
        c1 = int(it.get("cuotas_f1") or 0)
        c2 = int(it.get("cuotas_f2") or 0)
        m1 = float(it.get("monto_f1") or 0)
        m2 = float(it.get("monto_f2") or 0)
        redujo = bool(it.get("redujo_impagas")) if "redujo_impagas" in it else c2 < c1
        abono = (
            bool(it.get("abono_parcial"))
            if "abono_parcial" in it
            else ((not redujo) and (m2 + 0.009 < m1))
        )
        if redujo:
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E8F5E9"))
            )
        elif abono:
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFF3E0"))
            )
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story, canvasmaker=_NumberedCanvas)
    return buf.getvalue()



@router.get("/exportar/cartera")
def exportar_cartera(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None, description="YYYY-MM-DD vencimiento desde"),
    fecha_hasta: Optional[str] = Query(None, description="YYYY-MM-DD vencimiento hasta"),
    cuotas_impagas_min: int = Query(
        1, ge=1, le=99, description="Minimo de cuotas en mora en el rango"
    ),
    cuotas_impagas_max: int = Query(
        99, ge=1, le=99, description="Maximo de cuotas en mora (99 = todas)"
    ),
    meses: int = Query(12, ge=1, le=24, description="Legacy Excel por mes"),
    anos: Optional[str] = Query(None, description="Legacy anos"),
    meses_list: Optional[str] = Query(None, description="Legacy meses 1-12"),
):
    """
    Exporta Cuentas por cobrar.
    Con fecha_hasta (corte): detalle por prestamo con cuotas en MORA oficial.
    Sin esas fechas (legacy): Excel por mes / PDF resumen clasico.
    """
    if fecha_hasta and not fecha_desde:
        fh = _parse_fecha(fecha_hasta)
        data = _datos_cuentas_por_cobrar(
            db, fh, fh, cuotas_impagas_min, cuotas_impagas_max, corte_unico=True
        )
        stamp = fh.isoformat()
        if formato == "excel":
            content = _generar_excel_cuentas_por_cobrar(data)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=cuentas_por_cobrar_{stamp}.xlsx"
                },
            )
        content = _generar_pdf_cuentas_por_cobrar(data)
        return Response(
            content=content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=cuentas_por_cobrar_{stamp}.pdf"
            },
        )

    if fecha_desde and fecha_hasta:
        fd = _parse_fecha(fecha_desde)
        fh = _parse_fecha(fecha_hasta)
        # Siempre menor -> mayor: columna izquierda = corte menor, derecha = mayor.
        if fd > fh:
            fd, fh = fh, fd
        data = _datos_cuentas_por_cobrar(db, fd, fh, cuotas_impagas_min, cuotas_impagas_max)
        stamp = f"{fd.isoformat()}_{fh.isoformat()}"
        if formato == "excel":
            content = _generar_excel_cuentas_por_cobrar(data)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=cuentas_por_cobrar_{stamp}.xlsx"
                },
            )
        content = _generar_pdf_cuentas_por_cobrar(data)
        return Response(
            content=content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=cuentas_por_cobrar_{stamp}.pdf"
            },
        )

    fc = _parse_fecha(fecha_corte)
    if formato == "excel":
        data_por_mes = _cartera_por_periodos(db, _periodos_desde_filtros(anos, meses_list, meses))
        content = _generar_excel_cartera_por_mes(data_por_mes)
        hoy_str = date.today().isoformat()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_cartera_{hoy_str}.xlsx"},
        )
    data = _datos_cartera(db, fc)
    content = _generar_pdf_cartera(data)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_cartera_{fc.isoformat()}.pdf"},
    )


@router.post("/aseguradora/sync")
def sync_aseguradora_universo(db: Session = Depends(get_db)):
    """Sincroniza cedulas desde el Google Sheet Aseguradora (solo columna Cedula)."""
    from fastapi import HTTPException
    from app.services.aseguradora_sheet_sync import sync_aseguradora_cedulas_desde_sheet

    try:
        return sync_aseguradora_cedulas_desde_sheet(db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e


@router.get("/aseguradora/meta")
def meta_aseguradora_universo(db: Session = Depends(get_db)):
    from app.services.aseguradora_sheet_sync import meta_universo_aseguradora

    return meta_universo_aseguradora(db)


# Corte fijo Aseguradora: solo cuotas impagas acumuladas hasta esta fecha.
ASEGURADORA_CORTE_FIJO = date(2026, 8, 1)
# Solo entran quienes tienen 4 o mas cuotas sin pagar a ese corte.
ASEGURADORA_IMPAGAS_MIN = 4


@router.get("/exportar/aseguradora")
def exportar_aseguradora(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_desde: Optional[str] = Query(
        None,
        description="Ignorado: Aseguradora usa corte fijo 2026-08-01",
    ),
    fecha_hasta: Optional[str] = Query(
        None,
        description="Ignorado: Aseguradora usa corte fijo 2026-08-01",
    ),
    cuotas_impagas_min: Optional[int] = Query(
        None,
        description="Ignorado: fijo >= 4 cuotas impagas",
    ),
    cuotas_impagas_max: Optional[int] = Query(
        None,
        description="Ignorado: fijo >= 4 cuotas impagas",
    ),
    sync: bool = Query(True, description="Releer cedulas del Google Sheet antes de exportar"),
):
    """
    Universo Sheet Aseguradora: solo cuotas impagas a corte fijo 2026-08-01
    y con 4 o mas cuotas sin pagar (3 o menos no entran).
    Sin filtros de fechas ni de rango de cuotas (parametros se ignoran).
    """
    from fastapi import HTTPException
    from app.services.aseguradora_sheet_sync import (
        claves_universo_aseguradora,
        sync_aseguradora_cedulas_desde_sheet,
    )

    _ = fecha_desde, fecha_hasta, cuotas_impagas_min, cuotas_impagas_max
    if sync:
        try:
            sync_aseguradora_cedulas_desde_sheet(db)
        except Exception as e:
            # Si ya hay universo cacheado, continuar; si no, fallar.
            claves = claves_universo_aseguradora(db)
            if not claves:
                raise HTTPException(
                    status_code=400,
                    detail=f"No se pudo sincronizar la hoja Aseguradora y no hay cedulas en cache: {e}",
                ) from e
    claves = claves_universo_aseguradora(db)
    if not claves:
        raise HTTPException(
            status_code=400,
            detail="Universo Aseguradora vacio. Sincronice la hoja (POST /reportes/aseguradora/sync).",
        )
    corte = ASEGURADORA_CORTE_FIJO
    # Reglas fijas solo de este informe: corte = 2026-08-01; cuotas impagas >= 4.
    # Sin tope max inventado (no usar 1-15 de cartera).
    snap = _agg_impagas_en_fecha_historico(db, corte, cedulas_norm=claves)
    incluidos = {
        int(pid): row
        for pid, row in snap.items()
        if int(row.get("cuotas") or 0) >= ASEGURADORA_IMPAGAS_MIN
    }
    cuota_std = _cuota_estandar_por_prestamo(db, list(incluidos.keys()))
    items = []
    for pid, row in incluidos.items():
        items.append(
            {
                "cedula": (row.get("cedula") or "").strip(),
                "cuota_unitaria": cuota_std.get(pid, 0.0),
                "cuotas": int(row.get("cuotas") or 0),
                "monto": round(_safe_float(row.get("monto") or 0), 2),
            }
        )
    items.sort(key=lambda x: (x.get("cedula") or "",))
    tot_m = round(sum(float(i.get("monto") or 0) for i in items), 2)
    data = {
        "titulo_informe": "Aseguradora - Cuotas impagas a corte 2026-08-01 (4 o mas)",
        "fecha_desde": corte.isoformat(),
        "fecha_hasta": corte.isoformat(),
        "fecha_corte": corte.isoformat(),
        "cuotas_impagas_min": ASEGURADORA_IMPAGAS_MIN,
        "cuotas_impagas_max": None,
        "incluye_cuota_unitaria": True,
        "universo_cedulas": len(claves),
        "cantidad": len(items),
        "total_monto": tot_m,
        "recobrado_usd": 0.0,
        "recobrado_pagos": 0,
        "corte_fijo": True,
        "items": items,
    }
    stamp = corte.isoformat().replace("-", "")
    if formato == "excel":
        content = _generar_excel_impagas_cedula(data)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=aseguradora_{stamp}.xlsx"
            },
        )
    content = _generar_pdf_impagas_cedula(data)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=aseguradora_{stamp}.pdf"},
    )


def _agg_impagas_en_periodo_historico(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    cedulas_norm: Optional[Set[str]] = None,
) -> dict:
    """
    Cuotas impagas cuyo vencimiento cae en [fecha_desde, fecha_hasta].
    Pagado evaluado a fecha_hasta (historico). No acumula mora anterior al rango.
    """
    if cedulas_norm is not None and len(cedulas_norm) == 0:
        return {}
    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    limite = fecha_hasta + timedelta(days=1)
    estado_pago = func.upper(func.trim(func.coalesce(Pago.estado, "")))
    pago_operativo = and_(
        ~estado_pago.like("ANULADO%"),
        estado_pago.is_distinct_from("DUPLICADO"),
    )
    pagado_subq = (
        select(
            CuotaPago.cuota_id.label("cuota_id"),
            func.coalesce(func.sum(CuotaPago.monto_aplicado), 0).label("pagado_asof"),
        )
        .select_from(CuotaPago)
        .join(Pago, Pago.id == CuotaPago.pago_id)
        .where(Pago.fecha_pago < limite, pago_operativo)
        .group_by(CuotaPago.cuota_id)
        .subquery()
    )
    pagado_join = func.coalesce(pagado_subq.c.pagado_asof, 0)
    pagado_asof = case(
        (
            and_(
                pagado_join <= 0.009,
                Cuota.fecha_pago.is_not(None),
                Cuota.fecha_pago <= fecha_hasta,
            ),
            Cuota.monto,
        ),
        else_=pagado_join,
    )
    impaga = and_(
        pagado_asof < (Cuota.monto - 0.01),
        Cuota.estado.is_distinct_from("CANCELADA"),
    )
    saldo_cuota = func.greatest(Cuota.monto - pagado_asof, 0)
    estado_prestamo = func.upper(func.trim(Prestamo.estado))
    where_parts = [
        Cliente.estado == "ACTIVO",
        estado_prestamo.in_(("APROBADO", "LIQUIDADO")),
        impaga,
        Cuota.fecha_vencimiento >= fecha_desde,
        Cuota.fecha_vencimiento <= fecha_hasta,
    ]
    if cedulas_norm is not None:
        where_parts.append(
            expr_cedula_normalizada_para_comparar(Prestamo.cedula).in_(list(cedulas_norm))
        )
    rows = db.execute(
        select(
            Prestamo.id.label("prestamo_id"),
            Prestamo.cedula,
            Prestamo.nombres,
            func.count(Cuota.id).label("cuotas"),
            func.coalesce(func.sum(saldo_cuota), 0).label("monto"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .outerjoin(pagado_subq, pagado_subq.c.cuota_id == Cuota.id)
        .where(*where_parts)
        .group_by(Prestamo.id, Prestamo.cedula, Prestamo.nombres)
    ).fetchall()
    out: dict = {}
    for r in rows:
        out[int(r.prestamo_id)] = {
            "prestamo_id": int(r.prestamo_id),
            "cedula": (r.cedula or "").strip(),
            "nombres": (r.nombres or "").strip(),
            "cuotas": int(r.cuotas or 0),
            "monto": round(_safe_float(r.monto), 2),
        }
    return out



def _total_recobrado_usd_periodo_aseguradora(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    cedulas_norm: Set[str],
) -> dict:
    """
    Suma en USD de pagos operativos del universo Aseguradora con fecha_pago en el periodo.
    Excluye ANULADO*/DUPLICADO. Es lo recobrado por gestion de cobranza en el rango.
    """
    if not cedulas_norm:
        return {"recobrado_usd": 0.0, "pagos_count": 0}
    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde
    limite = fecha_hasta + timedelta(days=1)
    estado_pago = func.upper(func.trim(func.coalesce(Pago.estado, "")))
    pago_operativo = and_(
        ~estado_pago.like("ANULADO%"),
        estado_pago.is_distinct_from("DUPLICADO"),
    )
    ced_pago = expr_cedula_normalizada_para_comparar(Pago.cedula_cliente)
    ced_prest = expr_cedula_normalizada_para_comparar(Prestamo.cedula)
    claves = list(cedulas_norm)
    row = db.execute(
        select(
            func.coalesce(func.sum(Pago.monto_pagado), 0).label("total"),
            func.count(Pago.id).label("n_pagos"),
        )
        .select_from(Pago)
        .outerjoin(Prestamo, Prestamo.id == Pago.prestamo_id)
        .where(
            Pago.fecha_pago >= fecha_desde,
            Pago.fecha_pago < limite,
            pago_operativo,
            or_(ced_pago.in_(claves), ced_prest.in_(claves)),
        )
    ).one()
    return {
        "recobrado_usd": round(_safe_float(row.total), 2),
        "pagos_count": int(row.n_pagos or 0),
    }


def _cuota_estandar_por_prestamo(
    db: Session,
    prestamo_ids: List[int],
) -> dict:
    """
    Cuota estandar que paga el cliente en cada prestamo. Siempre devuelve un valor
    para cada prestamo pedido (nunca vacio ni cero), en este orden:

    1. Monto que mas se repite en la tabla de cuotas (mode) - lo realmente pactado.
    2. total_financiamiento / numero_cuotas - si no hay cuotas o vienen en cero.
    3. prestamos.cuota_periodo - ultimo recurso (puede estar desactualizado).
    """
    if not prestamo_ids:
        return {}
    ids = [int(x) for x in prestamo_ids]
    modal_subq = (
        select(
            Cuota.prestamo_id.label("prestamo_id"),
            func.mode().within_group(Cuota.monto.asc()).label("modal"),
        )
        .where(Cuota.prestamo_id.in_(ids))
        .group_by(Cuota.prestamo_id)
        .subquery()
    )
    rows = db.execute(
        select(
            Prestamo.id.label("prestamo_id"),
            modal_subq.c.modal,
            Prestamo.total_financiamiento,
            Prestamo.numero_cuotas,
            Prestamo.cuota_periodo,
        )
        .select_from(Prestamo)
        .outerjoin(modal_subq, modal_subq.c.prestamo_id == Prestamo.id)
        .where(Prestamo.id.in_(ids))
    ).fetchall()

    out: dict = {}
    for r in rows:
        modal = _safe_float(r.modal)
        if modal > 0:
            out[int(r.prestamo_id)] = round(modal, 2)
            continue
        n_cuotas = int(r.numero_cuotas or 0)
        financiado = _safe_float(r.total_financiamiento)
        if n_cuotas > 0 and financiado > 0:
            out[int(r.prestamo_id)] = round(financiado / n_cuotas, 2)
            continue
        out[int(r.prestamo_id)] = round(_safe_float(r.cuota_periodo), 2)
    return out


def _datos_impagas_cedula_aseguradora(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    cuotas_impagas_min: int,
    cuotas_impagas_max: int,
    cedulas_norm: Set[str],
) -> dict:
    """
    Universo Aseguradora: cedula + cuotas impagas acumuladas hasta fecha_hasta (corte).
    fecha_desde solo aplica al KPI de recobrado del periodo.
    """
    min_n = max(1, min(15, int(cuotas_impagas_min)))
    max_n = max(1, min(15, int(cuotas_impagas_max)))
    if min_n > max_n:
        min_n, max_n = max_n, min_n
    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde
    # Acumulado a corte (fecha_hasta): mora desde el inicio hasta el corte.
    snap = _agg_impagas_en_fecha_historico(
        db, fecha_hasta, cedulas_norm=cedulas_norm
    )
    items: List[dict] = []
    for _pid, row in snap.items():
        c = int(row.get("cuotas") or 0)
        if not (min_n <= c <= max_n):
            continue
        items.append(
            {
                "cedula": (row.get("cedula") or "").strip(),
                "cuotas": c,
                "monto": round(_safe_float(row.get("monto") or 0), 2),
            }
        )
    items.sort(key=lambda x: (x.get("cedula") or "",))
    tot_m = round(sum(float(i.get("monto") or 0) for i in items), 2)
    recobro = _total_recobrado_usd_periodo_aseguradora(
        db, fecha_desde, fecha_hasta, cedulas_norm
    )
    return {
        "titulo_informe": "Impagas por cedula",
        "fecha_desde": fecha_desde.isoformat(),
        "fecha_hasta": fecha_hasta.isoformat(),
        "fecha_corte": fecha_hasta.isoformat(),
        "cuotas_impagas_min": min_n,
        "cuotas_impagas_max": max_n,
        "universo_cedulas": len(cedulas_norm),
        "cantidad": len(items),
        "total_monto": tot_m,
        "recobrado_usd": recobro["recobrado_usd"],
        "recobrado_pagos": recobro["pagos_count"],
        "items": items,
    }


def _filas_tres_columnas_cedula_cuotas(items: List[dict]) -> List[list]:
    """Reparte en 3 bloques verticales: Cedula|Cuotas|Monto x3."""
    n = len(items)
    if n == 0:
        return []
    n_cols = 3
    per_col = (n + n_cols - 1) // n_cols
    cols = [items[i * per_col : (i + 1) * per_col] for i in range(n_cols)]
    max_len = max(len(c) for c in cols)
    rows: List[list] = []
    for i in range(max_len):
        row: list = []
        for col in cols:
            if i < len(col):
                row.append(col[i].get("cedula", ""))
                row.append(int(col[i].get("cuotas") or 0))
                row.append(round(float(col[i].get("monto") or 0), 2))
            else:
                row.extend(["", "", ""])
        rows.append(row)
    return rows


def _filas_tres_columnas_cedula_cuota_unitaria(items: List[dict]) -> List[list]:
    """3 bloques verticales: Cedula|Cuota|Impagas|Monto x3."""
    n = len(items)
    if n == 0:
        return []
    n_cols = 3
    per_col = (n + n_cols - 1) // n_cols
    cols = [items[i * per_col : (i + 1) * per_col] for i in range(n_cols)]
    max_len = max(len(c) for c in cols)
    rows: List[list] = []
    for i in range(max_len):
        row: list = []
        for col in cols:
            if i < len(col):
                row.append(col[i].get("cedula", ""))
                row.append(round(float(col[i].get("cuota_unitaria") or 0), 2))
                row.append(int(col[i].get("cuotas") or 0))
                row.append(round(float(col[i].get("monto") or 0), 2))
            else:
                row.extend(["", "", "", ""])
        rows.append(row)
    return rows


def _generar_excel_impagas_cedula(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Impagas cedula"
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    meta_font = Font(name="Calibri", size=9, color="666666")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    fd = data.get("fecha_desde") or data.get("fecha_corte", "")
    fh = data.get("fecha_hasta") or data.get("fecha_corte", "")
    ws.append([data.get("titulo_informe") or "Impagas por cedula"])
    ws["A1"].font = title_font
    rec_usd = float(data.get("recobrado_usd") or 0)
    rec_n = int(data.get("recobrado_pagos") or 0)
    rec_font = Font(name="Calibri", size=12, bold=True, color="166534")
    if data.get("corte_fijo"):
        ws.append(
            [
                f"CORTE FIJO: solo cuotas no pagadas hasta {fh}"
                f"   |   Condicion: {data.get('cuotas_impagas_min')}+ cuotas impagas"
            ]
        )
        ws["A2"].font = rec_font
        ws.append(
            [
                f"Corte impagas: {fh}   |   "
                f"Solo {data.get('cuotas_impagas_min')}+ cuotas sin pagar   |   "
                f"Universo hoja: {data.get('universo_cedulas')}   |   Registros: {data.get('cantidad', 0)}   |   "
                f"Pendiente acumulado a corte: ${data.get('total_monto', 0):,.2f}"
            ]
        )
    else:
        ws.append(
            [
                f"RECOBRADO EN EL PERIODO (gestion cobranza): ${rec_usd:,.2f} USD"
                f"   |   Pagos: {rec_n}"
            ]
        )
        ws["A2"].font = rec_font
        ws.append(
            [
                f"Recobrado periodo: {fd} a {fh}   |   Corte impagas: {fh}   |   "
                f"Filtro cuotas a corte: {data.get('cuotas_impagas_min')}-{data.get('cuotas_impagas_max')}   |   "
                f"Universo hoja: {data.get('universo_cedulas')}   |   Registros: {data.get('cantidad', 0)}   |   "
                f"Pendiente acumulado a corte: ${data.get('total_monto', 0):,.2f}"
            ]
        )
    ws["A3"].font = meta_font
    ws.append([])
    con_cuota = bool(data.get("incluye_cuota_unitaria"))
    if con_cuota:
        headers = ["Cedula", "Cuota", "Impagas", "Monto"] * 3
        anchos = "ABCDEFGHIJKL"
        ancho_por_col = (12, 11, 9, 12) * 3
        campos = 4
        filas = _filas_tres_columnas_cedula_cuota_unitaria(
            list(data.get("items") or [])
        )
    else:
        headers = ["Cedula", "Cuotas", "Monto"] * 3
        anchos = "ABCDEFGHI"
        ancho_por_col = (12, 8, 11, 12, 8, 11, 12, 8, 11)
        campos = 3
        filas = _filas_tres_columnas_cedula_cuotas(list(data.get("items") or []))
    ws.append(headers)
    for cell in ws[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for row in filas:
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = thin
            pos = (col - 1) % campos  # 0 cedula, luego cuota/impagas/monto
            es_conteo = pos == (2 if con_cuota else 1)
            es_monto = pos in ((1, 3) if con_cuota else (2,))
            if es_conteo:
                cell.alignment = Alignment(horizontal="center")
            if es_monto and cell.value != "":
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    for col, w in zip(anchos, ancho_por_col):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_impagas_cedula(data: dict) -> bytes:
    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    fd = data.get("fecha_desde") or data.get("fecha_corte", "")
    fh = data.get("fecha_hasta") or data.get("fecha_corte", "")
    fc = fh
    items = list(data.get("items") or [])
    n = int(data.get("cantidad") or len(items))
    filtro_min = data.get("cuotas_impagas_min")
    filtro_max = data.get("cuotas_impagas_max")
    univ = data.get("universo_cedulas")
    tot_m = float(data.get("total_monto") or 0)
    rec_usd = float(data.get("recobrado_usd") or 0)
    rec_n = int(data.get("recobrado_pagos") or 0)
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")
    titulo = (data.get("titulo_informe") or "Impagas por cedula").strip()

    page = landscape(letter)
    page_w, page_h = page
    top_m = 0.7 * inch
    bottom_m = 0.5 * inch
    side_m = 0.4 * inch

    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=side_m,
        rightMargin=side_m,
        topMargin=top_m,
        bottomMargin=bottom_m,
    )
    styles = getSampleStyleSheet()
    meta_style = ParagraphStyle(
        "imp_meta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#44546A"),
        leading=10,
        spaceAfter=6,
    )
    recobrado_style = ParagraphStyle(
        "imp_recobro",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#166534"),
        leading=14,
        spaceAfter=4,
        alignment=1,
    )
    header_style = ParagraphStyle(
        "imp_hdr",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.white,
        alignment=1,
        leading=10,
    )
    cell_style = ParagraphStyle(
        "imp_cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        textColor=colors.HexColor("#1A1A1A"),
        leading=9,
    )
    empty_style = ParagraphStyle(
        "imp_empty",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        textColor=colors.HexColor("#C00000"),
        alignment=1,
        spaceBefore=20,
    )

    class _NumberedCanvas(pdf_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            pdf_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page_decor(page_count)
                pdf_canvas.Canvas.showPage(self)
            pdf_canvas.Canvas.save(self)

        def _draw_page_decor(self, page_count):
            self.saveState()
            self.setFillColor(colors.HexColor("#1F4E79"))
            self.rect(0, page_h - 0.42 * inch, page_w, 0.42 * inch, fill=1, stroke=0)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 11)
            self.drawCentredString(page_w / 2.0, page_h - 0.27 * inch, titulo.upper()[:50])
            self.setStrokeColor(colors.HexColor("#D0D5DD"))
            self.setLineWidth(0.6)
            self.line(side_m, 0.35 * inch, page_w - side_m, 0.35 * inch)
            self.setFillColor(colors.HexColor("#667085"))
            self.setFont("Helvetica", 7.5)
            self.drawString(side_m, 0.2 * inch, f"Generado: {generado}")
            self.drawCentredString(
                page_w / 2.0, 0.2 * inch, f"Pagina {self._pageNumber} de {page_count}"
            )
            self.drawRightString(page_w - side_m, 0.2 * inch, "Confidencial - uso interno")
            self.restoreState()

    story = []
    if data.get("corte_fijo"):
        story.append(
            Paragraph(
                f"CORTE FIJO: solo cuotas no pagadas hasta <b>{fh}</b>"
                f" &nbsp;|&nbsp; Condicion: <b>{filtro_min}+ cuotas impagas</b>",
                recobrado_style,
            )
        )
        story.append(
            Paragraph(
                f"Corte impagas: <b>{fh}</b> &nbsp;|&nbsp; "
                f"Solo <b>{filtro_min}+</b> cuotas sin pagar &nbsp;|&nbsp; "
                f"Universo hoja: <b>{univ}</b> &nbsp;|&nbsp; "
                f"Registros: <b>{n:,}</b> &nbsp;|&nbsp; "
                f"Pendiente acumulado a corte: <b>${tot_m:,.2f}</b>",
                meta_style,
            )
        )
    else:
        story.append(
            Paragraph(
                f"RECOBRADO EN EL PERIODO (gestion cobranza): "
                f"<font color='#166534'>${rec_usd:,.2f} USD</font>"
                f" &nbsp;&nbsp;|&nbsp;&nbsp; Pagos: {rec_n}",
                recobrado_style,
            )
        )
        story.append(
            Paragraph(
                f"Recobrado periodo: <b>{fd}</b> a <b>{fh}</b> &nbsp;|&nbsp; "
                f"Corte impagas: <b>{fh}</b> &nbsp;|&nbsp; "
                f"Filtro cuotas a corte: <b>{filtro_min}-{filtro_max}</b> &nbsp;|&nbsp; "
                f"Universo hoja: <b>{univ}</b> &nbsp;|&nbsp; "
                f"Registros: <b>{n:,}</b> &nbsp;|&nbsp; "
                f"Pendiente acumulado a corte: <b>${tot_m:,.2f}</b>",
                meta_style,
            )
        )
    story.append(Spacer(1, 4))

    usable_w = page_w - 2 * side_m
    if not items:
        story.append(
            Paragraph(
                "Sin resultados para la fecha y el filtro de cuotas impagas seleccionados.",
                empty_style,
            )
        )
        doc.build(story, canvasmaker=_NumberedCanvas)
        return buf.getvalue()

    def _cel_cuota(v):
        return Paragraph("" if v == "" else str(v), cell_style)

    def _cel_monto(v):
        if v == "" or v is None:
            return Paragraph("", cell_style)
        return Paragraph(f"${float(v):,.2f}", cell_style)

    con_cuota = bool(data.get("incluye_cuota_unitaria"))
    if con_cuota:
        # 3 bloques: Cedula | Cuota | Impagas | Monto
        etiquetas = ("Cedula", "Cuota", "Impagas", "Monto")
        fracciones = (0.098, 0.078, 0.052, 0.095)
        filas_raw = _filas_tres_columnas_cedula_cuota_unitaria(items)
        constructores = (None, _cel_monto, _cel_cuota, _cel_monto)
    else:
        # 3 bloques: Cedula | Cuotas | Monto
        etiquetas = ("Cedula", "Cuotas", "Monto")
        fracciones = (0.12, 0.055, 0.105)
        filas_raw = _filas_tres_columnas_cedula_cuotas(items)
        constructores = (None, _cel_cuota, _cel_monto)

    campos = len(etiquetas)
    col_w = [usable_w * f for _ in range(3) for f in fracciones]
    col_w[-1] += usable_w - sum(col_w)

    header = [Paragraph(txt, header_style) for _ in range(3) for txt in etiquetas]
    rows = [header]
    for raw in filas_raw:
        fila = []
        for idx, valor in enumerate(raw):
            constructor = constructores[idx % campos]
            if constructor is None:
                fila.append(Paragraph(str(valor or ""), cell_style))
            else:
                fila.append(constructor(valor))
        rows.append(fila)

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for bloque in range(3):
        base = bloque * campos
        if con_cuota:
            style_cmds.append(("ALIGN", (base + 1, 0), (base + 1, -1), "RIGHT"))
            style_cmds.append(("ALIGN", (base + 2, 0), (base + 2, -1), "CENTER"))
            style_cmds.append(("ALIGN", (base + 3, 0), (base + 3, -1), "RIGHT"))
        else:
            style_cmds.append(("ALIGN", (base + 1, 0), (base + 1, -1), "CENTER"))
            style_cmds.append(("ALIGN", (base + 2, 0), (base + 2, -1), "RIGHT"))
        if bloque:
            style_cmds.append(
                ("LINEBEFORE", (base, 0), (base, -1), 1.0, colors.HexColor("#1F4E79"))
            )
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story, canvasmaker=_NumberedCanvas)
    return buf.getvalue()


@router.get("/exportar/aseguradora-impagas")
def exportar_aseguradora_impagas(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_desde: Optional[str] = Query(None, description="YYYY-MM-DD inicio periodo vencimiento"),
    fecha_hasta: Optional[str] = Query(None, description="YYYY-MM-DD fin periodo vencimiento"),
    cuotas_impagas_min: int = Query(1, ge=1, le=15),
    cuotas_impagas_max: int = Query(15, ge=1, le=15),
    sync: bool = Query(True, description="Releer cedulas del Google Sheet antes de exportar"),
):
    """
    Universo Aseguradora: Cedula + cuotas/monto impagos acumulados hasta fecha_hasta (corte).
    fecha_desde..fecha_hasta define el recobrado del periodo.
    """
    from fastapi import HTTPException
    from app.services.aseguradora_sheet_sync import (
        claves_universo_aseguradora,
        sync_aseguradora_cedulas_desde_sheet,
    )

    if not fecha_desde or not fecha_hasta:
        raise HTTPException(status_code=400, detail="Indique fecha_desde y fecha_hasta.")
    if sync:
        try:
            sync_aseguradora_cedulas_desde_sheet(db)
        except Exception as e:
            claves = claves_universo_aseguradora(db)
            if not claves:
                raise HTTPException(
                    status_code=400,
                    detail=f"No se pudo sincronizar la hoja Aseguradora y no hay cedulas en cache: {e}",
                ) from e
    claves = claves_universo_aseguradora(db)
    if not claves:
        raise HTTPException(
            status_code=400,
            detail="Universo Aseguradora vacio. Sincronice la hoja (POST /reportes/aseguradora/sync).",
        )
    fd = _parse_fecha(fecha_desde)
    fh = _parse_fecha(fecha_hasta)
    if fd > fh:
        fd, fh = fh, fd
    data = _datos_impagas_cedula_aseguradora(
        db, fd, fh, cuotas_impagas_min, cuotas_impagas_max, claves
    )
    stamp = f"{fd.isoformat()}_{fh.isoformat()}"
    if formato == "excel":
        content = _generar_excel_impagas_cedula(data)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=impagas_cedula_{stamp}.xlsx"
            },
        )
    content = _generar_pdf_impagas_cedula(data)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=impagas_cedula_{stamp}.pdf"
        },
    )


@router.post("/cuotas-hoja-periodo/actualizar")
def actualizar_cuotas_hoja_periodo(
    db: Session = Depends(get_db),
    fecha_desde: str = Query(..., description="YYYY-MM-DD inicio periodo"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD fin periodo"),
    dry_run: bool = Query(False, description="Si true, calcula sin escribir en Sheets"),
):
    """
    Actualiza columna Cuotas del Google Sheet por cedula:
    nuevo = max(0, base + impagas_periodo - cerradas_previas).
    """
    from fastapi import HTTPException
    from app.services.cuotas_hoja_periodo_sync import actualizar_cuotas_hoja_por_periodo

    try:
        fd = _parse_fecha(fecha_desde)
        fh = _parse_fecha(fecha_hasta)
        if fd > fh:
            fd, fh = fh, fd
        return actualizar_cuotas_hoja_por_periodo(db, fd, fh, dry_run=dry_run)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

