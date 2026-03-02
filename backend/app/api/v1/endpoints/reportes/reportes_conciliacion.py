"""
Reportes de Conciliación mejorado: con filtros, PDF export y optimizaciones.
"""
import io
import re
import calendar
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, Query, Body, File
from fastapi.responses import Response
from sqlalchemy import func, select, delete, and_, or_
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.cliente import Cliente
from app.models.conciliacion_temporal import ConciliacionTemporal
from app.models.cuota import Cuota
from app.models.pago import Pago
from app.models.prestamo import Prestamo

from app.api.v1.endpoints.reportes_utils import _safe_float

router = APIRouter(dependencies=[Depends(get_current_user)])

# Validación cédula: al menos 5 dígitos/letras (ajustar según reglas de negocio)
CEDULA_PATTERN = re.compile(r"^[A-Za-z0-9\-]{5,20}$")


def _validar_cedula(cedula: Any) -> bool:
    if cedula is None:
        return False
    s = str(cedula).strip()
    # Allow "cedula no encontrada" as special valid value
    if s.lower() == "cedula no encontrada":
        return True
    return bool(s and CEDULA_PATTERN.match(s))


def _validar_numero(val: Any) -> bool:
    if val is None:
        return False
    try:
        f = float(val)
        return f >= 0
    except (TypeError, ValueError):
        return False


def _parse_numero(val: Any) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal("0")


def _normalizar_cedula(cedula: str) -> str:
    """
    Normaliza cédula para comparación exacta.
    - Elimina espacios y tabs
    - Convierte a mayúsculas
    - Elimina caracteres especiales comunes
    
    Ejemplos:
        "V 12345678"  → "V12345678"
        "E 98765432"  → "E98765432"
        "  v12345678" → "V12345678"
    """
    if not cedula:
        return ""
    return cedula.replace(" ", "").replace("\t", "").strip().upper()



@router.post("/conciliacion/cargar-excel-debug")
async def cargar_conciliacion_excel_debug(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """DEBUG: Carga Excel y muestra exactamente qué lee"""
    import logging
    logger = logging.getLogger(__name__)
    
    logger.info(f"[DEBUG] Archivo recibido: {file.filename}")
    
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        logger.error(f"[DEBUG] Archivo no es Excel: {file.filename}")
        return {"error": "Debe subir un archivo Excel (.xlsx o .xls)"}
    
    try:
        import openpyxl
        content = await file.read()
        logger.info(f"[DEBUG] Tamaño archivo: {len(content)} bytes")
        
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        logger.info(f"[DEBUG] Hoja activa: {ws.title}")
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        logger.info(f"[DEBUG] Total filas (sin header): {len(rows)}")
        
        # Log primeras 5 filas
        for i, row in enumerate(rows[:5]):
            logger.info(f"[DEBUG] Fila {i+2}: {row}")
            if row and len(row) >= 3:
                logger.info(f"[DEBUG]   -> [0]={row[0]} ({type(row[0]).__name__}), [1]={row[1]} ({type(row[1]).__name__}), [2]={row[2]} ({type(row[2]).__name__})")
        
        return {
            "debug": True,
            "archivo": file.filename,
            "tamaño_bytes": len(content),
            "hoja": ws.title,
            "total_filas": len(rows),
            "primeras_filas": [dict(cedula=row[0], tf=row[1], ta=row[2]) if row and len(row) >= 3 else None for row in rows[:5]],
        }
        
    except Exception as e:
        logger.error(f"[DEBUG] Error: {str(e)}", exc_info=True)
        return {"error": str(e)}


@router.post("/conciliacion/cargar-excel")
async def cargar_conciliacion_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Carga archivo Excel de conciliación (cédula, total_financiamiento, total_abonos).
    Valida datos y los almacena en tabla temporal para generación de reportes.
    
    Formato esperado (3 columnas):
    - Columna A: Cédula (V12345678 o E98765432) - solo para concatenación
    - Columna B: Total Financiamiento (monto numérico)
    - Columna C: Total Abonos (monto numérico)
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Debe subir un archivo Excel (.xlsx o .xls)")
    
    try:
        import openpyxl
        import io
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="Archivo Excel vacío")
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="No se encontraron filas en el archivo")
        
        errores: List[str] = []
        filas_ok: List[dict] = []
        
        for i, row in enumerate(rows):
            if not row or all(cell is None for cell in row[:3]):
                continue
            
            try:
                cedula = row[0] if len(row) > 0 and row[0] is not None else None
                tf = row[1] if len(row) > 1 and row[1] is not None else None
                ta = row[2] if len(row) > 2 and row[2] is not None else None
                
                if not _validar_cedula(cedula):
                    errores.append(f"Fila {i + 2}: cédula inválida ({cedula})")
                    continue
                if not _validar_numero(tf):
                    errores.append(f"Fila {i + 2}: total financiamiento debe ser un número ≥ 0")
                    continue
                if not _validar_numero(ta):
                    errores.append(f"Fila {i + 2}: total abonos debe ser un número ≥ 0")
                    continue
                
                filas_ok.append({
                    "cedula": _normalizar_cedula(str(cedula).strip()),
                    "total_financiamiento": _parse_numero(tf),
                    "total_abonos": _parse_numero(ta),
                })
            except Exception as e:
                errores.append(f"Fila {i + 2}: {str(e)}")
        
        # Log para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[Conciliacion] Procesadas {len(rows)} filas del Excel")
        logger.info(f"[Conciliacion] Filas válidas: {len(filas_ok)}, Errores: {len(errores)}")
        if errores:
            logger.warning(f"[Conciliacion] Errores encontrados: {errores[:5]}")
        
        if errores:
            return {
                "ok": False,
                "mensaje": f"Se encontraron {len(errores)} errores de validación",
                "errores": errores[:50],
                "filas_ok": 0,
                "filas_con_error": len(errores),
            }
        
        if not filas_ok:
            logger.error(f"[Conciliacion] Ninguna fila válida para procesar")
            return {
                "ok": False,
                "mensaje": "No se encontraron filas válidas para procesar",
                "filas_ok": 0,
                "filas_con_error": len(rows),
                "errores": ["Todas las filas fueron rechazadas"],
            }
        
        # Borrar datos previos e insertar nuevos
        db.execute(delete(ConciliacionTemporal))
        for f in filas_ok:
            db.add(ConciliacionTemporal(
                cedula=f["cedula"],
                total_financiamiento=f["total_financiamiento"],
                total_abonos=f["total_abonos"],
            ))
        db.commit()
        logger.info(f"[Conciliacion] {len(filas_ok)} filas guardadas exitosamente en BD")
        
        return {
            "ok": True,
            "mensaje": f"Carga exitosa: {len(filas_ok)} filas procesadas",
            "filas_ok": len(filas_ok),
            "filas_con_error": 0,
            "errores": [],
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar Excel: {str(e)}") from e
@router.post("/conciliacion/cargar")
def cargar_conciliacion_temporal(
    body: List[dict] = Body(...),
    db: Session = Depends(get_db),
):
    """
    Recibe lista de filas: cedula, total_financiamiento, total_abonos, columna_e (opcional), columna_f (opcional).
    Valida cédula y cantidades; si todo es válido borra datos previos e inserta los nuevos.
    """
    if not body or not isinstance(body, list):
        raise HTTPException(status_code=400, detail="Se requiere una lista de filas")
    errores: List[str] = []
    filas_ok: List[dict] = []
    for i, row in enumerate(body):
        if not isinstance(row, dict):
            errores.append(f"Fila {i + 1}: debe ser un objeto con cedula, total_financiamiento, total_abonos")
            continue
        cedula = row.get("cedula")
        tf = row.get("total_financiamiento")
        ta = row.get("total_abonos")
        if not _validar_cedula(cedula):
            errores.append(f"Fila {i + 1}: cedula invalida")
            continue
        if not _validar_numero(tf):
            errores.append(f"Fila {i + 1}: total financiamiento debe ser un numero >= 0")
            continue
        if not _validar_numero(ta):
            errores.append(f"Fila {i + 1}: total abonos debe ser un numero >= 0")
            continue
        filas_ok.append({
            "cedula": str(cedula).strip(),
            "total_financiamiento": _parse_numero(tf),
            "total_abonos": _parse_numero(ta),
            "columna_e": str(row.get("columna_e") or "").strip() or None,
            "columna_f": str(row.get("columna_f") or "").strip() or None,
        })
    if errores:
        raise HTTPException(status_code=422, detail={"errores": errores, "mensaje": "Corrija los errores antes de guardar"})
    db.execute(delete(ConciliacionTemporal))
    for f in filas_ok:
        db.add(ConciliacionTemporal(
            cedula=f["cedula"],
            total_financiamiento=f["total_financiamiento"],
            total_abonos=f["total_abonos"],
            columna_e=f.get("columna_e"),
            columna_f=f.get("columna_f"),
        ))
    db.commit()
    return {"ok": True, "filas_guardadas": len(filas_ok)}


def _generar_excel_conciliacion(
    db: Session,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    cedulas_filter: Optional[List[str]] = None,
) -> bytes:
    """Genera Excel reporte Conciliación con filtros opcionales."""
    import openpyxl

    # Construir query base
    query = select(Prestamo).where(Prestamo.estado == "APROBADO")
    
    # Filtro por fecha de aprobación (si está disponible)
    if fecha_inicio:
        query = query.where(Prestamo.fecha_aprobacion >= datetime.combine(fecha_inicio, datetime.min.time()))
    if fecha_fin:
        query = query.where(Prestamo.fecha_aprobacion <= datetime.combine(fecha_fin, datetime.max.time()))
    
    query = query.order_by(Prestamo.id)
    prestamos = db.execute(query).scalars().all()

    # Filtro por cédulas si se proporciona (con normalización)
    if cedulas_filter:
        cedulas_filter_set = {_normalizar_cedula(c) for c in cedulas_filter}
        prestamos = [p for p in prestamos if _normalizar_cedula(p.cedula or "") in cedulas_filter_set]

    # Mapa cedula -> conciliacion_temporal (normalizado)
    concilia_rows = db.execute(select(ConciliacionTemporal)).fetchall()
    concilia_por_cedula: dict = {}
    for r in concilia_rows:
        c = r[0] if hasattr(r, "__getitem__") else r
        cedula_norm = _normalizar_cedula(c.cedula)
        if cedula_norm and cedula_norm not in concilia_por_cedula:
            concilia_por_cedula[cedula_norm] = c

    ids = [p.id for p in prestamos]
    total_pagos_map: dict = {}
    total_cuotas_num_map: dict = {}
    cuotas_pagadas_num_map: dict = {}
    cuotas_pagadas_monto_map: dict = {}
    cuotas_pendientes_num_map: dict = {}
    cuotas_pendientes_monto_map: dict = {}

    if ids:
        total_pagos_map = dict(
            db.execute(
                select(Pago.prestamo_id, func.coalesce(func.sum(Pago.monto_pagado), 0))
                .where(Pago.prestamo_id.isnot(None), Pago.prestamo_id.in_(ids))
                .group_by(Pago.prestamo_id)
            ).fetchall()
        )
        rows_tot = db.execute(
            select(Cuota.prestamo_id, func.count())
            .where(Cuota.prestamo_id.in_(ids))
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        for r in rows_tot:
            total_cuotas_num_map[r[0]] = int(r[1])
        rows_pag = db.execute(
            select(Cuota.prestamo_id, func.count(), func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado == "PAGADO")
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        for r in rows_pag:
            cuotas_pagadas_num_map[r[0]] = int(r[1])
            cuotas_pagadas_monto_map[r[0]] = _safe_float(r[2])
        rows_pend = db.execute(
            select(Cuota.prestamo_id, func.count(), func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado != "PAGADO")
            .group_by(Cuota.prestamo_id)
        ).fetchall()
        for r in rows_pend:
            cuotas_pendientes_num_map[r[0]] = int(r[1])
            cuotas_pendientes_monto_map[r[0]] = _safe_float(r[2])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliacion"
    
    # Headers - 12 columnas base
    headers = [
        "Nombre",                          # A
        "Cedula",                          # B
        "Numero de Prestamo",              # C
        "Total Financiamiento Excel",      # D
        "Total Financiamiento Sistema",    # E
        "Abonos Excel",                    # F
        "Abonos Sistema",                  # G
        "Total Cuotas",                    # H
        "Cuotas Pagadas (num)",            # I
        "Cuotas Pagadas ($)",              # J
        "Cuotas Pendientes (num)",         # K
        "Cuotas Pendientes ($)",           # L
    ]
    ws.append(headers)
    
    # Procesar cada préstamo
    for p in prestamos:
        cedula = (p.cedula or "").strip()
        cedula_normalizada = _normalizar_cedula(cedula)
        
        # Buscar cliente por cédula (usar cédula exacta de prestamo, no normalizada)
        try:
            cliente = db.execute(select(Cliente).where(Cliente.cedula == cedula)).scalar()
            nombre = (cliente.nombres or "").strip() if cliente else ""
        except Exception as e:
            nombre = ""
        
        # Buscar datos del Excel por cédula
        concilia = concilia_por_cedula.get(cedula_normalizada) if cedula_normalizada else None
        try:
            tf_excel = _safe_float(concilia.total_financiamiento) if concilia else 0
            abonos_excel = _safe_float(concilia.total_abonos) if concilia else 0
        except Exception:
            tf_excel = 0
            abonos_excel = 0
        
        # Datos del Sistema (tabla prestamos)
        try:
            tf_sistema = _safe_float(p.total_financiamiento) if p.total_financiamiento else 0
            abonos_sistema = _safe_float(p.total_abonos) if p.total_abonos else 0
        except Exception:
            tf_sistema = 0
            abonos_sistema = 0
        
        # Datos de cuotas
        tot_cuotas = total_cuotas_num_map.get(p.id, p.numero_cuotas or 0)
        pag_num = cuotas_pagadas_num_map.get(p.id, 0)
        pag_monto = cuotas_pagadas_monto_map.get(p.id, 0)
        pend_num = cuotas_pendientes_num_map.get(p.id, 0)
        pend_monto = cuotas_pendientes_monto_map.get(p.id, 0)
        
        # Crear fila base (12 columnas)
        try:
            row = [
                nombre,                                    # A
                cedula,                                    # B
                p.id,                                      # C
                round(tf_excel, 2) if tf_excel > 0 else "",   # D
                round(tf_sistema, 2) if tf_sistema > 0 else "", # E
                round(abonos_excel, 2) if abonos_excel > 0 else "",     # F
                round(abonos_sistema, 2) if abonos_sistema > 0 else "",  # G
                int(tot_cuotas) if tot_cuotas else 0,     # H
                int(pag_num) if pag_num else 0,           # I
                round(pag_monto, 2) if pag_monto else 0,  # J
                int(pend_num) if pend_num else 0,         # K
                round(pend_monto, 2) if pend_monto else 0, # L
            ]
            
            # Agregar columna "error TC" si hay diferencia en Total Financiamiento
            if tf_excel > 0 and tf_sistema > 0 and tf_excel != tf_sistema:
                row.append(f"error TC: {round(abs(tf_excel - tf_sistema), 2)}")
            
            # Agregar columna "error E" si hay diferencia en Abonos
            if abonos_excel > 0 and abonos_sistema > 0 and abonos_excel != abonos_sistema:
                row.append(f"error E: {round(abs(abonos_excel - abonos_sistema), 2)}")
        except Exception as e:
            continue
        
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_conciliacion(
    db: Session,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> bytes:
    """Genera PDF reporte Conciliación con resumen y gráficos."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=1,  # CENTER
    )
    story.append(Paragraph("Reporte de Conciliación", title_style))
    
    # Rango de fechas
    fecha_txt = "Todas las fechas"
    if fecha_inicio and fecha_fin:
        fecha_txt = f"{fecha_inicio.isoformat()} a {fecha_fin.isoformat()}"
    elif fecha_inicio:
        fecha_txt = f"Desde {fecha_inicio.isoformat()}"
    elif fecha_fin:
        fecha_txt = f"Hasta {fecha_fin.isoformat()}"
    story.append(Paragraph(f"<b>Periodo:</b> {fecha_txt}", styles['Normal']))
    story.append(Paragraph(f"<b>Generado:</b> {date.today().isoformat()}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Resumen general
    story.append(Paragraph("<b>Resumen General</b>", styles['Heading2']))
    
    query = select(Prestamo).where(Prestamo.estado == "APROBADO")
    if fecha_inicio:
        query = query.where(Prestamo.fecha_aprobacion >= datetime.combine(fecha_inicio, datetime.min.time()))
    if fecha_fin:
        query = query.where(Prestamo.fecha_aprobacion <= datetime.combine(fecha_fin, datetime.max.time()))
    
    prestamos = db.execute(query).scalars().all()
    ids = [p.id for p in prestamos]
    
    total_financiamiento = sum(_safe_float(p.total_financiamiento) for p in prestamos) if prestamos else 0
    total_pagos_general = 0
    total_cuotas_pagadas_general = 0
    total_cuotas_pendientes_general = 0
    
    if ids:
        total_pagos_general = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Pago.monto_pagado), 0))
            .where(Pago.prestamo_id.in_(ids))
        ))
        total_cuotas_pagadas_general = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado == "PAGADO")
        ))
        total_cuotas_pendientes_general = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado != "PAGADO")
        ))
    
    resumen_data = [
        ["Metrica", "Valor"],
        ["Total Prestamos", str(len(prestamos))],
        ["Total Financiamiento", f"${total_financiamiento:,.2f}"],
        ["Total Pagos", f"${total_pagos_general:,.2f}"],
        ["Cuotas Pagadas ($)", f"${total_cuotas_pagadas_general:,.2f}"],
        ["Cuotas Pendientes ($)", f"${total_cuotas_pendientes_general:,.2f}"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 3*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Percentaje cobrado
    if total_financiamiento > 0:
        pct_cobrado = (total_pagos_general / total_financiamiento) * 100
    else:
        pct_cobrado = 0
    
    story.append(Paragraph(f"<b>Porcentaje Cobrado:</b> {pct_cobrado:.2f}%", styles['Normal']))
    
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


@router.get("/exportar/conciliacion")
def exportar_conciliacion(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_inicio: Optional[str] = Query(None, description="YYYY-MM-DD"),
    fecha_fin: Optional[str] = Query(None, description="YYYY-MM-DD"),
    cedulas: Optional[str] = Query(None, description="Cedulas separadas por coma"),
):
    """
    Genera reporte Conciliación en Excel o PDF con filtros opcionales.
    Al descargar Excel se eliminan automáticamente los datos temporales.
    
    Filtros:
    - fecha_inicio, fecha_fin: rango de fechas de aprobación (YYYY-MM-DD)
    - cedulas: lista de cédulas separadas por coma (ej: V12345678,E98765432)
    """
    # Parsear fechas
    fi = None
    ff = None
    if fecha_inicio:
        try:
            fi = date.fromisoformat(fecha_inicio)
        except ValueError:
            pass
    if fecha_fin:
        try:
            ff = date.fromisoformat(fecha_fin)
        except ValueError:
            pass
    
    # Parsear cédulas
    cedulas_list = None
    if cedulas:
        cedulas_list = [c.strip() for c in cedulas.split(",") if c.strip()]
    
    hoy_str = date.today().isoformat()
    
    try:
        if formato == "excel":
            content = _generar_excel_conciliacion(db, fi, ff, cedulas_list)
            # Borra temporales al descargar Excel
            db.execute(delete(ConciliacionTemporal))
            db.commit()
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=reporte_conciliacion_{hoy_str}.xlsx"},
            )
        else:  # PDF
            content = _generar_pdf_conciliacion(db, fi, ff)
            return Response(
                content=content,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=reporte_conciliacion_{hoy_str}.pdf"},
            )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generando reporte conciliacion: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error generando reporte: {str(e)}"
        )


@router.get("/conciliacion/resumen")
def get_conciliacion_resumen(
    db: Session = Depends(get_db),
    fecha_inicio: Optional[str] = Query(None, description="YYYY-MM-DD"),
    fecha_fin: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """
    Obtiene resumen de conciliación sin descargar reporte.
    Útil para mostrar en dashboard o vista previa.
    """
    fi = None
    ff = None
    if fecha_inicio:
        try:
            fi = date.fromisoformat(fecha_inicio)
        except ValueError:
            pass
    if fecha_fin:
        try:
            ff = date.fromisoformat(fecha_fin)
        except ValueError:
            pass
    
    query = select(Prestamo).where(Prestamo.estado == "APROBADO")
    if fi:
        query = query.where(Prestamo.fecha_aprobacion >= datetime.combine(fi, datetime.min.time()))
    if ff:
        query = query.where(Prestamo.fecha_aprobacion <= datetime.combine(ff, datetime.max.time()))
    
    prestamos = db.execute(query).scalars().all()
    ids = [p.id for p in prestamos]
    
    total_financiamiento = sum(_safe_float(p.total_financiamiento) for p in prestamos) if prestamos else 0
    total_pagos = 0
    total_cuotas_pagadas = 0
    total_cuotas_pendientes = 0
    
    if ids:
        total_pagos = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Pago.monto_pagado), 0))
            .where(Pago.prestamo_id.in_(ids))
        ))
        total_cuotas_pagadas = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado == "PAGADO")
        ))
        total_cuotas_pendientes = _safe_float(db.scalar(
            select(func.coalesce(func.sum(Cuota.monto), 0))
            .where(Cuota.prestamo_id.in_(ids), Cuota.estado != "PAGADO")
        ))
    
    porcentaje_cobrado = (total_pagos / total_financiamiento * 100) if total_financiamiento > 0 else 0
    
    return {
        "fecha_inicio": fi.isoformat() if fi else None,
        "fecha_fin": ff.isoformat() if ff else None,
        "total_filas": len(prestamos),
        "filas_procesadas": len(prestamos),
        "monto_total_financiamiento": round(total_financiamiento, 2),
        "monto_total_abonos": round(total_pagos, 2),
        "diferencia_total": round(total_financiamiento - total_pagos, 2),
        "cedulas_unicas": len(set(p.cedula for p in prestamos if p.cedula)),
        "filas_con_discrepancia": len([p for p in prestamos if (p.total_financiamiento or 0) != (total_cuotas_pagadas or 0)]),
    }

