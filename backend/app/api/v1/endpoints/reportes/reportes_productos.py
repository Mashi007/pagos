"""
Reportes por productos.
"""
import calendar
import io
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.cliente import Cliente
from app.models.cuota import Cuota
from app.models.prestamo import Prestamo

from app.api.v1.endpoints.reportes_utils import _safe_float, _parse_fecha, _periodos_desde_filtros

router = APIRouter(dependencies=[Depends(get_current_user)])


@router.get("/productos/por-mes")
def get_productos_por_mes(
    db: Session = Depends(get_db),
    meses: int = Query(12, ge=1, le=24, description="Cantidad de meses hacia atrás"),
    anos: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Productos por mes: una pestaña por mes. Columnas: Modelo vehículo, Total financiamiento, Valor activo (70%)."""
    resultado: dict = {"meses": []}
    periodos = _periodos_desde_filtros(anos, meses_list, meses)
    fecha_ref = func.coalesce(func.date(Prestamo.fecha_aprobacion), func.date(Prestamo.fecha_registro))

    for (ano, mes) in periodos:
        inicio = date(ano, mes, 1)
        _, ultimo = calendar.monthrange(ano, mes)
        fin = date(ano, mes, ultimo)

        rows = db.execute(
            select(
                Prestamo.modelo_vehiculo,
                func.coalesce(func.sum(Prestamo.total_financiamiento), 0).label("total_financiamiento"),
            )
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.estado == "APROBADO",
                fecha_ref >= inicio,
                fecha_ref <= fin,
            )
            .group_by(Prestamo.modelo_vehiculo)
            .order_by(func.sum(Prestamo.total_financiamiento).desc())
        ).fetchall()

        items = []
        for r in rows:
            total_fin = round(_safe_float(r.total_financiamiento), 2)
            valor_activo = round(total_fin * 0.70, 2)
            items.append({
                "modelo": r.modelo_vehiculo or "Sin modelo",
                "total_financiamiento": total_fin,
                "valor_activo": valor_activo,
            })

        resultado["meses"].append({
            "mes": mes,
            "ano": ano,
            "label": f"{mes:02d}/{ano}",
            "items": items,
        })

    return resultado


@router.get("/productos")
def get_reporte_productos(
    db: Session = Depends(get_db),
    fecha_corte: Optional[str] = Query(None),
):
    """Reporte por producto. Datos reales desde BD."""
    fc = _parse_fecha(fecha_corte)
    productos = db.execute(
        select(Prestamo.producto)
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        .distinct()
    ).fetchall()
    resumen_por_producto: List[dict] = []
    for (producto,) in productos:
        if not producto:
            continue
        prestamos = db.execute(
            select(Prestamo.id)
            .select_from(Prestamo)
            .join(Cliente, Prestamo.cliente_id == Cliente.id)
            .where(
                Cliente.estado == "ACTIVO",
                Prestamo.producto == producto,
                Prestamo.estado == "APROBADO",
            )
        ).fetchall()
        ids = [x[0] for x in prestamos]
        total_prestamos = len(ids)
        total_clientes = db.scalar(select(func.count(func.distinct(Prestamo.cliente_id))).select_from(Prestamo).where(Prestamo.id.in_(ids))) or 0
        cartera_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.prestamo_id.in_(ids)))) or 0
        promedio_prestamo = _safe_float(db.scalar(select(func.avg(Prestamo.total_financiamiento)).select_from(Prestamo).where(Prestamo.id.in_(ids)))) or 0
        total_cobrado = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.isnot(None), Cuota.prestamo_id.in_(ids)))) or 0
        morosidad_total = _safe_float(db.scalar(select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < fc, Cuota.prestamo_id.in_(ids)))) or 0
        porcentaje_cobrado = (total_cobrado / (total_cobrado + cartera_total) * 100) if (total_cobrado + cartera_total) else 0
        resumen_por_producto.append({
            "producto": producto,
            "total_prestamos": total_prestamos,
            "total_clientes": total_clientes,
            "cartera_total": cartera_total,
            "promedio_prestamo": round(promedio_prestamo, 2),
            "total_cobrado": total_cobrado,
            "morosidad_total": morosidad_total,
            "porcentaje_cobrado": round(porcentaje_cobrado, 2),
        })
    productos_por_concesionario: List[dict] = []
    concesionarios = db.execute(
        select(Prestamo.concesionario)
        .select_from(Prestamo)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Cliente.estado == "ACTIVO", Prestamo.estado == "APROBADO")
        .distinct()
    ).fetchall()
    for (conc,) in concesionarios:
        if not conc:
            continue
        for (producto,) in productos:
            if not producto:
                continue
            cnt = db.scalar(
                select(func.count())
                .select_from(Prestamo)
                .join(Cliente, Prestamo.cliente_id == Cliente.id)
                .where(
                    Cliente.estado == "ACTIVO",
                    Prestamo.concesionario == conc,
                    Prestamo.producto == producto,
                    Prestamo.estado == "APROBADO",
                )
            ) or 0
            if cnt:
                monto = _safe_float(
                    db.scalar(
                        select(func.sum(Prestamo.total_financiamiento))
                        .select_from(Prestamo)
                        .join(Cliente, Prestamo.cliente_id == Cliente.id)
                        .where(
                            Cliente.estado == "ACTIVO",
                            Prestamo.concesionario == conc,
                            Prestamo.producto == producto,
                            Prestamo.estado == "APROBADO",
                        )
                    )
                    or 0
                )
                productos_por_concesionario.append({"concesionario": conc, "producto": producto, "cantidad_prestamos": cnt, "monto_total": monto})
    return {
        "fecha_corte": fc.isoformat(),
        "resumen_por_producto": resumen_por_producto,
        "productos_por_concesionario": productos_por_concesionario,
        "tendencia_mensual": [],
    }


def _generar_excel_productos_por_mes(data_por_mes: dict) -> bytes:
    """Genera Excel con una pestaña por mes."""
    import openpyxl

    wb = openpyxl.Workbook()
    meses_data = data_por_mes.get("meses", [])
    meses_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    for idx, mes_data in enumerate(meses_data):
        mes = mes_data.get("mes", 1)
        ano = mes_data.get("ano", date.today().year)
        mes_nombre = meses_es.get(mes, "")
        label = mes_data.get("label", f"{mes:02d}/{ano}")
        sheet_name = f"{mes_nombre} {ano}"[:31]

        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.append(["Reporte de Productos", label])
        ws.append([])
        ws.append(["Modelo de vehículo", "Total financiamiento ($)", "Valor del activo ($)"])
        items = mes_data.get("items", [])
        if items:
            for r in items:
                ws.append([r.get("modelo", ""), r.get("total_financiamiento", 0), r.get("valor_activo", 0)])
        else:
            ws.append(["Sin datos", 0, 0])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_productos(data: dict) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Productos", styles["Title"]))
    story.append(Paragraph(f"Fecha de corte: {data.get('fecha_corte', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    rows = [["Producto", "Préstamos", "Clientes", "Cartera", "Promedio", "Cobrado", "Mora", "% Cobrado"]]
    for r in data.get("resumen_por_producto", []):
        rows.append([r.get("producto", ""), str(r.get("total_prestamos", 0)), str(r.get("total_clientes", 0)), str(r.get("cartera_total", 0)), str(r.get("promedio_prestamo", 0)), str(r.get("total_cobrado", 0)), str(r.get("morosidad_total", 0)), str(r.get("porcentaje_cobrado", 0))])
    t = Table(rows)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
    story.append(t)
    prod_conc = data.get("productos_por_concesionario", [])
    if prod_conc:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Productos por concesionario", styles["Heading2"]))
        rows2 = [["Concesionario", "Producto", "Cant. préstamos", "Monto total"]]
        for r in prod_conc:
            rows2.append([r.get("concesionario", ""), r.get("producto", ""), str(r.get("cantidad_prestamos", 0)), str(r.get("monto_total", 0))])
        t2 = Table(rows2)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), "#e0e0e0"), ("GRID", (0, 0), (-1, -1), 0.5, "#ccc")]))
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


@router.get("/exportar/productos")
def exportar_productos(
    db: Session = Depends(get_db),
    formato: str = Query("excel", pattern="^(excel|pdf)$"),
    fecha_corte: Optional[str] = Query(None),
    meses: int = Query(12, ge=1, le=24, description="Para Excel: cantidad de meses"),
    anos: Optional[str] = Query(None),
    meses_list: Optional[str] = Query(None),
):
    """Exporta reporte productos. Excel: una pestaña por mes. PDF: clásico."""
    if formato == "excel":
        data_por_mes = get_productos_por_mes(db=db, meses=meses, anos=anos, meses_list=meses_list)
        content = _generar_excel_productos_por_mes(data_por_mes)
        hoy_str = date.today().isoformat()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_productos_{hoy_str}.xlsx"},
        )
    data = get_reporte_productos(db=db, fecha_corte=fecha_corte)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["Producto", "Préstamos", "Clientes", "Cartera", "Promedio", "Cobrado", "Mora", "% Cobrado"])
    for r in data.get("resumen_por_producto", []):
        ws.append([r.get("producto", ""), r.get("total_prestamos", 0), r.get("total_clientes", 0), r.get("cartera_total", 0), r.get("promedio_prestamo", 0), r.get("total_cobrado", 0), r.get("morosidad_total", 0), r.get("porcentaje_cobrado", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    if formato == "pdf":
        pdf_content = _generar_pdf_productos(data)
        return Response(content=pdf_content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_productos_{data['fecha_corte']}.pdf"})
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_productos_{data['fecha_corte']}.xlsx"})
