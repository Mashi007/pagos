"""
Endpoints de cobranzas. Datos reales desde BD (Cuota, Prestamo, Cliente).
Estructura de respuestas compatible con cobranzasService.ts.
Informes: JSON, PDF y Excel generados con datos reales.
"""
import io
from datetime import date, datetime, timedelta
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.cliente import Cliente
from app.models.cuota import Cuota
from app.models.prestamo import Prestamo

router = APIRouter(dependencies=[Depends(get_current_user)])


class MLImpagoUpdateBody(BaseModel):
    """Body para PUT /prestamos/{id}/ml-impago (marcar ML impago manual)."""
    nivel_riesgo: str = Field(..., min_length=1, max_length=50)
    probabilidad_impago: float = Field(..., ge=0, le=1)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _generar_excel_clientes_atrasados(clientes: List[dict], resumen: dict) -> bytes:
    """Genera Excel para informe clientes atrasados."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes Atrasados"
    ws.append(["Cédula", "Nombres", "Teléfono", "Email", "Estado"])
    for c in clientes:
        ws.append([
            c.get("cedula") or "",
            c.get("nombres") or "",
            c.get("telefono") or "",
            c.get("email") or "",
            c.get("estado") or "",
        ])
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Indicador", "Valor"])
    ws2.append(["Total cuotas vencidas", resumen.get("total_cuotas_vencidas", 0)])
    ws2.append(["Monto total adeudado", resumen.get("monto_total_adeudado", 0)])
    ws2.append(["Clientes atrasados", resumen.get("clientes_atrasados", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_rendimiento_analista(datos: List[dict]) -> bytes:
    """Genera Excel para informe rendimiento por analista."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendimiento por Analista"
    ws.append(["Analista", "Cantidad clientes", "Monto total"])
    for r in datos:
        ws.append([
            r.get("analista") or "",
            r.get("cantidad_clientes", 0),
            r.get("monto_total", 0),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_montos_periodo(datos: List[dict]) -> bytes:
    """Genera Excel para informe montos vencidos por período."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Montos por período"
    ws.append(["Mes", "Mes (display)", "Cantidad cuotas", "Monto total"])
    for r in datos:
        ws.append([
            r.get("mes") or "",
            r.get("mes_display") or "",
            r.get("cantidad_cuotas", 0),
            r.get("monto_total", 0),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_antiguedad_saldos(datos: List[dict]) -> bytes:
    """Genera Excel para informe antigüedad de saldos (puede estar vacío)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antigüedad de saldos"
    ws.append(["Rango", "Cantidad cuotas", "Monto total", "Porcentaje"])
    for r in (datos or []):
        ws.append([
            r.get("rango") or "",
            r.get("cantidad_cuotas", 0),
            r.get("monto_total", 0),
            r.get("porcentaje", 0),
        ])
    if not datos:
        ws.append(["Sin datos", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_excel_resumen_ejecutivo(resumen: dict) -> bytes:
    """Genera Excel para resumen ejecutivo."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen ejecutivo"
    ws.append(["Indicador", "Valor"])
    ws.append(["Total cuotas vencidas", resumen.get("total_cuotas_vencidas", 0)])
    ws.append(["Monto total adeudado", resumen.get("monto_total_adeudado", 0)])
    ws.append(["Clientes atrasados", resumen.get("clientes_atrasados", 0)])
    if resumen.get("diagnosticos"):
        ws.append([])
        ws.append(["Diagnóstico"])
        for k, v in resumen["diagnosticos"].items():
            ws.append([str(k), str(v)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generar_pdf_clientes_atrasados(clientes: List[dict], resumen: dict) -> bytes:
    """Genera PDF para informe clientes atrasados."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Informe: Clientes Atrasados", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Resumen", styles["Heading2"]))
    resumen_data = [
        ["Total cuotas vencidas", str(resumen.get("total_cuotas_vencidas", 0))],
        ["Monto total adeudado", str(resumen.get("monto_total_adeudado", 0))],
        ["Clientes atrasados", str(resumen.get("clientes_atrasados", 0))],
    ]
    t1 = Table(resumen_data)
    t1.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke)]))
    story.append(t1)
    story.append(Spacer(1, 16))
    story.append(Paragraph("Detalle de clientes", styles["Heading2"]))
    if clientes:
        headers = ["Cédula", "Nombres", "Teléfono", "Email"]
        rows = [headers] + [[str(c.get(k, "")) for k in ["cedula", "nombres", "telefono", "email"]] for c in clientes]
        t2 = Table(rows)
        t2.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
        story.append(t2)
    else:
        story.append(Paragraph("No hay clientes atrasados.", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


def _generar_pdf_rendimiento_analista(datos: List[dict]) -> bytes:
    """Genera PDF para informe rendimiento por analista."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph("Informe: Rendimiento por Analista", styles["Title"]), Spacer(1, 12)]
    if datos:
        rows = [["Analista", "Cantidad clientes", "Monto total"]] + [[str(r.get("analista", "")), str(r.get("cantidad_clientes", 0)), str(r.get("monto_total", 0))] for r in datos]
        t = Table(rows)
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke)]))
        story.append(t)
    else:
        story.append(Paragraph("No hay datos.", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


def _generar_pdf_montos_periodo(datos: List[dict]) -> bytes:
    """Genera PDF para informe montos vencidos por período."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph("Informe: Montos vencidos por período", styles["Title"]), Spacer(1, 12)]
    if datos:
        rows = [["Mes", "Cantidad cuotas", "Monto total"]] + [[str(r.get("mes_display") or r.get("mes", "")), str(r.get("cantidad_cuotas", 0)), str(r.get("monto_total", 0))] for r in datos]
        t = Table(rows)
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke)]))
        story.append(t)
    else:
        story.append(Paragraph("No hay datos.", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


def _generar_pdf_antiguedad_saldos(datos: List[dict]) -> bytes:
    """Genera PDF para informe antigüedad de saldos."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph("Informe: Antigüedad de saldos", styles["Title"]), Spacer(1, 12)]
    if datos:
        rows = [["Rango", "Cantidad cuotas", "Monto total", "Porcentaje"]] + [[str(r.get("rango", "")), str(r.get("cantidad_cuotas", 0)), str(r.get("monto_total", 0)), str(r.get("porcentaje", 0))] for r in datos]
        t = Table(rows)
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke)]))
        story.append(t)
    else:
        story.append(Paragraph("No hay datos.", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


def _generar_pdf_resumen_ejecutivo(resumen: dict) -> bytes:
    """Genera PDF para resumen ejecutivo."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph("Resumen ejecutivo - Cobranzas", styles["Title"]), Spacer(1, 12)]
    rows = [
        ["Total cuotas vencidas", str(resumen.get("total_cuotas_vencidas", 0))],
        ["Monto total adeudado", str(resumen.get("monto_total_adeudado", 0))],
        ["Clientes atrasados", str(resumen.get("clientes_atrasados", 0))],
    ]
    t = Table(rows)
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke)]))
    story.append(t)
    if resumen.get("diagnosticos"):
        story.append(Spacer(1, 12))
        story.append(Paragraph("Diagnóstico", styles["Heading2"]))
        diag = resumen["diagnosticos"]
        rows2 = [[str(k), str(v)] for k, v in diag.items()]
        t2 = Table(rows2)
        story.append(t2)
    doc.build(story)
    return buf.getvalue()


def _nombre_archivo_informe(slug: str, ext: str) -> str:
    """Nombre de archivo para descarga: informe_<slug>_YYYY-MM-DD.<ext>."""
    return f"informe_{slug}_{date.today().isoformat()}.{ext}"


# ---------------------------------------------------------------------------
# Resumen y diagnóstico
# ---------------------------------------------------------------------------

@router.get("/resumen")
def get_resumen(
    incluir_admin: bool = Query(False),
    incluir_diagnostico: bool = Query(False),
    db: Session = Depends(get_db),
):
    """Resumen de cobranzas desde BD: total cuotas vencidas, monto adeudado, clientes atrasados."""
    hoy = date.today()
    total_cuotas_vencidas = db.scalar(
        select(func.count()).select_from(Cuota).where(
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < hoy,
        )
    ) or 0
    monto_total_adeudado = db.scalar(
        select(func.coalesce(func.sum(Cuota.monto), 0)).select_from(Cuota).where(
            Cuota.fecha_pago.is_(None)
        )
    ) or 0
    # Clientes atrasados: distinct clientes con al menos una cuota vencida (vía préstamo si cuotas.cliente_id no está poblado)
    subq_clientes = (
        select(Prestamo.cliente_id)
        .select_from(Prestamo)
        .join(Cuota, Cuota.prestamo_id == Prestamo.id)
        .where(
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < hoy,
            Prestamo.cliente_id.isnot(None),
        )
        .distinct()
    )
    clientes_atrasados = db.scalar(select(func.count()).select_from(subq_clientes.subquery())) or 0
    data: dict[str, Any] = {
        "total_cuotas_vencidas": total_cuotas_vencidas,
        "monto_total_adeudado": _safe_float(monto_total_adeudado),
        "clientes_atrasados": clientes_atrasados,
    }
    if incluir_diagnostico:
        data["diagnosticos"] = _diagnostico_desde_bd(db)
    return data


@router.get("/diagnostico")
def get_diagnostico(db: Session = Depends(get_db)):
    """Diagnóstico de cobranzas desde BD (cuotas, vencidas, estados)."""
    diag = _diagnostico_desde_bd(db)
    return {
        "diagnosticos": diag,
        "analisis_filtros": {
            "cuotas_perdidas_por_estado": 0,
            "cuotas_perdidas_por_admin": 0,
            "cuotas_perdidas_por_user_admin": 0,
        },
    }


def ejecutar_actualizacion_reportes(db: Session) -> dict[str, Any]:
    """
    Ejecuta la lógica de actualización de reportes de cobranzas (resumen + diagnóstico).
    Usado por el scheduler a las 6:00 y 13:00. Retorna el resumen para logging.
    """
    resumen = get_resumen(incluir_admin=False, incluir_diagnostico=True, db=db)
    return resumen


def _diagnostico_desde_bd(db: Session) -> dict[str, Any]:
    """Calcula diagnóstico desde tabla cuotas."""
    hoy = date.today()
    total_cuotas_bd = db.scalar(select(func.count()).select_from(Cuota)) or 0
    cuotas_vencidas_solo_fecha = db.scalar(
        select(func.count()).select_from(Cuota).where(
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < hoy,
        )
    ) or 0
    cuotas_pago_incompleto = 0
    cuotas_vencidas_incompletas = 0
    rows_estado = db.execute(
        select(Prestamo.estado, func.count(func.distinct(Prestamo.id)))
        .select_from(Prestamo)
        .join(Cuota, Cuota.prestamo_id == Prestamo.id)
        .where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < hoy)
        .group_by(Prestamo.estado)
    ).all()
    estados_prestamos_con_cuotas_vencidas = {r[0]: r[1] for r in rows_estado}
    return {
        "total_cuotas_bd": total_cuotas_bd,
        "cuotas_vencidas_solo_fecha": cuotas_vencidas_solo_fecha,
        "cuotas_pago_incompleto": cuotas_pago_incompleto,
        "cuotas_vencidas_incompletas": cuotas_vencidas_incompletas,
        "estados_prestamos_con_cuotas_vencidas": estados_prestamos_con_cuotas_vencidas,
    }


# ---------------------------------------------------------------------------
# Clientes atrasados
# ---------------------------------------------------------------------------

def _listar_clientes_atrasados(
    db: Session,
    *,
    dias_retraso: Optional[int] = None,
    dias_retraso_min: Optional[int] = None,
    dias_retraso_max: Optional[int] = None,
    analista_filtro: Optional[str] = None,
    incluir_ml: bool = True,
) -> List[dict]:
    """
    Lista de préstamos con cuotas atrasadas (una fila por préstamo).
    Reutilizado por GET /clientes-atrasados y GET /por-analista/{analista}/clientes.
    """
    hoy = date.today()
    q = (
        select(
            Prestamo.id.label("prestamo_id"),
            Cliente.cedula,
            Cliente.nombres,
            Cliente.telefono,
            Cliente.email,
            Cliente.estado,
            func.coalesce(Prestamo.analista, "Sin analista").label("analista"),
            func.count().label("cuotas_vencidas"),
            func.coalesce(func.sum(Cuota.monto), 0).label("total_adeudado"),
            func.min(Cuota.fecha_vencimiento).label("fecha_primera_vencida"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < hoy,
        )
        .group_by(Prestamo.id, Prestamo.analista, Cliente.id, Cliente.cedula, Cliente.nombres, Cliente.telefono, Cliente.email, Cliente.estado)
    )
    if dias_retraso is not None:
        limite = hoy - timedelta(days=dias_retraso)
        q = q.where(Cuota.fecha_vencimiento <= limite)
    if dias_retraso_min is not None:
        limite_min = hoy - timedelta(days=dias_retraso_min)
        q = q.where(Cuota.fecha_vencimiento >= limite_min)
    if dias_retraso_max is not None:
        limite_max = hoy - timedelta(days=dias_retraso_max)
        q = q.where(Cuota.fecha_vencimiento <= limite_max)
    if analista_filtro is not None:
        if analista_filtro == "Sin analista":
            q = q.where(Prestamo.analista.is_(None))
        else:
            q = q.where(Prestamo.analista == analista_filtro)
    rows = db.execute(q).all()
    out: List[dict] = []
    for r in rows:
        item = {
            "cedula": r.cedula,
            "nombres": r.nombres,
            "analista": r.analista,
            "telefono": r.telefono or "",
            "email": r.email or "",
            "estado": r.estado or "",
            "prestamo_id": r.prestamo_id,
            "cuotas_vencidas": r.cuotas_vencidas,
            "total_adeudado": _safe_float(r.total_adeudado),
            "fecha_primera_vencida": r.fecha_primera_vencida.isoformat() if r.fecha_primera_vencida else None,
        }
        if incluir_ml:
            prestamo = db.get(Prestamo, r.prestamo_id)
            if prestamo:
                nivel = getattr(prestamo, "ml_impago_nivel_riesgo_manual", None) or getattr(prestamo, "ml_impago_nivel_riesgo_calculado", None)
                prob = getattr(prestamo, "ml_impago_probabilidad_manual", None) or getattr(prestamo, "ml_impago_probabilidad_calculada", None)
                if nivel is not None or prob is not None:
                    item["ml_impago"] = {
                        "nivel_riesgo": nivel or "N/A",
                        "probabilidad_impago": _safe_float(prob),
                        "prediccion": f"Riesgo {nivel or 'N/A'}",
                        "es_manual": getattr(prestamo, "ml_impago_nivel_riesgo_manual", None) is not None,
                    }
                else:
                    item["ml_impago"] = None
            else:
                item["ml_impago"] = None
        out.append(item)
    return out


@router.get("/clientes-atrasados")
def get_clientes_atrasados(
    dias_retraso: Optional[int] = Query(None),
    dias_retraso_min: Optional[int] = Query(None),
    dias_retraso_max: Optional[int] = Query(None),
    analista: Optional[str] = Query(None),
    incluir_admin: bool = Query(False),
    incluir_ml: bool = Query(True),
    db: Session = Depends(get_db),
):
    """
    Lista de préstamos con cuotas atrasadas desde BD.
    Una fila por préstamo: cedula, nombres, analista, cuotas_vencidas, total_adeudado, fecha_primera_vencida.
    Compatible con la pestaña 'Clientes con Cuotas Impagas'. Opcional: filtrar por analista (informes).
    """
    return _listar_clientes_atrasados(
        db,
        dias_retraso=dias_retraso,
        dias_retraso_min=dias_retraso_min,
        dias_retraso_max=dias_retraso_max,
        analista_filtro=analista,
        incluir_ml=incluir_ml,
    )


@router.get("/clientes-por-cantidad-pagos")
def get_clientes_por_cantidad_pagos(
    cantidad_pagos: int = Query(...),
    db: Session = Depends(get_db),
):
    """Clientes con exactamente N cuotas atrasadas (desde BD)."""
    hoy = date.today()
    subq = (
        select(Cuota.cliente_id, func.count().label("cnt"))
        .where(
            Cuota.fecha_pago.is_(None),
            Cuota.fecha_vencimiento < hoy,
            Cuota.cliente_id.isnot(None),
        )
        .group_by(Cuota.cliente_id)
        .having(func.count() == cantidad_pagos)
    )
    rows = db.execute(subq).all()
    ids = [r[0] for r in rows]
    if not ids:
        return []
    clientes = db.execute(select(Cliente).where(Cliente.id.in_(ids))).scalars().all()
    return [
        {"id": c.id, "cedula": c.cedula, "nombres": c.nombres, "telefono": c.telefono, "email": c.email}
        for c in clientes
    ]


# ---------------------------------------------------------------------------
# Por analista
# ---------------------------------------------------------------------------

@router.get("/por-analista")
def get_cobranzas_por_analista(
    incluir_admin: bool = Query(False),
    db: Session = Depends(get_db),
):
    """Cobranzas agrupadas por analista desde BD. Cuenta clientes por Prestamo.cliente_id (más fiable que Cuota.cliente_id)."""
    hoy = date.today()
    q = (
        select(
            func.coalesce(Prestamo.analista, "Sin analista").label("analista"),
            func.count(func.distinct(Prestamo.cliente_id)).label("cantidad_clientes"),
            func.coalesce(func.sum(Cuota.monto), 0).label("monto_total"),
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento < hoy)
        .group_by(Prestamo.analista)
    )
    rows = db.execute(q).all()
    return [
        {
            "analista": r.analista,
            "cantidad_clientes": r.cantidad_clientes,
            "monto_total": _safe_float(r.monto_total),
        }
        for r in rows
    ]


@router.get("/por-analista/{analista}/clientes")
def get_clientes_por_analista(analista: str, db: Session = Depends(get_db)):
    """Clientes atrasados de un analista desde BD. Misma estructura que GET /clientes-atrasados (una fila por préstamo)."""
    return _listar_clientes_atrasados(
        db,
        analista_filtro=analista,
        incluir_ml=False,
    )


# ---------------------------------------------------------------------------
# Montos por mes
# ---------------------------------------------------------------------------

@router.get("/montos-por-mes")
def get_montos_por_mes(
    incluir_admin: bool = Query(False),
    db: Session = Depends(get_db),
):
    """Montos vencidos por mes desde BD (cuotas no pagadas con vencimiento en ese mes)."""
    q = (
        select(
            func.to_char(func.date_trunc("month", Cuota.fecha_vencimiento), "YYYY-MM").label("mes"),
            func.count().label("cantidad_cuotas"),
            func.coalesce(func.sum(Cuota.monto), 0).label("monto_total"),
        )
        .select_from(Cuota)
        .where(Cuota.fecha_pago.is_(None), Cuota.fecha_vencimiento.isnot(None))
        .group_by(func.date_trunc("month", Cuota.fecha_vencimiento))
        .order_by(func.date_trunc("month", Cuota.fecha_vencimiento).desc())
        .limit(12)
    )
    rows = db.execute(q).all()
    nombres = ("Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic")
    out = []
    for r in rows:
        try:
            y, m = r.mes.split("-")
            mes_display = f"{nombres[int(m) - 1]} {y}"
        except Exception:
            mes_display = r.mes
        out.append({
            "mes": r.mes,
            "mes_display": mes_display,
            "cantidad_cuotas": r.cantidad_cuotas,
            "monto_total": _safe_float(r.monto_total),
        })
    return out


# ---------------------------------------------------------------------------
# Informes (JSON, PDF y Excel desde BD)
# ---------------------------------------------------------------------------

@router.get("/informes/clientes-atrasados")
def get_informe_clientes_atrasados(
    dias_retraso_min: Optional[int] = Query(None),
    dias_retraso_max: Optional[int] = Query(None),
    analista: Optional[str] = Query(None),
    formato: Optional[str] = Query("json"),
    db: Session = Depends(get_db),
):
    """Informe clientes atrasados desde BD. formato=json|pdf|excel. Opcional: filtrar por analista."""
    clientes = _listar_clientes_atrasados(
        db,
        dias_retraso_min=dias_retraso_min,
        dias_retraso_max=dias_retraso_max,
        analista_filtro=analista,
        incluir_ml=False,
    )
    resumen = get_resumen(incluir_diagnostico=False, db=db)
    if formato == "excel":
        content = _generar_excel_clientes_atrasados(clientes, resumen)
        filename = _nombre_archivo_informe("clientes_atrasados", "xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if formato == "pdf":
        content = _generar_pdf_clientes_atrasados(clientes, resumen)
        filename = _nombre_archivo_informe("clientes_atrasados", "pdf")
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return {"clientes": clientes, "resumen": resumen}


@router.get("/informes/rendimiento-analista")
def get_informe_rendimiento_analista(
    formato: str = Query("json"),
    db: Session = Depends(get_db),
):
    """Informe rendimiento por analista. formato=json|pdf|excel."""
    datos = get_cobranzas_por_analista(db=db)
    if formato == "excel":
        content = _generar_excel_rendimiento_analista(datos)
        filename = _nombre_archivo_informe("rendimiento_analista", "xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if formato == "pdf":
        content = _generar_pdf_rendimiento_analista(datos)
        filename = _nombre_archivo_informe("rendimiento_analista", "pdf")
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return datos


@router.get("/informes/montos-vencidos-periodo")
def get_informe_montos_periodo(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    formato: Optional[str] = Query("json"),
    db: Session = Depends(get_db),
):
    """Informe montos vencidos por período. formato=json|pdf|excel."""
    datos = get_montos_por_mes(db=db)
    if formato == "excel":
        content = _generar_excel_montos_periodo(datos)
        filename = _nombre_archivo_informe("montos_periodo", "xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if formato == "pdf":
        content = _generar_pdf_montos_periodo(datos)
        filename = _nombre_archivo_informe("montos_periodo", "pdf")
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return datos


@router.get("/informes/antiguedad-saldos")
def get_informe_antiguedad_saldos(formato: str = Query("json"), db: Session = Depends(get_db)):
    """Informe antigüedad de saldos. formato=json|pdf|excel. JSON devuelve lista (puede estar vacía)."""
    datos: List[dict] = []  # Sin consulta específica de rangos por ahora; estructura lista vacía
    if formato == "excel":
        content = _generar_excel_antiguedad_saldos(datos)
        filename = _nombre_archivo_informe("antiguedad_saldos", "xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if formato == "pdf":
        content = _generar_pdf_antiguedad_saldos(datos)
        filename = _nombre_archivo_informe("antiguedad_saldos", "pdf")
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return datos


@router.get("/informes/resumen-ejecutivo")
def get_informe_resumen_ejecutivo(formato: str = Query("json"), db: Session = Depends(get_db)):
    """Informe resumen ejecutivo. formato=json|pdf|excel."""
    resumen = get_resumen(incluir_diagnostico=True, db=db)
    if formato == "excel":
        content = _generar_excel_resumen_ejecutivo(resumen)
        filename = _nombre_archivo_informe("resumen_ejecutivo", "xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if formato == "pdf":
        content = _generar_pdf_resumen_ejecutivo(resumen)
        filename = _nombre_archivo_informe("resumen_ejecutivo", "pdf")
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return resumen


# ---------------------------------------------------------------------------
# Notificaciones y ML
# ---------------------------------------------------------------------------

@router.post("/notificaciones/atrasos")
def procesar_notificaciones_atrasos(db: Session = Depends(get_db)):
    """Procesar notificaciones de atrasos. Devuelve estadísticas desde BD."""
    resumen = get_resumen(incluir_diagnostico=False, db=db)
    return {
        "mensaje": "Procesamiento de notificaciones.",
        "estadisticas": resumen,
    }


@router.put("/prestamos/{prestamo_id}/ml-impago")
def actualizar_ml_impago(prestamo_id: int, body: MLImpagoUpdateBody, db: Session = Depends(get_db)):
    """Marcar ML impago manual y persistir en BD."""
    from fastapi import HTTPException
    p = db.get(Prestamo, prestamo_id)
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    p.ml_impago_nivel_riesgo_manual = body.nivel_riesgo
    p.ml_impago_probabilidad_manual = body.probabilidad_impago
    db.commit()
    db.refresh(p)
    return {"ok": True, "prestamo_id": prestamo_id}


@router.delete("/prestamos/{prestamo_id}/ml-impago")
def eliminar_ml_impago_manual(prestamo_id: int, db: Session = Depends(get_db)):
    """Quitar ML impago manual (volver a valores calculados) y persistir en BD."""
    from fastapi import HTTPException
    p = db.get(Prestamo, prestamo_id)
    if not p:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")
    p.ml_impago_nivel_riesgo_manual = None
    p.ml_impago_probabilidad_manual = None
    db.commit()
    db.refresh(p)
    return {"ok": True, "prestamo_id": prestamo_id}
