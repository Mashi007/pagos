# -*- coding: utf-8 -*-
"""
Excel «Saldos menores 200»: deudores cuyo saldo final del préstamo es ≤ 200 USD
(con ese monto terminan de pagar).

Columnas: cédula, nombres, teléfono, email, saldo final, cuotas vencidas, cuotas mora.

Cuotas vencidas / mora: se cuentan solo filas con saldo pendiente > 0 y ≤ 200 USD
(estado VENCIDO o MORA según `clasificar_estado_cuota`).
"""
from __future__ import annotations

import io
import logging
from decimal import Decimal
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from app.models.cliente import Cliente
from app.models.cuota import Cuota
from app.models.prestamo import Prestamo
from app.services.cuota_estado import clasificar_estado_cuota, hoy_negocio

logger = logging.getLogger(__name__)

SALDO_MAX_USD = Decimal("200.00")


def _dec(x: Any) -> Decimal:
    if x is None:
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def _saldo_cuota(monto: Any, total_pagado: Any) -> Decimal:
    sal = (_dec(monto) - _dec(total_pagado)).quantize(Decimal("0.01"))
    return sal if sal > Decimal("0.00") else Decimal("0.00")


def _cuenta_como_enumerable(saldo: Decimal) -> bool:
    """Saldo a pagar > 0 y ≤ 200 USD."""
    return Decimal("0.00") < saldo <= SALDO_MAX_USD


def construir_filas_saldos_menores_200(db: Session) -> List[Dict[str, Any]]:
    """
    Préstamos APROBADO con cliente ACTIVO (si hay estado), saldo final en (0, 200].
    Una fila por préstamo.
    """
    hoy = hoy_negocio()

    # Candidatos: saldo total pendiente ≤ 200 (SQL) para no cargar toda la cartera.
    m = func.coalesce(Cuota.monto, 0)
    tp = func.coalesce(Cuota.total_pagado, 0)
    per = func.greatest(0, m - tp)
    saldo_sq = (
        select(
            Cuota.prestamo_id.label("prestamo_id"),
            func.sum(per).label("saldo_final"),
        )
        .group_by(Cuota.prestamo_id)
        .having(
            and_(
                func.sum(per) > 0,
                func.sum(per) <= float(SALDO_MAX_USD),
            )
        )
        .subquery()
    )

    q = (
        select(
            Prestamo.id,
            Prestamo.cedula,
            Prestamo.nombres,
            Cliente.telefono,
            Cliente.email,
            saldo_sq.c.saldo_final,
        )
        .join(saldo_sq, saldo_sq.c.prestamo_id == Prestamo.id)
        .join(Cliente, Cliente.id == Prestamo.cliente_id)
        .where(Prestamo.estado == "APROBADO")
    )
    # Si la columna existe y se usa en cartera, respetar ACTIVO.
    if hasattr(Cliente, "estado"):
        q = q.where(func.upper(func.coalesce(Cliente.estado, "")) == "ACTIVO")

    q = q.order_by(Prestamo.cedula.asc(), Prestamo.id.asc())
    rows = list(db.execute(q).all())
    if not rows:
        return []

    pids = [int(r[0]) for r in rows]
    cuotas_por: Dict[int, List[Tuple[Any, Any, Any]]] = {pid: [] for pid in pids}
    for cu in db.execute(
        select(Cuota.prestamo_id, Cuota.monto, Cuota.total_pagado, Cuota.fecha_vencimiento)
        .where(Cuota.prestamo_id.in_(pids))
        .order_by(Cuota.prestamo_id.asc(), Cuota.fecha_vencimiento.asc())
    ).all():
        pid = int(cu[0])
        if pid in cuotas_por:
            cuotas_por[pid].append((cu[1], cu[2], cu[3]))

    out: List[Dict[str, Any]] = []
    for r in rows:
        pid = int(r[0])
        saldo_final = _dec(r[5]).quantize(Decimal("0.01"))
        if not (Decimal("0.00") < saldo_final <= SALDO_MAX_USD):
            continue
        n_venc = 0
        n_mora = 0
        for monto, tot_pag, fv in cuotas_por.get(pid, []):
            sal = _saldo_cuota(monto, tot_pag)
            if not _cuenta_como_enumerable(sal):
                continue
            est = clasificar_estado_cuota(
                float(_dec(tot_pag)),
                float(_dec(monto)),
                fv,
                hoy,
            )
            if est == "VENCIDO":
                n_venc += 1
            elif est == "MORA":
                n_mora += 1
        out.append(
            {
                "prestamo_id": pid,
                "cedula": (r[1] or "").strip(),
                "nombres": (r[2] or "").strip(),
                "telefono": (r[3] or "").strip() if r[3] is not None else "",
                "email": (r[4] or "").strip() if r[4] is not None else "",
                "saldo_final": float(saldo_final),
                "cuotas_vencidas": n_venc,
                "cuotas_mora": n_mora,
            }
        )
    return out


def construir_excel_saldos_menores_200(db: Session) -> Tuple[bytes, int]:
    """Devuelve (xlsx_bytes, n_filas_datos)."""
    filas = construir_filas_saldos_menores_200(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Saldos menores 200"

    headers = [
        "Cédula",
        "Nombres",
        "Teléfono",
        "Email",
        "Saldo final",
        "Cuotas vencidas",
        "Cuotas mora",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for f in filas:
        ws.append(
            [
                f["cedula"],
                f["nombres"],
                f["telefono"],
                f["email"],
                f["saldo_final"],
                f["cuotas_vencidas"],
                f["cuotas_mora"],
            ]
        )

    # Anchos legibles
    widths = [16, 36, 16, 32, 14, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(filas)
