"""
Excel Cédula | Cuota | cuotas en mora (1 jun y hoy) | pagos parciales a mora | saldo total, hoja Drive.

Cuota: prestamos.cuota_periodo (APROBADO; si no hay, LIQUIDADO, DESISTIMIENTO u otro).
Cédula hoja E84491751 cruza con V84491751 o 84491751 en sistema.
Solo cuotas del préstamo APROBADO (misma vista que el front); no mezcla LIQUIDADO u otros.

Columnas:
  - Email, Teléfono (cliente en sistema; cruce E/V por dígitos como la cédula).
  - Cuotas en mora al 1 jun (FECHA_PUNTO_1): solo estado MORA (no VENCIDO); conteo 0..N.
    Incluye cuotas en mora con abono parcial (columna D no aplica esa exclusión).
  - Cuotas en mora hoy: solo estado MORA; conteo real (1, 2, … 15, …) sin ocultar.
    No cuenta mora con abono parcial ≥ 0.10 USD (solo columna E).
  - Pagos parciales a mora (1 jun–hoy) ($): aplicaciones en cuota_pagos con fecha_pago
    en [1 jun, hoy] a cuotas que hoy siguen en MORA con saldo (abono parcial, no 100%).
  - Saldo total préstamo ($): suma pendiente de todas las cuotas (mora, vencido y por vencer).

APROBADO: mora al 1 jun = todas las MORA; mora hoy = sin abono parcial (conteo real; incluye 0).
"""
from __future__ import annotations

import csv
import io
import logging
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from sqlalchemy import Date, cast, func, select
from sqlalchemy.orm import Session

from app.constants.prestamo_estados import ESTADOS_PRESTAMO_DESISTIMIENTO_VARIANTES
from app.models.prestamo import Prestamo
from app.utils.cedula_almacenamiento import (
    expr_cedula_normalizada_para_comparar,
    texto_cedula_comparable_bd,
)

logger = logging.getLogger(__name__)

_SHEET_EXPORT = (
    "https://docs.google.com/spreadsheets/d/{sid}/export?format=csv"
)
_ESTADOS_EXCLUIDOS = ("DRAFT", "RECHAZADO")

# Dos puntos de análisis (cuotas en mora + dólares en cada fecha).
FECHA_PUNTO_1 = date(2026, 6, 1)
FECHA_INICIO_PAGOS_P2 = FECHA_PUNTO_1 + timedelta(days=1)  # 2 jun 2026
# Alias históricos
FECHA_CORTE_JUNIO = FECHA_PUNTO_1
FECHA_FIN_PAGOS_JUNIO = FECHA_PUNTO_1
MIN_CUOTAS_MORA_APROBADO = 4  # legacy; el Excel ya no oculta conteos < 4

# Item: (cuota_id|None, nro, fv, monto, pagado, fecha_pago_cuota, tot)
ItemCuota = Tuple[Any, Any, Any, Any, Any, Any, Any]
# Pago del préstamo: (fecha_pago, monto_pagado)
PagoVentana = Tuple[date, Decimal]
# Aplicación a cuota: (fecha_pago del pago, monto_aplicado, cuota_id, es_pago_completo)
AppCuotaVentana = Tuple[date, Decimal, int, bool]
_TOL_PARCIAL = Decimal("0.01")
# Umbral de abono parcial para excluir la cuota del conteo de mora del Excel.
MIN_ABONO_PARCIAL_EXCLUYE_MORA = Decimal("0.10")


def _sheet_id() -> str:
    from app.core.config import settings

    return (
        getattr(settings, "REPORTE_CUOTAS_JUN_AGO_SHEET_ID", None)
        or "1_Qean5MoSc1vWy6hMAAqOcMJeZzn9iUspJTqsOqZEqs"
    ).strip()


def _a_decimal(val: Any) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    try:
        d = Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return None
    return d


def cuota_unica_de_prestamos(
    prestamos: Sequence[Tuple[str, Any]],
) -> Optional[Decimal]:
    """
    Un solo monto si todos los préstamos elegibles coinciden.

    Prioriza APROBADO. Si no hay, LIQUIDADO/DESISTIMIENTO u otro estado con un solo monto.
    Distintos montos → None (no se elige ni se promedia).
    """
    if not prestamos:
        return None
    aprobados = [
        _a_decimal(m)
        for est, m in prestamos
        if str(est or "").strip().upper() == "APROBADO"
    ]
    aprobados = [m for m in aprobados if m is not None]
    if aprobados:
        montos = set(aprobados)
        if len(montos) != 1:
            return None
        return montos.pop()
    montos = {_a_decimal(m) for _e, m in prestamos}
    montos.discard(None)
    if len(montos) != 1:
        return None
    return montos.pop()


def estado_actual_de_prestamos(
    prestamos: Sequence[Tuple[str, Any]],
) -> Optional[str]:
    """
    Estado actual del préstamo. Si hay APROBADO, ese.
    Si no, un único estado; si hay varios reales, se listan separados por ' / '.
    """
    estados: List[str] = []
    for est, _m in prestamos:
        e = str(est or "").strip().upper()
        if e and e not in estados:
            estados.append(e)
    if not estados:
        return None
    if "APROBADO" in estados:
        return "APROBADO"
    if len(estados) == 1:
        return estados[0]
    orden = ("LIQUIDADO", "DESISTIMIENTO")
    ordenados = [e for e in orden if e in estados]
    ordenados.extend(e for e in estados if e not in orden)
    return " / ".join(ordenados)


def _digitos_cedula(ced: str) -> str:
    return "".join(ch for ch in (ced or "") if ch.isdigit())


def _claves_equivalentes_cedula(ced: str) -> List[str]:
    """E84491751, V84491751 y 84491751 se tratan como la misma persona."""
    full = texto_cedula_comparable_bd(ced)
    digits = _digitos_cedula(full)
    keys: List[str] = []
    for k in (full, digits, *(f"{p}{digits}" for p in "VEJG" if digits)):
        if k and k not in keys:
            keys.append(k)
    return keys


def _expandir_claves_sql(cedulas: Iterable[str]) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()
    for c in cedulas:
        for k in _claves_equivalentes_cedula(c):
            if k not in seen:
                seen.add(k)
                out.append(k)
    return out


def _prestamos_para_cedula_hoja(
    ced_hoja: str,
    prestamos_por_norm: Dict[str, List[Tuple[str, Any]]],
) -> List[Tuple[str, Any]]:
    k = texto_cedula_comparable_bd(ced_hoja)
    if k in prestamos_por_norm:
        return prestamos_por_norm[k]
    digits = _digitos_cedula(k)
    if not digits:
        return []
    merged: List[Tuple[str, Any]] = []
    seen: set[Tuple[str, str]] = set()
    for mk, prests in prestamos_por_norm.items():
        if _digitos_cedula(mk) != digits:
            continue
        for est, monto in prests:
            ident = (str(est), str(monto))
            if ident in seen:
                continue
            seen.add(ident)
            merged.append((est, monto))
    return merged


def _items_para_cedula_hoja(
    ced_hoja: str,
    items_por_norm: Dict[str, List[ItemCuota]],
) -> List[ItemCuota]:
    k = texto_cedula_comparable_bd(ced_hoja)
    if k in items_por_norm:
        return items_por_norm[k]
    digits = _digitos_cedula(k)
    if not digits:
        return []
    for mk, items in items_por_norm.items():
        if _digitos_cedula(mk) == digits:
            return items
    return []


def _apps_para_cedula_hoja(
    ced_hoja: str,
    apps_por_norm: Dict[str, List[PagoVentana]],
) -> List[PagoVentana]:
    k = texto_cedula_comparable_bd(ced_hoja)
    if k in apps_por_norm:
        return apps_por_norm[k]
    digits = _digitos_cedula(k)
    if not digits:
        return []
    for mk, apps in apps_por_norm.items():
        if _digitos_cedula(mk) == digits:
            return apps
    return []


def _apps_cuota_para_cedula_hoja(
    ced_hoja: str,
    apps_por_norm: Dict[str, List[AppCuotaVentana]],
) -> List[AppCuotaVentana]:
    k = texto_cedula_comparable_bd(ced_hoja)
    if k in apps_por_norm:
        return apps_por_norm[k]
    digits = _digitos_cedula(k)
    if not digits:
        return []
    for mk, apps in apps_por_norm.items():
        if _digitos_cedula(mk) == digits:
            return apps
    return []


def _contacto_para_cedula_hoja(
    ced_hoja: str,
    contacto_por_norm: Dict[str, Tuple[Optional[str], Optional[str]]],
) -> Tuple[Optional[str], Optional[str]]:
    k = texto_cedula_comparable_bd(ced_hoja)
    if k in contacto_por_norm:
        return contacto_por_norm[k]
    digits = _digitos_cedula(k)
    if not digits:
        return None, None
    for mk, contacto in contacto_por_norm.items():
        if _digitos_cedula(mk) == digits:
            return contacto
    return None, None


def _norm_item(item: Sequence[Any]) -> ItemCuota:
    """Acepta tupla legacy de 6 campos o ItemCuota de 7."""
    if len(item) >= 7:
        return (
            item[0],
            item[1],
            item[2],
            item[3],
            item[4],
            item[5],
            item[6],
        )
    nro, fv, monto, pagado, fp, tot = item[:6]
    return (None, nro, fv, monto, pagado, fp, tot)


def _as_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if hasattr(v, "date") and callable(getattr(v, "date")):
        try:
            d = v.date()
            return d if isinstance(d, date) else None
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()[:10]
        try:
            return date.fromisoformat(s)
        except ValueError:
            return None
    return None


def pendiente_vencido(
    monto: Any,
    total_pagado: Any,
    fecha_vencimiento: Optional[date],
    fecha_pago: Optional[date] = None,
    fecha_ref: Optional[date] = None,
) -> Optional[Decimal]:
    """Saldo pendiente solo si el estado real es VENCIDO o MORA. None si no aplica."""
    from app.services.cuota_estado import clasificar_estado_cuota, hoy_negocio

    m = _a_decimal(monto)
    if m is None or fecha_vencimiento is None:
        return None
    # Solo cuenta lo realmente pagado; fecha_pago sola no implica 100%.
    _ = fecha_pago
    pagado = _a_decimal(total_pagado) or Decimal("0")
    ref = fecha_ref or hoy_negocio()
    est = clasificar_estado_cuota(float(pagado), float(m), fecha_vencimiento, ref)
    if est not in ("VENCIDO", "MORA"):
        return None
    saldo = (m - pagado).quantize(Decimal("0.01"))
    if saldo <= Decimal("0.00"):
        return None
    return saldo


def _estado_item_al(
    fv: Any,
    monto: Any,
    pagado: Any,
    fp: Any,
    as_of: date,
) -> Optional[str]:
    from app.services.cuota_estado import clasificar_estado_cuota

    fv_d = _as_date(fv)
    if fv_d is None:
        return None
    m = _a_decimal(monto)
    if m is None:
        return None
    fp_d = _as_date(fp)
    # Pagos con fecha posterior al corte no existen aún a as_of.
    if fp_d is not None and fp_d > as_of:
        pag_eff = Decimal("0")
    else:
        # Usar total_pagado real. No inventar 100% solo porque hay fecha_pago.
        pag_eff = _a_decimal(pagado) or Decimal("0")
    return clasificar_estado_cuota(float(pag_eff), float(m), fv_d, as_of)


def _saldo_mora_item_al(
    fv: Any,
    monto: Any,
    pagado: Any,
    fp: Any,
    as_of: date,
) -> Optional[Decimal]:
    """Pendiente de la cuota solo si a as_of está en MORA."""
    if _estado_item_al(fv, monto, pagado, fp, as_of) != "MORA":
        return None
    fv_d = _as_date(fv)
    m = _a_decimal(monto)
    if fv_d is None or m is None:
        return None
    fp_d = _as_date(fp)
    if fp_d is not None and fp_d > as_of:
        pag_eff, fp_eff = Decimal("0"), None
    else:
        pag_eff, fp_eff = pagado, fp_d
    return pendiente_vencido(m, pag_eff, fv_d, fp_eff, as_of)


def _pagado_efectivo_al(pagado: Any, fp: Any, monto: Decimal, as_of: date) -> Decimal:
    """Pagado de la cuota a as_of (ignora pagos posteriores al corte)."""
    _ = monto
    fp_d = _as_date(fp)
    if fp_d is not None and fp_d > as_of:
        return Decimal("0")
    return _a_decimal(pagado) or Decimal("0")


def _tiene_abono_parcial_al(
    monto: Any,
    pagado: Any,
    fp: Any,
    as_of: date,
    *,
    minimo: Decimal = MIN_ABONO_PARCIAL_EXCLUYE_MORA,
) -> bool:
    """
    True si a as_of hay abono ≥ minimo y la cuota no está cubierta al 100%.
    (0.10 o más = abono parcial para el conteo de mora del informe.)
    """
    m = _a_decimal(monto)
    if m is None or m <= Decimal("0.00"):
        return False
    pag_eff = _pagado_efectivo_al(pagado, fp, m, as_of)
    if pag_eff < minimo:
        return False
    if (m - pag_eff) <= _TOL_PARCIAL:
        return False
    return True


def conteo_cuotas_en_mora(
    items: Sequence[Sequence[Any]],
    as_of: date,
    *,
    excluir_abono_parcial: bool = False,
) -> int:
    """
    Cantidad de cuotas con estado MORA a la fecha (1, 2, … N). No incluye VENCIDO.
    Si excluir_abono_parcial (columna E / mora hoy): no cuenta mora con abono ≥ 0.10.
    Columna D (1 jun) llama con excluir_abono_parcial=False.
    """
    n = 0
    for item in items:
        _cid, _nro, fv, monto, pagado, fp, _tot = _norm_item(item)
        if _estado_item_al(fv, monto, pagado, fp, as_of) != "MORA":
            continue
        if excluir_abono_parcial and _tiene_abono_parcial_al(monto, pagado, fp, as_of):
            continue
        n += 1
    return n


def saldo_total_prestamo(
    items: Sequence[Sequence[Any]],
    as_of: date,
) -> Decimal:
    """
    Cuánto debe en total del préstamo: suma (monto − pagado) de todas las cuotas
    con saldo > 0 a as_of (incluye mora, vencido y cuotas por vencer).
    """
    total = Decimal("0")
    for item in items:
        _cid, _nro, fv, monto, pagado, fp, _tot = _norm_item(item)
        m = _a_decimal(monto)
        if m is None or m <= Decimal("0.00"):
            continue
        pag_eff = _pagado_efectivo_al(pagado, fp, m, as_of)
        sal = (m - pag_eff).quantize(Decimal("0.01"))
        if sal > Decimal("0.00"):
            total += sal
    return total.quantize(Decimal("0.01"))


def saldo_vencido_solo_mora(
    items: Sequence[Sequence[Any]],
    as_of: date,
) -> Decimal:
    """Saldo vencido = suma pendiente solo de cuotas en MORA."""
    total = Decimal("0")
    for item in items:
        _cid, _nro, fv, monto, pagado, fp, _tot = _norm_item(item)
        sal = _saldo_mora_item_al(fv, monto, pagado, fp, as_of)
        if sal is not None:
            total += sal
    return total.quantize(Decimal("0.01"))


def _cuotas_mora_parcial_ids(
    items: Sequence[Sequence[Any]],
    as_of: date,
) -> set:
    """
    IDs de cuotas en MORA a as_of con abono parcial (0 < pagado < monto).
    Sin id (tuplas legacy) no entran; el fallback sin cuota_pagos usa otra ruta.
    """
    ids: set = set()
    for item in items:
        cid, _nro, fv, monto, pagado, fp, _tot = _norm_item(item)
        if cid is None:
            continue
        if _estado_item_al(fv, monto, pagado, fp, as_of) != "MORA":
            continue
        m = _a_decimal(monto)
        if m is None or m <= Decimal("0.00"):
            continue
        pag_eff = _pagado_efectivo_al(pagado, fp, m, as_of)
        if pag_eff <= _TOL_PARCIAL:
            continue
        if (m - pag_eff) <= _TOL_PARCIAL:
            continue
        try:
            ids.add(int(cid))
        except (TypeError, ValueError):
            continue
    return ids


def pagos_parciales_a_cuotas_en_mora(
    items: Sequence[Sequence[Any]],
    aplicaciones: Sequence[AppCuotaVentana],
    *,
    fecha_desde: date,
    fecha_hasta: date,
    as_of: date,
) -> Decimal:
    """
    Suma monto_aplicado de aplicaciones cuyo pago tiene fecha en [desde, hasta]
    hacia cuotas que a as_of están en MORA con saldo (parcial).

    Solo cuenta filas con es_pago_completo=False (el abono no cerró la cuota).
    """
    mora_parcial = _cuotas_mora_parcial_ids(items, as_of)
    if not mora_parcial:
        return Decimal("0.00")
    total = Decimal("0")
    for row in aplicaciones:
        if len(row) < 4:
            continue
        fpago, monto_ap, cuota_id, es_completo = row[0], row[1], row[2], row[3]
        if es_completo:
            continue
        try:
            cid = int(cuota_id)
        except (TypeError, ValueError):
            continue
        if cid not in mora_parcial:
            continue
        fd = _as_date(fpago)
        if fd is None or fd < fecha_desde or fd > fecha_hasta:
            continue
        val = _a_decimal(monto_ap)
        if val is None or val <= Decimal("0.00"):
            continue
        total += val
    return total.quantize(Decimal("0.01"))


def hay_pagos_parciales_a_cuotas_en_mora(
    items: Sequence[Sequence[Any]],
    aplicaciones: Sequence[AppCuotaVentana],
    *,
    fecha_desde: date,
    fecha_hasta: date,
    as_of: date,
) -> bool:
    return (
        pagos_parciales_a_cuotas_en_mora(
            items,
            aplicaciones,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            as_of=as_of,
        )
        > Decimal("0.00")
    )


# Alias: nombre histórico apuntaba a todo el crédito; ahora solo MORA.
saldo_vencido_credito = saldo_vencido_solo_mora
saldo_vencido_en_mora = saldo_vencido_solo_mora


def pagos_en_ventana(
    pagos: Sequence[PagoVentana],
    *,
    fecha_desde: Optional[date],
    fecha_hasta: date,
) -> Decimal:
    """Suma pagos con fecha_pago (día) en [desde, hasta] inclusive."""
    total = Decimal("0")
    for fpago, monto in pagos:
        fd = _as_date(fpago)
        if fd is None:
            continue
        if fecha_desde is not None and fd < fecha_desde:
            continue
        if fd > fecha_hasta:
            continue
        val = _a_decimal(monto)
        if val is None or val <= Decimal("0.00"):
            continue
        total += val
    return total.quantize(Decimal("0.01"))


def sumar_pagos_ventanas_exclusivas(
    pagos: Sequence[PagoVentana],
    *,
    corte_junio: date = FECHA_PUNTO_1,
    fecha_hoy: date,
) -> Tuple[Decimal, Decimal]:
    """
    Parte cada pago en una sola ventana (sin solape):
    - Punto 1: fecha_pago ≤ 1 jun (corte_junio)
    - Punto 2: fecha_pago desde 2 jun hasta hoy
    """
    fin_p1 = corte_junio
    inicio_p2 = corte_junio + timedelta(days=1)
    p1 = Decimal("0")
    p2 = Decimal("0")
    for fpago, monto in pagos:
        fd = _as_date(fpago)
        if fd is None:
            continue
        val = _a_decimal(monto)
        if val is None or val <= Decimal("0.00"):
            continue
        if fd <= fin_p1:
            p1 += val
        elif inicio_p2 <= fd <= fecha_hoy:
            p2 += val
    return p1.quantize(Decimal("0.01")), p2.quantize(Decimal("0.01"))


def saldo_a_pagar(
    saldo_mora: Optional[Decimal],
    pagos_hasta_mayo: Decimal,
    pagos_desde_junio: Decimal,
) -> Optional[Decimal]:
    """Saldo a pagar = dinero en mora − todos los pagos de ambas ventanas."""
    return saldo_neto_mora_menos_pagos(
        saldo_mora,
        (pagos_hasta_mayo or Decimal("0")) + (pagos_desde_junio or Decimal("0")),
    )


def saldo_neto_mora_menos_pagos(
    saldo_mora: Optional[Decimal],
    pagos_hasta_corte: Decimal,
) -> Optional[Decimal]:
    """
    Deuda mora menos pagos hasta el corte.
    Si el neto es negativo (pagó de más), se muestra 0.
    """
    mora = saldo_mora if saldo_mora is not None else Decimal("0")
    pagos = pagos_hasta_corte or Decimal("0")
    if mora <= Decimal("0.00") and pagos <= Decimal("0.00"):
        return None
    neto = (mora - pagos).quantize(Decimal("0.01"))
    if neto < Decimal("0.00"):
        return Decimal("0.00")
    return neto


# Alias de compatibilidad con tests / nombre anterior.
def pagos_aplicados_a_vencido_o_mora(
    items: Sequence[Sequence[Any]],
    aplicaciones: Sequence[Any],
    *,
    as_of: date,
    fecha_desde: Optional[date],
    fecha_hasta: date,
) -> Decimal:
    _ = items, as_of
    normalizados: List[PagoVentana] = []
    for row in aplicaciones:
        if len(row) >= 2:
            fp, monto = row[0], row[1]
            fp_d = _as_date(fp)
            val = _a_decimal(monto)
            if fp_d is not None and val is not None:
                normalizados.append((fp_d, val))
    return pagos_en_ventana(
        normalizados, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta
    )


def metricas_corte_mora(
    items: Sequence[Sequence[Any]],
    as_of: date,
    *,
    es_aprobado: bool,
    pagos: Optional[Sequence[PagoVentana]] = None,
    pagos_desde: Optional[date] = None,
    pagos_hasta: Optional[date] = None,
    # compat: nombre antiguo
    aplicaciones: Optional[Sequence[Any]] = None,
    excluir_abono_parcial: bool = False,
) -> Tuple[Optional[int], Optional[Decimal], Optional[Decimal]]:
    """
    (cuotas_en_mora, saldo_solo_mora, pagos_en_ventana).
    Cuotas en mora: conteo real 0..N (1, 2, … 15, …); no se oculta por umbral.
    excluir_abono_parcial: solo mora hoy (columna E).
    Saldo y pagos: siempre si > 0.
    """
    _ = es_aprobado  # ya no se oculta el conteo por umbral 4+
    n = conteo_cuotas_en_mora(
        items, as_of, excluir_abono_parcial=excluir_abono_parcial
    )
    # Siempre el número (incluido 0) para que el Excel muestre cuántas hay.
    n_out: Optional[int] = int(n)
    sal = saldo_vencido_solo_mora(items, as_of)
    sal_out: Optional[Decimal] = sal if sal > Decimal("0.00") else None
    pag_out: Optional[Decimal] = None
    fuente = pagos if pagos is not None else aplicaciones
    if fuente is not None and pagos_hasta is not None:
        if pagos is not None:
            pag = pagos_en_ventana(
                pagos, fecha_desde=pagos_desde, fecha_hasta=pagos_hasta
            )
        else:
            pag = pagos_aplicados_a_vencido_o_mora(
                items,
                aplicaciones or (),
                as_of=as_of,
                fecha_desde=pagos_desde,
                fecha_hasta=pagos_hasta,
            )
        pag_out = pag if pag > Decimal("0.00") else None
    return n_out, sal_out, pag_out


def parsear_cedulas_csv(raw: bytes) -> List[str]:
    if (
        not raw
        or raw.lstrip()[:15].lower().startswith(b"<!doctype")
        or b"<html" in raw[:400].lower()
    ):
        raise RuntimeError(
            "La hoja Drive no devolvio CSV (inicie sesion o publique la hoja)."
        )
    text = raw.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    out: List[str] = []
    for i, row in enumerate(rows):
        if not row:
            continue
        cell = str(row[0] or "").strip()
        if not cell:
            continue
        if i == 0 and not any(ch.isdigit() for ch in cell):
            continue
        out.append(cell)
    return out


def leer_cedulas_hoja(spreadsheet_id: Optional[str] = None) -> List[str]:
    sid = (spreadsheet_id or _sheet_id()).strip()
    if not sid:
        raise RuntimeError("REPORTE_CUOTAS_JUN_AGO_SHEET_ID no configurado.")
    url = _SHEET_EXPORT.format(sid=sid)
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "pagos-reportes-cedulas-cuota/1.0"},
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read()
    except urllib.error.URLError as e:
        raise RuntimeError(f"No se pudo leer la hoja Drive: {e}") from e
    return parsear_cedulas_csv(raw)


def _cuotas_bd_por_cedula_norm(
    db: Session, cedulas_norm: Iterable[str]
) -> Dict[str, List[Tuple[str, Any]]]:
    claves = _expandir_claves_sql(cedulas_norm)
    if not claves:
        return {}
    ced_expr = expr_cedula_normalizada_para_comparar(Prestamo.cedula)
    estado_u = func.upper(func.trim(Prestamo.estado))
    rows = db.execute(
        select(ced_expr, Prestamo.estado, Prestamo.cuota_periodo).where(
            ced_expr.in_(claves),
            estado_u.notin_(_ESTADOS_EXCLUIDOS),
        )
    ).all()
    out: Dict[str, List[Tuple[str, Any]]] = {}
    for ced, est, cuota in rows:
        k = texto_cedula_comparable_bd(ced or "")
        if not k:
            continue
        out.setdefault(k, []).append((str(est or ""), cuota))
    return out


def _contacto_bd_por_cedula_norm(
    db: Session, cedulas_norm: Iterable[str]
) -> Dict[str, Tuple[Optional[str], Optional[str]]]:
    """ced_norm -> (email, telefono) desde tabla clientes."""
    from app.models.cliente import Cliente

    claves = _expandir_claves_sql(cedulas_norm)
    if not claves:
        return {}
    ced_expr = expr_cedula_normalizada_para_comparar(Cliente.cedula)
    rows = db.execute(
        select(ced_expr, Cliente.email, Cliente.telefono).where(ced_expr.in_(claves))
    ).all()
    out: Dict[str, Tuple[Optional[str], Optional[str]]] = {}
    for ced, email, telefono in rows:
        k = texto_cedula_comparable_bd(ced or "")
        if not k:
            continue
        em = str(email or "").strip() or None
        tel = str(telefono or "").strip() or None
        out[k] = (em, tel)
    return out


def _items_cuotas_para_informe(
    k: str,
    items_all: Dict[str, List[ItemCuota]],
    items_aprobado: Dict[str, List[ItemCuota]],
    estados_por_k: Dict[str, set],
) -> List[ItemCuota]:
    """APROBADO: solo cuotas del préstamo activo; desistimiento: sin mora en informe."""
    estados = estados_por_k.get(k, set())
    if "APROBADO" in estados:
        return items_aprobado.get(k, [])
    if estados & ESTADOS_PRESTAMO_DESISTIMIENTO_VARIANTES:
        return []
    return items_all.get(k, [])


def _cargar_items_cuotas_por_cedula(
    db: Session,
    cedulas_norm: Iterable[str],
) -> Tuple[Dict[str, List[ItemCuota]], Dict[str, set]]:
    """ced_norm -> cuotas del préstamo relevante; también estados por cédula."""
    from app.models.cuota import Cuota

    claves = _expandir_claves_sql(cedulas_norm)
    if not claves:
        return {}, {}
    ced_expr = expr_cedula_normalizada_para_comparar(Prestamo.cedula)
    estado_u = func.upper(func.trim(Prestamo.estado))
    rows = db.execute(
        select(
            ced_expr,
            Cuota.id,
            Cuota.fecha_vencimiento,
            Cuota.monto,
            Cuota.total_pagado,
            Cuota.fecha_pago,
            Cuota.numero_cuota,
            Prestamo.numero_cuotas,
            Prestamo.estado,
        )
        .select_from(Cuota)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .where(
            ced_expr.in_(claves),
            estado_u.notin_(_ESTADOS_EXCLUIDOS),
        )
    ).all()
    items_all: Dict[str, List[ItemCuota]] = {}
    items_aprobado: Dict[str, List[ItemCuota]] = {}
    estados_por_k: Dict[str, set] = {}
    for ced, cid, fv, monto, pagado, f_pago, nro, tot, est in rows:
        k = texto_cedula_comparable_bd(ced or "")
        if not k or fv is None:
            continue
        fv_d = fv.date() if hasattr(fv, "date") else fv
        item: ItemCuota = (cid, nro, fv_d, monto, pagado, f_pago, tot)
        items_all.setdefault(k, []).append(item)
        e = str(est or "").strip().upper()
        if e == "APROBADO":
            items_aprobado.setdefault(k, []).append(item)
        if e:
            estados_por_k.setdefault(k, set()).add(e)
    out: Dict[str, List[ItemCuota]] = {}
    for k in items_all:
        out[k] = _items_cuotas_para_informe(k, items_all, items_aprobado, estados_por_k)
    return out, estados_por_k


def _cargar_pagos_por_cedula(
    db: Session,
    cedulas_norm: Iterable[str],
    *,
    fecha_hasta: date,
) -> Dict[str, List[PagoVentana]]:
    """ced_norm -> pagos del préstamo relevante (fecha_pago <= hasta)."""
    from app.models.pago import Pago

    claves = _expandir_claves_sql(cedulas_norm)
    if not claves:
        return {}
    ced_expr = expr_cedula_normalizada_para_comparar(Prestamo.cedula)
    estado_u = func.upper(func.trim(Prestamo.estado))
    estado_pago = func.upper(func.trim(func.coalesce(Pago.estado, "")))
    rows = db.execute(
        select(
            ced_expr,
            Prestamo.estado,
            Pago.id,
            cast(Pago.fecha_pago, Date).label("dia_pago"),
            Pago.monto_pagado,
        )
        .select_from(Pago)
        .join(Prestamo, Pago.prestamo_id == Prestamo.id)
        .where(
            ced_expr.in_(claves),
            estado_u.notin_(_ESTADOS_EXCLUIDOS),
            Pago.fecha_pago.isnot(None),
            cast(Pago.fecha_pago, Date) <= fecha_hasta,
            ~estado_pago.like("ANULADO%"),
            estado_pago.is_distinct_from("DUPLICADO"),
            Pago.monto_pagado.isnot(None),
        )
    ).all()
    all_pagos: Dict[str, List[PagoVentana]] = {}
    aprob_pagos: Dict[str, List[PagoVentana]] = {}
    estados_por_k: Dict[str, set] = {}
    vistos: Dict[str, set] = {}
    for ced, est, pid, fp, monto in rows:
        k = texto_cedula_comparable_bd(ced or "")
        if not k or fp is None or pid is None:
            continue
        val = _a_decimal(monto)
        if val is None or val <= Decimal("0.00"):
            continue
        vistos.setdefault(k, set())
        if int(pid) in vistos[k]:
            continue
        vistos[k].add(int(pid))
        fp_d = _as_date(fp)
        if fp_d is None:
            continue
        pago: PagoVentana = (fp_d, val.quantize(Decimal("0.01")))
        all_pagos.setdefault(k, []).append(pago)
        e = str(est or "").strip().upper()
        if e == "APROBADO":
            aprob_pagos.setdefault(k, []).append(pago)
        if e:
            estados_por_k.setdefault(k, set()).add(e)
    out: Dict[str, List[PagoVentana]] = {}
    for k, pagos in all_pagos.items():
        if "APROBADO" in estados_por_k.get(k, set()):
            out[k] = aprob_pagos.get(k, [])
        else:
            out[k] = pagos
    return out


# Alias de compatibilidad.
_cargar_aplicaciones_por_cedula = _cargar_pagos_por_cedula


def _cargar_apps_parciales_cuota_por_cedula(
    db: Session,
    cedulas_norm: Iterable[str],
    *,
    fecha_desde: date,
    fecha_hasta: date,
) -> Dict[str, List[AppCuotaVentana]]:
    """
    ced_norm -> aplicaciones cuota_pagos con fecha_pago del pago en [desde, hasta].
    Solo préstamos relevantes (APROBADO si existe).
    """
    from app.models.cuota import Cuota
    from app.models.cuota_pago import CuotaPago
    from app.models.pago import Pago

    claves = _expandir_claves_sql(cedulas_norm)
    if not claves:
        return {}
    ced_expr = expr_cedula_normalizada_para_comparar(Prestamo.cedula)
    estado_u = func.upper(func.trim(Prestamo.estado))
    estado_pago = func.upper(func.trim(func.coalesce(Pago.estado, "")))
    rows = db.execute(
        select(
            ced_expr,
            Prestamo.estado,
            CuotaPago.id,
            cast(Pago.fecha_pago, Date).label("dia_pago"),
            CuotaPago.monto_aplicado,
            CuotaPago.cuota_id,
            CuotaPago.es_pago_completo,
        )
        .select_from(CuotaPago)
        .join(Pago, Pago.id == CuotaPago.pago_id)
        .join(Cuota, Cuota.id == CuotaPago.cuota_id)
        .join(Prestamo, Prestamo.id == Cuota.prestamo_id)
        .where(
            ced_expr.in_(claves),
            estado_u.notin_(_ESTADOS_EXCLUIDOS),
            Pago.fecha_pago.isnot(None),
            cast(Pago.fecha_pago, Date) >= fecha_desde,
            cast(Pago.fecha_pago, Date) <= fecha_hasta,
            ~estado_pago.like("ANULADO%"),
            estado_pago.is_distinct_from("DUPLICADO"),
            CuotaPago.monto_aplicado.isnot(None),
        )
    ).all()
    all_apps: Dict[str, List[AppCuotaVentana]] = {}
    aprob_apps: Dict[str, List[AppCuotaVentana]] = {}
    estados_por_k: Dict[str, set] = {}
    vistos: Dict[str, set] = {}
    for ced, est, cp_id, fp, monto, cuota_id, es_completo in rows:
        k = texto_cedula_comparable_bd(ced or "")
        if not k or fp is None or cp_id is None or cuota_id is None:
            continue
        val = _a_decimal(monto)
        if val is None or val <= Decimal("0.00"):
            continue
        vistos.setdefault(k, set())
        if int(cp_id) in vistos[k]:
            continue
        vistos[k].add(int(cp_id))
        fp_d = _as_date(fp)
        if fp_d is None:
            continue
        app: AppCuotaVentana = (
            fp_d,
            val.quantize(Decimal("0.01")),
            int(cuota_id),
            bool(es_completo),
        )
        all_apps.setdefault(k, []).append(app)
        e = str(est or "").strip().upper()
        if e == "APROBADO":
            aprob_apps.setdefault(k, []).append(app)
        if e:
            estados_por_k.setdefault(k, set()).add(e)
    out: Dict[str, List[AppCuotaVentana]] = {}
    for k, apps in all_apps.items():
        if "APROBADO" in estados_por_k.get(k, set()):
            out[k] = aprob_apps.get(k, [])
        else:
            out[k] = apps
    return out


def filas_cedula_cuota(
    cedulas_hoja: Sequence[str],
    prestamos_por_norm: Dict[str, List[Tuple[str, Any]]],
    items_por_norm: Optional[Dict[str, List[Any]]] = None,
    apps_por_norm: Optional[Dict[str, List[PagoVentana]]] = None,
    apps_cuota_por_norm: Optional[Dict[str, List[AppCuotaVentana]]] = None,
    contacto_por_norm: Optional[Dict[str, Tuple[Optional[str], Optional[str]]]] = None,
    *,
    fecha_junio: date = FECHA_PUNTO_1,
    fecha_hoy: Optional[date] = None,
) -> List[Dict[str, Any]]:
    from app.services.cuota_estado import hoy_negocio

    items_por_norm = items_por_norm or {}
    apps_cuota_por_norm = apps_cuota_por_norm or {}
    contacto_por_norm = contacto_por_norm or {}
    _ = apps_por_norm  # compat firma
    hoy = fecha_hoy or hoy_negocio()
    punto_1 = fecha_junio
    filas: List[Dict[str, Any]] = []
    for raw in cedulas_hoja:
        prests = _prestamos_para_cedula_hoja(raw, prestamos_por_norm)
        estado = estado_actual_de_prestamos(prests)
        es_aprobado = estado == "APROBADO"
        cuota = cuota_unica_de_prestamos(prests)
        items = _items_para_cedula_hoja(raw, items_por_norm)
        apps_cuota = _apps_cuota_para_cedula_hoja(raw, apps_cuota_por_norm)
        n_p1, _, _ = metricas_corte_mora(
            items, punto_1, es_aprobado=es_aprobado, excluir_abono_parcial=False
        )
        n_hoy, _, _ = metricas_corte_mora(
            items, hoy, es_aprobado=es_aprobado, excluir_abono_parcial=True
        )
        sal_tot = saldo_total_prestamo(items, hoy)
        sal_out = float(sal_tot) if sal_tot > Decimal("0.00") else None
        pag_parcial = pagos_parciales_a_cuotas_en_mora(
            items,
            apps_cuota,
            fecha_desde=punto_1,
            fecha_hasta=hoy,
            as_of=hoy,
        )
        pag_parcial_out = (
            float(pag_parcial) if pag_parcial > Decimal("0.00") else None
        )
        hay_parcial = pag_parcial > Decimal("0.00")
        email, telefono = _contacto_para_cedula_hoja(raw, contacto_por_norm)
        filas.append(
            {
                "cedula": raw,
                "email": email,
                "telefono": telefono,
                "estado": estado,
                "cuota": float(cuota) if cuota is not None else None,
                "fecha_punto_1": punto_1,
                "fecha_junio": punto_1,
                "fecha_hoy": hoy,
                "mora_junio": n_p1,
                "mora_hoy": n_hoy,
                "pagos_parciales_mora": pag_parcial_out,
                "hay_pagos_parciales_mora": hay_parcial,
                "saldo_total_prestamo": sal_out,
                # Alias histórico del Excel (última columna).
                "saldo_a_pagar": sal_out,
            }
        )
    return filas


def generar_excel_cedulas_cuota(filas: Sequence[Dict[str, Any]]) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cedula cuota"
    bold = Font(bold=True)

    fecha_p1 = next(
        (
            f.get("fecha_punto_1") or f.get("fecha_junio")
            for f in filas
            if f.get("fecha_punto_1") or f.get("fecha_junio")
        ),
        FECHA_PUNTO_1,
    )
    fecha_hoy = next((f.get("fecha_hoy") for f in filas if f.get("fecha_hoy")), None)
    year = fecha_p1.year if isinstance(fecha_p1, date) else 2026
    label_m1 = f"Cuotas en mora al 1 jun {year}"
    label_m2 = (
        f"Cuotas en mora hoy ({fecha_hoy.isoformat()}, sin abono parcial ≥0.10)"
        if isinstance(fecha_hoy, date)
        else "Cuotas en mora hoy (sin abono parcial ≥0.10)"
    )
    label_parcial = (
        f"Pagos parciales a mora 1 jun–hoy ({fecha_hoy.isoformat()})"
        if isinstance(fecha_hoy, date)
        else "Pagos parciales a mora 1 jun–hoy"
    )

    headers = [
        "Cédula",
        "Email",
        "Teléfono",
        "Estado",
        "Cuota",
        label_m1,
        label_m2,
        label_parcial,
        "Saldo total préstamo ($)",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for f in filas:
        sal = f.get("saldo_total_prestamo")
        if sal is None:
            sal = f.get("saldo_a_pagar")
        hay = f.get("hay_pagos_parciales_mora")
        monto_p = f.get("pagos_parciales_mora")
        if hay is None and monto_p is not None:
            hay = float(monto_p or 0) > 0
        if hay and monto_p is not None:
            celda_parcial = f"Sí (${float(monto_p):,.2f})"
        elif hay:
            celda_parcial = "Sí"
        else:
            celda_parcial = "No"
        row = [
            f.get("cedula") or "",
            f.get("email") or None,
            f.get("telefono") or None,
            f.get("estado") or None,
            f.get("cuota") if f.get("cuota") is not None else None,
            f.get("mora_junio") if f.get("mora_junio") is not None else None,
            f.get("mora_hoy") if f.get("mora_hoy") is not None else None,
            celda_parcial,
            sal if sal is not None else None,
        ]
        ws.append(row)
        r = ws.max_row
        for col in (5, 9):
            cell = ws.cell(r, col)
            if cell.value is not None:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            cell.alignment = Alignment(horizontal="right")
        for col in (6, 7, 8):
            ws.cell(r, col).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 32
    ws.column_dimensions["I"].width = 24
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    _ = get_column_letter
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_excel_cedulas_cuota_hoja(
    db: Session,
    *,
    spreadsheet_id: Optional[str] = None,
    cedulas: Optional[Sequence[str]] = None,
) -> Tuple[bytes, int, int]:
    """
    Devuelve (xlsx, filas, filas_con_cuota).
    `cedulas` solo para tests; en runtime se leen de la hoja.
    """
    from app.services.cuota_estado import hoy_negocio

    lista = list(cedulas) if cedulas is not None else leer_cedulas_hoja(spreadsheet_id)
    if not lista:
        raise RuntimeError("La hoja no tiene cédulas.")
    normas = [texto_cedula_comparable_bd(c) for c in lista]
    mapa = _cuotas_bd_por_cedula_norm(db, normas)
    contacto = _contacto_bd_por_cedula_norm(db, normas)
    items, _estados = _cargar_items_cuotas_por_cedula(db, normas)
    hoy = hoy_negocio()
    apps_cuota = _cargar_apps_parciales_cuota_por_cedula(
        db,
        normas,
        fecha_desde=FECHA_PUNTO_1,
        fecha_hasta=hoy,
    )
    filas = filas_cedula_cuota(
        lista,
        mapa,
        items,
        apps_cuota_por_norm=apps_cuota,
        contacto_por_norm=contacto,
        fecha_hoy=hoy,
    )
    con_cuota = sum(1 for f in filas if f.get("cuota") is not None)
    logger.info(
        "[reporte_cedulas_cuota_hoja] filas=%s con_cuota=%s sin_cuota=%s",
        len(filas),
        con_cuota,
        len(filas) - con_cuota,
    )
    return generar_excel_cedulas_cuota(filas), len(filas), con_cuota
