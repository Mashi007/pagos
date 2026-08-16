"""Analisis de cobranzas alineado a segmentos del dashboard/menu."""

from __future__ import annotations

import copy
import io
import logging
import threading
import time as time_mod
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
from typing import Any, Optional, Sequence

from fastapi import HTTPException, UploadFile
from sqlalchemy import Date as SADate
from sqlalchemy import cast, func
from sqlalchemy.orm import Session

from app.constants.prestamo_estados import ESTADOS_PRESTAMO_EXCLUIDOS_COBRANZA_NOTIF
from app.models.cobranza_universo import CobranzaUniversoCedula, CobranzaUniversoDesempenoDiario
from app.models.cuota import Cuota
from app.models.cuota_pago import CuotaPago
from app.models.pago import Pago
from app.models.prestamo import Prestamo
from app.services.cuota_estado import (
    TZ_NEGOCIO,
    dias_retraso_desde_vencimiento,
    hoy_negocio,
)
from app.services.notificaciones_exclusion_desistimiento import sql_cliente_sin_desistimiento
from app.services.desempeno_1_cuota_stock import (
    SEG_MAX_N_EXACTO,
    SEG_MIN_DIAS_CUOTA_1,
    _cumple_ventana_6plus,
    _cumple_ventana_segmento,
    _cuotas_atrasadas_para_segmento,
    _load_cuotas_meta,
    _stock_1_cuota_cobranzas_at,
    _stock_2_cuotas_at,
    _stock_3_cuotas_at,
    _stock_4_cuotas_at,
    _stock_5_cuotas_at,
    _stock_6plus_cuotas_at,
    _stock_exact_n_cuotas_at,
    _t_fin_dia,
)
from app.utils.cedula_almacenamiento import (
    expr_cedula_normalizada_para_comparar,
    normalizar_cedula_almacenamiento,
    texto_cedula_comparable_bd,
)

logger = logging.getLogger(__name__)

_ANALISIS_CACHE_TTL_SEC = 180.0  # 3 min: repeat GET without re-scanning cartera
_analisis_cache: dict[str, tuple[float, dict[str, Any]]] = {}
_analisis_cache_lock = threading.Lock()

_HEADER_CELLS = frozenset({"cedula", "cedulas", "documento", "id"})
# Gráficos diarios (serie): 1..5 exactas + 6+ (>=6 cuotas).
_BUCKET_KEYS = ("1", "2", "3", "4", "5", "6plus")
# Tabla + detalle: solo 1..15 exactas por conteo de cuotas atrasadas.
_TABLA_BUCKET_KEYS = tuple(str(i) for i in range(1, SEG_MAX_N_EXACTO + 1))
_RESTO_6PLUS_KEY = "resto6plus"  # compat tests / helpers (≥16); no va a la tabla
# Distribución de atraso (mismo tramo que el dashboard): 20 bins de 30 días + >600.
_ATRASO_BIN_DIAS = 30
_ATRASO_N_BINS = 20
_ATRASO_MAX_DIAS = _ATRASO_N_BINS * _ATRASO_BIN_DIAS


def expr_prestamo_activo_cobranzas():
    """Misma cartera que dashboard/notificaciones: no LIQUIDADO ni DESISTIMIENTO."""
    est = func.upper(func.trim(func.coalesce(Prestamo.estado, "")))
    excl = tuple(str(e).strip().upper() for e in ESTADOS_PRESTAMO_EXCLUIDOS_COBRANZA_NOTIF)
    return est.notin_(excl)


def prestamo_es_activo_cobranzas(estado: Optional[str]) -> bool:
    return (estado or "").strip().upper() not in {
        str(e).strip().upper() for e in ESTADOS_PRESTAMO_EXCLUIDOS_COBRANZA_NOTIF
    }


def _texto_celda_a(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return str(raw).strip()
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw == int(raw):
            return str(int(raw))
        return str(raw).strip()
    return str(raw).strip()


def _es_celda_header(text: str) -> bool:
    t = (text or "").strip().lower()
    if not t:
        return False
    # quitar acentos basicos en encabezados tipicos
    t = (
        t.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    return t in _HEADER_CELLS


def parse_cedulas_desde_excel(content: bytes) -> list[str]:
    """Lee cedulas de la columna A (openpyxl). Omite encabezados y duplicados."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail="Dependencia openpyxl no disponible en el servidor.",
        ) from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        out: list[str] = []
        seen: set[str] = set()
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            raw = row[0] if row else None
            text = _texto_celda_a(raw)
            if not text or _es_celda_header(text):
                continue
            store = normalizar_cedula_almacenamiento(text)
            if not store:
                continue
            key = texto_cedula_comparable_bd(store)
            if not key or key in seen:
                continue
            seen.add(key)
            out.append(store)
        return out
    finally:
        wb.close()


def reemplazar_universo(
    db: Session,
    cedulas: Sequence[str],
    usuario_id: Optional[int] = None,
) -> int:
    """Reemplazo total (interno). Prefiera fusionar_universo para carga Excel."""
    db.query(CobranzaUniversoDesempenoDiario).delete()
    db.query(CobranzaUniversoCedula).delete()
    n = 0
    for c in cedulas:
        store = normalizar_cedula_almacenamiento(c)
        if not store:
            continue
        db.add(CobranzaUniversoCedula(cedula=store, usuario_id=usuario_id))
        n += 1
    db.commit()
    return n


def listar_cedulas_universo(db: Session) -> list[str]:
    """Cedulas del universo, ordenadas."""
    rows = db.query(CobranzaUniversoCedula.cedula).all()
    out = [str(c) for (c,) in rows if c]
    return sorted(out)


def fusionar_universo(
    db: Session,
    cedulas: Sequence[str],
    usuario_id: Optional[int] = None,
) -> dict[str, Any]:
    """Inserta solo cedulas nuevas (por clave comparable). No borra snapshots."""
    existentes = claves_universo(db)
    agregadas = 0
    ya_existian = 0
    seen_batch: set[str] = set()
    for c in cedulas:
        store = normalizar_cedula_almacenamiento(c)
        if not store:
            continue
        key = texto_cedula_comparable_bd(store)
        if not key:
            continue
        if key in existentes or key in seen_batch:
            ya_existian += 1
            continue
        seen_batch.add(key)
        db.add(CobranzaUniversoCedula(cedula=store, usuario_id=usuario_id))
        agregadas += 1
    if agregadas:
        db.commit()
    cantidad = contar_universo(db)
    return {
        "agregadas": agregadas,
        "ya_existian": ya_existian,
        "cantidad": cantidad,
        "meta": meta_universo(db),
    }


def agregar_cedula_universo(
    db: Session,
    cedula_raw: str,
    usuario_id: Optional[int] = None,
) -> dict[str, Any]:
    store = normalizar_cedula_almacenamiento(cedula_raw)
    if not store:
        raise HTTPException(status_code=400, detail="Cedula invalida o vacia.")
    key = texto_cedula_comparable_bd(store)
    if not key:
        raise HTTPException(status_code=400, detail="Cedula invalida o vacia.")
    if key in claves_universo(db):
        return {
            "cedula": store,
            "agregada": False,
            "cantidad": contar_universo(db),
        }
    db.add(CobranzaUniversoCedula(cedula=store, usuario_id=usuario_id))
    db.commit()
    return {
        "cedula": store,
        "agregada": True,
        "cantidad": contar_universo(db),
    }


def eliminar_cedula_universo(db: Session, cedula_raw: str) -> dict[str, Any]:
    store = normalizar_cedula_almacenamiento(cedula_raw)
    if not store:
        raise HTTPException(status_code=400, detail="Cedula invalida o vacia.")
    key = texto_cedula_comparable_bd(store)
    if not key:
        raise HTTPException(status_code=400, detail="Cedula invalida o vacia.")
    eliminada = False
    rows = db.query(CobranzaUniversoCedula).all()
    for row in rows:
        if texto_cedula_comparable_bd(row.cedula) == key:
            db.delete(row)
            eliminada = True
    if eliminada:
        db.commit()
    return {
        "cedula": store,
        "eliminada": eliminada,
        "cantidad": contar_universo(db),
    }


async def upload_universo_excel(
    db: Session,
    file: UploadFile,
    usuario_id: Optional[int] = None,
) -> dict[str, Any]:
    nombre = (file.filename or "").lower()
    if not nombre.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Debe subir un archivo Excel (.xlsx o .xls)",
        )
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacio.")
    cedulas = parse_cedulas_desde_excel(content)
    if not cedulas:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron cedulas en la columna A del Excel.",
        )
    return fusionar_universo(db, cedulas, usuario_id=usuario_id)


def contar_universo(db: Session) -> int:
    return int(db.query(CobranzaUniversoCedula).count() or 0)


def claves_universo(db: Session) -> set[str]:
    rows = db.query(CobranzaUniversoCedula.cedula).all()
    out: set[str] = set()
    for (ced,) in rows:
        k = texto_cedula_comparable_bd(ced)
        if k:
            out.add(k)
    return out


def cedula_en_universo(db: Session, cedula_raw: str) -> bool:
    """Compat: ya no hay filtro Excel; cualquier cedula valida puede buscarse."""
    del db  # no usado
    key = texto_cedula_comparable_bd(cedula_raw)
    return bool(key)


def meta_universo(db: Session) -> dict[str, Any]:
    rows = (
        db.query(CobranzaUniversoCedula)
        .order_by(CobranzaUniversoCedula.creado_en.desc())
        .limit(1)
        .all()
    )
    cantidad = contar_universo(db)
    if not rows:
        return {"cantidad": 0, "cargado_en": None, "usuario_id": None}
    r = rows[0]
    return {
        "cantidad": cantidad,
        "cargado_en": r.creado_en,
        "usuario_id": r.usuario_id,
    }


def limpiar_universo(db: Session) -> dict[str, Any]:
    n_ced = db.query(CobranzaUniversoCedula).delete()
    db.query(CobranzaUniversoDesempenoDiario).delete()
    db.commit()
    return {"eliminados": int(n_ced or 0)}


def _bucket_clave(n_vencidas: int) -> Optional[str]:
    """Compat: conteo exacto sin filtro de dias (tests / callers simples)."""
    if n_vencidas <= 0:
        return None
    if n_vencidas == 1:
        return "1"
    if n_vencidas == 2:
        return "2"
    if n_vencidas == 3:
        return "3"
    if n_vencidas == 4:
        return "4"
    if n_vencidas == 5:
        return "5"
    return "6plus"


def _bucket_clave_desde_atrasos(dias_atraso: list[int]) -> Optional[str]:
    """Segmento solo por conteo de cuotas atrasadas: 1..5 exactas, >=6 → 6plus."""
    n = len(dias_atraso)
    if n <= 0:
        return None
    if n <= 5 and _cumple_ventana_segmento(dias_atraso, n):
        return str(n)
    if n >= 6 and _cumple_ventana_6plus(dias_atraso):
        return "6plus"
    return None


def _aplicar_exclusion_cliente_bucket_1(
    filas: list[tuple[int, Optional[int], str, float]],
) -> list[tuple[int, Optional[int], str, float]]:
    """Quita del bucket 1 a prestamos cuyo cliente tiene otro con >=2 atrasadas.

    Misma exclusion mutua que `_stock_1_cuota_excluyendo_prejudicial_at`.
    `filas`: (prestamo_id, cliente_id, bucket, saldo_usd).
    """
    clientes_ge2: set[int] = set()
    for _pid, cid, bucket, _saldo in filas:
        if bucket in ("2", "3", "4", "5", "6plus") and cid is not None:
            clientes_ge2.add(int(cid))
    if not clientes_ge2:
        return filas
    out: list[tuple[int, Optional[int], str, float]] = []
    for pid, cid, bucket, saldo in filas:
        if bucket == "1" and cid is not None and int(cid) in clientes_ge2:
            continue
        out.append((pid, cid, bucket, saldo))
    return out


def _empty_buckets() -> dict[str, dict[str, Any]]:
    return {
        k: {"clave": k, "cantidad": 0, "monto_usd": 0.0, "items": []}
        for k in _TABLA_BUCKET_KEYS
    }


def _upsert_snapshot_hoy(
    db: Session,
    hoy: date,
    montos: dict[str, Decimal],
    cantidades: dict[str, int],
) -> None:
    for b in _BUCKET_KEYS:
        row = (
            db.query(CobranzaUniversoDesempenoDiario)
            .filter(
                CobranzaUniversoDesempenoDiario.fecha == hoy,
                CobranzaUniversoDesempenoDiario.bucket == b,
            )
            .first()
        )
        monto = montos.get(b, Decimal("0"))
        cant = int(cantidades.get(b, 0) or 0)
        if row:
            row.monto_usd = monto
            row.cantidad_prestamos = cant
        else:
            db.add(
                CobranzaUniversoDesempenoDiario(
                    fecha=hoy,
                    bucket=b,
                    monto_usd=monto,
                    cantidad_prestamos=cant,
                )
            )
    db.commit()


_TOL_SALDO_COBRANZAS = 0.01


def _as_date_caracas(value: date | datetime | None, z: ZoneInfo) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(z).date()
    return value


def _pagado_al_dia(
    *,
    monto: float,
    fecha_pago: date | None,
    eventos: list[tuple[date, float]],
    dia: date,
    total_pagado_actual: float,
    es_hoy: bool,
) -> float:
    """Monto ya aplicado a la cuota al cierre de `dia` (as-of).

    - Suma cuota_pagos con fecha_aplicacion <= dia.
    - Si fecha_pago <= dia y aun no cubre, trata como liquidada ese dia.
    - Solo en `es_hoy` usa total_pagado actual (puede ir adelante de eventos).
      Asi un cobro de hoy baja HOY pero no reescribe AYER/lunes.
    """
    paid = 0.0
    for fd, m in eventos:
        if fd <= dia:
            paid += float(m or 0)
    monto_f = float(monto or 0)
    if (
        fecha_pago is not None
        and fecha_pago <= dia
        and paid + _TOL_SALDO_COBRANZAS < monto_f
    ):
        paid = monto_f
    if es_hoy:
        paid = max(paid, float(total_pagado_actual or 0))
    return paid


def _cuota_vencida_saldo_en_fecha(
    monto: float,
    fecha_vencimiento: date | None,
    fecha_pago: date | None,
    dia: date,
    pagado_al_dia: float,
) -> Optional[float]:
    """Saldo residual de cuota vencida en `dia` (as-of), o None si no cuenta.

    Usa `pagado_al_dia` (aplicaciones hasta ese dia), no el total_pagado futuro.
    Asi se ve mejora: ayer incluye deuda cobrada hoy; hoy ya no.
    """
    if dias_retraso_desde_vencimiento(fecha_vencimiento, dia) < 1:
        return None
    monto_f = float(monto or 0)
    paid = float(pagado_al_dia or 0)
    if paid + _TOL_SALDO_COBRANZAS >= monto_f:
        return None
    return max(0.0, monto_f - paid)


def _load_eventos_por_cuota(
    db: Session, cuota_ids: Sequence[int], z: ZoneInfo
) -> dict[int, list[tuple[date, float]]]:
    """fecha_aplicacion (Caracas) + monto_aplicado por cuota_id."""
    out: dict[int, list[tuple[date, float]]] = {int(i): [] for i in cuota_ids}
    if not cuota_ids:
        return out
    chunk = 2000
    ids = [int(i) for i in cuota_ids]
    for i in range(0, len(ids), chunk):
        batch = ids[i : i + chunk]
        rows = (
            db.query(
                CuotaPago.cuota_id,
                CuotaPago.fecha_aplicacion,
                CuotaPago.monto_aplicado,
            )
            .filter(CuotaPago.cuota_id.in_(batch))
            .all()
        )
        for cid, fa, mon in rows:
            fd = _as_date_caracas(fa, z)
            if fd is None:
                continue
            out.setdefault(int(cid), []).append((fd, float(mon or 0)))
    for cid in out:
        out[cid].sort(key=lambda x: x[0])
    return out


def _metricas_prestamo_en_fecha(
    cuotas: list[Cuota],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    dia: date,
    hoy: date,
) -> tuple[Optional[str], float, list[int]]:
    """Devuelve (bucket|None, saldo_usd, dias_atraso_por_cuota_vencida)."""
    es_hoy = dia == hoy
    dias: list[int] = []
    saldo = 0.0
    for c in cuotas:
        monto = float(c.monto or 0)
        paid_act = float(c.total_pagado or 0)
        fv = c.fecha_vencimiento
        fp = c.fecha_pago if isinstance(c.fecha_pago, date) else None
        da = dias_retraso_desde_vencimiento(fv, dia)
        if da < SEG_MIN_DIAS_CUOTA_1:
            continue
        pagado = _pagado_al_dia(
            monto=monto,
            fecha_pago=fp,
            eventos=eventos_por_cuota.get(int(c.id), []),
            dia=dia,
            total_pagado_actual=paid_act,
            es_hoy=es_hoy,
        )
        s = _cuota_vencida_saldo_en_fecha(monto, fv, fp, dia, pagado)
        if s is None:
            continue
        dias.append(int(da))
        saldo += s
    return _bucket_clave_desde_atrasos(dias), round(saldo, 2), dias


_STOCK_FN_POR_BUCKET = {
    "1": _stock_1_cuota_cobranzas_at,
    "2": _stock_2_cuotas_at,
    "3": _stock_3_cuotas_at,
    "4": _stock_4_cuotas_at,
    "5": _stock_5_cuotas_at,
    "6plus": _stock_6plus_cuotas_at,
}


def _stock_resto6plus_at(
    cuotas_meta: list[dict[str, Any]], t_ref: datetime, z: ZoneInfo
) -> set[int]:
    """16 o más cuotas atrasadas (clave API resto6plus; partición vs 1..15)."""
    overdue = _cuotas_atrasadas_para_segmento(
        cuotas_meta, t_ref, z, min_dias=SEG_MIN_DIAS_CUOTA_1
    )
    return {
        pid
        for pid, dias_list in overdue.items()
        if len(dias_list) >= SEG_MAX_N_EXACTO + 1
    }


def _stock_fn_tabla(key: str):
    if key == "1":
        return _stock_1_cuota_cobranzas_at
    if key == _RESTO_6PLUS_KEY:
        return _stock_resto6plus_at
    n = int(key)
    return lambda meta, t, z, n=n: _stock_exact_n_cuotas_at(meta, t, z, n)


_STOCK_FN_TABLA = {k: _stock_fn_tabla(k) for k in _TABLA_BUCKET_KEYS}


def _sets_fin_dia_por_bucket(
    cuotas_meta: list[dict[str, Any]],
    dia: date,
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
    *,
    stock_fns: Optional[dict[str, Any]] = None,
    as_of_fin_only: bool = False,
) -> dict[str, set[int]]:
    """Sets por segmento as-of fin de día (hoy = ahora). Misma foto que la tabla."""
    fns = stock_fns or _STOCK_FN_POR_BUCKET
    t_fin = _t_fin_dia(dia, hoy, now_z, z)
    out: dict[str, set[int]] = {}
    for key, fn in fns.items():
        if not cuotas_meta:
            out[key] = set()
            continue
        set_fin = fn(cuotas_meta, t_fin, z)
        if as_of_fin_only:
            out[key] = set_fin
            continue
        t0 = datetime.combine(dia, time(0, 0, 0), tzinfo=z)
        out[key] = fn(cuotas_meta, t0, z) & set_fin
    return out


def _saldo_usd_prestamo_en_fecha(
    cuotas: list[Cuota],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    dia: date,
    hoy: date,
) -> tuple[float, int, int, int]:
    """Saldo residual as-of + n cuotas + min/max días atraso (fv)."""
    _bucket, saldo, dias = _metricas_prestamo_en_fecha(
        cuotas, eventos_por_cuota, dia, hoy
    )
    if not dias:
        return float(saldo or 0), 0, 0, 0
    return (
        float(saldo or 0),
        len(dias),
        int(min(dias)),
        int(max(dias)),
    )


def _buckets_metricas_en_fecha(
    by_pid: dict[int, list[Cuota]],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    cuotas_meta: list[dict[str, Any]],
    dia: date,
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
    *,
    bucket_keys: Sequence[str] = _BUCKET_KEYS,
    stock_fns: Optional[dict[str, Any]] = None,
) -> tuple[dict[str, float], dict[str, int], dict[str, set[int]]]:
    """Cantidad por segmento; monto = saldo as-of de esos prestamos."""
    keys = tuple(bucket_keys)
    montos: dict[str, float] = {k: 0.0 for k in keys}
    cants: dict[str, int] = {k: 0 for k in keys}
    sets = _sets_fin_dia_por_bucket(
        cuotas_meta,
        dia,
        hoy,
        now_z,
        z,
        stock_fns=stock_fns,
        as_of_fin_only=True,
    )
    for key in keys:
        pids = sets.get(key) or set()
        cants[key] = len(pids)
        total = 0.0
        for pid in pids:
            saldo, _n, _dmin, _dmax = _saldo_usd_prestamo_en_fecha(
                by_pid.get(int(pid), []), eventos_por_cuota, dia, hoy
            )
            total += saldo
        montos[key] = round(total, 2)
    return montos, cants, sets


def _load_recaudo_por_prestamo_dia(
    db: Session,
    prestamo_ids: Sequence[int],
    desde: date,
    hasta: date,
) -> dict[tuple[int, date], float]:
    """Suma Pago.monto_pagado (USD) por préstamo y día de fecha_pago.

    Misma fecha que el gráfico de pagos ingresados: date(pagos.fecha_pago).
    """
    out: dict[tuple[int, date], float] = defaultdict(float)
    ids = [int(i) for i in prestamo_ids if i is not None]
    if not ids:
        return out
    dia_expr = cast(Pago.fecha_pago, SADate)
    chunk = 2000
    for i in range(0, len(ids), chunk):
        batch = ids[i : i + chunk]
        rows = (
            db.query(
                Pago.prestamo_id,
                dia_expr.label("dia"),
                func.coalesce(func.sum(Pago.monto_pagado), 0),
            )
            .filter(
                Pago.prestamo_id.in_(batch),
                Pago.fecha_pago.isnot(None),
                dia_expr >= desde,
                dia_expr <= hasta,
            )
            .group_by(Pago.prestamo_id, dia_expr)
            .all()
        )
        for pid, dia_pago, mon in rows:
            if pid is None or dia_pago is None:
                continue
            fd = dia_pago.date() if isinstance(dia_pago, datetime) else dia_pago
            if not isinstance(fd, date):
                continue
            out[(int(pid), fd)] += float(mon or 0)
    return out


def _recaudo_por_bucket_en_dia(
    recaudo_pid_dia: dict[tuple[int, date], float],
    sets_inicio: dict[str, set[int]],
    dia: date,
    bucket_keys: Sequence[str],
) -> dict[str, float]:
    """Recaudo del día (tabla pagos), según el segmento al inicio de ese día."""
    out: dict[str, float] = {k: 0.0 for k in bucket_keys}
    for key in bucket_keys:
        acc = 0.0
        for pid in sets_inicio.get(key) or set():
            acc += float(recaudo_pid_dia.get((int(pid), dia), 0.0) or 0)
        out[key] = round(acc, 2)
    return out


def _punto_serie_vacio(d: date) -> dict[str, Any]:
    return {
        "fecha": d,
        "monto_1": 0.0,
        "monto_2": 0.0,
        "monto_3": 0.0,
        "monto_4": 0.0,
        "monto_5": 0.0,
        "monto_6plus": 0.0,
        "monto_total": 0.0,
        "cantidad_1": 0,
        "cantidad_2": 0,
        "cantidad_3": 0,
        "cantidad_4": 0,
        "cantidad_5": 0,
        "cantidad_6plus": 0,
        "cantidad_total": 0,
        "cobrado_1": 0.0,
        "cobrado_2": 0.0,
        "cobrado_3": 0.0,
        "cobrado_4": 0.0,
        "cobrado_5": 0.0,
        "cobrado_6": 0.0,
        "cobrado_7": 0.0,
        "cobrado_8": 0.0,
        "cobrado_9": 0.0,
        "cobrado_10": 0.0,
        "cobrado_11": 0.0,
        "cobrado_12": 0.0,
        "cobrado_13": 0.0,
        "cobrado_14": 0.0,
        "cobrado_15": 0.0,
        "cobrado_6plus": 0.0,
        "cobrado_total": 0.0,
    }


def _punto_serie_desde_metricas(
    d: date,
    montos: dict[str, float],
    cants: dict[str, int],
    cobrado: Optional[dict[str, float]] = None,
) -> dict[str, Any]:
    """Totales = misma suma que la tabla (segmentos 1–15). 6plus aquí = 6–15."""
    keys_6_15 = tuple(str(i) for i in range(6, SEG_MAX_N_EXACTO + 1))
    monto_6_15 = round(sum(float(montos.get(k, 0) or 0) for k in keys_6_15), 2)
    cant_6_15 = int(sum(int(cants.get(k, 0) or 0) for k in keys_6_15))
    monto_total = round(
        sum(float(montos.get(k, 0) or 0) for k in _TABLA_BUCKET_KEYS), 2
    )
    cant_total = int(sum(int(cants.get(k, 0) or 0) for k in _TABLA_BUCKET_KEYS))
    cob = cobrado or {}
    cobrado_n = {
        f"cobrado_{n}": float(cob.get(str(n), 0) or 0)
        for n in range(1, SEG_MAX_N_EXACTO + 1)
    }
    cobrado_6_15 = round(
        sum(cobrado_n[f"cobrado_{n}"] for n in range(6, SEG_MAX_N_EXACTO + 1)), 2
    )
    cobrado_total = round(
        sum(float(cob.get(k, 0) or 0) for k in _TABLA_BUCKET_KEYS), 2
    )
    return {
        "fecha": d,
        "monto_1": float(montos.get("1", 0) or 0),
        "monto_2": float(montos.get("2", 0) or 0),
        "monto_3": float(montos.get("3", 0) or 0),
        "monto_4": float(montos.get("4", 0) or 0),
        "monto_5": float(montos.get("5", 0) or 0),
        "monto_6plus": monto_6_15,
        "monto_total": monto_total,
        "cantidad_1": int(cants.get("1", 0) or 0),
        "cantidad_2": int(cants.get("2", 0) or 0),
        "cantidad_3": int(cants.get("3", 0) or 0),
        "cantidad_4": int(cants.get("4", 0) or 0),
        "cantidad_5": int(cants.get("5", 0) or 0),
        "cantidad_6plus": cant_6_15,
        "cantidad_total": cant_total,
        **cobrado_n,
        "cobrado_6plus": cobrado_6_15,
        "cobrado_total": cobrado_total,
    }


def _serie_diaria_30_vacia(hoy: date) -> list[dict[str, Any]]:
    """30 dias calendario terminando en hoy, todos en cero."""
    return [_punto_serie_vacio(hoy - timedelta(days=29 - i)) for i in range(30)]


def _serie_diaria_30_desde_universo(
    by_pid: dict[int, list[Cuota]],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    cuotas_meta: list[dict[str, Any]],
    recaudo_pid_dia: dict[tuple[int, date], float],
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
) -> list[dict[str, Any]]:
    """30 dias: barras = recaudo (pagos); linea = saldo vencido total."""
    serie: list[dict[str, Any]] = []
    dia_antes = hoy - timedelta(days=30)
    _m0, _c0, prev_sets = _buckets_metricas_en_fecha(
        by_pid,
        eventos_por_cuota,
        cuotas_meta,
        dia_antes,
        hoy,
        now_z,
        z,
        bucket_keys=_TABLA_BUCKET_KEYS,
        stock_fns=_STOCK_FN_TABLA,
    )
    for i in range(30):
        dia = hoy - timedelta(days=29 - i)
        montos, cants, sets_dia = _buckets_metricas_en_fecha(
            by_pid,
            eventos_por_cuota,
            cuotas_meta,
            dia,
            hoy,
            now_z,
            z,
            bucket_keys=_TABLA_BUCKET_KEYS,
            stock_fns=_STOCK_FN_TABLA,
        )
        recaudo = _recaudo_por_bucket_en_dia(
            recaudo_pid_dia, prev_sets, dia, _TABLA_BUCKET_KEYS
        )
        serie.append(
            _punto_serie_desde_metricas(dia, montos, cants, cobrado=recaudo)
        )
        prev_sets = sets_dia
    return serie


def _pct_var(actual: float, base: float) -> Optional[float]:
    if abs(base) < 0.005:
        if abs(actual) < 0.005:
            return 0.0
        return None
    return round(((actual - base) / abs(base)) * 100.0, 2)


_MESES_LECTURA = (
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
)


def _mes_anterior_dia_1(d: date) -> date:
    if d.month == 1:
        return date(d.year - 1, 12, 1)
    return date(d.year, d.month - 1, 1)


def _fechas_3_meses_ayer_hoy(hoy: date) -> list[date]:
    """Día 1 de los 3 meses más recientes (sin repetir ayer/hoy) + ayer + hoy."""
    ayer = hoy - timedelta(days=1)
    ocupadas = {ayer, hoy}
    meses: list[date] = []
    cursor = date(hoy.year, hoy.month, 1)
    while len(meses) < 3:
        if cursor not in ocupadas:
            meses.append(cursor)
        cursor = _mes_anterior_dia_1(cursor)
    meses.reverse()
    return meses + [ayer, hoy]


def _etiqueta_lectura(d: date, hoy: date) -> str:
    dd = d.strftime("%d/%m")
    if d == hoy:
        return f"Hoy {dd}"
    if d == hoy - timedelta(days=1):
        return f"Ayer {dd}"
    if d.day == 1:
        return f"1 de {_MESES_LECTURA[d.month - 1]}"
    return dd


def _ultimo_viernes_del_mes(year: int, month: int) -> date:
    """Último viernes calendario del mes (weekday: lun=0 … vie=4)."""
    if month == 12:
        last = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(year, month + 1, 1) - timedelta(days=1)
    return last - timedelta(days=(last.weekday() - 4) % 7)


def _ultimos_viernes_cierre_meses(hoy: date, n: int = 2) -> list[date]:
    """Último viernes de los n meses anteriores, del más reciente al más viejo."""
    y, m = hoy.year, hoy.month
    out: list[date] = []
    for _ in range(max(0, int(n))):
        if m == 1:
            y, m = y - 1, 12
        else:
            m -= 1
        out.append(_ultimo_viernes_del_mes(y, m))
    return out


def _etiqueta_bin_atraso(i: int) -> str:
    if i >= _ATRASO_N_BINS:
        return ">600 días"
    desde = i * _ATRASO_BIN_DIAS + 1
    hasta = min((i + 1) * _ATRASO_BIN_DIAS, _ATRASO_MAX_DIAS)
    return f"{desde}–{hasta}"


def _idx_bin_atraso(dias: int) -> int:
    d = max(1, int(dias or 0))
    if d > _ATRASO_MAX_DIAS:
        return _ATRASO_N_BINS
    return min(_ATRASO_N_BINS - 1, (d - 1) // _ATRASO_BIN_DIAS)


def _bins_atraso_vacios() -> list[dict[str, Any]]:
    n_bins = _ATRASO_N_BINS + 1
    return [
        {"label": _etiqueta_bin_atraso(i), "casos": 0, "monto_usd": 0.0}
        for i in range(n_bins)
    ]


def _distribucion_atraso_en_fecha(
    by_pid: dict[int, list[Cuota]],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    cuotas_meta: list[dict[str, Any]],
    dia: date,
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
) -> list[dict[str, Any]]:
    """Casos por tramo de días de atraso (hoy/fecha − fv más antigua), segmentos 1–15."""
    n_bins = _ATRASO_N_BINS + 1
    casos = [0] * n_bins
    montos = [0.0] * n_bins
    sets = _sets_fin_dia_por_bucket(
        cuotas_meta,
        dia,
        hoy,
        now_z,
        z,
        stock_fns=_STOCK_FN_TABLA,
        as_of_fin_only=True,
    )
    seen: set[int] = set()
    for key in _TABLA_BUCKET_KEYS:
        for pid in sets.get(key) or set():
            ipid = int(pid)
            if ipid in seen:
                continue
            seen.add(ipid)
            saldo, _n, _dmin, dmax = _saldo_usd_prestamo_en_fecha(
                by_pid.get(ipid, []), eventos_por_cuota, dia, hoy
            )
            idx = _idx_bin_atraso(int(dmax or 0))
            casos[idx] += 1
            montos[idx] += float(saldo or 0)
    return [
        {
            "label": _etiqueta_bin_atraso(i),
            "casos": casos[i],
            "monto_usd": round(montos[i], 2),
        }
        for i in range(n_bins)
    ]


def _dist_atraso_viernes_cierre(
    by_pid: dict[int, list[Cuota]],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    cuotas_meta: list[dict[str, Any]],
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
    n: int = 2,
) -> list[dict[str, Any]]:
    """Distribución as-of último viernes de los n meses anteriores."""
    fechas = _ultimos_viernes_cierre_meses(hoy, n)
    out: list[dict[str, Any]] = []
    for dia in fechas:
        mes = _MESES_LECTURA[dia.month - 1]
        out.append(
            {
                "fecha": dia.isoformat(),
                "etiqueta": mes,
                "bins": _distribucion_atraso_en_fecha(
                    by_pid,
                    eventos_por_cuota,
                    cuotas_meta,
                    dia,
                    hoy,
                    now_z,
                    z,
                ),
            }
        )
    return out


def _lecturas_lunes_desempeno(
    by_pid: dict[int, list[Cuota]],
    eventos_por_cuota: dict[int, list[tuple[date, float]]],
    cuotas_meta: list[dict[str, Any]],
    hoy: date,
    now_z: datetime,
    z: ZoneInfo,
) -> dict[str, Any]:
    """Cantidad = N cuotas atrasadas; monto = saldo as-of. Filas 1..15."""
    fechas = _fechas_3_meses_ayer_hoy(hoy)
    ayer = hoy - timedelta(days=1)
    snaps: list[tuple[date, dict[str, float], dict[str, int]]] = []
    for dia in fechas:
        montos, cants, _sets = _buckets_metricas_en_fecha(
            by_pid,
            eventos_por_cuota,
            cuotas_meta,
            dia,
            hoy,
            now_z,
            z,
            bucket_keys=_TABLA_BUCKET_KEYS,
            stock_fns=_STOCK_FN_TABLA,
        )
        snaps.append((dia, montos, cants))

    columnas = [
        {
            "fecha": dia.isoformat(),
            "etiqueta": _etiqueta_lectura(dia, hoy),
            "es_hoy": dia == hoy,
            "es_ayer": dia == ayer,
        }
        for dia, _, _ in snaps
    ]

    buckets_out: dict[str, Any] = {}
    for b in _TABLA_BUCKET_KEYS:
        lecturas = []
        for dia, montos, cants in snaps:
            lecturas.append(
                {
                    "fecha": dia.isoformat(),
                    "cantidad": int(cants.get(b, 0) or 0),
                    "monto_usd": float(montos.get(b, 0) or 0),
                }
            )
        buckets_out[b] = {"clave": b, "lecturas": lecturas}

    total_lecturas = []
    for dia, montos, cants in snaps:
        total_lecturas.append(
            {
                "fecha": dia.isoformat(),
                "cantidad": int(
                    sum(int(cants.get(k, 0) or 0) for k in _TABLA_BUCKET_KEYS)
                ),
                "monto_usd": round(
                    sum(float(montos.get(k, 0) or 0) for k in _TABLA_BUCKET_KEYS), 2
                ),
            }
        )

    return {
        "columnas": columnas,
        "buckets": buckets_out,
        "total": {"clave": "total", "lecturas": total_lecturas},
    }


def analizar_universo(db: Session) -> dict[str, Any]:
    """Cobranzas: CANTIDAD = Fin dia del dashboard; MONTO = saldo as-of USD.

    Sin tocar dashboard. Cartera sin LIQUIDADO/DESISTIMIENTO.
    """
    hoy = hoy_negocio()
    cache_key = f"universo_analisis:{hoy.isoformat()}"
    now_mono = time_mod.monotonic()
    with _analisis_cache_lock:
        hit = _analisis_cache.get(cache_key)
        if hit is not None and now_mono - hit[0] < _ANALISIS_CACHE_TTL_SEC:
            return copy.deepcopy(hit[1])

    buckets = _empty_buckets()
    sin_vencidas = 0
    z = ZoneInfo(TZ_NEGOCIO)
    now_z = datetime.now(z)

    prestamos = (
        db.query(Prestamo)
        .filter(expr_prestamo_activo_cobranzas())
        .filter(sql_cliente_sin_desistimiento())
        .all()
    )
    meta = {
        "cantidad": len(prestamos),
        "cargado_en": None,
        "usuario_id": None,
        "fuente": "bd_completa",
        "segmentacion": "conteo_cuotas_atrasadas",
        "cantidad_origen": "n_cuotas_vencidas_sin_pagar",
        "monto_origen": "saldo_asof_usd",
    }
    pids = [int(p.id) for p in prestamos]
    prestamo_by_id: dict[int, Prestamo] = {int(p.id): p for p in prestamos}

    by_pid: dict[int, list[Cuota]] = defaultdict(list)
    cuota_ids: list[int] = []
    if pids:
        for c in db.query(Cuota).filter(Cuota.prestamo_id.in_(pids)).all():
            by_pid[int(c.prestamo_id)].append(c)
            cuota_ids.append(int(c.id))

    eventos_por_cuota = _load_eventos_por_cuota(db, cuota_ids, z)

    # Meta de stock identica al dashboard (1 cuota / 2 / 3 / 4+).
    fv_max = hoy - timedelta(days=1)
    cuotas_meta = _load_cuotas_meta(db, fv_min=None, fv_max=fv_max, z=z)

    montos_snap: dict[str, Decimal] = {k: Decimal("0") for k in _BUCKET_KEYS}
    cant_snap: dict[str, int] = {k: 0 for k in _BUCKET_KEYS}

    for pid in pids:
        _b, _saldo, dias = _metricas_prestamo_en_fecha(
            by_pid.get(pid, []), eventos_por_cuota, hoy, hoy
        )
        if not dias:
            sin_vencidas += 1

    # Detalle + totales alineados a la tabla (solo 1..15).
    montos_hoy, cants_hoy, sets_hoy = _buckets_metricas_en_fecha(
        by_pid,
        eventos_por_cuota,
        cuotas_meta,
        hoy,
        hoy,
        now_z,
        z,
        bucket_keys=_TABLA_BUCKET_KEYS,
        stock_fns=_STOCK_FN_TABLA,
    )

    for key in _TABLA_BUCKET_KEYS:
        for pid in sorted(sets_hoy.get(key) or set()):
            p = prestamo_by_id.get(int(pid))
            if p is None:
                continue
            saldo_r, n_venc, dmin, dmax = _saldo_usd_prestamo_en_fecha(
                by_pid.get(int(pid), []), eventos_por_cuota, hoy, hoy
            )
            item = {
                "prestamo_id": int(pid),
                "cedula": p.cedula or "",
                "nombres": p.nombres,
                "cuotas_vencidas": int(n_venc),
                "saldo_vencido_usd": float(saldo_r),
                "dias_atraso_min": int(dmin),
                "dias_atraso_max": int(dmax),
            }
            buckets[key]["items"].append(item)
        buckets[key]["cantidad"] = int(cants_hoy.get(key, 0) or 0)
        buckets[key]["monto_usd"] = float(montos_hoy.get(key, 0) or 0)

    # Snapshot/serie diaria de gráficos: sigue 1..5 + 6plus.
    montos_chart, cants_chart, _sets_chart = _buckets_metricas_en_fecha(
        by_pid, eventos_por_cuota, cuotas_meta, hoy, hoy, now_z, z
    )
    for key in _BUCKET_KEYS:
        montos_snap[key] = Decimal(str(montos_chart.get(key, 0) or 0))
        cant_snap[key] = int(cants_chart.get(key, 0) or 0)

    _upsert_snapshot_hoy(db, hoy, montos_snap, cant_snap)

    recaudo_pid_dia = _load_recaudo_por_prestamo_dia(
        db, pids, hoy - timedelta(days=29), hoy
    )
    serie = _serie_diaria_30_desde_universo(
        by_pid, eventos_por_cuota, cuotas_meta, recaudo_pid_dia, hoy, now_z, z
    )

    meta["cantidad"] = len(prestamos)
    result = {
        "buckets": buckets,
        "sin_vencidas": sin_vencidas,
        "serie_diaria": serie,
        "desempeno_lecturas": _lecturas_lunes_desempeno(
            by_pid, eventos_por_cuota, cuotas_meta, hoy, now_z, z
        ),
        "dist_atraso_viernes_cierre": _dist_atraso_viernes_cierre(
            by_pid, eventos_por_cuota, cuotas_meta, hoy, now_z, z
        ),
        "meta": meta,
    }
    with _analisis_cache_lock:
        _analisis_cache.clear()
        _analisis_cache[cache_key] = (now_mono, result)
    return copy.deepcopy(result)
