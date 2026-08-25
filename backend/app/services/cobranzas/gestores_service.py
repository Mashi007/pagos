# -*- coding: utf-8 -*-
"""
Gestores de cobranza: asignacion fija, Excel en vivo y dashboard.

- Universo: prestamos APROBADO con fecha_aprobacion >= 2026-03-01 y <= hoy (Caracas)
  y al menos una cuota VENCIDO/MORA con vencimiento <= hoy.
- Unidad de asignacion: el **prestamo completo** (nunca se parte un prestamo entre
  gestores). UNIQUE(prestamo_id). Ademas, todos los prestamos de la misma cedula
  van al mismo gestor.
- Reparto equilibrado por dolares vencidos+mora y cantidad de cuotas,
  **por cedula** (todos los prestamos de la misma persona van al mismo gestor).
- Asignacion sticky: no se rebalancea ni se agregan casos nuevos tras el primer cierre;
  si una cedula quedara partida entre gestores, se consolida al abrir/usar el modulo.
  Antes de Excel/correo se audita integridad (prestamo y cedula en un solo gestor).
- Si un prestamo pasa a LIQUIDADO (u otro estado distinto de APROBADO), o su
  fecha_aprobacion queda fuera del rango, sale de la lista Excel/dashboard
  (la asignacion historica se conserva).
- Excel / montos: siempre recalculados desde BD (pagos actualizan al instante).
"""
from __future__ import annotations

import io
import logging
import re
import threading
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import Date, cast, func, select
from sqlalchemy.orm import Session

from app.models.cliente import Cliente
from app.models.cobranza_gestor import (
    CobranzaGestorAsignacion,
    CobranzaGestorDesempenoDiario,
)
from app.models.configuracion import Configuracion
from app.models.cuota import Cuota
from app.models.prestamo import Prestamo
from app.services.cobranzas.gestores_constantes import (
    EMAIL_GESTORES_BCC,
    EMAIL_GESTORES_TO,
    FECHA_INICIO_APROBACION_GESTORES,
    FECHA_INICIO_CARTERA_GESTORES,
    GESTOR_NOMBRES,
    GESTOR_SLUGS,
    GESTORES,
)
from app.services.cuota_estado import hoy_negocio
from app.utils.cedula_almacenamiento import texto_cedula_comparable_bd

logger = logging.getLogger(__name__)

CLAVE_ASIGNACION_CERRADA = "cobranza_gestores_asignacion_cerrada_v2_aprob_marzo"
ESTADOS_ATRASO = ("VENCIDO", "MORA")

_asignacion_bg_lock = threading.Lock()
_asignacion_bg_running = False


def _totales_vacios() -> List[Dict[str, Any]]:
    return [
        {
            "slug": slug,
            "nombre": nombre,
            "cantidad_casos": 0,
            "total_cobranza_usd": 0.0,
            "usd_vencidas": 0.0,
            "usd_mora": 0.0,
        }
        for slug, nombre in GESTORES
    ]


def _f(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _residual(monto: Any, total_pagado: Any) -> float:
    return max(0.0, _f(monto) - _f(total_pagado))


def listar_gestores() -> List[Dict[str, str]]:
    return [{"slug": s, "nombre": n} for s, n in GESTORES]


def _fecha_aprobacion_date(prestamo: Prestamo) -> Optional[date]:
    fa = getattr(prestamo, "fecha_aprobacion", None)
    if fa is None:
        return None
    if isinstance(fa, datetime):
        return fa.date()
    if isinstance(fa, date):
        return fa
    try:
        return date.fromisoformat(str(fa)[:10])
    except ValueError:
        return None


def _prestamo_elegible_gestores(prestamo: Prestamo, *, hoy: date) -> bool:
    """APROBADO con fecha_aprobacion entre 1-mar y hoy (Caracas)."""
    if (prestamo.estado or "").strip().upper() != "APROBADO":
        return False
    fa = _fecha_aprobacion_date(prestamo)
    if fa is None:
        return False
    return FECHA_INICIO_APROBACION_GESTORES <= fa <= hoy


def _clave_cedula_persona(
    cliente: Optional[Cliente], prestamo: Prestamo
) -> str:
    """Clave comparable de la persona; si no hay cedula, cae a prestamo:<id> (no se agrupa)."""
    raw = ""
    if cliente is not None and getattr(cliente, "cedula", None):
        raw = str(cliente.cedula)
    elif getattr(prestamo, "cedula", None):
        raw = str(prestamo.cedula)
    clave = texto_cedula_comparable_bd(raw)
    if clave:
        return clave
    return f"prestamo:{int(prestamo.id)}"


def _asignacion_cerrada(db: Session) -> bool:
    row = db.get(Configuracion, CLAVE_ASIGNACION_CERRADA)
    if not row or not row.valor:
        return False
    return str(row.valor).strip().lower() in ("1", "true", "si", "yes")


def _marcar_asignacion_cerrada(db: Session) -> None:
    row = db.get(Configuracion, CLAVE_ASIGNACION_CERRADA)
    if row is None:
        db.add(Configuracion(clave=CLAVE_ASIGNACION_CERRADA, valor="true"))
    else:
        row.valor = "true"


def _metricas_cuotas_atraso(
    cuotas: Sequence[Cuota],
    *,
    hasta: date,
    desde: Optional[date] = None,
) -> Dict[str, float]:
    """
    Cuotas VENCIDO/MORA con vencimiento <= hoy.
    `desde` es opcional (legacy); el filtro de ingreso al modulo es fecha_aprobacion.
    """
    cant_venc = 0
    usd_venc = 0.0
    cant_mora = 0
    usd_mora = 0.0
    total_pagado = 0.0
    for c in cuotas:
        total_pagado += _f(c.total_pagado)
        fv = c.fecha_vencimiento
        if fv is None or fv > hasta:
            continue
        if desde is not None and fv < desde:
            continue
        est = (c.estado or "").strip().upper()
        if est not in ESTADOS_ATRASO:
            continue
        res = _residual(c.monto, c.total_pagado)
        if est == "MORA":
            cant_mora += 1
            usd_mora += res
        else:
            cant_venc += 1
            usd_venc += res
    return {
        "cant_vencidas": float(cant_venc),
        "usd_vencidas": usd_venc,
        "cant_mora": float(cant_mora),
        "usd_mora": usd_mora,
        "total_pagado": total_pagado,
        "carga_usd": usd_venc + usd_mora,
        "carga_cuotas": float(cant_venc + cant_mora),
    }


def _cargar_universo_inicial(db: Session) -> List[Dict[str, Any]]:
    """
    Prestamos APROBADO con fecha_aprobacion en [1-mar .. hoy] y al menos
    una cuota VENCIDO/MORA vencida hasta hoy.
    """
    hoy = hoy_negocio()
    desde_apr = FECHA_INICIO_APROBACION_GESTORES
    rows = db.execute(
        select(Prestamo, Cliente)
        .join(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(
            func.upper(func.trim(Prestamo.estado)) == "APROBADO",
            Prestamo.fecha_aprobacion.isnot(None),
            cast(Prestamo.fecha_aprobacion, Date) >= desde_apr,
            cast(Prestamo.fecha_aprobacion, Date) <= hoy,
        )
        .order_by(Prestamo.id.asc())
    ).all()
    if not rows:
        return []

    pids = [int(p.id) for p, _c in rows]
    cuotas_by_pid: Dict[int, List[Cuota]] = {pid: [] for pid in pids}
    for c in (
        db.execute(
            select(Cuota)
            .where(Cuota.prestamo_id.in_(pids))
            .order_by(Cuota.prestamo_id.asc(), Cuota.numero_cuota.asc())
        )
        .scalars()
        .all()
    ):
        cuotas_by_pid.setdefault(int(c.prestamo_id), []).append(c)

    out: List[Dict[str, Any]] = []
    for prestamo, cliente in rows:
        if not _prestamo_elegible_gestores(prestamo, hoy=hoy):
            continue
        m = _metricas_cuotas_atraso(
            cuotas_by_pid.get(int(prestamo.id), []),
            hasta=hoy,
        )
        if m["carga_cuotas"] <= 0:
            continue
        out.append(
            {
                "prestamo_id": int(prestamo.id),
                "cedula_clave": _clave_cedula_persona(cliente, prestamo),
                "carga_usd": m["carga_usd"],
                "carga_cuotas": m["carga_cuotas"],
            }
        )
    return out


def _agrupar_universo_por_cedula(
    items: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Agrupa prestamos del universo por cedula (misma persona = un bloque)."""
    groups: Dict[str, Dict[str, Any]] = {}
    for it in items:
        clave = str(it.get("cedula_clave") or f"prestamo:{it['prestamo_id']}")
        g = groups.get(clave)
        if g is None:
            g = {
                "cedula_clave": clave,
                "items": [],
                "carga_usd": 0.0,
                "carga_cuotas": 0.0,
            }
            groups[clave] = g
        g["items"].append(it)
        g["carga_usd"] += float(it["carga_usd"])
        g["carga_cuotas"] += float(it["carga_cuotas"])
    return list(groups.values())


def _gestor_mayoria_en_grupo(slugs: Sequence[str]) -> str:
    """Gestor con mas prestamos del grupo; desempate por slug estable."""
    cnt = Counter(s for s in slugs if s)
    return min(cnt.keys(), key=lambda s: (-cnt[s], s))


def _consolidar_asignaciones_por_cedula(db: Session) -> int:
    """
    Si la misma cedula tiene prestamos en varios gestores, mueve todos al gestor
    que ya tiene mas prestamos de esa persona. Devuelve cuantas filas se movieron.
    """
    asigs = list(db.execute(select(CobranzaGestorAsignacion)).scalars().all())
    if not asigs:
        return 0

    pids = [int(a.prestamo_id) for a in asigs]
    prestamos = {
        int(p.id): p
        for p in db.execute(select(Prestamo).where(Prestamo.id.in_(pids)))
        .scalars()
        .all()
    }
    cids = {int(p.cliente_id) for p in prestamos.values() if p.cliente_id}
    clientes: Dict[int, Cliente] = {}
    if cids:
        clientes = {
            int(c.id): c
            for c in db.execute(select(Cliente).where(Cliente.id.in_(cids)))
            .scalars()
            .all()
        }

    by_ced: Dict[str, List[CobranzaGestorAsignacion]] = defaultdict(list)
    for asg in asigs:
        prestamo = prestamos.get(int(asg.prestamo_id))
        if not prestamo:
            continue
        cliente = clientes.get(int(prestamo.cliente_id)) if prestamo.cliente_id else None
        clave = _clave_cedula_persona(cliente, prestamo)
        by_ced[clave].append(asg)

    moved = 0
    for clave, group in by_ced.items():
        if clave.startswith("prestamo:"):
            continue
        slugs = [a.gestor_slug for a in group]
        if len(set(slugs)) <= 1:
            continue
        target = _gestor_mayoria_en_grupo(slugs)
        for asg in group:
            if asg.gestor_slug != target:
                asg.gestor_slug = target
                moved += 1

    if moved:
        db.commit()
        logger.info(
            "[gestores] consolidacion por cedula movidos=%s grupos_revisados=%s",
            moved,
            len(by_ced),
        )
    return moved


def _auditar_integridad_asignaciones(db: Session) -> Dict[str, Any]:
    """
    Verifica:
    - Cada prestamo_id aparece una sola vez (prestamo completo → 1 gestor).
    - Cada cedula comparable aparece en un solo gestor.
    """
    asigs = list(db.execute(select(CobranzaGestorAsignacion)).scalars().all())
    pids = [int(a.prestamo_id) for a in asigs]
    dup_pids = sorted(pid for pid, n in Counter(pids).items() if n > 1)

    prestamos: Dict[int, Prestamo] = {}
    clientes: Dict[int, Cliente] = {}
    if asigs:
        prestamos = {
            int(p.id): p
            for p in db.execute(select(Prestamo).where(Prestamo.id.in_(pids)))
            .scalars()
            .all()
        }
        cids = {int(p.cliente_id) for p in prestamos.values() if p.cliente_id}
        if cids:
            clientes = {
                int(c.id): c
                for c in db.execute(select(Cliente).where(Cliente.id.in_(cids)))
                .scalars()
                .all()
            }

    by_ced: Dict[str, List[CobranzaGestorAsignacion]] = defaultdict(list)
    for asg in asigs:
        prestamo = prestamos.get(int(asg.prestamo_id))
        if not prestamo:
            continue
        cliente = clientes.get(int(prestamo.cliente_id)) if prestamo.cliente_id else None
        clave = _clave_cedula_persona(cliente, prestamo)
        by_ced[clave].append(asg)

    cedulas_partidas: List[Dict[str, Any]] = []
    for clave, group in by_ced.items():
        if clave.startswith("prestamo:"):
            continue
        slugs = sorted({(a.gestor_slug or "").strip().lower() for a in group})
        if len(slugs) <= 1:
            continue
        cedulas_partidas.append(
            {
                "cedula_clave": clave,
                "gestores": slugs,
                "prestamo_ids": sorted(int(a.prestamo_id) for a in group),
            }
        )

    return {
        "ok": not dup_pids and not cedulas_partidas,
        "total_asignaciones": len(asigs),
        "prestamos_duplicados": dup_pids,
        "cedulas_partidas": cedulas_partidas,
    }


def _garantizar_integridad_listas(db: Session) -> Dict[str, Any]:
    """
    Consolida cedulas partidas y exige:
    - prestamo completo en un solo gestor (nunca partido entre 2 personas);
    - misma cedula en un solo gestor.
    """
    consolidados = _consolidar_asignaciones_por_cedula(db)
    report = _auditar_integridad_asignaciones(db)
    if not report["ok"]:
        consolidados += _consolidar_asignaciones_por_cedula(db)
        report = _auditar_integridad_asignaciones(db)

    if report["prestamos_duplicados"]:
        logger.error(
            "[gestores] prestamos duplicados en asignacion: %s",
            report["prestamos_duplicados"][:30],
        )
        raise RuntimeError(
            "Integridad gestores: un prestamo no puede estar en dos listas "
            f"(duplicados={report['prestamos_duplicados'][:10]})"
        )
    if report["cedulas_partidas"]:
        logger.error(
            "[gestores] cedulas partidas tras consolidar: %s",
            report["cedulas_partidas"][:10],
        )
        raise RuntimeError(
            "Integridad gestores: una cedula no puede estar en dos listas "
            f"(ej={report['cedulas_partidas'][:3]})"
        )

    report["consolidados"] = consolidados
    return report


def asegurar_asignaciones(
    db: Session, *, verificar_integridad: bool = False
) -> Dict[str, Any]:
    """
    Si la asignacion no esta cerrada: reparte el universo entre los 9 gestores
    por bloques de cedula (greedy por menor carga USD) y cierra.
    Si ya esta cerrada: no-op rapido (integridad solo si verificar_integridad=True).
    """
    if _asignacion_cerrada(db):
        n = db.scalar(select(func.count()).select_from(CobranzaGestorAsignacion)) or 0
        out: Dict[str, Any] = {
            "cerrada": True,
            "asignados": int(n),
            "nuevos": 0,
            "consolidados": 0,
            "integridad_ok": True,
        }
        if verificar_integridad:
            integridad = _garantizar_integridad_listas(db)
            out["consolidados"] = int(integridad.get("consolidados") or 0)
            out["integridad_ok"] = bool(integridad.get("ok", True))
        return out

    existentes_rows = list(db.execute(select(CobranzaGestorAsignacion)).scalars().all())
    asig_por_prestamo = {int(a.prestamo_id): a.gestor_slug for a in existentes_rows}
    existentes = set(asig_por_prestamo.keys())

    universo = _cargar_universo_inicial(db)
    grupos = _agrupar_universo_por_cedula(universo)

    cargas: Dict[str, Dict[str, float]] = {
        s: {"usd": 0.0, "cuotas": 0.0, "n": 0.0} for s in GESTOR_SLUGS
    }
    # Solo conteo de ya asignados (rapido); el reparto nuevo usa carga del universo.
    for asg in existentes_rows:
        slug = asg.gestor_slug
        if slug in cargas:
            cargas[slug]["n"] += 1.0

    pendientes: List[Dict[str, Any]] = []
    for g in grupos:
        items_pend = [it for it in g["items"] if int(it["prestamo_id"]) not in existentes]
        if not items_pend:
            continue
        forced: Optional[str] = None
        for it in g["items"]:
            slug_prev = asig_por_prestamo.get(int(it["prestamo_id"]))
            if slug_prev:
                forced = slug_prev
                break
        pendientes.append(
            {
                "cedula_clave": g["cedula_clave"],
                "items": items_pend,
                "carga_usd": sum(float(it["carga_usd"]) for it in items_pend),
                "carga_cuotas": sum(float(it["carga_cuotas"]) for it in items_pend),
                "forced": forced,
                "min_prestamo_id": min(int(it["prestamo_id"]) for it in items_pend),
            }
        )

    pendientes.sort(
        key=lambda x: (x["carga_usd"], x["carga_cuotas"], -x["min_prestamo_id"]),
        reverse=True,
    )

    nuevos = 0
    for bloque in pendientes:
        slug = bloque["forced"] or min(
            GESTOR_SLUGS,
            key=lambda s: (cargas[s]["usd"], cargas[s]["cuotas"], cargas[s]["n"], s),
        )
        for item in bloque["items"]:
            pid = int(item["prestamo_id"])
            if pid in existentes:
                continue
            db.add(
                CobranzaGestorAsignacion(
                    prestamo_id=pid,
                    gestor_slug=slug,
                )
            )
            existentes.add(pid)
            nuevos += 1
        cargas[slug]["usd"] += float(bloque["carga_usd"])
        cargas[slug]["cuotas"] += float(bloque["carga_cuotas"])
        cargas[slug]["n"] += len(bloque["items"])

    _marcar_asignacion_cerrada(db)
    db.commit()
    if verificar_integridad:
        integridad = _garantizar_integridad_listas(db)
        consolidados = int(integridad.get("consolidados") or 0)
    else:
        consolidados = _consolidar_asignaciones_por_cedula(db)
    total = db.scalar(select(func.count()).select_from(CobranzaGestorAsignacion)) or 0
    logger.info(
        "[gestores] asignacion cerrada (por cedula) nuevos=%s total=%s consolidados=%s",
        nuevos,
        total,
        consolidados,
    )
    return {
        "cerrada": True,
        "asignados": int(total),
        "nuevos": nuevos,
        "consolidados": consolidados,
        "integridad_ok": True,
    }


def _kick_asignacion_background() -> bool:
    """Lanza el reparto inicial en hilo daemon; evita timeout del dashboard."""
    global _asignacion_bg_running
    with _asignacion_bg_lock:
        if _asignacion_bg_running:
            return True
        _asignacion_bg_running = True

    def _worker() -> None:
        global _asignacion_bg_running
        from app.core.database import SessionLocal

        db = SessionLocal()
        try:
            asegurar_asignaciones(db, verificar_integridad=True)
        except Exception:
            logger.exception("[gestores] asignacion background fallo")
        finally:
            db.close()
            with _asignacion_bg_lock:
                _asignacion_bg_running = False

    threading.Thread(
        target=_worker, name="gestores-asignacion-bg", daemon=True
    ).start()
    return True


def asignacion_en_progreso() -> bool:
    with _asignacion_bg_lock:
        return _asignacion_bg_running


def _fila_caso(
    prestamo: Prestamo,
    cliente: Cliente,
    cuotas: Sequence[Cuota],
    *,
    hasta: date,
) -> Dict[str, Any]:
    m = _metricas_cuotas_atraso(cuotas, hasta=hasta)
    cant_vencidas = int(m["cant_vencidas"] + m["cant_mora"])
    monto_vencido = round(m["carga_usd"], 2)
    return {
        "prestamo_id": int(prestamo.id),
        "cedula": (cliente.cedula or prestamo.cedula or "").strip(),
        "nombres": (cliente.nombres or prestamo.nombres or "").strip(),
        "telefono": (cliente.telefono or "").strip(),
        "email": (cliente.email or "").strip(),
        "cant_cuotas_vencidas": cant_vencidas,
        "monto_vencido_usd": monto_vencido,
        # Compat dashboard / snapshots (mismo monto unificado).
        "total_cobranza_usd": monto_vencido,
        "usd_cuotas_vencidas": round(m["usd_vencidas"], 2),
        "usd_cuotas_mora": round(m["usd_mora"], 2),
        "cant_cuotas_mora": int(m["cant_mora"]),
    }


def _append_filas_cartera_excel(ws, filas: Sequence[Dict[str, Any]]) -> None:
    """Columnas operativas: cedula, nombres, telefono, email, cuotas vencidas, monto USD."""
    headers = [
        "Cedula",
        "Nombres",
        "Telefono",
        "Email",
        "Cuotas vencidas",
        "Monto vencido (USD)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for f in filas:
        ws.append(
            [
                f["cedula"],
                f["nombres"],
                f["telefono"],
                f["email"],
                int(f.get("cant_cuotas_vencidas") or 0),
                round(float(f.get("monto_vencido_usd") or f.get("total_cobranza_usd") or 0), 2),
            ]
        )


def filas_gestor(db: Session, gestor_slug: str) -> List[Dict[str, Any]]:
    slug = (gestor_slug or "").strip().lower()
    if slug not in GESTOR_NOMBRES:
        raise ValueError(f"Gestor desconocido: {gestor_slug}")
    asegurar_asignaciones(db)
    return _filas_gestor_sin_asegurar(db, slug)


def _cargar_asignaciones_vivas(
    db: Session, *, gestor_slug: Optional[str] = None
) -> Tuple[
    List[CobranzaGestorAsignacion],
    Dict[int, Prestamo],
    Dict[int, Cliente],
    Dict[int, List[Cuota]],
]:
    """Carga asignaciones + prestamos APROBADO + clientes + cuotas en pocas queries."""
    q = select(CobranzaGestorAsignacion).order_by(
        CobranzaGestorAsignacion.gestor_slug.asc(),
        CobranzaGestorAsignacion.prestamo_id.asc(),
    )
    if gestor_slug:
        q = q.where(CobranzaGestorAsignacion.gestor_slug == gestor_slug)
    asigs = list(db.execute(q).scalars().all())
    if not asigs:
        return [], {}, {}, {}

    pids = [int(a.prestamo_id) for a in asigs]
    prestamos = {
        int(p.id): p
        for p in db.execute(
            select(Prestamo).where(
                Prestamo.id.in_(pids),
                func.upper(func.trim(Prestamo.estado)) == "APROBADO",
            )
        )
        .scalars()
        .all()
    }
    if not prestamos:
        return asigs, {}, {}, {}

    cids = {int(p.cliente_id) for p in prestamos.values() if p.cliente_id}
    clientes: Dict[int, Cliente] = {}
    if cids:
        clientes = {
            int(c.id): c
            for c in db.execute(select(Cliente).where(Cliente.id.in_(cids)))
            .scalars()
            .all()
        }

    cuotas_by_pid: Dict[int, List[Cuota]] = {pid: [] for pid in prestamos}
    for c in (
        db.execute(
            select(Cuota)
            .where(Cuota.prestamo_id.in_(list(prestamos.keys())))
            .order_by(Cuota.prestamo_id.asc(), Cuota.numero_cuota.asc())
        )
        .scalars()
        .all()
    ):
        cuotas_by_pid.setdefault(int(c.prestamo_id), []).append(c)

    return asigs, prestamos, clientes, cuotas_by_pid


def _filas_gestor_sin_asegurar(db: Session, slug: str) -> List[Dict[str, Any]]:
    hoy = hoy_negocio()
    asigs, prestamos, clientes, cuotas_by_pid = _cargar_asignaciones_vivas(
        db, gestor_slug=slug
    )
    filas: List[Dict[str, Any]] = []
    vistos: set[int] = set()
    for asg in asigs:
        pid = int(asg.prestamo_id)
        if pid in vistos:
            # Defensa: un prestamo completo nunca debe aparecer dos veces.
            continue
        vistos.add(pid)
        prestamo = prestamos.get(pid)
        if not prestamo or not _prestamo_elegible_gestores(prestamo, hoy=hoy):
            # Fuera de rango de aprobacion / no APROBADO: sale de la lista.
            continue
        cliente = clientes.get(int(prestamo.cliente_id)) if prestamo.cliente_id else None
        if not cliente:
            continue
        filas.append(
            _fila_caso(
                prestamo,
                cliente,
                cuotas_by_pid.get(int(prestamo.id), []),
                hasta=hoy,
            )
        )
    return filas


def excel_gestor_bytes(db: Session, gestor_slug: str) -> Tuple[bytes, str, str]:
    slug = (gestor_slug or "").strip().lower()
    nombre = GESTOR_NOMBRES.get(slug)
    if not nombre:
        raise ValueError(f"Gestor desconocido: {gestor_slug}")
    filas = filas_gestor(db, slug)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera"
    _append_filas_cartera_excel(ws, filas)
    buf = io.BytesIO()
    wb.save(buf)
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", nombre).strip("_") or slug
    fname = f"gestor_{safe}_{hoy_negocio().isoformat()}.xlsx"
    return buf.getvalue(), fname, nombre


def excel_informe_diario_gestor_bytes(
    db: Session, gestor_slug: str
) -> Tuple[bytes, str, str]:
    """
    Informe Excel por gestor para Cobranza (gerente/admin):
    - Resumen_hoy: totales del dia (Caracas), se actualiza en cada descarga.
    - Por_dia: historial dia a dia (snapshot) con variacion vs dia anterior.
    - Cartera_hoy: prestamos vivos de la lista (misma logica que el Excel operativo).
    """
    slug = (gestor_slug or "").strip().lower()
    nombre = GESTOR_NOMBRES.get(slug)
    if not nombre:
        raise ValueError(f"Gestor desconocido: {gestor_slug}")

    asegurar_asignaciones(db)
    try:
        persistir_snapshot_diario(db)
    except Exception:
        db.rollback()
        logger.exception("[gestores] snapshot previo a informe diario %s", slug)

    hoy = hoy_negocio()
    filas = _filas_gestor_sin_asegurar(db, slug)
    total_cobranza = round(sum(f["total_cobranza_usd"] for f in filas), 2)
    total_cuotas = sum(int(f.get("cant_cuotas_vencidas") or 0) for f in filas)

    hist = (
        db.execute(
            select(CobranzaGestorDesempenoDiario)
            .where(CobranzaGestorDesempenoDiario.gestor_slug == slug)
            .order_by(CobranzaGestorDesempenoDiario.fecha.asc())
        )
        .scalars()
        .all()
    )

    wb = openpyxl.Workbook()

    # --- Resumen_hoy ---
    ws0 = wb.active
    ws0.title = "Resumen_hoy"
    ws0.append(["Informe diario gestores de cobranza"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.append(["Gestor", nombre])
    ws0.append(["Fecha negocio (Caracas)", hoy.isoformat()])
    ws0.append(["Casos en lista (APROBADO)", len(filas)])
    ws0.append(["Cuotas vencidas (total)", total_cuotas])
    ws0.append(["Monto vencido USD (total)", total_cobranza])
    ws0.append([])
    ws0.append(
        [
            "Nota",
            "Los montos se recalculan al descargar. La hoja Por_dia guarda el historial dia a dia.",
        ]
    )

    # --- Por_dia ---
    ws1 = wb.create_sheet("Por_dia")
    ws1.append(
        [
            "Fecha",
            "Total cobranza USD",
            "Cantidad casos",
            "Variacion USD vs dia anterior",
            "Variacion casos vs dia anterior",
        ]
    )
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    prev_usd: Optional[float] = None
    prev_n: Optional[int] = None
    for r in hist:
        usd = float(r.total_cobranza_usd or 0)
        n = int(r.cantidad_casos or 0)
        var_usd = "" if prev_usd is None else round(usd - prev_usd, 2)
        var_n = "" if prev_n is None else n - prev_n
        ws1.append(
            [
                r.fecha.isoformat() if r.fecha else "",
                round(usd, 2),
                n,
                var_usd,
                var_n,
            ]
        )
        prev_usd = usd
        prev_n = n

    # --- Cartera_hoy ---
    ws2 = wb.create_sheet("Cartera_hoy")
    _append_filas_cartera_excel(ws2, filas)

    buf = io.BytesIO()
    wb.save(buf)
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", nombre).strip("_") or slug
    fname = f"informe_diario_{safe}_{hoy.isoformat()}.xlsx"
    return buf.getvalue(), fname, nombre


def totales_vivos_por_gestor(
    db: Session, *, asegurar: bool = True
) -> List[Dict[str, Any]]:
    """Totales por gestor con una sola pasada batch (sin N+1)."""
    if asegurar:
        asegurar_asignaciones(db)
    hoy = hoy_negocio()
    asigs, prestamos, clientes, cuotas_by_pid = _cargar_asignaciones_vivas(db)

    agg: Dict[str, Dict[str, float]] = {
        s: {"n": 0.0, "usd": 0.0, "venc": 0.0, "mora": 0.0} for s in GESTOR_SLUGS
    }
    for asg in asigs:
        prestamo = prestamos.get(int(asg.prestamo_id))
        if not prestamo or not _prestamo_elegible_gestores(prestamo, hoy=hoy):
            continue
        cliente = clientes.get(int(prestamo.cliente_id)) if prestamo.cliente_id else None
        if not cliente:
            continue
        slug = (asg.gestor_slug or "").strip().lower()
        if slug not in agg:
            continue
        m = _metricas_cuotas_atraso(
            cuotas_by_pid.get(int(prestamo.id), []),
            hasta=hoy,
        )
        agg[slug]["n"] += 1
        agg[slug]["usd"] += m["carga_usd"]
        agg[slug]["venc"] += m["usd_vencidas"]
        agg[slug]["mora"] += m["usd_mora"]

    return [
        {
            "slug": slug,
            "nombre": nombre,
            "cantidad_casos": int(agg[slug]["n"]),
            "total_cobranza_usd": round(agg[slug]["usd"], 2),
            "usd_vencidas": round(agg[slug]["venc"], 2),
            "usd_mora": round(agg[slug]["mora"], 2),
        }
        for slug, nombre in GESTORES
    ]


def persistir_snapshot_diario(
    db: Session,
    *,
    fecha: Optional[date] = None,
    commit: bool = True,
    totales: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Guarda totales vivos del dia (Caracas) para la linea de tendencia."""
    dia = fecha or hoy_negocio()
    if totales is None:
        totales = totales_vivos_por_gestor(db)
    for t in totales:
        row = db.get(
            CobranzaGestorDesempenoDiario,
            {"fecha": dia, "gestor_slug": t["slug"]},
        )
        if row is None:
            row = CobranzaGestorDesempenoDiario(
                fecha=dia,
                gestor_slug=t["slug"],
                total_cobranza_usd=t["total_cobranza_usd"],
                cantidad_casos=t["cantidad_casos"],
            )
            db.add(row)
        else:
            row.total_cobranza_usd = t["total_cobranza_usd"]
            row.cantidad_casos = t["cantidad_casos"]
    if commit:
        db.commit()
    else:
        db.flush()
    return {"fecha": dia.isoformat(), "gestores": len(totales)}


def refrescar_desempeno_tras_pago(
    db: Session, prestamo_id: Optional[int]
) -> None:
    """
    Si el prestamo esta en una lista de gestores, actualiza el snapshot de hoy
    (todos los gestores) para que los graficos bajen al instante tras el pago.
    Nunca propaga excepciones al flujo de pagos.
    """
    if prestamo_id is None:
        return
    try:
        pid = int(prestamo_id)
    except (TypeError, ValueError):
        return
    if pid <= 0:
        return
    try:
        existe = db.scalar(
            select(CobranzaGestorAsignacion.id)
            .where(CobranzaGestorAsignacion.prestamo_id == pid)
            .limit(1)
        )
        if not existe:
            return
        persistir_snapshot_diario(db, commit=False)
    except Exception:
        logger.exception(
            "[gestores] refrescar_desempeno_tras_pago prestamo_id=%s",
            prestamo_id,
        )


def dashboard_gestores(db: Session) -> Dict[str, Any]:
    """
    Dashboard rapido: si la asignacion aun no cerro, la lanza en background
    y responde al instante (sin timeout). Con asignacion cerrada calcula totales
    en batch (sin auditoria pesada en cada poll).
    """
    en_progreso = False
    if not _asignacion_cerrada(db):
        _kick_asignacion_background()
        en_progreso = True
        # Tendencia historica si ya hubiera snapshots de intentos previos.
        rows = db.execute(
            select(CobranzaGestorDesempenoDiario)
            .order_by(
                CobranzaGestorDesempenoDiario.fecha.asc(),
                CobranzaGestorDesempenoDiario.gestor_slug.asc(),
            )
        ).scalars().all()
        by_fecha: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            key = r.fecha.isoformat()
            if key not in by_fecha:
                by_fecha[key] = {"fecha": key}
            by_fecha[key][r.gestor_slug] = float(r.total_cobranza_usd or 0)
        return {
            "gestores": listar_gestores(),
            "totales": _totales_vacios(),
            "tendencia": list(by_fecha.values()),
            "asignacion_cerrada": False,
            "asignacion_en_progreso": True,
            "fecha_inicio_cartera": FECHA_INICIO_APROBACION_GESTORES.isoformat(),
            "filtro": "fecha_aprobacion",
            "fecha_negocio": hoy_negocio().isoformat(),
        }

    asegurar_asignaciones(db)  # no-op rapido si ya cerrada
    totales = totales_vivos_por_gestor(db, asegurar=False)
    try:
        persistir_snapshot_diario(db, totales=totales)
    except Exception:
        db.rollback()
        logger.exception("[gestores] snapshot diario falló")
        totales = totales_vivos_por_gestor(db, asegurar=False)

    rows = db.execute(
        select(CobranzaGestorDesempenoDiario)
        .order_by(
            CobranzaGestorDesempenoDiario.fecha.asc(),
            CobranzaGestorDesempenoDiario.gestor_slug.asc(),
        )
    ).scalars().all()

    by_fecha = {}
    for r in rows:
        key = r.fecha.isoformat()
        if key not in by_fecha:
            by_fecha[key] = {"fecha": key}
        by_fecha[key][r.gestor_slug] = float(r.total_cobranza_usd or 0)

    return {
        "gestores": listar_gestores(),
        "totales": totales,
        "tendencia": list(by_fecha.values()),
        "asignacion_cerrada": True,
        "asignacion_en_progreso": en_progreso or asignacion_en_progreso(),
        "fecha_inicio_cartera": FECHA_INICIO_APROBACION_GESTORES.isoformat(),
        "filtro": "fecha_aprobacion",
        "fecha_negocio": hoy_negocio().isoformat(),
    }


def enviar_listas_gestores_email(db: Session) -> Dict[str, Any]:
    """
    Regenera las 9 listas Excel al momento (sin liquidados) y las envia.
    To: operaciones@  BCC: itmaster@
    Cada Excel contiene prestamos completos (nunca un prestamo partido entre gestores).
    """
    from app.core.email import send_email

    asegurar_asignaciones(db, verificar_integridad=True)
    integridad = _garantizar_integridad_listas(db)
    persistir_snapshot_diario(db)

    # Adjuntos = listas recalculadas justo antes del envio.
    attachments: List[Tuple[str, bytes]] = []
    for slug, _nombre in GESTORES:
        data, fname, _ = excel_gestor_bytes(db, slug)
        attachments.append((fname, data))

    hoy = hoy_negocio().isoformat()
    asunto = f"Listas actualizadas {hoy}"
    cuerpo = "Eduardo: Adjunto listas actualizadas"
    ok, err = send_email(
        [EMAIL_GESTORES_TO],
        asunto,
        cuerpo,
        bcc_emails=[EMAIL_GESTORES_BCC],
        attachments=attachments,
        servicio="notificaciones",
        tipo_tab="cobranza_gestores",
        aplicar_cco_automatica=False,
    )
    if not ok:
        logger.error("[gestores] fallo email listas: %s", err)
        return {"ok": False, "error": err, "adjuntos": len(attachments)}
    logger.info(
        "[gestores] email listas enviado to=%s bcc=%s adjuntos=%s asunto=%s asignaciones=%s",
        EMAIL_GESTORES_TO,
        EMAIL_GESTORES_BCC,
        len(attachments),
        asunto,
        integridad.get("total_asignaciones"),
    )
    return {
        "ok": True,
        "adjuntos": len(attachments),
        "to": EMAIL_GESTORES_TO,
        "bcc": EMAIL_GESTORES_BCC,
        "asunto": asunto,
        "prestamos_asignados": integridad.get("total_asignaciones"),
        "integridad_ok": True,
    }
