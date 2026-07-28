"""Conciliacion Bancos: carga Excel, match vs numero_documento, decision y aplicacion segura."""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from difflib import SequenceMatcher
from typing import Any, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, or_, select, text
from sqlalchemy.exc import IntegrityError, ProgrammingError, SQLAlchemyError
from sqlalchemy.orm import Session

from app.core.documento import normalize_documento, split_numero_documento_almacenado
from app.models.conciliacion_banco_ocr import (
    ConciliacionBancoExtracto,
    ConciliacionBancoOcrBanco,
    ConciliacionBancoOcrLote,
    ConciliacionBancoOcrResultado,
)
from app.models.pago import Pago
from app.services.cuota_pago_integridad import pago_tiene_aplicaciones_cuotas
from app.services.pago_numero_documento import numero_documento_ya_registrado
from app.services.tasa_cambio_service import (
    obtener_tasa_por_fecha,
    tasa_y_equivalente_usd_excel,
)

logger = logging.getLogger(__name__)

SIMILITUD_MINIMA = 70.0
MONTO_TOL = 0.02
# Capacidad de un lote Excel (extracto banco): hasta 25_000 filas de datos.
MAX_FILAS_EXCEL_LOTE = 25000
BULK_INSERT_CHUNK = 2000


def _clave_paquete_fecha_monto(
    fecha: Optional[date], monto: Optional[float]
) -> Optional[tuple[date, float]]:
    """Clave de indice para match parcial (paquete fecha+monto a 2 decimales)."""
    if fecha is None or monto is None:
        return None
    return (fecha, round(float(monto), 2))


def _candidatos_payload(pagos: list[Pago]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for p in pagos:
        out.append(
            {
                "pago_id": int(p.id),
                "cedula": p.cedula_cliente,
                "prestamo_id": int(p.prestamo_id) if p.prestamo_id is not None else None,
                "monto": float(p.monto_pagado) if p.monto_pagado is not None else None,
                "institucion_categoria": categoria_pago_conciliacion(p),
                "referencia_bd": p.numero_documento,
            }
        )
    return out


def _detalle_ambiguo_serial(pagos: list[Pago]) -> str:
    cats = {categoria_pago_conciliacion(p) for p in pagos}
    mercantil = cats == {"Mercantil"} or "Mercantil" in cats
    lines = []
    for p in pagos:
        ced = p.cedula_cliente or "?"
        pre = f"#{p.prestamo_id}" if p.prestamo_id is not None else "#?"
        lines.append(f"{ced} {pre} (pago {p.id})")
    base = (
        f"Mismo serial en {len(pagos)} pagos"
        + (" (patron tipico Mercantil)" if mercantil else "")
        + ". Discernimiento manual."
    )
    return base + " Candidatos: " + "; ".join(lines)



BANCOS_CATEGORIAS = ("Mercantil", "BNC", "Binance", "BNV", "Recibos", "Drive", "Otros")



# Asientos sinteticos Drive / Notificaciones (a menudo suma de pagos)
_PREFIJOS_DRIVE_ABONOS = (
    "ABONOS-NOTIF-",
    "ABONOS-DRIVE-",
)


def _es_asiento_drive_abonos(numero_documento: Optional[str]) -> bool:
    """True si es referencia sintetica tipo ABONOS-NOTIF-911-B5 (Drive/Notificaciones)."""
    s = (numero_documento or "").strip().upper()
    if not s:
        return False
    return any(s.startswith(pref) for pref in _PREFIJOS_DRIVE_ABONOS)



def _clave_caso_banco(
    referencia: Optional[str],
    fecha_banco: Optional[date],
    monto_usd: Optional[float],
) -> str:
    """
    Identifica un movimiento banco ya gestionado (Visto/confirmado):
    serial digitos + fecha + monto. Asi BNV puede reusar el mismo serial
    en otra fecha/monto y ese otro caso SI aparece.
    """
    dig = _ref_solo_digitos(referencia)
    if not dig or fecha_banco is None or monto_usd is None:
        return ""
    return f"{dig}|{fecha_banco.isoformat()}|{round(float(monto_usd), 2):.2f}"


def _paquete_banco_coherente_con_pago(
    pago: Pago,
    *,
    fecha_banco: Optional[date],
    monto_usd: Optional[float],
) -> bool:
    """
    Coincidencia exacta de paquete: monto Y fecha (ademas del serial fuera).
    Sin fecha o monto en banco no se considera match exacto/parcial firme.
    """
    if monto_usd is None or fecha_banco is None:
        return False
    if abs(float(pago.monto_pagado or 0) - float(monto_usd)) > MONTO_TOL:
        return False
    fd = _pago_fecha(pago)
    if fd is None or fd != fecha_banco:
        return False
    return True


def categoria_institucion(inst: Optional[str]) -> str:
    """Misma clasificacion que dashboard pagos por institucion."""
    s = (inst or "").strip().lower()
    if "mercantil" in s:
        return "Mercantil"
    if "bnc" in s or s == "banco nacional de credito":
        return "BNC"
    if "binance" in s:
        return "Binance"
    if "bnv" in s or "bdv" in s or "banco de venezuela" in s:
        return "BNV"
    if "recibo" in s:
        return "Recibos"
    if "drive" in s:
        return "Drive"
    return "Otros"


def categoria_pago_conciliacion(pago: Pago) -> str:
    """Categoria para filtro Conciliacion Bancos (Drive = ABONOS-NOTIF/DRIVE)."""
    if _es_asiento_drive_abonos(getattr(pago, "numero_documento", None)):
        return "Drive"
    return categoria_institucion(getattr(pago, "institucion_bancaria", None))


def pago_ids_conciliacion_bancaria_confirmada(
    db: Session, pago_ids: list[int]
) -> set[int]:
    """IDs con confirmacion en Conciliacion Bancos (CORREGIR + aplicado)."""
    if not pago_ids:
        return set()
    q = (
        select(ConciliacionBancoOcrResultado.pago_id)
        .where(
            ConciliacionBancoOcrResultado.pago_id.in_(pago_ids),
            ConciliacionBancoOcrResultado.decision == "CORREGIR",
            ConciliacionBancoOcrResultado.aplicado.is_(True),
        )
        .distinct()
    )
    return {int(x[0]) for x in db.execute(q).all() if x[0] is not None}


def pago_ids_ambiguo_bancario_multi(db: Session, pago_ids: list[int]) -> set[int]:
    """Solo pagos confirmados con tipo_novedad AMBIGUO (CORREGIR+aplicado)."""
    if not pago_ids:
        return set()
    ids = sorted({int(x) for x in pago_ids if x is not None})
    if not ids:
        return set()
    rows = (
        db.execute(
            select(ConciliacionBancoOcrResultado.pago_id).where(
                ConciliacionBancoOcrResultado.pago_id.in_(ids),
                ConciliacionBancoOcrResultado.tipo_novedad == "AMBIGUO",
                ConciliacionBancoOcrResultado.decision == "CORREGIR",
                ConciliacionBancoOcrResultado.aplicado.is_(True),
            )
        )
        .all()
    )
    return {int(r[0]) for r in rows if r[0] is not None}


def contar_conciliacion_bancaria_prestamo(db: Session, prestamo_id: int) -> int:
    """Cuantos pagos del prestamo tienen confirmacion bancaria."""
    q = (
        select(func.count(func.distinct(ConciliacionBancoOcrResultado.pago_id)))
        .select_from(ConciliacionBancoOcrResultado)
        .join(Pago, Pago.id == ConciliacionBancoOcrResultado.pago_id)
        .where(
            Pago.prestamo_id == int(prestamo_id),
            ConciliacionBancoOcrResultado.decision == "CORREGIR",
            ConciliacionBancoOcrResultado.aplicado.is_(True),
        )
    )
    return int(db.scalar(q) or 0)


def normalizar_bancos_filtro(bancos: Optional[list[str]]) -> list[str]:
    allowed = set(BANCOS_CATEGORIAS)
    out: list[str] = []
    for b in bancos or []:
        name = (b or "").strip()
        if name in allowed and name not in out:
            out.append(name)
    return out


def _guardar_bancos_en_lote(lote: ConciliacionBancoOcrLote, bancos: list[str]) -> None:
    payload: dict[str, Any] = {}
    raw = (lote.notas or "").strip()
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                payload = data
        except Exception:
            payload = {}
    payload["bancos_filtro"] = bancos
    lote.notas = json.dumps(payload, ensure_ascii=True)


def _leer_bancos_de_lote(lote: ConciliacionBancoOcrLote) -> list[str]:
    raw = (lote.notas or "").strip()
    if not raw:
        return []
    try:
        data = json.loads(raw)
        if isinstance(data, dict):
            return normalizar_bancos_filtro(data.get("bancos_filtro") or [])
    except Exception:
        pass
    return []


def _institucion_objetivo_desde_lote(lote: ConciliacionBancoOcrLote) -> Optional[str]:
    """
    Banco del extracto (filtro del lote) para pagos.institucion_bancaria con Ref. Banco.
    Ignora Otros/Recibos/Drive al desambiguar: filtro BNV+Recibos+Otros -> BNV.
    """
    bancos = _leer_bancos_de_lote(lote)
    if not bancos:
        return None
    # Bancos de extracto Excel (no auxiliares de busqueda)
    extracto = ("Mercantil", "BNC", "Binance", "BNV")
    candidatos = [b for b in bancos if b in extracto]
    if len(candidatos) == 1:
        return candidatos[0]
    if len(bancos) == 1:
        cat = bancos[0]
        if cat == "Otros":
            return None
        if cat == "Recibos":
            return "Recibo"
        if cat == "Drive":
            return "Drive"
        return cat
    return None


def _aplicar_institucion_desde_lote(pago: Pago, lote: ConciliacionBancoOcrLote) -> bool:
    """True si cambio institucion_bancaria del pago."""
    if _es_asiento_drive_abonos(getattr(pago, "numero_documento", None)):
        return False
    objetivo = _institucion_objetivo_desde_lote(lote)
    if not objetivo:
        return False
    cat_actual = categoria_institucion(getattr(pago, "institucion_bancaria", None))
    cat_obj = categoria_institucion(objetivo)
    if cat_actual == cat_obj:
        return False
    pago.institucion_bancaria = objetivo
    return True


# Prefijos/etiquetas frecuentes que operadores agregan al serial
# Extension Visto (control 5 / carga masiva): _A#### o _P#### al final.
# No debe entrar en la clave de match: si no, el mismo serial en varios prestamos
# no colapsa a AMBIGUO (los 4 digitos del sufijo ensucian la clave).
_SUFIJO_VISTO_DOC_RE = re.compile(r"_[AP]\d{4}$", re.IGNORECASE)

_REF_RUIDO_PREFIX = re.compile(
    r"^(?:"
    r"(?:bs\.?\s*)?(?:bnc|binance|mercantil|bnv|bdv|ve|zelle|paypal|banco)\s*"
    r"(?:/\s*|[-ΓÇôΓÇö]\s*|\s+)"
    r"(?:ref\.?\s*)?"
    r"|ref\.?\s*|nro\.?\s*|n[u├║]m(?:ero)?\.?\s*|doc\.?\s*|comp(?:robante)?\.?\s*"
    r")+",
    re.IGNORECASE,
)


def _ref_solo_digitos(val: Optional[str]) -> str:
    """
    Clave de match/similitud: solo digitos del comprobante.

    - Ignora letras y signos agregados por digitacion (REF-, BNC/, puntos, guiones, etc.).
    - Quita extension Visto _A####/_P#### y sufijo interno ┬ºCD: (codigo desambiguador) para no contaminar la clave.
    - Maneja notacion cientifica via normalize_documento.
    - Quita ceros a la izquierda (BD/OCR a veces guarda 00000019197881 == 19197881).
    """
    if val is None or val == "":
        return ""
    base, _codigo = split_numero_documento_almacenado(val)
    s = normalize_documento(base) or (base or str(val)).strip()
    # Quitar extension Visto _A####/_P#### (aprobacion manual); AMBIGUO por serial real.
    s = _SUFIJO_VISTO_DOC_RE.sub("", s).strip() or s
    if not s:
        return ""
    # Quitar etiquetas al inicio en bucle (BNC/ REF. ...)
    prev = None
    while prev != s:
        prev = s
        s2 = _REF_RUIDO_PREFIX.sub("", s).strip()
        s = s2 if s2 else s
        break  # una pasada basta; el bucle evita bucles raros si regex vacia
    # Solo digitos: letras/signos no participan en similitud
    digitos = re.sub(r"\D+", "", s)
    if not digitos:
        return ""
    # Normalizar padding de ceros (Rapi/OCR): 00000019197881 -> 19197881
    sin_ceros = digitos.lstrip("0")
    return sin_ceros if sin_ceros else "0"


def _similitud(a: str, b: str) -> float:
    da, db = _ref_solo_digitos(a), _ref_solo_digitos(b)
    if not da or not db:
        return 0.0
    if da == db:
        return 100.0
    # Contencion: "90694665" vs "00090694665" / extras de digitacion
    shorter, longer = (da, db) if len(da) <= len(db) else (db, da)
    if len(shorter) >= 6 and shorter in longer:
        return round(100.0 * (len(shorter) / len(longer)), 2)
    return round(SequenceMatcher(None, da, db).ratio() * 100.0, 2)



def _normalizar_banco_categoria(val: Any) -> Optional[str]:
    """Mapea texto Excel a categoria conocida (Mercantil, BNC, ...)."""
    s = str(val or "").strip()
    if not s:
        return None
    low = s.casefold()
    for b in BANCOS_CATEGORIAS:
        if low == b.casefold():
            return b
    return None


def _clave_natural_extracto(
    *,
    fecha: Optional[date],
    referencia_norm: str,
    monto: Optional[Decimal],
) -> str:
    """Unicidad BD historica: serial (referencia_norm) + fecha + monto."""
    ref = (referencia_norm or "").strip()
    f = fecha.isoformat() if fecha else ""
    m = f"{float(monto):.2f}" if monto is not None else ""
    return f"{ref}|{f}|{m}"


def _ensure_tabla_extracto(db: Session) -> None:
    """Crea conciliacion_banco_extracto e indices si no existen (prod/DBeaver)."""
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS conciliacion_banco_extracto (
                id              SERIAL PRIMARY KEY,
                banco           VARCHAR(40) NOT NULL,
                fecha           DATE NULL,
                referencia      TEXT NOT NULL,
                referencia_norm TEXT NULL,
                monto           NUMERIC(14, 2) NULL,
                moneda          VARCHAR(3) NOT NULL DEFAULT 'USD',
                clave_natural   TEXT NOT NULL,
                lote_origen_id  INTEGER NULL
                    REFERENCES conciliacion_banco_ocr_lote(id) ON DELETE SET NULL,
                archivo_nombre  VARCHAR(255) NULL,
                creado_en       TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                actualizado_en  TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_conciliacion_banco_extracto_clave "
            "ON conciliacion_banco_extracto (clave_natural)"
        )
    )
    db.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_conciliacion_banco_extracto_banco "
            "ON conciliacion_banco_extracto (banco)"
        )
    )
    db.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_conciliacion_banco_extracto_fecha "
            "ON conciliacion_banco_extracto (fecha)"
        )
    )
    db.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_conciliacion_banco_extracto_referencia_norm "
            "ON conciliacion_banco_extracto (referencia_norm)"
        )
    )
    db.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_conciliacion_banco_extracto_lote_origen "
            "ON conciliacion_banco_extracto (lote_origen_id)"
        )
    )
    _migrar_claves_extracto_serial_fecha_monto(db)
    db.flush()


def _migrar_claves_extracto_serial_fecha_monto(db: Session) -> None:
    """Reclava a serial|fecha|monto y elimina duplicados entre versiones de Excel."""
    db.execute(
        text(
            """
            UPDATE conciliacion_banco_extracto
            SET clave_natural =
                COALESCE(NULLIF(TRIM(referencia_norm), ''), TRIM(referencia), '')
                || '|' || COALESCE(to_char(fecha, 'YYYY-MM-DD'), '')
                || '|' || CASE
                    WHEN monto IS NULL THEN ''
                    ELSE TRIM(to_char(monto, 'FM999999999990.00'))
                END
            WHERE clave_natural IS DISTINCT FROM (
                COALESCE(NULLIF(TRIM(referencia_norm), ''), TRIM(referencia), '')
                || '|' || COALESCE(to_char(fecha, 'YYYY-MM-DD'), '')
                || '|' || CASE
                    WHEN monto IS NULL THEN ''
                    ELSE TRIM(to_char(monto, 'FM999999999990.00'))
                END
            )
            """
        )
    )
    db.execute(
        text(
            """
            DELETE FROM conciliacion_banco_extracto a
            USING conciliacion_banco_extracto b
            WHERE a.clave_natural = b.clave_natural
              AND a.id < b.id
            """
        )
    )


def _upsert_extracto_filas(
    db: Session,
    filas: list[dict[str, Any]],
    *,
    lote_id: int,
    archivo_nombre: Optional[str],
) -> int:
    """
    Inserta/actualiza en conciliacion_banco_extracto (BD historica).
    Usa INSERT ... ON CONFLICT para lotes grandes (~25k).
    """
    if not filas:
        return 0
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    _ensure_tabla_extracto(db)
    now = datetime.utcnow()
    by_clave: dict[str, dict[str, Any]] = {}
    for r in filas:
        by_clave[str(r["clave_natural"])] = r
    unicos = list(by_clave.values())
    chunk = 1000
    for i in range(0, len(unicos), chunk):
        batch = unicos[i : i + chunk]
        rows = [
            {
                "banco": r["banco"],
                "fecha": r["fecha"],
                "referencia": r["referencia"],
                "referencia_norm": r["referencia_norm"],
                "monto": r["monto"],
                "moneda": r["moneda"],
                "clave_natural": r["clave_natural"],
                "lote_origen_id": lote_id,
                "archivo_nombre": (archivo_nombre or None),
                "creado_en": now,
                "actualizado_en": now,
            }
            for r in batch
        ]
        stmt = pg_insert(ConciliacionBancoExtracto).values(rows)
        stmt = stmt.on_conflict_do_update(
            index_elements=["clave_natural"],
            set_={
                "banco": stmt.excluded.banco,
                "fecha": stmt.excluded.fecha,
                "referencia": stmt.excluded.referencia,
                "referencia_norm": stmt.excluded.referencia_norm,
                "monto": stmt.excluded.monto,
                "moneda": stmt.excluded.moneda,
                "lote_origen_id": stmt.excluded.lote_origen_id,
                "archivo_nombre": stmt.excluded.archivo_nombre,
                "actualizado_en": stmt.excluded.actualizado_en,
            },
        )
        db.execute(stmt)
        db.flush()
    return len(unicos)


def resumen_extracto_historico(
    db: Session,
    *,
    bancos: Optional[list[str]] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    moneda: Optional[str] = None,
) -> dict[str, Any]:
    """Resumen BD historica por variable banco: cantidad de filas + monto total."""
    q = select(
        ConciliacionBancoExtracto.banco,
        func.count(),
        func.coalesce(func.sum(ConciliacionBancoExtracto.monto), 0),
        func.min(ConciliacionBancoExtracto.fecha),
        func.max(ConciliacionBancoExtracto.fecha),
    )
    if bancos:
        cats = [_normalizar_banco_categoria(b) or b for b in bancos]
        cats = [c for c in cats if c]
        if cats:
            q = q.where(ConciliacionBancoExtracto.banco.in_(cats))
    if fecha_desde is not None:
        q = q.where(
            or_(
                ConciliacionBancoExtracto.fecha.is_(None),
                ConciliacionBancoExtracto.fecha >= fecha_desde,
            )
        )
    if fecha_hasta is not None:
        q = q.where(
            or_(
                ConciliacionBancoExtracto.fecha.is_(None),
                ConciliacionBancoExtracto.fecha <= fecha_hasta,
            )
        )
    if moneda:
        q = q.where(ConciliacionBancoExtracto.moneda == moneda.strip().upper())
    q = q.group_by(ConciliacionBancoExtracto.banco).order_by(
        ConciliacionBancoExtracto.banco.asc()
    )
    por_banco = []
    total = 0
    monto_total = 0.0
    for banco, n, monto_sum, fmin, fmax in db.execute(q).all():
        nn = int(n or 0)
        mm = float(monto_sum or 0)
        total += nn
        monto_total += mm
        por_banco.append(
            {
                "banco": banco,
                "filas": nn,
                "monto_total": round(mm, 2),
                "fecha_min": fmin.isoformat() if fmin else None,
                "fecha_max": fmax.isoformat() if fmax else None,
            }
        )
    for row in por_banco:
        row["pct_filas"] = (
            round(100.0 * row["filas"] / total, 2) if total else 0.0
        )
        row["pct_monto"] = (
            round(100.0 * row["monto_total"] / monto_total, 2)
            if monto_total
            else 0.0
        )
    return {
        "ok": True,
        "total": total,
        "monto_total": round(monto_total, 2),
        "bancos": len(por_banco),
        "por_banco": por_banco,
    }



def listar_lotes_recientes(db: Session, *, limit: int = 40) -> list[dict[str, Any]]:
    lim = max(1, min(int(limit or 40), 100))
    rows = (
        db.execute(
            select(ConciliacionBancoOcrLote)
            .order_by(ConciliacionBancoOcrLote.id.desc())
            .limit(lim)
        )
        .scalars()
        .all()
    )
    out: list[dict[str, Any]] = []
    for lote in rows:
        sin_bd = int(
            db.scalar(
                select(func.count())
                .select_from(ConciliacionBancoOcrResultado)
                .where(
                    ConciliacionBancoOcrResultado.lote_id == int(lote.id),
                    ConciliacionBancoOcrResultado.tipo_novedad == "SIN_BD",
                    ConciliacionBancoOcrResultado.decision == "PENDIENTE",
                )
            )
            or 0
        )
        out.append(
            {
                "id": int(lote.id),
                "archivo_nombre": lote.archivo_nombre,
                "estado": lote.estado,
                "fecha_desde": lote.fecha_desde.isoformat() if lote.fecha_desde else None,
                "fecha_hasta": lote.fecha_hasta.isoformat() if lote.fecha_hasta else None,
                "creado_en": lote.creado_en.isoformat() if lote.creado_en else None,
                "bancos_filtro": _leer_bancos_de_lote(lote),
                "sin_bd": sin_bd,
            }
        )
    return out


def resumen_sin_bd_por_banco(
    db: Session,
    *,
    lote_id: Optional[int] = None,
) -> dict[str, Any]:
    # Dashboard SIN_BD: sin match en pagos, clasificado por variable Banco.
    lote: Optional[ConciliacionBancoOcrLote] = None
    if lote_id is not None:
        lote = db.get(ConciliacionBancoOcrLote, int(lote_id))
    else:
        lote = (
            db.execute(
                select(ConciliacionBancoOcrLote)
                .where(ConciliacionBancoOcrLote.estado == "COMPARADO")
                .order_by(ConciliacionBancoOcrLote.id.desc())
                .limit(1)
            )
            .scalars()
            .first()
        )
    if lote is None:
        return {
            "ok": True,
            "tipo": "SIN_BD",
            "lote_id": None,
            "total": 0,
            "monto_total": 0.0,
            "bancos": 0,
            "por_banco": [],
            "message": "No hay lote COMPARADO con SIN_BD",
        }

    lid = int(lote.id)
    bancos_filtro = _leer_bancos_de_lote(lote)
    banco_default = bancos_filtro[0] if bancos_filtro else "Otros"

    rows = db.execute(
        select(ConciliacionBancoOcrResultado, ConciliacionBancoOcrBanco)
        .outerjoin(
            ConciliacionBancoOcrBanco,
            ConciliacionBancoOcrBanco.id == ConciliacionBancoOcrResultado.banco_id,
        )
        .where(
            ConciliacionBancoOcrResultado.lote_id == lid,
            ConciliacionBancoOcrResultado.tipo_novedad == "SIN_BD",
            ConciliacionBancoOcrResultado.decision == "PENDIENTE",
        )
    ).all()

    claves: list[str] = []
    meta: list[tuple[Optional[ConciliacionBancoOcrBanco], ConciliacionBancoOcrResultado, str]] = []
    for res, banco_row in rows:
        ref_norm = ""
        fecha_b = res.fecha_banco
        monto_clave = None
        if banco_row is not None:
            ref_norm = (
                (banco_row.ref_banco_norm or "").strip()
                or (banco_row.referencia_banco or "").strip()
            )
            fecha_b = banco_row.fecha_banco or res.fecha_banco
            monto_clave = (
                banco_row.monto_banco_original
                if banco_row.monto_banco_original is not None
                else banco_row.monto_banco
            )
        else:
            ref_norm = (res.referencia_banco or "").strip()
            monto_clave = res.monto_banco
        if not ref_norm:
            ref_norm = normalize_documento(res.referencia_banco or "") or ""
        if monto_clave is not None and not isinstance(monto_clave, Decimal):
            try:
                monto_clave = Decimal(str(monto_clave))
            except Exception:
                monto_clave = None
        clave = _clave_natural_extracto(
            fecha=fecha_b,
            referencia_norm=ref_norm,
            monto=monto_clave,
        )
        claves.append(clave)
        meta.append((banco_row, res, clave))

    extracto_banco: dict[str, str] = {}
    uniq = sorted({c for c in claves if c})
    for i in range(0, len(uniq), 2000):
        chunk = uniq[i : i + 2000]
        for er in db.execute(
            select(
                ConciliacionBancoExtracto.clave_natural,
                ConciliacionBancoExtracto.banco,
            ).where(ConciliacionBancoExtracto.clave_natural.in_(chunk))
        ).all():
            if er[0] and er[1]:
                extracto_banco[str(er[0])] = str(er[1])

    agg: dict[str, dict[str, Any]] = {}
    for _banco_row, res, clave in meta:
        banco = extracto_banco.get(clave) or banco_default or "Otros"
        slot = agg.setdefault(
            banco,
            {
                "banco": banco,
                "filas": 0,
                "monto_total": 0.0,
                "fecha_min": None,
                "fecha_max": None,
            },
        )
        slot["filas"] += 1
        if res.monto_banco is not None:
            slot["monto_total"] += float(res.monto_banco)
        fd = res.fecha_banco
        if fd is not None:
            if slot["fecha_min"] is None or fd < slot["fecha_min"]:
                slot["fecha_min"] = fd
            if slot["fecha_max"] is None or fd > slot["fecha_max"]:
                slot["fecha_max"] = fd

    por_banco = sorted(agg.values(), key=lambda x: (-int(x["filas"]), str(x["banco"])))
    total = sum(int(x["filas"]) for x in por_banco)
    monto_total = sum(float(x["monto_total"]) for x in por_banco)
    out_rows = []
    for x in por_banco:
        mm = round(float(x["monto_total"]), 2)
        nn = int(x["filas"])
        out_rows.append(
            {
                "banco": x["banco"],
                "filas": nn,
                "monto_total": mm,
                "pct_filas": round(100.0 * nn / total, 2) if total else 0.0,
                "pct_monto": round(100.0 * mm / monto_total, 2) if monto_total else 0.0,
                "fecha_min": x["fecha_min"].isoformat() if x["fecha_min"] else None,
                "fecha_max": x["fecha_max"].isoformat() if x["fecha_max"] else None,
            }
        )
    return {
        "ok": True,
        "tipo": "SIN_BD",
        "lote_id": lid,
        "archivo_nombre": lote.archivo_nombre,
        "estado": lote.estado,
        "bancos_filtro": bancos_filtro,
        "total": total,
        "monto_total": round(monto_total, 2),
        "bancos": len(out_rows),
        "por_banco": out_rows,
    }


def _parse_fecha(val: Any) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_monto(val: Any) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(round(float(val), 2)))
    s = str(val).strip().replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return Decimal(str(round(float(s), 2)))
    except Exception:
        return None


def _pago_fecha(p: Pago) -> Optional[date]:
    fp = p.fecha_pago
    if fp is None:
        return None
    return fp.date() if hasattr(fp, "date") else fp


def _snapshot_pago(p: Pago) -> dict[str, Any]:
    return {
        "pago_id": p.id,
        "fecha_pago": _pago_fecha(p).isoformat() if _pago_fecha(p) else None,
        "monto_pagado": float(p.monto_pagado or 0),
        "numero_documento": p.numero_documento,
        "institucion_bancaria": p.institucion_bancaria,
        "institucion_categoria": categoria_pago_conciliacion(p),
        "conciliado": bool(p.conciliado),
        "verificado_concordancia": p.verificado_concordancia,
        "moneda_registro": p.moneda_registro,
        "monto_bs_original": float(p.monto_bs_original) if p.monto_bs_original is not None else None,
        "tasa_cambio_bs_usd": float(p.tasa_cambio_bs_usd) if p.tasa_cambio_bs_usd is not None else None,
        "fecha_tasa_referencia": p.fecha_tasa_referencia.isoformat() if p.fecha_tasa_referencia else None,
    }



def buscar_serial_en_extracto(
    db: Session,
    *,
    serial: str,
    moneda: Optional[str] = None,
) -> dict[str, Any]:
    """Busca serial/referencia en BD historica. Indica tambien si hay pago en pagos."""
    raw = (serial or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Indique el serial / referencia")
    norm = normalize_documento(raw) or raw
    dig = _ref_solo_digitos(norm) or _ref_solo_digitos(raw)
    if not dig and not norm:
        raise HTTPException(status_code=400, detail="Serial invalido")

    conds = [
        ConciliacionBancoExtracto.referencia == raw,
        ConciliacionBancoExtracto.referencia == norm,
    ]
    if dig:
        conds.extend(
            [
                ConciliacionBancoExtracto.referencia_norm == dig,
                ConciliacionBancoExtracto.referencia_norm == norm,
                ConciliacionBancoExtracto.referencia == dig,
            ]
        )
    q = select(ConciliacionBancoExtracto).where(or_(*conds))
    if moneda:
        q = q.where(ConciliacionBancoExtracto.moneda == moneda.strip().upper())
    filas = list(
        db.execute(q.order_by(ConciliacionBancoExtracto.id.asc())).scalars().all()
    )

    n_pagos = 0
    pago_conds = [
        Pago.numero_documento == raw,
        Pago.numero_documento == norm,
    ]
    if dig:
        pago_conds.extend(
            [
                Pago.numero_documento == dig,
                Pago.numero_documento.like(f"%{dig}%"),
            ]
        )
    n_pagos = int(
        db.scalar(
            select(func.count()).select_from(Pago).where(or_(*pago_conds))
        )
        or 0
    )

    items = [
        {
            "id": int(r.id),
            "banco": r.banco,
            "fecha": r.fecha.isoformat() if r.fecha else None,
            "referencia": r.referencia,
            "referencia_norm": r.referencia_norm,
            "monto": float(r.monto) if r.monto is not None else None,
            "moneda": r.moneda,
        }
        for r in filas
    ]
    pendientes = _filtrar_extracto_sin_cerrados(db, filas)
    n_cerradas = len(filas) - len(pendientes)
    items_pend = [
        {
            "id": int(r.id),
            "banco": r.banco,
            "fecha": r.fecha.isoformat() if r.fecha else None,
            "referencia": r.referencia,
            "referencia_norm": r.referencia_norm,
            "monto": float(r.monto) if r.monto is not None else None,
            "moneda": r.moneda,
        }
        for r in pendientes
    ]
    return {
        "ok": True,
        "encontrado": len(items) > 0,
        "serial": raw,
        "serial_norm": dig or norm,
        "en_extracto": len(items) > 0,
        "filas_extracto": len(items),
        "filas_pendientes": len(items_pend),
        "filas_ya_cerradas": n_cerradas,
        "ya_visto_o_conciliado": len(items) > 0 and len(items_pend) == 0,
        "items": items_pend,
        "en_pagos": n_pagos > 0,
        "pagos_count": n_pagos,
    }


def crear_lote_desde_serial(
    db: Session,
    *,
    serial: str,
    moneda_carga: str,
    usuario_id: Optional[int],
    bancos: Optional[list[str]] = None,
) -> ConciliacionBancoOcrLote:
    """Si el serial esta en BD historica, crea lote listo para conciliar."""
    mon = (moneda_carga or "USD").strip().upper()
    if mon not in ("USD", "BS"):
        raise HTTPException(status_code=400, detail="moneda_carga debe ser USD o BS")
    info = buscar_serial_en_extracto(db, serial=serial, moneda=mon)
    if not info["encontrado"]:
        raise HTTPException(
            status_code=404,
            detail=f"Serial no esta en BD historica: {serial.strip()}",
        )
    ids = [int(x["id"]) for x in info["items"]]
    rows = list(
        db.execute(
            select(ConciliacionBancoExtracto)
            .where(ConciliacionBancoExtracto.id.in_(ids))
            .order_by(ConciliacionBancoExtracto.id.asc())
        )
        .scalars()
        .all()
    )
    rows = _filtrar_extracto_sin_cerrados(db, rows)
    if not rows:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Serial {serial.strip()} ya esta VISTO/conciliado; "
                "no se vuelve a cargar."
            ),
        )
    cats: list[str] = []
    if bancos:
        for b in bancos:
            nb = _normalizar_banco_categoria(b)
            if nb and nb not in cats:
                cats.append(nb)
        if cats:
            rows = [r for r in rows if r.banco in cats]
            if not rows:
                raise HTTPException(
                    status_code=404,
                    detail="Serial en historica pero no en los bancos del filtro",
                )
    else:
        for r in rows:
            if r.banco and r.banco not in cats:
                cats.append(r.banco)

    fechas = [r.fecha for r in rows if r.fecha is not None]
    f_desde = min(fechas) if fechas else date.today()
    f_hasta = max(fechas) if fechas else date.today()
    dig = info.get("serial_norm") or serial.strip()
    lote = ConciliacionBancoOcrLote(
        usuario_id=usuario_id,
        archivo_nombre=f"SERIAL_{dig}"[:255],
        fecha_desde=f_desde,
        fecha_hasta=f_hasta,
        estado="CARGADO",
        moneda_carga=mon,
    )
    db.add(lote)
    db.flush()
    batch = [
        {
            "lote_id": lote.id,
            "fila_excel": i,
            "fecha_banco": r.fecha,
            "referencia_banco": r.referencia,
            "ref_banco_norm": r.referencia_norm or r.referencia,
            "monto_banco": r.monto if mon == "USD" else None,
            "monto_banco_original": r.monto,
            "moneda_fila": mon,
        }
        for i, r in enumerate(rows, start=1)
    ]
    db.bulk_insert_mappings(ConciliacionBancoOcrBanco, batch)
    if cats:
        _guardar_bancos_en_lote(lote, cats)
    lote.notas = json.dumps(
        {
            "filas_banco": len(rows),
            "fuente_carga": "serial",
            "serial": serial.strip(),
            "bancos": cats,
            "en_pagos": info.get("en_pagos"),
            "pagos_count": info.get("pagos_count"),
        },
        ensure_ascii=True,
    )
    db.commit()
    db.refresh(lote)
    return lote

def _casos_banco_cerrados_globales(db: Session) -> set[str]:
    """Claves dig|fecha|monto ya VISTO / OMITIR / CORREGIR+aplicado (cualquier lote)."""
    rows = db.execute(
        select(
            ConciliacionBancoOcrResultado.referencia_banco,
            ConciliacionBancoOcrResultado.fecha_banco,
            ConciliacionBancoOcrResultado.monto_banco,
        ).where(
            or_(
                ConciliacionBancoOcrResultado.decision == "VISTO",
                ConciliacionBancoOcrResultado.decision == "OMITIR",
                (
                    (ConciliacionBancoOcrResultado.decision == "CORREGIR")
                    & (ConciliacionBancoOcrResultado.aplicado.is_(True))
                ),
            )
        )
    ).all()
    out: set[str] = set()
    for r in rows:
        mb = float(r.monto_banco) if r.monto_banco is not None else None
        clave = _clave_caso_banco(r.referencia_banco, r.fecha_banco, mb)
        if clave:
            out.add(clave)
    return out


def _filtrar_extracto_sin_cerrados(
    db: Session, rows: list
) -> list:
    """Quita filas de extracto ya vistas/conciliadas en conciliaciones previas."""
    if not rows:
        return rows
    cerrados = _casos_banco_cerrados_globales(db)
    if not cerrados:
        return rows
    out = []
    for r in rows:
        mb = float(r.monto) if getattr(r, "monto", None) is not None else None
        clave = _clave_caso_banco(
            getattr(r, "referencia", None),
            getattr(r, "fecha", None),
            mb,
        )
        if clave and clave in cerrados:
            continue
        out.append(r)
    return out
def crear_lote_desde_extracto(
    db: Session,
    *,
    bancos: list[str],
    fecha_desde: date,
    fecha_hasta: date,
    moneda_carga: str,
    usuario_id: Optional[int],
) -> ConciliacionBancoOcrLote:
    """Arma un lote de conciliacion desde conciliacion_banco_extracto (BD historica)."""
    mon = (moneda_carga or "USD").strip().upper()
    if mon not in ("USD", "BS"):
        raise HTTPException(status_code=400, detail="moneda_carga debe ser USD o BS")
    if fecha_hasta < fecha_desde:
        raise HTTPException(status_code=400, detail="fecha_hasta debe ser >= fecha_desde")
    cats = []
    for b in bancos or []:
        nb = _normalizar_banco_categoria(b)
        if nb and nb not in cats:
            cats.append(nb)
    if not cats:
        raise HTTPException(
            status_code=400,
            detail="Indique al menos un banco (Mercantil, BNC, ...)",
        )

    q = (
        select(ConciliacionBancoExtracto)
        .where(ConciliacionBancoExtracto.banco.in_(cats))
        .where(ConciliacionBancoExtracto.moneda == mon)
        .where(
            or_(
                ConciliacionBancoExtracto.fecha.is_(None),
                ConciliacionBancoExtracto.fecha >= fecha_desde,
            )
        )
        .where(
            or_(
                ConciliacionBancoExtracto.fecha.is_(None),
                ConciliacionBancoExtracto.fecha <= fecha_hasta,
            )
        )
        .order_by(
            ConciliacionBancoExtracto.fecha.asc().nulls_last(),
            ConciliacionBancoExtracto.id.asc(),
        )
    )
    rows = list(db.execute(q).scalars().all())
    rows = _filtrar_extracto_sin_cerrados(db, rows)
    if not rows:
        raise HTTPException(
            status_code=404,
            detail=(
                "BD historica sin filas pendientes (bancos/fechas/moneda); "
                "las ya VISTO/conciliadas no se vuelven a cargar."
            ),
        )
    if len(rows) > MAX_FILAS_EXCEL_LOTE:
        raise HTTPException(
            status_code=400,
            detail=(
                f"BD historica tiene {len(rows)} filas (tope {MAX_FILAS_EXCEL_LOTE}). "
                "Acote fechas o bancos."
            ),
        )

    lote = ConciliacionBancoOcrLote(
        usuario_id=usuario_id,
        archivo_nombre="BD_HISTORICA",
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado="CARGADO",
        moneda_carga=mon,
    )
    db.add(lote)
    db.flush()

    batch: list[dict[str, Any]] = []
    for i, r in enumerate(rows, start=1):
        batch.append(
            {
                "lote_id": lote.id,
                "fila_excel": i,
                "fecha_banco": r.fecha,
                "referencia_banco": r.referencia,
                "ref_banco_norm": r.referencia_norm or r.referencia,
                "monto_banco": r.monto if mon == "USD" else None,
                "monto_banco_original": r.monto,
                "moneda_fila": mon,
            }
        )
        if len(batch) >= BULK_INSERT_CHUNK:
            db.bulk_insert_mappings(ConciliacionBancoOcrBanco, batch)
            batch = []
    if batch:
        db.bulk_insert_mappings(ConciliacionBancoOcrBanco, batch)

    _guardar_bancos_en_lote(lote, cats)
    lote.notas = json.dumps(
        {
            "filas_banco": len(rows),
            "fuente_carga": "historica",
            "bancos": cats,
            "max_filas_lote": MAX_FILAS_EXCEL_LOTE,
        },
        ensure_ascii=True,
    )
    db.commit()
    db.refresh(lote)
    logger.info(
        "[conciliacion-bancos] lote historica id=%s filas=%s bancos=%s",
        lote.id,
        len(rows),
        cats,
    )
    return lote


def crear_lote_desde_excel(
    db: Session,
    *,
    file: UploadFile,
    content: bytes,
    moneda_carga: str,
    fecha_desde: date,
    fecha_hasta: date,
    usuario_id: Optional[int],
    banco: Optional[str] = None,
) -> ConciliacionBancoOcrLote:
    mon = (moneda_carga or "USD").strip().upper()
    if mon not in ("USD", "BS"):
        raise HTTPException(status_code=400, detail="moneda_carga debe ser USD o BS")
    if fecha_hasta < fecha_desde:
        raise HTTPException(status_code=400, detail="fecha_hasta debe ser >= fecha_desde")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Debe subir un Excel (.xlsx o .xls)")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel invalido: {e}") from e
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel sin hoja activa")

    lote = ConciliacionBancoOcrLote(
        usuario_id=usuario_id,
        archivo_nombre=(file.filename or "banco.xlsx")[:255],
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado="CARGADO",
        moneda_carga=mon,
    )
    db.add(lote)
    db.flush()

    n = 0
    fechas_excel: list[date] = []
    batch: list[dict[str, Any]] = []
    extracto_filas: list[dict[str, Any]] = []
    t0 = datetime.utcnow()
    excede_tope = False

    def _flush_batch() -> None:
        nonlocal batch
        if not batch:
            return
        db.bulk_insert_mappings(ConciliacionBancoOcrBanco, batch)
        batch = []

    try:
        banco_form = _normalizar_banco_categoria(banco)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            # Formato: Banco | Fecha | Referencia | Monto
            # Compat: Fecha | Referencia | Monto (+ banco form o Otros)
            banco_celda = _normalizar_banco_categoria(
                row[0] if len(row) > 0 else None
            )
            if banco_celda is not None:
                fecha_b = _parse_fecha(row[1] if len(row) > 1 else None)
                ref_raw = (
                    str(row[2]).strip()
                    if len(row) > 2 and row[2] is not None
                    else ""
                )
                monto_raw = _parse_monto(row[3] if len(row) > 3 else None)
                banco_fila = banco_celda
            else:
                fecha_b = _parse_fecha(row[0] if len(row) > 0 else None)
                ref_raw = (
                    str(row[1]).strip()
                    if len(row) > 1 and row[1] is not None
                    else ""
                )
                monto_raw = _parse_monto(row[2] if len(row) > 2 else None)
                banco_fila = banco_form or "Otros"
            if not ref_raw and monto_raw is None and fecha_b is None:
                continue
            if not ref_raw:
                continue
            if n >= MAX_FILAS_EXCEL_LOTE:
                excede_tope = True
                break
            ref_norm = normalize_documento(ref_raw) or ref_raw.strip()
            batch.append(
                {
                    "lote_id": lote.id,
                    "fila_excel": i,
                    "fecha_banco": fecha_b,
                    "referencia_banco": ref_raw,
                    "ref_banco_norm": ref_norm,
                    "monto_banco": monto_raw if mon == "USD" else None,
                    "monto_banco_original": monto_raw,
                    "moneda_fila": mon,
                }
            )
            extracto_filas.append(
                {
                    "banco": banco_fila,
                    "fecha": fecha_b,
                    "referencia": ref_raw,
                    "referencia_norm": ref_norm,
                    "monto": monto_raw,
                    "moneda": mon,
                    "clave_natural": _clave_natural_extracto(
                        fecha=fecha_b,
                        referencia_norm=ref_norm,
                        monto=monto_raw,
                    ),
                }
            )
            n += 1
            if fecha_b is not None:
                fechas_excel.append(fecha_b)
            if len(batch) >= BULK_INSERT_CHUNK:
                _flush_batch()
        _flush_batch()
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if n == 0:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="No hay filas validas (Fecha, Referencia, Monto)",
        )

    if excede_tope:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                f"El Excel tiene mas de {MAX_FILAS_EXCEL_LOTE} filas validas. "
                f"Divida el extracto en lotes de hasta {MAX_FILAS_EXCEL_LOTE} filas."
            ),
        )

    if fechas_excel:
        excel_min = min(fechas_excel)
        excel_max = max(fechas_excel)
        lote.fecha_desde = min(lote.fecha_desde, excel_min)
        lote.fecha_hasta = max(lote.fecha_hasta, excel_max)

    if not extracto_filas:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "No se armó ninguna fila para BD historica. "
                "Excel debe ser Banco|Fecha|Referencia|Monto (o Fecha|Referencia|Monto)."
            ),
        )
    try:
        n_extracto = _upsert_extracto_filas(
            db,
            extracto_filas,
            lote_id=int(lote.id),
            archivo_nombre=lote.archivo_nombre,
        )
        n_hist_lote = int(
            db.scalar(
                select(func.count())
                .select_from(ConciliacionBancoExtracto)
                .where(ConciliacionBancoExtracto.lote_origen_id == int(lote.id))
            )
            or 0
        )
    except (ProgrammingError, SQLAlchemyError, Exception) as e:
        msg = str(getattr(e, "orig", None) or e)
        logger.exception(
            "[conciliacion-bancos] upsert extracto fallo lote_id=%s: %s",
            lote.id,
            msg,
        )
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=(
                "No se pudo guardar en BD historica (conciliacion_banco_extracto): "
                f"{msg[:240]}"
            ),
        ) from e

    if n_extracto <= 0 or n_hist_lote <= 0:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=(
                "El Excel se leyo pero BD historica quedo en 0 filas. "
                "Revise permisos/tabla conciliacion_banco_extracto."
            ),
        )

    lote.notas = json.dumps(
        {
            "filas_banco": n,
            "filas_extracto_upsert": n_extracto,
            "filas_extracto_lote_origen": n_hist_lote,
            "max_filas_lote": MAX_FILAS_EXCEL_LOTE,
            "fuente_carga": "excel",
            "extracto_ok": True,
        },
        ensure_ascii=True,
    )

    db.commit()
    db.refresh(lote)
    elapsed_ms = int((datetime.utcnow() - t0).total_seconds() * 1000)
    logger.info(
        "[conciliacion-bancos] crear_lote id=%s filas=%s extracto=%s hist_lote=%s elapsed_ms=%s archivo=%s",
        lote.id,
        n,
        n_extracto,
        n_hist_lote,
        elapsed_ms,
        lote.archivo_nombre,
    )
    return lote




def sanear_comparando_huerfano(db: Session, lote: ConciliacionBancoOcrLote) -> bool:
    """
    Tras deploy/restart el hilo muere y el lote queda en COMPARANDO para siempre.
    Si no hay hilo vivo, marca ERROR_COMPARAR para que el front deje de hacer poll
    y el usuario pueda pulsar Conciliar de nuevo.
    """
    from app.services.conciliacion_bancos_bg_runner import comparar_activo

    if (lote.estado or "").strip().upper() != "COMPARANDO":
        return False
    if comparar_activo(int(lote.id)):
        return False
    payload: dict[str, Any] = {}
    raw = (lote.notas or "").strip()
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                payload = data
        except Exception:
            payload = {}
    payload["comparar_error"] = (
        "Comparacion interrumpida (reinicio del servidor o proceso caido). "
        "Pulse Conciliar de nuevo."
    )
    payload["comparar_huerfano"] = True
    lote.notas = json.dumps(payload, ensure_ascii=True)
    lote.estado = "ERROR_COMPARAR"
    db.commit()
    db.refresh(lote)
    logger.warning(
        "[conciliacion-bancos] COMPARANDO huerfano saneado lote_id=%s -> ERROR_COMPARAR",
        lote.id,
    )
    return True


def iniciar_comparar_lote(
    db: Session,
    lote_id: int,
    *,
    bancos_filtro: Optional[list[str]] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
) -> dict[str, Any]:
    """
    Valida y lanza comparar en background. Responde al instante (evita timeout HTTP).
    Polling: GET /lotes/{id} hasta estado COMPARADO | ERROR_COMPARAR.
    """
    from app.services.conciliacion_bancos_bg_runner import (
        comparar_activo,
        spawn_comparar_lote,
    )

    lote = db.get(ConciliacionBancoOcrLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    # Deploy/restart mata el hilo: liberar COMPARANDO huerfano para poder relanzar.
    sanear_comparando_huerfano(db, lote)
    db.refresh(lote)
    if comparar_activo(lote_id):
        return {
            "ok": True,
            "async": True,
            "lote_id": lote_id,
            "estado": "COMPARANDO",
            "message": "Comparacion ya en curso; espere o recargue el lote.",
        }

    bancos_sel = normalizar_bancos_filtro(bancos_filtro)
    if not bancos_sel:
        raise HTTPException(
            status_code=400,
            detail="Seleccione al menos un banco (Mercantil, BNC, Binance, BNV, Recibos, Otros).",
        )
    if fecha_desde is not None or fecha_hasta is not None:
        fd = fecha_desde or lote.fecha_desde
        fh = fecha_hasta or lote.fecha_hasta
        if fh < fd:
            raise HTTPException(status_code=400, detail="fecha_hasta debe ser >= fecha_desde")
        lote.fecha_desde = fd
        lote.fecha_hasta = fh

    _guardar_bancos_en_lote(lote, bancos_sel)
    lote.estado = "COMPARANDO"
    db.commit()

    started = spawn_comparar_lote(
        lote_id,
        bancos_filtro=bancos_sel,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    if not started:
        return {
            "ok": True,
            "async": True,
            "lote_id": lote_id,
            "estado": "COMPARANDO",
            "message": "Comparacion ya en curso.",
        }
    return {
        "ok": True,
        "async": True,
        "lote_id": lote_id,
        "estado": "COMPARANDO",
        "message": "Comparacion iniciada en segundo plano.",
        "fecha_desde": lote.fecha_desde.isoformat() if lote.fecha_desde else None,
        "fecha_hasta": lote.fecha_hasta.isoformat() if lote.fecha_hasta else None,
        "bancos_filtro": bancos_sel,
    }


def comparar_lote(
    db: Session,
    lote_id: int,
    *,
    bancos_filtro: Optional[list[str]] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
) -> dict[str, Any]:
    lote = db.get(ConciliacionBancoOcrLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    if fecha_desde is not None or fecha_hasta is not None:
        fd = fecha_desde or lote.fecha_desde
        fh = fecha_hasta or lote.fecha_hasta
        if fh < fd:
            raise HTTPException(status_code=400, detail="fecha_hasta debe ser >= fecha_desde")
        lote.fecha_desde = fd
        lote.fecha_hasta = fh

    bancos_sel = normalizar_bancos_filtro(bancos_filtro)
    if not bancos_sel:
        raise HTTPException(
            status_code=400,
            detail="Seleccione al menos un banco (Mercantil, BNC, Binance, BNV, Recibos, Otros).",
        )
    _guardar_bancos_en_lote(lote, bancos_sel)
    lote.estado = "COMPARANDO"
    # Limpiar error previo de corrida fallida
    try:
        payload_n: dict[str, Any] = {}
        raw_n = (lote.notas or "").strip()
        if raw_n:
            data_n = json.loads(raw_n)
            if isinstance(data_n, dict):
                payload_n = data_n
        payload_n.pop("comparar_error", None)
        payload_n["comparar_started_at"] = datetime.utcnow().isoformat() + "Z"
        lote.notas = json.dumps(payload_n, ensure_ascii=True)
    except Exception:
        pass
    db.commit()
    db.refresh(lote)

    t0 = datetime.utcnow()
    # Comparar 25k filas puede superar el statement_timeout global (5 min).
    try:
        db.execute(text("SET LOCAL statement_timeout = '900000'"))  # 15 min
    except Exception:
        logger.warning(
            "[conciliacion-bancos] no se pudo subir statement_timeout lote_id=%s",
            lote_id,
        )

    # 1) Cerrados GLOBALES: VISTO / OMITIR / CORREGIR+aplicado de CUALQUIER lote.
    # Al reconcialiar con BD historica no deben reaparecer.
    cerrados_global = db.execute(
        select(
            ConciliacionBancoOcrResultado.lote_id,
            ConciliacionBancoOcrResultado.pago_id,
            ConciliacionBancoOcrResultado.banco_id,
            ConciliacionBancoOcrResultado.decision,
            ConciliacionBancoOcrResultado.aplicado,
            ConciliacionBancoOcrResultado.referencia_banco,
            ConciliacionBancoOcrResultado.fecha_banco,
            ConciliacionBancoOcrResultado.monto_banco,
        ).where(
            or_(
                ConciliacionBancoOcrResultado.decision == "VISTO",
                ConciliacionBancoOcrResultado.decision == "OMITIR",
                (
                    (ConciliacionBancoOcrResultado.decision == "CORREGIR")
                    & (ConciliacionBancoOcrResultado.aplicado.is_(True))
                ),
            ),
        )
    ).all()
    # 2) Cerrados de ESTE lote: excluir tambien por banco_id interno del lote
    cerrados_lote = [c for c in cerrados_global if c.lote_id == lote_id]

    excluir_pago_ids: set[int] = set()
    excluir_banco_ids: set[int] = set()
    excluir_casos_banco: set[str] = set()
    confirmados_conservados = 0

    for c in cerrados_global:
        if c.pago_id and c.decision == "CORREGIR" and bool(c.aplicado):
            excluir_pago_ids.add(int(c.pago_id))
        mb = float(c.monto_banco) if c.monto_banco is not None else None
        clave = _clave_caso_banco(c.referencia_banco, c.fecha_banco, mb)
        if clave:
            excluir_casos_banco.add(clave)
        if (
            c.lote_id == lote_id
            and c.decision == "CORREGIR"
            and bool(c.aplicado)
        ):
            confirmados_conservados += 1

    for c in cerrados_lote:
        if c.banco_id:
            excluir_banco_ids.add(int(c.banco_id))

    db.execute(
        delete(ConciliacionBancoOcrResultado).where(
            ConciliacionBancoOcrResultado.lote_id == lote_id,
            ConciliacionBancoOcrResultado.decision == "PENDIENTE",
        )
    )

    bancos = (
        db.execute(
            select(ConciliacionBancoOcrBanco).where(
                ConciliacionBancoOcrBanco.lote_id == lote_id
            )
        )
        .scalars()
        .all()
    )

    def _monto_usd_fila_banco(brow: ConciliacionBancoOcrBanco) -> Optional[float]:
        if lote.moneda_carga == "USD":
            if brow.monto_banco_original is not None:
                return float(brow.monto_banco_original)
            if brow.monto_banco is not None:
                return float(brow.monto_banco)
            return None
        if brow.monto_banco is not None:
            return float(brow.monto_banco)
        return None

    bancos = [
        b
        for b in bancos
        if int(b.id) not in excluir_banco_ids
        and (
            _clave_caso_banco(
                b.referencia_banco,
                b.fecha_banco,
                _monto_usd_fila_banco(b),
            )
            not in excluir_casos_banco
        )
    ]

    fp_ini = datetime.combine(lote.fecha_desde, time.min)
    fp_fin_excl = datetime.combine(lote.fecha_hasta + timedelta(days=1), time.min)
    pagos_raw = (
        db.execute(
            select(Pago).where(
                Pago.fecha_pago >= fp_ini,
                Pago.fecha_pago < fp_fin_excl,
                Pago.numero_documento.isnot(None),
                Pago.numero_documento != "",
            )
        )
        .scalars()
        .all()
    )
    cats_match = set(bancos_sel)
    if any(b != "Otros" for b in bancos_sel):
        cats_match.add("Otros")
    pagos = [
        p
        for p in pagos_raw
        if categoria_pago_conciliacion(p) in cats_match
        and int(p.id) not in excluir_pago_ids
    ]

    by_digits: dict[str, list[Pago]] = {}
    by_paquete: dict[tuple[date, float], list[Pago]] = {}
    for p in pagos:
        if _es_asiento_drive_abonos(p.numero_documento):
            continue
        key = _ref_solo_digitos(
            normalize_documento(p.numero_documento) or p.numero_documento or ""
        )
        if key:
            by_digits.setdefault(key, []).append(p)
        pk = _clave_paquete_fecha_monto(_pago_fecha(p), float(p.monto_pagado or 0))
        if pk:
            by_paquete.setdefault(pk, []).append(p)

    tasas_memo: dict[date, Any] = {}

    def _tasas_por_fecha_memo(f: date):
        if f not in tasas_memo:
            tasas_memo[f] = obtener_tasa_por_fecha(db, f)
        return {f: tasas_memo[f]}

    matched_pago_ids: set[int] = set(excluir_pago_ids)
    stats = {
        "MATCH_EXACTO": 0,
        "MATCH_PARCIAL": 0,
        "SIN_BD": 0,
        "SIN_BANCO": 0,
        "AMBIGUO": 0,
        "SIN_TASA": 0,
        "CONCILIADOS": 0,
    }

    resultados_batch: list[dict[str, Any]] = []

    def _sanear_pago_ids_batch(batch: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Quita pago_id huérfanos (FK) para no tumbar todo el bulk insert."""
        ids = {
            int(r["pago_id"])
            for r in batch
            if r.get("pago_id") is not None
        }
        if not ids:
            return batch
        existentes = set(
            int(x)
            for x in db.scalars(select(Pago.id).where(Pago.id.in_(list(ids)))).all()
        )
        faltan = ids - existentes
        if not faltan:
            return batch
        logger.warning(
            "[conciliacion-bancos] pago_id inexistentes omitidos en insert lote_id=%s n=%s ej=%s",
            lote_id,
            len(faltan),
            sorted(faltan)[:8],
        )
        out: list[dict[str, Any]] = []
        for r in batch:
            pid = r.get("pago_id")
            if pid is None or int(pid) in existentes:
                out.append(r)
                continue
            # Match que apunto a pago borrado -> tratar como SIN_BD
            if r.get("banco_id") is not None:
                rr = dict(r)
                rr["pago_id"] = None
                rr["fecha_bd"] = None
                rr["referencia_bd"] = None
                rr["monto_bd"] = None
                rr["similitud_pct"] = None
                rr["tipo_novedad"] = "SIN_BD"
                rr["detalle_aplicacion"] = (
                    f"pago_id {pid} ya no existe en pagos; marcado SIN_BD."
                )
                rr.pop("valores_antes", None)
                out.append(rr)
                stats["SIN_BD"] += 1
                tipo_prev = str(r.get("tipo_novedad") or "")
                if tipo_prev in stats and stats[tipo_prev] > 0:
                    stats[tipo_prev] -= 1
            # SIN_BANCO sin pago valido: descartar fila y ajustar KPI
            tipo_prev = str(r.get("tipo_novedad") or "")
            if tipo_prev in stats and stats[tipo_prev] > 0:
                stats[tipo_prev] -= 1
        return out

    def _flush_resultados(*, commit: bool = True) -> None:
        nonlocal resultados_batch
        if not resultados_batch:
            return
        batch = _sanear_pago_ids_batch(resultados_batch)
        resultados_batch = []
        if not batch:
            return
        try:
            db.bulk_insert_mappings(ConciliacionBancoOcrResultado, batch)
            if commit:
                db.commit()
        except IntegrityError as ie:
            logger.warning(
                "[conciliacion-bancos] bulk_insert IntegrityError lote_id=%s; reintento fila a fila: %s",
                lote_id,
                str(ie.orig)[:200] if getattr(ie, "orig", None) else str(ie)[:200],
            )
            try:
                db.rollback()
            except Exception:
                pass
            # Revalidar y insertar uno a uno para no perder el lote entero
            batch2 = _sanear_pago_ids_batch(batch)
            for row in batch2:
                try:
                    db.bulk_insert_mappings(ConciliacionBancoOcrResultado, [row])
                    if commit:
                        db.commit()
                except IntegrityError:
                    try:
                        db.rollback()
                    except Exception:
                        pass
                    logger.warning(
                        "[conciliacion-bancos] fila omitida por FK lote_id=%s pago_id=%s banco_id=%s",
                        lote_id,
                        row.get("pago_id"),
                        row.get("banco_id"),
                    )

    def _add_resultado(**fields: Any) -> None:
        resultados_batch.append(fields)
        if len(resultados_batch) >= BULK_INSERT_CHUNK:
            _flush_resultados()

    for b in bancos:
        ref_b = b.ref_banco_norm or b.referencia_banco
        dig_b = _ref_solo_digitos(ref_b)
        fecha_b = b.fecha_banco
        monto_orig = (
            float(b.monto_banco_original) if b.monto_banco_original is not None else None
        )
        monto_usd: Optional[float] = None
        tipo_extra = None

        if lote.moneda_carga == "USD":
            monto_usd = monto_orig
        elif monto_orig is not None and fecha_b is not None:
            _tasa, usd = tasa_y_equivalente_usd_excel(
                db,
                fecha_b,
                monto_orig,
                "BS",
                tasas_por_fecha=_tasas_por_fecha_memo(fecha_b),
            )
            if usd is None:
                tipo_extra = "SIN_TASA"
            else:
                monto_usd = usd
                b.monto_banco = Decimal(str(usd))
        elif monto_orig is not None:
            tipo_extra = "SIN_TASA"

        candidatos_exactos = [
            p
            for p in (by_digits.get(dig_b, []) if dig_b else [])
            if int(p.id) not in matched_pago_ids
        ]

        if tipo_extra == "SIN_TASA" and lote.moneda_carga == "BS":
            _add_resultado(
                lote_id=lote_id,
                banco_id=b.id,
                pago_id=None,
                fecha_banco=fecha_b,
                referencia_banco=b.referencia_banco,
                monto_banco=None,
                similitud_pct=None,
                tipo_novedad="SIN_TASA",
                decision="PENDIENTE",
                detalle_aplicacion=(
                    "Sin tasa Bs/USD para la fecha del banco; no se puede comparar monto."
                ),
            )
            stats["SIN_TASA"] += 1
            continue

        if len(candidatos_exactos) == 1:
            p = candidatos_exactos[0]
            if _paquete_banco_coherente_con_pago(
                p, fecha_banco=fecha_b, monto_usd=monto_usd
            ):
                matched_pago_ids.add(p.id)
                _add_resultado(
                    lote_id=lote_id,
                    banco_id=b.id,
                    pago_id=p.id,
                    fecha_banco=fecha_b,
                    fecha_bd=_pago_fecha(p),
                    referencia_banco=b.referencia_banco,
                    referencia_bd=p.numero_documento,
                    monto_banco=Decimal(str(monto_usd)) if monto_usd is not None else None,
                    monto_bd=p.monto_pagado,
                    similitud_pct=Decimal("100"),
                    tipo_novedad="MATCH_EXACTO",
                    decision="PENDIENTE",
                )
                stats["MATCH_EXACTO"] += 1
                continue

        if len(candidatos_exactos) > 1:
            pool = list(candidatos_exactos)
            cands = _candidatos_payload(pool)
            _add_resultado(
                lote_id=lote_id,
                banco_id=b.id,
                pago_id=None,
                fecha_banco=fecha_b,
                referencia_banco=b.referencia_banco,
                monto_banco=Decimal(str(monto_usd)) if monto_usd is not None else None,
                similitud_pct=Decimal("100"),
                tipo_novedad="AMBIGUO",
                decision="PENDIENTE",
                detalle_aplicacion=_detalle_ambiguo_serial(pool),
                valores_antes=json.dumps({"candidatos": cands}, ensure_ascii=True),
            )
            stats["AMBIGUO"] += 1
            continue

        pk_b = _clave_paquete_fecha_monto(fecha_b, monto_usd)
        pool_parcial: list[Pago] = []
        if pk_b:
            seen_p: set[int] = set()
            f_pk, m_pk = pk_b
            for dm in (0.0, -0.01, 0.01, -0.02, 0.02):
                bucket = by_paquete.get((f_pk, round(m_pk + dm, 2)), [])
                for p in bucket:
                    if p.id not in seen_p:
                        seen_p.add(p.id)
                        pool_parcial.append(p)
        mejores: list[tuple[float, Pago]] = []
        for p in pool_parcial:
            if p.id in matched_pago_ids:
                continue
            if _es_asiento_drive_abonos(p.numero_documento):
                continue
            if not _paquete_banco_coherente_con_pago(
                p, fecha_banco=fecha_b, monto_usd=monto_usd
            ):
                continue
            sim = _similitud(ref_b, p.numero_documento or "")
            if sim < SIMILITUD_MINIMA:
                continue
            mejores.append((sim, p))
        mejores.sort(key=lambda x: x[0], reverse=True)

        if not mejores:
            _add_resultado(
                lote_id=lote_id,
                banco_id=b.id,
                pago_id=None,
                fecha_banco=fecha_b,
                referencia_banco=b.referencia_banco,
                monto_banco=Decimal(str(monto_usd)) if monto_usd is not None else None,
                similitud_pct=None,
                tipo_novedad="SIN_BD",
                decision="PENDIENTE",
                detalle_aplicacion=(
                    "Referencia banco sin match en BD (voucher no digitalizado / no reportado)."
                ),
            )
            stats["SIN_BD"] += 1
            continue

        if len(mejores) > 1 and abs(mejores[0][0] - mejores[1][0]) < 0.5:
            top = [p for sim, p in mejores if abs(sim - mejores[0][0]) < 0.5][:8]
            cands = _candidatos_payload(top)
            _add_resultado(
                lote_id=lote_id,
                banco_id=b.id,
                pago_id=None,
                fecha_banco=fecha_b,
                referencia_banco=b.referencia_banco,
                monto_banco=Decimal(str(monto_usd)) if monto_usd is not None else None,
                similitud_pct=Decimal(str(mejores[0][0])),
                tipo_novedad="AMBIGUO",
                decision="PENDIENTE",
                detalle_aplicacion=_detalle_ambiguo_serial(top),
                valores_antes=json.dumps({"candidatos": cands}, ensure_ascii=True),
            )
            stats["AMBIGUO"] += 1
            continue

        sim, p = mejores[0]
        matched_pago_ids.add(p.id)
        _add_resultado(
            lote_id=lote_id,
            banco_id=b.id,
            pago_id=p.id,
            fecha_banco=fecha_b,
            fecha_bd=_pago_fecha(p),
            referencia_banco=b.referencia_banco,
            referencia_bd=p.numero_documento,
            monto_banco=Decimal(str(monto_usd)) if monto_usd is not None else None,
            monto_bd=p.monto_pagado,
            similitud_pct=Decimal(str(sim)),
            tipo_novedad="MATCH_PARCIAL",
            decision="PENDIENTE",
        )
        stats["MATCH_PARCIAL"] += 1

    for p in pagos:
        if p.id in matched_pago_ids:
            continue
        _add_resultado(
            lote_id=lote_id,
            banco_id=None,
            pago_id=p.id,
            fecha_bd=_pago_fecha(p),
            referencia_bd=p.numero_documento,
            monto_bd=p.monto_pagado,
            similitud_pct=None,
            tipo_novedad="SIN_BANCO",
            decision="PENDIENTE",
            detalle_aplicacion="Pago en BD del rango sin fila correspondiente en Excel banco.",
        )
        stats["SIN_BANCO"] += 1

    _flush_resultados(commit=False)
    stats["CONCILIADOS"] = confirmados_conservados

    lote = db.get(ConciliacionBancoOcrLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado tras comparar")
    elapsed_ms = int((datetime.utcnow() - t0).total_seconds() * 1000)
    try:
        payload_f: dict[str, Any] = {}
        raw_f = (lote.notas or "").strip()
        if raw_f:
            data_f = json.loads(raw_f)
            if isinstance(data_f, dict):
                payload_f = data_f
        payload_f["stats"] = stats
        payload_f["pagos_universo"] = len(pagos)
        payload_f["comparar_elapsed_ms"] = elapsed_ms
        payload_f.pop("comparar_error", None)
        lote.notas = json.dumps(payload_f, ensure_ascii=True)
    except Exception:
        pass
    lote.estado = "COMPARADO"
    db.commit()
    logger.info(
        "[conciliacion-bancos] comparar_lote lote_id=%s bancos=%s pagos=%s elapsed_ms=%s stats=%s",
        lote_id,
        len(bancos),
        len(pagos),
        elapsed_ms,
        stats,
    )
    return {
        "lote_id": lote_id,
        "estado": lote.estado,
        "stats": stats,
        "bancos_filtro": bancos_sel,
        "fecha_desde": lote.fecha_desde.isoformat() if lote.fecha_desde else None,
        "fecha_hasta": lote.fecha_hasta.isoformat() if lote.fecha_hasta else None,
        "pagos_universo": len(pagos),
        "confirmados_conservados": confirmados_conservados,
        "pagos_excluidos_conciliados": len(excluir_pago_ids),
        "casos_banco_excluidos": len(excluir_casos_banco),
        "excluidos_por_confirmacion": len(excluir_pago_ids) + len(excluir_casos_banco),
        "elapsed_ms": elapsed_ms,
    }


def _serial_choca_unique_pagos(
    db: Session, serial_new: str, *, exclude_pago_id: int
) -> bool:
    """True si grabar serial_new violaria ux_pagos_numero_documento_btrim u otro pago."""
    if numero_documento_ya_registrado(db, serial_new, exclude_pago_id=exclude_pago_id):
        return True
    sn = (serial_new or "").strip()
    if not sn:
        return False
    q = (
        select(Pago.id)
        .where(
            func.btrim(Pago.numero_documento) == sn,
            Pago.id != int(exclude_pago_id),
        )
        .limit(1)
    )
    return db.scalar(q) is not None


def _paquetes_iguales(
    fecha_a: Optional[date],
    monto_a: Optional[float],
    serial_a: Optional[str],
    fecha_b: Optional[date],
    monto_b: Optional[float],
    serial_b: Optional[str],
) -> bool:
    sa = _ref_solo_digitos(normalize_documento(serial_a) or serial_a or "")
    sb = _ref_solo_digitos(normalize_documento(serial_b) or serial_b or "")
    if sa != sb:
        return False
    if fecha_a != fecha_b:
        return False
    ma = float(monto_a or 0)
    mb = float(monto_b or 0)
    return abs(ma - mb) < MONTO_TOL


def _candidatos_ids_desde_resultado(res: ConciliacionBancoOcrResultado) -> set[int]:
    ids: set[int] = set()
    raw = (res.valores_antes or "").strip()
    if not raw:
        return ids
    try:
        data = json.loads(raw)
        if isinstance(data, dict) and isinstance(data.get("candidatos"), list):
            for c in data["candidatos"]:
                if isinstance(c, dict) and c.get("pago_id"):
                    ids.add(int(c["pago_id"]))
    except Exception:
        return ids
    return ids


def _normalizar_pago_ids_elegidos(
    pago_id_elegido: Optional[int],
    pago_ids_elegidos: Optional[list[int]],
) -> list[int]:
    out: list[int] = []
    seen: set[int] = set()
    for raw in list(pago_ids_elegidos or []) + (
        [pago_id_elegido] if pago_id_elegido is not None else []
    ):
        try:
            pid = int(raw)
        except (TypeError, ValueError):
            continue
        if pid <= 0 or pid in seen:
            continue
        seen.add(pid)
        out.append(pid)
    return out


def _clonar_resultado_ambiguo_para_pago(
    db: Session,
    origen: ConciliacionBancoOcrResultado,
    pago_id: int,
) -> ConciliacionBancoOcrResultado:
    """Nueva fila resultado (mismo caso banco) para aprobar otro candidato AMBIGUO."""
    clone = ConciliacionBancoOcrResultado(
        lote_id=origen.lote_id,
        banco_id=origen.banco_id,
        pago_id=None,
        fecha_banco=origen.fecha_banco,
        referencia_banco=origen.referencia_banco,
        monto_banco=origen.monto_banco,
        similitud_pct=origen.similitud_pct,
        tipo_novedad="AMBIGUO",
        decision="PENDIENTE",
        aplicado=False,
        detalle_aplicacion=origen.detalle_aplicacion,
        valores_antes=origen.valores_antes,
    )
    db.add(clone)
    db.flush()
    return clone



def _resultado_bloqueado_permanente(res: ConciliacionBancoOcrResultado) -> Optional[str]:
    """
    Regla general: VISTO / OMITIR / CORREGIR+aplicado son definitivos.
    VISTO: una vez marcado, la fila queda bloqueada y no se puede cambiar.
    Retorna mensaje de error o None si sigue editable.
    """
    dec = (res.decision or "").strip().upper()
    if dec == "VISTO":
        return (
            "Regla: decision VISTO es definitiva. La fila queda bloqueada "
            "y no se puede cambiar."
        )
    if dec == "OMITIR":
        return "Esta fila ya fue omitida y queda bloqueada."
    if dec == "CORREGIR" and bool(res.aplicado):
        return "Esta fila ya fue conciliada (CORREGIR) y queda bloqueada."
    return None

def decidir_y_aplicar(
    db: Session,
    resultado_id: int,
    *,
    decision: str,
    fuente_elegida: Optional[str],
    usuario_id: Optional[int],
    pago_id_elegido: Optional[int] = None,
    pago_ids_elegidos: Optional[list[int]] = None,
) -> dict[str, Any]:
    res = db.get(ConciliacionBancoOcrResultado, resultado_id)
    if not res:
        raise HTTPException(status_code=404, detail="Resultado no encontrado")
    bloqueo = _resultado_bloqueado_permanente(res)
    if bloqueo:
        raise HTTPException(status_code=400, detail=bloqueo)

    dec = (decision or "").strip().upper()
    if dec not in ("VISTO", "CORREGIR", "OMITIR"):
        raise HTTPException(status_code=400, detail="decision invalida")

    pago_ids = _normalizar_pago_ids_elegidos(pago_id_elegido, pago_ids_elegidos)

    # AMBIGUO multi: uno, varios o todos los candidatos ΓåÆ aprobar conciliacion en cada uno
    if (
        dec == "CORREGIR"
        and res.tipo_novedad == "AMBIGUO"
        and len(pago_ids) > 1
    ):
        permitidos = _candidatos_ids_desde_resultado(res)
        if not permitidos or any(pid not in permitidos for pid in pago_ids):
            raise HTTPException(
                status_code=400,
                detail="pago_ids_elegidos deben estar entre los candidatos AMBIGUO",
            )
        # Clonar antes de aplicar: cada decide hace commit; asi no dependemos
        # del estado post-commit del resultado original.
        targets: list[tuple[int, int]] = [(int(resultado_id), int(pago_ids[0]))]
        for pid in pago_ids[1:]:
            clone = _clonar_resultado_ambiguo_para_pago(db, res, int(pid))
            targets.append((int(clone.id), int(pid)))
        db.flush()

        outs: list[dict[str, Any]] = []
        cambios = 0
        errores_multi: list[dict[str, Any]] = []
        for target_id, pid in targets:
            try:
                r = decidir_y_aplicar(
                    db,
                    target_id,
                    decision="CORREGIR",
                    fuente_elegida=fuente_elegida,
                    usuario_id=usuario_id,
                    pago_id_elegido=pid,
                    pago_ids_elegidos=None,
                )
                outs.append(r)
                if r.get("cambio"):
                    cambios += 1
            except HTTPException as he:
                errores_multi.append(
                    {
                        "resultado_id": target_id,
                        "pago_id": pid,
                        "ok": False,
                        "error": he.detail,
                    }
                )
        if not outs and errores_multi:
            raise HTTPException(
                status_code=400,
                detail=errores_multi[0].get("error")
                or "No se pudo aplicar AMBIGUO multi",
            )
        return {
            "ok": len(errores_multi) == 0,
            "multiple": True,
            "resultado_id": int(resultado_id),
            "pago_ids": pago_ids,
            "aplicados": len(outs),
            "cambio": cambios > 0,
            "detalle": outs,
            "errores": errores_multi,
        }

    # AMBIGUO: el operador elige el pago/prestamo entre candidatos (p.ej. Mercantil)
    if dec == "CORREGIR" and res.tipo_novedad == "AMBIGUO":
        if not pago_ids:
            raise HTTPException(
                status_code=400,
                detail="AMBIGUO: elija uno o mas prestamos/pagos candidatos antes de confirmar.",
            )
        permitidos = _candidatos_ids_desde_resultado(res)
        pid = int(pago_ids[0])
        if not permitidos or pid not in permitidos:
            raise HTTPException(
                status_code=400,
                detail="pago_id_elegido no esta entre los candidatos AMBIGUO de esta fila",
            )
        pago_chk = db.get(Pago, pid)
        if not pago_chk:
            raise HTTPException(status_code=404, detail="Pago elegido no encontrado")
        res.pago_id = pid
        res.referencia_bd = pago_chk.numero_documento
        res.fecha_bd = _pago_fecha(pago_chk)
        res.monto_bd = pago_chk.monto_pagado

    now = datetime.now()
    res.usuario_decision_id = usuario_id
    res.decidido_en = now

    if dec == "OMITIR":
        res.decision = "OMITIR"
        res.aplicado = False
        res.detalle_aplicacion = "Omitido por administrador"
        db.commit()
        return {"ok": True, "resultado_id": res.id, "decision": res.decision, "aplicado": False}

    if dec == "VISTO":
        res.decision = "VISTO"
        res.aplicado = False
        res.fuente_elegida = None
        res.detalle_aplicacion = (
            "Visto: sin cambios en BD. Bloqueado definitivo (no se puede cambiar)."
        )
        db.commit()
        return {
            "ok": True,
            "resultado_id": res.id,
            "decision": res.decision,
            "aplicado": False,
            "bloqueado": True,
        }

    # CORREGIR
    fuente = (fuente_elegida or "").strip().upper()
    if fuente not in ("BD", "BANCO"):
        raise HTTPException(
            status_code=400,
            detail="Para CORREGIR indique fuente_elegida=BD o BANCO",
        )
    res.decision = "CORREGIR"
    res.fuente_elegida = fuente

    if res.tipo_novedad in ("SIN_BD", "SIN_TASA"):
        raise HTTPException(
            status_code=400,
            detail="No hay pago BD vinculado; no se puede corregir (no se crean pagos).",
        )
    if not res.pago_id:
        if res.tipo_novedad == "AMBIGUO":
            raise HTTPException(
                status_code=400,
                detail="AMBIGUO: elija el prestamo/pago candidato antes de confirmar.",
            )
        raise HTTPException(
            status_code=400,
            detail="No hay pago BD vinculado; no se puede corregir (no se crean pagos).",
        )

    pago = db.get(Pago, res.pago_id)
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    lote = db.get(ConciliacionBancoOcrLote, res.lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    antes = _snapshot_pago(pago)
    res.valores_antes = json.dumps(antes, ensure_ascii=True)

    if fuente == "BD":
        # Mantener paquete BD. No toca pagos.conciliado (autoconciliacion/cuotas).
        # La confirmacion bancaria queda en este resultado (CORREGIR+aplicado).
        res.aplicado = True
        res.detalle_aplicacion = (
            "Referencia RapiC/BD: sin cambios de paquete. "
            "Confirmacion bancaria registrada (no altera autoconciliacion)."
        )
        res.valores_despues = json.dumps(antes, ensure_ascii=True)
        lote.estado = "APLICADO"
        db.commit()
        return {
            "ok": True,
            "resultado_id": res.id,
            "decision": res.decision,
            "fuente_elegida": fuente,
            "aplicado": True,
            "cambio": False,
        }

    # fuente BANCO
    fecha_new = res.fecha_banco or _pago_fecha(pago)
    # Serial como en banco (Excel); normalize_documento solo limpia notacion cientifica
    serial_new = (
        normalize_documento(res.referencia_banco)
        or (res.referencia_banco or "").strip()
    )
    if not serial_new:
        raise HTTPException(status_code=400, detail="Referencia banco vacia")

    monto_new_usd: Optional[float]
    monto_bs_orig = None
    tasa_usada = None
    if lote.moneda_carga == "BS":
        banco = db.get(ConciliacionBancoOcrBanco, res.banco_id) if res.banco_id else None
        monto_orig = (
            float(banco.monto_banco_original)
            if banco and banco.monto_banco_original is not None
            else (float(res.monto_banco) if res.monto_banco is not None else None)
        )
        if fecha_new is None or monto_orig is None:
            raise HTTPException(status_code=400, detail="Falta fecha/monto banco para convertir Bs")
        tasa_usada, monto_new_usd = tasa_y_equivalente_usd_excel(db, fecha_new, monto_orig, "BS")
        if monto_new_usd is None:
            raise HTTPException(
                status_code=400,
                detail=f"No hay tasa Bs/USD para {fecha_new.isoformat()}",
            )
        monto_bs_orig = monto_orig
    else:
        monto_new_usd = float(res.monto_banco) if res.monto_banco is not None else float(pago.monto_pagado or 0)

    paquetes_iguales = _paquetes_iguales(
        _pago_fecha(pago),
        float(pago.monto_pagado or 0),
        pago.numero_documento,
        fecha_new,
        monto_new_usd,
        serial_new,
    )

    if paquetes_iguales:
        # Digitos iguales: alinear texto serial + institucion. No toca pagos.conciliado.
        # Si el texto banco (ej. 20582) ya esta en otro pago, no reescribir (unique).
        serial_changed = False
        serial_omitido_dup = False
        if (pago.numero_documento or "").strip() != serial_new:
            if _serial_choca_unique_pagos(db, serial_new, exclude_pago_id=int(pago.id)):
                serial_omitido_dup = True
            else:
                pago.numero_documento = serial_new[:100]
                serial_changed = True
        inst_changed = _aplicar_institucion_desde_lote(pago, lote)
        if not inst_changed and not serial_changed:
            res.aplicado = True
            extra_dup = (
                f" Serial banco '{serial_new}' ya existe en otro pago; se mantuvo "
                f"'{pago.numero_documento}'."
                if serial_omitido_dup
                else ""
            )
            res.detalle_aplicacion = (
                "Paquete banco coincide con BD: sin cambios de datos. "
                "Confirmacion bancaria registrada."
                + extra_dup
            )
            res.valores_despues = json.dumps(antes, ensure_ascii=True)
            try:
                db.commit()
            except IntegrityError as ie:
                db.rollback()
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "No se pudo confirmar: el serial choca con otro pago "
                        f"({serial_new}). Discernimiento manual."
                    ),
                ) from ie
            return {
                "ok": True,
                "resultado_id": res.id,
                "decision": res.decision,
                "fuente_elegida": fuente,
                "aplicado": True,
                "cambio": False,
            }
        despues = _snapshot_pago(pago)
        res.valores_despues = json.dumps(despues, ensure_ascii=True)
        res.aplicado = True
        res.referencia_bd = pago.numero_documento
        partes = []
        if serial_changed:
            partes.append(f"serial -> {pago.numero_documento}")
        if serial_omitido_dup:
            partes.append(
                f"serial banco '{serial_new}' omitido (ya existe en otro pago)"
            )
        if inst_changed:
            partes.append(
                f"Institucion actualizada a {pago.institucion_bancaria}"
            )
        res.detalle_aplicacion = (
            "; ".join(partes)
            + " (fecha/monto ya coincidian). Confirmacion bancaria registrada."
        )
        lote.estado = "APLICADO"
        try:
            db.commit()
        except IntegrityError as ie:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=(
                    "No se pudo confirmar: el serial choca con otro pago "
                    f"({serial_new}). Discernimiento manual."
                ),
            ) from ie
        return {
            "ok": True,
            "resultado_id": res.id,
            "decision": res.decision,
            "fuente_elegida": fuente,
            "aplicado": True,
            "cambio": True,
            "antes": antes,
            "despues": despues,
        }

    # Conflicto de serial en otro pago
    if _serial_choca_unique_pagos(db, serial_new, exclude_pago_id=int(pago.id)):
        # MATCH_EXACTO y MATCH_PARCIAL: mismo mecanismo (confirmar sin reescribir serial).
        # AMBIGUO: igual confirmacion, pero revision marca Ambiguo (tipo_novedad).
        if res.tipo_novedad in ("MATCH_EXACTO", "MATCH_PARCIAL", "AMBIGUO"):
            inst_changed = _aplicar_institucion_desde_lote(pago, lote)
            res.decision = "CORREGIR"
            res.fuente_elegida = fuente
            res.aplicado = True
            if res.tipo_novedad == "AMBIGUO":
                partes = [
                    "AMBIGUO: confirmacion bancaria registrada",
                    f"serial banco '{serial_new}' no reescrito (ya en otro pago)",
                ]
            else:
                partes = [
                    f"{res.tipo_novedad}: confirmacion bancaria registrada",
                    f"serial banco '{serial_new}' no reescrito (ya en otro pago)",
                ]
            if inst_changed:
                partes.append(
                    f"Institucion actualizada a {pago.institucion_bancaria}"
                )
            res.detalle_aplicacion = "; ".join(partes)
            res.valores_despues = json.dumps(_snapshot_pago(pago), ensure_ascii=True)
            lote.estado = "APLICADO"
            try:
                db.commit()
            except IntegrityError as ie:
                db.rollback()
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"No se pudo confirmar {res.tipo_novedad}: conflicto al guardar. "
                        "Discernimiento manual."
                    ),
                ) from ie
            return {
                "ok": True,
                "resultado_id": res.id,
                "decision": res.decision,
                "fuente_elegida": fuente,
                "aplicado": True,
                "cambio": bool(inst_changed),
                "ambiguo": res.tipo_novedad == "AMBIGUO",
            }
        # Otros tipos: no forzar
        res.decision = "PENDIENTE"
        res.fuente_elegida = None
        res.aplicado = False
        res.detalle_aplicacion = (
            "Serial banco ya existe en otro pago. Discernimiento manual (puede marcar Visto)."
        )
        db.commit()
        raise HTTPException(
            status_code=409,
            detail=res.detalle_aplicacion,
        )

    # Ref. Banco: alinear institucion al banco del extracto (ej. Otros -> BNV)
    inst_changed = _aplicar_institucion_desde_lote(pago, lote)

    had_cp = pago_tiene_aplicaciones_cuotas(db, pago.id)
    old_fecha = _pago_fecha(pago)
    old_monto = float(pago.monto_pagado or 0)

    # Actualizar pago
    if fecha_new is not None:
        pago.fecha_pago = datetime.combine(fecha_new, datetime.min.time())
    pago.monto_pagado = Decimal(str(round(float(monto_new_usd), 2)))
    pago.numero_documento = serial_new[:100]
    if lote.moneda_carga == "BS":
        pago.moneda_registro = "BS"
        pago.monto_bs_original = Decimal(str(round(float(monto_bs_orig), 2))) if monto_bs_orig is not None else None
        pago.tasa_cambio_bs_usd = Decimal(str(tasa_usada)) if tasa_usada is not None else None
        pago.fecha_tasa_referencia = fecha_new
    else:
        if not pago.moneda_registro:
            pago.moneda_registro = "USD"

    db.flush()

    fecha_changed = had_cp and old_fecha != _pago_fecha(pago)
    monto_changed = had_cp and abs(old_monto - float(pago.monto_pagado or 0)) >= MONTO_TOL
    cascada_info = None

    if (fecha_changed or monto_changed) and pago.prestamo_id:
        # Misma ruta segura que actualizar_pago cuando articula cuotas
        from app.services.pagos_cuotas_reaplicacion import reset_y_reaplicar_cascada_prestamo
        from app.services.pago_huella_funcional import primer_par_huella_duplicada_prestamo

        par = primer_par_huella_duplicada_prestamo(db, int(pago.prestamo_id))
        if par is not None:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Huella funcional duplicada en prestamo {pago.prestamo_id} "
                    f"(pagos {par[0]} / {par[1]}). No se aplico el cambio."
                ),
            )
        r = reset_y_reaplicar_cascada_prestamo(db, int(pago.prestamo_id))
        if not r.get("ok"):
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=f"No se pudo reaplicar cascada: {r.get('error')}",
            )
        cascada_info = r
    elif (
        not had_cp
        and pago.prestamo_id
        and float(pago.monto_pagado or 0) > 0
    ):
        from app.services.pagos_cascada_aplicacion import _aplicar_pago_a_cuotas_interno

        cc, cp = _aplicar_pago_a_cuotas_interno(pago, db)
        cascada_info = {"cuotas_completadas": cc, "cuotas_parciales": cp}

    despues = _snapshot_pago(pago)
    res.valores_despues = json.dumps(despues, ensure_ascii=True)
    res.aplicado = True
    res.referencia_bd = pago.numero_documento
    res.fecha_bd = _pago_fecha(pago)
    res.monto_bd = pago.monto_pagado
    extra_inst = (
        f" Institucion -> {pago.institucion_bancaria}."
        if inst_changed
        else ""
    )
    res.detalle_aplicacion = (
        "Actualizado con paquete banco (fecha/monto/serial) "
        + ("y cascada reaplicada." if cascada_info else "sin rearticulacion de cuotas.")
        + extra_inst
        + " Confirmacion bancaria registrada (sin alterar autoconciliacion)."
    )
    lote.estado = "APLICADO"
    try:
        db.commit()
    except IntegrityError as ie:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "No se pudo confirmar: el serial choca con otro pago "
                f"({serial_new}). Discernimiento manual."
            ),
        ) from ie
    return {
        "ok": True,
        "resultado_id": res.id,
        "decision": res.decision,
        "fuente_elegida": fuente,
        "aplicado": True,
        "cambio": True,
        "cascada": cascada_info,
        "antes": antes,
        "despues": despues,
    }



def decidir_masivo(
    db: Session,
    items: list[dict[str, Any]],
    *,
    usuario_id: Optional[int],
    fuente_default: str = "BANCO",
) -> dict[str, Any]:
    """
    Confirma varias filas. Cada una usa la misma logica que decidir_y_aplicar
    (transaccion independiente): un fallo no detiene el resto.
    """
    if not items:
        raise HTTPException(status_code=400, detail="No hay filas seleccionadas")
    # Tope por request HTTP (Cloudflare/Render ~100s). El front manda tandas de 50.
    if len(items) > 100:
        raise HTTPException(
            status_code=400,
            detail="Maximo 100 filas por request masivo; envie en tandas.",
        )

    ok = 0
    errores = 0
    sin_pago = 0
    cambios = 0
    detalle: list[dict[str, Any]] = []
    fuente_def = (fuente_default or "BANCO").strip().upper()
    if fuente_def not in ("BD", "BANCO"):
        fuente_def = "BANCO"

    for raw in items:
        rid = int(raw.get("resultado_id") or 0)
        if rid <= 0:
            errores += 1
            detalle.append({"resultado_id": rid, "ok": False, "error": "id invalido"})
            continue
        fuente = (raw.get("fuente_elegida") or fuente_def or "BANCO").strip().upper()
        if fuente not in ("BD", "BANCO"):
            fuente = fuente_def
        res = db.get(ConciliacionBancoOcrResultado, rid)
        if not res:
            errores += 1
            detalle.append({"resultado_id": rid, "ok": False, "error": "no encontrado"})
            continue
        try:
            pago_eleg = raw.get("pago_id_elegido")
            pago_eleg_i = int(pago_eleg) if pago_eleg not in (None, "") else None
            raw_ids = raw.get("pago_ids_elegidos")
            pago_ids_l = None
            if isinstance(raw_ids, list) and raw_ids:
                pago_ids_l = []
                for x in raw_ids:
                    try:
                        pago_ids_l.append(int(x))
                    except (TypeError, ValueError):
                        pass
            if res.tipo_novedad == "AMBIGUO":
                if not pago_ids_l and pago_eleg_i is None:
                    # Masivo sin eleccion: todos los candidatos -> Ambiguo
                    pago_ids_l = sorted(_candidatos_ids_desde_resultado(res))
                if not pago_ids_l and pago_eleg_i is None:
                    errores += 1
                    detalle.append(
                        {
                            "resultado_id": rid,
                            "ok": False,
                            "error": "AMBIGUO sin candidatos",
                        }
                    )
                    continue
                r = decidir_y_aplicar(
                    db,
                    rid,
                    decision="CORREGIR",
                    fuente_elegida=fuente,
                    usuario_id=usuario_id,
                    pago_id_elegido=pago_eleg_i,
                    pago_ids_elegidos=pago_ids_l,
                )
                ok += 1
                if r.get("cambio"):
                    cambios += 1
                detalle.append(
                    {
                        "resultado_id": rid,
                        "ok": True,
                        "modo": "CORREGIR",
                        "fuente": fuente,
                        "pago_id": pago_eleg_i,
                        "cambio": bool(r.get("cambio")),
                    }
                )
            elif not res.pago_id or res.tipo_novedad in ("SIN_BD", "SIN_TASA"):
                r = decidir_y_aplicar(
                    db,
                    rid,
                    decision="VISTO",
                    fuente_elegida=None,
                    usuario_id=usuario_id,
                )
                sin_pago += 1
                ok += 1
                detalle.append({"resultado_id": rid, "ok": True, "modo": "VISTO", **{k: r.get(k) for k in ("decision", "aplicado")}})
            else:
                r = decidir_y_aplicar(
                    db,
                    rid,
                    decision="CORREGIR",
                    fuente_elegida=fuente,
                    usuario_id=usuario_id,
                    pago_id_elegido=pago_eleg_i,
                )
                ok += 1
                if r.get("cambio"):
                    cambios += 1
                detalle.append(
                    {
                        "resultado_id": rid,
                        "ok": True,
                        "modo": "CORREGIR",
                        "fuente": fuente,
                        "cambio": bool(r.get("cambio")),
                    }
                )
        except HTTPException as he:
            errores += 1
            detalle.append(
                {
                    "resultado_id": rid,
                    "ok": False,
                    "error": str(he.detail),
                }
            )
            try:
                db.rollback()
            except Exception:
                pass
        except Exception as e:
            errores += 1
            logger.exception("decidir_masivo fallo resultado_id=%s", rid)
            detalle.append({"resultado_id": rid, "ok": False, "error": str(e)[:300]})
            try:
                db.rollback()
            except Exception:
                pass

    return {
        "ok": errores == 0,
        "total": len(items),
        "exitosos": ok,
        "errores": errores,
        "sin_pago_vistos": sin_pago,
        "con_cambio": cambios,
        "detalle": detalle[:200],
    }



def _resultado_a_dict(
    r: ConciliacionBancoOcrResultado, pago: Optional[Pago]
) -> dict[str, Any]:
    candidatos = None
    if not pago and r.valores_antes:
        try:
            raw = json.loads(r.valores_antes)
            if isinstance(raw, dict) and isinstance(raw.get("candidatos"), list):
                candidatos = raw["candidatos"]
        except Exception:
            candidatos = None
    return {
        "id": r.id,
        "lote_id": r.lote_id,
        "banco_id": r.banco_id,
        "pago_id": r.pago_id,
        "cedula": (pago.cedula_cliente if pago else None),
        "prestamo_id": (pago.prestamo_id if pago else None),
        "institucion_bancaria": (pago.institucion_bancaria if pago else None),
        "institucion_categoria": (
            categoria_pago_conciliacion(pago) if pago else None
        ),
        "fecha_banco": r.fecha_banco.isoformat() if r.fecha_banco else None,
        "fecha_bd": r.fecha_bd.isoformat() if r.fecha_bd else None,
        "referencia_banco": r.referencia_banco,
        "referencia_bd": r.referencia_bd,
        "monto_banco": float(r.monto_banco) if r.monto_banco is not None else None,
        "monto_bd": float(r.monto_bd) if r.monto_bd is not None else None,
        "similitud_pct": float(r.similitud_pct) if r.similitud_pct is not None else None,
        "tipo_novedad": r.tipo_novedad,
        "decision": r.decision,
        "fuente_elegida": r.fuente_elegida,
        "aplicado": bool(r.aplicado),
        "detalle_aplicacion": r.detalle_aplicacion,
        "candidatos": candidatos,
    }


def kpis_vivos_lote(db: Session, lote_id: int) -> dict[str, int]:
    """
    KPIs actuales del lote (no el snapshot de comparar):
    - por tipo_novedad: solo filas decision=PENDIENTE
    - CONCILIADOS: decision=CORREGIR y aplicado=true
    Asi al confirmar, MATCH_*/AMBIGUO bajan y CONCILIADOS sube.
    """
    out: dict[str, int] = {
        "MATCH_EXACTO": 0,
        "MATCH_PARCIAL": 0,
        "SIN_BD": 0,
        "SIN_BANCO": 0,
        "AMBIGUO": 0,
        "SIN_TASA": 0,
        "CONCILIADOS": 0,
    }
    rows = db.execute(
        select(
            ConciliacionBancoOcrResultado.tipo_novedad,
            func.count(),
        )
        .where(
            ConciliacionBancoOcrResultado.lote_id == lote_id,
            ConciliacionBancoOcrResultado.decision == "PENDIENTE",
        )
        .group_by(ConciliacionBancoOcrResultado.tipo_novedad)
    ).all()
    for tipo, n in rows:
        k = str(tipo or "").upper()
        if k in out:
            out[k] = int(n or 0)
    n_ok = db.scalar(
        select(func.count())
        .select_from(ConciliacionBancoOcrResultado)
        .where(
            ConciliacionBancoOcrResultado.lote_id == lote_id,
            ConciliacionBancoOcrResultado.decision == "CORREGIR",
            ConciliacionBancoOcrResultado.aplicado.is_(True),
        )
    )
    out["CONCILIADOS"] = int(n_ok or 0)
    return out


def listar_resultados(
    db: Session,
    lote_id: int,
    *,
    page: int = 1,
    per_page: int = 1000,
    tipos: Optional[list[str]] = None,
    decision: Optional[str] = None,
) -> dict[str, Any]:
    """Lista paginada (lotes de 25k no caben en una sola respuesta HTTP)."""
    page = max(1, int(page or 1))
    per_page = min(1000, max(1, int(per_page or 1000)))
    q = select(ConciliacionBancoOcrResultado).where(
        ConciliacionBancoOcrResultado.lote_id == lote_id
    )
    tipos_n = [str(x).strip().upper() for x in (tipos or []) if str(x).strip()]
    quiere_conciliados = "CONCILIADOS" in tipos_n
    tipos_n = [x for x in tipos_n if x != "CONCILIADOS"]
    if quiere_conciliados and not tipos_n:
        q = q.where(
            ConciliacionBancoOcrResultado.decision == "CORREGIR",
            ConciliacionBancoOcrResultado.aplicado.is_(True),
        )
    elif quiere_conciliados and tipos_n:
        q = q.where(
            or_(
                (
                    (ConciliacionBancoOcrResultado.decision == "CORREGIR")
                    & (ConciliacionBancoOcrResultado.aplicado.is_(True))
                ),
                (
                    (ConciliacionBancoOcrResultado.decision == "PENDIENTE")
                    & (ConciliacionBancoOcrResultado.tipo_novedad.in_(tipos_n))
                ),
            )
        )
    elif tipos_n:
        q = q.where(ConciliacionBancoOcrResultado.tipo_novedad.in_(tipos_n))
    if decision:
        q = q.where(ConciliacionBancoOcrResultado.decision == decision.strip().upper())

    total = int(
        db.scalar(select(func.count()).select_from(q.order_by(None).subquery()))
        or 0
    )
    rows = (
        db.execute(
            q.order_by(ConciliacionBancoOcrResultado.id.asc())
            .offset((page - 1) * per_page)
            .limit(per_page)
        )
        .scalars()
        .all()
    )
    pago_ids = [int(r.pago_id) for r in rows if r.pago_id]
    pagos_map: dict[int, Pago] = {}
    if pago_ids:
        for p in db.execute(select(Pago).where(Pago.id.in_(pago_ids))).scalars().all():
            pagos_map[int(p.id)] = p
    items = [
        _resultado_a_dict(r, pagos_map.get(int(r.pago_id)) if r.pago_id else None)
        for r in rows
    ]
    pages = (total + per_page - 1) // per_page if per_page else 1
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "stats": kpis_vivos_lote(db, lote_id),
    }


def listar_resultados_todos(db: Session, lote_id: int) -> list[dict[str, Any]]:
    """Solo para export Excel: carga completa (puede ser pesado)."""
    rows = (
        db.execute(
            select(ConciliacionBancoOcrResultado)
            .where(ConciliacionBancoOcrResultado.lote_id == lote_id)
            .order_by(ConciliacionBancoOcrResultado.id.asc())
        )
        .scalars()
        .all()
    )
    pago_ids = [int(r.pago_id) for r in rows if r.pago_id]
    pagos_map: dict[int, Pago] = {}
    if pago_ids:
        # Chunk IN para no saturar
        for i in range(0, len(pago_ids), 2000):
            chunk = pago_ids[i : i + 2000]
            for p in db.execute(select(Pago).where(Pago.id.in_(chunk))).scalars().all():
                pagos_map[int(p.id)] = p
    return [
        _resultado_a_dict(r, pagos_map.get(int(r.pago_id)) if r.pago_id else None)
        for r in rows
    ]


_EXPORT_HEADERS = [
    "referencia_banco",
    "referencia_bd",
    "similitud_pct",
    "cedula",
    "prestamo_id",
    "institucion_bancaria",
    "institucion_categoria",
    "fecha_banco",
    "fecha_bd",
    "monto_banco_usd",
    "monto_bd_usd",
    "tipo_novedad",
    "decision",
    "fuente_elegida",
    "aplicado",
    "pago_id",
    "detalle",
]

# Pestanas de novedad (pendientes) + CONCILIADOS (aprobados bancarios)
_EXPORT_HOJAS_NOVEDAD = (
    "MATCH_EXACTO",
    "MATCH_PARCIAL",
    "SIN_BD",
    "SIN_BANCO",
    "AMBIGUO",
    "SIN_TASA",
    "CONCILIADOS",
)


def _es_resultado_conciliado_bancario(r: dict[str, Any]) -> bool:
    return (r.get("decision") or "").strip().upper() == "CORREGIR" and bool(
        r.get("aplicado")
    )


def _fila_export_resultado(r: dict[str, Any]) -> list[Any]:
    return [
        r.get("referencia_banco"),
        r.get("referencia_bd"),
        r.get("similitud_pct"),
        r.get("cedula"),
        r.get("prestamo_id"),
        r.get("institucion_bancaria"),
        r.get("institucion_categoria"),
        r.get("fecha_banco"),
        r.get("fecha_bd"),
        r.get("monto_banco"),
        r.get("monto_bd"),
        r.get("tipo_novedad"),
        r.get("decision"),
        r.get("fuente_elegida"),
        r.get("aplicado"),
        r.get("pago_id"),
        r.get("detalle_aplicacion"),
    ]


def exportar_excel_lote(db: Session, lote_id: int) -> bytes:
    """Excel: una hoja por novedad (pendientes) + pestana CONCILIADOS."""
    rows = listar_resultados_todos(db, lote_id)
    by_tipo: dict[str, list[dict[str, Any]]] = {k: [] for k in _EXPORT_HOJAS_NOVEDAD}
    for r in rows:
        if _es_resultado_conciliado_bancario(r):
            by_tipo["CONCILIADOS"].append(r)
            continue
        # Hojas de novedad = trabajo pendiente (alineado a chips KPI)
        if (r.get("decision") or "").strip().upper() != "PENDIENTE":
            # VISTO/OMITIR u otros cerrados: hoja auxiliar
            tipo = "CERRADOS_OTROS"
        else:
            tipo = (r.get("tipo_novedad") or "").strip() or "OTROS"
        if tipo not in by_tipo:
            by_tipo[tipo] = []
        by_tipo[tipo].append(r)

    wb = Workbook()
    first = True
    orden = list(_EXPORT_HOJAS_NOVEDAD) + [
        k for k in by_tipo.keys() if k not in _EXPORT_HOJAS_NOVEDAD
    ]
    for tipo in orden:
        items = by_tipo.get(tipo, [])
        # Siempre crear CONCILIADOS aunque vacia; otras hojas solo si hay filas
        if tipo != "CONCILIADOS" and not items:
            continue
        title = tipo[:31]
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)
        ws.append(list(_EXPORT_HEADERS))
        for r in items:
            ws.append(_fila_export_resultado(r))

    if first:
        # Lote sin filas: hoja vacia
        ws = wb.active
        ws.title = "CONCILIADOS"
        ws.append(list(_EXPORT_HEADERS))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
