# -*- coding: utf-8 -*-
"""
Importación extracto (faltantes): parse Excel banco → comparar vs pagos en prestamos APROBADO.

Solo APROBADO. LIQUIDADO y DESISTIMIENTO (y alias) no entran en lista ni comparación.
Varios APROBADO misma cédula → el de fecha_aprobacion más reciente.
- IGUAL_100: mismo serial ya en pagos del préstamo → no se lista (solo stats).
- PRESTAMO_PAGADO: última cuota vencida en Pagado (préstamo al día) → se lista sin OK.
- SE_PUEDE_IMPORTAR: serial ausente → % = 100% confiabilidad de importación.
- SEMEJANTE: serial parecido (≥70%) → % = similitud; importable con OK bajo criterio manual.
Importar (OK): pago con fecha/serial/monto + imagen placeholder;
marca el préstamo APROBADO con requiere_revision=SI.
Un lote = un banco (Mercantil, BNC, Binance, Zelle, BNV) elegido al subir el Excel.

Comparación crítica (evita falsos +/-): cédula canónica V/E/G/J + dígitos;
serial solo dígitos (prefijos BNC/ ignorados); serial compuesto indexado por partes;
similitud serial ≥70% (misma regla que Conciliación Bancos).
"""
from __future__ import annotations

import base64
import copy
import io
import logging
import re
import threading
import time as time_mod
import uuid
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from typing import Any, Optional
from zoneinfo import ZoneInfo

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.orm import Session

from app.constants.prestamo_estados import (
    ESTADOS_PRESTAMO_EXCLUIDOS_COBRANZA_NOTIF,
    prestamo_estado_es_desistimiento,
)
from app.core.documento import compose_numero_documento_almacenado, normalize_documento
from app.models.cliente import Cliente
from app.models.importacion_extracto import (
    ImportacionExtractoFila,
    ImportacionExtractoLote,
    ImportacionExtractoPagoConfirmado,
)
from app.models.pago import Pago
from app.models.pago_comprobante_imagen import PagoComprobanteImagen
from app.models.prestamo import Prestamo
from app.services.cuota_estado import estado_ultima_cuota_por_vencimiento
from app.services.cobranzas.universo_analisis_service import (
    invalidate_universo_analisis_cache,
)
from app.services.pago_autoconciliacion import (
    INSTITUCION_BANCARIA_DRIVE,
    es_referencia_abonos_drive_notif,
    forzar_institucion_drive_si_abonos,
    marcar_pago_autoconciliado,
)
from app.services.pago_numero_documento import numero_documento_ya_registrado
from app.services.pagos_gmail.comprobante_bd import url_comprobante_imagen_absoluta
from app.api.v1.endpoints.pagos.constants import TZ_NEGOCIO
from app.utils.cedula_almacenamiento import (
    normalizar_cedula_almacenamiento,
    resolver_cedula_almacenada_en_clientes,
    texto_cedula_comparable_bd,
)

# No disponibles para comparación ni lista de importación extracto.
_ESTADOS_EXCLUIDOS_IMPORTACION = frozenset(ESTADOS_PRESTAMO_EXCLUIDOS_COBRANZA_NOTIF)

# Marcas canónicas en columna Observación (normalizadas al listar y al crear lote).
_MARCA_OBS_DRIVE = "Drive"
_MARCA_OBS_SERIAL_COMPUESTO = "Serial compuesto"
# Alias legacy en lotes ya guardados (solo detección al listar).
_MARCA_OBS_DRIVE_LEGACY = ("banco drive",)
_MARCA_OBS_SERIAL_COMPUESTO_LEGACY = ("serial mixto",)
# Corridas de dígitos típicas de serial bancario (evita ruido corto).
_RE_SERIAL_DIGIT_RUN = re.compile(r"\d{5,}")
# Separadores que un humano usa al juntar 2+ seriales en un solo Nº documento.
_RE_SERIAL_MIXTO_SPLIT = re.compile(
    r"\s*[-–—|;]+\s*|\s+y\s+|\s*/\s*(?=BNC|BINANCE|MERCANTIL|BNV|BDV|REF)",
    re.IGNORECASE,
)

# Filas que el usuario puede autorizar con OK (individual o lote).
_ESTADOS_OK_IMPORTAR = frozenset({"SE_PUEDE_IMPORTAR", "SEMEJANTE", "VISTO"})
# Filtro solo_importables / UI «Se puede importar»: solo 100% confianza (serial ausente en cartera).
_ESTADO_FILTRO_100_IMPORTABLE = "SE_PUEDE_IMPORTAR"

# Bancos admitidos en extracto (un archivo = un banco, elegido en cabecera antes de subir).
_BANCOS_EXTRACTO_PERMITIDOS = frozenset({"Mercantil", "BNC", "Binance", "Zelle", "BNV"})

logger = logging.getLogger(__name__)

# Caché índice serial global (modo solo Serial): evita re-leer todos los pagos en cada request.
_serial_cartera_cache_lock = threading.Lock()
_serial_cartera_cache: Optional[tuple[float, dict[str, Any]]] = None
_SERIAL_CARTERA_CACHE_TTL_SEC = 180.0
_PAGOS_SERIAL_CHUNK = 4000

MAX_FILAS = 100_000
_FILAS_LISTAR_DEFAULT = 200
_FILAS_LISTAR_MAX = 500
_INSERT_CHUNK = 1000
_EVAL_LOG_CADA = 5000
_LOTE_BG_MIN_FILAS = 2000
_SERIAL_SQL_BATCH = 400
_SKIP_SEMEJANTE_MIN_FILAS = 2000
USUARIO_REGISTRO = "importacion-extracto@sistema.rapicredit.com"
# Umbral similitud serial (alineado con conciliacion_bancos_service.SIMILITUD_MINIMA).
_SIMILITUD_SERIAL_MINIMA = 70.0
_MIN_DIGITOS_SERIAL = 5

# PNG 1x1 blanco (placeholder genérico; no inventa comprobante real).
_PNG_BLANCO_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)

_RE_CEDULA = re.compile(
    r"(?:^|[:\s])([VEJG])\s*-?\s*0*(\d{5,12})\b",
    re.IGNORECASE,
)
_RE_CEDULA_ALT = re.compile(
    r"\b([VEJG])\s*-?\s*0*(\d{5,12})\b",
    re.IGNORECASE,
)


def ensure_schema(db: Session) -> None:
    """CREATE TABLE IF NOT EXISTS vía metadata (idempotente)."""
    from app.core.database import engine

    ImportacionExtractoLote.__table__.create(bind=engine, checkfirst=True)
    ImportacionExtractoFila.__table__.create(bind=engine, checkfirst=True)
    ImportacionExtractoPagoConfirmado.__table__.create(bind=engine, checkfirst=True)
    try:
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_lote "
                "ADD COLUMN IF NOT EXISTS banco VARCHAR(50)"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_lote "
                "ADD COLUMN IF NOT EXISTS modo_cedula BOOLEAN NOT NULL DEFAULT true"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_lote "
                "ADD COLUMN IF NOT EXISTS modo_serial BOOLEAN NOT NULL DEFAULT false"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_fila "
                "ADD COLUMN IF NOT EXISTS oculto BOOLEAN NOT NULL DEFAULT false"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_fila "
                "ADD COLUMN IF NOT EXISTS destino_importacion VARCHAR(20)"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_pago_confirmado "
                "ADD COLUMN IF NOT EXISTS prestamo_id INTEGER"
            )
        )
        db.execute(
            text(
                "ALTER TABLE importacion_extracto_pago_confirmado "
                "ADD COLUMN IF NOT EXISTS cedula VARCHAR(32)"
            )
        )
        db.commit()
    except Exception:
        db.rollback()


def _normalizar_banco_extracto(raw: Optional[str]) -> Optional[str]:
    s = (raw or "").strip()
    if not s:
        return None
    for b in sorted(_BANCOS_EXTRACTO_PERMITIDOS):
        if b.lower() == s.lower():
            return b
    return None


def _banco_lote_fila(db: Session, f: ImportacionExtractoFila) -> str:
    """Banco del lote; legacy sin columna → Mercantil."""
    lote = db.get(ImportacionExtractoLote, int(f.lote_id)) if f.lote_id else None
    banco = _normalizar_banco_extracto(getattr(lote, "banco", None) if lote else None)
    return banco or "Mercantil"


def _es_pago_banco_drive(
    institucion: Optional[str],
    numero_documento: Optional[str] = None,
    referencia_pago: Optional[str] = None,
) -> bool:
    """True si el pago es institución Drive / ABONOS-DRIVE / ABONOS-NOTIF."""
    if (institucion or "").strip().lower() == INSTITUCION_BANCARIA_DRIVE.lower():
        return True
    if es_referencia_abonos_drive_notif(numero_documento):
        return True
    if es_referencia_abonos_drive_notif(referencia_pago):
        return True
    return False


def _detalle_tiene_marca_drive(det: Optional[str]) -> bool:
    d = (det or "").lower()
    if not d:
        return False
    if _MARCA_OBS_DRIVE.lower() in d:
        return True
    return any(leg in d for leg in _MARCA_OBS_DRIVE_LEGACY)


def _detalle_tiene_marca_serial_compuesto(det: Optional[str]) -> bool:
    d = (det or "").lower()
    if not d:
        return False
    if _MARCA_OBS_SERIAL_COMPUESTO.lower() in d:
        return True
    return any(leg in d for leg in _MARCA_OBS_SERIAL_COMPUESTO_LEGACY)


def _normalizar_detalle_observaciones(det: Optional[str]) -> Optional[str]:
    """Unifica marcas Drive / Serial compuesto en texto de observación."""
    if not det:
        return det
    s = str(det).strip()
    for leg in _MARCA_OBS_DRIVE_LEGACY:
        s = re.sub(re.escape(leg), _MARCA_OBS_DRIVE, s, flags=re.IGNORECASE)
    for leg in _MARCA_OBS_SERIAL_COMPUESTO_LEGACY:
        s = re.sub(re.escape(leg), _MARCA_OBS_SERIAL_COMPUESTO, s, flags=re.IGNORECASE)
    return s.strip() or None


def _anexar_marca_observacion(det: Optional[str], marca: str) -> str:
    """Agrega marca canónica sin duplicar (Drive / Serial compuesto)."""
    base = _normalizar_detalle_observaciones(det) or ""
    marca = (marca or "").strip()
    if not marca:
        return base
    ml = marca.lower()
    if ml.startswith(_MARCA_OBS_DRIVE.lower()) and _detalle_tiene_marca_drive(base):
        return base
    if (
        marca == _MARCA_OBS_SERIAL_COMPUESTO
        or ml == _MARCA_OBS_SERIAL_COMPUESTO.lower()
    ) and _detalle_tiene_marca_serial_compuesto(base):
        return base
    if marca.lower() in base.lower():
        return base
    return f"{base} | {marca}".strip(" |") if base else marca


def _texto_obs_banco_drive(pago_ids: list[int]) -> str:
    n = len(pago_ids)
    if n <= 0:
        return ""
    muestra = ",".join(str(i) for i in pago_ids[:5])
    extra = f"+{n - 5}" if n > 5 else ""
    return f"{_MARCA_OBS_DRIVE} ({n} pago(s): {muestra}{extra})"


def _anotar_banco_drive(
    ev: dict[str, Any], idx: dict[str, Any], prestamo_id: Optional[int]
) -> dict[str, Any]:
    """Si el préstamo tiene pagos Drive, anexa observación y marca alerta."""
    if not prestamo_id:
        return ev
    drive_ids = list(idx.get("drive_by_prestamo", {}).get(int(prestamo_id), []) or [])
    if not drive_ids:
        return ev
    note = _texto_obs_banco_drive(drive_ids)
    ev["detalle"] = _anexar_marca_observacion(ev.get("detalle"), note)
    ev["alerta_banco_drive"] = True
    return ev


def _solo_digitos(s: Optional[str]) -> str:
    return re.sub(r"\D+", "", (s or "").strip())


def _serial_norm_comparacion(val: Optional[str]) -> str:
    """
    Clave de match extracto ↔ pagos del préstamo.

    Ignora letras y signos a la izquierda (BNC/, REF., guiones, etc.);
    compara solo la parte numérica (24803998). Misma regla que Conciliación Bancos.
    También limpia sufijo Excel ``.0``.
    """
    if not val:
        return ""
    s = str(val).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    from app.services.conciliacion_bancos_service import _ref_solo_digitos

    return _ref_solo_digitos(s) or _solo_digitos(s)


def _partes_serial_texto(val: Optional[str]) -> list[str]:
    """
    Partes numéricas de un Nº documento.

    Detecta serial mixto: ``BNC/125201931 - BNC/103917175`` →
    ``['125201931', '103917175']`` (no concatena en un solo blob).
    """
    if val is None or val == "":
        return []
    raw = str(val).strip()
    if not raw:
        return []
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".", 1)[0]

    out: list[str] = []
    seen: set[str] = set()

    def _add(dig: str) -> None:
        if dig and dig not in seen:
            seen.add(dig)
            out.append(dig)

    chunks = [c.strip() for c in _RE_SERIAL_MIXTO_SPLIT.split(raw) if c and c.strip()]
    if len(chunks) <= 1:
        chunks = [raw]

    for ch in chunks:
        runs = _RE_SERIAL_DIGIT_RUN.findall(ch)
        if len(runs) >= 2:
            for r in runs:
                _add(_serial_norm_comparacion(r) or (r.lstrip("0") or "0"))
            continue
        dig = _serial_norm_comparacion(ch)
        if dig:
            # Si la normalización juntó varias corridas (sin separador claro), partir.
            runs2 = _RE_SERIAL_DIGIT_RUN.findall(ch)
            if len(runs2) >= 2:
                for r in runs2:
                    _add(_serial_norm_comparacion(r) or (r.lstrip("0") or "0"))
            else:
                _add(dig)

    if not out:
        for r in _RE_SERIAL_DIGIT_RUN.findall(raw):
            _add(_serial_norm_comparacion(r) or (r.lstrip("0") or "0"))
    return out


def _es_serial_mixto_texto(val: Optional[str]) -> bool:
    """True si el humano juntó 2+ seriales en un solo campo (justifican 1 pago)."""
    return len(_partes_serial_texto(val)) >= 2


def _seriales_norm_desde_campos(*cands: Optional[str]) -> list[str]:
    """
    Claves numéricas distintas para indexar un pago.

    Si un campo es serial mixto, indexa cada parte por separado (evita falso
    'Se puede importar' cuando el extracto trae uno de los seriales del mixto).
    """
    out: list[str] = []
    seen: set[str] = set()
    for cand in cands:
        if not cand:
            continue
        partes = _partes_serial_texto(cand)
        if not partes:
            dig = _serial_norm_comparacion(cand)
            if dig and dig not in seen:
                seen.add(dig)
                out.append(dig)
            continue
        for dig in partes:
            if dig not in seen:
                seen.add(dig)
                out.append(dig)
    return out


def _anotar_serial_mixto(
    ev: dict[str, Any], idx: dict[str, Any], prestamo_id: Optional[int], pago_id: Optional[int]
) -> dict[str, Any]:
    """Si el match involucra un pago con Nº documento compuesto, marca observación rosa."""
    if not prestamo_id or pago_id is None:
        return ev
    mixto_ids = set(idx.get("mixto_by_prestamo", {}).get(int(prestamo_id), []) or [])
    if int(pago_id) not in mixto_ids:
        return ev
    ev["detalle"] = _anexar_marca_observacion(
        ev.get("detalle"), _MARCA_OBS_SERIAL_COMPUESTO
    )
    ev["alerta_serial_mixto"] = True
    return ev


def _verif_cedula_serial(cedula: Optional[str], serial_norm: Optional[str]) -> str:
    """Texto corto de auditoría para observación (factores críticos)."""
    c = _cedula_canon_match(cedula) if cedula else ""
    s = _serial_norm_comparacion(serial_norm) if serial_norm else (serial_norm or "").strip()
    partes = []
    if c:
        partes.append(f"cedula={c}")
    if s:
        partes.append(f"serial={s}")
    return " ".join(partes)


def _cedula_coincide_prestamo(
    db: Session, prest: Prestamo, cedula_fila: str
) -> bool:
    """True si la cédula del extracto alinea con prestamo.cedula o cliente.cedula."""
    target = _cedula_canon_match(cedula_fila)
    if not target:
        return False
    candidatos: list[str] = []
    if prest.cedula:
        candidatos.append(str(prest.cedula))
    if prest.cliente_id:
        cli = db.get(Cliente, int(prest.cliente_id))
        if cli and cli.cedula:
            candidatos.append(str(cli.cedula))
    for raw in candidatos:
        if _cedula_canon_match(raw) == target:
            return True
        # Fallback dígitos solo si misma letra o uno sin letra canónica
        if _digitos_cedula_canon(raw) == _digitos_cedula_canon(target):
            a = _cedula_canon_match(raw)
            if a and target and a[0] == target[0]:
                return True
    return False


def _cedula_canon_match(value: Optional[str]) -> str:
    """
    Clave de match extracto ↔ sistema.

    Excel banco suele traer ceros tras la letra (V-015276832); en cartera suele
    estar V15276832. Se ignoran guión/espacios y ceros a la izquierda del número.
    """
    s = texto_cedula_comparable_bd(value)
    if not s:
        return ""
    if len(s) >= 2 and s[0] in ("V", "E", "G", "J") and s[1:].isdigit():
        return s[0] + (s[1:].lstrip("0") or "0")
    if s.isdigit():
        return "V" + (s.lstrip("0") or "0")
    return s


def _digitos_cedula_canon(value: Optional[str]) -> str:
    """Solo dígitos de la cédula canónica (sin ceros a la izquierda)."""
    c = _cedula_canon_match(value)
    if len(c) >= 2 and c[0] in ("V", "E", "G", "J"):
        return c[1:]
    return _solo_digitos(c).lstrip("0") or ""


def _parse_fecha(val: Any) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


_RE_SERIAL_CIENTIFICO = re.compile(r"^\d\.\d+[Ee]\+\d+$")


def _texto_serial_excel(val: Any) -> str:
    """Referencia/serial del extracto: entero, sin .0 ni notación científica de Excel."""
    if val is None or val == "":
        return ""
    if isinstance(val, bool):
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, (float, Decimal)):
        f = float(val)
        if f != f:  # NaN
            return ""
        # Seriales largos: preferir notación fija (evita 7.40E+14 en str intermedios).
        if abs(f) >= 1e10 and abs(f - round(f)) < 1e-3:
            try:
                return str(int(Decimal(repr(f)).to_integral_value()))
            except Exception:
                pass
            s_fix = format(f, ".0f")
            if s_fix and "e" not in s_fix.lower():
                return s_fix.lstrip("0") or "0"
        if abs(f - round(f)) < 1e-6:
            return str(int(round(f)))
        s = format(f, ".15g").replace(",", ".")
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s
    s = str(val).strip()
    if not s or s.upper() in ("NA", "N/A", "NONE", "NULL"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    if _RE_SERIAL_CIENTIFICO.match(s) or re.search(r"[Ee][+-]?\d+", s):
        try:
            d = Decimal(s.replace(",", "."))
            if d == d.to_integral_value():
                return str(int(d))
            s_fix = format(float(d), ".0f")
            if s_fix and "e" not in s_fix.lower():
                return s_fix
        except Exception:
            pass
    if re.fullmatch(r"\d+\.\d+", s):
        try:
            f = float(s)
            if abs(f - round(f)) < 1e-6:
                return str(int(round(f)))
        except ValueError:
            pass
    return s


def _serial_excel_parece_corrupto(serial_raw: str) -> bool:
    """True si Excel guardó el serial como número (notación científica / pocos dígitos)."""
    s = (serial_raw or "").strip()
    if not s:
        return False
    if _RE_SERIAL_CIENTIFICO.match(s) or re.search(r"[Ee][+-]\d+", s):
        return True
    dig = _solo_digitos(s)
    # Referencias bancarias típicas ≥12 dígitos; 6 dígitos visibles = redondeo Excel.
    if dig and len(dig) < 10:
        return True
    return False


def _validar_seriales_solo_serial(parsed: list[dict[str, Any]]) -> None:
    """Rechaza lotes donde Excel destruyó los seriales (todos 7.40E+14, etc.)."""
    seriales = [
        _texto_serial_excel(it.get("serial_raw"))
        for it in parsed
        if _texto_serial_excel(it.get("serial_raw"))
    ]
    if not seriales:
        return
    corruptos = sum(1 for s in seriales if _serial_excel_parece_corrupto(s))
    if corruptos >= max(3, len(seriales) // 10):
        raise HTTPException(
            status_code=400,
            detail=(
                "Referencia corrupta: Excel guardó los seriales como número "
                "(notación 7.40E+14). Antes de pegar: seleccione columna Referencia → "
                "Formato Texto, pegue de nuevo, o guarde como CSV. "
                "Si ya pegó como número, debe reexportar desde el origen."
            ),
        )
    from collections import Counter

    top, cnt = Counter(seriales).most_common(1)[0]
    if cnt > max(5, len(seriales) // 2) and _serial_excel_parece_corrupto(top):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Demasiadas filas con la misma Referencia truncada ({top}). "
                "Excel perdió dígitos al guardar como número. "
                "Use columna Referencia en formato Texto antes de pegar los seriales."
            ),
        )


def _leer_celda_referencia_excel(cell: Any) -> Any:
    """Lee celda openpyxl priorizando texto almacenado (no float de Excel)."""
    v = getattr(cell, "value", cell)
    if v is None:
        return None
    if isinstance(v, str):
        return _texto_serial_excel(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return _texto_serial_excel(v)
    return _texto_serial_excel(v)


def _parse_monto(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return round(float(val), 2)
    s = str(val).strip().replace("+", "").replace(" ", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def extraer_cedula_descripcion(texto: Optional[str]) -> Optional[str]:
    """Extrae cédula de 'DP:V-019200177 JOSE…' sin inventar datos.

    Normaliza a letra + número sin ceros a la izquierda (V-015276832 → V15276832).
    """
    raw = (texto or "").strip()
    if not raw:
        return None
    m = _RE_CEDULA.search(raw) or _RE_CEDULA_ALT.search(raw)
    if not m:
        return None
    clave = f"{m.group(1).upper()}{m.group(2)}"
    canon = _cedula_canon_match(clave)
    return canon or normalizar_cedula_almacenamiento(clave) or clave


def _cedulas_filtro_indice(*cedulas: Optional[str]) -> set[str]:
    """Variantes de cédula para acotar el índice APROBADO al subir/revalidar."""
    out: set[str] = set()
    for raw in cedulas:
        if not raw:
            continue
        out.add(str(raw).strip())
        canon = _cedula_canon_match(raw)
        if canon:
            out.add(canon)
        dig = _digitos_cedula_canon(canon or raw)
        if dig:
            out.add(dig)
            if canon and len(canon) >= 2 and canon[0] in ("V", "E", "G", "J"):
                out.add(f"{canon[0]}{dig}")
                out.add(f"{canon[0]}:{dig}")
    return {x for x in out if x}


def _similitud_serial(a: str, b: str) -> float:
    """Similitud serial numérica (misma regla que Conciliación Bancos)."""
    from app.services.conciliacion_bancos_service import _similitud as _sim_cb

    return float(_sim_cb(a, b))


def _seriales_extracto_comparar(serial_raw: str, serial_norm: str) -> list[str]:
    """
    Claves del extracto a comparar vs pagos del préstamo.

    Un solo serial → una clave. Celda con 2+ partes → cada parte (evita falso
    negativo si el humano pegó varios en el Excel).
    """
    partes = _partes_serial_texto(serial_raw)
    if len(partes) >= 2:
        return partes
    if serial_norm:
        return [serial_norm]
    return []


def _buscar_igual_100_en_prestamo(
    pagos: list[tuple[int, str]], seriales: list[str]
) -> Optional[tuple[int, str]]:
    """Primer pago_id con match exacto cédula+serial (cualquier clave del extracto)."""
    if not seriales:
        return None
    targets = set(seriales)
    seen: set[int] = set()
    for pago_id, sp in pagos:
        if sp in targets and pago_id not in seen:
            seen.add(pago_id)
            return int(pago_id), sp
    return None


def _mejor_similitud_serial_en_prestamo(
    pagos: list[tuple[int, str]], seriales: list[str]
) -> tuple[float, Optional[int], Optional[str]]:
    """Mejor % similitud serial vs pagos del préstamo (todas las claves extracto)."""
    best_pct = 0.0
    best_pid: Optional[int] = None
    best_sp: Optional[str] = None
    seen: set[tuple[int, str]] = set()
    for sn in seriales:
        if not sn or len(sn) < _MIN_DIGITOS_SERIAL:
            continue
        for pago_id, sp in pagos:
            if not sp or len(sp) < _MIN_DIGITOS_SERIAL:
                continue
            key = (pago_id, sp)
            if key in seen:
                continue
            seen.add(key)
            pct = _similitud_serial(sn, sp)
            if pct > best_pct:
                best_pct = pct
                best_pid = int(pago_id)
                best_sp = sp
    return best_pct, best_pid, best_sp


def _lote_modo_confirmado(lote: ImportacionExtractoLote) -> bool:
    """True si el lote importa solo a estadística Pagos confirmados (serial sin cédula)."""
    return bool(getattr(lote, "modo_serial", False)) and not bool(
        getattr(lote, "modo_cedula", True)
    )


def invalidate_serial_cartera_cache() -> None:
    """Tras importar pago/confirmado, el índice serial global debe refrescarse."""
    global _serial_cartera_cache
    with _serial_cartera_cache_lock:
        _serial_cartera_cache = None


def _construir_indice_serial_cartera_sql(
    db: Session, filtro: set[str]
) -> dict[str, Any]:
    """Índice serial acotado por SQL (evita full scan de pagos con lotes 20k+)."""
    pagos_global: dict[str, list[tuple[int, Optional[int]]]] = {}
    confirmados_activos: dict[str, list[int]] = {}
    keys = sorted(s for s in filtro if s and len(s) >= _MIN_DIGITOS_SERIAL)
    if not keys:
        return {"pagos_global": pagos_global, "confirmados_activos": confirmados_activos}

    for i in range(0, len(keys), _SERIAL_SQL_BATCH):
        batch = keys[i : i + _SERIAL_SQL_BATCH]
        rows = db.execute(
            select(
                Pago.id,
                Pago.prestamo_id,
                Pago.numero_documento,
                Pago.referencia_pago,
                Pago.ref_norm,
                Pago.doc_canon_numero,
                Pago.doc_canon_referencia,
                Pago.institucion_bancaria,
            ).where(
                or_(
                    Pago.ref_norm.in_(batch),
                    Pago.doc_canon_numero.in_(batch),
                    Pago.doc_canon_referencia.in_(batch),
                    Pago.referencia_pago.in_(batch),
                    Pago.numero_documento.in_(batch),
                )
            )
        ).all()
        for (
            pago_id,
            prestamo_id,
            num_doc,
            ref,
            ref_n,
            doc_c,
            doc_cr,
            institucion,
        ) in rows:
            if _es_pago_banco_drive(institucion, num_doc, ref):
                continue
            ipago = int(pago_id)
            ipid = int(prestamo_id) if prestamo_id is not None else None
            for dig in _seriales_norm_desde_campos(num_doc, ref, ref_n, doc_c, doc_cr):
                if dig in filtro:
                    pagos_global.setdefault(dig, []).append((ipago, ipid))

        conf_rows = db.execute(
            select(
                ImportacionExtractoPagoConfirmado.id,
                ImportacionExtractoPagoConfirmado.serial_norm,
                ImportacionExtractoPagoConfirmado.serial,
            ).where(
                ImportacionExtractoPagoConfirmado.estado == "ACTIVO",
                or_(
                    ImportacionExtractoPagoConfirmado.serial_norm.in_(batch),
                    ImportacionExtractoPagoConfirmado.serial.in_(batch),
                ),
            )
        ).all()
        for conf_id, serial_norm, serial_raw in conf_rows:
            sn = _serial_norm_comparacion(serial_norm or serial_raw)
            if not sn:
                continue
            if sn in filtro:
                confirmados_activos.setdefault(sn, []).append(int(conf_id))
            for part in _seriales_extracto_comparar(serial_raw or "", sn):
                if part in filtro:
                    confirmados_activos.setdefault(part, []).append(int(conf_id))

    return {
        "pagos_global": pagos_global,
        "confirmados_activos": confirmados_activos,
    }


def _construir_indice_serial_cartera(
    db: Session,
    *,
    force_refresh: bool = False,
    serials_filtro: Optional[set[str]] = None,
) -> dict[str, Any]:
    """Índice global de seriales en cartera (pagos reales + confirmados ACTIVO).

    Excluye pagos Drive/ABONOS. Usado en modo solo Serial del extracto.
    Con ``serials_filtro`` solo indexa claves del extracto (más rápido en lotes grandes).
    """
    global _serial_cartera_cache
    use_cache = serials_filtro is None and not force_refresh
    if use_cache:
        with _serial_cartera_cache_lock:
            hit = _serial_cartera_cache
            if hit is not None and time_mod.monotonic() - hit[0] < _SERIAL_CARTERA_CACHE_TTL_SEC:
                return copy.deepcopy(hit[1])

    pagos_global: dict[str, list[tuple[int, Optional[int]]]] = {}
    confirmados_activos: dict[str, list[int]] = {}
    filtro = serials_filtro or None

    if filtro is not None and len(filtro) >= 500:
        return _construir_indice_serial_cartera_sql(db, filtro)

    last_id = 0
    while True:
        rows = db.execute(
            select(
                Pago.id,
                Pago.prestamo_id,
                Pago.numero_documento,
                Pago.referencia_pago,
                Pago.ref_norm,
                Pago.doc_canon_numero,
                Pago.doc_canon_referencia,
                Pago.institucion_bancaria,
            )
            .where(Pago.id > last_id)
            .order_by(Pago.id)
            .limit(_PAGOS_SERIAL_CHUNK)
        ).all()
        if not rows:
            break
        for (
            pago_id,
            prestamo_id,
            num_doc,
            ref,
            ref_n,
            doc_c,
            doc_cr,
            institucion,
        ) in rows:
            last_id = int(pago_id)
            if _es_pago_banco_drive(institucion, num_doc, ref):
                continue
            ipago = int(pago_id)
            ipid = int(prestamo_id) if prestamo_id is not None else None
            for dig in _seriales_norm_desde_campos(num_doc, ref, ref_n, doc_c, doc_cr):
                if filtro is not None and dig not in filtro:
                    continue
                pagos_global.setdefault(dig, []).append((ipago, ipid))

    conf_rows = db.execute(
        select(
            ImportacionExtractoPagoConfirmado.id,
            ImportacionExtractoPagoConfirmado.serial_norm,
            ImportacionExtractoPagoConfirmado.serial,
        ).where(ImportacionExtractoPagoConfirmado.estado == "ACTIVO")
    ).all()
    for conf_id, serial_norm, serial_raw in conf_rows:
        sn = _serial_norm_comparacion(serial_norm or serial_raw)
        if not sn:
            continue
        if filtro is not None:
            parts = set(_seriales_extracto_comparar(serial_raw or "", sn))
            if sn not in filtro and not (parts & filtro):
                continue
        confirmados_activos.setdefault(sn, []).append(int(conf_id))
        for part in _seriales_extracto_comparar(serial_raw or "", sn):
            if filtro is None or part in filtro:
                confirmados_activos.setdefault(part, []).append(int(conf_id))

    result = {
        "pagos_global": pagos_global,
        "confirmados_activos": confirmados_activos,
    }
    if use_cache:
        with _serial_cartera_cache_lock:
            _serial_cartera_cache = (time_mod.monotonic(), copy.deepcopy(result))
    return result


def _buscar_igual_100_global(
    idx: dict[str, Any], seriales: list[str]
) -> Optional[tuple[str, Optional[int], Optional[int], Optional[int]]]:
    """Match exacto serial vs cartera global o confirmado ACTIVO.

    Retorna (clave_match, pago_id, prestamo_id, confirmado_id).
    """
    if not seriales:
        return None
    targets = set(seriales)
    pagos_global = idx.get("pagos_global") or {}
    confirmados = idx.get("confirmados_activos") or {}
    for sp in seriales:
        if sp in pagos_global:
            pago_id, prestamo_id = pagos_global[sp][0]
            return sp, int(pago_id), prestamo_id, None
        if sp in confirmados:
            conf_id = int(confirmados[sp][0])
            return sp, None, None, conf_id
    for sp in targets:
        if sp in pagos_global:
            pago_id, prestamo_id = pagos_global[sp][0]
            return sp, int(pago_id), prestamo_id, None
        if sp in confirmados:
            conf_id = int(confirmados[sp][0])
            return sp, None, None, conf_id
    return None


def _mejor_similitud_serial_global(
    idx: dict[str, Any], seriales: list[str]
) -> tuple[float, Optional[int], Optional[str]]:
    """Mejor % similitud serial vs todos los pagos de cartera (global)."""
    best_pct = 0.0
    best_pid: Optional[int] = None
    best_sp: Optional[str] = None
    pagos_global = idx.get("pagos_global") or {}
    for sn in seriales:
        if not sn or len(sn) < _MIN_DIGITOS_SERIAL:
            continue
        for sp, lst in pagos_global.items():
            if not sp or len(sp) < _MIN_DIGITOS_SERIAL:
                continue
            pct = _similitud_serial(sn, sp)
            if pct > best_pct:
                best_pct = pct
                best_pid = int(lst[0][0]) if lst else None
                best_sp = sp
    return best_pct, best_pid, best_sp


def _evaluar_fila_serial_cartera(
    idx: dict[str, Any],
    *,
    fecha: Optional[date],
    serial_raw: str,
    monto: Optional[float],
    allow_semejante: bool = True,
) -> dict[str, Any]:
    """Modo solo Serial: comparar vs cartera global (sin cédula ni préstamo)."""
    serial_norm = _serial_norm_comparacion(serial_raw)
    seriales_cmp = _seriales_extracto_comparar(serial_raw, serial_norm)

    if not serial_norm or not seriales_cmp:
        return {
            "cedula": None,
            "serial_norm": None,
            "estado": "PARSE_ERROR",
            "detalle": "Referencia/serial vacío",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
            "destino_importacion": None,
        }
    if any(len(s) < _MIN_DIGITOS_SERIAL for s in seriales_cmp):
        return {
            "cedula": None,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": f"Serial demasiado corto (mín {_MIN_DIGITOS_SERIAL} dígitos)",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
            "destino_importacion": None,
        }
    if monto is None or monto <= 0:
        return {
            "cedula": None,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": "Monto Haber inválido o <= 0",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
            "destino_importacion": None,
        }
    if fecha is None:
        return {
            "cedula": None,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": "Fecha inválida",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
            "destino_importacion": None,
        }

    match = _buscar_igual_100_global(idx, seriales_cmp)
    if match is not None:
        sp_match, pago_id, _prestamo_id, conf_id = match
        det_extra = ""
        if len(seriales_cmp) > 1:
            det_extra = f"; clave extracto={serial_norm} match parcial={sp_match}"
        if conf_id is not None:
            det = (
                f"100% serial ya confirmado sin cédula (confirmado_id={conf_id}"
                f"{det_extra})"
            )
        else:
            det = (
                f"100% serial ya en cartera (pago_id={pago_id}"
                f"; prefijos BNC/letras ignorados{det_extra})"
            )
        return {
            "cedula": None,
            "serial_norm": serial_norm,
            "estado": "IGUAL_100",
            "detalle": det,
            "similitud_pct": 100.0,
            "pago_id_match": pago_id,
            "prestamo_id": None,
            "destino_importacion": None,
            "omitir_lista": True,
        }

    best_pct, best_pid, _best_sp = (0.0, None, None)
    if allow_semejante:
        best_pct, best_pid, _best_sp = _mejor_similitud_serial_global(idx, seriales_cmp)
    if allow_semejante and best_pct >= _SIMILITUD_SERIAL_MINIMA:
        return {
            "cedula": None,
            "serial_norm": serial_norm,
            "estado": "SEMEJANTE",
            "detalle": (
                f"Serial semejante {best_pct}% al pago_id={best_pid} "
                f"(cartera global); importar con OK bajo criterio"
            ),
            "similitud_pct": best_pct,
            "pago_id_match": best_pid,
            "prestamo_id": None,
            "destino_importacion": "CONFIRMADO",
        }

    return {
        "cedula": None,
        "serial_norm": serial_norm,
        "estado": "SE_PUEDE_IMPORTAR",
        "detalle": (
            "Serial ausente en cartera global; confiabilidad confirmación 100%"
        ),
        "similitud_pct": 100.0,
        "pago_id_match": None,
        "prestamo_id": None,
        "destino_importacion": "CONFIRMADO",
    }


def _retirar_confirmado_activo_por_serial(
    db: Session,
    serial_raw: str,
    serial_norm: str,
    *,
    pago_id: int,
    fila_id: Optional[int] = None,
    prestamo_id: Optional[int] = None,
    cedula: Optional[str] = None,
    monto_usd: Optional[float] = None,
) -> dict[str, Any]:
    """Marca el confirmado ACTIVO que corresponde al serial (1 pago ↔ 1 confirmado).

    Empareja por claves de serial (partes compuestas) y prefiere monto más cercano.
    Persiste prestamo_id y cédula para auditoría y evitar confusión de casos.
    """
    seriales_cmp = set(_seriales_extracto_comparar(serial_raw, serial_norm))
    sn = _serial_norm_comparacion(serial_norm)
    if sn:
        seriales_cmp.add(sn)
    if not seriales_cmp:
        return {"retirados": 0, "confirmado_ids": [], "advertencias": []}

    candidatos = (
        db.execute(
            select(ImportacionExtractoPagoConfirmado).where(
                ImportacionExtractoPagoConfirmado.estado == "ACTIVO",
            )
        )
        .scalars()
        .all()
    )
    matched: list[ImportacionExtractoPagoConfirmado] = []
    for c in candidatos:
        c_sn = _serial_norm_comparacion(c.serial_norm or c.serial)
        if c_sn and c_sn in seriales_cmp:
            matched.append(c)
            continue
        c_parts = set(_seriales_extracto_comparar(c.serial or "", c_sn or ""))
        if seriales_cmp & c_parts:
            matched.append(c)

    advertencias: list[str] = []
    if not matched:
        return {"retirados": 0, "confirmado_ids": [], "advertencias": advertencias}

    if monto_usd is not None:
        matched.sort(
            key=lambda c: (
                abs(float(c.monto_usd or 0) - float(monto_usd)),
                -int(c.id),
            )
        )
        best = matched[0]
        diff = abs(float(best.monto_usd or 0) - float(monto_usd))
        if diff > 0.05:
            advertencias.append(
                f"monto confirmado id={best.id} (${float(best.monto_usd):.2f}) "
                f"≠ pago (${float(monto_usd):.2f})"
            )
        aplicar = [best]
        if len(matched) > 1:
            advertencias.append(
                f"{len(matched)} confirmados ACTIVO mismo serial; "
                f"aplicado id={best.id}"
            )
    else:
        aplicar = [matched[0]]
        if len(matched) > 1:
            advertencias.append(
                f"{len(matched)} confirmados ACTIVO mismo serial; "
                f"aplicado id={matched[0].id}"
            )

    cedula_canon = _cedula_canon_match(cedula) if cedula else None
    ahora = datetime.utcnow()
    ids: list[int] = []
    for c in aplicar:
        c.estado = "APLICADO_PRESTAMO"
        c.pago_id = int(pago_id)
        c.fila_id_aplicacion = int(fila_id) if fila_id else None
        if prestamo_id is not None:
            c.prestamo_id = int(prestamo_id)
        if cedula_canon:
            c.cedula = cedula_canon
        c.aplicado_en = ahora
        nota = (
            f"Aplicado prestamo_id={prestamo_id} pago_id={pago_id}"
            + (f" cedula={cedula_canon}" if cedula_canon else "")
        )
        c.detalle = ((c.detalle or "") + f" | {nota}").strip(" |")
        ids.append(int(c.id))

    return {"retirados": len(ids), "confirmado_ids": ids, "advertencias": advertencias}


def _load_confirmados_activos(db: Session) -> list[ImportacionExtractoPagoConfirmado]:
    return list(
        db.execute(
            select(ImportacionExtractoPagoConfirmado).where(
                ImportacionExtractoPagoConfirmado.estado == "ACTIVO",
            )
        )
        .scalars()
        .all()
    )


def _construir_indice_confirmados_activos(
    db: Session,
) -> dict[str, Any]:
    """Índice O(1) serial → confirmados ACTIVO (evita O(filas×confirmados) en lote)."""
    activos = _load_confirmados_activos(db)
    por_serial: dict[str, list[ImportacionExtractoPagoConfirmado]] = {}
    for c in activos:
        c_sn = _serial_norm_comparacion(c.serial_norm or c.serial)
        keys = set(_seriales_extracto_comparar(c.serial or "", c_sn or ""))
        if c_sn:
            keys.add(c_sn)
        for k in keys:
            if not k:
                continue
            por_serial.setdefault(k, []).append(c)
    return {"activos": activos, "por_serial": por_serial}


def _confirmados_activos_para_seriales(
    activos: list[ImportacionExtractoPagoConfirmado],
    seriales_cmp: list[str],
    *,
    idx_confirmados: Optional[dict[str, Any]] = None,
) -> list[ImportacionExtractoPagoConfirmado]:
    """Confirmados ACTIVO cuyo serial coincide (exacto o parte compuesta)."""
    if not seriales_cmp:
        return []
    if idx_confirmados is not None:
        por_serial = idx_confirmados.get("por_serial") or {}
        seen: set[int] = set()
        out: list[ImportacionExtractoPagoConfirmado] = []
        for sp in seriales_cmp:
            for c in por_serial.get(sp, []):
                cid = int(c.id)
                if cid in seen:
                    continue
                seen.add(cid)
                out.append(c)
        out.sort(key=lambda c: int(c.id))
        return out
    targets = set(seriales_cmp)
    out: list[ImportacionExtractoPagoConfirmado] = []
    seen: set[int] = set()
    for c in activos:
        cid = int(c.id)
        if cid in seen:
            continue
        c_sn = _serial_norm_comparacion(c.serial_norm or c.serial)
        parts = set(_seriales_extracto_comparar(c.serial or "", c_sn or ""))
        if (c_sn and c_sn in targets) or (parts & targets):
            seen.add(cid)
            out.append(c)
    out.sort(key=lambda c: int(c.id))
    return out


def _anotar_confirmado_pendiente_en_ev(
    ev: dict[str, Any],
    pendientes: list[ImportacionExtractoPagoConfirmado],
    prestamo_id: Optional[int],
) -> dict[str, Any]:
    if not pendientes:
        return ev
    ids = [int(c.id) for c in pendientes]
    montos = [float(c.monto_usd or 0) for c in pendientes]
    ev = dict(ev)
    ev["alerta_confirmado_pendiente"] = True
    ev["confirmado_pendiente_ids"] = ids
    ev["confirmado_pendiente_montos"] = montos
    pid_txt = f"prestamo_id={prestamo_id}" if prestamo_id else "préstamo APROBADO"
    ev["detalle"] = (
        (ev.get("detalle") or "")
        + (
            f"; Serial confirmado sin cédula (confirmado_id={ids[0]} "
            f"${montos[0]:.2f}"
            + (f", +{len(ids)-1} más" if len(ids) > 1 else "")
            + f"); al OK → {pid_txt} y sale de Pagos confirmados"
        )
    ).strip("; ")
    return ev


def _enriquecer_filas_confirmado_pendiente(
    db: Session,
    filas: list[ImportacionExtractoFila],
    *,
    idx_confirmados: Optional[dict[str, Any]] = None,
) -> dict[int, dict[str, Any]]:
    """Batch: filas con cédula+serial que tienen confirmado ACTIVO pendiente."""
    pendientes = [f for f in filas if not f.importado and f.serial_norm]
    if not pendientes:
        return {}
    if idx_confirmados is None:
        idx_confirmados = _construir_indice_confirmados_activos(db)
    if not idx_confirmados.get("activos"):
        return {}
    out: dict[int, dict[str, Any]] = {}
    for f in pendientes:
        seriales = _seriales_extracto_comparar(
            f.serial or "", _serial_norm_comparacion(f.serial_norm or f.serial)
        )
        pend = _confirmados_activos_para_seriales(
            idx_confirmados["activos"], seriales, idx_confirmados=idx_confirmados
        )
        if not pend:
            continue
        out[int(f.id)] = {
            "alerta_confirmado_pendiente": True,
            "confirmado_pendiente_ids": [int(c.id) for c in pend],
            "confirmado_pendiente_montos": [float(c.monto_usd or 0) for c in pend],
            "prestamo_destino_id": int(f.prestamo_id) if f.prestamo_id else None,
        }
    return out


def _cell(row: tuple, idx: int) -> Any:
    return row[idx] if len(row) > idx else None


def _magic_excel(raw: bytes) -> str:
    """Clasifica bytes del archivo para mensajes y motor de lectura."""
    if not raw:
        return "vacio"
    head = raw[:8]
    if head[:2] == b"PK":
        return "xlsx_zip"
    if head[:4] == b"\xd0\xcf\x11\xe0":
        return "xls_ole"
    soft = raw.lstrip()[:200].lower()
    if soft.startswith(b"<?xml") or b"spreadsheetml" in soft or b"<workbook" in soft:
        return "xls_xml"
    if soft.startswith(b"<html") or soft.startswith(b"<!doctype html"):
        return "html"
    # CSV / TSV texto
    try:
        sample = raw[:4096].decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            sample = raw[:4096].decode("latin-1")
        except Exception:
            return "desconocido"
    if "," in sample or ";" in sample or "\t" in sample:
        return "csv"
    return "desconocido"


def _rows_from_openpyxl(
    raw: bytes, *, columnas_serial: Optional[set[int]] = None
) -> list[tuple]:
    """Lee xlsx; columnas_serial lee Referencia como texto (evita 7.40E+14)."""
    cols = columnas_serial or set()
    if cols:
        wb = load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                raise ValueError("Excel sin hoja activa")
            out: list[tuple] = []
            for row in ws.iter_rows():
                cells: list[Any] = []
                for i, cell in enumerate(row):
                    if i in cols:
                        cells.append(_leer_celda_referencia_excel(cell))
                    else:
                        cells.append(cell.value)
                out.append(tuple(cells))
            return out
        finally:
            wb.close()
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel sin hoja activa")
        out: list[tuple] = []
        for row in ws.iter_rows(values_only=True):
            out.append(tuple(row) if row is not None else tuple())
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _rows_from_xlrd(raw: bytes) -> list[tuple]:
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    out: list[tuple] = []
    for r in range(sheet.nrows):
        vals: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            # xlrd date type
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    vals.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    vals.append(cell.value)
            else:
                vals.append(cell.value)
        out.append(tuple(vals))
    return out


def _rows_from_pandas(raw: bytes, filename: str) -> list[tuple]:
    import pandas as pd

    name = (filename or "").lower()
    bio = io.BytesIO(raw)
    if name.endswith(".csv") or _magic_excel(raw) == "csv":
        # Separador ; típico en exports LatAm / banco
        try:
            df = pd.read_csv(bio, sep=None, engine="python")
        except Exception:
            bio.seek(0)
            df = pd.read_csv(bio, sep=";", engine="python")
    else:
        engine = "xlrd" if name.endswith(".xls") and not name.endswith(".xlsx") else None
        try:
            df = pd.read_excel(bio, engine=engine)
        except Exception:
            bio.seek(0)
            # Reintento forzado xlrd (bancos renombran .xls → .xlsx)
            df = pd.read_excel(bio, engine="xlrd")
    # Incluir encabezado como fila 0 para que min_row=2 siga siendo datos
    header = tuple("" if c is None else str(c) for c in df.columns.tolist())
    rows: list[tuple] = [header]
    for _, series in df.iterrows():
        rows.append(tuple(None if (isinstance(v, float) and v != v) else v for v in series.tolist()))
    return rows


def _rows_from_html(raw: bytes) -> list[tuple]:
    import pandas as pd

    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    tables = pd.read_html(io.StringIO(text))
    if not tables:
        raise ValueError("HTML sin tablas")
    df = tables[0]
    header = tuple("" if c is None else str(c) for c in df.columns.tolist())
    rows: list[tuple] = [header]
    for _, series in df.iterrows():
        rows.append(tuple(None if (isinstance(v, float) and v != v) else v for v in series.tolist()))
    return rows


def _cargar_filas_excel(
    raw: bytes, filename: Optional[str], *, solo_serial: bool = False
) -> list[tuple]:
    """
    Lee extracto bancario en varios formatos reales:
    .xlsx (OOXML), .xls (OLE), XML Spreadsheet, HTML-as-xls, CSV.
    """
    serial_cols = {2, 6} if solo_serial else {6}
    kind = _magic_excel(raw)
    name = (filename or "extracto.xlsx").strip()
    errors: list[str] = []

    # 1) OOXML verdadero
    if kind == "xlsx_zip":
        try:
            return _rows_from_openpyxl(raw, columnas_serial=serial_cols)
        except Exception as e:
            errors.append(f"openpyxl: {e}")
            try:
                return _rows_from_pandas(raw, name)
            except Exception as e2:
                errors.append(f"pandas: {e2}")

    # 2) Excel 97-2003 OLE (.xls)
    if kind == "xls_ole" or name.lower().endswith(".xls"):
        try:
            return _rows_from_xlrd(raw)
        except Exception as e:
            errors.append(f"xlrd: {e}")
            try:
                return _rows_from_pandas(raw, name if name.lower().endswith(".xls") else "a.xls")
            except Exception as e2:
                errors.append(f"pandas-xls: {e2}")

    # 3) HTML exportado como .xls/.xlsx
    if kind == "html" or (kind == "desconocido" and b"<table" in raw[:8000].lower()):
        try:
            return _rows_from_html(raw)
        except Exception as e:
            errors.append(f"html: {e}")

    # 4) CSV / XML / último recurso pandas
    if kind in ("csv", "xls_xml", "desconocido", "xlsx_zip"):
        try:
            return _rows_from_pandas(raw, name)
        except Exception as e:
            errors.append(f"pandas-fallback: {e}")

    if kind == "xls_ole":
        raise HTTPException(
            status_code=400,
            detail=(
                "El archivo es Excel antiguo (.xls). No se pudo leer. "
                "Ábralo en Excel y guarde como .xlsx, o reintente tras el deploy con soporte .xls. "
                f"Detalle: {'; '.join(errors)[:240]}"
            ),
        )
    if kind == "xlsx_zip":
        raise HTTPException(
            status_code=400,
            detail=(
                "Excel .xlsx dañado o incompleto (sin workbook válido). "
                "Vuelva a exportar/guardar como Libro de Excel (.xlsx). "
                f"Detalle: {'; '.join(errors)[:240]}"
            ),
        )
    raise HTTPException(
        status_code=400,
        detail=(
            "No se pudo leer el archivo como Excel/CSV. "
            "Use .xlsx (recomendado), .xls o CSV del extracto bancario. "
            f"Tipo detectado={kind}. Detalle: {'; '.join(errors)[:240]}"
        ),
    )


def _prestamos_aprobados_cedula(db: Session, cedula: str) -> list[Prestamo]:
    """Lookup puntual (p.ej. importar una fila). Preferir índice en lote Excel."""
    idx = _construir_indice_aprobado(db)
    pids = _prestamo_ids_para_cedula(idx, cedula)
    if not pids:
        return []
    return list(
        db.execute(select(Prestamo).where(Prestamo.id.in_(pids)).order_by(Prestamo.id))
        .scalars()
        .all()
    )


def _construir_indice_aprobado(
    db: Session,
    *,
    cedulas_filtro: Optional[set[str]] = None,
) -> dict[str, Any]:
    """
    Solo préstamos **APROBADO** (+ pagos).

    Excluidos de lista y comparación (no disponibles): LIQUIDADO, DESISTIMIENTO
    y alias (DESESTIMADO, DESISTIDO). Cualquier otro estado tampoco entra.

    Misma cédula con varios APROBADO → se elige el de ``fecha_aprobacion`` más
    reciente (empate: id mayor). Si ``cedulas_filtro`` (Excel), solo carga pagos
    de esos préstamos.

    ``pids_al_dia`` marca los préstamos cuya última cuota vencida está Pagado
    (mismo estado que la columna Estado del listado de Préstamos). El extracto
    no importa depósitos contra ellos.
    """
    rows = db.execute(
        select(
            Prestamo.id,
            Prestamo.cedula,
            Cliente.cedula,
            Prestamo.fecha_aprobacion,
        )
        .outerjoin(Cliente, Prestamo.cliente_id == Cliente.id)
        .where(Prestamo.estado == "APROBADO")
        .order_by(Prestamo.id)
    ).all()

    fecha_por_pid: dict[int, datetime] = {}
    by_norm: dict[str, list[int]] = {}
    by_dig: dict[str, list[int]] = {}
    by_dig_letra: dict[str, list[int]] = {}

    def _add(key_map: dict[str, list[int]], key: str, pid: int) -> None:
        if not key:
            return
        lst = key_map.get(key)
        if lst is None:
            key_map[key] = [pid]
        elif pid not in lst:
            lst.append(pid)

    for pid, pced, cced, fapr in rows:
        ipid = int(pid)
        if isinstance(fapr, datetime):
            fecha_por_pid[ipid] = fapr
        elif fapr is not None:
            try:
                fecha_por_pid[ipid] = datetime.combine(fapr, datetime.min.time())
            except Exception:
                fecha_por_pid[ipid] = datetime.min
        else:
            fecha_por_pid[ipid] = datetime.min
        for raw in (pced, cced):
            if not raw:
                continue
            n = _cedula_canon_match(str(raw))
            _add(by_norm, n, ipid)
            dig = _digitos_cedula_canon(n)
            if dig:
                _add(by_dig, dig, ipid)
                if len(n) >= 2 and n[0] in ("V", "E", "G", "J"):
                    _add(by_dig_letra, f"{n[0]}:{dig}", ipid)

    def _elegir_aprobado_mas_reciente(pids: list[int]) -> Optional[int]:
        if not pids:
            return None
        return max(pids, key=lambda p: (fecha_por_pid.get(p, datetime.min), p))

    # 1 APROBADO por cédula = fecha_aprobacion más actual.
    by_norm_one: dict[str, list[int]] = {}
    for key, pids in by_norm.items():
        best = _elegir_aprobado_mas_reciente(pids)
        if best is not None:
            by_norm_one[key] = [best]
    by_dig_one: dict[str, list[int]] = {}
    for key, pids in by_dig.items():
        best = _elegir_aprobado_mas_reciente(pids)
        if best is not None:
            by_dig_one[key] = [best]
    by_dig_letra_one: dict[str, list[int]] = {}
    for key, pids in by_dig_letra.items():
        best = _elegir_aprobado_mas_reciente(pids)
        if best is not None:
            by_dig_letra_one[key] = [best]
    by_norm, by_dig, by_dig_letra = by_norm_one, by_dig_one, by_dig_letra_one

    if cedulas_filtro:
        filtro_norm: set[str] = set()
        filtro_dig: set[str] = set()
        filtro_dig_letra: set[str] = set()
        for raw in cedulas_filtro:
            n = _cedula_canon_match(raw)
            if n:
                filtro_norm.add(n)
            d = _digitos_cedula_canon(n or raw)
            if d:
                filtro_dig.add(d)
                if n and len(n) >= 2 and n[0] in ("V", "E", "G", "J"):
                    filtro_dig_letra.add(f"{n[0]}:{d}")
        pids_set: set[int] = set()
        for n in filtro_norm:
            for pid in by_norm.get(n, []):
                pids_set.add(pid)
        for d in filtro_dig:
            for pid in by_dig.get(d, []):
                pids_set.add(pid)
        for k in filtro_dig_letra:
            for pid in by_dig_letra.get(k, []):
                pids_set.add(pid)
        all_pids = sorted(pids_set)
    else:
        all_pids = sorted({pid for lst in by_norm.values() for pid in lst})

    estados_ultima_cuota = estado_ultima_cuota_por_vencimiento(db, all_pids)
    pids_al_dia = {
        pid for pid, (codigo, _et) in estados_ultima_cuota.items() if codigo == "PAGADO"
    }

    pagos_by_prestamo: dict[int, list[tuple[int, str]]] = {pid: [] for pid in all_pids}
    drive_by_prestamo: dict[int, list[int]] = {pid: [] for pid in all_pids}
    mixto_by_prestamo: dict[int, list[int]] = {pid: [] for pid in all_pids}
    if all_pids:
        chunk = 800
        for i in range(0, len(all_pids), chunk):
            part = all_pids[i : i + chunk]
            pago_rows = db.execute(
                select(
                    Pago.id,
                    Pago.prestamo_id,
                    Pago.numero_documento,
                    Pago.referencia_pago,
                    Pago.ref_norm,
                    Pago.doc_canon_numero,
                    Pago.doc_canon_referencia,
                    Pago.institucion_bancaria,
                ).where(Pago.prestamo_id.in_(part))
            ).all()
            for (
                pago_id,
                prestamo_id,
                num_doc,
                ref,
                ref_n,
                doc_c,
                doc_cr,
                institucion,
            ) in pago_rows:
                if prestamo_id is None:
                    continue
                ipid = int(prestamo_id)
                ipago = int(pago_id)
                if _es_pago_banco_drive(institucion, num_doc, ref):
                    drive_by_prestamo.setdefault(ipid, []).append(ipago)
                # Serial mixto: 2+ seriales en un Nº documento (humano juntó justificación).
                if any(
                    _es_serial_mixto_texto(x)
                    for x in (num_doc, ref, ref_n, doc_c, doc_cr)
                    if x
                ):
                    mixto_by_prestamo.setdefault(ipid, []).append(ipago)
                for dig in _seriales_norm_desde_campos(
                    num_doc, ref, ref_n, doc_c, doc_cr
                ):
                    pagos_by_prestamo.setdefault(ipid, []).append((ipago, dig))

    return {
        "by_norm": by_norm,
        "by_dig": by_dig,
        "by_dig_letra": by_dig_letra,
        "pagos_by_prestamo": pagos_by_prestamo,
        "drive_by_prestamo": drive_by_prestamo,
        "mixto_by_prestamo": mixto_by_prestamo,
        "fecha_por_pid": fecha_por_pid,
        "pids_al_dia": pids_al_dia,
    }


def _parse_fila_excel_row(
    row: tuple,
    fila_excel: int,
    *,
    solo_serial: bool = False,
) -> Optional[dict[str, Any]]:
    """Extrae campos de una fila de Excel sin tocar BD. None si vacía.

    solo_serial: plantilla Fecha | cedula (vacía) | Referencia | Monto (4 columnas).
    extracto banco: Fecha | Descripción | … | Referencia (col G) | Haber (col H).
    """
    if not row:
        return None

    if solo_serial:
        c0 = str(_cell(row, 0) or "").strip().lower()
        c2 = str(_cell(row, 2) or "").strip().lower()
        if c0 in ("fecha", "date") and c2 in ("referencia", "serial", "ref"):
            return None
        fecha = _parse_fecha(_cell(row, 0))
        serial_raw = _texto_serial_excel(_cell(row, 2))
        monto = _parse_monto(_cell(row, 3))
        if not serial_raw and monto is None and fecha is None:
            return None
        return {
            "fila_excel": fila_excel,
            "fecha": fecha,
            "desc": "",
            "serial_raw": serial_raw,
            "monto": monto,
        }

    fecha = _parse_fecha(_cell(row, 0))
    desc = str(_cell(row, 1) or "").strip()
    serial_raw = _texto_serial_excel(_cell(row, 6))
    monto = _parse_monto(_cell(row, 7))
    if not serial_raw and len(row) >= 4:
        maybe_ced = str(_cell(row, 1) or "").strip()
        maybe_ser = _texto_serial_excel(_cell(row, 2))
        maybe_mon = _parse_monto(_cell(row, 3))
        if maybe_ser and maybe_mon is not None:
            desc = maybe_ced if ":" not in maybe_ced else desc
            if re.match(r"^[VEJG]-?\d", maybe_ced, re.I):
                desc = f"DP:{maybe_ced}"
            serial_raw = maybe_ser
            monto = maybe_mon
    if not desc and not serial_raw and monto is None and fecha is None:
        return None
    return {
        "fila_excel": fila_excel,
        "fecha": fecha,
        "desc": desc,
        "serial_raw": serial_raw,
        "monto": monto,
    }


def _prestamo_ids_para_cedula(idx: dict[str, Any], cedula: str) -> list[int]:
    c = _cedula_canon_match(cedula)
    if not c:
        return []
    by_norm: dict[str, list[int]] = idx["by_norm"]
    by_dig: dict[str, list[int]] = idx["by_dig"]
    by_dig_letra: dict[str, list[int]] = idx.get("by_dig_letra") or {}
    found = by_norm.get(c)
    if found:
        return list(found)
    dig = _digitos_cedula_canon(c)
    if dig and len(c) >= 2 and c[0] in ("V", "E", "G", "J"):
        found = by_dig_letra.get(f"{c[0]}:{dig}")
        if found:
            return list(found)
    if dig:
        found = by_dig.get(dig)
        if found:
            return list(found)
    return []


def _pagos_aprobados_cedula(db: Session, cedula: str, prestamo_ids: list[int]) -> list[Pago]:
    if not prestamo_ids:
        return []
    c = normalizar_cedula_almacenamiento(cedula) or cedula
    return list(
        db.execute(
            select(Pago).where(
                Pago.prestamo_id.in_(prestamo_ids),
                or_(Pago.cedula_cliente == c, Pago.cedula_cliente.is_(None)),
            )
        )
        .scalars()
        .all()
    )


def _serial_pago_digitos(p: Pago) -> str:
    """Primera clave numérica comparable del pago (prefijos BNC/ ignorados)."""
    serials = _seriales_norm_desde_campos(
        p.numero_documento,
        p.referencia_pago,
        p.ref_norm,
        p.doc_canon_numero,
        p.doc_canon_referencia,
    )
    return serials[0] if serials else ""


def _evaluar_fila_con_indice(
    idx: dict[str, Any],
    *,
    fecha: Optional[date],
    desc: str,
    serial_raw: str,
    monto: Optional[float],
    confirmados_activos: Optional[list[ImportacionExtractoPagoConfirmado]] = None,
    idx_confirmados: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    cedula = extraer_cedula_descripcion(desc)
    cedula = _cedula_canon_match(cedula) if cedula else None
    serial_norm = _serial_norm_comparacion(serial_raw)
    seriales_cmp = _seriales_extracto_comparar(serial_raw, serial_norm)

    if not cedula:
        return {
            "cedula": None,
            "serial_norm": serial_norm or None,
            "estado": "PARSE_ERROR",
            "detalle": "No se pudo extraer cédula de Descripción",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
        }
    if not serial_norm or not seriales_cmp:
        return {
            "cedula": cedula,
            "serial_norm": None,
            "estado": "PARSE_ERROR",
            "detalle": "Referencia/serial vacío",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
        }
    if any(len(s) < _MIN_DIGITOS_SERIAL for s in seriales_cmp):
        return {
            "cedula": cedula,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": f"Serial demasiado corto (mín {_MIN_DIGITOS_SERIAL} dígitos)",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
        }
    if monto is None or monto <= 0:
        return {
            "cedula": cedula,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": "Monto Haber inválido o <= 0",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
        }
    if fecha is None:
        return {
            "cedula": cedula,
            "serial_norm": serial_norm,
            "estado": "PARSE_ERROR",
            "detalle": "Fecha inválida",
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
        }

    pids = _prestamo_ids_para_cedula(idx, cedula)
    if not pids:
        return {
            "cedula": cedula,
            "serial_norm": serial_norm,
            "estado": "SIN_PRESTAMO",
            "detalle": (
                "Sin préstamo APROBADO (LIQUIDADO y DESISTIMIENTO no están "
                "disponibles para comparación; no se incluye en la lista)"
            ),
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": None,
            "omitir_lista": True,
        }

    # Índice ya deja 1 APROBADO por cédula (fecha_aprobacion más reciente).
    prestamo_id = int(pids[0])
    pagos = idx["pagos_by_prestamo"].get(prestamo_id, [])
    verif = _verif_cedula_serial(cedula, serial_norm)

    match = _buscar_igual_100_en_prestamo(pagos, seriales_cmp)
    if match is not None:
        pago_id, sp_match = match
        det_extra = ""
        if len(seriales_cmp) > 1:
            det_extra = f"; clave extracto={serial_norm} match parcial en pago={sp_match}"
        ev = _anotar_banco_drive(
            {
                "cedula": cedula,
                "serial_norm": serial_norm,
                "estado": "IGUAL_100",
                "detalle": (
                    f"100% cédula+serial ({verif}; pago_id={pago_id}; "
                    f"prefijos BNC/letras ignorados{det_extra})"
                ),
                "similitud_pct": 100.0,
                "pago_id_match": pago_id,
                "prestamo_id": prestamo_id,
            },
            idx,
            prestamo_id,
        )
        return _anotar_serial_mixto(ev, idx, prestamo_id, pago_id)

    # Al día (última cuota vencida Pagado): no hay cuota que cubrir, no se importa.
    if prestamo_id in (idx.get("pids_al_dia") or set()):
        return {
            "cedula": cedula,
            "serial_norm": serial_norm,
            "estado": "PRESTAMO_PAGADO",
            "detalle": (
                f"Préstamo pagado: la última cuota vencida del prestamo_id="
                f"{prestamo_id} está Pagado ({verif}); depósito no importable"
            ),
            "similitud_pct": None,
            "pago_id_match": None,
            "prestamo_id": prestamo_id,
        }

    best_pct, best_pid, _best_sp = _mejor_similitud_serial_en_prestamo(
        pagos, seriales_cmp
    )

    # Similares: ≥70% (Conciliación Bancos) → revisión manual (Visto).
    if best_pct >= _SIMILITUD_SERIAL_MINIMA:
        ev = _anotar_banco_drive(
            {
                "cedula": cedula,
                "serial_norm": serial_norm,
                "estado": "SEMEJANTE",
                "detalle": (
                    f"Serial semejante {best_pct}% al pago_id={best_pid} "
                    f"({verif}; prestamo_id={prestamo_id}); importar con OK bajo criterio"
                ),
                "similitud_pct": best_pct,
                "pago_id_match": best_pid,
                "prestamo_id": prestamo_id,
            },
            idx,
            prestamo_id,
        )
        return _aplicar_nota_confirmado_pendiente(
            _anotar_serial_mixto(ev, idx, prestamo_id, best_pid),
            idx_confirmados,
            seriales_cmp,
            prestamo_id,
        )

    # Serial no existe en pagos del APROBADO → se puede importar.
    # % = confiabilidad de importación (100% = no hay ese pago en el préstamo).
    return _aplicar_nota_confirmado_pendiente(
        _anotar_banco_drive(
            {
                "cedula": cedula,
                "serial_norm": serial_norm,
                "estado": "SE_PUEDE_IMPORTAR",
                "detalle": (
                    f"Serial ausente en pagos del APROBADO prestamo_id={prestamo_id}; "
                    f"{verif}; confiabilidad importación 100%"
                ),
                "similitud_pct": 100.0,
                "pago_id_match": None,
                "prestamo_id": prestamo_id,
            },
            idx,
            prestamo_id,
        ),
        idx_confirmados,
        seriales_cmp,
        prestamo_id,
    )


def _aplicar_nota_confirmado_pendiente(
    ev: dict[str, Any],
    idx_confirmados: Optional[dict[str, Any]],
    seriales_cmp: list[str],
    prestamo_id: Optional[int],
) -> dict[str, Any]:
    if not idx_confirmados or not idx_confirmados.get("activos"):
        return ev
    pend = _confirmados_activos_para_seriales(
        idx_confirmados["activos"],
        seriales_cmp,
        idx_confirmados=idx_confirmados,
    )
    if not pend:
        return ev
    return _anotar_confirmado_pendiente_en_ev(ev, pend, prestamo_id)


def _evaluar_fila(
    db: Session,
    *,
    fecha: Optional[date],
    desc: str,
    serial_raw: str,
    monto: Optional[float],
    cedula_hint: Optional[str] = None,
) -> dict[str, Any]:
    """Evalúa una fila; índice acotado a la cédula si se conoce (revalidación OK)."""
    filtro = _cedulas_filtro_indice(cedula_hint, extraer_cedula_descripcion(desc))
    idx = _construir_indice_aprobado(db, cedulas_filtro=filtro or None)
    idx_confirmados = _construir_indice_confirmados_activos(db)
    return _evaluar_fila_con_indice(
        idx,
        fecha=fecha,
        desc=desc,
        serial_raw=serial_raw,
        monto=monto,
        idx_confirmados=idx_confirmados,
    )


def comparar_filas_lote(
    db: Session,
    lote_id: int,
    parsed: list[dict[str, Any]],
    *,
    solo_serial: bool,
    t0: Optional[datetime] = None,
    raise_on_empty: bool = True,
) -> dict[str, Any]:
    """Evalúa filas parseadas y persiste resultados (sync o background)."""
    t0 = t0 or datetime.utcnow()
    lote = db.get(ImportacionExtractoLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    cedulas_excel: set[str] = set()
    if not solo_serial:
        for item in parsed:
            ced = extraer_cedula_descripcion(item.get("desc") or "")
            cedulas_excel.update(_cedulas_filtro_indice(ced))

    seriales_excel: set[str] = set()
    if solo_serial:
        for item in parsed:
            sn = _serial_norm_comparacion(item.get("serial_raw") or "")
            seriales_excel.update(
                _seriales_extracto_comparar(item.get("serial_raw") or "", sn or "")
            )

    allow_semejante = len(parsed) < _SKIP_SEMEJANTE_MIN_FILAS
    idx_confirmados: Optional[dict[str, Any]] = None
    if solo_serial:
        idx = _construir_indice_serial_cartera(
            db, serials_filtro=seriales_excel or None
        )
        logger.info(
            "[IMPORT_EXTRACTO] indice serial scoped=%s pagos=%s confirmados=%s (%.1fs)",
            len(seriales_excel),
            len(idx.get("pagos_global") or {}),
            len(idx.get("confirmados_activos") or {}),
            (datetime.utcnow() - t0).total_seconds(),
        )
    else:
        idx = _construir_indice_aprobado(db, cedulas_filtro=cedulas_excel or None)
        idx_confirmados = _construir_indice_confirmados_activos(db)
        logger.info(
            "[IMPORT_EXTRACTO] indice scoped cedulas_excel=%s prestamos_pagos=%s confirmados_activos=%s (%.1fs)",
            len(cedulas_excel),
            len(idx["pagos_by_prestamo"]),
            len(idx_confirmados.get("activos") or []),
            (datetime.utcnow() - t0).total_seconds(),
        )

    stats: dict[str, int] = {}
    pending: list[dict[str, Any]] = []
    n = 0
    eval_n = 0
    commit_cada = len(parsed) >= _LOTE_BG_MIN_FILAS
    for item in parsed:
        eval_n += 1
        if eval_n % _EVAL_LOG_CADA == 0:
            logger.info(
                "[IMPORT_EXTRACTO] evaluadas %s/%s filas (%.1fs)",
                eval_n,
                len(parsed),
                (datetime.utcnow() - t0).total_seconds(),
            )
        if solo_serial:
            ev = _evaluar_fila_serial_cartera(
                idx,
                fecha=item["fecha"],
                serial_raw=item["serial_raw"],
                monto=item["monto"],
                allow_semejante=allow_semejante,
            )
        else:
            ev = _evaluar_fila_con_indice(
                idx,
                fecha=item["fecha"],
                desc=item["desc"],
                serial_raw=item["serial_raw"],
                monto=item["monto"],
                idx_confirmados=idx_confirmados,
            )
        if ev.get("omitir_lista") or ev.get("estado") == "SIN_PRESTAMO":
            stats["OMITIDO_SIN_APROBADO"] = stats.get("OMITIDO_SIN_APROBADO", 0) + 1
            continue
        if ev.get("estado") == "IGUAL_100":
            stats["IGUAL_100"] = stats.get("IGUAL_100", 0) + 1
            continue
        pending.append(
            _mapping_fila_desde_ev(lote.id, item, ev, solo_serial=solo_serial)
        )
        stats[ev["estado"]] = stats.get(ev["estado"], 0) + 1
        if len(pending) >= _INSERT_CHUNK:
            db.bulk_insert_mappings(ImportacionExtractoFila, pending)
            n += len(pending)
            pending.clear()
            if commit_cada:
                db.commit()

    if pending:
        db.bulk_insert_mappings(ImportacionExtractoFila, pending)
        n += len(pending)
        pending.clear()

    if n == 0:
        omit = int(stats.get("OMITIDO_SIN_APROBADO") or 0)
        igual = int(stats.get("IGUAL_100") or 0)
        lote.estado = "ERROR"
        lote.notas = str({"omitido": omit, "igual_100": igual, "error": "sin_filas"})
        db.commit()
        if raise_on_empty:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Ninguna fila para mostrar. "
                    f"sin APROBADO/omitidas: {omit}; 100% iguales: {igual}."
                ),
            )
        return {
            "lote": _lote_dict(lote),
            "stats": stats,
            "filas": 0,
        }

    lote.estado = "COMPARADO"
    lote.notas = str(stats)
    db.commit()
    db.refresh(lote)
    logger.info(
        "[IMPORT_EXTRACTO] lote_id=%s filas=%s stats=%s total=%.1fs",
        lote.id,
        n,
        stats,
        (datetime.utcnow() - t0).total_seconds(),
    )
    return {
        "lote": _lote_dict(lote, stats),
        "stats": stats,
        "filas": n,
    }


def crear_lote_desde_excel(
    db: Session,
    archivo: UploadFile,
    usuario_id: Optional[int],
    *,
    banco: Optional[str] = None,
    modo_cedula: bool = True,
    modo_serial: bool = False,
) -> dict[str, Any]:
    ensure_schema(db)
    if not modo_cedula and not modo_serial:
        raise HTTPException(
            status_code=400,
            detail="Marque al menos Cédula o Serial para continuar con la importación.",
        )
    banco_norm = _normalizar_banco_extracto(banco)
    if not banco_norm:
        raise HTTPException(
            status_code=400,
            detail=(
                "Seleccione el banco del extracto antes de subir el archivo "
                f"({', '.join(sorted(_BANCOS_EXTRACTO_PERMITIDOS))})."
            ),
        )
    t0 = datetime.utcnow()
    solo_serial_upload = bool(modo_serial) and not bool(modo_cedula)
    raw = archivo.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    rows = _cargar_filas_excel(raw, archivo.filename, solo_serial=solo_serial_upload)
    logger.info(
        "[IMPORT_EXTRACTO] archivo=%s bytes=%s filas_excel=%s",
        archivo.filename,
        len(raw),
        max(0, len(rows) - 1),
    )

    # Pasada 1: parsear Excel en memoria y recoger cédulas (sin BD).
    parsed: list[dict[str, Any]] = []
    cedulas_excel: set[str] = set()
    for i, row in enumerate(rows[1:], start=2):
        item = _parse_fila_excel_row(row, i, solo_serial=solo_serial_upload)
        if item is None:
            continue
        if len(parsed) >= MAX_FILAS:
            break
        parsed.append(item)
        if not solo_serial_upload:
            ced = extraer_cedula_descripcion(item["desc"])
            cedulas_excel.update(_cedulas_filtro_indice(ced))

    if solo_serial_upload:
        _validar_seriales_solo_serial(parsed)

    if not parsed:
        plantilla = (
            "Plantilla solo Serial: Fecha | cedula (vacía) | Referencia | Monto"
            if solo_serial_upload
            else "Plantilla extracto banco: Fecha | Descripción | … | Referencia | Haber"
        )
        raise HTTPException(
            status_code=400,
            detail=f"No hay filas válidas. {plantilla}",
        )

    lote = ImportacionExtractoLote(
        usuario_id=usuario_id,
        archivo_nombre=(archivo.filename or "extracto.xlsx")[:255],
        estado="PROCESANDO" if len(parsed) >= _LOTE_BG_MIN_FILAS else "COMPARADO",
        banco=banco_norm,
        modo_cedula=bool(modo_cedula),
        modo_serial=bool(modo_serial),
    )
    db.add(lote)
    db.commit()
    db.refresh(lote)

    solo_serial = _lote_modo_confirmado(lote)

    if len(parsed) >= _LOTE_BG_MIN_FILAS:
        from app.services.importacion_extracto_bg_runner import spawn_comparar_extracto

        spawn_comparar_extracto(lote.id, parsed, solo_serial=solo_serial)
        return {
            "async": True,
            "lote": _lote_dict(lote, {"pendiente": len(parsed)}),
            "stats": {},
            "filas": 0,
            "message": (
                f"Comparando {len(parsed)} filas en segundo plano "
                "(no bloquea el resto del sistema)."
            ),
        }

    return comparar_filas_lote(
        db, lote.id, parsed, solo_serial=solo_serial, t0=t0
    )


def _lote_dict(lote: ImportacionExtractoLote, stats: Optional[dict] = None) -> dict:
    st = stats
    if st is None and lote.notas:
        try:
            import ast

            parsed = ast.literal_eval(lote.notas)
            if isinstance(parsed, dict):
                st = parsed
        except Exception:
            st = None
    return {
        "id": lote.id,
        "archivo_nombre": lote.archivo_nombre,
        "banco": getattr(lote, "banco", None),
        "modo_cedula": bool(getattr(lote, "modo_cedula", True)),
        "modo_serial": bool(getattr(lote, "modo_serial", False)),
        "estado": lote.estado,
        "usuario_id": lote.usuario_id,
        "creado_en": lote.creado_en.isoformat() if lote.creado_en else None,
        "stats": st,
    }


def listar_lotes(db: Session, limit: int = 30) -> list[dict]:
    ensure_schema(db)
    rows = (
        db.execute(
            select(ImportacionExtractoLote)
            .order_by(ImportacionExtractoLote.id.desc())
            .limit(limit)
        )
        .scalars()
        .all()
    )
    return [_lote_dict(r) for r in rows]


def _mapping_fila_desde_ev(
    lote_id: int,
    item: dict[str, Any],
    ev: dict[str, Any],
    *,
    solo_serial: bool,
) -> dict[str, Any]:
    serial_raw = item["serial_raw"]
    monto = item["monto"]
    fecha = item["fecha"]
    desc = item["desc"]
    serial_store = (
        ev.get("serial_norm") or _serial_norm_comparacion(serial_raw) or serial_raw
    )[:100]
    return {
        "lote_id": lote_id,
        "fila_excel": item["fila_excel"],
        "fecha_deposito": fecha,
        "descripcion_raw": desc[:2000] if desc else None,
        "cedula": _cedula_canon_match(ev.get("cedula")) or ev.get("cedula"),
        "serial": serial_store or None,
        "serial_norm": ev.get("serial_norm") or _serial_norm_comparacion(serial_raw) or None,
        "monto_usd": Decimal(str(monto)) if monto is not None else None,
        "estado": ev["estado"],
        "similitud_pct": (
            Decimal(str(ev["similitud_pct"]))
            if ev.get("similitud_pct") is not None
            else None
        ),
        "pago_id_match": ev.get("pago_id_match"),
        "prestamo_id": ev.get("prestamo_id"),
        "detalle": (ev.get("detalle") or "")[:2000],
        "visto": False,
        "importado": False,
        "destino_importacion": ev.get("destino_importacion")
        or ("CONFIRMADO" if solo_serial else "PRESTAMO"),
    }


def _aplicar_filtros_filas_q(
    q,
    *,
    solo_ocultos: bool,
    solo_importables: bool,
    estado: Optional[str],
    excluir_drive: bool,
):
    if solo_ocultos:
        q = q.where(ImportacionExtractoFila.oculto.is_(True))
    else:
        q = q.where(ImportacionExtractoFila.oculto.is_(False))
    if solo_importables:
        q = q.where(
            ImportacionExtractoFila.estado == _ESTADO_FILTRO_100_IMPORTABLE,
            ImportacionExtractoFila.importado.is_(False),
        )
    elif estado:
        q = q.where(ImportacionExtractoFila.estado == estado)
        if estado in _ESTADOS_OK_IMPORTAR:
            q = q.where(ImportacionExtractoFila.importado.is_(False))
    if excluir_drive and not solo_ocultos:
        q = q.where(
            or_(
                ImportacionExtractoFila.detalle.is_(None),
                and_(
                    ~ImportacionExtractoFila.detalle.ilike("%Drive%"),
                    ~ImportacionExtractoFila.detalle.ilike("%banco drive%"),
                ),
            )
        )
    return q


def listar_filas(
    db: Session,
    lote_id: int,
    *,
    estado: Optional[str] = None,
    solo_importables: bool = False,
    solo_ocultos: bool = False,
    limit: int = _FILAS_LISTAR_DEFAULT,
    offset: int = 0,
    excluir_drive: bool = True,
    enriquecer_confirmados: bool = True,
) -> dict[str, Any]:
    ensure_schema(db)
    lote = db.get(ImportacionExtractoLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    lim = max(1, min(int(limit or _FILAS_LISTAR_DEFAULT), _FILAS_LISTAR_MAX))
    off = max(0, int(offset or 0))

    base = select(ImportacionExtractoFila).where(
        ImportacionExtractoFila.lote_id == lote_id,
    )
    base = _aplicar_filtros_filas_q(
        base,
        solo_ocultos=solo_ocultos,
        solo_importables=solo_importables,
        estado=estado,
        excluir_drive=excluir_drive,
    )

    total = int(
        db.scalar(select(func.count()).select_from(base.subquery())) or 0
    )

    q = base.order_by(ImportacionExtractoFila.fila_excel).offset(off).limit(lim)
    rows = db.execute(q).scalars().all()

    enrich: dict[int, dict[str, Any]] = {}
    if (
        enriquecer_confirmados
        and rows
        and any(not f.importado and f.serial_norm for f in rows)
    ):
        idx_conf = _construir_indice_confirmados_activos(db)
        enrich = _enriquecer_filas_confirmado_pendiente(
            db, rows, idx_confirmados=idx_conf
        )

    out: list[dict] = []
    for f in rows:
        d = _fila_dict(f)
        extra = enrich.get(int(f.id))
        if extra:
            d.update(extra)
        out.append(d)

    return {
        "lote_id": lote_id,
        "filas": out,
        "total": total,
        "limit": lim,
        "offset": off,
        "has_more": off + len(out) < total,
    }


def listar_filas_ids(
    db: Session,
    lote_id: int,
    *,
    estado: Optional[str] = None,
    solo_importables: bool = False,
    solo_ocultos: bool = False,
    excluir_drive: bool = True,
    limit: Optional[int] = None,
    offset: int = 0,
) -> dict[str, Any]:
    """Solo IDs (OK masivo sin transferir 25k filas completas)."""
    ensure_schema(db)
    lote = db.get(ImportacionExtractoLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    off = max(0, int(offset or 0))
    base = select(ImportacionExtractoFila.id).where(
        ImportacionExtractoFila.lote_id == lote_id,
    )
    base = _aplicar_filtros_filas_q(
        base,
        solo_ocultos=solo_ocultos,
        solo_importables=solo_importables,
        estado=estado,
        excluir_drive=excluir_drive,
    )
    total = int(
        db.scalar(select(func.count()).select_from(base.subquery())) or 0
    )
    q = base.order_by(ImportacionExtractoFila.fila_excel).offset(off)
    if limit is not None:
        q = q.limit(max(1, min(int(limit), 50_000)))
    ids = [int(x) for x in db.execute(q).scalars().all()]
    return {
        "lote_id": lote_id,
        "ids": ids,
        "total": total,
        "offset": off,
        "has_more": (off + len(ids)) < total if limit is not None else False,
    }


def _fila_dict(f: ImportacionExtractoFila) -> dict:
    detalle = _normalizar_detalle_observaciones(f.detalle)
    alerta_drive = _detalle_tiene_marca_drive(detalle)
    alerta_mixto = _detalle_tiene_marca_serial_compuesto(detalle)
    return {
        "id": f.id,
        "lote_id": f.lote_id,
        "fila_excel": f.fila_excel,
        "fecha_deposito": f.fecha_deposito.isoformat() if f.fecha_deposito else None,
        "descripcion_raw": f.descripcion_raw,
        "cedula": f.cedula,
        "serial": f.serial,
        "serial_norm": f.serial_norm,
        "monto_usd": float(f.monto_usd) if f.monto_usd is not None else None,
        "estado": f.estado,
        "similitud_pct": float(f.similitud_pct) if f.similitud_pct is not None else None,
        "pago_id_match": f.pago_id_match,
        "prestamo_id": f.prestamo_id,
        "pago_id_creado": f.pago_id_creado,
        "detalle": detalle,
        "alerta_banco_drive": alerta_drive,
        "alerta_serial_mixto": alerta_mixto,
        "visto": bool(f.visto),
        "importado": bool(f.importado),
        "oculto": bool(getattr(f, "oculto", False)),
        "destino_importacion": getattr(f, "destino_importacion", None),
        "puede_ok_importar": (
            f.estado in _ESTADOS_OK_IMPORTAR
            and not f.importado
            and not bool(getattr(f, "oculto", False))
        ),
    }


def marcar_visto(db: Session, fila_ids: list[int]) -> dict[str, Any]:
    ensure_schema(db)
    ok = 0
    for fid in fila_ids:
        f = db.get(ImportacionExtractoFila, int(fid))
        if not f:
            continue
        if f.estado not in ("SEMEJANTE", "IGUAL_100", "SE_PUEDE_IMPORTAR"):
            continue
        if f.importado:
            continue
        f.visto = True
        f.estado = "VISTO"
        f.detalle = ((f.detalle or "") + " | Marcado Visto (revisión manual)").strip(" |")
        ok += 1
    db.commit()
    return {"ok": True, "marcados": ok}


def ocultar_filas(db: Session, fila_ids: list[int]) -> dict[str, Any]:
    """Oculta filas del listado de auditoría (no borra datos del lote)."""
    ensure_schema(db)
    ok = 0
    for fid in fila_ids:
        f = db.get(ImportacionExtractoFila, int(fid))
        if not f or bool(getattr(f, "oculto", False)):
            continue
        f.oculto = True
        nota = "Oculto en auditoría"
        if f.detalle and nota not in f.detalle:
            f.detalle = ((f.detalle or "") + f" | {nota}").strip(" |")
        ok += 1
    db.commit()
    return {"ok": True, "ocultados": ok}


def _guardar_placeholder_imagen(db: Session) -> str:
    img_id = uuid.uuid4().hex
    db.add(
        PagoComprobanteImagen(
            id=img_id,
            content_type="image/png",
            imagen_data=_PNG_BLANCO_1X1,
        )
    )
    db.flush()
    return img_id


def _cuotas_pendientes_prestamo(db: Session, prestamo_id: int) -> int:
    from app.models.cuota import Cuota

    return int(
        db.scalar(
            select(func.count())
            .select_from(Cuota)
            .where(
                Cuota.prestamo_id == int(prestamo_id),
                or_(Cuota.total_pagado.is_(None), Cuota.total_pagado < Cuota.monto - 0.01),
            )
        )
        or 0
    )


def _aplicar_cascada_importacion_extracto(
    db: Session, pago: Pago, prestamo_id: int
) -> tuple[bool, str, int, int]:
    """Misma secuencia que carga masiva / export batch: cascada + post-conciliacion."""
    from app.api.v1.endpoints import pagos as pagos_ep

    cc = cp = 0
    try:
        cc, cp = pagos_ep._aplicar_pago_a_cuotas_interno(pago, db)
        pago.estado = pagos_ep._estado_conciliacion_post_cascada(pago, cc, cp)
        if cc + cp > 0:
            return True, f"cuotas_ok={cc} parciales={cp}", cc, cp
        pend = _cuotas_pendientes_prestamo(db, prestamo_id)
        if pend == 0:
            return True, "sin cuotas pendientes (cupo cubierto)", cc, cp
        return False, f"sin aplicar ({pend} cuota(s) pendientes)", cc, cp
    except Exception as e:
        logger.exception(
            "[importacion-extracto] cascada pago_id=%s prestamo_id=%s",
            getattr(pago, "id", None),
            prestamo_id,
        )
        return False, f"error: {str(e)[:200]}", cc, cp


def _crear_confirmado_desde_fila(
    db: Session,
    f: ImportacionExtractoFila,
    *,
    idx: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Modo solo Serial: registra pago confirmado (KPI) sin préstamo ni cascada."""
    estado_inicial = f.estado
    importacion_manual = estado_inicial in ("SEMEJANTE", "VISTO")
    if estado_inicial not in _ESTADOS_OK_IMPORTAR or f.importado:
        return {"ok": False, "motivo": "no_importable", "fila_id": f.id}
    if bool(getattr(f, "oculto", False)):
        return {"ok": False, "motivo": "fila_oculta", "fila_id": f.id}
    if not f.serial_norm or f.monto_usd is None or not f.fecha_deposito:
        return {"ok": False, "motivo": "datos_incompletos", "fila_id": f.id}

    banco = _banco_lote_fila(db, f)
    serial_norm = _serial_norm_comparacion(f.serial_norm or f.serial)
    if not serial_norm:
        return {"ok": False, "motivo": "serial_invalido", "fila_id": f.id}
    f.serial_norm = serial_norm
    if not f.serial:
        f.serial = serial_norm

    serial_raw = f.serial or serial_norm
    monto_f = float(f.monto_usd)
    if idx is None:
        idx = _construir_indice_serial_cartera(db)
    ev = _evaluar_fila_serial_cartera(
        idx,
        fecha=f.fecha_deposito,
        serial_raw=serial_raw,
        monto=monto_f,
    )
    if ev.get("serial_norm"):
        f.serial_norm = ev["serial_norm"]
        serial_norm = ev["serial_norm"]

    ev_estado = str(ev.get("estado") or "")
    if ev_estado == "IGUAL_100" or ev.get("omitir_lista"):
        f.estado = ev_estado or f.estado
        f.pago_id_match = ev.get("pago_id_match")
        if ev.get("similitud_pct") is not None:
            f.similitud_pct = Decimal(str(ev["similitud_pct"]))
        f.detalle = ev.get("detalle") or f.detalle
        return {
            "ok": False,
            "motivo": "igual_100" if ev_estado == "IGUAL_100" else ev_estado.lower(),
            "fila_id": f.id,
        }

    if not importacion_manual and ev_estado != "SE_PUEDE_IMPORTAR":
        f.estado = ev_estado
        f.pago_id_match = ev.get("pago_id_match")
        if ev.get("similitud_pct") is not None:
            f.similitud_pct = Decimal(str(ev["similitud_pct"]))
        f.detalle = ev.get("detalle") or f.detalle
        return {"ok": False, "motivo": ev_estado.lower(), "fila_id": f.id}

    seriales_cmp = _seriales_extracto_comparar(serial_raw, serial_norm)
    match = _buscar_igual_100_global(idx, seriales_cmp)
    if match is not None:
        f.estado = "IGUAL_100"
        f.detalle = ev.get("detalle") or "Serial ya en cartera al confirmar"
        return {"ok": False, "motivo": "igual_100", "fila_id": f.id}

    monto = Decimal(str(round(float(f.monto_usd), 2)))
    conf = ImportacionExtractoPagoConfirmado(
        fila_id=int(f.id),
        lote_id=int(f.lote_id),
        serial=(f.serial or serial_norm)[:100],
        serial_norm=serial_norm[:100],
        monto_usd=monto,
        fecha_deposito=f.fecha_deposito,
        banco=banco,
        estado="ACTIVO",
        detalle=(
            "[IMPORTACION_EXTRACTO] confirmado sin cédula; KPI Pagos confirmados"
            + ("; criterio manual semejante" if importacion_manual else "")
        ),
    )
    db.add(conf)
    db.flush()

    f.importado = True
    f.estado = "IMPORTADO"
    f.destino_importacion = "CONFIRMADO"
    f.detalle = (
        f"Confirmado id={conf.id} (estadística Pagos confirmados); banco={banco}; "
        f"serial={serial_norm}"
        + ("; criterio manual semejante" if importacion_manual else "")
    )
    return {
        "ok": True,
        "fila_id": f.id,
        "confirmado_id": int(conf.id),
        "destino": "CONFIRMADO",
    }


def _crear_pago_desde_fila(
    db: Session,
    f: ImportacionExtractoFila,
    *,
    idx: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Crea pago con datos del Excel + placeholder. No inventa fecha/serial/monto.

    Si ``idx`` viene de ``importar_filas``, se reutiliza (evita reconstruir índice
    APROBADO por cada fila — causa típica de timeout en OK lote).
    """
    estado_inicial = f.estado
    importacion_manual = estado_inicial in ("SEMEJANTE", "VISTO")
    if estado_inicial not in _ESTADOS_OK_IMPORTAR or f.importado:
        return {"ok": False, "motivo": "no_importable", "fila_id": f.id}
    if bool(getattr(f, "oculto", False)):
        return {"ok": False, "motivo": "fila_oculta", "fila_id": f.id}
    if not f.cedula or not f.serial_norm or f.monto_usd is None or not f.fecha_deposito:
        return {"ok": False, "motivo": "datos_incompletos", "fila_id": f.id}
    if not f.prestamo_id:
        return {"ok": False, "motivo": "sin_prestamo", "fila_id": f.id}

    banco = _banco_lote_fila(db, f)

    # Factores críticos: re-normalizar cédula/serial antes de persistir.
    cedula_canon = _cedula_canon_match(f.cedula) or (f.cedula or "").strip().upper()
    serial_norm = _serial_norm_comparacion(f.serial_norm or f.serial)
    if not cedula_canon or not serial_norm:
        return {"ok": False, "motivo": "cedula_o_serial_invalido", "fila_id": f.id}
    f.cedula = cedula_canon
    f.serial_norm = serial_norm
    if not f.serial:
        f.serial = serial_norm

    # Revalidar cédula+serial vs APROBADO (misma regla de comparación).
    desc = f.descripcion_raw or f"DP:{cedula_canon}"
    serial_raw = f.serial or serial_norm
    monto_f = float(f.monto_usd)
    if idx is not None:
        conf_act = idx.get("_idx_confirmados")
        if conf_act is None:
            conf_act = _construir_indice_confirmados_activos(db)
        ev = _evaluar_fila_con_indice(
            idx,
            fecha=f.fecha_deposito,
            desc=desc,
            serial_raw=serial_raw,
            monto=monto_f,
            idx_confirmados=conf_act,
        )
    else:
        ev = _evaluar_fila(
            db,
            fecha=f.fecha_deposito,
            desc=desc,
            serial_raw=serial_raw,
            monto=monto_f,
            cedula_hint=cedula_canon,
        )
    if ev.get("cedula"):
        f.cedula = _cedula_canon_match(ev["cedula"]) or cedula_canon
    if ev.get("serial_norm"):
        f.serial_norm = ev["serial_norm"]
        serial_norm = ev["serial_norm"]

    ev_estado = str(ev.get("estado") or "")
    if ev_estado == "IGUAL_100" or ev.get("omitir_lista") or ev_estado == "SIN_PRESTAMO":
        f.estado = ev_estado or f.estado
        f.pago_id_match = ev.get("pago_id_match")
        if ev.get("similitud_pct") is not None:
            f.similitud_pct = Decimal(str(ev["similitud_pct"]))
        f.detalle = ev.get("detalle") or f.detalle
        if ev.get("prestamo_id"):
            f.prestamo_id = int(ev["prestamo_id"])
        motivo = "igual_100" if ev_estado == "IGUAL_100" else ev_estado.lower()
        return {"ok": False, "motivo": motivo, "fila_id": f.id}

    if not importacion_manual and ev_estado != "SE_PUEDE_IMPORTAR":
        f.estado = ev_estado
        f.pago_id_match = ev.get("pago_id_match")
        if ev.get("similitud_pct") is not None:
            f.similitud_pct = Decimal(str(ev["similitud_pct"]))
        f.detalle = ev.get("detalle") or f.detalle
        if ev.get("prestamo_id"):
            f.prestamo_id = int(ev["prestamo_id"])
        return {
            "ok": False,
            "motivo": ev_estado.lower(),
            "fila_id": f.id,
        }

    if ev.get("prestamo_id"):
        f.prestamo_id = int(ev["prestamo_id"])

    prest = db.get(Prestamo, int(f.prestamo_id))
    est = str(prest.estado or "").upper() if prest else ""
    if (
        not prest
        or est != "APROBADO"
        or est in _ESTADOS_EXCLUIDOS_IMPORTACION
        or prestamo_estado_es_desistimiento(est)
    ):
        return {"ok": False, "motivo": "prestamo_no_aprobado", "fila_id": f.id}

    if not _cedula_coincide_prestamo(db, prest, cedula_canon):
        f.detalle = (
            f"Cédula extracto no alinea con préstamo "
            f"({_verif_cedula_serial(cedula_canon, serial_norm)}; "
            f"prestamo_id={prest.id})"
        )
        return {"ok": False, "motivo": "cedula_no_coincide_prestamo", "fila_id": f.id}

    numero_doc = compose_numero_documento_almacenado(
        serial_norm, None, institucion=banco
    )
    if not numero_doc:
        return {"ok": False, "motivo": "serial_invalido", "fila_id": f.id}

    duplicado = numero_documento_ya_registrado(db, numero_doc)
    if not duplicado and banco == "BNC":
        duplicado = numero_documento_ya_registrado(db, f"BNC/{serial_norm}")
    if duplicado:
        f.estado = "IGUAL_100"
        f.detalle = (
            f"Serial ya registrado al importar "
            f"({_verif_cedula_serial(cedula_canon, serial_norm)}; no se duplicó)"
        )
        return {"ok": False, "motivo": "duplicado_documento", "fila_id": f.id}

    cedula_pago = (
        resolver_cedula_almacenada_en_clientes(db, cedula_canon) or cedula_canon
    )
    institucion = forzar_institucion_drive_si_abonos(numero_doc, banco) or banco

    img_id = _guardar_placeholder_imagen(db)
    link_comprobante = url_comprobante_imagen_absoluta(img_id)
    fecha_dt = datetime.combine(f.fecha_deposito, dt_time(12, 0, 0))
    monto = Decimal(str(round(float(f.monto_usd), 2)))
    verif = _verif_cedula_serial(cedula_canon, serial_norm)
    ahora_conc = datetime.now(ZoneInfo(TZ_NEGOCIO))

    pago = Pago(
        prestamo_id=int(f.prestamo_id),
        cedula_cliente=cedula_pago,
        fecha_pago=fecha_dt,
        monto_pagado=monto,
        numero_documento=numero_doc[:100],
        referencia_pago=serial_norm[:100],
        institucion_bancaria=institucion,
        estado="PAGADO",
        conciliado=True,
        verificado_concordancia="SI",
        fecha_conciliacion=ahora_conc,
        usuario_registro=USUARIO_REGISTRO,
        notas=(
            "[IMPORTACION_EXTRACTO] placeholder imagen; origen Excel extracto; "
            f"banco={banco}; {verif}"
            + ("; importación manual (sem.)" if importacion_manual else "")
        ),
        documento_nombre="placeholder-extracto.png",
        documento_tipo="image/png",
        documento_ruta=link_comprobante,
        link_comprobante=link_comprobante,
        moneda_registro="USD",
    )
    db.add(pago)
    db.flush()
    marcar_pago_autoconciliado(pago, ahora_conc)

    from app.services.conciliacion_bancos_service import (
        registrar_conciliacion_bancaria_importacion_extracto,
    )

    cascada_ok, cascada_det, _cc, _cp = _aplicar_cascada_importacion_extracto(
        db, pago, int(f.prestamo_id)
    )
    if not cascada_ok:
        f.detalle = (
            f"No importado: cascada falló ({cascada_det}); "
            f"{verif}; banco={banco}"
        )
        db.delete(pago)
        db.flush()
        return {
            "ok": False,
            "motivo": "cascada_fallida",
            "fila_id": f.id,
            "detalle": cascada_det,
        }

    conciliacion_ok = False
    try:
        registrar_conciliacion_bancaria_importacion_extracto(
            db,
            pago=pago,
            importacion_lote_id=int(f.lote_id),
            fecha_banco=f.fecha_deposito,
            referencia_banco=serial_norm or numero_doc,
            monto_usd=float(monto),
        )
        conciliacion_ok = True
    except Exception:
        logger.exception(
            "[importacion-extracto] conciliacion bancaria pago_id=%s prestamo_id=%s",
            pago.id,
            f.prestamo_id,
        )

    if not bool(getattr(prest, "requiere_revision", False)):
        prest.requiere_revision = True

    f.pago_id_creado = int(pago.id)
    f.importado = True
    f.estado = "IMPORTADO"
    f.destino_importacion = "PRESTAMO"
    retiro = _retirar_confirmado_activo_por_serial(
        db,
        serial_raw,
        serial_norm,
        pago_id=int(pago.id),
        fila_id=int(f.id),
        prestamo_id=int(f.prestamo_id),
        cedula=cedula_canon,
        monto_usd=float(monto),
    )
    retirados = int(retiro.get("retirados") or 0)
    conf_ids = retiro.get("confirmado_ids") or []
    advertencias = retiro.get("advertencias") or []
    if conf_ids:
        pago.notas = (
            (pago.notas or "")
            + f"; confirmado_aplicado={','.join(str(i) for i in conf_ids)}"
            + f"; prestamo_id={f.prestamo_id}"
        )
    f.detalle = (
        f"Importado pago_id={pago.id} (placeholder); banco={banco}; {verif}; "
        f"cascada {cascada_det}; "
        f"conciliación bancaria {'SI' if conciliacion_ok else 'NO'}; "
        f"prestamo_id={f.prestamo_id}; "
        "requiere_revision=SI"
        + ("; criterio manual semejante" if importacion_manual else "")
        + (
            f"; confirmado(s) aplicado(s)={','.join(str(i) for i in conf_ids)}"
            if conf_ids
            else ""
        )
        + (f"; {' | '.join(advertencias)}" if advertencias else "")
    )
    return {
        "ok": True,
        "fila_id": f.id,
        "pago_id": int(pago.id),
        "cascada_ok": cascada_ok,
        "conciliacion_ok": conciliacion_ok,
        "confirmados_retirados": retirados,
        "confirmado_ids": conf_ids,
        "prestamo_id": int(f.prestamo_id),
        "destino": "PRESTAMO",
    }


def importar_filas(db: Session, fila_ids: list[int]) -> dict[str, Any]:
    """Autoriza importación (OK individual o lote). SE_PUEDE_IMPORTAR, SEMEJANTE o VISTO.

    Optimizaciones anti-timeout (Render / proxy 5 min):
    - Un solo índice APROBADO para todo el lote (no por fila).
    - Commit tras cada fila OK para no perder progreso si el cliente corta.
    - Actualiza el índice en memoria tras cada alta (IGUAL_100 en siguientes del lote).
    """
    ensure_schema(db)
    ids = [int(x) for x in fila_ids if x is not None]
    filas: list[ImportacionExtractoFila] = []
    lotes: dict[int, ImportacionExtractoLote] = {}
    for fid in ids:
        f = db.get(ImportacionExtractoFila, fid)
        if f:
            filas.append(f)
            if f.lote_id and int(f.lote_id) not in lotes:
                lot = db.get(ImportacionExtractoLote, int(f.lote_id))
                if lot:
                    lotes[int(f.lote_id)] = lot

    need_cedula = any(
        lot and not _lote_modo_confirmado(lot) for lot in lotes.values()
    )
    need_serial = any(
        lot and _lote_modo_confirmado(lot) for lot in lotes.values()
    )

    filtro: set[str] = set()
    for f in filas:
        filtro |= _cedulas_filtro_indice(
            f.cedula, extraer_cedula_descripcion(f.descripcion_raw or "")
        )
    idx_cedula = (
        _construir_indice_aprobado(db, cedulas_filtro=filtro or None)
        if need_cedula
        else None
    )
    if idx_cedula is not None:
        idx_cedula["_idx_confirmados"] = _construir_indice_confirmados_activos(db)
    idx_serial = _construir_indice_serial_cartera(db) if need_serial else None

    resultados: list[dict[str, Any]] = []
    for f in filas:
        fid = int(f.id)
        if f.importado:
            resultados.append(
                {"ok": False, "motivo": "ya_importado", "fila_id": fid}
            )
            continue
        lote = lotes.get(int(f.lote_id)) if f.lote_id else None
        modo_conf = _lote_modo_confirmado(lote) if lote else False
        try:
            with db.begin_nested():
                if modo_conf:
                    r = _crear_confirmado_desde_fila(db, f, idx=idx_serial)
                else:
                    r = _crear_pago_desde_fila(db, f, idx=idx_cedula)
                resultados.append(r)
                if not r.get("ok"):
                    raise RuntimeError(str(r.get("motivo") or "import_fallido"))
            pago_id = r.get("pago_id")
            confirmado_id = r.get("confirmado_id")
            prestamo_id = int(f.prestamo_id) if f.prestamo_id else None
            serial_dig = _serial_norm_comparacion(f.serial_norm or f.serial)
            # Persistir fila a fila: si el HTTP corta a mitad, lo ya importado queda.
            try:
                db.commit()
            except Exception as e:
                db.rollback()
                logger.exception(
                    "[importacion-extracto] commit tras fila %s", fid
                )
                resultados[-1] = {
                    "ok": False,
                    "fila_id": fid,
                    "motivo": "error_commit",
                    "detalle": str(e)[:200],
                }
                continue
            # Refrescar índice: siguientes filas ven el serial recién creado.
            if modo_conf and confirmado_id and serial_dig and idx_serial is not None:
                idx_serial.setdefault("confirmados_activos", {}).setdefault(
                    serial_dig, []
                ).append(int(confirmado_id))
            elif pago_id and prestamo_id and serial_dig and idx_cedula is not None:
                pagos_lst = idx_cedula.setdefault("pagos_by_prestamo", {}).setdefault(
                    prestamo_id, []
                )
                pagos_lst.append((int(pago_id), serial_dig))
                conf_ids_aplicados = {
                    int(x) for x in (r.get("confirmado_ids") or [])
                }
                if conf_ids_aplicados and idx_cedula.get("_idx_confirmados"):
                    activos = idx_cedula["_idx_confirmados"].get("activos") or []
                    idx_cedula["_idx_confirmados"]["activos"] = [
                        c
                        for c in activos
                        if int(c.id) not in conf_ids_aplicados
                    ]
                    por_serial = idx_cedula["_idx_confirmados"].get("por_serial") or {}
                    for k in list(por_serial.keys()):
                        por_serial[k] = [
                            c for c in por_serial[k] if int(c.id) not in conf_ids_aplicados
                        ]
                invalidate_serial_cartera_cache()
                if idx_serial is not None:
                    idx_serial.setdefault("pagos_global", {}).setdefault(
                        serial_dig, []
                    ).append((int(pago_id), prestamo_id))
                    idx_serial.setdefault("confirmados_activos", {}).pop(
                        serial_dig, None
                    )
            elif modo_conf and confirmado_id:
                invalidate_serial_cartera_cache()
        except Exception as e:
            logger.exception("[importacion-extracto] importar fila %s", fid)
            if not resultados or resultados[-1].get("fila_id") != fid:
                resultados.append(
                    {
                        "ok": False,
                        "fila_id": fid,
                        "motivo": "error",
                        "detalle": str(e)[:200],
                    }
                )
            try:
                db.rollback()
            except Exception:
                pass

    # Filas pedidas que no existían en BD.
    vistos = {int(r["fila_id"]) for r in resultados if r.get("fila_id") is not None}
    for fid in ids:
        if fid not in vistos:
            resultados.append({"ok": False, "fila_id": fid, "motivo": "no_existe"})

    ok_n = sum(1 for r in resultados if r.get("ok"))
    confirmados_n = sum(
        1
        for r in resultados
        if r.get("ok") and r.get("destino") == "CONFIRMADO"
    )
    if ok_n > 0:
        invalidate_universo_analisis_cache()
    return {
        "ok": True,
        "importados": ok_n,
        "confirmados": confirmados_n,
        "resultados": resultados,
    }
