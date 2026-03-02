#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para debuggear la lectura del Excel
"""
import sys
import os
import io
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl

# Crear un Excel de prueba
print("Creando Excel de prueba...")
wb = openpyxl.Workbook()
ws = wb.active

# Agregar header
ws.append(["Cedula", "Total Financiamiento", "Total Abonos"])

# Agregar datos
ws.append(["V23107415", 864, 810])
ws.append(["E98765432", 2000, 1500])
ws.append(["V12345678", 1000, 500])

# Guardar a archivo temporal
import tempfile
test_file = os.path.join(tempfile.gettempdir(), "test_conciliacion.xlsx")
wb.save(test_file)
print(f"Excel creado: {test_file}")

# Ahora leer el Excel como lo hace el código
print("\n=== LEYENDO EXCEL ===")
wb_read = openpyxl.load_workbook(test_file, read_only=True, data_only=True)
ws_read = wb_read.active

print(f"\nHoja activa: {ws_read.title}")

# Leer todas las filas
print("\nLeyendo todas las filas (incluyendo header):")
all_rows = list(ws_read.iter_rows(values_only=True))
for i, row in enumerate(all_rows):
    print(f"  Fila {i}: {row}")

# Leer desde fila 2 (como hace el código)
print("\nLeyendo desde fila 2 (sin header, como el código):")
rows = list(ws_read.iter_rows(min_row=2, values_only=True))
for i, row in enumerate(rows):
    print(f"  Fila {i} (índice real {i+2}): {row}")
    if row and len(row) >= 3:
        cedula = row[0]
        tf = row[1]
        ta = row[2]
        print(f"    -> Cedula: {cedula} (tipo: {type(cedula).__name__})")
        print(f"    -> TF: {tf} (tipo: {type(tf).__name__})")
        print(f"    -> TA: {ta} (tipo: {type(ta).__name__})")
